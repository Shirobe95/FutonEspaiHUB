import json
import os
import re
import sqlite3
import sys
import warnings
import unicodedata
import tkinter as tk
import tkinter.font as tkfont
from tkinter import filedialog, messagebox, ttk

import openpyxl


# ---------------------------------------------------------------------------
# Rutas y datos compartidos con Calculo de Coste
# ---------------------------------------------------------------------------
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
CONSTANTES_NEGOCIO_PATH = os.path.join(_THIS_DIR, "constantes_negocio.json")
DATA_XLSX_PATH = os.path.join(_THIS_DIR, "data.xlsx")


def load_workbook_silencioso(path, **kwargs):
    """Carga Excel ocultando avisos cosméticos de áreas de impresión inválidas.

    Algunos pedidos antiguos traen nombres de área de impresión que openpyxl
    no puede reconstruir. Es un aviso visual del libro, no afecta a los datos
    de pedido ni al cálculo.
    """
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r"Print area cannot be set.*",
            category=UserWarning,
            module=r"openpyxl\.reader\.workbook",
        )
        return openpyxl.load_workbook(path, **kwargs)



def auto_fit_tree_columns(tree_widget, *, min_width=70, max_width=620, padding=34):
    """Centra y ajusta columnas al contenido cargado, manteniendo scroll horizontal."""
    try:
        body_font = tkfont.nametofont("TkDefaultFont")
    except tk.TclError:
        body_font = None
    try:
        heading_font = tkfont.nametofont("TkHeadingFont")
    except tk.TclError:
        heading_font = body_font

    columns_local = list(tree_widget["columns"])
    for column in columns_local:
        heading = str(tree_widget.heading(column, "text") or column)
        width = (heading_font.measure(heading) if heading_font else len(heading) * 8) + padding
        index = columns_local.index(column)
        for item_id in tree_widget.get_children(""):
            values = tree_widget.item(item_id, "values")
            if index < len(values):
                text = str(values[index] or "")
                measured = (body_font.measure(text) if body_font else len(text) * 8) + padding
                width = max(width, measured)
        tree_widget.heading(column, anchor="center")
        tree_widget.column(column, width=max(min_width, min(width, max_width)), minwidth=min_width, anchor="center", stretch=False)

def center_window(window, width, height):
    """Centra una ventana Tk/Toplevel en la pantalla."""
    window.update_idletasks()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = max((screen_width - width) // 2, 0)
    y = max((screen_height - height) // 2, 0)
    window.geometry(f"{width}x{height}+{x}+{y}")


# ---------------------------------------------------------------------------
# Constantes del negocio
# ---------------------------------------------------------------------------
IMPORTE_DESCARGA_MT = 250
PC_GASTOS_MANIPULACION = 7
PC_GASTOS_FINANCIACION = 7
IMPORTES_VARIOS = 100
COSTE_TOTAL_DESCARGA_FUTONES_IVA = 302.5
COSTE_DESCARGA_FUTONES_UNIDAD = 1.69
IVA_RECARGO_EQUIVALENCIA = 26.20 / 100
COSTE_DIARIO_ALMACENAJE_M3 = 0.3743
PC_PLUS = 0

_CONSTANTES_DEFAULT = {
    "IMPORTE_DESCARGA_MT": IMPORTE_DESCARGA_MT,
    "PC_GASTOS_MANIPULACION": PC_GASTOS_MANIPULACION,
    "PC_GASTOS_FINANCIACION": PC_GASTOS_FINANCIACION,
    "IMPORTES_VARIOS": IMPORTES_VARIOS,
    "COSTE_TOTAL_DESCARGA_FUTONES_IVA": COSTE_TOTAL_DESCARGA_FUTONES_IVA,
    "COSTE_DESCARGA_FUTONES_UNIDAD": COSTE_DESCARGA_FUTONES_UNIDAD,
    "IVA_RECARGO_EQUIVALENCIA": IVA_RECARGO_EQUIVALENCIA,
    "COSTE_DIARIO_ALMACENAJE_M3": COSTE_DIARIO_ALMACENAJE_M3,
}


def cargar_constantes_negocio():
    global IMPORTE_DESCARGA_MT, PC_GASTOS_MANIPULACION, PC_GASTOS_FINANCIACION
    global IMPORTES_VARIOS, COSTE_TOTAL_DESCARGA_FUTONES_IVA
    global COSTE_DESCARGA_FUTONES_UNIDAD, IVA_RECARGO_EQUIVALENCIA
    global COSTE_DIARIO_ALMACENAJE_M3

    valores = dict(_CONSTANTES_DEFAULT)
    if os.path.exists(CONSTANTES_NEGOCIO_PATH):
        try:
            with open(CONSTANTES_NEGOCIO_PATH, "r", encoding="utf-8") as fh:
                guardados = json.load(fh)
            for nombre in valores:
                if nombre in guardados:
                    valores[nombre] = float(guardados[nombre])
        except Exception as exc:
            print(f"Aviso: no se pudieron cargar constantes de negocio: {exc}")

    IMPORTE_DESCARGA_MT = valores["IMPORTE_DESCARGA_MT"]
    PC_GASTOS_MANIPULACION = valores["PC_GASTOS_MANIPULACION"]
    PC_GASTOS_FINANCIACION = valores["PC_GASTOS_FINANCIACION"]
    IMPORTES_VARIOS = valores["IMPORTES_VARIOS"]
    COSTE_TOTAL_DESCARGA_FUTONES_IVA = valores["COSTE_TOTAL_DESCARGA_FUTONES_IVA"]
    COSTE_DESCARGA_FUTONES_UNIDAD = valores["COSTE_DESCARGA_FUTONES_UNIDAD"]
    IVA_RECARGO_EQUIVALENCIA = valores["IVA_RECARGO_EQUIVALENCIA"]
    COSTE_DIARIO_ALMACENAJE_M3 = valores["COSTE_DIARIO_ALMACENAJE_M3"]
    return valores


# ---------------------------------------------------------------------------
# Inventario local
# item = [index, item_id, name, cubic_meters, rotation_c, packages,
#         primary_supplier_price, pascal_price]
# ---------------------------------------------------------------------------
data = []
INVENTORY_DB_PATH = None
SUPPLIER_PRICES_BY_ID = {}
SUPPLIER_LEGACY_FALLBACK = {
    "ekomat": 6,
    "pascal": 7,
    "hemei": 6,
    "heimei": 6,
    "cipta": 6,
}
# Artículos completados manualmente desde la ventana de pedido.
# Permite calcular pedidos a medida sin obligar a guardarlos en WooCommerce.
_MANUAL_ITEMS_BY_REF = {}
_inventory_db_candidates = [
    os.environ.get("FUTON_INVENTORY_DB_PATH", ""),
    os.path.join(_THIS_DIR, "..", "GestorWoo", "data", "gestorwoo.sqlite3"),
    os.path.join(os.getcwd(), "..", "GestorWoo", "data", "gestorwoo.sqlite3"),
]

for _db_path in _inventory_db_candidates:
    if not _db_path:
        continue
    try:
        if os.path.exists(_db_path):
            # IMPORTANTE: al abrir Coste de Pedido desde el HUB, la base puede estar
            # abierta por otras ventanas. En el arranque solo leemos; no migramos ni
            # escribimos supplier_prices aquí para evitar bloqueos de SQLite.
            with sqlite3.connect(_db_path, timeout=1.0) as _conn:
                _rows = _conn.execute(
                    """
                    SELECT
                        item_id,
                        name,
                        cubic_meters,
                        rotation_c,
                        packages,
                        primary_supplier_price,
                        pascal_price
                    FROM inventory_items
                    ORDER BY item_id
                    """
                ).fetchall()
                try:
                    _supplier_rows = _conn.execute(
                        "SELECT item_id, supplier, price FROM supplier_prices"
                    ).fetchall()
                except sqlite3.Error:
                    _supplier_rows = []
            for _item_id, _supplier, _price in _supplier_rows:
                SUPPLIER_PRICES_BY_ID.setdefault(int(_item_id), {})[str(_supplier).lower()] = _price
            for _i, _row in enumerate(_rows, start=1):
                data.append([_i, *_row])
            if data:
                INVENTORY_DB_PATH = _db_path
                print(f"Datos cargados desde inventario local: {_db_path}")
                break
    except Exception as exc:
        print(f"Aviso: no se pudo cargar inventario local {_db_path}: {exc}")

def _codigo_key_para_data(valor):
    raw = str(valor or "").strip()
    if raw.endswith(".0"):
        raw = raw[:-2]
    if raw.replace(".", "", 1).isdigit():
        try:
            return int(raw.lstrip("0") or "0")
        except Exception:
            pass
    return raw


def _precio_excel_limpio(valor):
    if valor in (None, ""):
        return "NO ESTA"
    texto = str(valor).strip()
    if not texto or texto.upper() in {"#N/A", "N/A", "NA", "NO ESTA", "NONE"}:
        return "NO ESTA"
    try:
        return float(texto.replace(",", "."))
    except Exception:
        return texto






def preparar_m3_inventario(ref, nuevo_m3, pending_updates):
    """Actualiza M3 en memoria y deja la escritura SQLite para un lote final."""
    if nuevo_m3 is None or nuevo_m3 <= 0:
        return False, None
    item = buscar_articulo_por_referencia(ref)
    if item is None:
        return False, None
    anterior = safe_float(item[3])
    codigo = codigo_desde_referencia(ref)
    item[3] = float(nuevo_m3)
    if isinstance(codigo, int):
        pending_updates[codigo] = float(nuevo_m3)
        return True, anterior
    return False, anterior


def preparar_precio_calculado_pedido(ref, precio_coste_final, pending_updates):
    """Deja la actualización de Precio Calculado de Pedido para un lote SQLite."""
    if precio_coste_final is None:
        return False
    codigo = codigo_desde_referencia(ref)
    if not isinstance(codigo, int):
        return False
    pending_updates[codigo] = float(precio_coste_final)
    return True


def aplicar_actualizaciones_calculo_en_lote(m3_updates, price_updates):
    """Guarda M3 y Precio Calculado de Pedido en una única transacción.

    Antes se escribía una vez por línea calculada. En pedidos grandes eso podía
    congelar la interfaz y chocar con SQLite si el HUB tenía otra ventana abierta.
    """
    if not INVENTORY_DB_PATH or (not m3_updates and not price_updates):
        return 0, 0
    m3_count = 0
    price_count = 0
    try:
        with sqlite3.connect(INVENTORY_DB_PATH, timeout=8.0) as conn:
            conn.execute("PRAGMA busy_timeout = 8000")
            existing = {row[1] for row in conn.execute("PRAGMA table_info(inventory_items)").fetchall()}
            if price_updates and "order_calculated_price" not in existing:
                conn.execute("ALTER TABLE inventory_items ADD COLUMN order_calculated_price REAL")
            if m3_updates:
                conn.executemany(
                    """
                    UPDATE inventory_items
                    SET cubic_meters = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE item_id = ?
                    """,
                    [(m3, item_id) for item_id, m3 in m3_updates.items()],
                )
                m3_count = len(m3_updates)
            if price_updates:
                conn.executemany(
                    """
                    UPDATE inventory_items
                    SET order_calculated_price = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE item_id = ?
                    """,
                    [(price, item_id) for item_id, price in price_updates.items()],
                )
                price_count = len(price_updates)
            conn.commit()
    except Exception as exc:
        print(f"Aviso: no se pudieron aplicar actualizaciones de cálculo en lote: {exc}")
    return m3_count, price_count



def registrar_log_seguridad_local(module, action, status="OK", entity_type="", entity_id="", details="", context=None, category="General"):
    """Registra un evento de seguridad desde CalculoCoste sin depender de imports del paquete."""
    if not INVENTORY_DB_PATH:
        return
    try:
        payload = json.dumps(context or {}, ensure_ascii=False, default=str) if context else ""
        with sqlite3.connect(INVENTORY_DB_PATH, timeout=8.0) as conn:
            conn.execute("PRAGMA busy_timeout = 5000")
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS security_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    created_at TEXT NOT NULL,
                    category TEXT NOT NULL DEFAULT 'General',
                    severity TEXT NOT NULL DEFAULT 'INFO',
                    module TEXT NOT NULL,
                    action TEXT NOT NULL,
                    status TEXT NOT NULL DEFAULT 'OK',
                    entity_type TEXT DEFAULT '',
                    entity_id TEXT DEFAULT '',
                    details TEXT DEFAULT '',
                    context_json TEXT DEFAULT '',
                    source TEXT DEFAULT 'HUB'
                )
                """
            )
            conn.execute(
                """
                INSERT INTO security_logs (
                    created_at, category, severity, module, action, status,
                    entity_type, entity_id, details, context_json, source
                ) VALUES (?, ?, 'INFO', ?, ?, ?, ?, ?, ?, ?, 'CalculoCoste')
                """,
                (
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    category,
                    module,
                    action,
                    status,
                    entity_type,
                    str(entity_id or ""),
                    details,
                    payload,
                ),
            )
            conn.commit()
    except Exception:
        return

def registrar_pedido_proveedor_en_maestro(resultados):
    """Registra las cantidades del pedido calculado como Pedido de Proveedor local.

    No toca WooCommerce ni Heca. Guarda un resumen por codigo en SQLite para que
    el Mapa Maestro muestre cuanto viene pedido por proveedor. Si se recalcula el
    mismo archivo/proveedor, reemplaza ese pedido para evitar duplicados.
    """
    if not INVENTORY_DB_PATH or not resultados:
        return False, 0
    proveedor = PROVEEDOR_CFG["nombre"]
    archivo = os.path.basename(archivo_pedido_actual or f"pedido_{proveedor.lower()}")
    agregados = 0
    try:
        with sqlite3.connect(INVENTORY_DB_PATH, timeout=3.0) as conn:
            existing = {row[1] for row in conn.execute("PRAGMA table_info(inventory_items)").fetchall()}
            for column, coltype in {
                "supplier_order_qty": "REAL",
                "supplier_order_provider": "TEXT",
                "supplier_order_file": "TEXT",
                "supplier_order_updated_at": "TEXT",
            }.items():
                if column not in existing:
                    conn.execute(f"ALTER TABLE inventory_items ADD COLUMN {column} {coltype}")
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS supplier_pending_order_items (
                    provider TEXT NOT NULL,
                    order_file TEXT NOT NULL,
                    item_id INTEGER NOT NULL,
                    item_code TEXT NOT NULL,
                    quantity REAL DEFAULT 0,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    PRIMARY KEY(provider, order_file, item_id)
                )
                """
            )
            # En el Mapa Maestro mostramos el pedido proveedor PENDIENTE actual por proveedor.
            # Durante pruebas, guardar el mismo pedido con nombres/rutas distintas podía dejar
            # duplicados acumulados y mostrar 3 -> 30, 7 -> 70, etc.
            # Por seguridad operativa, al aceptar un nuevo pedido de este proveedor
            # reemplazamos todo lo pendiente de ese proveedor.
            conn.execute(
                "DELETE FROM supplier_pending_order_items WHERE provider = ?",
                (proveedor,),
            )
            debug_pedido_proveedor("SAVE_START", "proveedor=", proveedor, "archivo=", archivo, "resultados=", len(resultados))
            for res in resultados:
                codigo = codigo_desde_referencia(res.get("referencia"))
                debug_pedido_proveedor(
                    "SAVE_READ",
                    "ref=", repr(res.get("referencia")),
                    "codigo=", codigo, type(codigo),
                    "unidades_raw=", repr(res.get("unidades_raw")), type(res.get("unidades_raw")),
                    "unidades=", repr(res.get("unidades")), type(res.get("unidades")),
                )
                if not isinstance(codigo, int):
                    debug_pedido_proveedor("SAVE_SKIP_CODIGO_INVALIDO", repr(res.get("referencia")), codigo)
                    continue
                unidades = unidades_reales_pedido(res)
                debug_pedido_proveedor("SAVE_NORMALIZED", "codigo=", codigo, "unidades_final=", unidades, type(unidades))
                if unidades <= 0:
                    debug_pedido_proveedor("SAVE_SKIP_UNIDADES", "codigo=", codigo, "unidades_final=", unidades)
                    continue
                conn.execute(
                    """
                    INSERT INTO supplier_pending_order_items (
                        provider, order_file, item_id, item_code, quantity, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    ON CONFLICT(provider, order_file, item_id) DO UPDATE SET
                        quantity = excluded.quantity,
                        updated_at = CURRENT_TIMESTAMP
                    """,
                    (proveedor, archivo, codigo, str(res.get("referencia") or codigo), unidades),
                )
                agregados += 1
                row_guardada = conn.execute(
                    "SELECT provider, order_file, item_id, item_code, quantity FROM supplier_pending_order_items WHERE provider = ? AND order_file = ? AND item_id = ?",
                    (proveedor, archivo, codigo),
                ).fetchone()
                debug_pedido_proveedor("SQL_AFTER_INSERT", row_guardada)

            # Recalcula el total pendiente por articulo para mostrarlo en el Mapa Maestro.
            conn.execute(
                """
                UPDATE inventory_items
                SET supplier_order_qty = NULL,
                    supplier_order_provider = NULL,
                    supplier_order_file = NULL,
                    supplier_order_updated_at = NULL
                """
            )
            totals = conn.execute(
                """
                SELECT
                    item_id,
                    SUM(quantity) AS total_qty,
                    GROUP_CONCAT(DISTINCT provider) AS providers,
                    GROUP_CONCAT(DISTINCT order_file) AS files
                FROM supplier_pending_order_items
                GROUP BY item_id
                """
            ).fetchall()
            debug_pedido_proveedor("SQL_TOTALS_COUNT", len(totals))
            for item_id, total_qty, providers, files in totals:
                debug_pedido_proveedor("SQL_TOTAL", "item_id=", item_id, "total_qty=", total_qty, "providers=", providers, "files=", files)
                conn.execute(
                    """
                    UPDATE inventory_items
                    SET supplier_order_qty = ?,
                        supplier_order_provider = ?,
                        supplier_order_file = ?,
                        supplier_order_updated_at = CURRENT_TIMESTAMP,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE item_id = ?
                    """,
                    (total_qty, providers, files, item_id),
                )
            conn.commit()
            debug_pedido_proveedor("SAVE_DONE", "agregados=", agregados)
        return True, agregados
    except Exception as exc:
        print(f"Aviso: no se pudo registrar Pedido de Proveedor en Mapa Maestro: {exc}")
        return False, agregados


def preguntar_registrar_pedido_proveedor():
    if not resultados_exportar:
        return
    proveedor = PROVEEDOR_CFG["nombre"]
    archivo = os.path.basename(archivo_pedido_actual or "")
    debug_pedido_proveedor("POPUP_START", "proveedor=", proveedor, "archivo=", archivo, "lineas=", len(resultados_exportar))
    for idx, r in enumerate(resultados_exportar[:30], start=1):
        debug_pedido_proveedor(
            "POPUP_LINEA", idx,
            "ref=", r.get("referencia"),
            "unidades_raw=", repr(r.get("unidades_raw")), type(r.get("unidades_raw")),
            "unidades=", repr(r.get("unidades")), type(r.get("unidades")),
            "unidades_reales=", unidades_reales_pedido(r),
        )
    if len(resultados_exportar) > 30:
        debug_pedido_proveedor("POPUP_LINEAS_OMITIDAS", len(resultados_exportar) - 30)
    total_unidades = sum(unidades_reales_pedido(r) for r in resultados_exportar)
    debug_pedido_proveedor("POPUP_TOTAL_UNIDADES", total_unidades)
    if not messagebox.askyesno(
        "Añadir Pedido de Proveedor al Mapa Maestro",
        (
            f"Pedido calculado para {proveedor}.\n\n"
            f"Archivo: {archivo or 'sin nombre'}\n"
            f"Líneas calculadas: {len(resultados_exportar)}\n"
            f"Unidades: {format_num(total_unidades, 0)}\n\n"
            "¿Quieres guardar estas cantidades en la sección de Pedido de Proveedor "
            "del Mapa Maestro?\n\n"
            "Esto reemplazará el pedido pendiente actual de este proveedor para evitar duplicados.\n"
            "No se modifica WooCommerce ni Heca."
        ),
    ):
        return
    ok, agregados = registrar_pedido_proveedor_en_maestro(resultados_exportar)
    if ok:
        messagebox.showinfo(
            "Pedido de Proveedor actualizado",
            f"Se añadieron {agregados} líneas al Mapa Maestro para {proveedor}.",
        )
    else:
        messagebox.showwarning(
            "Pedido no guardado",
            "No se pudo actualizar la sección de Pedido de Proveedor. Revisa la consola.",
        )




# ---------------------------------------------------------------------------
# Pedidos de Proveedor: registro histórico + pendientes para recepción
# ---------------------------------------------------------------------------
def _recalcular_pendientes_pedidos_proveedor(conn):
    conn.execute(
        """
        UPDATE inventory_items
        SET supplier_order_qty = NULL,
            supplier_order_provider = NULL,
            supplier_order_file = NULL,
            supplier_order_updated_at = NULL
        """
    )
    rows = conn.execute(
        """
        SELECT
            soi.item_id,
            SUM(MAX(soi.quantity_ordered - COALESCE(soi.quantity_received, 0), 0)) AS pending_qty,
            GROUP_CONCAT(DISTINCT so.provider) AS providers,
            GROUP_CONCAT(DISTINCT so.order_file) AS files
        FROM supplier_order_items AS soi
        JOIN supplier_orders AS so ON so.order_id = soi.order_id
        WHERE so.status NOT IN ('Recibido', 'Cancelado')
        GROUP BY soi.item_id
        """
    ).fetchall()
    for item_id, pending_qty, providers, files in rows:
        if safe_float(pending_qty) <= 0:
            continue
        conn.execute(
            """
            UPDATE inventory_items
            SET supplier_order_qty = ?,
                supplier_order_provider = ?,
                supplier_order_file = ?,
                supplier_order_updated_at = CURRENT_TIMESTAMP,
                updated_at = CURRENT_TIMESTAMP
            WHERE item_id = ?
            """,
            (pending_qty, providers, files, item_id),
        )


def registrar_pedido_proveedor_en_maestro(resultados):
    """Guarda el pedido calculado como Pedido de Proveedor pendiente.

    No actualiza stock ni coste promedio. Eso se hace luego desde Mapa Maestro → Pedidos
    al confirmar la recepción real de la mercancía.
    """
    if not INVENTORY_DB_PATH or not resultados:
        return False, 0
    proveedor = PROVEEDOR_CFG["nombre"]
    archivo = os.path.basename(archivo_pedido_actual or f"pedido_{proveedor.lower()}")
    agregados = 0
    try:
        with sqlite3.connect(INVENTORY_DB_PATH, timeout=8.0) as conn:
            existing = {row[1] for row in conn.execute("PRAGMA table_info(inventory_items)").fetchall()}
            for column, coltype in {
                "supplier_order_qty": "REAL",
                "supplier_order_provider": "TEXT",
                "supplier_order_file": "TEXT",
                "supplier_order_updated_at": "TEXT",
                "order_calculated_price": "REAL",
                "weighted_average_cost": "REAL",
            }.items():
                if column not in existing:
                    conn.execute(f"ALTER TABLE inventory_items ADD COLUMN {column} {coltype}")
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS supplier_orders (
                    order_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    provider TEXT NOT NULL,
                    order_file TEXT NOT NULL,
                    status TEXT NOT NULL DEFAULT 'Pendiente',
                    total_items REAL DEFAULT 0,
                    total_cost REAL DEFAULT 0,
                    notes TEXT,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(provider, order_file)
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS supplier_order_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    order_id INTEGER NOT NULL,
                    item_id INTEGER NOT NULL,
                    item_code TEXT NOT NULL,
                    item_name TEXT,
                    quantity_ordered REAL DEFAULT 0,
                    quantity_received REAL DEFAULT 0,
                    unit_cost REAL DEFAULT 0,
                    line_cost REAL DEFAULT 0,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(order_id) REFERENCES supplier_orders(order_id) ON DELETE CASCADE
                )
                """
            )
            old = conn.execute(
                "SELECT order_id FROM supplier_orders WHERE provider = ? AND order_file = ?",
                (proveedor, archivo),
            ).fetchone()
            if old:
                conn.execute("DELETE FROM supplier_order_items WHERE order_id = ?", (old[0],))
                conn.execute("DELETE FROM supplier_orders WHERE order_id = ?", (old[0],))
            total_items = sum(unidades_reales_pedido(r) for r in resultados)
            total_cost = sum(safe_float(r.get("precio_coste_final")) * unidades_reales_pedido(r) for r in resultados)
            cur = conn.execute(
                """
                INSERT INTO supplier_orders (provider, order_file, status, total_items, total_cost, notes, updated_at)
                VALUES (?, ?, 'Pendiente', ?, ?, ?, CURRENT_TIMESTAMP)
                """,
                (proveedor, archivo, total_items, total_cost, "Creado desde Calculo de Coste de Pedido"),
            )
            order_id = int(cur.lastrowid)
            for res in resultados:
                codigo = codigo_desde_referencia(res.get("referencia"))
                if not isinstance(codigo, int):
                    continue
                unidades = unidades_reales_pedido(res)
                if unidades <= 0:
                    continue
                unit_cost = safe_float(res.get("precio_coste_final"), 0)
                nombre = str(res.get("denominacion") or res.get("producto") or res.get("referencia") or codigo)
                conn.execute(
                    """
                    INSERT INTO supplier_order_items (
                        order_id, item_id, item_code, item_name, quantity_ordered,
                        quantity_received, unit_cost, line_cost, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, 0, ?, ?, CURRENT_TIMESTAMP)
                    """,
                    (order_id, codigo, str(res.get("referencia") or codigo), nombre, unidades, unit_cost, round(unidades * unit_cost, 4)),
                )
                agregados += 1
            _recalcular_pendientes_pedidos_proveedor(conn)
            conn.commit()
        registrar_log_seguridad_local(
            "Coste de Pedido",
            "Añadir pedido de proveedor",
            status="OK",
            entity_type="Pedido proveedor",
            entity_id=f"{proveedor} · {archivo}",
            details=f"Pedido añadido a Pedidos de Proveedor. Proveedor: {proveedor}. Líneas guardadas: {agregados}. Unidades: {total_items:g}.",
            context={
                "proveedor": proveedor,
                "archivo": archivo,
                "lineas_guardadas": agregados,
                "total_unidades": total_items,
                "total_coste": total_cost,
                "items": [
                    {
                        "codigo": str(r.get("referencia") or codigo_desde_referencia(r.get("referencia")) or ""),
                        "nombre": str(r.get("denominacion") or r.get("producto") or r.get("referencia") or ""),
                        "unidades": unidades_reales_pedido(r),
                        "coste_unitario": safe_float(r.get("precio_coste_final"), 0),
                    }
                    for r in resultados
                    if unidades_reales_pedido(r) > 0
                ],
            },
            category="Pedidos de Proveedores",
        )
        return True, agregados
    except Exception as exc:
        print(f"Aviso: no se pudo registrar Pedido de Proveedor: {exc}")
        return False, agregados


def preguntar_registrar_pedido_proveedor():
    if not resultados_exportar:
        return
    proveedor = PROVEEDOR_CFG["nombre"]
    archivo = os.path.basename(archivo_pedido_actual or "")
    total_unidades = sum(unidades_reales_pedido(r) for r in resultados_exportar)
    total_coste = sum(safe_float(r.get("precio_coste_final")) * unidades_reales_pedido(r) for r in resultados_exportar)
    if not messagebox.askyesno(
        "Añadir a Pedidos de Proveedor",
        (
            f"Pedido calculado para {proveedor}.\n\n"
            f"Archivo: {archivo or 'sin nombre'}\n"
            f"Líneas calculadas: {len(resultados_exportar)}\n"
            f"Unidades: {format_num(total_unidades, 0)}\n"
            f"Coste estimado: {format_money(total_coste)}\n\n"
            "¿Quieres añadir este cálculo a la sección Pedidos del Mapa Maestro?\n\n"
            "Esto lo guardará como pedido pendiente. No suma stock ni pondera costes todavía.\n"
            "La actualización de stock y Coste Promedio Ponderado se hará cuando confirmes la recepción desde Mapa Maestro → Pedidos.\n\n"
            "No se modifica WooCommerce ni Heca."
        ),
    ):
        return
    ok, agregados = registrar_pedido_proveedor_en_maestro(resultados_exportar)
    if ok:
        messagebox.showinfo(
            "Pedido añadido",
            f"Se añadió el pedido a Pedidos de Proveedor. Líneas guardadas: {agregados}.",
        )
    else:
        messagebox.showwarning(
            "Pedido no guardado",
            "No se pudo añadir el pedido. Revisa la consola.",
        )

def precio_disponible(valor):
    return valor not in (None, "", "NO ESTA")


def _registrar_precio_proveedor_memoria(item_id, proveedor_key, precio):
    """Actualiza solo el cache en memoria. No escribe SQLite durante el arranque."""
    if not precio_disponible(precio):
        return
    try:
        item_id = int(item_id)
    except Exception:
        return
    key = str(proveedor_key or "").strip().lower()
    if key == "heimei":
        key = "hemei"
    SUPPLIER_PRICES_BY_ID.setdefault(item_id, {})[key] = precio


def _merge_data_xlsx():
    """Carga/mezcla CalculoCoste/data.xlsx como fuente actualizada de precios.

    La base SQLite mantiene el inventario local. El Excel se usa como capa
    actualizada para los cálculos: si un artículo existe, actualiza nombre, M3,
    rotación, bultos y precios en memoria; si no existe, lo añade solo para el
    cálculo del pedido. No toca WooCommerce.
    """
    if not os.path.exists(DATA_XLSX_PATH):
        return 0, 0
    try:
        wb = load_workbook_silencioso(DATA_XLSX_PATH, data_only=True)
        ws = wb.active
        headers = {str(ws.cell(1, c).value or "").strip().lower(): c for c in range(1, ws.max_column + 1)}
        def c(*names):
            for name in names:
                if name.lower() in headers:
                    return headers[name.lower()]
            return None
        c_id = c("id", "item_id", "codigo", "código")
        c_den = c("denominacion", "denominación", "name", "nombre")
        c_m3 = c("m3", "cubic_meters")
        c_rot = c("rotacionc", "rotacion_c", "rotación c")
        c_bultos = c("bultos", "packages")
        c_p1 = c("precio_1", "precio1", "primary_supplier_price")
        c_p2 = c("precio_2", "precio2", "pascal_price")
        if not c_id:
            return 0, 0

        by_key = {_codigo_key_para_data(item[1]): item for item in data}
        updated = 0
        added = 0
        for row in range(2, ws.max_row + 1):
            raw_id = ws.cell(row, c_id).value
            if raw_id in (None, ""):
                continue
            key = _codigo_key_para_data(raw_id)
            nombre = ws.cell(row, c_den).value if c_den else ""
            m3 = ws.cell(row, c_m3).value if c_m3 else 0
            rot = ws.cell(row, c_rot).value if c_rot else 0
            bultos = ws.cell(row, c_bultos).value if c_bultos else 1
            p1 = _precio_excel_limpio(ws.cell(row, c_p1).value if c_p1 else None)
            p2 = _precio_excel_limpio(ws.cell(row, c_p2).value if c_p2 else None)
            item = by_key.get(key)
            if item is None:
                item = [len(data) + 1, key, nombre, m3, rot, bultos, p1, p2]
                data.append(item)
                by_key[key] = item
                added += 1
            else:
                item[2] = nombre or item[2]
                item[3] = m3 if m3 not in (None, "") else item[3]
                item[4] = rot if rot not in (None, "") else item[4]
                item[5] = bultos if bultos not in (None, "") else item[5]
                item[6] = p1
                item[7] = p2
                updated += 1

            # Precios por proveedor real en memoria y SQLite.
            text = str(nombre or (item[2] if item else "")).lower()
            supplier = "hemei" if "tatami" in text else ("cipta" if ("cama" in text or "sofa" in text or "sofá" in text) else "ekomat")
            _registrar_precio_proveedor_memoria(key, supplier, p1)
            _registrar_precio_proveedor_memoria(key, "pascal", p2)
        return updated, added
    except Exception as exc:
        print(f"Aviso: no se pudo cargar data.xlsx actualizado: {exc}")
        return 0, 0



# ---------------------------------------------------------------------------
# Helpers de cálculo
# ---------------------------------------------------------------------------
def safe_float(v, d=0.0):
    try:
        if v in (None, "", "NO ESTA"):
            return d
        return float(str(v).strip().replace(",", "."))
    except Exception:
        return d


def parse_units_value(value, default=0.0):
    """Lee unidades de Excel/PDF sin arrastrar formato ni textos tipo '3 Pc'.

    Para Pedido de Proveedor, la unidad real del documento es fuente de verdad.
    Esto evita casos donde una cantidad de un dígito termine escalada al guardar
    en el Mapa Maestro.
    """
    if value in (None, "", "NO ESTA"):
        return default
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return default
    text = str(value).strip().replace(" ", " ").replace(",", ".")
    if not text:
        return default
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if match:
        try:
            return float(match.group(0))
        except Exception:
            return default
    return safe_float(text, default)


def unidades_reales_pedido(row):
    """Devuelve las unidades reales del pedido para guardar Pedido Proveedor.

    Si la línea conserva el valor crudo leído del Excel/PDF, se usa ese valor
    antes que cualquier valor calculado/mostrado. Así evitamos conversiones
    accidentales como 3 -> 30 o 7 -> 70 al registrar en Mapa Maestro.
    """
    raw = row.get("unidades_raw")
    raw_units = parse_units_value(raw, None) if raw not in (None, "") else None
    if raw_units is not None and raw_units > 0:
        return raw_units
    return parse_units_value(row.get("unidades"), 0.0)



def debug_pedido_proveedor(*parts):
    # Debug desactivado en version estable. Mantener la funcion evita romper llamadas residuales.
    return None


def _supplier_key(nombre):
    value = str(nombre or "").strip().lower()
    aliases = {"heimei": "hemei"}
    return aliases.get(value, value)


def precio_proveedor_item(item, proveedor_key=None):
    """Devuelve precio por proveedor real, con fallback legacy precio1/precio2.

    Motor nuevo: supplier_prices.
    Respaldo seguro: columnas antiguas para no bloquear el flujo si aún no se migró.
    """
    if item is None:
        return None
    key = _supplier_key(proveedor_key or PROVEEDOR_KEY)
    try:
        item_id = int(item[1])
    except Exception:
        item_id = None
    if item_id is not None:
        price = SUPPLIER_PRICES_BY_ID.get(item_id, {}).get(key)
        if precio_disponible(price):
            return price
    legacy_index = SUPPLIER_LEGACY_FALLBACK.get(key)
    if legacy_index is not None and len(item) > legacy_index:
        return item[legacy_index]
    return None


def guardar_precio_proveedor_local(item_id, proveedor_key, precio, source="Coste de Pedido"):
    key = _supplier_key(proveedor_key)
    if not precio_disponible(precio):
        return
    try:
        item_id = int(item_id)
    except Exception:
        return
    SUPPLIER_PRICES_BY_ID.setdefault(item_id, {})[key] = precio
    if not INVENTORY_DB_PATH:
        return
    try:
        with sqlite3.connect(INVENTORY_DB_PATH, timeout=1.0) as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS supplier_prices (
                    item_id INTEGER NOT NULL,
                    supplier TEXT NOT NULL,
                    price TEXT,
                    currency TEXT DEFAULT 'EUR',
                    source TEXT,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    PRIMARY KEY(item_id, supplier)
                )
                """
            )
            conn.execute(
                """
                INSERT INTO supplier_prices (item_id, supplier, price, currency, source, updated_at)
                VALUES (?, ?, ?, 'EUR', ?, CURRENT_TIMESTAMP)
                ON CONFLICT(item_id, supplier) DO UPDATE SET
                    price = excluded.price,
                    source = excluded.source,
                    updated_at = CURRENT_TIMESTAMP
                """,
                (item_id, {"ekomat": "Ekomat", "pascal": "Pascal", "hemei": "Hemei", "cipta": "Cipta"}.get(key, key.title()), str(precio), source),
            )
            conn.commit()
    except Exception as exc:
        print(f"Aviso: no se pudo guardar precio {key} de {item_id}: {exc}")



_xlsx_updated, _xlsx_added = _merge_data_xlsx()
if _xlsx_updated or _xlsx_added:
    print(f"Datos complementados desde data.xlsx: {_xlsx_updated} actualizados, {_xlsx_added} añadidos")

if not data:
    print("Aviso: no se encontraron datos en inventario local ni en data.xlsx.")

_DATA_BY_ID = {}
for _item in data:
    try:
        _DATA_BY_ID[int(_item[1])] = _item
    except Exception:
        _DATA_BY_ID[str(_item[1]).strip()] = _item


def safe_int(v, d=1):
    try:
        return int(float(v))
    except Exception:
        return d


def format_money(v):
    if v in (None, ""):
        return "—"
    return f"{float(v):.2f} €"


def format_num(v, decimals=3):
    if v in (None, ""):
        return "—"
    value = float(v)
    if int(decimals) <= 0:
        # Importante: para unidades enteras NO se deben quitar ceros finales.
        # Antes: format_num(30, 0) -> "30" -> rstrip("0") -> "3".
        return f"{value:.0f}"
    txt = f"{value:.{decimals}f}"
    return txt.rstrip("0").rstrip(".")


def parse_float_text(raw):
    if raw is None:
        return None
    raw = str(raw).strip().replace(",", ".")
    if not raw:
        return None
    return float(raw)


def normalizar_referencia(ref):
    raw = str(ref).strip()
    if raw.endswith(".0"):
        raw = raw[:-2]
    return raw


EXCLUIR_REPARTO_DESCARGA_KEYWORDS = (
    "funda",
    "cover",
    "topper",
    "pillow",
    "pillows",
    "almohada",
    "almohadas",
)

# Estos artículos contienen palabras como "pillows" en la descripción,
# pero son productos grandes y sí cuentan para repartir la descarga.
FORZAR_REPARTO_DESCARGA_CODIGOS = {
    "0727007",
    "0730009",
    "1242001",
    "1242002",
    "1243001",
    "1244001",
    "1245001",
    "1249001",
}


def normalizar_codigo_reparto(ref):
    """Normaliza códigos para comparar aunque Excel quite ceros iniciales.

    Ejemplo: el código 0730009 puede llegar desde Excel como número 730009.
    Para las excepciones de reparto comparamos ambos como 730009.
    """
    raw = normalizar_referencia(ref)
    if raw.replace(".", "", 1).isdigit():
        if raw.endswith(".0"):
            raw = raw[:-2]
        return raw.lstrip("0") or "0"
    return raw.strip()


FORZAR_REPARTO_DESCARGA_CODIGOS_NORMALIZADOS = {
    normalizar_codigo_reparto(codigo) for codigo in FORZAR_REPARTO_DESCARGA_CODIGOS
}


def normalizar_texto_busqueda(texto):
    texto = str(texto or "").lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto


def referencia_en_forzar_reparto(ref):
    codigo = normalizar_codigo_reparto(ref)
    return codigo in FORZAR_REPARTO_DESCARGA_CODIGOS_NORMALIZADOS


def referencia_parece_codigo(ref):
    raw = normalizar_referencia(ref)
    return raw.lstrip("0").isdigit()


def cuenta_para_reparto_descarga(linea):
    """Devuelve False para artículos pequeños que no cuentan en el reparto.

    Se siguen mostrando y calculando, pero sus unidades no forman parte de
    la cantidad total usada para repartir la descarga del pedido.
    Hay excepciones por referencia para productos grandes que incluyen
    palabras como "pillows" en la descripción.
    """
    if referencia_en_forzar_reparto(linea.get("referencia", "")):
        return True

    texto = " ".join(
        str(linea.get(campo, ""))
        for campo in ("producto", "composicion", "medida", "color", "referencia")
    )
    texto_norm = normalizar_texto_busqueda(texto)
    return not any(keyword in texto_norm for keyword in EXCLUIR_REPARTO_DESCARGA_KEYWORDS)


CUENTA_REPARTO_DESCARGA_OPCIONES = ("Sí", "No")


def valor_visible_cuenta_reparto_descarga(linea):
    """Devuelve Sí/No usando la regla automática cuando la línea aún no fue editada."""
    return "Sí" if linea.get("cuenta_reparto_descarga", cuenta_para_reparto_descarga(linea)) else "No"


def motivo_cuenta_reparto_descarga(linea):
    """Explica por qué una línea cuenta o no para el reparto de descarga."""
    ref = linea.get("referencia", "")
    if referencia_en_forzar_reparto(ref):
        return "Sí por excepción definida: producto grande que cuenta para descarga."

    texto = " ".join(
        str(linea.get(campo, ""))
        for campo in ("producto", "composicion", "medida", "color", "referencia")
    )
    texto_norm = normalizar_texto_busqueda(texto)
    encontrados = [kw for kw in EXCLUIR_REPARTO_DESCARGA_KEYWORDS if kw in texto_norm]
    if encontrados:
        return "No por regla automática: " + ", ".join(encontrados) + " no reparte descarga."
    return "Sí por regla automática: producto principal para reparto de descarga."


def aplicar_regla_cuenta_reparto_descarga(linea):
    linea["cuenta_reparto_descarga"] = cuenta_para_reparto_descarga(linea)
    return linea["cuenta_reparto_descarga"]


def total_unidades_para_reparto(lineas):
    return sum(
        safe_float(linea.get("unidades"), 0)
        for linea in lineas
        if linea.get("cuenta_reparto_descarga", cuenta_para_reparto_descarga(linea))
    )


def codigo_desde_referencia(ref):
    raw = normalizar_referencia(ref)
    limpio = raw.lstrip("0") or "0"
    try:
        return int(limpio)
    except ValueError:
        return raw


def buscar_articulo_por_referencia(ref):
    ref_norm = normalizar_referencia(ref)
    manual = _MANUAL_ITEMS_BY_REF.get(ref_norm)
    if manual is not None:
        return manual

    codigo = codigo_desde_referencia(ref)
    item = _DATA_BY_ID.get(codigo)
    if item is None:
        item = _DATA_BY_ID.get(str(codigo))
    if item is None:
        item = _DATA_BY_ID.get(ref_norm)
    return item


def registrar_item_manual(ref, nombre, m3, rotacion_c, bultos, precio1=None, precio2=None):
    """Crea o actualiza un artículo manual usable en el pedido actual.

    Si la referencia es numérica y existe la base local, también actualiza/crea
    el registro en SQLite. Si la referencia es un nombre de cliente, queda como
    override de sesión para poder calcular el pedido sin contaminar el inventario
    con códigos no reales.
    """
    ref_norm = normalizar_referencia(ref)
    codigo = codigo_desde_referencia(ref_norm)
    item_id_para_memoria = codigo if isinstance(codigo, int) else ref_norm

    item = buscar_articulo_por_referencia(ref_norm)
    if item is None:
        item = [len(data) + 1, item_id_para_memoria, nombre, m3, rotacion_c, int(bultos), precio1, precio2]
        data.append(item)
    else:
        item[2] = nombre
        item[3] = m3
        item[4] = rotacion_c
        item[5] = int(bultos)
        item[6] = precio1
        item[7] = precio2

    _MANUAL_ITEMS_BY_REF[ref_norm] = item
    if isinstance(codigo, int):
        _DATA_BY_ID[codigo] = item
        _DATA_BY_ID[str(codigo)] = item

    guardado_en_db = False
    if INVENTORY_DB_PATH and isinstance(codigo, int):
        try:
            with sqlite3.connect(INVENTORY_DB_PATH, timeout=1.0) as conn:
                conn.execute(
                    """
                    INSERT INTO inventory_items (
                        item_id, name, cubic_meters, rotation_c, packages,
                        primary_supplier_price, pascal_price, source, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                    ON CONFLICT(item_id) DO UPDATE SET
                        name = excluded.name,
                        cubic_meters = excluded.cubic_meters,
                        rotation_c = excluded.rotation_c,
                        packages = excluded.packages,
                        primary_supplier_price = excluded.primary_supplier_price,
                        pascal_price = excluded.pascal_price,
                        updated_at = CURRENT_TIMESTAMP
                    """,
                    (
                        codigo,
                        nombre,
                        float(m3),
                        float(rotacion_c),
                        int(bultos),
                        precio1 if precio_disponible(precio1) else "NO ESTA",
                        precio2 if precio_disponible(precio2) else "NO ESTA",
                        "Coste de Pedido - edición rápida",
                    ),
                )
                conn.commit()
            guardado_en_db = True
        except Exception as exc:
            print(f"Aviso: no se pudo guardar item {ref_norm} en inventario local: {exc}")

    return item, guardado_en_db


def actualizar_m3_inventario(ref, nuevo_m3):
    """Actualiza el M3 del artículo en el inventario local SQLite y en memoria.

    El pedido Excel se considera la fuente más reciente para cubic_meters.
    No toca WooCommerce.
    """
    if not INVENTORY_DB_PATH or nuevo_m3 is None or nuevo_m3 <= 0:
        return False, None

    item = buscar_articulo_por_referencia(ref)
    if item is None:
        return False, None

    anterior = safe_float(item[3])
    codigo = codigo_desde_referencia(ref)

    # Los pedidos a medida pueden usar nombres de cliente como referencia.
    # Esos overrides se calculan en la sesión, pero no se guardan como item_id
    # en SQLite para no ensuciar el inventario con códigos no reales.
    if not isinstance(codigo, int):
        item[3] = float(nuevo_m3)
        return False, anterior

    try:
        with sqlite3.connect(INVENTORY_DB_PATH, timeout=1.0) as conn:
            conn.execute(
                """
                UPDATE inventory_items
                SET cubic_meters = ?, updated_at = CURRENT_TIMESTAMP
                WHERE item_id = ?
                """,
                (float(nuevo_m3), codigo),
            )
            conn.commit()
        item[3] = float(nuevo_m3)
        return True, anterior
    except Exception as exc:
        print(f"Aviso: no se pudo actualizar M3 de {ref}: {exc}")
        return False, anterior



def actualizar_precio_calculado_pedido(ref, precio_coste_final):
    """Guarda el último precio coste final calculado en el Mapa Maestro local.

    No toca WooCommerce. Sirve para comparar visualmente el precio calculado
    del último pedido contra el precio Woo enlazado.
    """
    if not INVENTORY_DB_PATH or precio_coste_final is None:
        return False
    codigo = codigo_desde_referencia(ref)
    if not isinstance(codigo, int):
        return False
    try:
        with sqlite3.connect(INVENTORY_DB_PATH, timeout=1.0) as conn:
            existing = {
                row[1]
                for row in conn.execute("PRAGMA table_info(inventory_items)").fetchall()
            }
            if "order_calculated_price" not in existing:
                conn.execute("ALTER TABLE inventory_items ADD COLUMN order_calculated_price REAL")
            conn.execute(
                """
                UPDATE inventory_items
                SET order_calculated_price = ?, updated_at = CURRENT_TIMESTAMP
                WHERE item_id = ?
                """,
                (float(precio_coste_final), codigo),
            )
            conn.commit()
        return True
    except Exception as exc:
        print(f"Aviso: no se pudo actualizar Precio Calculado de Pedido de {ref}: {exc}")
        return False




def registrar_pedido_proveedor_en_maestro(resultados):
    """Registra las cantidades del pedido calculado como Pedido de Proveedor local.

    No toca WooCommerce ni Heca. Guarda un resumen por codigo en SQLite para que
    el Mapa Maestro muestre cuanto viene pedido por proveedor. Si se recalcula el
    mismo archivo/proveedor, reemplaza ese pedido para evitar duplicados.
    """
    if not INVENTORY_DB_PATH or not resultados:
        return False, 0
    proveedor = PROVEEDOR_CFG["nombre"]
    archivo = os.path.basename(archivo_pedido_actual or f"pedido_{proveedor.lower()}")
    agregados = 0
    try:
        with sqlite3.connect(INVENTORY_DB_PATH, timeout=3.0) as conn:
            existing = {row[1] for row in conn.execute("PRAGMA table_info(inventory_items)").fetchall()}
            for column, coltype in {
                "supplier_order_qty": "REAL",
                "supplier_order_provider": "TEXT",
                "supplier_order_file": "TEXT",
                "supplier_order_updated_at": "TEXT",
            }.items():
                if column not in existing:
                    conn.execute(f"ALTER TABLE inventory_items ADD COLUMN {column} {coltype}")
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS supplier_pending_order_items (
                    provider TEXT NOT NULL,
                    order_file TEXT NOT NULL,
                    item_id INTEGER NOT NULL,
                    item_code TEXT NOT NULL,
                    quantity REAL DEFAULT 0,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    PRIMARY KEY(provider, order_file, item_id)
                )
                """
            )
            # En el Mapa Maestro mostramos el pedido proveedor PENDIENTE actual por proveedor.
            # Durante pruebas, guardar el mismo pedido con nombres/rutas distintas podía dejar
            # duplicados acumulados y mostrar 3 -> 30, 7 -> 70, etc.
            # Por seguridad operativa, al aceptar un nuevo pedido de este proveedor
            # reemplazamos todo lo pendiente de ese proveedor.
            conn.execute(
                "DELETE FROM supplier_pending_order_items WHERE provider = ?",
                (proveedor,),
            )
            debug_pedido_proveedor("SAVE_START", "proveedor=", proveedor, "archivo=", archivo, "resultados=", len(resultados))
            for res in resultados:
                codigo = codigo_desde_referencia(res.get("referencia"))
                debug_pedido_proveedor(
                    "SAVE_READ",
                    "ref=", repr(res.get("referencia")),
                    "codigo=", codigo, type(codigo),
                    "unidades_raw=", repr(res.get("unidades_raw")), type(res.get("unidades_raw")),
                    "unidades=", repr(res.get("unidades")), type(res.get("unidades")),
                )
                if not isinstance(codigo, int):
                    debug_pedido_proveedor("SAVE_SKIP_CODIGO_INVALIDO", repr(res.get("referencia")), codigo)
                    continue
                unidades = unidades_reales_pedido(res)
                debug_pedido_proveedor("SAVE_NORMALIZED", "codigo=", codigo, "unidades_final=", unidades, type(unidades))
                if unidades <= 0:
                    debug_pedido_proveedor("SAVE_SKIP_UNIDADES", "codigo=", codigo, "unidades_final=", unidades)
                    continue
                conn.execute(
                    """
                    INSERT INTO supplier_pending_order_items (
                        provider, order_file, item_id, item_code, quantity, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    ON CONFLICT(provider, order_file, item_id) DO UPDATE SET
                        quantity = excluded.quantity,
                        updated_at = CURRENT_TIMESTAMP
                    """,
                    (proveedor, archivo, codigo, str(res.get("referencia") or codigo), unidades),
                )
                agregados += 1
                row_guardada = conn.execute(
                    "SELECT provider, order_file, item_id, item_code, quantity FROM supplier_pending_order_items WHERE provider = ? AND order_file = ? AND item_id = ?",
                    (proveedor, archivo, codigo),
                ).fetchone()
                debug_pedido_proveedor("SQL_AFTER_INSERT", row_guardada)

            # Recalcula el total pendiente por articulo para mostrarlo en el Mapa Maestro.
            conn.execute(
                """
                UPDATE inventory_items
                SET supplier_order_qty = NULL,
                    supplier_order_provider = NULL,
                    supplier_order_file = NULL,
                    supplier_order_updated_at = NULL
                """
            )
            totals = conn.execute(
                """
                SELECT
                    item_id,
                    SUM(quantity) AS total_qty,
                    GROUP_CONCAT(DISTINCT provider) AS providers,
                    GROUP_CONCAT(DISTINCT order_file) AS files
                FROM supplier_pending_order_items
                GROUP BY item_id
                """
            ).fetchall()
            debug_pedido_proveedor("SQL_TOTALS_COUNT", len(totals))
            for item_id, total_qty, providers, files in totals:
                debug_pedido_proveedor("SQL_TOTAL", "item_id=", item_id, "total_qty=", total_qty, "providers=", providers, "files=", files)
                conn.execute(
                    """
                    UPDATE inventory_items
                    SET supplier_order_qty = ?,
                        supplier_order_provider = ?,
                        supplier_order_file = ?,
                        supplier_order_updated_at = CURRENT_TIMESTAMP,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE item_id = ?
                    """,
                    (total_qty, providers, files, item_id),
                )
            conn.commit()
            debug_pedido_proveedor("SAVE_DONE", "agregados=", agregados)
        return True, agregados
    except Exception as exc:
        print(f"Aviso: no se pudo registrar Pedido de Proveedor en Mapa Maestro: {exc}")
        return False, agregados


def preguntar_registrar_pedido_proveedor():
    if not resultados_exportar:
        return
    proveedor = PROVEEDOR_CFG["nombre"]
    archivo = os.path.basename(archivo_pedido_actual or "")
    debug_pedido_proveedor("POPUP_START", "proveedor=", proveedor, "archivo=", archivo, "lineas=", len(resultados_exportar))
    for idx, r in enumerate(resultados_exportar[:30], start=1):
        debug_pedido_proveedor(
            "POPUP_LINEA", idx,
            "ref=", r.get("referencia"),
            "unidades_raw=", repr(r.get("unidades_raw")), type(r.get("unidades_raw")),
            "unidades=", repr(r.get("unidades")), type(r.get("unidades")),
            "unidades_reales=", unidades_reales_pedido(r),
        )
    if len(resultados_exportar) > 30:
        debug_pedido_proveedor("POPUP_LINEAS_OMITIDAS", len(resultados_exportar) - 30)
    total_unidades = sum(unidades_reales_pedido(r) for r in resultados_exportar)
    debug_pedido_proveedor("POPUP_TOTAL_UNIDADES", total_unidades)
    if not messagebox.askyesno(
        "Añadir Pedido de Proveedor al Mapa Maestro",
        (
            f"Pedido calculado para {proveedor}.\n\n"
            f"Archivo: {archivo or 'sin nombre'}\n"
            f"Líneas calculadas: {len(resultados_exportar)}\n"
            f"Unidades: {format_num(total_unidades, 0)}\n\n"
            "¿Quieres guardar estas cantidades en la sección de Pedido de Proveedor "
            "del Mapa Maestro?\n\n"
            "Esto reemplazará el pedido pendiente actual de este proveedor para evitar duplicados.\n"
            "No se modifica WooCommerce ni Heca."
        ),
    ):
        return
    ok, agregados = registrar_pedido_proveedor_en_maestro(resultados_exportar)
    if ok:
        messagebox.showinfo(
            "Pedido de Proveedor actualizado",
            f"Se añadieron {agregados} líneas al Mapa Maestro para {proveedor}.",
        )
    else:
        messagebox.showwarning(
            "Pedido no guardado",
            "No se pudo actualizar la sección de Pedido de Proveedor. Revisa la consola.",
        )




# ---------------------------------------------------------------------------
# Pedidos de Proveedor: registro histórico + pendientes para recepción
# ---------------------------------------------------------------------------
def _recalcular_pendientes_pedidos_proveedor(conn):
    conn.execute(
        """
        UPDATE inventory_items
        SET supplier_order_qty = NULL,
            supplier_order_provider = NULL,
            supplier_order_file = NULL,
            supplier_order_updated_at = NULL
        """
    )
    rows = conn.execute(
        """
        SELECT
            soi.item_id,
            SUM(MAX(soi.quantity_ordered - COALESCE(soi.quantity_received, 0), 0)) AS pending_qty,
            GROUP_CONCAT(DISTINCT so.provider) AS providers,
            GROUP_CONCAT(DISTINCT so.order_file) AS files
        FROM supplier_order_items AS soi
        JOIN supplier_orders AS so ON so.order_id = soi.order_id
        WHERE so.status NOT IN ('Recibido', 'Cancelado')
        GROUP BY soi.item_id
        """
    ).fetchall()
    for item_id, pending_qty, providers, files in rows:
        if safe_float(pending_qty) <= 0:
            continue
        conn.execute(
            """
            UPDATE inventory_items
            SET supplier_order_qty = ?,
                supplier_order_provider = ?,
                supplier_order_file = ?,
                supplier_order_updated_at = CURRENT_TIMESTAMP,
                updated_at = CURRENT_TIMESTAMP
            WHERE item_id = ?
            """,
            (pending_qty, providers, files, item_id),
        )


def registrar_pedido_proveedor_en_maestro(resultados):
    """Guarda el pedido calculado como Pedido de Proveedor pendiente.

    No actualiza stock ni coste promedio. Eso se hace luego desde Mapa Maestro → Pedidos
    al confirmar la recepción real de la mercancía.
    """
    if not INVENTORY_DB_PATH or not resultados:
        return False, 0
    proveedor = PROVEEDOR_CFG["nombre"]
    archivo = os.path.basename(archivo_pedido_actual or f"pedido_{proveedor.lower()}")
    agregados = 0
    try:
        with sqlite3.connect(INVENTORY_DB_PATH, timeout=8.0) as conn:
            existing = {row[1] for row in conn.execute("PRAGMA table_info(inventory_items)").fetchall()}
            for column, coltype in {
                "supplier_order_qty": "REAL",
                "supplier_order_provider": "TEXT",
                "supplier_order_file": "TEXT",
                "supplier_order_updated_at": "TEXT",
                "order_calculated_price": "REAL",
                "weighted_average_cost": "REAL",
            }.items():
                if column not in existing:
                    conn.execute(f"ALTER TABLE inventory_items ADD COLUMN {column} {coltype}")
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS supplier_orders (
                    order_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    provider TEXT NOT NULL,
                    order_file TEXT NOT NULL,
                    status TEXT NOT NULL DEFAULT 'Pendiente',
                    total_items REAL DEFAULT 0,
                    total_cost REAL DEFAULT 0,
                    notes TEXT,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(provider, order_file)
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS supplier_order_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    order_id INTEGER NOT NULL,
                    item_id INTEGER NOT NULL,
                    item_code TEXT NOT NULL,
                    item_name TEXT,
                    quantity_ordered REAL DEFAULT 0,
                    quantity_received REAL DEFAULT 0,
                    unit_cost REAL DEFAULT 0,
                    line_cost REAL DEFAULT 0,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(order_id) REFERENCES supplier_orders(order_id) ON DELETE CASCADE
                )
                """
            )
            old = conn.execute(
                "SELECT order_id FROM supplier_orders WHERE provider = ? AND order_file = ?",
                (proveedor, archivo),
            ).fetchone()
            if old:
                conn.execute("DELETE FROM supplier_order_items WHERE order_id = ?", (old[0],))
                conn.execute("DELETE FROM supplier_orders WHERE order_id = ?", (old[0],))
            total_items = sum(unidades_reales_pedido(r) for r in resultados)
            total_cost = sum(safe_float(r.get("precio_coste_final")) * unidades_reales_pedido(r) for r in resultados)
            cur = conn.execute(
                """
                INSERT INTO supplier_orders (provider, order_file, status, total_items, total_cost, notes, updated_at)
                VALUES (?, ?, 'Pendiente', ?, ?, ?, CURRENT_TIMESTAMP)
                """,
                (proveedor, archivo, total_items, total_cost, "Creado desde Calculo de Coste de Pedido"),
            )
            order_id = int(cur.lastrowid)
            for res in resultados:
                codigo = codigo_desde_referencia(res.get("referencia"))
                if not isinstance(codigo, int):
                    continue
                unidades = unidades_reales_pedido(res)
                if unidades <= 0:
                    continue
                unit_cost = safe_float(res.get("precio_coste_final"), 0)
                nombre = str(res.get("denominacion") or res.get("producto") or res.get("referencia") or codigo)
                conn.execute(
                    """
                    INSERT INTO supplier_order_items (
                        order_id, item_id, item_code, item_name, quantity_ordered,
                        quantity_received, unit_cost, line_cost, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, 0, ?, ?, CURRENT_TIMESTAMP)
                    """,
                    (order_id, codigo, str(res.get("referencia") or codigo), nombre, unidades, unit_cost, round(unidades * unit_cost, 4)),
                )
                agregados += 1
            _recalcular_pendientes_pedidos_proveedor(conn)
            conn.commit()
        registrar_log_seguridad_local(
            "Coste de Pedido",
            "Añadir pedido de proveedor",
            status="OK",
            entity_type="Pedido proveedor",
            entity_id=f"{proveedor} · {archivo}",
            details=f"Pedido añadido a Pedidos de Proveedor. Proveedor: {proveedor}. Líneas guardadas: {agregados}. Unidades: {total_items:g}.",
            context={
                "proveedor": proveedor,
                "archivo": archivo,
                "lineas_guardadas": agregados,
                "total_unidades": total_items,
                "total_coste": total_cost,
                "items": [
                    {
                        "codigo": str(r.get("referencia") or codigo_desde_referencia(r.get("referencia")) or ""),
                        "nombre": str(r.get("denominacion") or r.get("producto") or r.get("referencia") or ""),
                        "unidades": unidades_reales_pedido(r),
                        "coste_unitario": safe_float(r.get("precio_coste_final"), 0),
                    }
                    for r in resultados
                    if unidades_reales_pedido(r) > 0
                ],
            },
            category="Pedidos de Proveedores",
        )
        return True, agregados
    except Exception as exc:
        print(f"Aviso: no se pudo registrar Pedido de Proveedor: {exc}")
        return False, agregados


def preguntar_registrar_pedido_proveedor():
    if not resultados_exportar:
        return
    proveedor = PROVEEDOR_CFG["nombre"]
    archivo = os.path.basename(archivo_pedido_actual or "")
    total_unidades = sum(unidades_reales_pedido(r) for r in resultados_exportar)
    total_coste = sum(safe_float(r.get("precio_coste_final")) * unidades_reales_pedido(r) for r in resultados_exportar)
    if not messagebox.askyesno(
        "Añadir a Pedidos de Proveedor",
        (
            f"Pedido calculado para {proveedor}.\n\n"
            f"Archivo: {archivo or 'sin nombre'}\n"
            f"Líneas calculadas: {len(resultados_exportar)}\n"
            f"Unidades: {format_num(total_unidades, 0)}\n"
            f"Coste estimado: {format_money(total_coste)}\n\n"
            "¿Quieres añadir este cálculo a la sección Pedidos del Mapa Maestro?\n\n"
            "Esto lo guardará como pedido pendiente. No suma stock ni pondera costes todavía.\n"
            "La actualización de stock y Coste Promedio Ponderado se hará cuando confirmes la recepción desde Mapa Maestro → Pedidos.\n\n"
            "No se modifica WooCommerce ni Heca."
        ),
    ):
        return
    ok, agregados = registrar_pedido_proveedor_en_maestro(resultados_exportar)
    if ok:
        messagebox.showinfo(
            "Pedido añadido",
            f"Se añadió el pedido a Pedidos de Proveedor. Líneas guardadas: {agregados}.",
        )
    else:
        messagebox.showwarning(
            "Pedido no guardado",
            "No se pudo añadir el pedido. Revisa la consola.",
        )

def precio_disponible(valor):
    return valor not in (None, "", "NO ESTA")


def proveedor_actual_info():
    """Devuelve clave y nombre visible del proveedor activo."""
    return _supplier_key(PROVEEDOR_KEY), PROVEEDOR_CFG["nombre"]


def estado_previo_linea(linea):
    """Mensaje claro de qué falta antes de poder calcular una línea."""
    item = buscar_articulo_por_referencia(linea["referencia"])
    if item is None:
        if not referencia_parece_codigo(linea["referencia"]):
            return "Falta inventario: pedido a medida", "warn", item
        return "Falta inventario: referencia no encontrada", "warn", item

    proveedor_key, proveedor_nombre = proveedor_actual_info()
    precio_actual = precio_proveedor_item(item, proveedor_key)
    if not precio_disponible(precio_actual):
        if CALCULO_TIPO == "tatamis" and safe_float(linea.get("precio_excel"), 0) > 0:
            return "Listo para calcular · precio desde pedido", "pending", item
        return f"Falta precio {proveedor_nombre}", "warn", item

    m3_excel = safe_float(linea.get("m3_und"), 0)
    if CALCULO_TIPO == "futones" and m3_excel <= 0:
        return "Falta M3/Und. en pedido", "warn", item

    if CALCULO_TIPO == "tatamis" and safe_float(item[3], 0) <= 0:
        return "Falta M3 en inventario local", "warn", item

    return "Listo para calcular", "pending", item


def validar_rentabilidad(valor, modo_rentabilidad):
    if valor is None:
        return True, None
    if modo_rentabilidad == 0 and valor <= 0:
        return False, "El precio de venta debe ser mayor que 0."
    if modo_rentabilidad == 1 and valor >= 100:
        return False, "La rentabilidad deseada debe ser menor que 100%."
    return True, None


def calcular_coste_unitario_pedido(coste_transporte_iva, coste_total_descarga_iva, m3_total_camion,
                                   cantidad_total_productos, precio_proveedor,
                                   item, m3_excel=None, unidades_referencia=1):
    """Calcula una línea de pedido usando la fórmula de Futones.

    En esta herramienta el M3 por unidad viene del Excel del pedido. Ese valor
    también se actualiza antes en el inventario local, pero aquí lo recibimos
    explícitamente para que el cálculo use siempre el dato recién cargado.
    """
    cargar_constantes_negocio()
    if m3_total_camion == 0 or cantidad_total_productos == 0:
        return None

    m3 = safe_float(m3_excel) if safe_float(m3_excel) > 0 else safe_float(item[3])
    rotacion_c = safe_float(item[4])
    n_bultos = safe_int(item[5])
    unidades_referencia = safe_float(unidades_referencia, 0)

    ct_m3 = round(coste_transporte_iva / m3_total_camion, 2)
    ct_m3_prod = round(ct_m3 * m3, 2)
    ct_total_ref = round(unidades_referencia * ct_m3_prod, 2)
    cd_prod_iva = round(coste_total_descarga_iva / cantidad_total_productos, 2)
    cd_total_ref = round(unidades_referencia * cd_prod_iva, 3)
    iva_re = round(precio_proveedor * IVA_RECARGO_EQUIVALENCIA, 2)
    precio_con_iva = round(precio_proveedor + iva_re, 2)
    coste_descarga = round(ct_m3_prod + cd_prod_iva + precio_con_iva, 2)

    coste_almacenaje_iva = round(COSTE_DIARIO_ALMACENAJE_M3 * m3 * rotacion_c * 1.21, 4)
    coste_picking_iva = round(((n_bultos * 0.3) + 4.12) * 1.21, 3)
    precio_coste_final = round(coste_descarga + coste_almacenaje_iva + coste_picking_iva, 2)

    return {
        "denominacion": item[2],
        "m3_bd": m3,
        "rotacion_c": rotacion_c,
        "n_bultos": n_bultos,
        "coste_transporte_iva": coste_transporte_iva,
        "coste_total_descarga_iva": coste_total_descarga_iva,
        "m3_total_camion": m3_total_camion,
        "cantidad_total_productos": cantidad_total_productos,
        "unidades_referencia": unidades_referencia,
        "ct_m3": ct_m3,
        "ct_m3_prod": ct_m3_prod,
        "ct_total_ref": ct_total_ref,
        "cd_prod_iva": cd_prod_iva,
        "cd_total_ref": cd_total_ref,
        "iva_re": iva_re,
        "precio_con_iva": precio_con_iva,
        "coste_descarga": coste_descarga,
        "coste_almacenaje_iva": coste_almacenaje_iva,
        "coste_picking_iva": coste_picking_iva,
        "precio_coste_final": precio_coste_final,
    }


def calcular_coste_unitario_tatamis_pedido(precio_dolares, precio_euros, factura_transporte,
                                          derechos_aranceles, precio_articulo, item):
    """Calcula una línea de pedido con la fórmula de Maderas/Tatamis."""
    cargar_constantes_negocio()
    if precio_euros == 0:
        return None

    m3 = safe_float(item[3])
    rotacion_c = safe_float(item[4])
    n_bultos = safe_int(item[5])

    tasa_cambio = round(precio_dolares / precio_euros, 6)
    importe_transporte = factura_transporte + derechos_aranceles
    pc_transporte = round((importe_transporte / precio_euros) * 100, 2)
    pc_descarga = round((IMPORTE_DESCARGA_MT * 100) / precio_euros, 2)
    pc_varios = round((IMPORTES_VARIOS / precio_euros) * 100, 2)
    pc_suma = round(
        pc_transporte + pc_descarga + PC_GASTOS_FINANCIACION +
        PC_GASTOS_MANIPULACION + pc_varios,
        2,
    )

    precio_euros_art = round(precio_articulo / tasa_cambio, 2)
    gastos_aplicables = round(precio_euros_art * pc_suma / 100, 2)
    coste_sin_almacenaje = round(precio_euros_art + gastos_aplicables, 2)
    coste_almacenaje_iva = round(COSTE_DIARIO_ALMACENAJE_M3 * m3 * rotacion_c * 1.21, 4)
    coste_picking_iva = round(((n_bultos * 0.3) + 4.12) * 1.21, 3)
    precio_coste_final = round(coste_sin_almacenaje + coste_almacenaje_iva + coste_picking_iva, 2)

    return {
        "denominacion": item[2],
        "m3_bd": m3,
        "rotacion_c": rotacion_c,
        "n_bultos": n_bultos,
        "precio_dolares": precio_dolares,
        "precio_euros": precio_euros,
        "factura_transporte": factura_transporte,
        "derechos_aranceles": derechos_aranceles,
        "tasa_cambio": tasa_cambio,
        "importe_transporte": importe_transporte,
        "pc_transporte": pc_transporte,
        "pc_descarga": pc_descarga,
        "pc_varios": pc_varios,
        "pc_manipulacion": PC_GASTOS_MANIPULACION,
        "pc_financiacion": PC_GASTOS_FINANCIACION,
        "pc_suma": pc_suma,
        "precio_euros_art": precio_euros_art,
        "gastos_aplicables": gastos_aplicables,
        "coste_sin_almacenaje": coste_sin_almacenaje,
        "coste_almacenaje_iva": coste_almacenaje_iva,
        "coste_picking_iva": coste_picking_iva,
        "precio_coste_final": precio_coste_final,
    }


cargar_constantes_negocio()


# ---------------------------------------------------------------------------
# Paleta de colores, igual que coste_1.py
# ---------------------------------------------------------------------------
C_BG = "#F6F7F9"
C_PANEL = "#FFFFFF"
C_PANEL_LINE = "#DDE3EA"
C_INFO = "#F1F5F9"
C_ENTRY = "#FFFFFF"
C_ENTRY_LINE = "#CBD5E1"
C_ENTRY_ERR_BG = "#FEF2F2"
C_BTN = "#2563EB"
C_BTN_HOVER = "#1D4ED8"
C_BTN_ACTIVE = "#1E40AF"
C_EXPORT = "#0F766E"
C_EXPORT_HOVER = "#0D9488"
C_EXPORT_ACTIVE = "#115E59"
C_BTN_FG = "#FFFFFF"
C_LBL = "#1F2937"
C_MUTED = "#64748B"
C_GREY = "#94A3B8"
C_ERR = "#B91C1C"
C_OK = "#15803D"
C_RIGHT = "#EEF2F7"
FONT_LBL = ("Segoe UI", 9)
FONT_LBL_B = ("Segoe UI", 9, "bold")
FONT_TABLE = ("Segoe UI", 9)
FONT_BTN = ("Segoe UI", 10, "bold")


# ---------------------------------------------------------------------------
# Estado de la ventana
# ---------------------------------------------------------------------------
pedido_lineas = []
resultados_exportar = []
archivo_pedido_actual = None


# ---------------------------------------------------------------------------
# Proveedor / modo de cálculo
# ---------------------------------------------------------------------------
PROVEEDORES = {
    "ekomat": {
        "nombre": "Ekomat",
        "titulo": "Pedido Ekomat · Futones",
        "precio_label": "Precio Ekomat",
        "calculo": "futones",
        "parser": "futones",
    },
    "pascal": {
        "nombre": "Pascal",
        "titulo": "Pedido Pascal · Futones",
        "precio_label": "Precio Pascal",
        "calculo": "futones",
        "parser": "futones",
    },
    "hemei": {
        "nombre": "Hemei",
        "titulo": "Pedido Hemei · Tatamis",
        "precio_label": "Precio Hemei",
        "calculo": "tatamis",
        "parser": "heimei",
    },
    # Alias interno para no romper comandos o accesos antiguos.
    "heimei": {
        "nombre": "Hemei",
        "titulo": "Pedido Hemei · Tatamis",
        "precio_label": "Precio Hemei",
        "calculo": "tatamis",
        "parser": "heimei",
    },
    "cipta": {
        "nombre": "Cipta",
        "titulo": "Pedido Cipta · Camas y bases",
        "precio_label": "Precio Cipta",
        "calculo": "tatamis",
        "parser": "cipta",
    },
}


def _proveedor_desde_argv():
    valor = "ekomat"
    for idx, arg in enumerate(sys.argv):
        if arg == "--proveedor" and idx + 1 < len(sys.argv):
            valor = sys.argv[idx + 1].strip().lower()
        elif arg.startswith("--proveedor="):
            valor = arg.split("=", 1)[1].strip().lower()
    # Compatibilidad: versiones antiguas del HUB llamaban al proveedor como
    # "heimei". Canonizamos a "hemei" para que todas las ramas internas
    # usen la misma clave y no se pierdan parsers/PDF/fórmulas.
    aliases = {"heimei": "hemei"}
    valor = aliases.get(valor, valor)
    return valor if valor in PROVEEDORES else "ekomat"


PROVEEDOR_KEY = _proveedor_desde_argv()
PROVEEDOR_CFG = PROVEEDORES[PROVEEDOR_KEY]
CALCULO_TIPO = PROVEEDOR_CFG["calculo"]


# ---------------------------------------------------------------------------
# Ventana principal
# ---------------------------------------------------------------------------
ventana_principal = tk.Tk()
ventana_principal.title(f"FutonSpai · Calculo de Coste de Pedido · {PROVEEDOR_CFG['nombre']}")
center_window(ventana_principal, 1180, 740)
ventana_principal.minsize(1080, 680)
ventana_principal.configure(bg=C_BG)
try:
    ventana_principal.deiconify()
    ventana_principal.state("normal")
    ventana_principal.update_idletasks()
except tk.TclError:
    pass
print(f"Ventana de Coste de Pedido creada para proveedor: {PROVEEDOR_CFG['nombre']}", flush=True)
# En Windows algunas aperturas desde el HUB pueden quedar detrás de la ventana
# principal aunque el proceso esté vivo. Forzamos un primer plano suave solo al
# arrancar para evitar el falso estado de "Traer al frente" sin ventana visible.
def _raise_startup_window():
    try:
        ventana_principal.lift()
        ventana_principal.attributes("-topmost", True)
        ventana_principal.after(350, lambda: ventana_principal.attributes("-topmost", False))
        ventana_principal.focus_force()
    except tk.TclError:
        pass

ventana_principal.after(150, _raise_startup_window)

style = ttk.Style()
style.theme_use("clam")
style.configure(
    "TCombobox",
    fieldbackground=C_ENTRY,
    background=C_ENTRY,
    foreground=C_LBL,
    bordercolor=C_ENTRY_LINE,
    arrowcolor=C_BTN,
    padding=6,
)
style.map("TCombobox", fieldbackground=[("readonly", C_ENTRY)])
style.configure(
    "Treeview",
    font=FONT_TABLE,
    rowheight=28,
    background="#FFFFFF",
    fieldbackground="#FFFFFF",
    borderwidth=1,
    relief="solid",
)
style.configure(
    "Treeview.Heading",
    font=FONT_LBL_B,
    background=C_INFO,
    foreground=C_LBL,
    borderwidth=1,
    relief="solid",
    padding=(4, 6),
)
style.map("Treeview", background=[("selected", "#DBEAFE")], foreground=[("selected", C_LBL)])


# ---------------------------------------------------------------------------
# Panel izquierdo
# ---------------------------------------------------------------------------
left_shell = tk.Frame(
    ventana_principal,
    width=330,
    bg=C_PANEL,
    highlightthickness=1,
    highlightbackground=C_PANEL_LINE,
)
left_shell.pack(side="left", fill="y", padx=(12, 6), pady=12)
left_shell.pack_propagate(False)

left_canvas = tk.Canvas(left_shell, bg=C_PANEL, highlightthickness=0)
left_scroll = tk.Scrollbar(left_shell, orient="vertical", command=left_canvas.yview)
left_canvas.configure(yscrollcommand=left_scroll.set)
left_scroll.pack(side="right", fill="y")
left_canvas.pack(side="top", fill="both", expand=True)

frame_izquierdo = tk.Frame(left_canvas, padx=14, pady=14, bg=C_PANEL)
left_window_id = left_canvas.create_window((0, 0), window=frame_izquierdo, anchor="nw")
left_canvas.bind("<Configure>", lambda e: left_canvas.itemconfig(left_window_id, width=e.width))
frame_izquierdo.bind("<Configure>", lambda e: left_canvas.configure(scrollregion=left_canvas.bbox("all")))
left_canvas.bind("<MouseWheel>", lambda e: left_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))


# ---------------------------------------------------------------------------
# Panel derecho
# ---------------------------------------------------------------------------
right_container = tk.Frame(ventana_principal, bg=C_BG)
right_container.pack(side="right", fill="both", expand=True, padx=(6, 12), pady=12)

right_header = tk.Frame(right_container, bg=C_RIGHT, padx=14, pady=12)
right_header.pack(fill="x")

lbl_titulo_tabla = tk.Label(
    right_header,
    text=f"{PROVEEDOR_CFG['titulo']} · Pedido cargado",
    bg=C_RIGHT,
    fg=C_LBL,
    font=("Segoe UI", 14, "bold"),
    anchor="w",
)
lbl_titulo_tabla.pack(side="left")

main_table_frame = tk.Frame(
    right_container,
    bg=C_PANEL_LINE,
    highlightthickness=1,
    highlightbackground=C_PANEL_LINE,
)
main_table_frame.pack(fill="both", expand=True)

tree_scroll_y = tk.Scrollbar(main_table_frame, orient="vertical")
tree_scroll_y.pack(side="right", fill="y")
tree_scroll_x = tk.Scrollbar(main_table_frame, orient="horizontal")
tree_scroll_x.pack(side="bottom", fill="x")

columns = (
    "referencia", "producto", "color", "und", "cuenta_reparto", "m3_und", "m3_total", "m3_inv_ant", "estado",
    "precio", "coste_transporte_iva", "m3_total_camion", "ct_m3", "ct_m3_prod", "ct_total_ref",
    "coste_total_descarga_iva", "cd_prod_iva", "cd_total_ref",
    "iva_re", "precio_con_iva", "coste_descarga", "alm_iva", "picking_iva",
    "coste_unit", "coste_total", "venta", "venta_total", "rentabilidad",
)
tree = ttk.Treeview(
    main_table_frame,
    columns=columns,
    show="headings",
    yscrollcommand=tree_scroll_y.set,
    xscrollcommand=tree_scroll_x.set,
)
tree_scroll_y.config(command=tree.yview)
tree_scroll_x.config(command=tree.xview)
tree.pack(fill="both", expand=True)

summary_frame = tk.Frame(
    right_container,
    bg=C_INFO,
    highlightthickness=1,
    highlightbackground=C_PANEL_LINE,
    padx=12,
    pady=8,
)
summary_frame.pack(fill="x", pady=(8, 0))

tk.Label(
    summary_frame,
    text="Resumen del pedido",
    bg=C_INFO,
    fg=C_LBL,
    font=FONT_LBL_B,
    anchor="w",
).pack(side="left")

lbl_resumen_tabla = tk.Label(
    summary_frame,
    text="Sin pedido cargado",
    bg=C_INFO,
    fg=C_MUTED,
    font=FONT_LBL_B,
    anchor="e",
    justify="right",
)
lbl_resumen_tabla.pack(side="right", fill="x", expand=True)

_headers = {
    "referencia": "Referencia",
    "producto": "Producto / Medida",
    "color": "Color",
    "und": "Und.",
    "cuenta_reparto": "Cuenta para descarga",
    "m3_und": "M3/Und.",
    "m3_total": "M3 total",
    "m3_inv_ant": "M3 inv. ant.",
    "estado": "Estado",
    "precio": "Precio proveedor",
    "coste_transporte_iva": "Coste total transporte + IVA",
    "m3_total_camion": "M3 camion",
    "ct_m3": "Coste transporte por M3",
    "ct_m3_prod": "Coste transporte por M3 de cada producto",
    "ct_total_ref": "Coste transporte total por referencia",
    "coste_total_descarga_iva": "Coste total descarga + IVA",
    "cd_prod_iva": "Coste descarga por producto + IVA",
    "cd_total_ref": "Coste descarga total por referencia",
    "iva_re": "Importe IVA + RE",
    "precio_con_iva": "Precio compra con IVA + RE",
    "coste_descarga": "Coste final con descarga",
    "alm_iva": "Coste almacenaje + IVA",
    "picking_iva": "Coste unitario picking + IVA",
    "coste_unit": "Precio coste final",
    "coste_total": "Coste total linea",
    "venta": "Precio venta unit.",
    "venta_total": "Venta total linea",
    "rentabilidad": "Rentabilidad",
}
if CALCULO_TIPO == "tatamis":
    _headers.update({
        "cuenta_reparto": "Cuenta para descarga",
        "m3_und": "M3 inventario",
        "m3_total": "M3 total estimado",
        "m3_inv_ant": "M3 inv.",
        "coste_transporte_iva": "Precio en dólares",
        "m3_total_camion": "Precio pagado en euros",
        "ct_m3": "Tasa cambio",
        "ct_m3_prod": "Precio artículo €",
        "ct_total_ref": "Factura transporte",
        "coste_total_descarga_iva": "Derechos aranceles",
        "cd_prod_iva": "% Transporte",
        "cd_total_ref": "% Descarga",
        "iva_re": "% Varios",
        "precio_con_iva": "% Manipulación",
        "coste_descarga": "% Financiación",
        "alm_iva": "Coste almacenaje + IVA",
        "picking_iva": "Coste unitario picking + IVA",
        "coste_unit": "Precio coste final",
    })

_widths = {
    "referencia": 90,
    "producto": 260,
    "color": 180,
    "und": 70,
    "cuenta_reparto": 120,
    "m3_und": 85,
    "m3_total": 90,
    "m3_inv_ant": 95,
    "estado": 185,
    "precio": 105,
    "coste_transporte_iva": 150,
    "m3_total_camion": 100,
    "ct_m3": 150,
    "ct_m3_prod": 190,
    "ct_total_ref": 190,
    "coste_total_descarga_iva": 150,
    "cd_prod_iva": 180,
    "cd_total_ref": 95,
    "iva_re": 85,
    "precio_con_iva": 105,
    "coste_descarga": 120,
    "alm_iva": 90,
    "picking_iva": 95,
    "coste_unit": 95,
    "coste_total": 100,
    "venta": 95,
    "venta_total": 100,
    "rentabilidad": 80,
}
for col in columns:
    tree.heading(col, text=_headers[col], anchor="center")
    tree.column(col, width=_widths[col], minwidth=60, anchor="center", stretch=False)

tree.tag_configure("ok", foreground=C_LBL, background="#FFFFFF")
tree.tag_configure("ok_alt", foreground=C_LBL, background="#F8FAFC")
tree.tag_configure("warn", foreground=C_ERR, background="#FEF2F2")
tree.tag_configure("warn_alt", foreground=C_ERR, background="#FEE2E2")
tree.tag_configure("pending", foreground=C_MUTED, background="#FFFFFF")
tree.tag_configure("pending_alt", foreground=C_MUTED, background="#F8FAFC")


# ---------------------------------------------------------------------------
# Helpers UI
# ---------------------------------------------------------------------------
def make_entry(parent):
    entry = tk.Entry(
        parent,
        bg=C_ENTRY,
        fg=C_LBL,
        relief="flat",
        highlightthickness=1,
        highlightbackground=C_ENTRY_LINE,
        highlightcolor=C_BTN,
        font=FONT_LBL,
        insertbackground=C_LBL,
    )
    return entry


def make_label(parent, text):
    return tk.Label(parent, text=text, bg=C_PANEL, fg=C_LBL, font=FONT_LBL, anchor="w")


def make_button(parent, text, command, bg, hover, active, pady=8):
    btn = tk.Button(
        parent,
        text=text,
        command=command,
        bg=bg,
        fg=C_BTN_FG,
        activebackground=active,
        activeforeground=C_BTN_FG,
        font=FONT_BTN,
        relief="flat",
        cursor="hand2",
        pady=pady,
        bd=0,
        highlightthickness=0,
    )
    btn.bind("<Enter>", lambda _e: btn.config(bg=hover))
    btn.bind("<Leave>", lambda _e: btn.config(bg=bg))
    return btn


def mark_input_error(entry):
    entry.config(bg=C_ENTRY_ERR_BG, highlightbackground=C_ERR, highlightcolor=C_ERR)


def reset_input_errors():
    for entry in ALL_INPUT_ENTRIES:
        entry.config(bg=C_ENTRY, highlightbackground=C_ENTRY_LINE, highlightcolor=C_BTN)


def parse_required_float(entry, label, errors, allow_zero=True):
    raw = entry.get().strip().replace(",", ".")
    if not raw:
        mark_input_error(entry)
        errors.append(f"{label}: campo obligatorio.")
        return None
    try:
        value = float(raw)
    except ValueError:
        mark_input_error(entry)
        errors.append(f"{label}: debe ser un valor numérico.")
        return None
    if not allow_zero and value == 0:
        mark_input_error(entry)
        errors.append(f"{label}: no puede ser 0.")
        return None
    return value


def show_input_errors(errors):
    if errors:
        messagebox.showerror("Revisa los campos", "Corrige estos errores:\n\n- " + "\n- ".join(errors))
        return True
    return False


def limpiar_tabla():
    for row_id in tree.get_children():
        tree.delete(row_id)


def pintar_pedido_en_tabla():
    limpiar_tabla()
    for idx, linea in enumerate(pedido_lineas):
        estado, tag_base, item = estado_previo_linea(linea)
        tag = f"{tag_base}_alt" if idx % 2 else tag_base
        producto = linea["producto"]
        if item is not None and item[2]:
            producto = f"{item[2]} | {linea['medida']}"
        tree.insert(
            "",
            "end",
            iid=str(idx),
            values=(
                linea["referencia"],
                producto,
                linea["color"],
                format_num(linea["unidades"], 0),
                valor_visible_cuenta_reparto_descarga(linea),
                format_num(linea["m3_und"], 6),
                format_num(linea["m3_total"], 3),
                format_num(safe_float(item[3]), 6) if item is not None else "—",
                estado,
                "—", "—", "—", "—", "—", "—", "—", "—", "—",
                "—", "—", "—", "—", "—", "—", "—", "—", "—", "—",
            ),
            tags=(tag,),
        )
    auto_fit_tree_columns(tree)
    actualizar_resumen_tabla()


def actualizar_resumen_tabla(extra=""):
    total_productos = len(pedido_lineas)
    total_unidades = sum(safe_float(l["unidades"]) for l in pedido_lineas)
    unidades_reparto = total_unidades_para_reparto(pedido_lineas)
    total_m3 = sum(safe_float(l["m3_total"]) for l in pedido_lineas)
    base = (
        f"Productos: {total_productos}    |    "
        f"Unidades: {format_num(total_unidades, 0)}    |    "
        f"Unidades reparto descarga: {format_num(unidades_reparto, 0)}    |    "
        f"M3 total: {format_num(total_m3, 3)}"
    )
    if extra:
        base += f"    |    {extra}"
    lbl_resumen_tabla.config(text=base if total_productos else "Sin pedido cargado")
    lbl_contador_productos.config(text=f"Productos: {total_productos}")
    lbl_contador_unidades.config(text=f"Unidades: {format_num(total_unidades, 0)}")
    lbl_contador_unidades_reparto.config(text=f"Unidades reparto descarga: {format_num(unidades_reparto, 0)}")
    lbl_contador_m3.config(text=f"M3 total: {format_num(total_m3, 3)}")




def abrir_editor_linea(event=None):
    """Abre una ventana para completar los datos faltantes de una línea."""
    seleccion = tree.selection()
    if not seleccion:
        return
    try:
        idx = int(seleccion[0])
        linea = pedido_lineas[idx]
    except Exception:
        return

    item = buscar_articulo_por_referencia(linea["referencia"])
    proveedor_key, proveedor_nombre = proveedor_actual_info()
    ref = normalizar_referencia(linea.get("referencia", ""))
    es_codigo = referencia_parece_codigo(ref)

    win = tk.Toplevel(ventana_principal)
    win.title(f"Completar datos · {ref}")
    win.configure(bg=C_BG)
    win.transient(ventana_principal)
    win.grab_set()
    center_window(win, 700, 720)
    win.minsize(640, 660)

    cont = tk.Frame(win, bg=C_PANEL, padx=18, pady=16, highlightthickness=1, highlightbackground=C_PANEL_LINE)
    cont.pack(fill="both", expand=True, padx=12, pady=12)

    tk.Label(
        cont,
        text="Completar datos para calcular",
        bg=C_PANEL,
        fg=C_LBL,
        font=("Segoe UI", 13, "bold"),
        anchor="w",
    ).pack(fill="x")
    subtitulo = (
        "Los campos marcados en rojo son necesarios. "
        "Si la referencia es un pedido a medida, se usará solo para este pedido."
    )
    tk.Label(cont, text=subtitulo, bg=C_PANEL, fg=C_MUTED, font=FONT_LBL, wraplength=560, justify="left", anchor="w").pack(fill="x", pady=(4, 12))

    form = tk.Frame(cont, bg=C_PANEL)
    form.pack(fill="both", expand=True)

    def fila(label, value="", readonly=False):
        row = tk.Frame(form, bg=C_PANEL)
        row.pack(fill="x", pady=4)
        tk.Label(row, text=label, bg=C_PANEL, fg=C_LBL, font=FONT_LBL, anchor="w", width=22).pack(side="left")
        ent = make_entry(row)
        ent.insert(0, "" if value is None else str(value))
        ent.pack(side="left", fill="x", expand=True)
        if readonly:
            ent.config(state="readonly", disabledbackground=C_INFO, disabledforeground=C_LBL)
        return ent

    nombre_base = item[2] if item is not None and item[2] else linea.get("producto", ref)
    m3_base = linea.get("m3_und") or (item[3] if item is not None else "")
    rot_base = item[4] if item is not None else ""
    bultos_base = item[5] if item is not None else ""
    precio1_base = item[6] if item is not None else ""
    precio2_base = item[7] if item is not None else ""

    entry_ref = fila("Referencia", ref, readonly=True)
    entry_nombre = fila("Descripción", nombre_base)
    entry_unidades = fila("Unidades", format_num(linea.get("unidades", 0), 0))
    entry_m3 = fila("M3/Und.", m3_base)
    entry_rot = fila("Rotación C", rot_base)
    entry_bultos = fila("Nº bultos", bultos_base)
    entry_precio1 = fila("Precio 1 (Ekomat / Hemei)", precio1_base)
    entry_precio2 = fila("Precio 2 (Pascal)", precio2_base)

    row_cuenta = tk.Frame(form, bg=C_PANEL)
    row_cuenta.pack(fill="x", pady=4)
    tk.Label(row_cuenta, text="Cuenta para descarga", bg=C_PANEL, fg=C_LBL, font=FONT_LBL, anchor="w", width=22).pack(side="left")
    selector_cuenta = ttk.Combobox(row_cuenta, values=CUENTA_REPARTO_DESCARGA_OPCIONES, state="readonly", font=FONT_LBL, width=12)
    selector_cuenta.set(valor_visible_cuenta_reparto_descarga(linea))
    selector_cuenta.pack(side="left", fill="x", expand=False)
    tk.Label(row_cuenta, text="▼", bg=C_PANEL, fg=C_MUTED, font=FONT_LBL_B).pack(side="left", padx=(3, 0))

    lbl_cuenta_motivo = tk.Label(form, text=motivo_cuenta_reparto_descarga(linea), bg=C_INFO, fg=C_MUTED, font=FONT_SMALL, wraplength=620, justify="left", anchor="w")
    lbl_cuenta_motivo.pack(fill="x", pady=(0, 6))

    def actualizar_motivo_cuenta_manual(_event=None):
        auto = motivo_cuenta_reparto_descarga(linea)
        lbl_cuenta_motivo.config(text=f"Selección actual: {selector_cuenta.get()}. Regla automática: {auto}")

    selector_cuenta.bind("<<ComboboxSelected>>", actualizar_motivo_cuenta_manual)

    def aplicar_regla_auto_cuenta():
        aplicar_regla_cuenta_reparto_descarga(linea)
        selector_cuenta.set(valor_visible_cuenta_reparto_descarga(linea))
        lbl_cuenta_motivo.config(text=motivo_cuenta_reparto_descarga(linea))

    tk.Button(row_cuenta, text="Regla auto", command=aplicar_regla_auto_cuenta, bg=C_BTN_ALT, fg=C_BTN_ALT_FG, relief="flat", padx=10).pack(side="left", padx=(8, 0))

    aviso = tk.Label(cont, text="", bg=C_PANEL, fg=C_ERR, font=FONT_LBL, wraplength=560, justify="left", anchor="w")
    aviso.pack(fill="x", pady=(8, 2))

    entradas = [entry_nombre, entry_unidades, entry_m3, entry_rot, entry_bultos, entry_precio1, entry_precio2]

    def limpiar_marcas():
        for ent in entradas:
            ent.config(bg=C_ENTRY, highlightbackground=C_ENTRY_LINE, highlightcolor=C_BTN)

    def marcar(ent):
        ent.config(bg=C_ENTRY_ERR_BG, highlightbackground=C_ERR, highlightcolor=C_ERR)

    def leer_float(ent):
        return parse_float_text(ent.get())

    def validar_y_marcar_inicial():
        limpiar_marcas()
        if not entry_nombre.get().strip():
            marcar(entry_nombre)
        if leer_float(entry_unidades) is None or safe_float(entry_unidades.get(), 0) <= 0:
            marcar(entry_unidades)
        if leer_float(entry_m3) is None or safe_float(entry_m3.get(), 0) <= 0:
            marcar(entry_m3)
        if leer_float(entry_rot) is None:
            marcar(entry_rot)
        if leer_float(entry_bultos) is None or safe_int(entry_bultos.get(), 0) <= 0:
            marcar(entry_bultos)
        proveedor_entry = entry_precio1 if proveedor_key != "pascal" else entry_precio2
        if leer_float(proveedor_entry) is None or safe_float(proveedor_entry.get(), 0) <= 0:
            marcar(proveedor_entry)

    def aceptar():
        limpiar_marcas()
        errores = []
        nombre = entry_nombre.get().strip()
        unidades = leer_float(entry_unidades)
        m3 = leer_float(entry_m3)
        rotacion = leer_float(entry_rot)
        bultos = leer_float(entry_bultos)
        precio1 = entry_precio1.get().strip().replace(",", ".")
        precio2 = entry_precio2.get().strip().replace(",", ".")

        if not nombre:
            marcar(entry_nombre)
            errores.append("Descripción")
        if unidades is None or unidades <= 0:
            marcar(entry_unidades)
            errores.append("Unidades")
        if m3 is None or m3 <= 0:
            marcar(entry_m3)
            errores.append("M3/Und.")
        if rotacion is None:
            marcar(entry_rot)
            errores.append("Rotación C")
        if bultos is None or bultos <= 0:
            marcar(entry_bultos)
            errores.append("Nº bultos")

        precio_proveedor = precio1 if proveedor_key != "pascal" else precio2
        if parse_float_text(precio_proveedor) is None or safe_float(precio_proveedor, 0) <= 0:
            marcar(entry_precio1 if proveedor_key != "pascal" else entry_precio2)
            errores.append(f"Precio {proveedor_nombre}")

        if errores:
            aviso.config(text="Faltan o son incorrectos: " + ", ".join(errores))
            return

        # Mantener valores no usados como NO ESTA para que la tabla avise si se cambia de proveedor.
        precio1_final = precio1 if precio1 else "NO ESTA"
        precio2_final = precio2 if precio2 else "NO ESTA"
        item_actualizado, guardado_en_db = registrar_item_manual(ref, nombre, float(m3), float(rotacion), int(float(bultos)), precio1_final, precio2_final)

        linea["producto"] = nombre
        linea["unidades"] = float(unidades)
        linea["m3_und"] = float(m3)
        linea["m3_total"] = round(float(m3) * float(unidades), 6)
        linea["cuenta_reparto_descarga"] = selector_cuenta.get() == "Sí"
        linea["codigo"] = codigo_desde_referencia(ref)
        linea["editado_manual"] = True
        linea["guardado_en_inventario_local"] = guardado_en_db

        global resultados_exportar
        resultados_exportar = []
        pintar_pedido_en_tabla()
        origen = "guardado en inventario local" if guardado_en_db else "guardado para este pedido"
        lbl_estado_calculo.config(text=f"Datos de {ref} actualizados ({origen}). Vuelve a calcular el pedido.", fg=C_MUTED)
        win.destroy()

    validar_y_marcar_inicial()

    buttons = tk.Frame(cont, bg=C_PANEL)
    buttons.pack(fill="x", pady=(12, 0))
    tk.Button(
        buttons,
        text="Cancelar",
        command=win.destroy,
        bg=C_GREY,
        fg=C_BTN_FG,
        activebackground="#64748B",
        activeforeground=C_BTN_FG,
        font=FONT_BTN,
        relief="flat",
        padx=16,
        pady=8,
    ).pack(side="right", padx=(8, 0))
    tk.Button(
        buttons,
        text="Aceptar",
        command=aceptar,
        bg=C_BTN,
        fg=C_BTN_FG,
        activebackground=C_BTN_ACTIVE,
        activeforeground=C_BTN_FG,
        font=FONT_BTN,
        relief="flat",
        padx=16,
        pady=8,
    ).pack(side="right")


def on_tree_double_click(event=None):
    row_id = tree.identify_row(event.y) if event is not None else None
    if row_id:
        tree.selection_set(row_id)
    abrir_editor_linea(event)

tree.bind("<Double-1>", on_tree_double_click)


def _normalizar_header(value):
    txt = normalizar_texto_busqueda(value)
    # Deja solo letras y números para tolerar saltos de línea, paréntesis,
    # puntos, barras y columnas bilingües como "CódigoArtículo (Item Code)".
    return re.sub(r"[^a-z0-9]+", "", txt)


def _detectar_header_pedido(ws):
    """Detecta cabeceras de Ekomat/Pascal y Hemei."""
    header_row = None
    header_map = {}
    for row_idx in range(1, min(ws.max_row, 40) + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, min(ws.max_column, 14) + 1)]
        normalized = [_normalizar_header(v) for v in values]
        has_ref_futon = "referencia" in normalized
        has_ref_heimei = "ref" in normalized
        has_ref_cipta = any(k in normalized for k in ("codigoarticuloitemcode", "codigoarticulo", "codigo", "itemcode"))
        has_und = any(k in normalized for k in ("und", "unds", "unidades", "cantidadquantity", "cantidad", "quantity"))
        if (has_ref_futon or has_ref_heimei or has_ref_cipta) and has_und:
            header_row = row_idx
            for i, name in enumerate(normalized):
                if name:
                    header_map[name] = i + 1
            break
    return header_row, header_map


def _m3_desde_medida_texto(medida):
    """Intenta convertir medidas tipo 80 * 200 * 5,5 cm a m3.

    Se usa solo como respaldo visual. Para Hemei/Tatamis el cálculo toma
    como fuente principal el M3 del inventario local.
    """
    texto = str(medida or "").lower().replace(",", ".")
    nums = re.findall(r"\d+(?:\.\d+)?", texto)
    if len(nums) >= 3:
        a, l, h = [float(x) for x in nums[:3]]
        if a > 0 and l > 0 and h > 0:
            return round((a * l * h) / 1_000_000, 6)
    return 0


def cargar_pedido_desde_excel(ruta):
    wb = load_workbook_silencioso(ruta, data_only=True)
    ws = wb.active

    header_row, header_map = _detectar_header_pedido(ws)
    if header_row is None:
        if PROVEEDOR_KEY == "hemei":
            raise ValueError("No se encontró una fila de cabecera con REF y UND.")
        raise ValueError("No se encontró una fila de cabecera con Referencia y Und.")

    def col(*names):
        for name in names:
            key = _normalizar_header(name)
            if key in header_map:
                return header_map[key]
        return None

    if PROVEEDOR_KEY == "hemei":
        c_ref = col("REF", "Referencia")
        c_comp = col("DESCRIPCIÓN", "DESCRIPCION", "Descripcion", "Descripción")
        c_medida = col("MEDIDA", "Medida")
        c_und = col("UND.", "UND", "Unidades")
        c_color = col("Color")
        c_precio_excel = col("Precio Compra unidad", "Precio compra unidad", "Precio Compra", "Precio")
        c_m3u = col("M3/Und.", "M3Und")
        c_m3t = col("M3/total.", "M3total")
    elif PROVEEDOR_KEY == "cipta":
        c_ref = col("CódigoArtículo (Item Code)", "CodigoArticulo (Item Code)", "CódigoArtículo", "CodigoArticulo", "Item Code", "Código", "Codigo", "REF")
        c_comp = col("Modelo (Model)", "Modelo", "Model", "Familia (Family)", "Familia")
        c_medida = col("Medida (Size)", "Medida", "Size")
        c_color = col("Color (Color)", "Color")
        c_und = col("Cantidad (Quantity)", "Cantidad", "Quantity", "UND", "Unidades")
        c_precio_excel = col("Precio Compra unidad", "Precio compra unidad", "Precio Compra", "Precio")
        c_m3u = col("m3/unit", "M3/Und.", "M3Und")
        c_m3t = col("m3/total", "M3/total.", "M3total")
    else:
        c_ref = col("Referencia")
        c_medida = col("Medida")
        c_comp = col("Composición", "Composicion")
        c_color = col("Color")
        c_und = col("Und.", "Und")
        c_precio_excel = None
        c_m3u = col("M3/Und.", "M3Und")
        c_m3t = col("M3/total.", "M3total")

    if not c_ref or not c_und:
        raise ValueError("El pedido debe tener al menos referencia/REF y Und.")

    lineas = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        ref = ws.cell(row_idx, c_ref).value
        unidades_raw = ws.cell(row_idx, c_und).value
        unidades = parse_units_value(unidades_raw, 0)

        # Ignora fila total, filas vacías y productos sin unidades.
        if ref in (None, ""):
            continue
        if "total" in normalizar_texto_busqueda(ref):
            continue
        if unidades <= 0:
            continue

        referencia = normalizar_referencia(ref)
        medida = str(ws.cell(row_idx, c_medida).value or "").strip() if c_medida else ""
        comp = str(ws.cell(row_idx, c_comp).value or "").strip() if c_comp else ""
        color = str(ws.cell(row_idx, c_color).value or "").strip() if c_color else ""
        precio_excel = safe_float(ws.cell(row_idx, c_precio_excel).value, 0) if c_precio_excel else 0
        m3_und = safe_float(ws.cell(row_idx, c_m3u).value, 0) if c_m3u else 0
        m3_total = safe_float(ws.cell(row_idx, c_m3t).value, 0) if c_m3t else 0

        if CALCULO_TIPO == "tatamis" and m3_und <= 0:
            item = buscar_articulo_por_referencia(referencia)
            if item is not None and safe_float(item[3], 0) > 0:
                m3_und = safe_float(item[3], 0)
            else:
                m3_und = _m3_desde_medida_texto(medida)

        if m3_total == 0 and m3_und:
            m3_total = round(m3_und * unidades, 6)

        # Cipta incluye líneas de componentes del mismo artículo (patas, lamas)
        # con cantidad. Si se calculan como artículos separados, duplican el
        # coste del producto principal. Las omitimos del pedido calculable y
        # nos quedamos con las líneas principales del artículo/pack.
        if PROVEEDOR_KEY == "cipta":
            comp_norm = normalizar_texto_busqueda(comp)
            if m3_und <= 0 and m3_total <= 0:
                continue
            if any(token in comp_norm for token in ("bed leg", "leg", "slat", "slats", "lamas", "patas")):
                continue

        producto = " · ".join([x for x in (comp, medida) if x]) or referencia
        linea = {
            "referencia": referencia,
            "codigo": codigo_desde_referencia(referencia),
            "medida": medida,
            "composicion": comp,
            "producto": producto,
            "color": color,
            "unidades": unidades,
            "unidades_raw": unidades_raw,
            "m3_und": m3_und,
            "m3_total": m3_total,
            "precio_excel": precio_excel,
            "fila_excel": row_idx,
        }
        # Aplicar la regla centralizada de reparto: excluye fundas/topper/pillows/almohadas,
        # salvo referencias definidas como excepción porque sí cuentan para descarga.
        linea["cuenta_reparto_descarga"] = cuenta_para_reparto_descarga(linea)
        lineas.append(linea)

    return lineas




def cargar_pedido_desde_pdf(ruta):
    """Lee pedidos PDF simples de Hemei.

    Para otros proveedores seguimos recomendando Excel porque el PDF puede
    cambiar de maquetación y perder estructura de columnas.
    """
    if PROVEEDOR_KEY != "hemei":
        raise ValueError("La lectura de PDF está preparada de momento para pedidos Hemei. Para este proveedor usa Excel.")
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ValueError(
            "Para leer pedidos PDF instala pypdf: pip install pypdf. "
            "Si prefieres evitar dependencias, usa el Excel del pedido."
        ) from exc

    reader = PdfReader(ruta)
    texto = "\n".join(page.extract_text() or "" for page in reader.pages)
    if not texto.strip():
        raise ValueError("No se pudo extraer texto del PDF. Usa el Excel del pedido o revisa que el PDF no sea una imagen escaneada.")

    patron = re.compile(
        r"(?m)^\s*(\d{6,8})\s+(.+?)\s+"
        r"((?:\d+(?:[,.]\d+)?\s*[xX*×]\s*){2}\d+(?:[,.]\d+)?\s*cm)\s+"
        r"(\d+(?:[,.]\d+)?)\s*(?:Pc|pc|PCS|Pcs)?\s*$"
    )
    lineas = []
    for match in patron.finditer(texto):
        ref_raw, descripcion, medida, unidades_raw = match.groups()
        referencia = normalizar_referencia(ref_raw)
        unidades = parse_units_value(unidades_raw, 0)
        if unidades <= 0:
            continue
        m3_und = 0
        item = buscar_articulo_por_referencia(referencia)
        if item is not None and safe_float(item[3], 0) > 0:
            m3_und = safe_float(item[3], 0)
        if m3_und <= 0:
            m3_und = _m3_desde_medida_texto(medida)
        m3_total = round(m3_und * unidades, 6) if m3_und else 0
        producto = " · ".join([x for x in (descripcion.strip(), medida.strip()) if x]) or referencia
        linea_pdf = {
            "referencia": referencia,
            "codigo": codigo_desde_referencia(referencia),
            "medida": medida.strip(),
            "composicion": descripcion.strip(),
            "producto": producto,
            "color": "",
            "unidades": unidades,
            "unidades_raw": unidades_raw,
            "m3_und": m3_und,
            "m3_total": m3_total,
            "precio_excel": 0,
            "fila_excel": 0,
        }
        linea_pdf["cuenta_reparto_descarga"] = cuenta_para_reparto_descarga(linea_pdf)
        lineas.append(linea_pdf)

    if not lineas:
        raise ValueError("No se encontraron líneas de pedido en el PDF. Usa el Excel o revisa el formato del PDF.")
    return lineas


# ---------------------------------------------------------------------------
# Acciones principales
# ---------------------------------------------------------------------------
def seleccionar_pedido():
    global pedido_lineas, archivo_pedido_actual, resultados_exportar

    ruta = filedialog.askopenfilename(
        title="Seleccionar pedido",
        filetypes=[("Pedidos", "*.xlsx *.xlsm *.pdf"), ("Excel", "*.xlsx *.xlsm"), ("PDF", "*.pdf"), ("Todos los archivos", "*.*")],
    )
    if not ruta:
        return

    try:
        if str(ruta).lower().endswith(".pdf"):
            lineas = cargar_pedido_desde_pdf(ruta)
        else:
            lineas = cargar_pedido_desde_excel(ruta)
    except Exception as exc:
        messagebox.showerror("No se pudo cargar el pedido", str(exc))
        return

    pedido_lineas = lineas
    resultados_exportar = []
    archivo_pedido_actual = ruta
    lbl_archivo.config(text=os.path.basename(ruta), fg=C_LBL)
    lbl_estado_calculo.config(text="Pedido cargado. Listo para calcular.", fg=C_MUTED)
    pintar_pedido_en_tabla()

    if not pedido_lineas:
        messagebox.showinfo("Pedido vacío", "No se encontraron líneas con unidades mayores que 0.")


def btn_calcular():
    global resultados_exportar

    reset_input_errors()
    errors = []
    if not pedido_lineas:
        messagebox.showinfo("Sin pedido", "Primero carga un Excel de pedido.")
        return

    if CALCULO_TIPO == "futones":
        coste_transporte_iva = parse_required_float(entry_coste_transporte_iva, "Coste total transporte + IVA (€)", errors)
        coste_total_descarga_iva = parse_required_float(entry_coste_total_descarga_iva, "Coste total descarga + IVA (€)", errors)
        m3_total_camion = sum(safe_float(linea.get("m3_total"), 0) for linea in pedido_lineas)
        cantidad_total_productos = total_unidades_para_reparto(pedido_lineas)
        if m3_total_camion <= 0:
            errors.append("M3 total del pedido: debe ser mayor que 0. Revisa el Excel cargado.")
        if cantidad_total_productos <= 0:
            errors.append("Cantidad total de productos para reparto de descarga: debe ser mayor que 0. Revisa el Excel cargado.")
        precio_dolares = precio_euros = factura_transporte = derechos_aranceles = None
    else:
        precio_dolares = parse_required_float(entry_precio_dolares, "Precio en Dólares del pedido", errors, allow_zero=False)
        precio_euros = parse_required_float(entry_precio_euros, "Precio pagado en Euros", errors, allow_zero=False)
        factura_transporte = parse_required_float(entry_factura_transporte, "Factura Transporte (€)", errors)
        derechos_aranceles = parse_required_float(entry_derechos_aranceles, "Derechos de Aranceles (€)", errors)
        coste_transporte_iva = coste_total_descarga_iva = None
        m3_total_camion = sum(safe_float(linea.get("m3_total"), 0) for linea in pedido_lineas)
        cantidad_total_productos = total_unidades_para_reparto(pedido_lineas)
        if cantidad_total_productos <= 0:
            errors.append("Cantidad total de productos para reparto de descarga: debe ser mayor que 0. Revisa qué líneas cuentan para descarga.")

    rent_modo = selector_rent_modo.current()
    raw_rent = entry_rent_entrada.get().strip() if rentabilidad_var.get() else ""
    valor_rent = None
    if rentabilidad_var.get() and raw_rent:
        try:
            valor_rent = float(raw_rent.replace(",", "."))
        except ValueError:
            mark_input_error(entry_rent_entrada)
            errors.append(f"{lbl_rent_entrada.cget('text')}: debe ser un valor numérico.")

    rent_ok, rent_error = validar_rentabilidad(valor_rent, rent_modo)
    if not rent_ok:
        mark_input_error(entry_rent_entrada)
        errors.append(rent_error)

    if show_input_errors(errors):
        return

    proveedor_key, proveedor_nombre = proveedor_actual_info()

    resultados_exportar = []
    ok_count = 0
    error_count = 0
    m3_update_count = 0
    pending_m3_updates = {}
    pending_price_updates = {}

    try:
        btn_calcular_widget.config(state="disabled")
    except Exception:
        pass
    lbl_estado_calculo.config(text="Calculando pedido... No cierres la ventana.", fg=C_MUTED)
    ventana_principal.update_idletasks()

    for idx, linea in enumerate(pedido_lineas):
        item = buscar_articulo_por_referencia(linea["referencia"])
        tag = "ok"
        estado = "Calculado"
        precio_raw = None
        res = None
        precio_venta = None
        venta_total = None
        rentabilidad = None
        m3_anterior = safe_float(item[3]) if item is not None else None
        m3_actualizado = False

        if item is None:
            if not referencia_parece_codigo(linea["referencia"]):
                estado = "Falta inventario: pedido a medida"
            else:
                estado = "Falta inventario: referencia no encontrada"
            tag = "warn"
            error_count += 1
        else:
            precio_raw = precio_proveedor_item(item, proveedor_key)
            # Algunos pedidos de Tatamis/Maderas traen el precio compra unidad; se usa
            # como respaldo para revisar/cargar pedidos aunque el inventario aún no tenga precio_1.
            if CALCULO_TIPO == "tatamis" and not precio_disponible(precio_raw) and safe_float(linea.get("precio_excel"), 0) > 0:
                precio_raw = safe_float(linea.get("precio_excel"), 0)
                _registrar_precio_proveedor_memoria(item[1], proveedor_key, precio_raw)
                estado = "Calculado · precio desde pedido"

            if not precio_disponible(precio_raw):
                estado = f"Falta precio {proveedor_nombre}"
                tag = "warn"
                error_count += 1
            else:
                precio_articulo = safe_float(precio_raw)
                unidades = safe_float(linea["unidades"])

                if CALCULO_TIPO == "futones":
                    m3_excel = safe_float(linea.get("m3_und"), 0)
                    m3_actualizado, m3_anterior = preparar_m3_inventario(linea["referencia"], m3_excel, pending_m3_updates)
                    res = calcular_coste_unitario_pedido(
                        coste_transporte_iva,
                        coste_total_descarga_iva,
                        m3_total_camion,
                        cantidad_total_productos,
                        precio_articulo,
                        item,
                        m3_excel=m3_excel,
                        unidades_referencia=unidades,
                    )
                else:
                    res = calcular_coste_unitario_tatamis_pedido(
                        precio_dolares,
                        precio_euros,
                        factura_transporte,
                        derechos_aranceles,
                        precio_articulo,
                        item,
                    )

                if res is None:
                    estado = "Error de cálculo"
                    tag = "warn"
                    error_count += 1
                else:
                    coste = res["precio_coste_final"]
                    if valor_rent is not None:
                        if rent_modo == 0:
                            precio_venta = valor_rent
                            rentabilidad = round((valor_rent - coste) / valor_rent * 100, 2)
                        else:
                            rentabilidad = valor_rent
                            precio_venta = round(coste / (1 - valor_rent / 100), 2)
                    coste_total = round(coste * unidades, 2)
                    venta_total = round(precio_venta * unidades, 2) if precio_venta is not None else None
                    resultado = {
                        **res,
                        **linea,
                        "tipo": "Pedido Futones" if CALCULO_TIPO == "futones" else f"Pedido {PROVEEDOR_CFG['nombre']} · Maderas/Tatamis",
                        "proveedor": proveedor_nombre,
                        "precio_proveedor": precio_articulo,
                        "m3_inventario_anterior": m3_anterior,
                        "m3_inventario_actualizado": m3_actualizado,
                        "cuenta_reparto_descarga": linea.get("cuenta_reparto_descarga", cuenta_para_reparto_descarga(linea)),
                        "precio_venta": precio_venta,
                        "rentabilidad": rentabilidad,
                        "coste_total_linea": coste_total,
                        "venta_total_linea": venta_total,
                    }
                    preparar_precio_calculado_pedido(linea["referencia"], coste, pending_price_updates)
                    resultados_exportar.append(resultado)
                    ok_count += 1
                    if m3_actualizado:
                        m3_update_count += 1

        precio_txt = format_money(precio_raw) if precio_disponible(precio_raw) else "—"
        coste_unit = format_money(res["precio_coste_final"]) if res else "—"
        coste_total = format_money(round(res["precio_coste_final"] * safe_float(linea["unidades"]), 2)) if res else "—"
        venta = format_money(precio_venta) if precio_venta is not None else "—"
        rent = f"{rentabilidad:.2f} %" if rentabilidad is not None else "—"
        producto = linea["producto"]
        if item is not None and item[2]:
            producto = f"{item[2]} | {linea['medida']}"

        display_tag = f"{tag}_alt" if idx % 2 else tag

        if CALCULO_TIPO == "futones":
            values = (
                linea["referencia"], producto, linea["color"], format_num(linea["unidades"], 0),
                valor_visible_cuenta_reparto_descarga(linea),
                format_num(linea["m3_und"], 6), format_num(linea["m3_total"], 3),
                format_num(m3_anterior, 6) if m3_anterior is not None else "—",
                estado + (" · M3 actualizado" if m3_actualizado else ""), precio_txt,
                format_money(res["coste_transporte_iva"]) if res else "—",
                format_num(res["m3_total_camion"], 3) if res else "—",
                format_money(res["ct_m3"]) if res else "—",
                format_money(res["ct_m3_prod"]) if res else "—",
                format_money(res["ct_total_ref"]) if res else "—",
                format_money(res["coste_total_descarga_iva"]) if res else "—",
                format_money(res["cd_prod_iva"]) if res else "—",
                format_money(res["cd_total_ref"]) if res else "—",
                format_money(res["iva_re"]) if res else "—",
                format_money(res["precio_con_iva"]) if res else "—",
                format_money(res["coste_descarga"]) if res else "—",
                format_money(res["coste_almacenaje_iva"]) if res else "—",
                format_money(res["coste_picking_iva"]) if res else "—",
                coste_unit, coste_total, venta,
                format_money(venta_total) if venta_total is not None else "—", rent,
            )
        else:
            values = (
                linea["referencia"], producto, linea["color"], format_num(linea["unidades"], 0),
                valor_visible_cuenta_reparto_descarga(linea), format_num(linea.get("m3_und"), 6), format_num(linea.get("m3_total"), 3),
                format_num(m3_anterior, 6) if m3_anterior is not None else "—", estado, precio_txt,
                format_money(res["precio_dolares"]) if res else "—",
                format_money(res["precio_euros"]) if res else "—",
                format_num(res["tasa_cambio"], 6) if res else "—",
                format_money(res["precio_euros_art"]) if res else "—",
                format_money(res["factura_transporte"]) if res else "—",
                format_money(res["derechos_aranceles"]) if res else "—",
                f"{res['pc_transporte']:.2f} %" if res else "—",
                f"{res['pc_descarga']:.2f} %" if res else "—",
                f"{res['pc_varios']:.2f} %" if res else "—",
                f"{res['pc_manipulacion']:.2f} %" if res else "—",
                f"{res['pc_financiacion']:.2f} %" if res else "—",
                format_money(res["coste_almacenaje_iva"]) if res else "—",
                format_money(res["coste_picking_iva"]) if res else "—",
                coste_unit, coste_total, venta,
                format_money(venta_total) if venta_total is not None else "—", rent,
            )
        tree.item(str(idx), values=values, tags=(display_tag,))

    total_coste = sum(r["coste_total_linea"] for r in resultados_exportar)
    m3_guardados, precios_guardados = aplicar_actualizaciones_calculo_en_lote(pending_m3_updates, pending_price_updates)
    if m3_guardados:
        m3_update_count = m3_guardados
    excluidas_reparto = sum(safe_float(l.get("unidades"), 0) for l in pedido_lineas if not l.get("cuenta_reparto_descarga", cuenta_para_reparto_descarga(l)))
    extra_reparto = f" · Unidades excluidas reparto: {format_num(excluidas_reparto, 0)}" if excluidas_reparto else ""
    auto_fit_tree_columns(tree)
    actualizar_resumen_tabla(f"Calculados: {ok_count} · Pendientes: {error_count}{extra_reparto} · M3 actualizados: {m3_update_count} · Coste: {format_money(total_coste)}")
    lbl_estado_calculo.config(
        text=f"Calculados {ok_count} artículos. Pendientes {error_count}." + (f" M3 actualizados en inventario: {m3_update_count}." if CALCULO_TIPO == "futones" else ""),
        fg=C_OK if error_count == 0 else C_ERR,
    )
    try:
        btn_calcular_widget.config(state="normal")
    except Exception:
        pass
    if ok_count > 0:
        ventana_principal.after(150, preguntar_registrar_pedido_proveedor)


def btn_exportar():
    if not resultados_exportar:
        messagebox.showinfo("Sin datos", "No hay artículos calculados para exportar.")
        return

    ruta = filedialog.asksaveasfilename(
        title="Guardar como",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
        initialfile=f"coste_pedido_{PROVEEDOR_CFG['nombre'].lower()}_futonspai.xlsx",
    )
    if not ruta:
        return

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coste Pedido"

    if CALCULO_TIPO == "tatamis":
        headers = [
            "Referencia", "Código", "Denominación", "Medida", "Descripción", "Color",
            "Unidades", "Cuenta para reparto descarga", "M3 Inventario", "Proveedor", "Precio proveedor",
            "Precio en dólares", "Precio pagado en euros", "Tasa cambio",
            "Precio artículo €", "Factura transporte", "Derechos aranceles",
            "% Transporte", "% Descarga", "% Varios", "% Manipulación", "% Financiación", "% Suma gastos",
            "Gastos aplicables", "Coste sin almacenaje", "Coste almacenaje + IVA",
            "Coste unitario picking + IVA", "Precio coste final", "Coste total linea",
            "Precio venta unit.", "Venta total linea", "Rentabilidad (%)",
        ]
    else:
        headers = [
            "Referencia", "Código", "Denominación", "Medida", "Composición", "Color",
            "Unidades", "Cuenta para reparto descarga", "M3/Und. Pedido", "M3 Total Pedido", "M3 Inventario Anterior",
            "M3 Inventario Actualizado", "Proveedor", "Precio proveedor",
            "Coste total transporte + IVA", "M3 camion", "Cantidad total productos para descarga",
            "Coste transporte por M3", "Coste transporte por M3 de cada producto",
            "Coste transporte total por referencia", "Coste total descarga + IVA",
            "Coste descarga por producto + IVA", "Coste descarga total por referencia",
            "Importe IVA + RE", "Precio compra con IVA + RE", "Coste final con descarga",
            "Coste almacenaje + IVA", "Coste unitario picking + IVA", "Precio coste final",
            "Coste total linea", "Precio venta unit.", "Venta total linea", "Rentabilidad (%)",
        ]
    ws.append(headers)

    for res in resultados_exportar:
        if CALCULO_TIPO == "tatamis":
            ws.append([
                res["referencia"], res["codigo"], res["denominacion"], res["medida"],
                res["composicion"], res["color"], res["unidades"],
                "Sí" if res.get("cuenta_reparto_descarga", True) else "No",
                res["m3_bd"], res["proveedor"], res["precio_proveedor"], res["precio_dolares"], res["precio_euros"],
                res["tasa_cambio"], res["precio_euros_art"], res["factura_transporte"],
                res["derechos_aranceles"], res["pc_transporte"], res["pc_descarga"], res["pc_varios"],
                res["pc_manipulacion"], res["pc_financiacion"], res["pc_suma"], res["gastos_aplicables"],
                res["coste_sin_almacenaje"], res["coste_almacenaje_iva"], res["coste_picking_iva"],
                res["precio_coste_final"], res["coste_total_linea"], res["precio_venta"],
                res["venta_total_linea"], res["rentabilidad"],
            ])
        else:
            ws.append([
                res["referencia"], res["codigo"], res["denominacion"], res["medida"],
                res["composicion"], res["color"], res["unidades"],
                "Sí" if res.get("cuenta_reparto_descarga", True) else "No",
                res["m3_und"], res["m3_total"], res["m3_inventario_anterior"],
                "Sí" if res.get("m3_inventario_actualizado") else "No",
                res["proveedor"], res["precio_proveedor"],
                res["coste_transporte_iva"], res["m3_total_camion"], res["cantidad_total_productos"],
                res["ct_m3"], res["ct_m3_prod"], res["ct_total_ref"],
                res["coste_total_descarga_iva"], res["cd_prod_iva"], res["cd_total_ref"], res["iva_re"],
                res["precio_con_iva"], res["coste_descarga"], res["coste_almacenaje_iva"],
                res["coste_picking_iva"], res["precio_coste_final"],
                res["coste_total_linea"], res["precio_venta"], res["venta_total_linea"],
                res["rentabilidad"],
            ])

    fill_hdr = PatternFill("solid", fgColor="2563EB")
    fill_alt = PatternFill("solid", fgColor="F8FAFC")
    font_hdr = Font(color="FFFFFF", bold=True, name="Segoe UI", size=10)
    font_body = Font(color="1F2937", name="Segoe UI", size=10)
    border = Border(bottom=Side(style="thin", color="E5E7EB"))

    for cell in ws[1]:
        cell.font = font_hdr
        cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = font_body
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row % 2 == 0:
                cell.fill = fill_alt

    money_headers = {
        "Precio proveedor", "Precio en dólares", "Precio pagado en euros", "Precio artículo €",
        "Factura transporte", "Derechos aranceles", "Gastos aplicables", "Coste sin almacenaje",
        "Coste total transporte + IVA", "Coste transporte por M3",
        "Coste transporte por M3 de cada producto", "Coste transporte total por referencia",
        "Coste total descarga + IVA", "Coste descarga por producto + IVA",
        "Coste descarga total por referencia", "Importe IVA + RE", "Precio compra con IVA + RE",
        "Coste final con descarga", "Coste almacenaje + IVA", "Coste unitario picking + IVA",
        "Precio coste final", "Coste total linea", "Precio venta unit.", "Venta total linea",
    }
    percent_headers = {"Rentabilidad (%)", "% Transporte", "% Descarga", "% Varios", "% Manipulación", "% Financiación", "% Suma gastos"}
    for col_idx, header in enumerate(headers, start=1):
        number_format = None
        if header in money_headers:
            number_format = '#,##0.00 €'
        elif header in percent_headers:
            number_format = '0.00'
        elif header.startswith("M3"):
            number_format = '0.000000'
        if number_format:
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for c in cell:
                    c.number_format = number_format

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 38)

    resumen = wb.create_sheet("Resumen")
    resumen_rows = [
        ["Archivo pedido", os.path.basename(archivo_pedido_actual or "")],
        ["Proveedor usado", PROVEEDOR_CFG["nombre"]],
        ["Tipo cálculo", "Tatamis / Maderas" if CALCULO_TIPO == "tatamis" else "Futones"],
        ["Productos calculados", len(resultados_exportar)],
        ["Unidades", sum(r["unidades"] for r in resultados_exportar)],
        ["M3 total pedido", sum(r.get("m3_total", 0) for r in resultados_exportar)],
        ["M3 actualizados en inventario", sum(1 for r in resultados_exportar if r.get("m3_inventario_actualizado"))],
        ["Coste total pedido", sum(r["coste_total_linea"] for r in resultados_exportar)],
        ["Venta total", sum((r["venta_total_linea"] or 0) for r in resultados_exportar)],
    ]
    if CALCULO_TIPO == "futones":
        resumen_rows.insert(5, ["Unidades para reparto descarga", total_unidades_para_reparto(resultados_exportar)])

    resumen.append(["Resumen", "Valor"])
    for row in resumen_rows:
        resumen.append(row)
    for cell in resumen[1]:
        cell.font = font_hdr
        cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center")
    for row in resumen.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    resumen.column_dimensions["A"].width = 30
    resumen.column_dimensions["B"].width = 32

    try:
        wb.save(ruta)
        messagebox.showinfo("Exportado", f"Archivo guardado en:\n{ruta}")
    except Exception as exc:
        messagebox.showerror("Error al guardar", str(exc))


# ---------------------------------------------------------------------------
# Widgets panel izquierdo
# ---------------------------------------------------------------------------
tk.Label(
    frame_izquierdo,
    text=f"Calculo de Coste de Pedido · {PROVEEDOR_CFG['nombre']}",
    bg=C_PANEL,
    fg=C_LBL,
    font=("Segoe UI", 13, "bold"),
    anchor="w",
).pack(fill="x", pady=(0, 12))

make_button(
    frame_izquierdo,
    "Cargar Pedido",
    seleccionar_pedido,
    C_BTN,
    C_BTN_HOVER,
    C_BTN_ACTIVE,
    pady=8,
).pack(fill="x", pady=(0, 10))

frame_info = tk.Frame(frame_izquierdo, bg=C_INFO, bd=0)
frame_info.pack(fill="x", pady=(0, 12))
tk.Frame(frame_info, bg=C_BTN, height=3).pack(fill="x")
lbl_archivo = tk.Label(frame_info, text="Ningún archivo seleccionado", bg=C_INFO, fg=C_GREY, font=FONT_LBL_B, anchor="w", wraplength=260)
lbl_archivo.pack(fill="x", padx=8, pady=(6, 2))
lbl_contador_productos = tk.Label(frame_info, text="Productos: 0", bg=C_INFO, fg=C_MUTED, font=FONT_LBL, anchor="w")
lbl_contador_productos.pack(fill="x", padx=8)
lbl_contador_unidades = tk.Label(frame_info, text="Unidades: 0", bg=C_INFO, fg=C_MUTED, font=FONT_LBL, anchor="w")
lbl_contador_unidades.pack(fill="x", padx=8)
lbl_contador_unidades_reparto = tk.Label(frame_info, text="Unidades reparto descarga: 0", bg=C_INFO, fg=C_MUTED, font=FONT_LBL, anchor="w")
lbl_contador_unidades_reparto.pack(fill="x", padx=8)
lbl_contador_m3 = tk.Label(frame_info, text="M3 total: 0", bg=C_INFO, fg=C_MUTED, font=FONT_LBL, anchor="w")
lbl_contador_m3.pack(fill="x", padx=8, pady=(0, 6))

frame_proveedor_info = tk.Frame(frame_izquierdo, bg=C_INFO, bd=0)
frame_proveedor_info.pack(fill="x", pady=(4, 10))
tk.Frame(frame_proveedor_info, bg=C_EXPORT, height=3).pack(fill="x")
tk.Label(
    frame_proveedor_info,
    text=f"Proveedor: {PROVEEDOR_CFG['nombre']}",
    bg=C_INFO,
    fg=C_LBL,
    font=FONT_LBL_B,
    anchor="w",
).pack(fill="x", padx=8, pady=(6, 1))
tk.Label(
    frame_proveedor_info,
    text=f"Precio usado: {PROVEEDOR_CFG['precio_label']}",
    bg=C_INFO,
    fg=C_MUTED,
    font=FONT_LBL,
    anchor="w",
    wraplength=260,
).pack(fill="x", padx=8, pady=(0, 6))

rentabilidad_var = tk.BooleanVar(value=False)
check_rentabilidad = tk.Checkbutton(
    frame_izquierdo,
    text="Rentabilidad",
    variable=rentabilidad_var,
    bg=C_PANEL,
    fg=C_LBL,
    activebackground=C_PANEL,
    activeforeground=C_LBL,
    selectcolor=C_PANEL,
    font=FONT_LBL_B,
    anchor="w",
    command=lambda: toggle_rentabilidad(),
)
check_rentabilidad.pack(fill="x", pady=(6, 1))

frame_rentabilidad = tk.Frame(frame_izquierdo, bg=C_PANEL)
selector_rent_modo = ttk.Combobox(
    frame_rentabilidad,
    values=["Tengo precio de venta", "Quiero rentabilidad %"],
    state="readonly",
    font=FONT_LBL,
)
selector_rent_modo.current(0)
selector_rent_modo.pack(fill="x", pady=(0, 4))

lbl_rent_entrada = make_label(frame_rentabilidad, "Precio de Venta (€)")
entry_rent_entrada = make_entry(frame_rentabilidad)
lbl_rent_entrada.pack(fill="x", pady=(4, 1))
entry_rent_entrada.pack(fill="x", pady=(0, 6))


def on_rent_modo_change(event=None):
    if selector_rent_modo.current() == 0:
        lbl_rent_entrada.config(text="Precio de Venta (€)")
    else:
        lbl_rent_entrada.config(text="% Rentabilidad deseada")


def toggle_rentabilidad():
    if rentabilidad_var.get():
        frame_rentabilidad.pack(fill="x", pady=(0, 6), after=check_rentabilidad)
    else:
        frame_rentabilidad.pack_forget()
        entry_rent_entrada.delete(0, tk.END)


selector_rent_modo.bind("<<ComboboxSelected>>", on_rent_modo_change)

frame_inputs_futones = tk.Frame(frame_izquierdo, bg=C_PANEL)
frame_inputs_tatamis = tk.Frame(frame_izquierdo, bg=C_PANEL)

make_label(frame_inputs_futones, "Coste total transporte + IVA (€)").pack(fill="x", pady=(8, 1))
entry_coste_transporte_iva = make_entry(frame_inputs_futones)
entry_coste_transporte_iva.pack(fill="x", pady=(0, 8))

make_label(frame_inputs_futones, "Coste total descarga + IVA (€)").pack(fill="x", pady=(6, 1))
entry_coste_total_descarga_iva = make_entry(frame_inputs_futones)
entry_coste_total_descarga_iva.insert(0, str(COSTE_TOTAL_DESCARGA_FUTONES_IVA))
entry_coste_total_descarga_iva.pack(fill="x", pady=(0, 8))

make_label(frame_inputs_tatamis, "Precio en Dólares del pedido").pack(fill="x", pady=(8, 1))
entry_precio_dolares = make_entry(frame_inputs_tatamis)
entry_precio_dolares.pack(fill="x", pady=(0, 8))

make_label(frame_inputs_tatamis, "Precio pagado en Euros").pack(fill="x", pady=(6, 1))
entry_precio_euros = make_entry(frame_inputs_tatamis)
entry_precio_euros.pack(fill="x", pady=(0, 8))

make_label(frame_inputs_tatamis, "Factura Transporte (€)").pack(fill="x", pady=(6, 1))
entry_factura_transporte = make_entry(frame_inputs_tatamis)
entry_factura_transporte.pack(fill="x", pady=(0, 8))

make_label(frame_inputs_tatamis, "Derechos de Aranceles (€)").pack(fill="x", pady=(6, 1))
entry_derechos_aranceles = make_entry(frame_inputs_tatamis)
entry_derechos_aranceles.pack(fill="x", pady=(0, 8))

if CALCULO_TIPO == "tatamis":
    frame_inputs_tatamis.pack(fill="x")
else:
    frame_inputs_futones.pack(fill="x")

lbl_estado_calculo = tk.Label(
    frame_izquierdo,
    text="Carga un pedido para empezar.",
    bg=C_PANEL,
    fg=C_MUTED,
    font=FONT_LBL,
    wraplength=260,
    justify="left",
    anchor="w",
)
lbl_estado_calculo.pack(fill="x", pady=(8, 4))

ALL_INPUT_ENTRIES = (
    entry_coste_transporte_iva,
    entry_coste_total_descarga_iva,
    entry_precio_dolares,
    entry_precio_euros,
    entry_factura_transporte,
    entry_derechos_aranceles,
    entry_rent_entrada,
)

boton_container = tk.Frame(left_shell, bg=C_PANEL, padx=14, pady=10)
boton_container.pack(side="bottom", fill="x")

btn_calcular_widget = make_button(
    boton_container,
    "Calcular",
    btn_calcular,
    C_BTN,
    C_BTN_HOVER,
    C_BTN_ACTIVE,
    pady=8,
)
btn_calcular_widget.pack(fill="x")

make_button(
    boton_container,
    "Exportar a Excel",
    btn_exportar,
    C_EXPORT,
    C_EXPORT_HOVER,
    C_EXPORT_ACTIVE,
    pady=7,
).pack(fill="x", pady=(8, 0))

actualizar_resumen_tabla()
ventana_principal.mainloop()
