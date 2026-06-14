import json
import os
import sqlite3
import sys
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl


# ---------------------------------------------------------------------------
# Carga de datos desde Inventario SQLite; Excel queda como respaldo temporal
# ---------------------------------------------------------------------------
data = []
SUPPLIER_PRICES_BY_ID = {}
INVENTORY_DB_PATH = None
_this_dir = os.path.dirname(os.path.abspath(__file__))
_inventory_db_candidates = [
    os.environ.get("FUTON_INVENTORY_DB_PATH", ""),
    os.path.join(_this_dir, "..", "GestorWoo", "data", "gestorwoo.sqlite3"),
    os.path.join(os.getcwd(), "..", "GestorWoo", "data", "gestorwoo.sqlite3"),
]

for _db_path in _inventory_db_candidates:
    if not _db_path:
        continue
    try:
        if os.path.exists(_db_path):
            with sqlite3.connect(_db_path) as _conn:
                _conn.execute(
                    """
                    CREATE TABLE IF NOT EXISTS supplier_prices (
                        item_id INTEGER NOT NULL,
                        supplier TEXT NOT NULL,
                        price TEXT,
                        currency TEXT DEFAULT 'EUR',
                        source TEXT,
                        updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        PRIMARY KEY(item_id, supplier)
                    )
                    """
                )
                _rows = _conn.execute(
                    """
                    SELECT
                        item_id,
                        name,
                        cubic_meters,
                        rotation_c,
                        packages,
                        primary_supplier_price,
                        pascal_price
                    FROM inventory_items
                    ORDER BY item_id
                    """
                ).fetchall()
                # Migracion defensiva desde Precio 1/Precio 2 a precios por proveedor.
                for _row in _rows:
                    _item_id, _name, _m3, _rot, _pkg, _p1, _p2 = _row
                    _text = str(_name or "").lower()
                    _primary_supplier = "Hemei" if "tatami" in _text else ("Cipta" if ("cama" in _text or "sofa" in _text or "sofá" in _text) else "Ekomat")
                    if _p1 not in (None, "", "NO ESTA"):
                        _conn.execute(
                            """
                            INSERT INTO supplier_prices (item_id, supplier, price, source, updated_at)
                            VALUES (?, ?, ?, 'Migrado desde Precio 1', CURRENT_TIMESTAMP)
                            ON CONFLICT(item_id, supplier) DO UPDATE SET price = excluded.price, source = excluded.source, updated_at = CURRENT_TIMESTAMP
                            """,
                            (_item_id, _primary_supplier, str(_p1)),
                        )
                    if _p2 not in (None, "", "NO ESTA"):
                        _conn.execute(
                            """
                            INSERT INTO supplier_prices (item_id, supplier, price, source, updated_at)
                            VALUES (?, 'Pascal', ?, 'Migrado desde Precio 2', CURRENT_TIMESTAMP)
                            ON CONFLICT(item_id, supplier) DO UPDATE SET price = excluded.price, source = excluded.source, updated_at = CURRENT_TIMESTAMP
                            """,
                            (_item_id, str(_p2)),
                        )
                _supplier_rows = _conn.execute("SELECT item_id, supplier, price FROM supplier_prices").fetchall()
                _conn.commit()
            for _item_id, _supplier, _price in _supplier_rows:
                SUPPLIER_PRICES_BY_ID.setdefault(int(_item_id), {})[str(_supplier).lower()] = _price
            for _i, _row in enumerate(_rows, start=1):
                data.append([_i, *_row])
            if data:
                INVENTORY_DB_PATH = _db_path
                print(f"Datos cargados desde inventario local: {_db_path}")
                break
    except Exception as exc:
        print(f"Aviso: no se pudo cargar inventario local {_db_path}: {exc}")

if not data:
    _excel_candidates = [
        os.path.join(_this_dir, 'data.xlsx'),
        os.path.join(_this_dir, '..', 'data.xlsx'),
        os.path.join(_this_dir, '..', 'Futon Spai', 'data.xlsx'),
        os.path.join(os.getcwd(), 'data.xlsx'),
        os.path.join(os.getcwd(), 'Futon Spai', 'data.xlsx'),
    ]

    for _path in _excel_candidates:
        try:
            if os.path.exists(_path):
                _wb = openpyxl.load_workbook(_path, data_only=True)
                _sheet = _wb.active
                for _i, _row in enumerate(_sheet.iter_rows(min_row=2, values_only=True), start=1):
                    data.append([_i] + list(_row))
                print(f"Datos cargados desde Excel de respaldo: {_path}")
                break
        except Exception as exc:
            print(f"Aviso: no se pudo cargar Excel {_path}: {exc}")

if not data:
    print("Aviso: no se encontraron datos en inventario local ni en Excel de respaldo.")



def precio_disponible(valor):
    return valor not in (None, "", "NO ESTA")


def precio_proveedor_item(item, proveedor):
    if item is None:
        return None
    key = str(proveedor or "").strip().lower()
    aliases = {"heimei": "hemei"}
    key = aliases.get(key, key)
    try:
        item_id = int(item[1])
    except Exception:
        item_id = None
    if item_id is not None:
        price = SUPPLIER_PRICES_BY_ID.get(item_id, {}).get(key)
        if precio_disponible(price):
            return price
    if key == "pascal":
        return item[7]
    return item[6]


# ---------------------------------------------------------------------------
# Constantes del negocio
# ---------------------------------------------------------------------------
IMPORTE_DESCARGA_MT              = 250
PC_GASTOS_MANIPULACION           = 7
PC_GASTOS_FINANCIACION           = 7
IMPORTES_VARIOS                  = 100
COSTE_TOTAL_DESCARGA_FUTONES_IVA = 302.5
COSTE_DESCARGA_FUTONES_UNIDAD    = 1.69
IVA_RECARGO_EQUIVALENCIA         = 26.20 / 100
COSTE_DIARIO_ALMACENAJE_M3       = 0.3743
PC_PLUS                          = 0   # Reservado


CONSTANTES_NEGOCIO_PATH = os.path.join(_this_dir, "constantes_negocio.json")

_CONSTANTES_NEGOCIO = {
    "IMPORTE_DESCARGA_MT": {
        "label": "Importe descarga Maderas y Tatamis",
        "suffix": "€",
        "kind": "currency",
        "default": IMPORTE_DESCARGA_MT,
    },
    "PC_GASTOS_MANIPULACION": {
        "label": "% gastos manipulacion",
        "suffix": "%",
        "kind": "percent_points",
        "default": PC_GASTOS_MANIPULACION,
    },
    "PC_GASTOS_FINANCIACION": {
        "label": "% gastos financiacion",
        "suffix": "%",
        "kind": "percent_points",
        "default": PC_GASTOS_FINANCIACION,
    },
    "IMPORTES_VARIOS": {
        "label": "Importes varios",
        "suffix": "€",
        "kind": "currency",
        "default": IMPORTES_VARIOS,
    },
    "COSTE_TOTAL_DESCARGA_FUTONES_IVA": {
        "label": "Coste total descarga futones + IVA",
        "suffix": "€",
        "kind": "currency",
        "default": COSTE_TOTAL_DESCARGA_FUTONES_IVA,
    },
    "COSTE_DESCARGA_FUTONES_UNIDAD": {
        "label": "Coste descarga futones por unidad",
        "suffix": "€",
        "kind": "currency",
        "default": COSTE_DESCARGA_FUTONES_UNIDAD,
    },
    "IVA_RECARGO_EQUIVALENCIA": {
        "label": "IVA + recargo equivalencia",
        "suffix": "%",
        "kind": "fraction_percent",
        "default": IVA_RECARGO_EQUIVALENCIA,
    },
    "COSTE_DIARIO_ALMACENAJE_M3": {
        "label": "Coste diario almacenaje por m³",
        "suffix": "€",
        "kind": "currency",
        "default": COSTE_DIARIO_ALMACENAJE_M3,
    },
}


def _format_constante_para_ui(nombre, valor):
    meta = _CONSTANTES_NEGOCIO[nombre]
    if meta["kind"] == "fraction_percent":
        valor = valor * 100
    return f"{valor:.4f}".rstrip("0").rstrip(".")


def _normalizar_valor_constante(nombre, valor_ui):
    meta = _CONSTANTES_NEGOCIO[nombre]
    valor = float(str(valor_ui).strip().replace(",", "."))
    if valor < 0:
        raise ValueError("El valor no puede ser negativo.")
    if meta["kind"] == "fraction_percent":
        valor = valor / 100
    return valor



def registrar_log_seguridad_local(module, action, status="OK", entity_type="", entity_id="", details="", context=None, category="General"):
    """Registra eventos de seguridad desde CalculoCoste. No debe romper el flujo principal."""
    if not INVENTORY_DB_PATH:
        return
    try:
        payload = json.dumps(context or {}, ensure_ascii=False, default=str) if context else ""
        with sqlite3.connect(INVENTORY_DB_PATH, timeout=8.0) as conn:
            conn.execute("PRAGMA busy_timeout = 5000")
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS security_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    created_at TEXT NOT NULL,
                    category TEXT NOT NULL DEFAULT 'General',
                    severity TEXT NOT NULL DEFAULT 'INFO',
                    module TEXT NOT NULL,
                    action TEXT NOT NULL,
                    status TEXT NOT NULL DEFAULT 'OK',
                    entity_type TEXT DEFAULT '',
                    entity_id TEXT DEFAULT '',
                    details TEXT DEFAULT '',
                    context_json TEXT DEFAULT '',
                    source TEXT DEFAULT 'HUB'
                )
                """
            )
            conn.execute(
                """
                INSERT INTO security_logs (
                    created_at, category, severity, module, action, status,
                    entity_type, entity_id, details, context_json, source
                ) VALUES (?, ?, 'INFO', ?, ?, ?, ?, ?, ?, ?, 'CalculoCoste')
                """,
                (
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    category,
                    module,
                    action,
                    status,
                    entity_type,
                    str(entity_id or ""),
                    details,
                    payload,
                ),
            )
            conn.commit()
    except Exception:
        return

def cargar_constantes_negocio():
    global IMPORTE_DESCARGA_MT, PC_GASTOS_MANIPULACION, PC_GASTOS_FINANCIACION
    global IMPORTES_VARIOS, COSTE_TOTAL_DESCARGA_FUTONES_IVA
    global COSTE_DESCARGA_FUTONES_UNIDAD, IVA_RECARGO_EQUIVALENCIA
    global COSTE_DIARIO_ALMACENAJE_M3

    valores = {nombre: meta["default"] for nombre, meta in _CONSTANTES_NEGOCIO.items()}
    if os.path.exists(CONSTANTES_NEGOCIO_PATH):
        try:
            with open(CONSTANTES_NEGOCIO_PATH, "r", encoding="utf-8") as fh:
                guardados = json.load(fh)
            for nombre in valores:
                if nombre in guardados:
                    valores[nombre] = float(guardados[nombre])
        except Exception as exc:
            print(f"Aviso: no se pudieron cargar constantes de negocio: {exc}")

    IMPORTE_DESCARGA_MT = valores["IMPORTE_DESCARGA_MT"]
    PC_GASTOS_MANIPULACION = valores["PC_GASTOS_MANIPULACION"]
    PC_GASTOS_FINANCIACION = valores["PC_GASTOS_FINANCIACION"]
    IMPORTES_VARIOS = valores["IMPORTES_VARIOS"]
    COSTE_TOTAL_DESCARGA_FUTONES_IVA = valores["COSTE_TOTAL_DESCARGA_FUTONES_IVA"]
    COSTE_DESCARGA_FUTONES_UNIDAD = valores["COSTE_DESCARGA_FUTONES_UNIDAD"]
    IVA_RECARGO_EQUIVALENCIA = valores["IVA_RECARGO_EQUIVALENCIA"]
    COSTE_DIARIO_ALMACENAJE_M3 = valores["COSTE_DIARIO_ALMACENAJE_M3"]
    return valores


def guardar_constantes_negocio(valores):
    os.makedirs(os.path.dirname(CONSTANTES_NEGOCIO_PATH), exist_ok=True)
    with open(CONSTANTES_NEGOCIO_PATH, "w", encoding="utf-8") as fh:
        json.dump(valores, fh, ensure_ascii=False, indent=2)
    cargar_constantes_negocio()


def abrir_ventana_constantes():
    valores_actuales = cargar_constantes_negocio()

    ventana = tk.Tk()
    ventana.title("FutonSpai · Constantes del negocio")
    center_window(ventana, 560, 520)
    ventana.minsize(520, 460)
    ventana.configure(bg=C_BG)

    contenedor = tk.Frame(ventana, bg=C_BG, padx=18, pady=18)
    contenedor.pack(fill="both", expand=True)

    panel = tk.Frame(
        contenedor,
        bg=C_PANEL,
        highlightthickness=1,
        highlightbackground=C_PANEL_LINE,
        padx=18,
        pady=16,
    )
    panel.pack(fill="both", expand=True)
    panel.columnconfigure(1, weight=1)

    tk.Label(
        panel,
        text="Constantes del negocio",
        bg=C_PANEL,
        fg=C_LBL,
        font=("Segoe UI", 14, "bold"),
        anchor="w",
    ).grid(row=0, column=0, columnspan=3, sticky="ew")

    tk.Label(
        panel,
        text="Estos valores se guardan al pulsar Aceptar y se usarán en los cálculos siguientes.",
        bg=C_PANEL,
        fg=C_MUTED,
        font=FONT_LBL,
        wraplength=470,
        justify="left",
        anchor="w",
    ).grid(row=1, column=0, columnspan=3, sticky="ew", pady=(4, 16))

    entradas = {}
    for fila, nombre in enumerate(_CONSTANTES_NEGOCIO, start=2):
        meta = _CONSTANTES_NEGOCIO[nombre]
        tk.Label(
            panel,
            text=meta["label"],
            bg=C_PANEL,
            fg=C_LBL,
            font=FONT_LBL,
            anchor="w",
        ).grid(row=fila, column=0, sticky="w", pady=5)

        entry = tk.Entry(
            panel,
            bg=C_ENTRY,
            fg=C_LBL,
            relief="flat",
            highlightthickness=1,
            highlightbackground=C_ENTRY_LINE,
            highlightcolor=C_BTN,
            font=FONT_LBL,
            insertbackground=C_LBL,
        )
        entry.insert(0, _format_constante_para_ui(nombre, valores_actuales[nombre]))
        entry.grid(row=fila, column=1, sticky="ew", padx=(12, 8), pady=5)
        entradas[nombre] = entry

        tk.Label(
            panel,
            text=meta["suffix"],
            bg=C_PANEL,
            fg=C_MUTED,
            font=FONT_LBL,
            anchor="w",
        ).grid(row=fila, column=2, sticky="w", pady=5)

    botones = tk.Frame(contenedor, bg=C_BG)
    botones.pack(fill="x", pady=(14, 0))

    def cancelar():
        ventana.destroy()

    def aceptar():
        nuevos_valores = {}
        errores = []
        for nombre, entry in entradas.items():
            try:
                nuevos_valores[nombre] = _normalizar_valor_constante(nombre, entry.get())
                entry.config(bg=C_ENTRY, highlightbackground=C_ENTRY_LINE, highlightcolor=C_BTN)
            except ValueError as exc:
                entry.config(bg=C_ENTRY_ERR_BG, highlightbackground=C_ERR, highlightcolor=C_ERR)
                errores.append(f"{_CONSTANTES_NEGOCIO[nombre]['label']}: {exc}")

        if errores:
            messagebox.showerror(
                "Revisa los valores",
                "Corrige estos errores:\n\n- " + "\n- ".join(errores),
                parent=ventana,
            )
            return

        cambios_constantes = []
        for nombre, nuevo in nuevos_valores.items():
            anterior = valores_actuales.get(nombre)
            if anterior != nuevo:
                cambios_constantes.append({
                    "constante": nombre,
                    "campo": _CONSTANTES_NEGOCIO[nombre]["label"],
                    "valor_anterior": anterior,
                    "valor_nuevo": nuevo,
                })

        try:
            guardar_constantes_negocio(nuevos_valores)
        except Exception as exc:
            messagebox.showerror("No se pudo guardar", str(exc), parent=ventana)
            return

        if cambios_constantes:
            registrar_log_seguridad_local(
                "Constantes de Cálculo de Coste",
                "Cambiar valores de constantes",
                status="OK",
                entity_type="Constantes",
                entity_id="constantes_negocio.json",
                details=f"Se modificaron {len(cambios_constantes)} constantes del negocio.",
                context={"changes": cambios_constantes},
                category="Constantes de Cálculo de Coste",
            )

        messagebox.showinfo("Constantes actualizadas", "Los valores se guardaron correctamente.", parent=ventana)
        ventana.destroy()

    btn_cancelar = tk.Button(
        botones,
        text="Cancelar",
        command=cancelar,
        bg=C_GREY,
        fg=C_BTN_FG,
        activebackground="#64748B",
        activeforeground=C_BTN_FG,
        relief="flat",
        bd=0,
        padx=18,
        pady=8,
        font=FONT_BTN,
        cursor="hand2",
    )
    btn_cancelar.pack(side="right")

    btn_aceptar = tk.Button(
        botones,
        text="Aceptar",
        command=aceptar,
        bg=C_BTN,
        fg=C_BTN_FG,
        activebackground=C_BTN_ACTIVE,
        activeforeground=C_BTN_FG,
        relief="flat",
        bd=0,
        padx=18,
        pady=8,
        font=FONT_BTN,
        cursor="hand2",
    )
    btn_aceptar.pack(side="right", padx=(0, 8))

    ventana.bind("<Escape>", lambda _event: cancelar())
    ventana.mainloop()


cargar_constantes_negocio()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def center_window(window, width, height):
    """Centra una ventana Tk/Toplevel en la pantalla."""
    window.update_idletasks()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = max((screen_width - width) // 2, 0)
    y = max((screen_height - height) // 2, 0)
    window.geometry(f"{width}x{height}+{x}+{y}")


def buscar_articulo(codigo):
    for item in data:
        try:
            if item[1] == codigo:
                return item
        except Exception:
            continue
    return None

def safe_float(v, d=0.0):
    try:    return float(v)
    except: return d

def safe_int(v, d=1):
    try:    return int(v)
    except: return d

def validar_rentabilidad(valor, modo_rentabilidad):
    if valor is None:
        return True, None
    if modo_rentabilidad == 0 and valor <= 0:
        return False, "El precio de venta debe ser mayor que 0."
    if modo_rentabilidad == 1 and valor >= 100:
        return False, "La rentabilidad deseada debe ser menor que 100%."
    return True, None


# ---------------------------------------------------------------------------
# Cálculo Maderas y Tatamis
# ---------------------------------------------------------------------------
def calcular_maderas_tatamis(precio_dolares, precio_euros, factura_transporte,
                              derechos_aranceles, precio_articulo, codigo):
    cargar_constantes_negocio()
    if precio_euros == 0:
        return None
    item = buscar_articulo(codigo)
    if item is None:
        return None

    m3         = safe_float(item[3])
    rotacion_c = safe_float(item[4])
    n_bultos   = safe_int(item[5])

    tasa_cambio        = round(precio_dolares / precio_euros, 6)
    importe_transporte = factura_transporte + derechos_aranceles
    pc_transporte      = round((importe_transporte / precio_euros) * 100, 2)
    pc_descarga        = round((IMPORTE_DESCARGA_MT * 100) / precio_euros, 2)
    pc_varios          = round((IMPORTES_VARIOS / precio_euros) * 100, 2)
    pc_suma            = round(pc_transporte + pc_descarga + PC_GASTOS_FINANCIACION
                               + PC_GASTOS_MANIPULACION + pc_varios, 2)

    precio_euros_art      = round(precio_articulo / tasa_cambio, 2)
    gastos_aplicables     = round(precio_euros_art * pc_suma / 100, 2)
    coste_sin_almacenaje  = round(precio_euros_art + gastos_aplicables, 2)
    coste_almacenaje_iva  = round(COSTE_DIARIO_ALMACENAJE_M3 * m3 * rotacion_c * 1.21, 4)
    coste_picking_iva     = round(((n_bultos * 0.3) + 4.12) * 1.21, 3)
    precio_coste_final    = round(coste_sin_almacenaje + coste_almacenaje_iva + coste_picking_iva, 2)

    return {
        "denominacion":          item[2],
        "tasa_cambio":           tasa_cambio,
        "importe_transporte":    importe_transporte,
        "pc_transporte":         pc_transporte,
        "pc_descarga":           pc_descarga,
        "pc_suma":               pc_suma,
        "precio_euros_art":      precio_euros_art,
        "gastos_aplicables":     gastos_aplicables,
        "coste_sin_almacenaje":  coste_sin_almacenaje,
        "coste_almacenaje_iva":  coste_almacenaje_iva,
        "coste_picking_iva":     coste_picking_iva,
        "precio_coste_final":    precio_coste_final,
    }


# ---------------------------------------------------------------------------
# Cálculo Futones
# ---------------------------------------------------------------------------
def calcular_futones(coste_transporte_iva, m3_total_camion, unidades_referencia,
                     cantidad_total_productos, precio_proveedor, codigo):
    cargar_constantes_negocio()
    if m3_total_camion == 0 or cantidad_total_productos == 0:
        return None
    item = buscar_articulo(codigo)
    if item is None:
        return None

    m3         = safe_float(item[3])
    rotacion_c = safe_float(item[4])
    n_bultos   = safe_int(item[5])

    ct_m3          = round(coste_transporte_iva / m3_total_camion, 2)
    ct_m3_prod     = round(ct_m3 * m3, 2)
    ct_total_ref   = round(unidades_referencia * ct_m3_prod, 2)
    cd_prod_iva    = round(COSTE_TOTAL_DESCARGA_FUTONES_IVA / cantidad_total_productos, 2)
    cd_total_ref   = round(unidades_referencia * cd_prod_iva, 3)
    iva_re         = round(precio_proveedor * IVA_RECARGO_EQUIVALENCIA, 2)
    precio_con_iva = round(precio_proveedor + iva_re, 2)
    coste_descarga = round(ct_m3_prod + cd_prod_iva + precio_con_iva, 2)

    coste_almacenaje_iva = round(COSTE_DIARIO_ALMACENAJE_M3 * m3 * rotacion_c * 1.21, 4)
    coste_picking_iva    = round(((n_bultos * 0.3) + 4.12) * 1.21, 3)
    precio_coste_final   = round(coste_descarga + coste_almacenaje_iva + coste_picking_iva, 2)

    return {
        "denominacion":       item[2],
        "ct_m3":              ct_m3,
        "ct_m3_prod":         ct_m3_prod,
        "ct_total_ref":       ct_total_ref,
        "cd_prod_iva":        cd_prod_iva,
        "cd_total_ref":       cd_total_ref,
        "iva_re":             iva_re,
        "precio_con_iva":     precio_con_iva,
        "coste_descarga":     coste_descarga,
        "coste_almacenaje_iva": coste_almacenaje_iva,
        "coste_picking_iva":  coste_picking_iva,
        "precio_coste_final": precio_coste_final,
    }


# ---------------------------------------------------------------------------
# Paleta de colores
# ---------------------------------------------------------------------------
C_BG          = "#F6F7F9"   # fondo general
C_PANEL       = "#FFFFFF"   # panel izquierdo
C_PANEL_LINE  = "#DDE3EA"
C_INFO        = "#F1F5F9"   # panel info artículo
C_ENTRY       = "#FFFFFF"
C_ENTRY_LINE  = "#CBD5E1"
C_ENTRY_ERR_BG = "#FEF2F2"
C_BTN         = "#2563EB"   # botón calcular
C_BTN_HOVER   = "#1D4ED8"
C_BTN_ACTIVE  = "#1E40AF"
C_EXPORT      = "#0F766E"
C_EXPORT_HOVER = "#0D9488"
C_EXPORT_ACTIVE = "#115E59"
C_BTN_FG      = "#FFFFFF"
C_LBL         = "#1F2937"
C_MUTED       = "#64748B"
C_GREY        = "#94A3B8"
C_ERR         = "#B91C1C"
C_OK          = "#15803D"
C_MT_BG       = "#FFFFFF"
C_MT_ROW      = "#F8FAFC"
C_MT_FG       = "#1D4ED8"
C_MT_ACCENT   = "#DBEAFE"
C_FT_BG       = "#FFFFFF"
C_FT_ROW      = "#F8FAFC"
C_FT_FG       = "#047857"
C_FT_ACCENT   = "#D1FAE5"
C_RIGHT       = "#EEF2F7"
FONT_LBL      = ("Segoe UI", 9)
FONT_LBL_B    = ("Segoe UI", 9, "bold")
FONT_TABLE    = ("Segoe UI", 9)
FONT_TABLE_B  = ("Segoe UI", 9, "bold")
FONT_BTN      = ("Segoe UI", 10, "bold")
FONT_INFO_B   = ("Segoe UI", 8, "bold")
FONT_INFO     = ("Segoe UI", 8)


if "--constantes" in sys.argv:
    abrir_ventana_constantes()
    sys.exit(0)


# ---------------------------------------------------------------------------
# Ventana principal
# ---------------------------------------------------------------------------
ventana_principal = tk.Tk()
ventana_principal.title("FutonSpai · Calculadora de Costes")
center_window(ventana_principal, 1020, 740)
ventana_principal.minsize(960, 680)
ventana_principal.configure(bg=C_BG)

# Estilo ttk
style = ttk.Style()
style.theme_use("clam")
style.configure("TCombobox",
                fieldbackground=C_ENTRY,
                background=C_ENTRY,
                foreground=C_LBL,
                bordercolor=C_ENTRY_LINE,
                arrowcolor=C_BTN,
                padding=6)
style.map("TCombobox", fieldbackground=[("readonly", C_ENTRY)])


# ---------------------------------------------------------------------------
# Panel izquierdo
# ---------------------------------------------------------------------------
left_shell = tk.Frame(ventana_principal, width=320, bg=C_PANEL,
                      highlightthickness=1, highlightbackground=C_PANEL_LINE)
left_shell.pack(side="left", fill="y", padx=(12, 6), pady=12)
left_shell.pack_propagate(False)

left_canvas = tk.Canvas(left_shell, bg=C_PANEL, highlightthickness=0)
left_scroll = tk.Scrollbar(left_shell, orient="vertical", command=left_canvas.yview)
left_canvas.configure(yscrollcommand=left_scroll.set)
left_scroll.pack(side="right", fill="y")
left_canvas.pack(side="top", fill="both", expand=True)

frame_izquierdo = tk.Frame(left_canvas, padx=14, pady=14, bg=C_PANEL)
left_window_id = left_canvas.create_window((0, 0), window=frame_izquierdo, anchor="nw")
left_canvas.bind("<Configure>", lambda e: left_canvas.itemconfig(left_window_id, width=e.width))
frame_izquierdo.bind("<Configure>", lambda e: left_canvas.configure(scrollregion=left_canvas.bbox("all")))
left_canvas.bind("<MouseWheel>", lambda e: left_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

# ---------------------------------------------------------------------------
# Panel derecho scrollable
# ---------------------------------------------------------------------------
right_container = tk.Frame(ventana_principal, bg=C_BG)
right_container.pack(side="right", fill="both", expand=True, padx=(6, 12), pady=12)

canvas   = tk.Canvas(right_container, bg=C_RIGHT, highlightthickness=0)
v_scroll = tk.Scrollbar(right_container, orient="vertical", command=canvas.yview)
canvas.configure(yscrollcommand=v_scroll.set)
v_scroll.pack(side="right", fill="y")
canvas.pack(side="left", fill="both", expand=True)

frame_derecho = tk.Frame(canvas, bg=C_RIGHT)
window_id     = canvas.create_window((0, 0), window=frame_derecho, anchor='nw')

canvas.bind('<Configure>',
            lambda e: canvas.itemconfig(window_id, width=e.width))
frame_derecho.bind('<Configure>',
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
canvas.bind_all("<MouseWheel>",
                lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))


# ---------------------------------------------------------------------------
# Helper para crear entries con estilo
# ---------------------------------------------------------------------------
def make_entry(parent):
    entry = tk.Entry(parent, bg=C_ENTRY, fg=C_LBL, relief="flat",
                     highlightthickness=1, highlightbackground=C_ENTRY_LINE,
                     highlightcolor=C_BTN, font=FONT_LBL,
                     insertbackground=C_LBL)
    return entry

def reset_input_errors():
    for entry in globals().get("ALL_INPUT_ENTRIES", ()): 
        entry.config(bg=C_ENTRY, highlightbackground=C_ENTRY_LINE, highlightcolor=C_BTN)

def mark_input_error(entry):
    entry.config(bg=C_ENTRY_ERR_BG, highlightbackground=C_ERR, highlightcolor=C_ERR)

def parse_required_float(entry, label, errors, allow_zero=True):
    raw = entry.get().strip().replace(",", ".")
    if not raw:
        mark_input_error(entry)
        errors.append(f"{label}: campo obligatorio.")
        return None
    try:
        value = float(raw)
    except ValueError:
        mark_input_error(entry)
        errors.append(f"{label}: debe ser un valor numerico.")
        return None
    if not allow_zero and value == 0:
        mark_input_error(entry)
        errors.append(f"{label}: no puede ser 0.")
        return None
    return value

def parse_required_int(entry, label, errors, min_value=None):
    raw = entry.get().strip()
    if not raw:
        mark_input_error(entry)
        errors.append(f"{label}: campo obligatorio.")
        return None
    try:
        value = int(raw)
    except ValueError:
        mark_input_error(entry)
        errors.append(f"{label}: debe ser un numero entero.")
        return None
    if min_value is not None and value < min_value:
        mark_input_error(entry)
        errors.append(f"{label}: debe ser mayor o igual que {min_value}.")
        return None
    return value

def show_input_errors(errors):
    if errors:
        messagebox.showerror(
            "Revisa los campos",
            "Corrige estos errores:\n\n- " + "\n- ".join(errors),
        )
        return True
    return False

def make_label(parent, text):
    return tk.Label(parent, text=text, bg=C_PANEL, fg=C_LBL,
                    font=FONT_LBL, anchor='w')

def make_button(parent, text, command, bg, hover, active, pady=8):
    btn = tk.Button(
        parent, text=text, command=command,
        bg=bg, fg=C_BTN_FG, activebackground=active, activeforeground=C_BTN_FG,
        font=FONT_BTN, relief="flat", cursor="hand2", pady=pady,
        bd=0, highlightthickness=0
    )
    btn.bind("<Enter>", lambda _e: btn.config(bg=hover))
    btn.bind("<Leave>", lambda _e: btn.config(bg=bg))
    return btn


# ---------------------------------------------------------------------------
# Widgets panel izquierdo
# ---------------------------------------------------------------------------

# Selector de modo
selector_calculo = ttk.Combobox(frame_izquierdo,
                                 values=["Maderas y Tatamis", "Futones"],
                                 state="readonly", font=FONT_LBL)
selector_calculo.current(0)
selector_calculo.pack(fill="x", pady=(0, 10))

# Campo código
make_label(frame_izquierdo, "Código del artículo").pack(fill="x")
entry_codigo = make_entry(frame_izquierdo)
entry_codigo.pack(fill="x", pady=(2, 6))

# --- Panel info artículo ---
frame_info = tk.Frame(frame_izquierdo, bg=C_INFO, bd=0)
frame_info.pack(fill="x", pady=(0, 10))

# Borde superior coloreado como acento
tk.Frame(frame_info, bg=C_BTN, height=3).pack(fill="x")

lbl_info_denominacion = tk.Label(
    frame_info, text="—", bg=C_INFO,
    wraplength=230, justify="left", anchor='w',
    font=FONT_INFO_B, fg=C_GREY
)
lbl_info_denominacion.pack(fill="x", padx=8, pady=(6, 2))

lbl_info_precio1 = tk.Label(
    frame_info, text="Precio (CIPTA): —",
    bg=C_INFO, anchor='w', font=FONT_INFO, fg=C_GREY
)
lbl_info_precio1.pack(fill="x", padx=8)

lbl_info_precio2 = tk.Label(
    frame_info, text="Precio 2 (Pascal): —",
    bg=C_INFO, anchor='w', font=FONT_INFO, fg=C_GREY
)
# precio2 se gestiona en _actualizar_panel_info, no se hace pack aquí

tk.Frame(frame_info, bg=C_INFO, height=6).pack(fill="x")  # espaciado inferior


# ---------------------------------------------------------------------------
# _actualizar_panel_info  — adapta etiquetas y visibilidad según modo
# ---------------------------------------------------------------------------
def _actualizar_panel_info():
    modo = selector_calculo.get()
    raw  = entry_codigo.get().strip()

    # Visibilidad de precio2
    if modo == "Maderas y Tatamis":
        lbl_info_precio2.pack_forget()
    else:
        lbl_info_precio2.pack(fill="x", padx=8, after=lbl_info_precio1)

    # Estado vacío
    if not raw:
        lbl_info_denominacion.config(text="—", fg=C_GREY)
        lbl_info_precio1.config(
            text="Precio (CIPTA): —" if modo == "Maderas y Tatamis"
                 else "Precio 1 (Ekomat): —",
            fg=C_GREY
        )
        lbl_info_precio2.config(text="Precio 2 (Pascal): —", fg=C_GREY)
        return

    # Código no numérico
    try:
        codigo = int(raw)
    except ValueError:
        lbl_info_denominacion.config(text="Código no válido", fg=C_ERR)
        lbl_info_precio1.config(
            text="Precio (CIPTA): —" if modo == "Maderas y Tatamis"
                 else "Precio 1 (Ekomat): —",
            fg=C_GREY
        )
        lbl_info_precio2.config(text="Precio 2 (Pascal): —", fg=C_GREY)
        return

    item = buscar_articulo(codigo)

    # No encontrado
    if item is None:
        lbl_info_denominacion.config(text="Artículo no encontrado", fg=C_ERR)
        lbl_info_precio1.config(
            text="Precio (CIPTA): —" if modo == "Maderas y Tatamis"
                 else "Precio 1 (Ekomat): —",
            fg=C_GREY
        )
        lbl_info_precio2.config(text="Precio 2 (Pascal): —", fg=C_GREY)
        return

    # Denominación
    lbl_info_denominacion.config(
        text=str(item[2]) if item[2] else "Sin nombre", fg=C_LBL
    )

    p1, p2 = item[6], item[7]

    if modo == "Maderas y Tatamis":
        if p1 in (None, "NO ESTA"):
            lbl_info_precio1.config(text="Precio (CIPTA): NO ESTÁ EN BD ⚠", fg=C_ERR)
        else:
            lbl_info_precio1.config(text=f"Precio (CIPTA): {p1} $", fg=C_LBL)
    else:
        if p1 in (None, "NO ESTA"):
            lbl_info_precio1.config(text="Precio 1 (Ekomat): NO ESTÁ EN BD ⚠", fg=C_ERR)
        else:
            lbl_info_precio1.config(text=f"Precio 1 (Ekomat): {p1} €", fg=C_LBL)
        if p2 in (None, "NO ESTA"):
            lbl_info_precio2.config(text="Precio 2 (Pascal): NO ESTÁ EN BD ⚠", fg=C_ERR)
        else:
            lbl_info_precio2.config(text=f"Precio 2 (Pascal): {p2} €", fg=C_LBL)


# ---------------------------------------------------------------------------
# Campos Maderas y Tatamis
# ---------------------------------------------------------------------------
lbl_precio_dolares       = make_label(frame_izquierdo, "Precio total en Dólares ($)")
entry_precio_dolares     = make_entry(frame_izquierdo)

lbl_precio_euros         = make_label(frame_izquierdo, "Precio total en Euros (€)")
entry_precio_euros       = make_entry(frame_izquierdo)

lbl_factura_transporte   = make_label(frame_izquierdo, "Factura de Transporte (€)")
entry_factura_transporte = make_entry(frame_izquierdo)

lbl_derechos_aranceles   = make_label(frame_izquierdo, "Derechos de Aranceles (€)")
entry_derechos_aranceles = make_entry(frame_izquierdo)

# ---------------------------------------------------------------------------
# Campos Futones
# ---------------------------------------------------------------------------
lbl_proveedor_f    = make_label(frame_izquierdo, "Proveedor")
selector_proveedor = ttk.Combobox(frame_izquierdo, values=["Ekomat", "Pascal"],
                                   state="readonly", font=FONT_LBL)
selector_proveedor.current(0)

lbl_coste_transporte_f   = make_label(frame_izquierdo, "Coste Transporte + IVA (€)")
entry_coste_transporte_f = make_entry(frame_izquierdo)

lbl_m3_total_camion_f    = make_label(frame_izquierdo, "M³ Total Camión")
entry_m3_total_camion_f  = make_entry(frame_izquierdo)

lbl_unidad               = make_label(frame_izquierdo, "Unidades por referencia")
entry_unidad             = make_entry(frame_izquierdo)

lbl_cantidad_productos   = make_label(frame_izquierdo, "Cantidad total de productos")
entry_cantidad_productos = make_entry(frame_izquierdo)

# ---------------------------------------------------------------------------
# Sección rentabilidad — selector de modo + entrada dinámica
# ---------------------------------------------------------------------------
tk.Frame(frame_izquierdo, bg=C_PANEL_LINE, height=1).pack(fill="x", pady=(12, 0))

selector_rent_modo = ttk.Combobox(
    frame_izquierdo,
    values=["Precio de Venta → % Rentabilidad", "% Rentabilidad → Precio de Venta"],
    state="readonly", font=FONT_LBL,
)
selector_rent_modo.current(0)
selector_rent_modo.pack(fill="x", pady=(8, 2))

lbl_rent_entrada   = make_label(frame_izquierdo, "Precio de Venta (€)")
entry_rent_entrada = make_entry(frame_izquierdo)
lbl_rent_entrada.pack(fill="x", pady=(4, 1))
entry_rent_entrada.pack(fill="x", pady=(0, 4))

ALL_INPUT_ENTRIES = (
    entry_codigo,
    entry_precio_dolares, entry_precio_euros,
    entry_factura_transporte, entry_derechos_aranceles,
    entry_coste_transporte_f, entry_m3_total_camion_f,
    entry_unidad, entry_cantidad_productos,
    entry_rent_entrada,
)

def on_rent_modo_change(event=None):
    if selector_rent_modo.current() == 0:
        lbl_rent_entrada.config(text="Precio de Venta (€)")
    else:
        lbl_rent_entrada.config(text="% Rentabilidad deseada")

selector_rent_modo.bind("<<ComboboxSelected>>", on_rent_modo_change)

# ---------------------------------------------------------------------------
# Show / hide
# ---------------------------------------------------------------------------
_MT_WIDGETS = (
    lbl_precio_dolares, entry_precio_dolares,
    lbl_precio_euros, entry_precio_euros,
    lbl_factura_transporte, entry_factura_transporte,
    lbl_derechos_aranceles, entry_derechos_aranceles,
)
_F_WIDGETS = (
    lbl_proveedor_f, selector_proveedor,
    lbl_coste_transporte_f, entry_coste_transporte_f,
    lbl_m3_total_camion_f, entry_m3_total_camion_f,
    lbl_unidad, entry_unidad,
    lbl_cantidad_productos, entry_cantidad_productos,
)

def _pack(w):
    if isinstance(w, tk.Label):
        w.pack(fill="x", pady=(8, 1))
    else:
        w.pack(fill="x", pady=(0, 2))

def _show_mt_fields():
    for w in _F_WIDGETS:  w.pack_forget()
    for w in _MT_WIDGETS: _pack(w)

def _show_futones_fields():
    for w in _MT_WIDGETS: w.pack_forget()
    for w in _F_WIDGETS:  _pack(w)


# ---------------------------------------------------------------------------
# Eventos
# ---------------------------------------------------------------------------
def on_selector_modo(event=None):
    if selector_calculo.get() == "Maderas y Tatamis":
        _show_mt_fields()
    else:
        _show_futones_fields()
    _actualizar_panel_info()

def on_selector_proveedor(event=None):
    _actualizar_panel_info()

def on_codigo_focusout(event=None):
    _actualizar_panel_info()

selector_calculo.bind("<<ComboboxSelected>>", on_selector_modo)
selector_proveedor.bind("<<ComboboxSelected>>", on_selector_proveedor)
entry_codigo.bind("<FocusOut>", on_codigo_focusout)
entry_codigo.bind("<Return>",   on_codigo_focusout)

# Estado inicial correcto
_show_mt_fields()
_actualizar_panel_info()   # ← fija las etiquetas al arrancar


# ---------------------------------------------------------------------------
# Lista acumuladora para exportar
# ---------------------------------------------------------------------------
resultados_exportar = []   # cada entrada: dict con todos los datos del ítem


# ---------------------------------------------------------------------------
# Botón Calcular
# ---------------------------------------------------------------------------
def btn_calcular():
    modo = selector_calculo.get()
    reset_input_errors()
    errors = []

    raw_codigo = entry_codigo.get().strip()
    if not raw_codigo:
        mark_input_error(entry_codigo)
        errors.append("Codigo del articulo: campo obligatorio.")
        codigo = None
    else:
        try:
            codigo = int(raw_codigo)
        except ValueError:
            mark_input_error(entry_codigo)
            errors.append("Codigo del articulo: debe ser numerico.")
            codigo = None

    if show_input_errors(errors):
        entry_codigo.focus_set()
        return

    item = buscar_articulo(codigo)
    if item is None:
        mark_input_error(entry_codigo)
        messagebox.showwarning(
            "Articulo no encontrado",
            f"No se encontro el codigo {codigo} en la base de datos.",
        )
        return

    rent_modo = selector_rent_modo.current()
    raw_rent = entry_rent_entrada.get().strip()

    precio_venta = None
    rentabilidad = None

    if raw_rent:
        try:
            val = float(raw_rent.replace(",", "."))
        except ValueError:
            mark_input_error(entry_rent_entrada)
            errors.append(f"{lbl_rent_entrada.cget('text')}: debe ser un valor numerico.")
            val = None
    else:
        val = None

    rent_ok, rent_error = validar_rentabilidad(val, rent_modo)
    if not rent_ok:
        mark_input_error(entry_rent_entrada)
        errors.append(rent_error)

    if show_input_errors(errors):
        return

    if modo == "Maderas y Tatamis":
        precio_raw = precio_proveedor_item(item, "Cipta")
        if precio_raw in (None, "", "NO ESTA"):
            messagebox.showwarning(
                "Precio no disponible",
                "Este articulo no tiene precio Cipta en la base de datos.",
            )
            return

        precio_dolares = parse_required_float(
            entry_precio_dolares, "Precio total en Dolares ($)", errors
        )
        precio_euros = parse_required_float(
            entry_precio_euros, "Precio total en Euros (EUR)", errors, allow_zero=False
        )
        factura_transporte = parse_required_float(
            entry_factura_transporte, "Factura de Transporte (EUR)", errors
        )
        derechos_aranceles = parse_required_float(
            entry_derechos_aranceles, "Derechos de Aranceles (EUR)", errors
        )
        if show_input_errors(errors):
            return

        res = calcular_maderas_tatamis(
            precio_dolares,
            precio_euros,
            factura_transporte,
            derechos_aranceles,
            float(precio_raw),
            codigo,
        )
        if res is None:
            mark_input_error(entry_precio_euros)
            messagebox.showwarning(
                "Error de calculo",
                "No se pudo calcular. Revisa que el precio total en euros no sea 0.",
            )
            return

        coste = res["precio_coste_final"]
        if val is not None:
            if rent_modo == 0:
                precio_venta = val
                rentabilidad = round((val - coste) / val * 100, 2)
            else:
                rentabilidad = val
                precio_venta = round(coste / (1 - val / 100), 2)

        res["precio_venta"] = precio_venta
        res["rentabilidad"] = rentabilidad
        res["tipo"] = "Maderas y Tatamis"
        res["codigo"] = codigo
        res["proveedor"] = "CIPTA"
        resultados_exportar.insert(0, res)
        _mostrar_resultado_mt(res, codigo)

    else:
        proveedor = selector_proveedor.get()
        precio_raw = precio_proveedor_item(item, proveedor)
        if precio_raw in (None, "NO ESTA"):
            messagebox.showwarning(
                "Precio no disponible",
                f"Este articulo no tiene precio para {proveedor}.",
            )
            return

        ct_iva = parse_required_float(
            entry_coste_transporte_f, "Coste Transporte + IVA (EUR)", errors
        )
        m3_camion = parse_required_float(
            entry_m3_total_camion_f, "M3 Total Camion", errors, allow_zero=False
        )
        unidades = parse_required_int(
            entry_unidad, "Unidades por referencia", errors, min_value=1
        )
        cantidad = parse_required_int(
            entry_cantidad_productos, "Cantidad total de productos", errors, min_value=1
        )
        if show_input_errors(errors):
            return

        res = calcular_futones(ct_iva, m3_camion, unidades, cantidad, float(precio_raw), codigo)
        if res is None:
            mark_input_error(entry_m3_total_camion_f)
            mark_input_error(entry_cantidad_productos)
            messagebox.showwarning(
                "Error de calculo",
                "No se pudo calcular. Revisa M3 Total Camion y Cantidad total de productos.",
            )
            return

        coste = res["precio_coste_final"]
        if val is not None:
            if rent_modo == 0:
                precio_venta = val
                rentabilidad = round((val - coste) / val * 100, 2)
            else:
                rentabilidad = val
                precio_venta = round(coste / (1 - val / 100), 2)

        res["precio_venta"] = precio_venta
        res["rentabilidad"] = rentabilidad
        res["tipo"] = "Futones"
        res["codigo"] = codigo
        res["proveedor"] = proveedor
        resultados_exportar.insert(0, res)
        _mostrar_resultado_futones(res, codigo, proveedor)


# ---------------------------------------------------------------------------
# Visualización de resultados
# ---------------------------------------------------------------------------
def _fila(parent, etiqueta, valor, bg):
    fila = tk.Frame(parent, bg=bg)
    fila.pack(fill="x", padx=2, pady=1)
    tk.Label(fila, text=etiqueta, bg=bg, fg=C_MUTED, anchor="w",
             font=FONT_TABLE).pack(side="left", fill="x", expand=True)
    tk.Label(fila, text=valor, bg=bg, fg=C_LBL, anchor="e",
             font=FONT_TABLE_B).pack(side="right")

def _separador(parent, bg):
    tk.Frame(parent, bg=C_PANEL_LINE, height=1).pack(fill="x", padx=2, pady=7)

def _fila_final(parent, etiqueta, valor, bg, fg):
    fila = tk.Frame(parent, bg=bg)
    fila.pack(fill="x", padx=2, pady=(4, 1))
    tk.Label(fila, text=etiqueta, bg=bg, fg=fg, anchor="w",
             font=FONT_TABLE_B).pack(side="left", fill="x", expand=True)
    tk.Label(fila, text=valor, bg=bg, fg=fg, anchor="e",
             font=("Segoe UI", 10, "bold")).pack(side="right")


def _insertar_al_principio(nuevo_frame):
    """Mueve el frame al inicio del panel de resultados tras crearlo."""
    # pack el frame normalmente primero para que exista en el gestor
    nuevo_frame.pack(fill="x", pady=6, padx=8)
    # Luego reordenar: usar tk.pack con before del primer hijo existente
    hijos = list(frame_derecho.pack_slaves())
    # hijos[-1] es el recién añadido; hijos[:-1] son los anteriores
    anteriores = hijos[:-1]
    if anteriores:
        # Desempacar y reempacar antes del primero
        nuevo_frame.pack_forget()
        nuevo_frame.pack(fill="x", pady=6, padx=8, before=anteriores[0])


def _mostrar_resultado_mt(res, codigo):
    nombre = res["denominacion"] or f"Artículo {codigo}"
    frame  = tk.LabelFrame(frame_derecho, text=f"  {nombre}  ",
                           padx=12, pady=10,
                           bg=C_MT_BG, fg=C_MT_FG,
                           font=("Segoe UI", 9, "bold"),
                           bd=1, relief="solid",
                           highlightthickness=1, highlightbackground=C_MT_ACCENT)
    _insertar_al_principio(frame)

    for i, (etiqueta, valor) in enumerate([
        ("Tasa de cambio", res["tasa_cambio"]),
        ("Importe transporte", f"{res['importe_transporte']} €"),
        ("% coste transporte", f"{res['pc_transporte']} %"),
        ("% coste descarga", f"{res['pc_descarga']} %"),
        ("% suma gastos", f"{res['pc_suma']} %"),
        ("Precio artículo", f"{res['precio_euros_art']} €"),
        ("Gastos aplicables", f"{res['gastos_aplicables']} €"),
        ("Coste sin almacenaje", f"{res['coste_sin_almacenaje']} €"),
        ("Coste almacenaje + IVA", f"{res['coste_almacenaje_iva']} €"),
        ("Coste picking + IVA", f"{res['coste_picking_iva']} €"),
    ]):
        _fila(frame, etiqueta, valor, C_MT_ROW if i % 2 else C_MT_BG)
    _separador(frame, C_MT_BG)
    _fila_final(frame, "PRECIO COSTE FINAL", f"{res['precio_coste_final']} €",
                C_MT_BG, C_MT_FG)

    if res.get("precio_venta") is not None or res.get("rentabilidad") is not None:
        pv   = res.get("precio_venta")
        rent = res.get("rentabilidad")
        color_rent = C_OK if rent is not None and rent >= 0 else C_ERR
        _separador(frame, C_MT_BG)
        _fila(frame, "Precio de venta", f"{pv} €" if pv is not None else "—", C_MT_BG)
        _fila_final(frame, "RENTABILIDAD", f"{rent} %" if rent is not None else "—",
                    C_MT_BG, color_rent)


def _mostrar_resultado_futones(res, codigo, proveedor):
    nombre = res["denominacion"] or f"Artículo {codigo}"
    frame  = tk.LabelFrame(frame_derecho, text=f"  {nombre}  [{proveedor}]  ",
                           padx=12, pady=10,
                           bg=C_FT_BG, fg=C_FT_FG,
                           font=("Segoe UI", 9, "bold"),
                           bd=1, relief="solid",
                           highlightthickness=1, highlightbackground=C_FT_ACCENT)
    _insertar_al_principio(frame)

    for i, (etiqueta, valor) in enumerate([
        ("Coste transporte por M³", f"{res['ct_m3']} €"),
        ("Coste transporte x producto", f"{res['ct_m3_prod']} €"),
        ("Coste transporte total ref.", f"{res['ct_total_ref']} €"),
        ("Coste descarga x prod. + IVA", f"{res['cd_prod_iva']} €"),
        ("Coste descarga total ref.", f"{res['cd_total_ref']} €"),
        ("Importe IVA + RE", f"{res['iva_re']} €"),
        ("Precio compra IVA + RE", f"{res['precio_con_iva']} €"),
        ("Coste final con descarga", f"{res['coste_descarga']} €"),
        ("Coste almacenaje + IVA", f"{res['coste_almacenaje_iva']} €"),
        ("Coste picking + IVA", f"{res['coste_picking_iva']} €"),
    ]):
        _fila(frame, etiqueta, valor, C_FT_ROW if i % 2 else C_FT_BG)
    _separador(frame, C_FT_BG)
    _fila_final(frame, "PRECIO COSTE FINAL", f"{res['precio_coste_final']} €",
                C_FT_BG, C_FT_FG)

    if res.get("precio_venta") is not None or res.get("rentabilidad") is not None:
        pv   = res.get("precio_venta")
        rent = res.get("rentabilidad")
        color_rent = C_OK if rent is not None and rent >= 0 else C_ERR
        _separador(frame, C_FT_BG)
        _fila(frame, "Precio de venta", f"{pv} €" if pv is not None else "—", C_FT_BG)
        _fila_final(frame, "RENTABILIDAD", f"{rent} %" if rent is not None else "—",
                    C_FT_BG, color_rent)


# ---------------------------------------------------------------------------
# Exportar a Excel
# ---------------------------------------------------------------------------
def btn_exportar():
    from tkinter import filedialog
    import openpyxl
    from openpyxl.styles import Border, Font, PatternFill, Alignment, Side

    if not resultados_exportar:
        messagebox.showinfo("Sin datos", "No hay artículos calculados para exportar.")
        return

    ruta = filedialog.asksaveasfilename(
        title="Guardar como",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
        initialfile="costes_futonspai.xlsx",
    )
    if not ruta:
        return   # cancelado

    wb = openpyxl.Workbook()

    # ---- Hoja Maderas y Tatamis ----
    ws_mt = wb.active
    ws_mt.title = "Maderas y Tatamis"

    hdrs_mt = [
        "Código", "Denominación", "Proveedor",
        "Tasa Cambio", "Importe Transporte (€)", "% Transporte",
        "% Descarga", "% Suma Gastos", "Precio Art. en €",
        "Gastos Aplicables (€)", "Coste sin Almacenaje (€)",
        "Coste Almacenaje+IVA (€)", "Coste Picking+IVA (€)",
        "Precio Coste Final (€)", "Precio Venta (€)", "Rentabilidad (%)",
    ]

    # ---- Hoja Futones ----
    ws_ft = wb.create_sheet("Futones")

    hdrs_ft = [
        "Código", "Denominación", "Proveedor",
        "Coste Transporte M³ (€)", "Coste Transporte x Prod. (€)",
        "Coste Transporte Total Ref. (€)", "Coste Descarga x Prod.+IVA (€)",
        "Coste Descarga Total Ref. (€)", "Importe IVA+RE (€)",
        "Precio Compra IVA+RE (€)", "Coste Final con Descarga (€)",
        "Coste Almacenaje+IVA (€)", "Coste Picking+IVA (€)",
        "Precio Coste Final (€)", "Precio Venta (€)", "Rentabilidad (%)",
    ]

    # Estilo cabecera
    fill_mt = PatternFill("solid", fgColor="1D4ED8")
    fill_ft = PatternFill("solid", fgColor="047857")
    fill_alt = PatternFill("solid", fgColor="F8FAFC")
    font_hdr = Font(color="FFFFFF", bold=True, name="Segoe UI", size=10)
    font_body = Font(color="1F2937", name="Segoe UI", size=10)
    border = Border(bottom=Side(style="thin", color="E5E7EB"))

    def _write_headers(ws, hdrs, fill):
        ws.append(hdrs)
        for cell in ws[1]:
            cell.font      = font_hdr
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border
        ws.row_dimensions[1].height = 22

    _write_headers(ws_mt, hdrs_mt, fill_mt)
    _write_headers(ws_ft, hdrs_ft, fill_ft)

    for res in resultados_exportar:
        rent = res.get("rentabilidad")
        pv   = res.get("precio_venta")

        if res["tipo"] == "Maderas y Tatamis":
            ws_mt.append([
                res["codigo"], res["denominacion"], res["proveedor"],
                res["tasa_cambio"], res["importe_transporte"], res["pc_transporte"],
                res["pc_descarga"], res["pc_suma"], res["precio_euros_art"],
                res["gastos_aplicables"], res["coste_sin_almacenaje"],
                res["coste_almacenaje_iva"], res["coste_picking_iva"],
                res["precio_coste_final"], pv, rent,
            ])
        else:
            ws_ft.append([
                res["codigo"], res["denominacion"], res["proveedor"],
                res["ct_m3"], res["ct_m3_prod"], res["ct_total_ref"],
                res["cd_prod_iva"], res["cd_total_ref"], res["iva_re"],
                res["precio_con_iva"], res["coste_descarga"],
                res["coste_almacenaje_iva"], res["coste_picking_iva"],
                res["precio_coste_final"], pv, rent,
            ])

    # Ajustar anchos automáticamente
    for ws in (ws_mt, ws_ft):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = font_body
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if cell.row % 2 == 0:
                    cell.fill = fill_alt

    try:
        wb.save(ruta)
        messagebox.showinfo("Exportado", f"Archivo guardado en:\n{ruta}")
    except Exception as e:
        messagebox.showerror("Error al guardar", str(e))


# ---------------------------------------------------------------------------
# Botón Calcular — anclado abajo
# ---------------------------------------------------------------------------
boton_container = tk.Frame(left_shell, bg=C_PANEL, padx=14, pady=10)
boton_container.pack(side="bottom", fill="x")

make_button(
    boton_container, "Calcular", btn_calcular,
    C_BTN, C_BTN_HOVER, C_BTN_ACTIVE, pady=8
).pack(fill="x")

make_button(
    boton_container, "Exportar a Excel", btn_exportar,
    C_EXPORT, C_EXPORT_HOVER, C_EXPORT_ACTIVE, pady=7
).pack(fill="x", pady=(8, 0))

ventana_principal.mainloop()
