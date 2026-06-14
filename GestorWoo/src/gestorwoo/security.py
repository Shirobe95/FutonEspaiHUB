from __future__ import annotations

import json
import sqlite3
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from gestorwoo.config import Settings
from gestorwoo.theme import (
    C_BG,
    C_ERR,
    C_EXPORT,
    C_INFO,
    C_LBL,
    C_OK,
    C_PANEL,
    C_PANEL_LINE,
    apply_theme,
)

try:
    from gestorwoo.windowing import center_window
except ModuleNotFoundError:
    def center_window(window: tk.Tk | tk.Toplevel, width: int, height: int) -> None:
        window.update_idletasks()
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()
        x = max((screen_width - width) // 2, 0)
        y = max((screen_height - height) // 2, 0)
        window.geometry(f"{width}x{height}+{x}+{y}")


LOG_COLUMNS = (
    "id",
    "created_at",
    "category",
    "severity",
    "module",
    "action",
    "status",
    "entity_type",
    "entity_id",
    "details",
)

LOG_CATEGORIES = [
    "Todos",
    "Cambio de Precios WooCommerce",
    "Pedidos de Proveedores",
    "Constantes de Cálculo de Coste",
    "Inventario WooCommerce",
    "Mapa Maestro",
    "Backups y Restauraciones",
    "Seguridad / Sistema",
    "General",
]


def infer_log_category(module: str = "", action: str = "", entity_type: str = "") -> str:
    text = f"{module} {action} {entity_type}".lower()
    if any(token in text for token in ["backup", "restaur", "restore"]):
        return "Backups y Restauraciones"
    if any(token in text for token in ["constante", "constantes"]):
        return "Constantes de Cálculo de Coste"
    if any(token in text for token in ["coste de pedido", "pedido proveedor", "proveedor", "ekomat", "pascal", "hemei", "cipta"]):
        return "Pedidos de Proveedores"
    if any(token in text for token in ["cambio de precio", "precio", "propuesta", "publicar cambios", "excel revisado"]):
        return "Cambio de Precios WooCommerce"
    if any(token in text for token in ["mapa maestro", "heca", "inventario local", "data.xlsx", "modificar ficha"]):
        return "Mapa Maestro"
    if any(token in text for token in ["woocommerce", "clasificar", "sincronizar", "catalogo", "catálogo"]):
        return "Inventario WooCommerce"
    if any(token in text for token in ["seguridad", "logs", "hub", "sistema"]):
        return "Seguridad / Sistema"
    return "General"


def _connect(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 5000")
    return conn


def init_security_schema(db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with _connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS security_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                category TEXT NOT NULL DEFAULT 'General',
                severity TEXT NOT NULL DEFAULT 'INFO',
                module TEXT NOT NULL,
                action TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'OK',
                entity_type TEXT DEFAULT '',
                entity_id TEXT DEFAULT '',
                details TEXT DEFAULT '',
                context_json TEXT DEFAULT '',
                source TEXT DEFAULT 'HUB'
            )
            """
        )
        cols = {row[1] for row in conn.execute("PRAGMA table_info(security_logs)").fetchall()}
        if "category" not in cols:
            conn.execute("ALTER TABLE security_logs ADD COLUMN category TEXT NOT NULL DEFAULT 'General'")
            rows = conn.execute("SELECT id, module, action, entity_type FROM security_logs").fetchall()
            for row in rows:
                category = infer_log_category(row["module"], row["action"], row["entity_type"])
                conn.execute("UPDATE security_logs SET category = ? WHERE id = ?", (category, row["id"]))
        conn.execute("CREATE INDEX IF NOT EXISTS idx_security_logs_created ON security_logs(created_at DESC)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_security_logs_category ON security_logs(category)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_security_logs_module ON security_logs(module)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_security_logs_status ON security_logs(status)")
        conn.commit()


def log_event(
    db_path: Path,
    *,
    module: str,
    action: str,
    category: str | None = None,
    status: str = "OK",
    severity: str = "INFO",
    entity_type: str = "",
    entity_id: str = "",
    details: str = "",
    context: dict[str, object] | None = None,
    source: str = "HUB",
) -> None:
    """Registra un evento operativo del HUB. Nunca debe romper la accion principal."""
    try:
        init_security_schema(db_path)
        payload = json.dumps(context or {}, ensure_ascii=False, default=str) if context else ""
        resolved_category = (category or infer_log_category(module, action, entity_type)).strip() or "General"
        with _connect(db_path) as conn:
            conn.execute(
                """
                INSERT INTO security_logs (
                    created_at, category, severity, module, action, status,
                    entity_type, entity_id, details, context_json, source
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    resolved_category,
                    severity.upper().strip() or "INFO",
                    module.strip() or "General",
                    action.strip() or "Accion",
                    status.upper().strip() or "OK",
                    str(entity_type or ""),
                    str(entity_id or ""),
                    str(details or ""),
                    payload,
                    str(source or "HUB"),
                ),
            )
            conn.commit()
    except Exception:
        # El log es capa de seguridad/observabilidad. No debe tumbar el flujo operativo.
        return


def list_logs(
    db_path: Path,
    *,
    category: str = "Todos",
    module: str = "Todos",
    status: str = "Todos",
    severity: str = "Todos",
    search: str = "",
    limit: int = 1000,
) -> list[sqlite3.Row]:
    init_security_schema(db_path)
    query = [
        "SELECT id, created_at, category, severity, module, action, status, entity_type, entity_id, details, context_json, source",
        "FROM security_logs WHERE 1=1",
    ]
    params: list[object] = []
    if category and category != "Todos":
        query.append("AND category = ?")
        params.append(category)
    if module and module != "Todos":
        query.append("AND module = ?")
        params.append(module)
    if status and status != "Todos":
        query.append("AND status = ?")
        params.append(status)
    if severity and severity != "Todos":
        query.append("AND severity = ?")
        params.append(severity)
    if search.strip():
        like = f"%{search.strip()}%"
        query.append(
            "AND (action LIKE ? OR details LIKE ? OR entity_id LIKE ? OR entity_type LIKE ? OR module LIKE ? OR category LIKE ?)"
        )
        params.extend([like, like, like, like, like, like])
    query.append("ORDER BY datetime(created_at) DESC, id DESC LIMIT ?")
    params.append(int(limit))
    with _connect(db_path) as conn:
        return list(conn.execute("\n".join(query), params))


def distinct_values(db_path: Path, column: str) -> list[str]:
    if column not in {"category", "module", "status", "severity"}:
        return []
    init_security_schema(db_path)
    with _connect(db_path) as conn:
        rows = conn.execute(
            f"SELECT DISTINCT {column} AS value FROM security_logs WHERE COALESCE({column}, '') <> '' ORDER BY {column}"
        ).fetchall()
    return [str(row["value"]) for row in rows]


def _pretty_json(value: str) -> str:
    if not value:
        return ""
    try:
        parsed = json.loads(value)
    except Exception:
        return value
    return json.dumps(parsed, ensure_ascii=False, indent=2, default=str)


def _load_context(value: str) -> object:
    if not value:
        return {}
    try:
        return json.loads(value)
    except Exception:
        return {}


def _short_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def _first_list_from_context(context: object) -> tuple[str, list[object]]:
    if not isinstance(context, dict):
        return "", []
    for key in ("changes", "cambios", "items", "lineas", "propuestas"):
        value = context.get(key)
        if isinstance(value, list) and value:
            return key, value
    return "", []


def _friendly_context_title(key: str) -> str:
    return {
        "changes": "Cambios detectados",
        "cambios": "Cambios detectados",
        "items": "Items afectados",
        "lineas": "Líneas afectadas",
        "propuestas": "Propuestas afectadas",
    }.get(key, "Detalle")


def _friendly_column_name(key: str) -> str:
    mapping = {
        "campo": "Campo",
        "field": "Campo",
        "field_name": "Campo",
        "valor_anterior": "Valor anterior",
        "old_value": "Valor anterior",
        "precio_anterior": "Precio anterior",
        "valor_nuevo": "Valor nuevo",
        "new_value": "Valor nuevo",
        "precio_nuevo": "Precio nuevo",
        "proveedor": "Proveedor",
        "supplier": "Proveedor",
        "codigo": "Código",
        "code": "Código",
        "sku": "SKU",
        "nombre": "Nombre",
        "name": "Nombre",
        "item_id": "ID local",
        "unidades": "Unidades",
        "quantity": "Cantidad",
        "qty": "Cantidad",
        "precio_actual": "Precio actual",
        "current_price": "Precio actual",
        "precio_propuesto": "Precio propuesto",
        "new_price": "Precio nuevo",
        "diferencia": "Diferencia",
        "status": "Estado",
        "estado": "Estado",
    }
    return mapping.get(key, key.replace("_", " ").title())


def export_logs_to_excel(db_path: Path, output_path: Path, rows: list[sqlite3.Row]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"

    headers = ["Fecha", "Sección", "Nivel", "Modulo", "Accion", "Estado", "Entidad", "ID", "Detalles"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["created_at"],
            row["category"],
            row["severity"],
            row["module"],
            row["action"],
            row["status"],
            row["entity_type"],
            row["entity_id"],
            row["details"],
        ])

    header_fill = PatternFill("solid", fgColor="EAF2FF")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    for idx, column in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 3, 12), 55)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


class SecurityLogsApp(tk.Tk):
    def __init__(self, settings: Settings) -> None:
        super().__init__()
        self.settings = settings
        self.db_path = settings.db_path
        init_security_schema(self.db_path)
        self.title("Futon Espai - Seguridad · Logs")
        center_window(self, 1180, 720)
        self.minsize(980, 620)
        self.configure(bg=C_BG)
        apply_theme(self)

        self.category_var = tk.StringVar(value="Todos")
        self.module_var = tk.StringVar(value="Todos")
        self.status_var = tk.StringVar(value="Todos")
        self.severity_var = tk.StringVar(value="Todos")
        self.search_var = tk.StringVar()
        self.limit_var = tk.StringVar(value="1000")
        self.rows: list[sqlite3.Row] = []

        self._build_layout()
        self._refresh_filter_values()
        self._load_logs()

    def _build_layout(self) -> None:
        root = ttk.Frame(self, padding=20)
        root.pack(fill=tk.BOTH, expand=True)

        header = ttk.Frame(root)
        header.pack(fill=tk.X, pady=(0, 16))
        ttk.Label(header, text="Seguridad · Logs", style="Title.TLabel").pack(anchor=tk.W)
        ttk.Label(
            header,
            text="Bitácoras separadas por área: precios, pedidos de proveedor, constantes, inventario, Mapa Maestro, backups y sistema.",
            style="Muted.TLabel",
        ).pack(anchor=tk.W, pady=(4, 0))

        filters = tk.Frame(root, bg=C_PANEL, highlightbackground=C_PANEL_LINE, highlightthickness=1)
        filters.pack(fill=tk.X, pady=(0, 12))
        for idx in range(10):
            filters.columnconfigure(idx, weight=1 if idx in {1, 3, 5, 7} else 0)

        def label(text: str, row: int, col: int) -> None:
            ttk.Label(filters, text=text, style="Panel.TLabel").grid(row=row, column=col, padx=(12, 4), pady=8, sticky="w")

        label("Bitácora", 0, 0)
        self.category_combo = ttk.Combobox(filters, textvariable=self.category_var, state="readonly", width=30)
        self.category_combo.grid(row=0, column=1, columnspan=3, padx=(0, 8), pady=8, sticky="ew")

        label("Buscar", 0, 4)
        ttk.Entry(filters, textvariable=self.search_var).grid(row=0, column=5, columnspan=3, padx=(0, 8), pady=8, sticky="ew")

        ttk.Button(filters, text="Actualizar", command=self._load_logs).grid(row=0, column=8, padx=(0, 8), pady=8)
        ttk.Button(filters, text="Limpiar", command=self._clear_filters).grid(row=0, column=9, padx=(0, 12), pady=8)

        label("Modulo", 1, 0)
        self.module_combo = ttk.Combobox(filters, textvariable=self.module_var, state="readonly", width=22)
        self.module_combo.grid(row=1, column=1, padx=(0, 8), pady=(0, 8), sticky="ew")

        label("Estado", 1, 2)
        self.status_combo = ttk.Combobox(filters, textvariable=self.status_var, state="readonly", width=14)
        self.status_combo.grid(row=1, column=3, padx=(0, 8), pady=(0, 8), sticky="ew")

        label("Nivel", 1, 4)
        self.severity_combo = ttk.Combobox(filters, textvariable=self.severity_var, state="readonly", width=14)
        self.severity_combo.grid(row=1, column=5, padx=(0, 8), pady=(0, 8), sticky="ew")

        table_frame = ttk.Frame(root)
        table_frame.pack(fill=tk.BOTH, expand=True)
        columns = ("created_at", "category", "severity", "module", "action", "status", "entity", "entity_id", "details")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="browse")
        headings = {
            "created_at": "Fecha",
            "category": "Bitácora",
            "severity": "Nivel",
            "module": "Modulo",
            "action": "Accion",
            "status": "Estado",
            "entity": "Entidad",
            "entity_id": "ID",
            "details": "Detalles",
        }
        for col, text in headings.items():
            self.tree.heading(col, text=text, anchor=tk.CENTER)
            self.tree.column(col, anchor=tk.CENTER, width=120, minwidth=80, stretch=False)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        self.tree.tag_configure("ERROR", foreground=C_ERR)
        self.tree.tag_configure("WARNING", foreground="#B45309")
        self.tree.tag_configure("OK", foreground=C_OK)
        self.tree.bind("<Double-1>", self._open_log_detail)

        bottom = ttk.Frame(root)
        bottom.pack(fill=tk.X, pady=(12, 0))
        self.status_label = ttk.Label(bottom, text="", style="Status.TLabel")
        self.status_label.pack(side=tk.LEFT)
        ttk.Button(bottom, text="Exportar Excel", style="Secondary.TButton", command=self._export_excel).pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(bottom, text="Cerrar", command=self.destroy).pack(side=tk.RIGHT)

    def _refresh_filter_values(self) -> None:
        categories = [cat for cat in LOG_CATEGORIES if cat == "Todos" or cat in set(distinct_values(self.db_path, "category")) or cat != "General"]
        modules = ["Todos", *distinct_values(self.db_path, "module")]
        statuses = ["Todos", *distinct_values(self.db_path, "status")]
        severities = ["Todos", *distinct_values(self.db_path, "severity")]
        self.category_combo.configure(values=categories)
        self.module_combo.configure(values=modules)
        self.status_combo.configure(values=statuses)
        self.severity_combo.configure(values=severities)

    def _clear_filters(self) -> None:
        self.category_var.set("Todos")
        self.module_var.set("Todos")
        self.status_var.set("Todos")
        self.severity_var.set("Todos")
        self.search_var.set("")
        self._load_logs()

    def _load_logs(self) -> None:
        try:
            limit = max(50, min(int(self.limit_var.get() or "1000"), 10000))
        except ValueError:
            limit = 1000
        self.rows = list_logs(
            self.db_path,
            category=self.category_var.get(),
            module=self.module_var.get(),
            status=self.status_var.get(),
            severity=self.severity_var.get(),
            search=self.search_var.get(),
            limit=limit,
        )
        self.tree.delete(*self.tree.get_children())
        for row in self.rows:
            tag = row["status"] if row["status"] in {"OK", "ERROR", "WARNING"} else row["severity"]
            self.tree.insert(
                "",
                tk.END,
                values=(
                    row["created_at"],
                    row["category"],
                    row["severity"],
                    row["module"],
                    row["action"],
                    row["status"],
                    row["entity_type"],
                    row["entity_id"],
                    row["details"],
                ),
                tags=(tag,),
            )
        self._auto_size_columns()
        self.status_label.configure(text=f"Logs mostrados: {len(self.rows)}")
        self._refresh_filter_values()

    def _auto_size_columns(self) -> None:
        widths = {
            "created_at": 155,
            "category": 230,
            "severity": 80,
            "module": 140,
            "action": 220,
            "status": 90,
            "entity": 120,
            "entity_id": 120,
            "details": 360,
        }
        for idx, col in enumerate(self.tree["columns"]):
            max_len = len(self.tree.heading(col)["text"])
            for item in self.tree.get_children("")[:200]:
                values = self.tree.item(item, "values")
                if idx < len(values):
                    max_len = max(max_len, len(str(values[idx])))
            width = min(max(widths.get(col, 120), max_len * 8 + 26), 520)
            self.tree.column(col, width=width)

    def _open_log_detail(self, _event=None) -> None:
        selected = self.tree.selection()
        if not selected:
            return
        try:
            index = self.tree.index(selected[0])
            row = self.rows[index]
        except Exception:
            return
        LogDetailDialog(self, row)

    def _export_excel(self) -> None:
        if not self.rows:
            messagebox.showinfo("Sin datos", "No hay logs para exportar.")
            return
        default_name = f"logs_hub_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Exportar logs",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        try:
            export_logs_to_excel(self.db_path, Path(path), self.rows)
        except Exception as exc:
            messagebox.showerror("No se pudo exportar", str(exc))
            return
        messagebox.showinfo("Exportacion completa", f"Logs exportados en:\n{path}")



class LogDetailDialog(tk.Toplevel):
    def __init__(self, parent: SecurityLogsApp, row: sqlite3.Row) -> None:
        super().__init__(parent)
        self.row = row
        self.context = _load_context(str(row["context_json"] or ""))
        self.title(f"Detalle log #{row['id']}")
        center_window(self, 980, 720)
        self.minsize(820, 560)
        self.configure(bg=C_BG)
        self.transient(parent)
        self.grab_set()
        self._build_layout()

    def _build_layout(self) -> None:
        container = ttk.Frame(self, padding=16)
        container.pack(fill=tk.BOTH, expand=True)
        ttk.Label(container, text="Detalle completo del registro", style="Title.TLabel").pack(anchor=tk.W)
        ttk.Label(
            container,
            text="Vista ampliada de la acción: resumen, entidad afectada y cambios registrados.",
            style="Muted.TLabel",
            wraplength=900,
        ).pack(anchor=tk.W, pady=(4, 12))

        fields = ttk.Frame(container, style="Panel.TFrame", padding=10)
        fields.pack(fill=tk.X, pady=(0, 12))
        items = [
            ("Fecha", self.row["created_at"]),
            ("Bitácora", self.row["category"]),
            ("Nivel", self.row["severity"]),
            ("Módulo", self.row["module"]),
            ("Acción", self.row["action"]),
            ("Estado", self.row["status"]),
            ("Entidad", self.row["entity_type"]),
            ("ID", self.row["entity_id"]),
            ("Fuente", self.row["source"]),
        ]
        for idx, (label, value) in enumerate(items):
            ttk.Label(fields, text=f"{label}:", style="Panel.TLabel").grid(row=idx//3, column=(idx%3)*2, sticky="w", padx=(0, 6), pady=3)
            ttk.Label(fields, text=str(value or "—"), style="Status.TLabel").grid(row=idx//3, column=(idx%3)*2+1, sticky="w", padx=(0, 18), pady=3)

        notebook = ttk.Notebook(container)
        notebook.pack(fill=tk.BOTH, expand=True)
        self._build_summary_tab(notebook)
        self._build_table_tab(notebook)
        self._build_json_tab(notebook)

        footer = ttk.Frame(container)
        footer.pack(fill=tk.X, pady=(12, 0))
        ttk.Button(footer, text="Cerrar", command=self.destroy).pack(side=tk.RIGHT)

    def _build_summary_tab(self, notebook: ttk.Notebook) -> None:
        frame = ttk.Frame(notebook, padding=10)
        notebook.add(frame, text="Resumen")
        text = tk.Text(frame, wrap="word", bg="white", fg=C_LBL, relief="flat", padx=10, pady=10, height=10)
        scroll = ttk.Scrollbar(frame, orient="vertical", command=text.yview)
        text.configure(yscrollcommand=scroll.set)
        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        text.insert("end", "DETALLES\n")
        text.insert("end", (str(self.row["details"] or "") or "Sin detalles") + "\n\n")
        key, rows = _first_list_from_context(self.context)
        if key and rows:
            text.insert("end", f"{_friendly_context_title(key).upper()}\n")
            text.insert("end", f"Total registros detallados: {len(rows)}\n")
        else:
            text.insert("end", "No hay tabla de cambios detallada para este registro.\n")
        text.configure(state="disabled")

    def _build_table_tab(self, notebook: ttk.Notebook) -> None:
        key, rows = _first_list_from_context(self.context)
        frame = ttk.Frame(notebook, padding=10)
        notebook.add(frame, text=_friendly_context_title(key) if key else "Cambios / Items")
        if not rows:
            ttk.Label(frame, text="Este log no contiene una lista de cambios o items.", style="Muted.TLabel").pack(anchor=tk.W)
            return

        normalized: list[dict[str, object]] = []
        columns: list[str] = []
        for row in rows:
            data = row if isinstance(row, dict) else {"valor": row}
            normalized.append(data)
            for col in data.keys():
                if col not in columns:
                    columns.append(col)
        preferred = [
            "campo", "field", "field_name", "proveedor", "codigo", "sku", "item_id", "nombre", "name",
            "valor_anterior", "old_value", "precio_anterior", "precio_actual",
            "valor_nuevo", "new_value", "precio_nuevo", "precio_propuesto",
            "unidades", "quantity", "qty", "diferencia", "status", "estado",
        ]
        ordered = [c for c in preferred if c in columns] + [c for c in columns if c not in preferred]
        tree = ttk.Treeview(frame, columns=ordered, show="headings", selectmode="browse")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        for col in ordered:
            tree.heading(col, text=_friendly_column_name(col), anchor=tk.CENTER)
            tree.column(col, anchor=tk.CENTER, width=150, minwidth=90, stretch=False)
        for data in normalized:
            tree.insert("", tk.END, values=[_short_value(data.get(col, "")) for col in ordered])
        for idx, col in enumerate(ordered):
            max_len = len(_friendly_column_name(col))
            for iid in tree.get_children("")[:300]:
                vals = tree.item(iid, "values")
                if idx < len(vals):
                    max_len = max(max_len, len(str(vals[idx])))
            tree.column(col, width=min(max(max_len * 8 + 28, 110), 360))

    def _build_json_tab(self, notebook: ttk.Notebook) -> None:
        frame = ttk.Frame(notebook, padding=10)
        notebook.add(frame, text="Contexto técnico")
        text = tk.Text(frame, wrap="none", bg="white", fg=C_LBL, relief="flat", padx=10, pady=10)
        yscroll = ttk.Scrollbar(frame, orient="vertical", command=text.yview)
        xscroll = ttk.Scrollbar(frame, orient="horizontal", command=text.xview)
        text.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        text.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        text.insert("end", _pretty_json(str(self.row["context_json"] or "")) or "Sin contexto adicional")
        text.configure(state="disabled")

def run_logs_app(settings: Settings) -> None:
    app = SecurityLogsApp(settings)
    app.mainloop()
