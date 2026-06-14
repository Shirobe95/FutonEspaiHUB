from __future__ import annotations

import json
import queue
import re
import threading
import tkinter as tk
import tkinter.font as tkfont
import unicodedata
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any

from gestorwoo.config import Settings
from gestorwoo.publish import publish_pending_price_changes
from gestorwoo.guard import OperationBlockedError
from gestorwoo.storage import ProductStore
from gestorwoo.sync import sync_products
from gestorwoo.theme import C_BG, C_PANEL, C_PANEL_LINE, apply_theme
from gestorwoo.woocommerce import WooCommerceError


def auto_fit_tree_columns(tree: ttk.Treeview, *, min_width: int = 70, max_width: int = 620, padding: int = 34) -> None:
    """Centra y ajusta las columnas visibles al texto cargado.

    Mantiene el scroll horizontal como red de seguridad para tablas anchas.
    """
    try:
        body_font = tkfont.nametofont("TkDefaultFont")
    except tk.TclError:
        body_font = None
    try:
        heading_font = tkfont.nametofont("TkHeadingFont")
    except tk.TclError:
        heading_font = body_font

    for column in tree["columns"]:
        heading = str(tree.heading(column, "text") or column)
        if heading_font is not None:
            width = heading_font.measure(heading) + padding
        else:
            width = len(heading) * 8 + padding

        index = list(tree["columns"]).index(column)
        for item_id in tree.get_children(""):
            values = tree.item(item_id, "values")
            if index < len(values):
                text = str(values[index] or "")
                measured = body_font.measure(text) + padding if body_font is not None else len(text) * 8 + padding
                width = max(width, measured)

        tree.heading(column, anchor=tk.CENTER)
        tree.column(
            column,
            width=max(min_width, min(width, max_width)),
            minwidth=min_width,
            anchor=tk.CENTER,
            stretch=False,
        )


class ProductTableApp(tk.Tk):
    FAMILIES = [
        "Futones",
        "Tatamis",
        "Camas Japonesas",
        "Sofás Cama",
        "Complementos",
        "Ofertas / Packs",
        "Outlet",
        "Otros / Sin clasificar",
    ]
    SUBGROUPS = [
        "Algodón",
        "Algodón + Látex",
        "Algodón + Coco",
        "Algodón + Lana",
        "Látex",
        "Coco",
        "Lana",
        "Mesitas",
        "Almohadas",
        "Fundas",
        "Toppers",
        "Cojines",
        "Accesorios",
        "Pack Futón + Tatami",
        "Pack Sofá Cama",
        "A medida",
        "Sin subgrupo",
        "Otros",
    ]
    COMMERCIAL_STATUSES = ["Normal", "Oferta", "Outlet"]

    def __init__(self, settings: Settings, mode: str = "inventory") -> None:
        super().__init__()
        self.settings = settings
        self.mode = mode if mode in {"inventory", "prices"} else "inventory"
        self.store = ProductStore(settings.db_path)
        self.events: queue.Queue[tuple[str, object]] = queue.Queue()

        self.title(
            "GestorWoo - Gestion de Inventario"
            if self.mode == "inventory"
            else "GestorWoo - Cambio de Precios"
        )
        self.geometry("1360x760")
        self.minsize(1120, 620)
        self.configure(bg=C_BG)
        apply_theme(self)

        self.status_text = tk.StringVar(value="Cargando articulos locales...")
        self.search_text = tk.StringVar()
        self.category_text = tk.StringVar(value="Todas")
        self.family_text = tk.StringVar(value="Todas")
        self.subgroup_text = tk.StringVar(value="Todos")
        self.color_text = tk.StringVar(value="Todos")
        self.commercial_status_text = tk.StringVar(value="Todos")
        self.reviewed_text = tk.StringVar(value="Todos")
        self.pack_text = tk.StringVar(value="Todos")
        self.catalog_items: list[dict[str, object]] = []
        self.filtered_items: list[dict[str, object]] = []
        self.busy_window: tk.Widget | None = None
        self.busy_message_text = tk.StringVar(value="")
        self.busy_detail_text = tk.StringVar(value="")
        self.busy_progress: ttk.Progressbar | None = None
        self._busy_configure_binding: str | None = None
        self._busy_job: str | None = None

        self._build_layout()
        self._load_products()

    def _build_layout(self) -> None:
        header = ttk.Frame(self, padding=(14, 12, 14, 8))
        header.pack(fill=tk.X)

        title_text = (
            "Gestor WooCommerce · Gestión de inventario"
            if self.mode == "inventory"
            else "Gestor WooCommerce · Cambio de precios seguro"
        )
        ttk.Label(
            header,
            text=title_text,
            style="Title.TLabel",
        ).pack(side=tk.LEFT)

        subtitle_text = (
            "Organiza, clasifica y revisa el catálogo local sin modificar WooCommerce."
            if self.mode == "inventory"
            else "Prepara, revisa y publica cambios de precio con controles de seguridad."
        )
        ttk.Label(
            header,
            text=subtitle_text,
            style="Muted.TLabel",
        ).pack(side=tk.LEFT, padx=(18, 0))

        toolbar = ttk.Frame(self, padding=(14, 8, 14, 6), style="Toolbar.TFrame")
        toolbar.pack(fill=tk.X)

        self.update_button: ttk.Button | None = None
        self.auto_classify_button: ttk.Button | None = None
        if self.mode == "inventory":
            update_button = ttk.Button(
                toolbar,
                text="Actualizar desde WooCommerce",
                command=self._sync_from_woocommerce,
            )
            update_button.pack(side=tk.LEFT)
            self.update_button = update_button

            auto_classify_button = ttk.Button(
                toolbar,
                text="Clasificar automaticamente",
                command=self._auto_classify_catalog,
            )
            auto_classify_button.pack(side=tk.LEFT, padx=(8, 0))
            self.auto_classify_button = auto_classify_button

            ttk.Button(
                toolbar,
                text="Editar clasificación",
                command=self._open_classification_dialog,
            ).pack(side=tk.LEFT, padx=(8, 0))

            ttk.Button(
                toolbar,
                text="Marcar revisado",
                command=self._mark_selected_reviewed,
            ).pack(side=tk.LEFT, padx=(8, 0))

            ttk.Button(
                toolbar,
                text="Exportar catálogo clasificado",
                command=self._export_classified_catalog,
            ).pack(side=tk.LEFT, padx=(8, 0))

            ttk.Button(
                toolbar,
                text="Crear pack",
                command=self._open_pack_dialog,
            ).pack(side=tk.LEFT, padx=(18, 0))
        else:
            ttk.Button(
                toolbar,
                text="Cambiar precio",
                command=self._open_price_dialog,
            ).pack(side=tk.LEFT, padx=(8, 0))

            ttk.Button(
                toolbar,
                text="Añadir filtro a propuesta",
                command=self._open_bulk_price_dialog,
            ).pack(side=tk.LEFT, padx=(8, 0))

            ttk.Button(
                toolbar,
                text="Ver propuestas",
                command=self._open_publish_dialog,
            ).pack(side=tk.LEFT, padx=(8, 0))

            ttk.Button(
                toolbar,
                text="Publicar cambios",
                style="Secondary.TButton",
                command=self._open_publish_dialog,
            ).pack(side=tk.LEFT, padx=(8, 0))

        filter_bar = ttk.Frame(self, padding=(14, 4, 14, 10), style="Toolbar.TFrame")
        filter_bar.pack(fill=tk.X)

        ttk.Label(filter_bar, text="Buscar:").pack(side=tk.LEFT, padx=(0, 6))
        search_entry = ttk.Entry(filter_bar, textvariable=self.search_text, width=26)
        search_entry.pack(side=tk.LEFT)
        search_entry.bind("<KeyRelease>", lambda _event: self._render_products())

        ttk.Label(filter_bar, text="Categoria Woo:").pack(side=tk.LEFT, padx=(12, 6))
        self.category_filter = ttk.Combobox(
            filter_bar,
            textvariable=self.category_text,
            state="readonly",
            width=22,
        )
        self.category_filter.pack(side=tk.LEFT)
        self.category_filter.bind("<<ComboboxSelected>>", lambda _event: self._render_products())

        ttk.Label(filter_bar, text="Familia:").pack(side=tk.LEFT, padx=(12, 6))
        self.family_filter = ttk.Combobox(
            filter_bar,
            textvariable=self.family_text,
            state="readonly",
            width=20,
        )
        self.family_filter.pack(side=tk.LEFT)
        self.family_filter.bind("<<ComboboxSelected>>", lambda _event: self._render_products())

        ttk.Label(filter_bar, text="Subgrupo:").pack(side=tk.LEFT, padx=(12, 6))
        self.subgroup_filter = ttk.Combobox(
            filter_bar,
            textvariable=self.subgroup_text,
            state="readonly",
            width=18,
        )
        self.subgroup_filter.pack(side=tk.LEFT)
        self.subgroup_filter.bind("<<ComboboxSelected>>", lambda _event: self._render_products())

        ttk.Label(filter_bar, text="Color:").pack(side=tk.LEFT, padx=(12, 6))
        self.color_filter = ttk.Combobox(
            filter_bar,
            textvariable=self.color_text,
            state="readonly",
            width=13,
        )
        self.color_filter.pack(side=tk.LEFT)
        self.color_filter.bind("<<ComboboxSelected>>", lambda _event: self._render_products())

        ttk.Label(filter_bar, text="Estado:").pack(side=tk.LEFT, padx=(12, 6))
        self.commercial_status_filter = ttk.Combobox(
            filter_bar,
            textvariable=self.commercial_status_text,
            state="readonly",
            width=12,
            values=["Todos", *self.COMMERCIAL_STATUSES],
        )
        self.commercial_status_filter.pack(side=tk.LEFT)
        self.commercial_status_filter.bind(
            "<<ComboboxSelected>>", lambda _event: self._render_products()
        )

        ttk.Label(filter_bar, text="Revisión:").pack(side=tk.LEFT, padx=(12, 6))
        self.reviewed_filter = ttk.Combobox(
            filter_bar,
            textvariable=self.reviewed_text,
            state="readonly",
            width=12,
            values=["Todos", "Revisados", "Pendientes"],
        )
        self.reviewed_filter.pack(side=tk.LEFT)
        self.reviewed_filter.bind("<<ComboboxSelected>>", lambda _event: self._render_products())

        ttk.Label(filter_bar, text="Pack:").pack(side=tk.LEFT, padx=(12, 6))
        self.pack_filter = ttk.Combobox(
            filter_bar,
            textvariable=self.pack_text,
            state="readonly",
            width=8,
            values=["Todos", "Sí", "No"],
        )
        self.pack_filter.pack(side=tk.LEFT)
        self.pack_filter.bind("<<ComboboxSelected>>", lambda _event: self._render_products())

        ttk.Button(filter_bar, text="Limpiar filtros", command=self._clear_search).pack(
            side=tk.LEFT, padx=(10, 0)
        )

        outer_table = tk.Frame(
            self,
            bg=C_PANEL,
            highlightbackground=C_PANEL_LINE,
            highlightthickness=1,
            bd=0,
        )
        outer_table.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0, 10))

        table_frame = ttk.Frame(outer_table, padding=(10, 10, 10, 8), style="Panel.TFrame")
        table_frame.pack(fill=tk.BOTH, expand=True)

        columns = (
            "item_kind",
            "woo_id",
            "sku",
            "name",
            "attributes_label",
            "categories_label",
            "family",
            "subgroup",
            "size",
            "materials",
            "colors",
            "commercial_status",
            "is_pack",
            "reviewed",
            "regular_price",
            "sale_price",
            "price",
            "stock_status",
            "synced_at",
        )
        self.table = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            selectmode="browse",
        )

        headings = {
            "item_kind": "Clase",
            "woo_id": "ID Woo",
            "sku": "SKU",
            "name": "Nombre",
            "attributes_label": "Variacion",
            "categories_label": "Categorias Woo",
            "family": "Familia interna",
            "subgroup": "Subgrupo",
            "size": "Medida A×L×H",
            "materials": "Materiales",
            "colors": "Colores",
            "commercial_status": "Estado comercial",
            "is_pack": "Pack",
            "reviewed": "Revisado",
            "regular_price": "Precio normal",
            "sale_price": "Precio oferta",
            "price": "Precio actual",
            "stock_status": "Stock",
            "synced_at": "Sincronizado",
        }
        widths = {
            "item_kind": 85,
            "woo_id": 78,
            "sku": 118,
            "name": 285,
            "attributes_label": 220,
            "categories_label": 220,
            "family": 155,
            "subgroup": 150,
            "size": 135,
            "materials": 130,
            "colors": 130,
            "commercial_status": 120,
            "is_pack": 70,
            "reviewed": 85,
            "regular_price": 105,
            "sale_price": 105,
            "price": 105,
            "stock_status": 95,
            "synced_at": 150,
        }

        for column in columns:
            self.table.heading(column, text=headings[column], anchor=tk.CENTER)
            self.table.column(column, width=widths[column], minwidth=60, anchor=tk.CENTER)

        self.table.tag_configure("pending", background="#fff8dc")
        self.table.tag_configure("reviewed", background="#edf7ed")
        self.table.tag_configure("outlet", background="#f3ecff")
        self.table.tag_configure("pack", background="#eef6ff")
        self.table.bind("<Double-1>", lambda _event: self._handle_table_double_click())

        y_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.table.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.table.xview)
        self.table.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        self.table.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        status_bar = ttk.Label(
            self,
            textvariable=self.status_text,
            padding=(10, 4),
            anchor=tk.W,
        )
        status_bar.pack(fill=tk.X)

    def _show_busy(self, title: str, detail: str = "") -> None:
        """Muestra una capa modal interna para evitar clics mientras el HUB trabaja.

        La capa se dibuja dentro de la ventana principal, no como ventana flotante.
        Así siempre queda anclada al HUB y evita aparecer en la posición 0,0 de Windows.
        """
        if self.busy_window is not None:
            self.busy_message_text.set(title)
            self.busy_detail_text.set(detail)
            self._position_busy_overlay()
            return

        self.busy_message_text.set(title)
        self.busy_detail_text.set(detail)
        self.configure(cursor="watch")

        overlay = tk.Frame(self, bg="#e7ebf2", bd=0, highlightthickness=0)
        overlay.place(x=0, y=0, relwidth=1, relheight=1)
        overlay.lift()

        # Captura clics y teclado para que el usuario no pulse controles mientras se trabaja.
        for sequence in ("<Button-1>", "<Button-2>", "<Button-3>", "<MouseWheel>", "<Key>"):
            overlay.bind(sequence, lambda _event: "break")

        card = tk.Frame(
            overlay,
            bg=C_PANEL,
            highlightbackground=C_PANEL_LINE,
            highlightthickness=1,
            bd=0,
        )
        card.place(relx=0.5, rely=0.5, anchor=tk.CENTER, width=590, height=235)

        ttk.Label(
            card,
            textvariable=self.busy_message_text,
            style="Title.TLabel",
            anchor=tk.CENTER,
            wraplength=520,
            justify=tk.CENTER,
        ).pack(fill=tk.X, padx=28, pady=(26, 10))
        ttk.Label(
            card,
            textvariable=self.busy_detail_text,
            anchor=tk.CENTER,
            justify=tk.CENTER,
            wraplength=510,
        ).pack(fill=tk.X, padx=30, pady=(0, 18))

        progress = ttk.Progressbar(card, mode="indeterminate")
        progress.pack(fill=tk.X, padx=46, pady=(0, 14))
        progress.start(12)

        ttk.Label(
            card,
            text="Por favor, espera. El HUB está trabajando y los botones quedan protegidos hasta terminar.",
            anchor=tk.CENTER,
            justify=tk.CENTER,
            wraplength=510,
        ).pack(fill=tk.X, padx=30)

        self.busy_window = overlay
        self.busy_progress = progress
        self._position_busy_overlay()
        self._busy_configure_binding = self.bind("<Configure>", lambda _event: self._position_busy_overlay(), add="+")
        overlay.focus_set()

    def _position_busy_overlay(self) -> None:
        if self.busy_window is None:
            return
        self.busy_window.place_configure(x=0, y=0, relwidth=1, relheight=1)
        self.busy_window.lift()

    def _update_busy(self, title: str | None = None, detail: str | None = None) -> None:
        if title is not None:
            self.busy_message_text.set(title)
        if detail is not None:
            self.busy_detail_text.set(detail)
        if self.busy_window is not None:
            self.busy_window.lift()

    def _hide_busy(self) -> None:
        if self.busy_progress is not None:
            try:
                self.busy_progress.stop()
            except tk.TclError:
                pass
            self.busy_progress = None
        if self._busy_configure_binding:
            try:
                self.unbind("<Configure>", self._busy_configure_binding)
            except tk.TclError:
                pass
            self._busy_configure_binding = None
        if self.busy_window is not None:
            self.busy_window.destroy()
            self.busy_window = None
        self._busy_job = None
        self.configure(cursor="")

    def _is_busy(self) -> bool:
        if self.busy_window is not None:
            self.busy_window.lift()
            return True
        return False

    def _load_products(self) -> None:
        self.store.init_schema()
        self.catalog_items = [dict(row) for row in self.store.list_catalog_items()]
        categories = ["Todas"] + [row["name"] for row in self.store.list_categories()]
        self.category_filter.configure(values=categories)
        if self.category_text.get() not in categories:
            self.category_text.set("Todas")
        self._refresh_classification_filters()
        self._render_products()

    def _refresh_classification_filters(self) -> None:
        families = sorted({str(item.get("family") or "") for item in self.catalog_items if item.get("family")})
        subgroups = sorted({str(item.get("subgroup") or "") for item in self.catalog_items if item.get("subgroup")})
        colors = sorted(
            {
                color.strip()
                for item in self.catalog_items
                for color in str(item.get("colors") or "").split(",")
                if color.strip()
            }
        )
        self.family_filter.configure(values=["Todas", *families])
        self.subgroup_filter.configure(values=["Todos", *subgroups])
        self.color_filter.configure(values=["Todos", *colors])
        if self.family_text.get() not in ["Todas", *families]:
            self.family_text.set("Todas")
        if self.subgroup_text.get() not in ["Todos", *subgroups]:
            self.subgroup_text.set("Todos")
        if self.color_text.get() not in ["Todos", *colors]:
            self.color_text.set("Todos")

    def _render_products(self) -> None:
        query = self._normalize(self.search_text.get())
        filtered = self.catalog_items
        category = self.category_text.get()
        if category and category != "Todas":
            filtered = [
                item
                for item in filtered
                if category in str(item.get("categories_label") or "").split(" / ")
            ]
        family = self.family_text.get()
        if family and family != "Todas":
            filtered = [item for item in filtered if str(item.get("family") or "") == family]
        subgroup = self.subgroup_text.get()
        if subgroup and subgroup != "Todos":
            filtered = [item for item in filtered if str(item.get("subgroup") or "") == subgroup]
        color = self.color_text.get()
        if color and color != "Todos":
            filtered = [
                item
                for item in filtered
                if color in [part.strip() for part in str(item.get("colors") or "").split(",")]
            ]
        commercial_status = self.commercial_status_text.get()
        if commercial_status and commercial_status != "Todos":
            filtered = [
                item
                for item in filtered
                if str(item.get("commercial_status") or "Normal") == commercial_status
            ]
        reviewed = self.reviewed_text.get()
        if reviewed == "Revisados":
            filtered = [item for item in filtered if int(item.get("reviewed") or 0) == 1]
        elif reviewed == "Pendientes":
            filtered = [item for item in filtered if int(item.get("reviewed") or 0) == 0]
        pack = self.pack_text.get()
        if pack == "Sí":
            filtered = [item for item in filtered if int(item.get("is_pack") or 0) == 1]
        elif pack == "No":
            filtered = [item for item in filtered if int(item.get("is_pack") or 0) == 0]
        if query:
            filtered = [
                item
                for item in filtered
                if query
                in self._normalize(
                    " ".join(
                        str(item.get(key) or "")
                        for key in (
                            "name",
                            "sku",
                            "woo_id",
                            "parent_woo_id",
                            "attributes_label",
                            "categories_label",
                            "family",
                            "subgroup",
                            "size",
                            "materials",
                            "colors",
                            "commercial_status",
                        )
                    )
                )
            ]

        self.filtered_items = filtered
        self.table.delete(*self.table.get_children())
        for item in filtered:
            tags = []
            if int(item.get("reviewed") or 0) == 1:
                tags.append("reviewed")
            else:
                tags.append("pending")
            if str(item.get("commercial_status") or "") == "Outlet":
                tags = ["outlet"]
            elif int(item.get("is_pack") or 0) == 1:
                tags = ["pack"]
            self.table.insert(
                "",
                tk.END,
                values=(
                    item.get("item_kind") or "",
                    item.get("woo_id") or "",
                    item.get("sku") or "",
                    item.get("name") or "",
                    item.get("attributes_label") or "",
                    item.get("categories_label") or "",
                    item.get("family") or "Otros / Sin clasificar",
                    item.get("subgroup") or "Sin subgrupo",
                    item.get("size") or "Sin medida",
                    item.get("materials") or "",
                    item.get("colors") or "",
                    item.get("commercial_status") or "Normal",
                    "Sí" if int(item.get("is_pack") or 0) == 1 else "No",
                    "Sí" if int(item.get("reviewed") or 0) == 1 else "No",
                    item.get("regular_price") or "",
                    item.get("sale_price") or "",
                    item.get("price") or "",
                    item.get("stock_status") or "",
                    item.get("synced_at") or "",
                ),
                tags=tags,
            )

        auto_fit_tree_columns(self.table)

        total = len(self.catalog_items)
        visible = len(filtered)
        pending = sum(1 for item in self.catalog_items if int(item.get("reviewed") or 0) == 0)
        outlets = sum(1 for item in self.catalog_items if str(item.get("commercial_status") or "") == "Outlet")
        packs = sum(1 for item in self.catalog_items if int(item.get("is_pack") or 0) == 1)
        self.status_text.set(
            f"Mostrando {visible} de {total} articulos locales. "
            f"Pendientes de revisar: {pending} | Outlet: {outlets} | Packs: {packs}. "
            "Organización local: no modifica WooCommerce."
        )

    def _clear_search(self) -> None:
        self.search_text.set("")
        self.category_text.set("Todas")
        self.family_text.set("Todas")
        self.subgroup_text.set("Todos")
        self.color_text.set("Todos")
        self.commercial_status_text.set("Todos")
        self.reviewed_text.set("Todos")
        self.pack_text.set("Todos")
        self._render_products()

    def _sync_from_woocommerce(self) -> None:
        if self._is_busy():
            return
        self._busy_job = "sync"
        if self.update_button is not None:
            self.update_button.configure(state=tk.DISABLED)
        self.status_text.set("Actualizando productos desde WooCommerce...")
        self._show_busy(
            "Actualizando desde WooCommerce",
            "Sincronizando categorías, productos y variaciones desde WooCommerce. "
            "No cierres la ventana ni pulses otros botones.",
        )

        thread = threading.Thread(target=self._sync_worker, daemon=True)
        thread.start()
        self.after(150, self._poll_sync_events)

    def _sync_worker(self) -> None:
        try:
            result = sync_products(self.settings)
        except OperationBlockedError as exc:
            self.events.put(("sync_error", str(exc)))
        except WooCommerceError as exc:
            self.events.put(("sync_error", str(exc)))
        except Exception as exc:  # noqa: BLE001
            self.events.put(("sync_error", f"Error inesperado: {exc}"))
        else:
            self.events.put(("sync_done", result))

    def _poll_sync_events(self) -> None:
        try:
            event, payload = self.events.get_nowait()
        except queue.Empty:
            self.after(150, self._poll_sync_events)
            return

        if self.update_button is not None:
            self.update_button.configure(state=tk.NORMAL)
        self._hide_busy()
        if event == "sync_error":
            self.status_text.set("No se pudo actualizar desde WooCommerce.")
            messagebox.showerror("Actualizacion fallida", str(payload), parent=self)
            return

        self._load_products()
        self.status_text.set(
            "Actualizacion completada. "
            f"Categorias: {payload.categories_imported}. "
            f"Productos: {payload.imported}. "
            f"Variaciones: {payload.variations_imported}. "
            f"Total local: {payload.total_local + payload.total_variations}."
        )

    def _auto_classify_catalog(self) -> None:
        if self._is_busy():
            return
        if not self.catalog_items:
            messagebox.showinfo("Sin articulos", "No hay articulos locales para clasificar.", parent=self)
            return
        self._busy_job = "classify"
        if self.auto_classify_button is not None:
            self.auto_classify_button.configure(state=tk.DISABLED)
        self.status_text.set("Clasificando catalogo local...")
        self._show_busy(
            "Clasificando catálogo local",
            "Analizando nombres, SKU, variaciones y categorías guardadas en el HUB. "
            "Esto no modifica WooCommerce.",
        )
        items_snapshot = list(self.catalog_items)
        thread = threading.Thread(
            target=self._auto_classify_worker,
            args=(items_snapshot,),
            daemon=True,
        )
        thread.start()
        self.after(150, self._poll_classification_events)

    def _auto_classify_worker(self, items: list[dict[str, object]]) -> None:
        try:
            reviewed_keys = self.store.existing_reviewed_classification_keys()
            created = 0
            skipped = 0
            total = len(items)
            for index, item in enumerate(items, start=1):
                key = (str(item.get("item_kind") or ""), int(item.get("woo_id") or 0))
                if key in reviewed_keys:
                    skipped += 1
                else:
                    classification = self._classify_item(item)
                    self.store.upsert_product_classification(classification)
                    created += 1
                if index == 1 or index % 25 == 0 or index == total:
                    self.events.put(("classify_progress", (index, total, created, skipped)))
        except Exception as exc:  # noqa: BLE001
            self.events.put(("classify_error", f"Error inesperado: {exc}"))
        else:
            self.events.put(("classify_done", (created, skipped)))

    def _poll_classification_events(self) -> None:
        try:
            event, payload = self.events.get_nowait()
        except queue.Empty:
            self.after(150, self._poll_classification_events)
            return

        if event == "classify_progress":
            index, total, created, skipped = payload
            self._update_busy(
                detail=(
                    f"Revisando {index} de {total} articulos. "
                    f"Clasificados: {created} | Manuales preservados: {skipped}."
                )
            )
            self.after(150, self._poll_classification_events)
            return

        if self.auto_classify_button is not None:
            self.auto_classify_button.configure(state=tk.NORMAL)
        self._hide_busy()
        if event == "classify_error":
            self.status_text.set("No se pudo clasificar el catalogo local.")
            messagebox.showerror("Clasificacion fallida", str(payload), parent=self)
            return

        created, skipped = payload
        self._load_products()
        self.status_text.set(
            f"Clasificacion completada. Clasificados/actualizados: {created}. "
            f"Revisados manuales preservados: {skipped}."
        )
        messagebox.showinfo(
            "Clasificacion completada",
            f"Clasificados/actualizados: {created}\nRevisados manuales preservados: {skipped}",
            parent=self,
        )

    def _classify_item(self, item: dict[str, object]) -> dict[str, Any]:
        """Clasifica usando nombre + datos ricos sincronizados desde WooCommerce.

        La sincronizacion ya guarda raw_json de productos y variaciones. Aqui lo usamos
        solo para organizar localmente el HUB: descripciones, descripcion corta,
        atributos, opciones de variacion, categorias, etiquetas, dimensiones e imagenes.
        No se escribe nada en WooCommerce.
        """
        text_raw = self._classification_text(item)
        text = self._normalize(text_raw)
        identity_raw = " ".join(
            str(item.get(key) or "")
            for key in ("name", "sku", "categories_label", "type")
        )
        identity_text = self._normalize(identity_raw)
        family = "Otros / Sin clasificar"
        subgroup = "Sin subgrupo"
        commercial_status = "Normal"
        is_pack = 0
        notes = "Auto: nombre + SKU + categorias + atributos + descripciones Woo. Medida prioriza la variacion y Ancho x Largo x Alto. Revisar antes de usar en cambios masivos."

        if "outlet" in text:
            commercial_status = "Outlet"
        elif any(word in text for word in ("oferta", "rebaja", "descuento", "liquidacion", "liquidación")):
            commercial_status = "Oferta"

        has_pack = any(word in identity_text for word in ("pack", "combo", "conjunto", "combinacion", "combinación"))
        has_futon = self._has_futon_intent(identity_text)
        has_clear_complement = self._has_clear_complement_intent(identity_text)
        has_sofa_cama = any(word in identity_text for word in ("sofa cama", "sofa-cama", "sofá cama"))

        # Prioridad importante:
        # - Si el producto es Futón, debe ganar Futones aunque aparezcan textos
        #   como "sin cojines", "con cojines", "funda de transporte" o categorias
        #   heredadas que mencionen sofa cama.
        # - Solo mandamos a Complementos cuando el item es claramente funda, topper,
        #   almohada, cojin, mesita, etc. y no un futón real.
        if has_clear_complement:
            family = "Complementos"
            subgroup = self._complement_subgroup(identity_text)
            is_pack = 1 if has_pack else 0
        elif has_pack and any(word in identity_text for word in ("futon", "futón", "tatami", "sofa cama", "sofá cama", "cama japonesa")):
            family = "Ofertas / Packs"
            is_pack = 1
            if has_sofa_cama and not has_futon:
                subgroup = "Pack Sofá Cama"
            elif "tatami" in identity_text and has_futon:
                futon_group = self._futon_subgroup(text)
                subgroup = f"Pack Futón {futon_group} + Tatami" if futon_group != "Otros" else "Pack Futón + Tatami"
            elif "tatami" in identity_text or has_futon:
                subgroup = "Pack Futón + Tatami"
            else:
                subgroup = "Otros"
        elif has_futon:
            family = "Futones"
            subgroup = self._futon_subgroup(text)
        elif "tatami" in text:
            family = "Tatamis"
            subgroup = "Tatami"
        elif has_sofa_cama:
            family = "Sofás Cama"
            subgroup = "Sofá Cama"
        elif "cama japonesa" in text or "camas japonesas" in text or "base para tatami" in text:
            family = "Camas Japonesas"
            subgroup = "Cama Japonesa"
        elif has_pack:
            family = "Ofertas / Packs"
            is_pack = 1
            subgroup = "Otros"
        elif commercial_status == "Outlet":
            family = "Outlet"
            notes = "Outlet detectado sin familia real clara. Revisar manualmente."

        materials = self._detect_materials(text)
        colors = self._detect_colors(text, item)
        size = self._detect_size(text, item)
        needs_review = 1
        return {
            "item_kind": item.get("item_kind") or "product",
            "item_woo_id": int(item.get("woo_id") or 0),
            "sku": item.get("sku") or None,
            "family": family,
            "subgroup": subgroup,
            "size": size,
            "materials": materials,
            "colors": colors,
            "commercial_status": commercial_status,
            "is_pack": is_pack,
            "reviewed": 0,
            "classification_source": "auto",
            "needs_review": needs_review,
            "notes": notes,
        }

    def _classification_text(self, item: dict[str, object]) -> str:
        parts = [
            str(item.get(key) or "")
            for key in ("name", "sku", "attributes_label", "categories_label", "type")
        ]
        # En variaciones, el raw_json propio debe tener prioridad sobre el padre.
        # Si el padre mide 140x200x14.5 pero una variacion trae otra medida,
        # no debemos copiar la medida del padre a todas.
        for raw_key in ("raw_json", "parent_raw_json"):
            raw = item.get(raw_key)
            if not raw:
                continue
            try:
                payload = json.loads(str(raw))
            except (TypeError, ValueError, json.JSONDecodeError):
                continue
            parts.extend(self._extract_woo_text_parts(payload))
        return " ".join(part for part in parts if part)

    def _extract_woo_text_parts(self, payload: dict[str, Any]) -> list[str]:
        parts: list[str] = []
        for key in ("name", "slug", "sku", "description", "short_description"):
            value = payload.get(key)
            if value:
                parts.append(self._strip_html(str(value)))
        dimensions = payload.get("dimensions")
        if isinstance(dimensions, dict):
            parts.extend(str(value) for value in dimensions.values() if value)
        for collection_key in ("categories", "tags", "brands", "attributes", "default_attributes"):
            values = payload.get(collection_key)
            if isinstance(values, list):
                for value in values:
                    if isinstance(value, dict):
                        attr_name = str(value.get("name") or "")
                        attr_slug = str(value.get("slug") or "")
                        parts.append(attr_name)
                        parts.append(attr_slug)
                        option = value.get("option")
                        if option:
                            option_text = str(option)
                            parts.append(option_text)
                            if attr_name:
                                parts.append(f"{attr_name}: {option_text}")
                        options = value.get("options")
                        if isinstance(options, list):
                            for option in options:
                                if option:
                                    option_text = str(option)
                                    parts.append(option_text)
                                    if attr_name:
                                        parts.append(f"{attr_name}: {option_text}")
                    elif value:
                        parts.append(str(value))
        image = payload.get("image")
        if isinstance(image, dict):
            parts.append(str(image.get("name") or ""))
            parts.append(str(image.get("alt") or ""))
        images = payload.get("images")
        if isinstance(images, list):
            for image_data in images[:3]:
                if isinstance(image_data, dict):
                    parts.append(str(image_data.get("name") or ""))
                    parts.append(str(image_data.get("alt") or ""))
        return [part for part in parts if part]

    def _raw_payloads(self, item: dict[str, object]) -> list[dict[str, Any]]:
        payloads: list[dict[str, Any]] = []
        # En variaciones, el raw_json propio debe tener prioridad sobre el padre.
        # Si el padre mide 140x200x14.5 pero una variacion trae otra medida,
        # no debemos copiar la medida del padre a todas.
        for raw_key in ("raw_json", "parent_raw_json"):
            raw = item.get(raw_key)
            if not raw:
                continue
            try:
                payload = json.loads(str(raw))
            except (TypeError, ValueError, json.JSONDecodeError):
                continue
            if isinstance(payload, dict):
                payloads.append(payload)
        return payloads

    def _attribute_values(self, item: dict[str, object], name_contains: tuple[str, ...]) -> list[str]:
        values: list[str] = []
        needles = tuple(self._normalize(value) for value in name_contains)
        for payload in self._raw_payloads(item):
            for key in ("attributes", "default_attributes"):
                attrs = payload.get(key)
                if not isinstance(attrs, list):
                    continue
                for attr in attrs:
                    if not isinstance(attr, dict):
                        continue
                    attr_name = self._normalize(" ".join(str(attr.get(k) or "") for k in ("name", "slug")))
                    if not any(needle in attr_name for needle in needles):
                        continue
                    option = attr.get("option")
                    if option:
                        option_text = str(option)
                        values.append(option_text)
                        if attr_name:
                            values.append(f"{attr_name}: {option_text}")
                    options = attr.get("options")
                    if isinstance(options, list):
                        for option in options:
                            if option:
                                option_text = str(option)
                                values.append(option_text)
                                if attr_name:
                                    values.append(f"{attr_name}: {option_text}")
        return values

    def _strip_html(self, text: str) -> str:
        return re.sub(r"<[^>]+>", " ", text)

    def _has_futon_intent(self, text: str) -> bool:
        """Detecta futones reales sin dejarse arrastrar por categorias secundarias.

        Las fundas/covers/topper/almohadas para futón deben ser Complementos,
        aunque contengan la palabra futón. En cambio, "futón sin cojines" o
        "futón portátil con funda" siguen siendo futones reales.
        """
        if "futon" not in text and "futón" not in text:
            return False
        normalized = self._normalize(text)
        non_futon_patterns = (
            r"\bfundas?\s+(?:de\s+|para\s+)?futones?\b",
            r"\bcovers?\s+(?:de\s+|para\s+)?futones?\b",
            r"\btopper\s+(?:de\s+|para\s+)?futones?\b",
            r"\balmohadas?\s+(?:de\s+|para\s+)?futones?\b",
            r"\bpillows?\s+(?:de\s+|para\s+)?futones?\b",
            r"\bcojines?\s+(?:de\s+|para\s+)?futones?\b",
            r"\bfutones?\s+(?:con\s+)?fundas?\b",
        )
        return not any(re.search(pattern, normalized) for pattern in non_futon_patterns)

    def _has_clear_complement_intent(self, text: str) -> bool:
        """Complementos claros, evitando falsos positivos de futones con cojines."""
        if "mesita" in text or "topper" in text or "almohada" in text or "pillow" in text:
            return True
        if "cojin" in text or "cojín" in text:
            # "Futon algodon sin cojines" no es complemento; sigue siendo futón.
            return not self._has_futon_intent(text)
        if "funda" in text or "cover" in text:
            # "Futón portátil con funda" o "futón sin funda" no debe ir a Fundas.
            if self._has_futon_intent(text) and any(word in text for word in ("portatil", "portátil", "portable")):
                return False
            normalized = self._normalize(text)
            return not self._has_futon_intent(text) or any(
                re.search(pattern, normalized)
                for pattern in (
                    r"\bfundas?\s+(?:de\s+|para\s+)?futones?\b",
                    r"\bcovers?\s+(?:de\s+|para\s+)?futones?\b",
                    r"\bfundas?\s+(?:de\s+|para\s+)?tatamis?\b",
                    r"\bfundas?\s+(?:de\s+|para\s+)?cojines?\b",
                )
            )
        return "complemento" in text and not self._has_futon_intent(text)

    def _futon_subgroup(self, text: str) -> str:
        has_algodon = "algodon" in text or "algodón" in text
        has_latex = "latex" in text or "látex" in text
        has_coco = "coco" in text
        has_lana = "lana" in text
        is_portable = any(word in text for word in ("portatil", "portátil", "portable"))
        latex_layers = self._detect_latex_layers(text)

        if is_portable:
            return "Portátil"
        if has_algodon and has_latex and latex_layers == 2:
            return "Algodón + 2 capas Látex"
        if has_algodon and has_latex and latex_layers == 1:
            return "Algodón + 1 capa Látex"
        if has_algodon and has_latex:
            return "Algodón + Látex"
        if has_algodon and has_coco:
            return "Algodón + Coco"
        if has_algodon and has_lana:
            return "Algodón + Lana"
        if has_algodon:
            return "Algodón"
        if has_latex and latex_layers == 2:
            return "2 capas Látex"
        if has_latex and latex_layers == 1:
            return "1 capa Látex"
        if has_latex:
            return "Látex"
        if has_coco:
            return "Coco"
        if "medida" in text or "a medida" in text:
            return "A medida"
        return "Otros"

    def _detect_latex_layers(self, text: str) -> int | None:
        """Detecta si el futón declara una o dos capas de látex."""
        normalized = self._normalize(text)
        two_layer_patterns = (
            r"(?:2|dos)\s*(?:capas?|layers?)\s*(?:de\s*)?latex",
            r"latex\s*(?:2|dos)\s*(?:capas?|layers?)",
            r"doble\s*(?:capa\s*)?latex",
        )
        one_layer_patterns = (
            r"(?:1|una)\s*(?:capa|layer)\s*(?:de\s*)?latex",
            r"latex\s*(?:1|una)\s*(?:capa|layer)",
            r"simple\s*(?:capa\s*)?latex",
        )
        if any(re.search(pattern, normalized) for pattern in two_layer_patterns):
            return 2
        if any(re.search(pattern, normalized) for pattern in one_layer_patterns):
            return 1
        return None

    def _complement_subgroup(self, text: str) -> str:
        if "mesita" in text:
            return "Mesitas"
        cojin_pos = min([pos for pos in (text.find("cojin"), text.find("cojín")) if pos >= 0], default=-1)
        almohada_pos = min([pos for pos in (text.find("almohada"), text.find("pillow")) if pos >= 0], default=-1)
        if cojin_pos >= 0 or almohada_pos >= 0:
            if almohada_pos >= 0 and (cojin_pos < 0 or almohada_pos < cojin_pos):
                return "Almohadas"
            return "Cojines"
        if "funda" in text or "cover" in text:
            return "Fundas"
        if "topper" in text:
            return "Toppers"
        return "Accesorios"

    def _detect_materials(self, text: str) -> str:
        materials = []
        checks = [
            # Materiales relevantes para la organizacion de FutonEspai.
            # No se incluyen Madera, Tatami ni Bambu: en la tienda actuan como
            # familias/tipos de producto, no como materiales operativos para futones.
            ("Algodón", ("algodon", "algodón", "cotton")),
            ("Látex", ("latex", "látex")),
            ("Lana", ("lana", "wool")),
            ("Coco", ("coco", "coconut")),
        ]
        for label, words in checks:
            if any(word in text for word in words):
                materials.append(label)
        return ", ".join(dict.fromkeys(materials))

    def _detect_colors(self, text: str, item: dict[str, object]) -> str:
        colors: list[str] = []
        for value in self._attribute_values(item, ("color", "colour", "acabado")):
            clean = value.strip()
            if clean:
                colors.append(clean)
        checks = [
            ("Crudo", ("crudo", "blanco crudo")),
            ("Natural", ("natural",)),
            ("Blanco", ("blanco", "white")),
            ("Negro", ("negro", "black")),
            ("Gris", ("gris", "grey", "gray")),
            ("Antracita", ("antracita",)),
            ("Marrón", ("marron", "marrón", "brown", "chocolate")),
            ("Beige", ("beige", "beig")),
            ("Azul", ("azul", "blue")),
            ("Verde", ("verde", "green")),
            ("Rojo", ("rojo", "red")),
            ("Naranja", ("naranja", "orange")),
            ("Amarillo", ("amarillo", "yellow", "mostaza")),
            ("Rosa", ("rosa", "pink")),
            ("Morado", ("morado", "violeta", "purple")),
            ("Terracota", ("terracota",)),
            ("Marfil", ("marfil", "ivory")),
            ("Visón", ("vison", "visón")),
        ]
        for label, words in checks:
            if any(word in text for word in words):
                colors.append(label)
        # Evita duplicados preservando orden, y no deja listas infinitas de opciones.
        return ", ".join(list(dict.fromkeys(colors))[:8])

    def _detect_size(self, text: str, item: dict[str, object] | None = None) -> str:
        """Devuelve la medida completa Ancho x Largo x Alto cuando sea posible.

        Regla clave para WooCommerce:
        - En variaciones, manda la medida de la variacion: attributes_label / raw_json propio.
        - En productos padre, manda el nombre del producto. Si el padre tiene muchas
          medidas como opciones de variacion, no cogemos la primera al azar: marcamos
          "Varias medidas" salvo que el nombre ya traiga una medida concreta.
        """
        item_kind = str(item.get("item_kind") or "product") if item is not None else "product"
        primary_candidates: list[str] = []
        secondary_candidates: list[str] = []

        if item is not None:
            name_candidate = str(item.get("name") or "")
            sku_candidate = str(item.get("sku") or "")
            attrs_label = str(item.get("attributes_label") or "")
            if item_kind == "variation":
                primary_candidates.extend([attrs_label, name_candidate, sku_candidate])
                primary_candidates.extend(self._attribute_values(
                    item,
                    (
                        "medida", "medidas", "medidas futon", "medida futon",
                        "dimension", "dimensiones", "tamaño", "tamano",
                        "ancho", "largo", "alto", "altura", "grosor", "espesor",
                    ),
                ))
                primary_candidates.extend(self._dimension_candidates(item))
            else:
                primary_candidates.extend([name_candidate, sku_candidate])
                primary_candidates.extend(self._dimension_candidates(item))
                secondary_candidates.extend(self._attribute_values(
                    item,
                    (
                        "medida", "medidas", "medidas futon", "medida futon",
                        "dimension", "dimensiones", "tamaño", "tamano",
                        "ancho", "largo", "alto", "altura", "grosor", "espesor",
                    ),
                ))
        primary_candidates.append(text)

        for candidate in primary_candidates:
            found = self._find_dimension_3d(str(candidate))
            if found:
                return found

        if item_kind != "variation" and secondary_candidates:
            found_options = [self._find_dimension_3d(str(candidate)) for candidate in secondary_candidates]
            found_options = [value for value in found_options if value]
            unique_options = list(dict.fromkeys(found_options))
            if len(unique_options) == 1:
                return unique_options[0]
            if len(unique_options) > 1:
                return "Varias medidas"

        all_candidates = primary_candidates + secondary_candidates
        two_part: tuple[str, str] | None = None
        height: str | None = None
        for candidate in all_candidates:
            normalized_candidate = str(candidate).replace("×", "x")
            if two_part is None:
                match_2d = re.search(
                    r"(\d{2,3}(?:[.,]\d+)?)\s*[xX]\s*(\d{2,3}(?:[.,]\d+)?)",
                    normalized_candidate,
                )
                if match_2d:
                    two_part = (match_2d.group(1), match_2d.group(2))
            if height is None:
                match_height = re.search(
                    r"(?:alto|altura|grosor|espesor)\s*[:=]?\s*(\d{1,3}(?:[.,]\d+)?)",
                    self._normalize(normalized_candidate),
                )
                if match_height:
                    height = match_height.group(1)
        if two_part and height:
            return self._format_dimension(two_part[0], two_part[1], height)
        if two_part:
            return f"{self._clean_dimension_number(two_part[0])}x{self._clean_dimension_number(two_part[1])}"

        if "a medida" in text or "medida especial" in text:
            return "A medida"
        return "Sin medida"

    def _find_dimension_3d(self, candidate: str) -> str | None:
        normalized_candidate = candidate.replace("×", "x")
        match = re.search(
            r"(\d{2,3}(?:[.,]\d+)?)\s*[xX]\s*(\d{2,3}(?:[.,]\d+)?)\s*[xX]\s*(\d{1,3}(?:[.,]\d+)?)",
            normalized_candidate,
        )
        if not match:
            return None
        return self._format_dimension(match.group(1), match.group(2), match.group(3))

    def _dimension_candidates(self, item: dict[str, object]) -> list[str]:
        candidates: list[str] = []
        for payload in self._raw_payloads(item):
            dimensions = payload.get("dimensions")
            if not isinstance(dimensions, dict):
                continue
            width = dimensions.get("width") or dimensions.get("ancho")
            length = dimensions.get("length") or dimensions.get("largo")
            height = dimensions.get("height") or dimensions.get("alto") or dimensions.get("altura")
            if width and length and height:
                candidates.append(f"{width}x{length}x{height}")
            elif width and length:
                candidates.append(f"{width}x{length}")
            candidates.extend(str(value) for value in dimensions.values() if value)
        return candidates

    def _format_dimension(self, width: str, length: str, height: str) -> str:
        return (
            f"{self._clean_dimension_number(width)}x"
            f"{self._clean_dimension_number(length)}x"
            f"{self._clean_dimension_number(height)}"
        )

    def _clean_dimension_number(self, value: str) -> str:
        clean = str(value).strip().replace(",", ".")
        try:
            number = float(clean)
        except ValueError:
            return clean
        if number.is_integer():
            return str(int(number))
        return f"{number:g}"

    def _open_classification_dialog(self) -> None:
        item = self._selected_catalog_item()
        if not item:
            messagebox.showwarning("Seleccion requerida", "Selecciona un articulo primero.", parent=self)
            return
        ClassificationDialog(self, self.store, item, self._load_products)

    def _mark_selected_reviewed(self) -> None:
        item = self._selected_catalog_item()
        if not item:
            messagebox.showwarning("Seleccion requerida", "Selecciona un articulo primero.", parent=self)
            return
        item_kind = str(item.get("item_kind") or "product")
        item_woo_id = int(item.get("woo_id") or 0)
        if item_woo_id <= 0:
            return
        if str(item.get("classification_source") or "") == "pendiente":
            classification = self._classify_item(item)
            classification["reviewed"] = 1
            classification["needs_review"] = 0
            classification["classification_source"] = "manual"
            self.store.upsert_product_classification(classification)
        else:
            self.store.mark_classification_reviewed(item_kind, item_woo_id)
        self._load_products()

    def _export_classified_catalog(self) -> None:
        if not self.filtered_items:
            messagebox.showwarning("Sin datos", "No hay articulos visibles para exportar.", parent=self)
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror(
                "Falta openpyxl",
                "Instala openpyxl para exportar: pip install openpyxl",
                parent=self,
            )
            return

        default_name = "catalogo_clasificado_futonespai.xlsx"
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Exportar catálogo clasificado",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Catalogo clasificado"
        columns = [
            ("Clase", "item_kind"),
            ("ID Woo", "woo_id"),
            ("SKU", "sku"),
            ("Nombre", "name"),
            ("Variación", "attributes_label"),
            ("Categorías Woo", "categories_label"),
            ("Familia interna", "family"),
            ("Subgrupo", "subgroup"),
            ("Medida A×L×H", "size"),
            ("Materiales", "materials"),
            ("Colores", "colors"),
            ("Estado comercial", "commercial_status"),
            ("Es pack", "is_pack"),
            ("Revisado", "reviewed"),
            ("Precio normal", "regular_price"),
            ("Precio oferta", "sale_price"),
            ("Precio actual", "price"),
            ("Stock", "stock_status"),
            ("Notas clasificación", "classification_notes"),
        ]
        for col_idx, (label, _key) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center")
        for row_idx, item in enumerate(self.filtered_items, start=2):
            for col_idx, (_label, key) in enumerate(columns, start=1):
                value = item.get(key) or ""
                if key in ("is_pack", "reviewed"):
                    value = "Sí" if int(item.get(key) or 0) == 1 else "No"
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 42)
        wb.save(path)
        messagebox.showinfo("Exportacion completada", f"Catalogo exportado:\n{path}", parent=self)

    def _handle_table_double_click(self) -> None:
        if self.mode == "prices":
            self._open_price_dialog()
        else:
            self._open_classification_dialog()

    def _open_pack_dialog(self) -> None:
        PackDialog(self, self.store, self.catalog_items)

    def _open_price_dialog(self) -> None:
        PriceChangeDialog(
            self,
            self.store,
            self.catalog_items,
            self._selected_catalog_item(),
            on_added=self._price_proposal_added,
        )

    def _open_bulk_price_dialog(self) -> None:
        if not self.filtered_items:
            messagebox.showinfo(
                "Sin articulos visibles",
                "Aplica un filtro o limpia la busqueda para tener articulos visibles antes de crear una propuesta masiva.",
                parent=self,
            )
            return
        BulkPriceChangeDialog(
            self,
            self.store,
            list(self.filtered_items),
            on_added=self._price_proposal_added,
        )

    def _price_proposal_added(self, added_count: int) -> None:
        self.status_text.set(
            f"Propuesta actualizada. Articulos/variaciones añadidos: {added_count}. "
            "No se ha modificado WooCommerce."
        )

    def _open_publish_dialog(self) -> None:
        PublishDialog(self, self.settings, self.store, on_done=self._load_products)

    def _selected_catalog_item(self) -> dict[str, object] | None:
        selection = self.table.selection()
        if not selection:
            return None
        values = self.table.item(selection[0], "values")
        if len(values) < 2:
            return None
        item_kind = str(values[0])
        try:
            woo_id = int(values[1])
        except ValueError:
            return None
        for item in self.catalog_items:
            if item.get("item_kind") == item_kind and int(item["woo_id"]) == woo_id:
                return item
        return None

    def _normalize(self, text: str) -> str:
        normalized = unicodedata.normalize("NFKD", text)
        return "".join(char for char in normalized if not unicodedata.combining(char)).lower()


class ClassificationDialog(tk.Toplevel):
    def __init__(
        self,
        parent: ProductTableApp,
        store: ProductStore,
        item: dict[str, object],
        on_saved: object,
    ) -> None:
        super().__init__(parent)
        self.parent = parent
        self.store = store
        self.item = item
        self.on_saved = on_saved
        self.title("Editar clasificación local")
        self.geometry("620x560")
        self.minsize(560, 500)
        self.transient(parent)
        self.grab_set()

        suggestion = parent._classify_item(item)
        self.family_text = tk.StringVar(value=str(item.get("family") or suggestion["family"]))
        self.subgroup_text = tk.StringVar(value=str(item.get("subgroup") or suggestion["subgroup"]))
        self.size_text = tk.StringVar(value=str(item.get("size") or suggestion["size"]))
        self.materials_text = tk.StringVar(value=str(item.get("materials") or suggestion["materials"]))
        self.colors_text = tk.StringVar(value=str(item.get("colors") or suggestion.get("colors", "")))
        self.commercial_status_text = tk.StringVar(
            value=str(item.get("commercial_status") or suggestion["commercial_status"])
        )
        self.is_pack_var = tk.BooleanVar(value=bool(int(item.get("is_pack") or suggestion["is_pack"])))
        self.reviewed_var = tk.BooleanVar(value=bool(int(item.get("reviewed") or 0)))
        self.notes_text = tk.StringVar(value=str(item.get("classification_notes") or ""))

        self._build_layout()

    def _build_layout(self) -> None:
        body = ttk.Frame(self, padding=(16, 14, 16, 10))
        body.pack(fill=tk.BOTH, expand=True)

        title = str(self.item.get("name") or "")
        variation = str(self.item.get("attributes_label") or "")
        sku = str(self.item.get("sku") or "")
        ttk.Label(body, text="Producto", style="Title.TLabel").grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 6)
        )
        ttk.Label(body, text=f"SKU: {sku or 'Sin SKU'} | ID Woo: {self.item.get('woo_id')}").grid(
            row=1, column=0, columnspan=2, sticky="w"
        )
        ttk.Label(body, text=title, wraplength=560).grid(
            row=2, column=0, columnspan=2, sticky="w", pady=(2, 0)
        )
        if variation:
            ttk.Label(body, text=f"Variación: {variation}", wraplength=560).grid(
                row=3, column=0, columnspan=2, sticky="w", pady=(2, 8)
            )

        row = 4
        self._label(body, "Familia interna", row)
        ttk.Combobox(
            body,
            textvariable=self.family_text,
            values=self.parent.FAMILIES,
            width=36,
        ).grid(row=row, column=1, sticky="ew", pady=4)

        row += 1
        self._label(body, "Subgrupo", row)
        ttk.Combobox(
            body,
            textvariable=self.subgroup_text,
            values=self.parent.SUBGROUPS,
            width=36,
        ).grid(row=row, column=1, sticky="ew", pady=4)

        row += 1
        self._label(body, "Medida A×L×H", row)
        ttk.Entry(body, textvariable=self.size_text, width=38).grid(
            row=row, column=1, sticky="ew", pady=4
        )

        row += 1
        self._label(body, "Materiales", row)
        ttk.Entry(body, textvariable=self.materials_text, width=38).grid(
            row=row, column=1, sticky="ew", pady=4
        )

        row += 1
        self._label(body, "Colores", row)
        ttk.Entry(body, textvariable=self.colors_text, width=38).grid(
            row=row, column=1, sticky="ew", pady=4
        )

        row += 1
        self._label(body, "Estado comercial", row)
        ttk.Combobox(
            body,
            textvariable=self.commercial_status_text,
            values=self.parent.COMMERCIAL_STATUSES,
            width=36,
        ).grid(row=row, column=1, sticky="ew", pady=4)

        row += 1
        ttk.Checkbutton(body, text="Es pack", variable=self.is_pack_var).grid(
            row=row, column=1, sticky="w", pady=4
        )

        row += 1
        ttk.Checkbutton(body, text="Clasificación revisada", variable=self.reviewed_var).grid(
            row=row, column=1, sticky="w", pady=4
        )

        row += 1
        self._label(body, "Notas", row)
        ttk.Entry(body, textvariable=self.notes_text, width=38).grid(
            row=row, column=1, sticky="ew", pady=4
        )

        body.columnconfigure(1, weight=1)

        footer = ttk.Frame(self, padding=(16, 0, 16, 14))
        footer.pack(fill=tk.X)
        ttk.Button(footer, text="Cancelar", command=self.destroy).pack(side=tk.RIGHT)
        ttk.Button(footer, text="Aceptar", command=self._save).pack(side=tk.RIGHT, padx=(0, 8))

    def _label(self, parent: ttk.Frame, text: str, row: int) -> None:
        ttk.Label(parent, text=text).grid(row=row, column=0, sticky="w", padx=(0, 12), pady=4)

    def _save(self) -> None:
        family = self.family_text.get().strip() or "Otros / Sin clasificar"
        subgroup = self.subgroup_text.get().strip() or "Sin subgrupo"
        size = self.size_text.get().strip() or "Sin medida"
        commercial_status = self.commercial_status_text.get().strip() or "Normal"
        reviewed = 1 if self.reviewed_var.get() else 0
        classification = {
            "item_kind": self.item.get("item_kind") or "product",
            "item_woo_id": int(self.item.get("woo_id") or 0),
            "sku": self.item.get("sku") or None,
            "family": family,
            "subgroup": subgroup,
            "size": size,
            "materials": self.materials_text.get().strip(),
            "colors": self.colors_text.get().strip(),
            "commercial_status": commercial_status,
            "is_pack": 1 if self.is_pack_var.get() else 0,
            "reviewed": reviewed,
            "classification_source": "manual",
            "needs_review": 0 if reviewed else 1,
            "notes": self.notes_text.get().strip(),
        }
        self.store.upsert_product_classification(classification)
        if callable(self.on_saved):
            self.on_saved()
        self.destroy()


def run_app(settings: Settings, mode: str = "inventory") -> None:
    app = ProductTableApp(settings, mode=mode)
    app.mainloop()


class PublishDialog(tk.Toplevel):
    def __init__(
        self,
        parent: tk.Tk,
        settings: Settings,
        store: ProductStore,
        on_done: object,
    ) -> None:
        super().__init__(parent)
        self.settings = settings
        self.store = store
        self.on_done = on_done
        self.changes = [dict(row) for row in self.store.list_pending_price_changes()]

        self.title("Propuestas y publicacion segura")
        self.geometry("1120x620")
        self.minsize(940, 500)
        self.transient(parent)
        self.grab_set()

        self.status_text = tk.StringVar(value="")
        self._build_layout()
        self._render_changes()

    def _build_layout(self) -> None:
        ttk.Label(
            self,
            text=(
                "Propuestas pendientes. Primero exporta/revisa si hace falta. "
                "WooCommerce solo se modifica al publicar."
            ),
            padding=(10, 10, 10, 6),
        ).pack(fill=tk.X)

        columns = ("id", "kind", "woo_id", "name", "old", "new", "delta", "risk", "notes")
        self.table = ttk.Treeview(self, columns=columns, show="headings")
        headings = {
            "id": ("ID local", 75),
            "kind": ("Clase", 80),
            "woo_id": ("ID Woo", 80),
            "name": ("Nombre", 310),
            "old": ("Actual", 90),
            "new": ("Nuevo", 90),
            "delta": ("Dif.", 90),
            "risk": ("Seguridad", 160),
            "notes": ("Notas", 210),
        }
        for column, (text, width) in headings.items():
            self.table.heading(column, text=text, anchor=tk.CENTER)
            self.table.column(column, width=width, minwidth=60, anchor=tk.CENTER)
        self.table.tag_configure("blocked", background="#ffdede")
        self.table.tag_configure("warning", background="#fff4c2")
        self.table.tag_configure("ok", background="#eaf7ea")
        self.table.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 8))

        ttk.Label(self, textvariable=self.status_text, padding=(10, 0, 10, 8)).pack(
            fill=tk.X
        )

        footer = ttk.Frame(self, padding=(10, 0, 10, 10))
        footer.pack(fill=tk.X)
        ttk.Button(footer, text="Cerrar", command=self.destroy).pack(side=tk.RIGHT)
        self.publish_button = ttk.Button(
            footer,
            text="Publicar en WooCommerce",
            command=self._publish,
        )
        self.publish_button.pack(side=tk.RIGHT, padx=(0, 8))
        ttk.Button(
            footer,
            text="Cargar Excel revisado",
            command=self._import_reviewed_excel,
        ).pack(side=tk.RIGHT, padx=(0, 8))
        ttk.Button(
            footer,
            text="Exportar propuestas a Excel",
            command=self._export_proposals_excel,
        ).pack(side=tk.RIGHT, padx=(0, 8))

    def _render_changes(self) -> None:
        self.changes = [dict(row) for row in self.store.list_pending_price_changes()]
        self.table.delete(*self.table.get_children())
        blocked = 0
        warnings = 0
        for change in self.changes:
            risk, tag = self._proposal_risk(change)
            if tag == "blocked":
                blocked += 1
            elif tag == "warning":
                warnings += 1
            self.table.insert(
                "",
                tk.END,
                values=(
                    change["id"],
                    change["item_kind"],
                    change["item_woo_id"],
                    change["name"],
                    self._format_money(change["old_price"]),
                    self._format_money(change["new_price"]),
                    self._format_signed(change["delta"]),
                    risk,
                    change["notes"] or "",
                ),
                tags=(tag,),
            )

        auto_fit_tree_columns(self.table)

        total_delta = sum(float(change["delta"]) for change in self.changes)
        self.status_text.set(
            f"Pendientes: {len(self.changes)} | Bloqueadas: {blocked} | Avisos: {warnings} | "
            f"Diferencia acumulada: {self._format_signed(total_delta)}"
        )
        self.publish_button.configure(state=tk.NORMAL)
        if not self.changes or blocked:
            self.publish_button.configure(state=tk.DISABLED)

    def _proposal_risk(self, change: dict[str, object]) -> tuple[str, str]:
        old_price = float(change.get("old_price") or 0)
        new_price = float(change.get("new_price") or 0)
        if new_price <= 0:
            return "BLOQUEADO: precio <= 0", "blocked"
        if old_price <= 0:
            return "AVISO: precio actual <= 0", "warning"
        decrease_pct = ((old_price - new_price) / old_price) * 100
        if decrease_pct > 30:
            return f"AVISO: baja {decrease_pct:.1f}%", "warning"
        if new_price < old_price:
            return f"AVISO: baja {decrease_pct:.1f}%", "warning"
        return "OK", "ok"

    def _export_proposals_excel(self) -> None:
        if not self.changes:
            messagebox.showwarning("Sin propuestas", "No hay propuestas pendientes para exportar.", parent=self)
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            messagebox.showerror(
                "Falta dependencia",
                "Instala openpyxl para exportar: pip install openpyxl",
                parent=self,
            )
            return
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Exportar propuestas de precio",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="propuestas_precios_futonespai.xlsx",
        )
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Propuestas"
        columns = [
            "ID Local",
            "Clase",
            "Nombre",
            "Precio Actual",
            "Precio Propuesto",
            "Diferencia €",
            "Diferencia %",
            "Notas",
        ]
        for col_idx, label in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center")
        for row_idx, change in enumerate(self.changes, start=2):
            risk, _tag = self._proposal_risk(change)
            old_price = float(change["old_price"] or 0)
            new_price = float(change["new_price"] or 0)
            delta = new_price - old_price
            delta_pct = (delta / old_price * 100) if old_price else 0
            values = [
                change["id"],
                change["item_kind"],
                change["name"],
                old_price,
                new_price,
                delta,
                delta_pct,
                change["notes"] or "",
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 48)
        wb.save(path)
        messagebox.showinfo("Exportacion completada", f"Propuestas exportadas:\n{path}", parent=self)

    def _import_reviewed_excel(self) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError:
            messagebox.showerror(
                "Falta dependencia",
                "Instala openpyxl para cargar Excel: pip install openpyxl",
                parent=self,
            )
            return
        path = filedialog.askopenfilename(
            parent=self,
            title="Cargar Excel revisado",
            filetypes=[("Excel", "*.xlsx *.xlsm")],
        )
        if not path:
            return
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        headers = {str(cell.value or "").strip().lower(): idx + 1 for idx, cell in enumerate(ws[1])}
        id_col = headers.get("id local")
        proposed_price_col = headers.get("precio propuesto")
        notes_col = headers.get("notas")
        current_price_col = headers.get("precio actual")
        if not id_col:
            messagebox.showerror("Excel no valido", "No encuentro la columna 'ID Local'.", parent=self)
            return
        if not proposed_price_col:
            messagebox.showerror("Excel no valido", "No encuentro la columna 'Precio Propuesto'.", parent=self)
            return

        proposals_by_id = {int(change["id"]): dict(change) for change in self.store.list_pending_price_changes()}
        updated = unchanged = skipped = blocked = missing = notes_changed = current_mismatch = 0
        change_lines: list[str] = []

        for row in range(2, ws.max_row + 1):
            proposal_id = self._safe_int(ws.cell(row=row, column=id_col).value)
            if not proposal_id:
                continue
            saved = proposals_by_id.get(proposal_id)
            if not saved:
                missing += 1
                continue

            raw_price = ws.cell(row=row, column=proposed_price_col).value
            reviewed_price = self._safe_float(raw_price) if raw_price not in (None, "") else None
            if reviewed_price is None:
                skipped += 1
                continue
            if reviewed_price <= 0:
                blocked += 1
                continue

            saved_old_price = float(saved.get("old_price") or 0)
            saved_new_price = float(saved.get("new_price") or 0)
            excel_notes = str(ws.cell(row=row, column=notes_col).value or "").strip() if notes_col else ""
            saved_notes = str(saved.get("notes") or "").strip()
            changes_for_row: list[str] = []

            if current_price_col:
                excel_current = self._safe_float(ws.cell(row=row, column=current_price_col).value)
                if excel_current is not None and abs(excel_current - saved_old_price) > 0.005:
                    current_mismatch += 1
                    changes_for_row.append(
                        f"precio actual Excel {excel_current:.2f} ≠ guardado {saved_old_price:.2f}"
                    )

            price_changed = abs(reviewed_price - saved_new_price) > 0.005
            note_changed = excel_notes != saved_notes
            final_notes = saved_notes
            if note_changed:
                notes_changed += 1
                final_notes = excel_notes
                changes_for_row.append("notas modificadas")
            if price_changed:
                changes_for_row.append(f"precio propuesto {saved_new_price:.2f} → {reviewed_price:.2f}")

            if price_changed or note_changed:
                self.store.update_price_change_proposal_review(
                    proposal_id,
                    new_price=reviewed_price if price_changed else None,
                    notes=final_notes if note_changed else None,
                    status="pending",
                )
                updated += 1
                if len(change_lines) < 12:
                    change_lines.append(
                        f"ID {proposal_id} · {saved.get('name')}: " + "; ".join(changes_for_row)
                    )
            else:
                unchanged += 1

        self._render_changes()
        detail = ""
        if change_lines:
            detail = "\n\nCambios detectados:\n" + "\n".join(change_lines)
            if updated > len(change_lines):
                detail += f"\n... y {updated - len(change_lines)} cambios mas."
        messagebox.showinfo(
            "Excel revisado cargado",
            f"Actualizadas por cambios detectados: {updated}\n"
            f"Sin cambios: {unchanged}\n"
            f"Omitidas sin precio propuesto: {skipped}\n"
            f"No encontradas en propuestas pendientes: {missing}\n"
            f"Bloqueadas por precio <= 0: {blocked}\n"
            f"Notas modificadas: {notes_changed}\n"
            f"Avisos de precio actual distinto: {current_mismatch}"
            + detail,
            parent=self,
        )

    def _publish(self) -> None:
        if not self.changes:
            return
        blocked = [change for change in self.changes if self._proposal_risk(change)[1] == "blocked"]
        if blocked:
            messagebox.showerror(
                "Publicacion bloqueada",
                "Hay propuestas con precio 0 o negativo. Corrigelas antes de publicar.",
                parent=self,
            )
            self._render_changes()
            return
        warnings = [change for change in self.changes if self._proposal_risk(change)[1] == "warning"]
        warning_text = ""
        if warnings:
            warning_text = f"\n\nAvisos detectados: {len(warnings)}. Revisa especialmente bajadas de precio."
        confirmed = messagebox.askyesno(
            "Confirmar publicacion",
            "Vas a modificar precios reales en WooCommerce. "
            "Esta accion debe hacerse solo con propuestas revisadas." + warning_text +
            "\n\n¿Quieres publicar todos los cambios pendientes?",
            parent=self,
        )
        if not confirmed:
            self.status_text.set("Publicacion denegada. No se ha modificado WooCommerce.")
            return

        self.publish_button.configure(state=tk.DISABLED)
        self.status_text.set("Publicando cambios en WooCommerce...")
        self.update_idletasks()

        try:
            result = publish_pending_price_changes(self.settings)
        except OperationBlockedError as exc:
            self.publish_button.configure(state=tk.NORMAL)
            self.status_text.set("Publicacion bloqueada por seguridad.")
            messagebox.showwarning("Operacion bloqueada", str(exc), parent=self)
            return
        except WooCommerceError as exc:
            self.publish_button.configure(state=tk.NORMAL)
            self.status_text.set("Publicacion fallida.")
            messagebox.showerror("Error WooCommerce", str(exc), parent=self)
            return

        if result.errors:
            messagebox.showwarning(
                "Publicacion parcial",
                "Algunos cambios fallaron:\n\n" + "\n".join(result.errors[:8]),
                parent=self,
            )
        else:
            messagebox.showinfo(
                "Publicacion completada",
                f"Cambios publicados: {result.published}",
                parent=self,
            )

        self._render_changes()
        if callable(self.on_done):
            self.on_done()

    def _safe_int(self, value: object) -> int | None:
        try:
            return int(float(str(value).replace(",", ".")))
        except (TypeError, ValueError):
            return None

    def _safe_float(self, value: object) -> float | None:
        try:
            return float(str(value).replace("€", "").replace(" ", "").replace(",", "."))
        except (TypeError, ValueError):
            return None

    def _format_money(self, value: object) -> str:
        return f"{float(value):.2f}"

    def _format_signed(self, value: object) -> str:
        return f"{float(value):+.2f}"

class PackDialog(tk.Toplevel):
    def __init__(
        self,
        parent: tk.Tk,
        store: ProductStore,
        catalog_items: list[dict[str, object]],
    ) -> None:
        super().__init__(parent)
        self.parent = parent
        self.store = store
        self.catalog_items = catalog_items
        self.filtered_items = catalog_items
        self.selected_items: dict[tuple[str, int], dict[str, Any]] = {}
        self.pack_candidates = [dict(row) for row in self.store.list_pack_candidates()]
        self.pack_by_label: dict[str, dict[str, Any]] = {}

        self.title("Crear pack")
        self.geometry("1180x720")
        self.minsize(980, 600)
        self.transient(parent)
        self.grab_set()

        self.pack_text = tk.StringVar()
        self.search_text = tk.StringVar()
        self.category_text = tk.StringVar(value="Todas")
        self.pricing_mode = tk.StringVar(value="current")
        self.percent_text = tk.StringVar(value="0")
        self.amount_text = tk.StringVar(value="0")
        self.total_text = tk.StringVar(value="0.00")
        self.final_text = tk.StringVar(value="0.00")
        self.pack_current_text = tk.StringVar(value="0.00")
        self.item_detail_text = tk.StringVar(value="Selecciona un item para ver precio y categorias.")

        self._build_layout()
        self._load_pack_candidates()
        self._render_available_items()
        self._recalculate()

    def _build_layout(self) -> None:
        header = ttk.Frame(self, padding=(10, 10, 10, 6))
        header.pack(fill=tk.X)

        ttk.Label(header, text="Producto pack:").pack(side=tk.LEFT)
        self.pack_combo = ttk.Combobox(
            header,
            textvariable=self.pack_text,
            state="readonly",
            width=72,
        )
        self.pack_combo.pack(side=tk.LEFT, padx=(6, 12))
        self.pack_combo.bind("<<ComboboxSelected>>", lambda _event: self._pack_changed())

        ttk.Label(header, text="Precio actual:").pack(side=tk.LEFT)
        ttk.Label(header, textvariable=self.pack_current_text, width=12).pack(
            side=tk.LEFT,
            padx=(6, 0),
        )

        body = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        body.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 8))

        available = ttk.Frame(body)
        selected = ttk.Frame(body)
        body.add(available, weight=3)
        body.add(selected, weight=2)

        filters = ttk.Frame(available)
        filters.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(filters, text="Buscar:").pack(side=tk.LEFT)
        search = ttk.Entry(filters, textvariable=self.search_text, width=34)
        search.pack(side=tk.LEFT, padx=(6, 10))
        search.bind("<KeyRelease>", lambda _event: self._render_available_items())

        ttk.Label(filters, text="Categoria:").pack(side=tk.LEFT)
        self.category_filter = ttk.Combobox(
            filters,
            textvariable=self.category_text,
            state="readonly",
            width=26,
        )
        self.category_filter.pack(side=tk.LEFT, padx=(6, 0))
        self.category_filter.bind(
            "<<ComboboxSelected>>",
            lambda _event: self._render_available_items(),
        )

        available_columns = ("kind", "id", "name", "variation", "price", "categories")
        self.available_table = ttk.Treeview(
            available,
            columns=available_columns,
            show="headings",
            selectmode="browse",
            height=16,
        )
        self._setup_table(
            self.available_table,
            {
                "kind": ("Clase", 80),
                "id": ("ID", 75),
                "name": ("Nombre", 230),
                "variation": ("Variacion", 180),
                "price": ("Precio", 80),
                "categories": ("Categorias", 210),
            },
        )
        self.available_table.pack(fill=tk.BOTH, expand=True)
        self.available_table.bind("<<TreeviewSelect>>", lambda _event: self._show_item_detail())
        self.available_table.bind("<Double-1>", lambda _event: self._add_selected_item())

        ttk.Label(
            available,
            textvariable=self.item_detail_text,
            padding=(0, 6, 0, 0),
        ).pack(fill=tk.X)

        available_actions = ttk.Frame(available)
        available_actions.pack(fill=tk.X, pady=(6, 0))
        ttk.Button(
            available_actions,
            text="Agregar seleccionado",
            command=self._add_selected_item,
        ).pack(side=tk.RIGHT)

        selected_columns = ("kind", "id", "name", "variation", "qty", "price", "line_total")
        self.selected_table = ttk.Treeview(
            selected,
            columns=selected_columns,
            show="headings",
            selectmode="browse",
            height=13,
        )
        self._setup_table(
            self.selected_table,
            {
                "kind": ("Clase", 75),
                "id": ("ID", 70),
                "name": ("Nombre", 220),
                "variation": ("Variacion", 170),
                "qty": ("Cant.", 65),
                "price": ("Precio", 80),
                "line_total": ("Total", 85),
            },
        )
        self.selected_table.pack(fill=tk.BOTH, expand=True)

        selected_actions = ttk.Frame(selected)
        selected_actions.pack(fill=tk.X, pady=(6, 8))
        ttk.Button(selected_actions, text="+ Cant.", command=self._increase_qty).pack(side=tk.LEFT)
        ttk.Button(selected_actions, text="- Cant.", command=self._decrease_qty).pack(
            side=tk.LEFT,
            padx=(6, 0),
        )
        ttk.Button(selected_actions, text="Quitar", command=self._remove_selected_item).pack(
            side=tk.LEFT,
            padx=(6, 0),
        )

        pricing = ttk.LabelFrame(selected, text="Precio del pack", padding=10)
        pricing.pack(fill=tk.X)

        ttk.Label(pricing, text="Suma fija componentes:").grid(row=0, column=0, sticky=tk.W)
        ttk.Label(pricing, textvariable=self.total_text).grid(row=0, column=1, sticky=tk.W)

        ttk.Radiobutton(
            pricing,
            text="1. Dejar el precio como esta",
            variable=self.pricing_mode,
            value="current",
            command=self._recalculate,
        ).grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(8, 0))

        ttk.Radiobutton(
            pricing,
            text="2. Rebaja por porcentaje",
            variable=self.pricing_mode,
            value="percent",
            command=self._recalculate,
        ).grid(row=2, column=0, sticky=tk.W, pady=(6, 0))
        percent = ttk.Entry(pricing, textvariable=self.percent_text, width=8)
        percent.grid(row=2, column=1, sticky=tk.W, pady=(6, 0))
        percent.bind("<KeyRelease>", lambda _event: self._recalculate())

        ttk.Radiobutton(
            pricing,
            text="3. Rebaja por monto",
            variable=self.pricing_mode,
            value="amount",
            command=self._recalculate,
        ).grid(row=3, column=0, sticky=tk.W, pady=(6, 0))
        amount = ttk.Entry(pricing, textvariable=self.amount_text, width=8)
        amount.grid(row=3, column=1, sticky=tk.W, pady=(6, 0))
        amount.bind("<KeyRelease>", lambda _event: self._recalculate())

        ttk.Label(pricing, text="Precio final calculado:").grid(
            row=4,
            column=0,
            sticky=tk.W,
            pady=(10, 0),
        )
        ttk.Label(pricing, textvariable=self.final_text).grid(
            row=4,
            column=1,
            sticky=tk.W,
            pady=(10, 0),
        )

        footer = ttk.Frame(self, padding=(10, 0, 10, 10))
        footer.pack(fill=tk.X)
        ttk.Button(footer, text="Cancelar", command=self.destroy).pack(side=tk.RIGHT)
        ttk.Button(footer, text="Guardar pack", command=self._save_pack).pack(
            side=tk.RIGHT,
            padx=(0, 8),
        )

    def _setup_table(self, table: ttk.Treeview, columns: dict[str, tuple[str, int]]) -> None:
        for column, (heading, width) in columns.items():
            table.heading(column, text=heading, anchor=tk.CENTER)
            table.column(column, width=width, minwidth=55, anchor=tk.CENTER)

    def _load_pack_candidates(self) -> None:
        labels = []
        for pack in self.pack_candidates:
            label = f"{pack['woo_id']} | {pack['price'] or '-'} | {pack['name']}"
            labels.append(label)
            self.pack_by_label[label] = pack
        self.pack_combo.configure(values=labels)
        if labels:
            self.pack_text.set(labels[0])
            self._pack_changed()

        categories = sorted(
            {
                category
                for item in self.catalog_items
                for category in str(item.get("categories_label") or "").split(" / ")
                if category
            }
        )
        self.category_filter.configure(values=["Todas", *categories])

    def _pack_changed(self) -> None:
        pack = self._selected_pack()
        self.pack_current_text.set(self._format_money(self._price_of(pack)))
        self._recalculate()

    def _selected_pack(self) -> dict[str, Any] | None:
        return self.pack_by_label.get(self.pack_text.get())

    def _render_available_items(self) -> None:
        query = self.search_text.get().strip().lower()
        category = self.category_text.get()
        items = self.catalog_items
        if category and category != "Todas":
            items = [
                item
                for item in items
                if category in str(item.get("categories_label") or "").split(" / ")
            ]
        if query:
            items = [
                item
                for item in items
                if query in str(item.get("name") or "").lower()
                or query in str(item.get("sku") or "").lower()
                or query in str(item.get("woo_id") or "").lower()
                or query in str(item.get("attributes_label") or "").lower()
                or query in str(item.get("categories_label") or "").lower()
            ]
        self.filtered_items = items

        self.available_table.delete(*self.available_table.get_children())
        selected_pack = self._selected_pack()
        selected_pack_id = selected_pack["woo_id"] if selected_pack else None
        for index, item in enumerate(items):
            if item.get("item_kind") == "product" and item.get("woo_id") == selected_pack_id:
                continue
            self.available_table.insert(
                "",
                tk.END,
                iid=str(index),
                values=(
                    item.get("item_kind") or "",
                    item.get("woo_id") or "",
                    item.get("name") or "",
                    item.get("attributes_label") or "",
                    item.get("price") or "",
                    item.get("categories_label") or "",
                ),
            )
        auto_fit_tree_columns(self.available_table)

    def _selected_available_item(self) -> dict[str, object] | None:
        selection = self.available_table.selection()
        if not selection:
            return None
        return self.filtered_items[int(selection[0])]

    def _show_item_detail(self) -> None:
        item = self._selected_available_item()
        if not item:
            return
        price = self._format_money(self._price_of(item))
        categories = item.get("categories_label") or "Sin categoria"
        variation = item.get("attributes_label") or "Sin variacion"
        self.item_detail_text.set(
            f"Precio: {price} | Categorias: {categories} | {variation}"
        )

    def _add_selected_item(self) -> None:
        item = self._selected_available_item()
        if not item:
            messagebox.showwarning("Seleccion necesaria", "Selecciona un item del catalogo.")
            return
        key = (str(item["item_kind"]), int(item["woo_id"]))
        if key in self.selected_items:
            self.selected_items[key]["quantity"] += 1
        else:
            self.selected_items[key] = {
                "item": item,
                "quantity": 1.0,
            }
        self._render_selected_items()

    def _render_selected_items(self) -> None:
        self.selected_table.delete(*self.selected_table.get_children())
        for key, data in self.selected_items.items():
            item = data["item"]
            quantity = float(data["quantity"])
            price = self._price_of(item)
            self.selected_table.insert(
                "",
                tk.END,
                iid=f"{key[0]}:{key[1]}",
                values=(
                    item.get("item_kind") or "",
                    item.get("woo_id") or "",
                    item.get("name") or "",
                    item.get("attributes_label") or "",
                    self._format_quantity(quantity),
                    self._format_money(price),
                    self._format_money(price * quantity),
                ),
            )
        auto_fit_tree_columns(self.selected_table)
        self._recalculate()

    def _selected_component_key(self) -> tuple[str, int] | None:
        selection = self.selected_table.selection()
        if not selection:
            return None
        kind, raw_id = selection[0].split(":", 1)
        return kind, int(raw_id)

    def _increase_qty(self) -> None:
        key = self._selected_component_key()
        if not key:
            return
        self.selected_items[key]["quantity"] += 1
        self._render_selected_items()

    def _decrease_qty(self) -> None:
        key = self._selected_component_key()
        if not key:
            return
        self.selected_items[key]["quantity"] -= 1
        if self.selected_items[key]["quantity"] <= 0:
            del self.selected_items[key]
        self._render_selected_items()

    def _remove_selected_item(self) -> None:
        key = self._selected_component_key()
        if not key:
            return
        del self.selected_items[key]
        self._render_selected_items()

    def _recalculate(self) -> None:
        total = sum(
            self._price_of(data["item"]) * float(data["quantity"])
            for data in self.selected_items.values()
        )
        self.total_text.set(self._format_money(total))

        mode = self.pricing_mode.get()
        if mode == "current":
            final = self._price_of(self._selected_pack())
        elif mode == "percent":
            final = total * (1 - self._safe_float(self.percent_text.get()) / 100)
        else:
            final = total - self._safe_float(self.amount_text.get())
        self.final_text.set(self._format_money(max(final, 0)))

    def _save_pack(self) -> None:
        pack = self._selected_pack()
        if not pack:
            messagebox.showwarning("Pack necesario", "Selecciona el producto pack.")
            return
        if not self.selected_items:
            messagebox.showwarning("Componentes necesarios", "Agrega al menos un componente.")
            return

        components_total = self._safe_float(self.total_text.get())
        final_price = self._safe_float(self.final_text.get())
        mode = self.pricing_mode.get()
        pack_id = self.store.save_manual_pack(
            {
                "pack_woo_id": pack["woo_id"],
                "name": pack["name"],
                "components_total": components_total,
                "pricing_mode": mode,
                "discount_percent": self._safe_float(self.percent_text.get())
                if mode == "percent"
                else None,
                "discount_amount": self._safe_float(self.amount_text.get())
                if mode == "amount"
                else None,
                "final_price": final_price,
                "notes": None,
            },
            [
                {
                    "item_kind": key[0],
                    "item_woo_id": key[1],
                    "quantity": float(data["quantity"]),
                    "notes": None,
                }
                for key, data in self.selected_items.items()
            ],
        )
        messagebox.showinfo("Pack guardado", f"Pack guardado con ID local {pack_id}.")
        self.destroy()

    def _price_of(self, item: dict[str, Any] | None) -> float:
        if not item:
            return 0
        return self._safe_float(item.get("price"))

    def _safe_float(self, value: object) -> float:
        if value is None:
            return 0
        text = str(value).strip().replace(",", ".")
        if not text:
            return 0
        try:
            return float(text)
        except ValueError:
            return 0

    def _format_money(self, value: float) -> str:
        return f"{value:.2f}"

    def _format_quantity(self, value: float) -> str:
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}"


class BulkPriceChangeDialog(tk.Toplevel):
    """Crea propuestas de precio usando los articulos visibles por los filtros del listado."""

    def __init__(
        self,
        parent: tk.Tk,
        store: ProductStore,
        filtered_items: list[dict[str, object]],
        on_added: object | None = None,
    ) -> None:
        super().__init__(parent)
        self.store = store
        self.filtered_items = filtered_items
        self.on_added = on_added
        self.adjust_mode = tk.StringVar(value="percent")
        self.adjust_value = tk.StringVar(value="0")
        self.notes_text = tk.StringVar(value="Subida por filtro desde Cambio de Precios")
        self.include_products = tk.BooleanVar(value=True)
        self.include_variations = tk.BooleanVar(value=True)
        self.exclude_outlet = tk.BooleanVar(value=True)
        self.status_text = tk.StringVar(value="")
        self.preview_rows: list[dict[str, object]] = []

        self.title("Añadir filtro a propuesta")
        self.geometry("1120x690")
        self.minsize(980, 560)
        self.transient(parent)
        self.grab_set()

        container = ttk.Frame(self, padding=14)
        container.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            container,
            text="Subida de precios por listado filtrado",
            style="Title.TLabel",
        ).pack(anchor=tk.W)
        ttk.Label(
            container,
            text=(
                "Usa exactamente los artículos visibles en la tabla de Cambio de Precios. "
                "Esto solo crea/actualiza propuestas locales; no modifica WooCommerce."
            ),
            style="Muted.TLabel",
            wraplength=980,
        ).pack(anchor=tk.W, pady=(4, 12))

        controls = ttk.LabelFrame(container, text="Ajuste a aplicar", padding=10)
        controls.pack(fill=tk.X)

        ttk.Radiobutton(
            controls,
            text="Subir por porcentaje",
            variable=self.adjust_mode,
            value="percent",
            command=self._render_preview,
        ).grid(row=0, column=0, sticky=tk.W)
        ttk.Radiobutton(
            controls,
            text="Subir importe fijo",
            variable=self.adjust_mode,
            value="amount",
            command=self._render_preview,
        ).grid(row=0, column=1, sticky=tk.W, padx=(18, 0))

        ttk.Label(controls, text="Valor:").grid(row=0, column=2, sticky=tk.E, padx=(24, 6))
        value_entry = ttk.Entry(controls, textvariable=self.adjust_value, width=12, justify=tk.CENTER)
        value_entry.grid(row=0, column=3, sticky=tk.W)
        value_entry.bind("<KeyRelease>", lambda _event: self._render_preview())

        ttk.Checkbutton(
            controls,
            text="Productos visibles",
            variable=self.include_products,
            command=self._render_preview,
        ).grid(row=1, column=0, sticky=tk.W, pady=(10, 0))
        ttk.Checkbutton(
            controls,
            text="Variaciones visibles",
            variable=self.include_variations,
            command=self._render_preview,
        ).grid(row=1, column=1, sticky=tk.W, padx=(18, 0), pady=(10, 0))
        ttk.Checkbutton(
            controls,
            text="Excluir Outlet",
            variable=self.exclude_outlet,
            command=self._render_preview,
        ).grid(row=1, column=2, sticky=tk.W, columnspan=2, pady=(10, 0))

        ttk.Label(controls, text="Notas:").grid(row=2, column=0, sticky=tk.W, pady=(10, 0))
        notes_entry = ttk.Entry(controls, textvariable=self.notes_text, width=72)
        notes_entry.grid(row=2, column=1, columnspan=3, sticky="ew", padx=(8, 0), pady=(10, 0))
        controls.columnconfigure(3, weight=1)

        table_frame = ttk.Frame(container, padding=(0, 12, 0, 0))
        table_frame.pack(fill=tk.BOTH, expand=True)
        columns = (
            "estado",
            "item_kind",
            "woo_id",
            "name",
            "attributes_label",
            "price",
            "new_price",
            "delta",
        )
        self.preview_table = ttk.Treeview(table_frame, columns=columns, show="headings")
        headings = {
            "estado": "Estado",
            "item_kind": "Clase",
            "woo_id": "ID Woo",
            "name": "Nombre",
            "attributes_label": "Variación",
            "price": "Precio actual",
            "new_price": "Precio propuesto",
            "delta": "Diferencia",
        }
        widths = {
            "estado": 170,
            "item_kind": 90,
            "woo_id": 80,
            "name": 310,
            "attributes_label": 230,
            "price": 115,
            "new_price": 130,
            "delta": 110,
        }
        for column in columns:
            self.preview_table.heading(column, text=headings[column], anchor=tk.CENTER)
            self.preview_table.column(column, width=widths[column], minwidth=70, anchor=tk.CENTER)
        self.preview_table.tag_configure("ok", background="#edf7ed")
        self.preview_table.tag_configure("warn", background="#fff8dc")
        self.preview_table.tag_configure("error", background="#fdecec")

        y_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.preview_table.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.preview_table.xview)
        self.preview_table.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.preview_table.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        footer = ttk.Frame(container)
        footer.pack(fill=tk.X, pady=(10, 0))
        ttk.Label(footer, textvariable=self.status_text, style="Muted.TLabel").pack(side=tk.LEFT)
        ttk.Button(footer, text="Cancelar", command=self.destroy).pack(side=tk.RIGHT)
        ttk.Button(
            footer,
            text="Añadir visibles válidos a propuesta",
            command=self._add_visible_to_proposal,
        ).pack(side=tk.RIGHT, padx=(0, 8))

        self._render_preview()

    def _render_preview(self) -> None:
        self.preview_table.delete(*self.preview_table.get_children())
        self.preview_rows = []
        raw_value = self._safe_float(self.adjust_value.get())
        mode = self.adjust_mode.get()
        total = len(self.filtered_items)
        valid = 0
        skipped = 0
        warnings = 0

        for item in self.filtered_items:
            row = self._preview_for_item(item, raw_value, mode)
            self.preview_rows.append(row)
            state = str(row["state"])
            if state == "ok":
                valid += 1
            elif state == "warn":
                warnings += 1
            else:
                skipped += 1
            self.preview_table.insert(
                "",
                tk.END,
                values=(
                    row["status"],
                    row["item_kind"],
                    row["woo_id"],
                    row["name"],
                    row["attributes_label"],
                    self._format_money(float(row["old_price"])),
                    self._format_money(float(row["new_price"])),
                    self._format_signed(float(row["delta"])),
                ),
                tags=(state,),
            )
        auto_fit_tree_columns(self.preview_table)
        self.status_text.set(
            f"Visibles por filtro: {total} | Listos para propuesta: {valid} | "
            f"Avisos: {warnings} | Omitidos: {skipped}. WooCommerce no se modifica."
        )

    def _preview_for_item(
        self,
        item: dict[str, object],
        value: float,
        mode: str,
    ) -> dict[str, object]:
        kind = str(item.get("item_kind") or "")
        old_price = self._safe_float(item.get("price"))
        new_price = old_price * (1 + value / 100) if mode == "percent" else old_price + value
        status = "Listo"
        state = "ok"

        if kind == "product" and not self.include_products.get():
            status = "Omitido: productos desmarcados"
            state = "error"
        elif kind == "variation" and not self.include_variations.get():
            status = "Omitido: variaciones desmarcadas"
            state = "error"
        elif self.exclude_outlet.get() and str(item.get("commercial_status") or "") == "Outlet":
            status = "Omitido: Outlet protegido"
            state = "error"
        elif old_price <= 0:
            status = "Omitido: sin precio actual"
            state = "error"
        elif new_price <= 0:
            status = "Bloqueado: precio <= 0"
            state = "error"
        elif abs(new_price - old_price) <= 0.005:
            status = "Sin cambio de precio"
            state = "error"
        elif new_price < old_price:
            status = "Aviso: bajada de precio"
            state = "warn"

        return {
            "item_kind": kind,
            "item_woo_id": int(item.get("woo_id") or 0),
            "woo_id": int(item.get("woo_id") or 0),
            "name": self._item_name(item),
            "attributes_label": str(item.get("attributes_label") or ""),
            "old_price": old_price,
            "new_price": max(new_price, 0),
            "delta": max(new_price, 0) - old_price,
            "status": status,
            "state": state,
        }

    def _add_visible_to_proposal(self) -> None:
        valid_rows = [row for row in self.preview_rows if row.get("state") in {"ok", "warn"}]
        if not valid_rows:
            messagebox.showwarning(
                "Sin cambios validos",
                "No hay articulos validos para añadir a propuesta. Revisa filtros, precio actual y valor del ajuste.",
                parent=self,
            )
            return
        warnings = sum(1 for row in valid_rows if row.get("state") == "warn")
        if warnings:
            proceed = messagebox.askyesno(
                "Avisos de bajada",
                f"Hay {warnings} articulos con bajada de precio.\n\n¿Quieres añadirlos igualmente a la propuesta local?",
                parent=self,
            )
            if not proceed:
                return

        note = self.notes_text.get().strip() or "Cambio por filtro desde Cambio de Precios"
        proposals = [
            {
                "item_kind": row["item_kind"],
                "item_woo_id": row["item_woo_id"],
                "name": row["name"],
                "old_price": row["old_price"],
                "new_price": row["new_price"],
                "delta": row["delta"],
                "notes": note,
            }
            for row in valid_rows
        ]
        added = self.store.upsert_pending_price_change_proposals(proposals)
        if callable(self.on_added):
            self.on_added(added)
        messagebox.showinfo(
            "Propuesta actualizada",
            f"Añadidos/actualizados en propuesta: {added}.\n\nWooCommerce no se ha modificado.",
            parent=self,
        )
        self.destroy()

    def _item_name(self, item: dict[str, object]) -> str:
        name = str(item.get("name") or "")
        variation = str(item.get("attributes_label") or "")
        if variation:
            return f"{name} - {variation}"
        return name

    def _safe_float(self, value: object) -> float:
        if value is None:
            return 0.0
        try:
            return float(str(value).strip().replace(",", ".") or "0")
        except ValueError:
            return 0.0

    def _format_money(self, value: float) -> str:
        return f"{value:.2f}"

    def _format_signed(self, value: float) -> str:
        return f"{value:+.2f}"


class PriceChangeDialog(tk.Toplevel):
    def __init__(
        self,
        parent: tk.Tk,
        store: ProductStore,
        catalog_items: list[dict[str, object]],
        initial_item: dict[str, object] | None,
        on_added: object | None = None,
    ) -> None:
        super().__init__(parent)
        self.store = store
        self.catalog_items = catalog_items
        self.on_added = on_added
        self.item_by_label: dict[str, dict[str, object]] = {}
        self.related_rows: dict[int, dict[str, object]] = {}
        self.related_proposals: dict[int, float] = {}
        self.marked_related: set[int] = set()
        self.variation_rows: dict[int, dict[str, object]] = {}
        self.variation_proposals: dict[int, float] = {}
        self.marked_variations: set[int] = set()

        self.title("Cambiar precio local")
        self.geometry("1160x720")
        self.minsize(980, 600)
        self.transient(parent)
        self.grab_set()

        self.item_text = tk.StringVar()
        self.current_price_text = tk.StringVar(value="0.00")
        self.new_price_text = tk.StringVar(value="0.00")
        self.percent_increase_text = tk.StringVar(value="5")
        self.related_price_text = tk.StringVar(value="")
        self.related_adjust_mode = tk.StringVar(value="percent")
        self.related_adjust_value = tk.StringVar(value="5")
        self.variation_price_text = tk.StringVar(value="")
        self.variation_adjust_mode = tk.StringVar(value="percent")
        self.variation_adjust_value = tk.StringVar(value="5")
        self.proposal_status_text = tk.StringVar(value="Selecciona un artículo, ajusta precios y pulsa Añadir a propuesta.")
        self.delta_text = tk.StringVar(value="0.00")
        self.detail_text = tk.StringVar(value="")

        self._build_layout()
        self._load_items(initial_item)
        self._recalculate()

    def _build_layout(self) -> None:
        top = ttk.Frame(self, padding=(10, 10, 10, 6))
        top.pack(fill=tk.X)

        ttk.Label(top, text="Articulo:").pack(side=tk.LEFT)
        self.item_combo = ttk.Combobox(
            top,
            textvariable=self.item_text,
            state="readonly",
            width=82,
        )
        self.item_combo.pack(side=tk.LEFT, padx=(6, 0))
        self.item_combo.bind("<<ComboboxSelected>>", lambda _event: self._item_changed())

        price_frame = ttk.Frame(self, padding=(10, 0, 10, 8))
        price_frame.pack(fill=tk.X)
        ttk.Label(price_frame, text="Precio actual:").grid(row=0, column=0, sticky=tk.W)
        ttk.Label(price_frame, textvariable=self.current_price_text, width=12).grid(
            row=0,
            column=1,
            sticky=tk.W,
            padx=(6, 20),
        )
        ttk.Label(price_frame, text="Nuevo precio:").grid(row=0, column=2, sticky=tk.W)
        new_price = ttk.Entry(price_frame, textvariable=self.new_price_text, width=12)
        new_price.grid(row=0, column=3, sticky=tk.W, padx=(6, 20))
        new_price.bind("<KeyRelease>", lambda _event: self._recalculate())
        ttk.Label(price_frame, text="Diferencia:").grid(row=0, column=4, sticky=tk.W)
        ttk.Label(price_frame, textvariable=self.delta_text, width=12).grid(
            row=0,
            column=5,
            sticky=tk.W,
            padx=(6, 0),
        )
        ttk.Label(price_frame, text="Subida %:").grid(row=0, column=6, sticky=tk.W, padx=(20, 0))
        ttk.Entry(price_frame, textvariable=self.percent_increase_text, width=8).grid(
            row=0,
            column=7,
            sticky=tk.W,
            padx=(6, 6),
        )
        ttk.Button(
            price_frame,
            text="Aplicar al articulo",
            command=self._apply_percent_to_main_item,
        ).grid(row=0, column=8, sticky=tk.W)

        ttk.Label(self, textvariable=self.detail_text, padding=(10, 0, 10, 8)).pack(
            fill=tk.X
        )

        notebook = ttk.Notebook(self)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 8))

        variations_frame = ttk.Frame(notebook)
        related_frame = ttk.Frame(notebook)
        packs_frame = ttk.Frame(notebook)
        notebook.add(variations_frame, text="Variaciones")
        notebook.add(related_frame, text="Relacionadas")
        notebook.add(packs_frame, text="Packs afectados")

        self.variations_table = ttk.Treeview(
            variations_frame,
            columns=("apply", "id", "variation", "current", "new", "delta", "stock"),
            show="headings",
            selectmode="browse",
        )
        self._setup_table(
            self.variations_table,
            {
                "apply": ("Añadir", 70),
                "id": ("ID", 80),
                "variation": ("Variacion", 360),
                "current": ("Actual", 95),
                "new": ("Nuevo", 95),
                "delta": ("Diferencia", 95),
                "stock": ("Stock", 100),
            },
        )
        self.variations_table.pack(fill=tk.BOTH, expand=True)
        self.variations_table.bind(
            "<<TreeviewSelect>>",
            lambda _event: self._variation_selection_changed(),
        )
        self.variations_table.bind("<Double-1>", lambda _event: self._toggle_variation_mark())

        variation_editor = ttk.Frame(variations_frame, padding=(0, 6, 0, 0))
        variation_editor.pack(fill=tk.X)
        ttk.Button(
            variation_editor,
            text="Marcar/desmarcar",
            command=self._toggle_variation_mark,
        ).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(
            variation_editor,
            text="Marcar todas",
            command=self._mark_all_variations,
        ).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(variation_editor, text="Precio propuesto:").pack(side=tk.LEFT)
        ttk.Entry(variation_editor, textvariable=self.variation_price_text, width=12).pack(
            side=tk.LEFT,
            padx=(6, 8),
        )
        ttk.Button(
            variation_editor,
            text="Actualizar variacion",
            command=self._update_variation_price,
        ).pack(side=tk.LEFT)
        ttk.Radiobutton(
            variation_editor,
            text="%",
            variable=self.variation_adjust_mode,
            value="percent",
        ).pack(side=tk.LEFT, padx=(12, 0))
        ttk.Radiobutton(
            variation_editor,
            text="Monto",
            variable=self.variation_adjust_mode,
            value="amount",
        ).pack(side=tk.LEFT)
        ttk.Entry(
            variation_editor,
            textvariable=self.variation_adjust_value,
            width=8,
        ).pack(side=tk.LEFT, padx=(6, 8))
        ttk.Button(
            variation_editor,
            text="Aplicar a marcadas",
            command=self._apply_adjustment_to_marked_variations,
        ).pack(side=tk.LEFT, padx=(8, 0))

        self.related_table = ttk.Treeview(
            related_frame,
            columns=(
                "apply",
                "id",
                "name",
                "variation",
                "current",
                "proposed",
                "delta",
                "reason",
            ),
            show="headings",
            selectmode="browse",
        )
        self._setup_table(
            self.related_table,
            {
                "apply": ("Aplicar", 70),
                "id": ("ID", 80),
                "name": ("Producto", 240),
                "variation": ("Variacion", 300),
                "current": ("Actual", 90),
                "proposed": ("Propuesto", 95),
                "delta": ("Diferencia", 95),
                "reason": ("Relacion", 150),
            },
        )
        self.related_table.pack(fill=tk.BOTH, expand=True)
        self.related_table.bind(
            "<<TreeviewSelect>>",
            lambda _event: self._related_selection_changed(),
        )
        self.related_table.bind("<Double-1>", lambda _event: self._toggle_related_mark())

        related_editor = ttk.Frame(related_frame, padding=(0, 6, 0, 0))
        related_editor.pack(fill=tk.X)
        ttk.Button(
            related_editor,
            text="Marcar/desmarcar",
            command=self._toggle_related_mark,
        ).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(related_editor, text="Precio propuesto:").pack(side=tk.LEFT)
        ttk.Entry(related_editor, textvariable=self.related_price_text, width=12).pack(
            side=tk.LEFT,
            padx=(6, 8),
        )
        ttk.Button(
            related_editor,
            text="Actualizar variacion",
            command=self._update_related_price,
        ).pack(side=tk.LEFT)
        ttk.Radiobutton(
            related_editor,
            text="%",
            variable=self.related_adjust_mode,
            value="percent",
        ).pack(side=tk.LEFT, padx=(12, 0))
        ttk.Radiobutton(
            related_editor,
            text="Monto",
            variable=self.related_adjust_mode,
            value="amount",
        ).pack(side=tk.LEFT)
        ttk.Entry(
            related_editor,
            textvariable=self.related_adjust_value,
            width=8,
        ).pack(side=tk.LEFT, padx=(6, 8))
        ttk.Button(
            related_editor,
            text="Aplicar a marcadas",
            command=self._apply_adjustment_to_marked_related,
        ).pack(side=tk.LEFT, padx=(8, 0))

        self.packs_table = ttk.Treeview(
            packs_frame,
            columns=(
                "id",
                "name",
                "qty",
                "components_current",
                "components_new",
                "pack_current",
                "pack_new",
                "mode",
            ),
            show="headings",
        )
        self._setup_table(
            self.packs_table,
            {
                "id": ("ID Woo", 80),
                "name": ("Pack", 280),
                "qty": ("Cant.", 70),
                "components_current": ("Suma actual", 105),
                "components_new": ("Suma nueva", 105),
                "pack_current": ("Pack actual", 105),
                "pack_new": ("Pack nuevo", 105),
                "mode": ("Regla", 90),
            },
        )
        self.packs_table.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            self,
            textvariable=self.proposal_status_text,
            padding=(10, 0, 10, 8),
            anchor=tk.W,
        ).pack(fill=tk.X)

        footer = ttk.Frame(self, padding=(10, 0, 10, 10))
        footer.pack(fill=tk.X)
        ttk.Button(footer, text="Cerrar", command=self.destroy).pack(side=tk.RIGHT)
        ttk.Button(footer, text="Añadir a propuesta", command=self._add_to_proposal).pack(
            side=tk.RIGHT,
            padx=(0, 8),
        )

    def _setup_table(self, table: ttk.Treeview, columns: dict[str, tuple[str, int]]) -> None:
        for column, (heading, width) in columns.items():
            table.heading(column, text=heading, anchor=tk.CENTER)
            table.column(column, width=width, minwidth=60, anchor=tk.CENTER)

    def _load_items(self, initial_item: dict[str, object] | None) -> None:
        labels = []
        initial_label = None
        for item in self.catalog_items:
            label = (
                f"{item['item_kind']} | {item['woo_id']} | "
                f"{item.get('price') or '-'} | {item.get('name')}"
            )
            if item.get("attributes_label"):
                label += f" | {item.get('attributes_label')}"
            labels.append(label)
            self.item_by_label[label] = item
            if initial_item and item == initial_item:
                initial_label = label
        self.item_combo.configure(values=labels)
        if initial_label:
            self.item_text.set(initial_label)
        elif labels:
            self.item_text.set(labels[0])
        self._item_changed()

    def _item_changed(self) -> None:
        item = self._selected_item()
        current = self._price_of(item)
        self.current_price_text.set(self._format_money(current))
        self.new_price_text.set(self._format_money(current))
        self._recalculate()

    def _selected_item(self) -> dict[str, object] | None:
        return self.item_by_label.get(self.item_text.get())

    def _recalculate(self) -> None:
        item = self._selected_item()
        if not item:
            return
        current = self._price_of(item)
        new = self._safe_float(self.new_price_text.get())
        self.delta_text.set(self._format_signed(new - current))
        self.detail_text.set(
            f"Categorias: {item.get('categories_label') or 'Sin categoria'} | "
            f"Variacion: {item.get('attributes_label') or 'Sin variacion'}"
        )
        self._render_variations(item, current, new)
        self._render_related_impacts(item, new)
        self._render_pack_impacts(item, new)

    def _apply_percent_to_main_item(self) -> None:
        item = self._selected_item()
        if not item:
            return
        current = self._price_of(item)
        percent = self._safe_float(self.percent_increase_text.get())
        self.new_price_text.set(self._format_money(current * (1 + percent / 100)))
        self._recalculate()

    def _render_variations(
        self,
        item: dict[str, object],
        current: float,
        new: float,
    ) -> None:
        self.variations_table.delete(*self.variations_table.get_children())
        self.variation_rows = {}
        self.variation_proposals = {}
        self.marked_variations = set()
        if item.get("item_kind") == "product":
            rows = self.store.list_variations_for_product(int(item["woo_id"]))
            for row in rows:
                woo_id = int(row["woo_id"])
                row_current = self._safe_float(row["price"])
                self.variation_rows[woo_id] = dict(row)
                self.variation_proposals[woo_id] = row_current
            self._render_variations_table()
            return

        woo_id = int(item["woo_id"])
        self.variation_rows[woo_id] = {
            "woo_id": woo_id,
            "attributes_label": item.get("attributes_label") or "",
            "price": current,
            "stock_status": item.get("stock_status") or "",
            "name": item.get("name") or "",
        }
        self.variation_proposals[woo_id] = new
        self.marked_variations = {woo_id} if abs(new - current) > 0.005 else set()
        self._render_variations_table()

    def _render_variations_table(self) -> None:
        self.variations_table.delete(*self.variations_table.get_children())
        for woo_id, row in self.variation_rows.items():
            current = self._safe_float(row.get("price"))
            proposed = self.variation_proposals.get(woo_id, current)
            self.variations_table.insert(
                "",
                tk.END,
                iid=str(woo_id),
                values=(
                    "[x]" if woo_id in self.marked_variations else "[ ]",
                    woo_id,
                    row.get("attributes_label") or "",
                    self._format_money(current),
                    self._format_money(proposed),
                    self._format_signed(proposed - current),
                    row.get("stock_status") or "",
                ),
            )
        auto_fit_tree_columns(self.variations_table)

    def _variation_selection_changed(self) -> None:
        selection = self.variations_table.selection()
        if not selection:
            self.variation_price_text.set("")
            return
        woo_id = int(selection[0])
        self.variation_price_text.set(
            self._format_money(self.variation_proposals.get(woo_id, 0))
        )

    def _toggle_variation_mark(self) -> None:
        selection = self.variations_table.selection()
        if not selection:
            messagebox.showwarning("Seleccion necesaria", "Selecciona una variación.")
            return
        woo_id = int(selection[0])
        current = self._safe_float(self.variation_rows[woo_id].get("price"))
        if woo_id in self.marked_variations:
            self.marked_variations.remove(woo_id)
            self.variation_proposals[woo_id] = current
        else:
            self.marked_variations.add(woo_id)
            if abs(self.variation_proposals.get(woo_id, current) - current) <= 0.005:
                percent = self._safe_float(self.percent_increase_text.get())
                self.variation_proposals[woo_id] = current * (1 + percent / 100)
        self._render_variations_table()
        self.variations_table.selection_set(str(woo_id))

    def _mark_all_variations(self) -> None:
        if not self.variation_rows:
            messagebox.showwarning("Sin variaciones", "Este artículo no tiene variaciones para marcar.")
            return
        percent = self._safe_float(self.percent_increase_text.get())
        for woo_id, row in self.variation_rows.items():
            current = self._safe_float(row.get("price"))
            self.marked_variations.add(woo_id)
            if abs(self.variation_proposals.get(woo_id, current) - current) <= 0.005:
                self.variation_proposals[woo_id] = current * (1 + percent / 100)
        self._render_variations_table()

    def _update_variation_price(self) -> None:
        selection = self.variations_table.selection()
        if not selection:
            messagebox.showwarning("Seleccion necesaria", "Selecciona una variación.")
            return
        woo_id = int(selection[0])
        proposed = self._safe_float(self.variation_price_text.get())
        if proposed < 0:
            messagebox.showwarning("Precio invalido", "El precio no puede ser negativo.")
            return
        self.marked_variations.add(woo_id)
        self.variation_proposals[woo_id] = proposed
        self._render_variations_table()
        self.variations_table.selection_set(str(woo_id))

    def _apply_adjustment_to_marked_variations(self) -> None:
        if not self.marked_variations:
            messagebox.showwarning(
                "Seleccion necesaria",
                "Marca al menos una variación antes de aplicar el ajuste.",
            )
            return
        value = self._safe_float(self.variation_adjust_value.get())
        mode = self.variation_adjust_mode.get()
        for woo_id in self.marked_variations:
            row = self.variation_rows[woo_id]
            current = self._safe_float(row.get("price"))
            if mode == "percent":
                proposed = current * (1 + value / 100)
            else:
                proposed = current + value
            self.variation_proposals[woo_id] = max(proposed, 0)
        self._render_variations_table()

    def _render_pack_impacts(self, item: dict[str, object], new_price: float) -> None:
        self.packs_table.delete(*self.packs_table.get_children())
        impacts = self.store.list_manual_pack_impacts(
            str(item["item_kind"]),
            int(item["woo_id"]),
            new_price,
        )
        for impact in impacts:
            self.packs_table.insert(
                "",
                tk.END,
                values=(
                    impact["pack_woo_id"],
                    impact["name"],
                    self._format_quantity(float(impact["quantity"])),
                    self._format_money(float(impact["current_components_total"])),
                    self._format_money(float(impact["new_components_total"])),
                    self._format_money(float(impact["current_final_price"])),
                    self._format_money(float(impact["new_final_price"])),
                    impact["pricing_mode"],
                ),
            )
        auto_fit_tree_columns(self.packs_table)

    def _render_related_impacts(self, item: dict[str, object], new_price: float) -> None:
        self.related_table.delete(*self.related_table.get_children())
        impacts = self.store.list_related_variation_impacts(
            str(item["item_kind"]),
            int(item["woo_id"]),
            new_price,
        )
        self.related_rows = {int(impact["woo_id"]): impact for impact in impacts}
        self.related_proposals = {
            int(impact["woo_id"]): float(impact["current_price"]) for impact in impacts
        }
        self.marked_related = set()
        self._render_related_table()

    def _render_related_table(self) -> None:
        self.related_table.delete(*self.related_table.get_children())
        for woo_id, impact in self.related_rows.items():
            current = float(impact["current_price"])
            proposed = self.related_proposals.get(woo_id, current)
            self.related_table.insert(
                "",
                tk.END,
                iid=str(woo_id),
                values=(
                    "[x]" if woo_id in self.marked_related else "[ ]",
                    woo_id,
                    impact["name"],
                    impact["attributes_label"],
                    self._format_money(current),
                    self._format_money(proposed),
                    self._format_signed(proposed - current),
                    impact["reason"],
                ),
            )
        auto_fit_tree_columns(self.related_table)

    def _related_selection_changed(self) -> None:
        selection = self.related_table.selection()
        if not selection:
            self.related_price_text.set("")
            return
        woo_id = int(selection[0])
        self.related_price_text.set(
            self._format_money(self.related_proposals.get(woo_id, 0))
        )

    def _toggle_related_mark(self) -> None:
        selection = self.related_table.selection()
        if not selection:
            messagebox.showwarning("Seleccion necesaria", "Selecciona una variacion relacionada.")
            return
        woo_id = int(selection[0])
        if woo_id in self.marked_related:
            self.marked_related.remove(woo_id)
            self.related_proposals[woo_id] = float(self.related_rows[woo_id]["current_price"])
        else:
            self.marked_related.add(woo_id)
        self._render_related_table()
        self.related_table.selection_set(str(woo_id))

    def _update_related_price(self) -> None:
        selection = self.related_table.selection()
        if not selection:
            messagebox.showwarning("Seleccion necesaria", "Selecciona una variacion relacionada.")
            return
        woo_id = int(selection[0])
        proposed = self._safe_float(self.related_price_text.get())
        if proposed < 0:
            messagebox.showwarning("Precio invalido", "El precio no puede ser negativo.")
            return
        self.marked_related.add(woo_id)
        self.related_proposals[woo_id] = proposed
        self._render_related_table()
        self.related_table.selection_set(str(woo_id))

    def _apply_adjustment_to_marked_related(self) -> None:
        if not self.marked_related:
            messagebox.showwarning(
                "Seleccion necesaria",
                "Marca al menos una variacion relacionada antes de aplicar el ajuste.",
            )
            return
        value = self._safe_float(self.related_adjust_value.get())
        mode = self.related_adjust_mode.get()
        for woo_id in self.marked_related:
            impact = self.related_rows[woo_id]
            current = float(impact["current_price"])
            if mode == "percent":
                proposed = current * (1 + value / 100)
            else:
                proposed = current + value
            self.related_proposals[woo_id] = max(proposed, 0)
        self._render_related_table()

    def _add_to_proposal(self) -> None:
        item = self._selected_item()
        if not item:
            return
        proposals: list[dict[str, object]] = []

        current = self._price_of(item)
        new = self._safe_float(self.new_price_text.get())
        if new < 0:
            messagebox.showwarning("Precio invalido", "El precio no puede ser negativo.")
            return
        if abs(new - current) > 0.005:
            proposals.append(
                {
                    "item_kind": item["item_kind"],
                    "item_woo_id": item["woo_id"],
                    "name": self._item_name(item),
                    "old_price": current,
                    "new_price": new,
                    "delta": new - current,
                    "notes": "Cambio añadido desde Cambio de Precios",
                }
            )

        for woo_id in self.marked_variations:
            row = self.variation_rows.get(woo_id)
            if not row:
                continue
            current_price = self._safe_float(row.get("price"))
            proposed = self.variation_proposals.get(woo_id, current_price)
            if proposed < 0:
                messagebox.showwarning("Precio invalido", "Una variación marcada tiene precio negativo.")
                return
            if abs(proposed - current_price) <= 0.005:
                continue
            proposals.append(
                {
                    "item_kind": "variation",
                    "item_woo_id": woo_id,
                    "name": self._item_name(
                        {
                            "name": item.get("name") or row.get("name") or "",
                            "attributes_label": row.get("attributes_label") or "",
                        }
                    ),
                    "old_price": current_price,
                    "new_price": proposed,
                    "delta": proposed - current_price,
                    "notes": "Variación marcada desde Cambio de Precios",
                }
            )

        for woo_id, impact in self.related_rows.items():
            if woo_id not in self.marked_related:
                continue
            current_price = float(impact["current_price"])
            proposed = self.related_proposals.get(woo_id, current_price)
            if proposed < 0:
                messagebox.showwarning("Precio invalido", "Una variación relacionada tiene precio negativo.")
                return
            if abs(proposed - current_price) <= 0.005:
                continue
            proposals.append(
                {
                    "item_kind": "variation",
                    "item_woo_id": woo_id,
                    "name": self._item_name(
                        {
                            "name": impact["name"],
                            "attributes_label": impact["attributes_label"],
                        }
                    ),
                    "old_price": current_price,
                    "new_price": proposed,
                    "delta": proposed - current_price,
                    "notes": f"Relación local: {impact['reason']}",
                }
            )

        if not proposals:
            messagebox.showwarning(
                "Sin cambios",
                "No hay precios cambiados ni variaciones marcadas para añadir.",
            )
            return

        added_count = self.store.upsert_pending_price_change_proposals(proposals)
        self.proposal_status_text.set(
            f"Añadidos/actualizados en propuesta: {added_count}. "
            "Puedes elegir otro artículo y seguir añadiendo. WooCommerce no se ha modificado."
        )
        if callable(self.on_added):
            self.on_added(added_count)
        # La ventana queda abierta para construir propuestas grandes sin entrar y salir.

    def _item_name(self, item: dict[str, object]) -> str:
        name = str(item.get("name") or "")
        variation = str(item.get("attributes_label") or "")
        if variation:
            return f"{name} - {variation}"
        return name

    def _price_of(self, item: dict[str, object] | None) -> float:
        if not item:
            return 0
        return self._safe_float(item.get("price"))

    def _safe_float(self, value: object) -> float:
        if value is None:
            return 0.0
        try:
            return float(str(value).strip().replace(",", ".") or "0")
        except ValueError:
            return 0.0

    def _format_money(self, value: float) -> str:
        return f"{value:.2f}"

    def _format_signed(self, value: float) -> str:
        return f"{value:+.2f}"

    def _format_quantity(self, value: float) -> str:
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}"
