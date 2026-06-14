from __future__ import annotations

import csv
import json
import os
import re
import tkinter as tk
import threading
import unicodedata
import warnings
from dataclasses import dataclass
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from futonhub.cloud.audit import list_audit_logs as legacy_list_audit_logs, list_operation_snapshots as legacy_list_operation_snapshots
from futonhub.cloud.services.security_logs import (
    build_before_after_diff,
    export_security_logs_excel,
    get_snapshot_by_operation,
    list_audit_logs as list_security_audit_logs,
    preview_restore_snapshot,
    restore_snapshot_to_previous_state,
    security_log_kpis,
)
from futonhub.cloud.auth import SupabaseAuthError, register_device_seen, sign_in_with_password
from futonhub.cloud.services.inventory import (
    create_cloud_inventory_item,
    fetch_inventory_item_history,
    fetch_inventory_pack_components,
    list_cloud_inventory_items,
    preview_create_cloud_inventory_item,
    preview_internal_inventory_update,
    search_cloud_inventory_items,
    update_inventory_item_fields,
)
from futonhub.cloud.services.price_proposals import (
    create_real_price_proposal,
    delete_real_price_proposal_group,
    format_existing_price_proposal_preview,
    list_real_price_proposals,
    preview_existing_price_proposal,
    review_latest_real_price_proposal,
)
from futonhub.cloud.services.orders import (
    cancel_supplier_order,
    create_supplier_order_draft,
    format_order_date,
    list_cloud_supplier_order_items,
    list_cloud_supplier_orders,
    order_display_name as cloud_order_display_name,
    preview_receive_supplier_order,
    receive_supplier_order,
    summarize_order_items,
    update_supplier_order_draft,
    update_supplier_order_calculation,
)
from futonhub.cloud.services.woocommerce_publish import publish_woocommerce_price
from futonhub.cloud.services.woocommerce_publish import format_woocommerce_publish_preview, preview_woocommerce_publish
from futonhub.cloud.services.woocommerce_sync_preview import (
    apply_manual_classification_edit,
    apply_manual_woo_link,
    build_sync_preview as build_woo_sync_preview,
    export_preview_json as export_woo_preview_json,
    format_manual_woo_link_preview,
    preview_manual_woo_link,
    search_manual_link_inventory_candidates,
)
from futonhub.cloud.services.supplier_prices import (
    get_supplier_price,
    list_supplier_price_inventory_items,
    update_supplier_price_inventory_item,
)
from futonhub.cloud.services.business_constants import DEFAULT_BUSINESS_CONSTANTS, list_business_constants, save_business_constants
from futonhub.core.config import load_settings
from futonhub.core.guard import active_locks, stale_locks
from futonhub.ui.theme import apply_theme
from futonhub.ui.windowing import center_window


BG = "#F8FAFC"
SIDEBAR = "#FFFFFF"
CARD = "#FFFFFF"
LINE = "#E2E8F0"
SOFT = "#F1F5F9"
TEXT = "#0F172A"
MUTED = "#64748B"
INDIGO = "#4F46E5"
INDIGO_SOFT = "#EEF2FF"
GREEN = "#059669"
GREEN_SOFT = "#ECFDF5"
BLUE = "#2563EB"
BLUE_SOFT = "#EFF6FF"
AMBER = "#D97706"
AMBER_SOFT = "#FFFBEB"
ORANGE = "#EA580C"
ORANGE_SOFT = "#FFF7ED"
ROSE = "#E11D48"
ROSE_SOFT = "#FFF1F2"


@dataclass(frozen=True)
class NavItem:
    key: str
    label: str
    group: str


@dataclass(frozen=True)
class InventoryItem:
    code: str
    name: str
    price: str
    stock: str
    status: str
    family: str
    provider: str
    m3: str
    sku_woo: str
    measures: str
    material: str
    sync_woo: str
    notes: str
    subgroup: str = "-"
    store_stock: str = "-"
    warehouse_stock: str = "-"
    stock_total: str = "-"
    woo_id: str = "-"
    woo_parent_id: str = "-"
    woo_name: str = "-"
    woo_price: str = "-"
    woo_categories: str = "-"
    woo_item_kind: str = "-"
    woo_link_status: str = "-"
    order_calculated_price: str = "-"
    weighted_average_cost: str = "-"
    supplier_order_qty: str = "-"
    supplier_order_provider: str = "-"
    status_reasons: tuple[str, ...] = ()
    raw: dict[str, Any] | None = None


@dataclass(frozen=True)
class ProposalLine:
    code: str
    name: str
    old_price: str
    new_price: str
    change: str
    direction: str


@dataclass(frozen=True)
class PriceProposal:
    name: str
    date: str
    items: int
    up: int
    down: int
    flat: int
    change: str
    status: str
    lines: tuple[ProposalLine, ...]
    raw: dict[str, Any] | None = None


@dataclass(frozen=True)
class OrderItem:
    code: str
    name: str
    quantity: int
    m3: str
    final_cost: str
    status: str
    raw: dict[str, Any] | None = None


@dataclass(frozen=True)
class SupplierOrder:
    order_id: str
    provider: str
    date: str
    items_count: int
    total_m3: str
    status: str
    total_cost: str
    notes: str
    items: tuple[OrderItem, ...]
    raw: dict[str, Any] | None = None


@dataclass(frozen=True)
class WooDifference:
    local_id: str
    woo_id: str
    name: str
    field: str
    local_value: str
    woo_value: str
    difference: str
    classification: str
    action: str
    status: str
    detail: str


@dataclass(frozen=True)
class ExportRecord:
    date: str
    report_type: str
    module: str
    format: str
    user_role: str
    status: str
    file_name: str
    rows: str
    reference: str
    filters: str
    columns: str
    path: str


@dataclass(frozen=True)
class SecurityEvent:
    date: str
    level: str
    module: str
    action: str
    user_role: str
    result: str
    reference: str
    message: str
    payload: str


@dataclass(frozen=True)
class SecurityLogRow:
    created_at: str
    user_email: str
    role: str
    visual_module: str
    visual_action: str
    status: str
    severity: str
    entity_type: str
    entity_id: str
    operation_id: str
    message: str
    raw: dict[str, Any] | None = None


NAV_ITEMS = [
    NavItem("dashboard", "Dashboard", "Principal"),
    NavItem("inventario", "Inventario", "Operaciones"),
    NavItem("precios", "Cambio de Precios", "Operaciones"),
    NavItem("calcular", "Pedidos", "Operaciones"),
    NavItem("woocommerce", "WooCommerce", "Gestion"),
    NavItem("precios_proveedor", "Precio Proveedores", "Gestion"),
    NavItem("informes", "Informes / Exportaciones", "Gestion"),
    NavItem("seguridad", "Seguridad / Logs", "Sistema"),
    NavItem("configuracion", "Configuracion", "Sistema"),
]


INVENTORY_ITEMS = [
    InventoryItem(
        "FUT-ALG-140",
        "Futon algodon 140x200",
        "249.00",
        "8",
        "OK",
        "Futones",
        "Ekomat",
        "0.42",
        "woo-3841",
        "140 x 200 x 14 cm",
        "Algodon",
        "Sin diferencias criticas",
        "Producto revisado. Sin bloqueo para propuesta de precio.",
    ),
    InventoryItem(
        "TAT-160",
        "Tatami 160x200",
        "315.00",
        "3",
        "Info",
        "Tatamis",
        "Heimei",
        "0.31",
        "woo-2210",
        "160 x 200 x 5.5 cm",
        "Paja de arroz",
        "Pendiente de comparacion",
        "Revisar condiciones del proveedor antes de pedido grande.",
    ),
    InventoryItem(
        "SOFA-IND-01",
        "Sofa cama individual",
        "0.00",
        "0",
        "Critical",
        "Sofas",
        "Pascal",
        "0.86",
        "woo-611",
        "90 x 200 cm",
        "Madera y algodon",
        "Precio Woo distinto al interno",
        "Precio interno en cero. Bloquea propuesta y publicacion.",
    ),
    InventoryItem(
        "FUNDA-90",
        "Funda futon 90 cm",
        "48.00",
        "21",
        "Warning",
        "Fundas",
        "Cipta",
        "0.08",
        "woo-5122",
        "90 x 200 cm",
        "Algodon",
        "Sincronizacion pendiente",
        "Confirmar color y variante antes de incluir en propuesta.",
    ),
    InventoryItem(
        "COLCH-LAT-150",
        "Colchon latex 150x200",
        "429.00",
        "2",
        "OK",
        "Colchones",
        "Ekomat",
        "0.58",
        "woo-7740",
        "150 x 200 x 18 cm",
        "Latex",
        "Sin diferencias criticas",
        "Stock bajo pero operativo.",
    ),
]


DEFAULT_PROPOSAL_LINES = (
    ProposalLine("FUT-ALG-140", "Futon algodon 140x200", "249.00", "269.00", "+8.03%", "up"),
    ProposalLine("TAT-160", "Tatami 160x200", "315.00", "329.00", "+4.44%", "up"),
    ProposalLine("FUNDA-90", "Funda futon 90 cm", "48.00", "51.00", "+6.25%", "up"),
    ProposalLine("OUTLET-023", "Base outlet exposicion", "220.00", "199.00", "-9.55%", "down"),
    ProposalLine("MES-001", "Mesita madera natural", "89.00", "89.00", "0.00%", "flat"),
    ProposalLine("COLCH-LAT-150", "Colchon latex 150x200", "429.00", "449.00", "+4.66%", "up"),
)


SAVED_PROPOSALS = [
    PriceProposal("Subida general junio", "28/05/2026", 23, 19, 1, 3, "+6.8%", "Warning", DEFAULT_PROPOSAL_LINES),
    PriceProposal(
        "Revision tatamis Heimei",
        "27/05/2026",
        8,
        8,
        0,
        0,
        "+4.1%",
        "OK",
        (
            ProposalLine("TAT-160", "Tatami 160x200", "315.00", "329.00", "+4.44%", "up"),
            ProposalLine("TAT-90", "Tatami 90x200", "179.00", "185.00", "+3.35%", "up"),
        ),
    ),
    PriceProposal(
        "Correccion productos criticos",
        "26/05/2026",
        3,
        0,
        0,
        3,
        "Bloq.",
        "Critical",
        (ProposalLine("SOFA-IND-01", "Sofa cama individual", "0.00", "389.00", "Bloqueado", "critical"),),
    ),
    PriceProposal("Complementos temporada", "24/05/2026", 11, 7, 0, 4, "+2.3%", "Info", DEFAULT_PROPOSAL_LINES[:3]),
    PriceProposal("Ajuste outlet", "22/05/2026", 6, 0, 6, 0, "-8.5%", "Error", DEFAULT_PROPOSAL_LINES[3:5]),
]


ORDER_ITEMS_HEIMEI = (
    OrderItem("1244001", "Tatami 90x200", 12, "3.12", "1248.00", "OK"),
    OrderItem("0730009", "Producto especial que cuenta", 4, "1.18", "516.00", "OK"),
    OrderItem("TAT-160", "Tatami 160x200", 6, "3.42", "1890.00", "Warning"),
    OrderItem("A MEDIDA", "Pedido cliente personalizado", 1, "Pendiente", "Bloqueado", "Error"),
)


SUPPLIER_ORDERS = [
    SupplierOrder("PED-HEI-028", "Heimei", "28/05/2026", 23, "8.42", "Warning", "3654.00", "Validacion pendiente por item a medida.", ORDER_ITEMS_HEIMEI),
    SupplierOrder(
        "PED-EKO-017",
        "Ekomat",
        "27/05/2026",
        41,
        "12.10",
        "OK",
        "6840.00",
        "Pedido calculado y listo para exportar.",
        (
            OrderItem("FUT-ALG-140", "Futon algodon 140x200", 8, "3.36", "1992.00", "OK"),
            OrderItem("COLCH-LAT-150", "Colchon latex 150x200", 5, "2.90", "2145.00", "OK"),
        ),
    ),
    SupplierOrder(
        "PED-PAS-011",
        "Pascal",
        "25/05/2026",
        14,
        "6.88",
        "Info",
        "2790.00",
        "Borrador guardado para revision.",
        (OrderItem("SOFA-IND-01", "Sofa cama individual", 3, "2.58", "1167.00", "Info"),),
    ),
    SupplierOrder(
        "PED-EKO-016",
        "Ekomat",
        "22/05/2026",
        29,
        "9.02",
        "Error",
        "Bloqueado",
        "Faltan datos M3 en una linea.",
        (OrderItem("A MEDIDA", "Pedido personalizado", 1, "Pendiente", "Bloqueado", "Error"),),
    ),
]


WOO_DIFFERENCES = [
    WooDifference(
        "FUT-ALG-140",
        "3841",
        "Futon algodon 140x200",
        "Precio",
        "249.00 EUR",
        "269.00 EUR",
        "Woo +20.00 EUR",
        "Futones - Algodon",
        "Actualizar local",
        "Warning",
        "WooCommerce tiene un precio mas alto que la base local. Pendiente de traer el cambio al HUB.",
    ),
    WooDifference(
        "TAT-160",
        "4102",
        "Tatami 160x200",
        "Stock",
        "2",
        "3",
        "Woo +1",
        "Tatamis",
        "Actualizar local",
        "Info",
        "Diferencia de stock leve. Puede actualizarse localmente tras revision.",
    ),
    WooDifference(
        "SOFA-IND-01",
        "611",
        "Sofa cama individual",
        "Familia",
        "Sin definir",
        "Sofa cama",
        "Clasificacion nueva",
        "Sofa camas",
        "Auto-clasificar",
        "Warning",
        "Woo aporta familia util para organizar la base local.",
    ),
    WooDifference(
        "FUNDA-90",
        "5120",
        "Funda futon 90 cm",
        "Material / grupo",
        "Sin clasificar",
        "Funda algodon",
        "Clasificacion nueva",
        "Complementos - Fundas",
        "Auto-clasificar",
        "OK",
        "La clasificacion propuesta es limpia y no presenta conflicto visible.",
    ),
    WooDifference(
        "PORT-001",
        "5201",
        "Futon portatil plegable",
        "Familia",
        "Fundas",
        "Futon portatil",
        "Posible mala familia",
        "Revisar manual",
        "Revisar",
        "Error",
        "La familia local parece incompatible con el producto Woo. Debe revisarse manualmente.",
    ),
    WooDifference(
        "SYNC-000",
        "0",
        "Producto Woo sin pareja local",
        "Relacion",
        "No existe",
        "Woo ID 7009",
        "Sin enlace local",
        "Pendiente",
        "Revisar manual",
        "Critical",
        "No se puede actualizar nada hasta enlazar o descartar este producto.",
    ),
]


EXPORT_RECORDS = [
    ExportRecord(
        "31/05/2026 19:05",
        "Inventario completo",
        "Inventario",
        "XLSX",
        "Admin",
        "OK",
        "inventario_20260531.xlsx",
        "842",
        "EXP-INV-1905",
        "Familias: todas - Estado: todos - Stock: todos",
        "ID, nombre, precio, stock, familia, proveedor, M3, estado",
        "/exports/2026/05/inventario_20260531.xlsx",
    ),
    ExportRecord(
        "31/05/2026 18:44",
        "Incidencias WooCommerce",
        "WooCommerce",
        "XLSX",
        "Admin",
        "Warning",
        "woo_incidencias_1844.xlsx",
        "27",
        "EXP-WOO-1844",
        "Estado: Warning/Error - Accion: revisar",
        "ID local, ID Woo, diferencia, clasificacion, accion, estado",
        "/exports/2026/05/woo_incidencias_1844.xlsx",
    ),
    ExportRecord(
        "31/05/2026 18:20",
        "Detalle de pedido",
        "Pedidos",
        "PDF",
        "Admin",
        "OK",
        "PED-HEI-029_detalle.pdf",
        "23",
        "EXP-PED-1820",
        "Proveedor: Heimei - Pedido: PED-HEI-029",
        "ID, nombre, cantidad, M3, coste final",
        "/exports/2026/05/PED-HEI-029_detalle.pdf",
    ),
    ExportRecord(
        "31/05/2026 17:58",
        "Propuesta de precios",
        "Cambio de Precios",
        "XLSX",
        "Admin",
        "OK",
        "propuesta_junio.xlsx",
        "23",
        "EXP-PRC-1758",
        "Propuesta: Subida general junio",
        "ID, nombre, precio anterior, precio nuevo, cambio, estado",
        "/exports/2026/05/propuesta_junio.xlsx",
    ),
    ExportRecord(
        "31/05/2026 17:35",
        "Auditoria de logs",
        "Seguridad / Logs",
        "PDF",
        "Admin",
        "OK",
        "auditoria_logs_1735.pdf",
        "128",
        "EXP-LOG-1735",
        "Nivel: todos - Modulo: todos",
        "fecha, nivel, modulo, accion, usuario, resultado, referencia",
        "/exports/2026/05/auditoria_logs_1735.pdf",
    ),
    ExportRecord(
        "31/05/2026 16:50",
        "Pedido calculado",
        "Pedidos",
        "XLSX",
        "Admin",
        "Error",
        "No generado",
        "0",
        "EXP-PED-1650",
        "Proveedor: Heimei - faltan M3",
        "No disponible",
        "No generado",
    ),
]


STATUS_STYLES = {
    "OK": (GREEN, GREEN_SOFT),
    "Info": (BLUE, BLUE_SOFT),
    "Warning": (AMBER, AMBER_SOFT),
    "Error": (ORANGE, ORANGE_SOFT),
    "Critical": (ROSE, ROSE_SOFT),
    "Pendiente": (AMBER, AMBER_SOFT),
    "Aprobada": (GREEN, GREEN_SOFT),
    "Rechazada": (ROSE, ROSE_SOFT),
    "Publicando": (BLUE, BLUE_SOFT),
    "Publicada": (GREEN, GREEN_SOFT),
    "Fallida": (ROSE, ROSE_SOFT),
    # Estados de pedidos guardados en Supabase.
    "Borrador": (BLUE, BLUE_SOFT),
    "Pendiente archivo": (AMBER, AMBER_SOFT),
    "Validacion": (AMBER, AMBER_SOFT),
    "Validación": (AMBER, AMBER_SOFT),
    "Calculado": (GREEN, GREEN_SOFT),
    "Guardado": (BLUE, BLUE_SOFT),
    "Recibido parcial": (AMBER, AMBER_SOFT),
    "Recibido completo": (GREEN, GREEN_SOFT),
    "Exportado": (GREEN, GREEN_SOFT),
    "Cancelado": (ROSE, ROSE_SOFT),
}


class FutonHubErpPrototype(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("FutonHUB - UI ERP Prototype")
        center_window(self, 1280, 760)
        self.minsize(1100, 680)
        self.configure(bg=BG)
        apply_theme(self)
        self._configure_erp_table_style()
        self._nav_buttons: dict[str, tk.Button] = {}
        self._content: tk.Frame | None = None
        self._status_area: tk.Frame | None = None
        self._cloud_session = None
        self._login_in_progress = False
        self._login_window: tk.Toplevel | None = None
        self._login_loading_window: tk.Toplevel | None = None
        self._current_key = "dashboard"
        self._inventory_items: list[InventoryItem] = []
        self._inventory_error = ""
        self._inventory_loading = False
        self._inventory_loaded_once = False
        self._inventory_query = ""
        self._selected_inventory_item: InventoryItem | None = None
        self._proposal_source_item: InventoryItem | None = None
        self._selected_price_proposal = SAVED_PROPOSALS[0]
        self._price_proposals = list(SAVED_PROPOSALS)
        self._price_loading = False
        self._price_loaded_once = False
        self._price_error = ""
        self._price_mode = "saved"
        self._price_edit_lines: list[ProposalLine] = []
        self._price_edit_initialized = False
        self._price_edit_selected_code = ""
        self._price_edit_notice = ""
        self._price_search_query = ""
        self._proposal_search_query = ""
        self._price_delete_target_id = ""
        self._price_delete_target_name = ""
        self._price_available_items: list[InventoryItem] = []
        self._price_items_loading = False
        self._price_items_error = ""
        self._price_line_sources: dict[str, dict[str, Any]] = {}
        self._supplier_orders: list[SupplierOrder] = []
        self._orders_loaded_once = False
        self._orders_loading = False
        self._orders_error = ""
        self._selected_supplier_order: SupplierOrder | None = None
        self._selected_woo_difference = WOO_DIFFERENCES[0]
        self._woo_sync_preview: dict[str, Any] | None = None
        self._woo_sync_rows: list[dict[str, Any]] = []
        self._woo_sync_loading = False
        self._woo_sync_error = ""
        self._woo_sync_filter_text = ""
        self._woo_sync_filter_review = "Todos"
        self._woo_sync_filter_status = "Todos"
        self._woo_sync_filter_link = "Todos"
        self._selected_export_record = EXPORT_RECORDS[0]
        self._settings_tab = "Generales"
        self._business_constants: dict[str, dict[str, Any]] = {key: dict(value) for key, value in DEFAULT_BUSINESS_CONSTANTS.items()}
        self._security_events: list[SecurityEvent] = []
        self._security_log_rows: list[dict[str, Any]] = []
        self._security_visible_rows: list[dict[str, Any]] = []
        self._security_snapshots: list[dict[str, Any]] = []
        self._security_local_locks = 0
        self._security_stale_locks = 0
        self._security_error = ""
        self._security_filter_text = ""
        self._security_filter_user = ""
        self._security_filter_module = ""
        self._security_filter_status = ""
        self._security_filter_severity = ""
        self._security_filter_date_from = ""
        self._security_filter_date_to = ""
        self.withdraw()
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.after(0, self._show_startup_login)

    def _configure_erp_table_style(self) -> None:
        """Apply the ERP table style: centered text and clearer row separation."""
        style = ttk.Style(self)
        style.configure(
            "Treeview",
            background=CARD,
            fieldbackground=CARD,
            foreground=TEXT,
            bordercolor=LINE,
            lightcolor=LINE,
            darkcolor=LINE,
            borderwidth=1,
            relief="solid",
            rowheight=30,
        )
        style.configure(
            "Treeview.Heading",
            background=SOFT,
            foreground=TEXT,
            font=("Segoe UI", 9, "bold"),
            bordercolor=LINE,
            lightcolor=LINE,
            darkcolor=LINE,
            borderwidth=1,
            relief="solid",
            padding=(8, 7),
        )
        # Importante: no fijamos un color de fondo para filas no seleccionadas.
        # Si lo hacemos, Tkinter/ttk pisa los colores por tags y no se ven las
        # filas rojas/amarillas de error o warning en tablas como Pedidos.
        style.map(
            "Treeview",
            background=[("selected", INDIGO)],
            foreground=[("selected", "#FFFFFF")],
        )

    def _build_shell(self) -> None:
        shell = tk.Frame(self, bg=BG)
        shell.pack(fill=tk.BOTH, expand=True)
        shell.columnconfigure(1, weight=1)
        shell.rowconfigure(0, weight=1)

        sidebar = tk.Frame(shell, bg=SIDEBAR, width=270, highlightbackground=LINE, highlightthickness=1)
        sidebar.grid(row=0, column=0, sticky="ns")
        sidebar.grid_propagate(False)
        self._build_sidebar(sidebar)

        main = tk.Frame(shell, bg=BG)
        main.grid(row=0, column=1, sticky="nsew")
        main.rowconfigure(1, weight=1)
        main.columnconfigure(0, weight=1)

        self._build_topbar(main)
        self._content = tk.Frame(main, bg=BG)
        self._content.grid(row=1, column=0, sticky="nsew", padx=24, pady=22)

    def _build_sidebar(self, parent: tk.Frame) -> None:
        brand = tk.Frame(parent, bg=SIDEBAR)
        brand.pack(fill=tk.X, padx=18, pady=(20, 28))
        logo = tk.Label(brand, text="F", bg=INDIGO, fg="white", font=("Segoe UI", 18, "bold"), width=2, height=1)
        logo.pack(side=tk.LEFT)
        text = tk.Frame(brand, bg=SIDEBAR)
        text.pack(side=tk.LEFT, padx=(12, 0))
        tk.Label(text, text="FutonHUB", bg=SIDEBAR, fg=TEXT, font=("Segoe UI", 15, "bold")).pack(anchor=tk.W)
        tk.Label(text, text="ERP privado - prototipo", bg=SIDEBAR, fg=MUTED, font=("Segoe UI", 9)).pack(anchor=tk.W)

        grouped: dict[str, list[NavItem]] = {}
        is_admin = self._cloud_session is not None and str(getattr(self._cloud_session, "role", "") or "").lower() == "admin"
        for item in NAV_ITEMS:
            if item.key == "seguridad" and not is_admin:
                continue
            grouped.setdefault(item.group, []).append(item)

        for group, items in grouped.items():
            tk.Label(
                parent,
                text=group.upper(),
                bg=SIDEBAR,
                fg="#94A3B8",
                font=("Segoe UI", 8, "bold"),
                anchor=tk.W,
            ).pack(fill=tk.X, padx=22, pady=(10, 4))
            for item in items:
                button = tk.Button(
                    parent,
                    text=item.label,
                    anchor=tk.W,
                    bd=0,
                    relief=tk.FLAT,
                    padx=14,
                    pady=10,
                    font=("Segoe UI", 10, "bold"),
                    command=lambda key=item.key: self._show_view(key),
                )
                button.pack(fill=tk.X, padx=14, pady=2)
                self._nav_buttons[item.key] = button

    def _build_topbar(self, parent: tk.Frame) -> None:
        topbar = tk.Frame(parent, bg=BG, highlightbackground=LINE, highlightthickness=1)
        topbar.grid(row=0, column=0, sticky="ew")
        topbar.columnconfigure(0, weight=1)

        search = tk.Frame(topbar, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        search.grid(row=0, column=0, sticky="ew", padx=20, pady=14)
        tk.Label(search, text="Buscar producto, proveedor, informe o incidencia...", bg=CARD, fg=MUTED, anchor=tk.W).pack(
            fill=tk.X,
            padx=14,
            pady=10,
        )

        status = tk.Frame(topbar, bg=BG)
        status.grid(row=0, column=1, padx=(0, 20), pady=14)
        self._status_area = status
        self._render_session_status()

    def _render_session_status(self) -> None:
        if self._status_area is None:
            return
        for child in self._status_area.winfo_children():
            child.destroy()
        role = self._cloud_session.role or "sin rol"
        self._status_chip(self._status_area, "Online", "OK").pack(side=tk.LEFT, padx=(0, 8))
        self._status_chip(self._status_area, role.title(), "Info").pack(side=tk.LEFT)

    def _show_startup_login(self) -> None:
        try:
            settings = load_settings()
            default_email = settings.hub_user_email or ""
        except Exception as exc:
            messagebox.showerror("Login Supabase", f"No se pudo leer configuracion.\n\n{exc}")
            self.destroy()
            return

        win = tk.Toplevel(self)
        self._login_window = win
        win.title("FutonHUB - Login")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.protocol("WM_DELETE_WINDOW", self.destroy)

        card = tk.Frame(win, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=22, pady=22)

        logo = tk.Label(card, text="F", bg=INDIGO, fg="white", font=("Segoe UI", 18, "bold"), width=2, height=1)
        logo.pack(anchor=tk.W, padx=24, pady=(24, 12))
        tk.Label(card, text="Entrar a FutonHUB", bg=CARD, fg=TEXT, font=("Segoe UI", 20, "bold")).pack(anchor=tk.W, padx=24)
        tk.Label(
            card,
            text="Valida usuario, rol y dispositivo antes de abrir el ERP.",
            bg=CARD,
            fg=MUTED,
            font=("Segoe UI", 10),
        ).pack(anchor=tk.W, padx=24, pady=(6, 18))

        form = tk.Frame(card, bg=CARD)
        form.pack(fill=tk.X, padx=24)
        email_var = tk.StringVar(value=default_email)
        password_var = tk.StringVar()
        self._field(form, "Usuario", email_var).pack(fill=tk.X, pady=(0, 12))
        self._field(form, "Contrasena", password_var, show="*").pack(fill=tk.X, pady=(0, 8))
        error_label = tk.Label(card, text="", bg=CARD, fg=ROSE, font=("Segoe UI", 9), anchor=tk.W, justify=tk.LEFT)
        error_label.pack(fill=tk.X, padx=24, pady=(0, 8))

        actions = tk.Frame(card, bg=CARD)
        actions.pack(fill=tk.X, padx=24, pady=(14, 24))
        login_button = self._button(
            actions,
            "Aceptar",
            primary=True,
            command=lambda: self._login_supabase(email_var, password_var, error_label, login_button),
        )
        login_button.configure(width=12)
        login_button.pack(side=tk.RIGHT)
        cancel_button = self._button(actions, "Cancelar", command=self.destroy)
        cancel_button.configure(width=12)
        cancel_button.pack(side=tk.RIGHT, padx=(0, 8))
        win.bind("<Return>", lambda _event: self._login_supabase(email_var, password_var, error_label, login_button))
        win.bind("<Escape>", lambda _event: self.destroy())
        win.update_idletasks()
        width = max(520, win.winfo_reqwidth() + 48)
        height = max(460, win.winfo_reqheight() + 48)
        center_window(win, width, height)
        win.grab_set()
        win.lift()
        win.focus_force()
        for child in form.winfo_children():
            entries = [grandchild for grandchild in child.winfo_children() if isinstance(grandchild, tk.Entry)]
            if entries:
                entries[0].focus_set()
                break

    def _login_supabase(
        self,
        email_var: tk.StringVar,
        password_var: tk.StringVar,
        error_label: tk.Label,
        login_button: tk.Button,
    ) -> None:
        if self._login_in_progress:
            return
        try:
            settings = load_settings()
        except Exception as exc:
            error_label.configure(text=f"No se pudo leer configuracion: {exc}")
            return
        if settings.app_mode != "supabase_guarded":
            error_label.configure(text="El ERP requiere modo supabase_guarded. Revisa GestorWoo/.env.")
            return
        email = email_var.get().strip()
        if not email:
            error_label.configure(text="Introduce usuario.")
            return
        password = password_var.get()
        if not password:
            error_label.configure(text="Introduce contrasena.")
            return
        self._login_in_progress = True
        error_label.configure(text="")
        login_button.configure(state=tk.DISABLED, text="Validando...")
        self._show_login_loading()

        def worker() -> None:
            try:
                session = sign_in_with_password(email, password, settings)
                register_device_seen(session, settings)
            except Exception as exc:
                self.after(0, lambda exc=exc: self._finish_login(None, exc, error_label, login_button))
                return
            self.after(0, lambda session=session: self._finish_login(session, None, error_label, login_button))

        threading.Thread(target=worker, daemon=True).start()

    def _finish_login(
        self,
        session,
        exc: Exception | None,
        error_label: tk.Label,
        login_button: tk.Button,
    ) -> None:
        self._login_in_progress = False
        self._hide_login_loading()
        if exc is not None:
            if isinstance(exc, SupabaseAuthError):
                error_label.configure(text=str(exc))
            else:
                error_label.configure(text=f"No se pudo iniciar sesion: {exc}")
            login_button.configure(state=tk.NORMAL, text="Aceptar")
            if self._login_window is not None and self._login_window.winfo_exists():
                self._login_window.grab_set()
                self._login_window.lift()
            return
        self._cloud_session = session
        if self._login_window is not None and self._login_window.winfo_exists():
            self._login_window.grab_release()
            self._login_window.destroy()
        self._login_window = None
        self._build_shell()
        self._show_view("dashboard")
        self.deiconify()
        self.lift()
        self.focus_force()

    def _show_login_loading(self) -> None:
        if self._login_loading_window is not None and self._login_loading_window.winfo_exists():
            return
        parent = self._login_window or self
        win = tk.Toplevel(parent)
        self._login_loading_window = win
        win.title("Validando usuario")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.transient(parent)
        win.protocol("WM_DELETE_WINDOW", lambda: None)

        card = tk.Frame(win, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=18, pady=18)
        tk.Label(card, text="Validando usuario", bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold")).pack(
            anchor=tk.W,
            padx=20,
            pady=(18, 6),
        )
        tk.Label(
            card,
            text="Comprobando credenciales, rol y dispositivo en Supabase.",
            bg=CARD,
            fg=MUTED,
            font=("Segoe UI", 10),
            justify=tk.LEFT,
        ).pack(anchor=tk.W, padx=20, pady=(0, 16))
        progress = ttk.Progressbar(card, mode="indeterminate")
        progress.pack(fill=tk.X, padx=20, pady=(0, 20))
        progress.start(12)
        win.update_idletasks()
        center_window(win, max(420, win.winfo_reqwidth() + 36), max(190, win.winfo_reqheight() + 36))
        win.grab_set()
        win.lift()

    def _hide_login_loading(self) -> None:
        win = self._login_loading_window
        self._login_loading_window = None
        try:
            if win is not None and win.winfo_exists():
                win.grab_release()
                win.destroy()
        except Exception:
            pass

    def _show_view(self, key: str) -> None:
        self._current_key = key
        for item_key, button in self._nav_buttons.items():
            if item_key == key:
                button.configure(bg=INDIGO, fg="white", activebackground=INDIGO, activeforeground="white")
            else:
                button.configure(bg=SIDEBAR, fg="#475569", activebackground=SOFT, activeforeground=TEXT)

        if self._content is None:
            return
        for child in self._content.winfo_children():
            child.destroy()

        builders = {
            "dashboard": self._build_dashboard,
            "inventario": self._build_inventory,
            "precios": self._build_prices,
            "calcular": self._build_order_calc,
            "woocommerce": self._build_woocommerce,
            "precios_proveedor": self._build_supplier_prices,
            "informes": self._build_reports,
            "configuracion": self._build_settings,
            "seguridad": self._build_security,
        }
        builders.get(key, self._build_dashboard)(self._content)

    def _page_header(self, parent: tk.Frame, tag: str, title: str, subtitle: str, actions: list[str] | None = None) -> None:
        header = tk.Frame(parent, bg=BG)
        header.pack(fill=tk.X, pady=(0, 18))
        header.columnconfigure(0, weight=1)
        left = tk.Frame(header, bg=BG)
        left.grid(row=0, column=0, sticky="ew")
        tk.Label(left, text=tag, bg=INDIGO_SOFT, fg=INDIGO, font=("Segoe UI", 9, "bold"), padx=10, pady=4).pack(anchor=tk.W)
        tk.Label(left, text=title, bg=BG, fg=TEXT, font=("Segoe UI", 24, "bold")).pack(anchor=tk.W, pady=(8, 2))
        tk.Label(left, text=subtitle, bg=BG, fg=MUTED, font=("Segoe UI", 10)).pack(anchor=tk.W)
        if actions:
            right = tk.Frame(header, bg=BG)
            right.grid(row=0, column=1, sticky="e")
            for action in actions:
                command = self._page_header_action_command(action)
                self._button(right, action, primary=action == actions[-1], command=command).pack(side=tk.LEFT, padx=(8, 0))

    def _page_header_action_command(self, action: str) -> object | None:
        if action == "Nueva propuesta":
            return self._start_new_price_proposal
        return None

    def _start_new_price_proposal(self) -> None:
        self._selected_price_proposal = None
        self._proposal_source_item = None
        self._price_reset_edit_state()
        self._price_mode = "edit"
        self._show_view("precios")

    def _build_dashboard(self, parent: tk.Frame) -> None:
        self._page_header(
            parent,
            "Principal",
            "Dashboard",
            "Cabina de mando: atención diaria, actividad reciente y estado del sistema.",
        )

        data = self._dashboard_collect_data()
        kpis = data.get("kpis", {})
        errors = data.get("errors", [])
        if errors:
            tk.Label(
                parent,
                text=" · ".join(errors[:3]),
                bg=AMBER_SOFT,
                fg=AMBER,
                anchor=tk.W,
                justify=tk.LEFT,
                padx=12,
                pady=9,
                wraplength=980,
            ).pack(fill=tk.X, pady=(0, 14))

        metrics = tk.Frame(parent, bg=BG)
        metrics.pack(fill=tk.X, pady=(0, 18))
        metric_data = [
            ("Pedidos abiertos", str(kpis.get("open_orders", 0)), "Info", "Pedidos abiertos", data.get("open_order_items", []), "calcular"),
            ("En validación", str(kpis.get("validation_orders", 0)), "Warning" if kpis.get("validation_orders", 0) else "OK", "Pedidos en validación", data.get("validation_order_items", []), "calcular"),
            ("Recepciones parciales", str(kpis.get("partial_receipts", 0)), "Warning" if kpis.get("partial_receipts", 0) else "OK", "Recepciones parciales", data.get("partial_receipt_items", []), "calcular"),
            ("Propuestas pendientes", str(kpis.get("pending_proposals", 0)), "Warning" if kpis.get("pending_proposals", 0) else "OK", "Propuestas pendientes", data.get("proposal_items", []), "precios"),
            ("Errores hoy", str(kpis.get("errors_today", 0)), "Error" if kpis.get("errors_today", 0) else "OK", "Errores hoy", data.get("error_items", []), "seguridad"),
        ]
        for i, (label, value, status, title, items, target) in enumerate(metric_data):
            metrics.columnconfigure(i, weight=1)
            self._metric(
                metrics,
                label,
                value,
                status,
                command=lambda title=title, items=items, target=target: self._dashboard_show_attention(title, items, target),
            ).grid(row=0, column=i, sticky="ew", padx=(0 if i == 0 else 8, 0))

        # Dashboard responsive:
        # el admin puede tener mucha actividad reciente; por eso el cuerpo va
        # dentro de un canvas con scroll vertical. Así ninguna tarjeta queda
        # "fuera del mapa" cuando los logs crecen.
        viewport = tk.Frame(parent, bg=BG)
        viewport.pack(fill=tk.BOTH, expand=True)
        viewport.columnconfigure(0, weight=1)
        viewport.rowconfigure(0, weight=1)

        canvas = tk.Canvas(viewport, bg=BG, highlightthickness=0)
        yscroll = ttk.Scrollbar(viewport, orient=tk.VERTICAL, command=canvas.yview)
        body = tk.Frame(canvas, bg=BG)
        body.columnconfigure(0, weight=3, minsize=580)
        body.columnconfigure(1, weight=2, minsize=390)

        body_window = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=yscroll.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")

        def _sync_dashboard_scroll(_event: object | None = None) -> None:
            canvas.configure(scrollregion=canvas.bbox("all"))
            try:
                canvas.itemconfigure(body_window, width=canvas.winfo_width())
            except Exception:
                pass

        body.bind("<Configure>", _sync_dashboard_scroll)
        canvas.bind("<Configure>", _sync_dashboard_scroll)
        canvas.bind_all("<MouseWheel>", lambda event: canvas.yview_scroll(int(-1 * (event.delta / 120)), "units") if self._current_key == "dashboard" else None)

        activity = self._dashboard_activity_card(body, data.get("activity", []))
        activity.grid(row=0, column=0, sticky="nsew", padx=(0, 14), pady=(0, 14))

        alerts = self._dashboard_attention_card(
            body,
            "Bloques de atención",
            [
                ("Pedidos que necesitan revisión", data.get("validation_order_items", []), "calcular", "Warning"),
                ("Propuestas pendientes", data.get("proposal_items", []), "precios", "Warning"),
                ("Últimos errores", data.get("error_items", []), "seguridad", "Error"),
            ],
        )
        alerts.grid(row=0, column=1, sticky="nsew", pady=(0, 14))

        orders = self._dashboard_compact_list_card(body, "Pedidos recientes", data.get("recent_order_items", []), "calcular")
        orders.grid(row=1, column=0, sticky="nsew", padx=(0, 14), pady=(0, 8))

        systems = self._dashboard_system_card(body, data)
        systems.grid(row=1, column=1, sticky="nsew", pady=(0, 8))

    def _dashboard_collect_data(self) -> dict[str, Any]:
        data: dict[str, Any] = {
            "kpis": {
                "open_orders": 0,
                "validation_orders": 0,
                "partial_receipts": 0,
                "pending_proposals": 0,
                "errors_today": 0,
            },
            "open_order_items": [],
            "validation_order_items": [],
            "partial_receipt_items": [],
            "recent_order_items": [],
            "proposal_items": [],
            "error_items": [],
            "activity": [],
            "systems": [],
            "errors": [],
        }
        if self._cloud_session is None:
            data["systems"] = [
                ("Supabase", "Sin sesión activa", "Warning"),
                ("WooCommerce", "Pendiente de validar desde módulo Woo", "Info"),
                ("Seguridad", "Sin logs hasta iniciar sesión", "Info"),
            ]
            return data

        try:
            orders_rows = list_cloud_supplier_orders(self._cloud_session, limit=80)
            open_orders = []
            validation_orders = []
            partial_orders = []
            recent_orders = []
            for row in orders_rows:
                status = str(row.get("status") or "").strip().lower()
                name = cloud_order_display_name(row)
                provider = str(row.get("provider") or "-")
                updated = self._format_datetime_short(row.get("updated_at") or row.get("created_at"))
                detail = f"{provider} · {str(row.get('status') or '-')} · {updated}"
                item = (name, detail, "Warning" if "valid" in status or "parcial" in status else "Info")
                if status not in {"recibido completo", "received_full", "cancelled", "canceled", "cancelado"}:
                    open_orders.append(item)
                if "valid" in status:
                    validation_orders.append((name, detail, "Warning"))
                if "parcial" in status:
                    partial_orders.append((name, detail, "Warning"))
                recent_orders.append(item)
            data["open_order_items"] = open_orders[:6]
            data["validation_order_items"] = validation_orders[:6]
            data["partial_receipt_items"] = partial_orders[:6]
            data["recent_order_items"] = recent_orders[:8]
            data["kpis"]["open_orders"] = len(open_orders)
            data["kpis"]["validation_orders"] = len(validation_orders)
            data["kpis"]["partial_receipts"] = len(partial_orders)
        except Exception as exc:
            data["errors"].append(f"Pedidos no cargados: {exc}")

        try:
            proposals = list_real_price_proposals(self._cloud_session, status="pending", limit=80)
            proposal_items = []
            for row in proposals:
                name = str(row.get("name") or row.get("proposal_name") or row.get("item_woo_id") or row.get("id") or "Propuesta")
                created = self._format_datetime_short(row.get("created_at"))
                detail = f"{row.get('status') or 'pending'} · {created}"
                proposal_items.append((name, detail, "Warning"))
            data["proposal_items"] = proposal_items[:6]
            data["kpis"]["pending_proposals"] = len(proposal_items)
        except Exception as exc:
            data["errors"].append(f"Propuestas no cargadas: {exc}")

        try:
            logs = list_security_audit_logs(self._cloud_session, filters={}, limit=80)
            activity = []
            error_items = []
            errors_today = 0
            from datetime import datetime, timezone
            today = datetime.now(timezone.utc).date()
            for row in logs:
                module = row.get("visual_module") or row.get("module") or "-"
                action = row.get("visual_action") or row.get("action") or "-"
                user = row.get("user_email") or "-"
                time = self._format_datetime_short(row.get("created_at"))
                status_raw = str(row.get("status") or row.get("severity") or "INFO").upper()
                status = self._normalize_security_level(status_raw)
                activity.append((f"{time} · {user}", f"{module} · {action}", status))
                is_error = status_raw in {"ERROR", "CRITICAL", "BLOCKED"} or str(row.get("severity") or "").upper() in {"ERROR", "CRITICAL"}
                if is_error:
                    error_items.append((f"{module} · {action}", str(row.get("message") or row.get("error_detail") or row.get("operation_id") or "-")[:120], "Error"))
                    try:
                        dt = datetime.fromisoformat(str(row.get("created_at")).replace("Z", "+00:00"))
                        if dt.date() == today:
                            errors_today += 1
                    except Exception:
                        pass
            data["activity"] = activity[:10]
            data["error_items"] = error_items[:6]
            data["kpis"]["errors_today"] = errors_today
        except Exception as exc:
            data["errors"].append(f"Actividad reciente no cargada: {exc}")

        data["systems"] = [
            ("Supabase", "Sesión activa y lectura operativa", "OK"),
            ("WooCommerce", "Pendiente de sincronización completa v48", "Info"),
            ("Seguridad", "Logs y rollback activos", "OK"),
        ]
        return data

    def _dashboard_show_attention(self, title: str, items: list[tuple[str, str, str]], target: str) -> None:
        win = tk.Toplevel(self)
        win.title(title)
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 700, 520)
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)
        tk.Label(win, text=title, bg=BG, fg=TEXT, font=("Segoe UI", 18, "bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(18, 8))
        card = self._card(win)
        card.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 12))
        card.columnconfigure(0, weight=1)
        if not items:
            self._status_row(card, "Sin pendientes", "No hay elementos que requieran atención en este bloque.", "OK").pack(fill=tk.X, padx=16, pady=16)
        for index, (label, detail, status) in enumerate(items[:12]):
            self._status_row(card, label, detail, status).pack(fill=tk.X, padx=16, pady=(16 if index == 0 else 6, 0))
        footer = tk.Frame(win, bg=BG)
        footer.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))
        self._button(footer, "Ir al módulo", primary=True, command=lambda: (win.destroy(), self._show_view(target))).pack(side=tk.RIGHT)
        self._button(footer, "Cerrar", command=win.destroy).pack(side=tk.RIGHT, padx=(0, 8))

    def _dashboard_activity_card(self, parent: tk.Misc, activity: list[tuple[str, str, str]]) -> tk.Frame:
        card = self._card(parent)
        card.columnconfigure(0, weight=1)
        tk.Label(card, text="Actividad reciente", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w", padx=16, pady=(16, 8))
        tk.Label(card, text="Últimas acciones auditadas, en formato reducido.", bg=CARD, fg=MUTED).grid(row=1, column=0, sticky="w", padx=16, pady=(0, 8))
        if not activity:
            self._status_row(card, "Sin actividad visible", "Todavía no hay logs recientes o el rol actual no puede leerlos.", "Info").grid(row=2, column=0, sticky="ew", padx=16, pady=12)
            return card
        for index, (headline, detail, status) in enumerate(activity[:8], start=2):
            self._status_row(card, headline, detail, status).grid(row=index, column=0, sticky="ew", padx=16, pady=(4 if index > 2 else 8, 0))
        return card

    def _dashboard_attention_card(self, parent: tk.Misc, title: str, sections: list[tuple[str, list[tuple[str, str, str]], str, str]]) -> tk.Frame:
        card = self._card(parent)
        card.columnconfigure(0, weight=1)
        tk.Label(card, text=title, bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w", padx=16, pady=(16, 10))
        row = 1
        for section_title, items, target, status in sections:
            frame = tk.Frame(card, bg=CARD)
            frame.grid(row=row, column=0, sticky="ew", padx=16, pady=(0, 10))
            frame.columnconfigure(0, weight=1)
            tk.Label(frame, text=section_title, bg=CARD, fg=TEXT, font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w")
            self._button(frame, "Ver", command=lambda t=target: self._show_view(t)).grid(row=0, column=1, sticky="e")
            if not items:
                tk.Label(frame, text="Sin elementos pendientes.", bg=GREEN_SOFT, fg=GREEN, anchor=tk.W, padx=10, pady=7).grid(row=1, column=0, columnspan=2, sticky="ew", pady=(6, 0))
            else:
                for idx, (label, detail, item_status) in enumerate(items[:2], start=1):
                    self._status_row(frame, label, detail, item_status or status).grid(row=idx, column=0, columnspan=2, sticky="ew", pady=(6, 0))
            row += 1
        return card

    def _dashboard_compact_list_card(self, parent: tk.Misc, title: str, items: list[tuple[str, str, str]], target: str) -> tk.Frame:
        card = self._card(parent)
        card.columnconfigure(0, weight=1)
        head = tk.Frame(card, bg=CARD)
        head.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 8))
        head.columnconfigure(0, weight=1)
        tk.Label(head, text=title, bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w")
        self._button(head, "Ver módulo", command=lambda: self._show_view(target)).grid(row=0, column=1, sticky="e")
        if not items:
            self._status_row(card, "Sin datos", "No hay elementos recientes para este bloque.", "Info").grid(row=1, column=0, sticky="ew", padx=16, pady=10)
            return card
        for index, (label, detail, status) in enumerate(items[:5], start=1):
            self._status_row(card, label, detail, status).grid(row=index, column=0, sticky="ew", padx=16, pady=(5, 0))
        return card

    def _dashboard_system_card(self, parent: tk.Misc, data: dict[str, Any]) -> tk.Frame:
        card = self._card(parent)
        card.columnconfigure(0, weight=1)
        tk.Label(card, text="Estado de sistemas", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w", padx=16, pady=(16, 8))
        systems = data.get("systems") or []
        for index, (label, detail, status) in enumerate(systems, start=1):
            self._status_row(card, label, detail, status).grid(row=index, column=0, sticky="ew", padx=16, pady=(6, 0))
        self._status_row(card, "Próximo frente", "WooCommerce completo: lectura, autoclasificación y preview de incidencias.", "Info").grid(row=len(systems) + 1, column=0, sticky="ew", padx=16, pady=(12, 16))
        return card


    def _build_inventory(self, parent: tk.Frame) -> None:
        if self._cloud_session is not None and not self._inventory_loaded_once and not self._inventory_loading:
            self.after(80, lambda: self._refresh_inventory(parent, self._inventory_query, allow_empty=True))

        top = tk.Frame(parent, bg=BG)
        top.pack(fill=tk.X, pady=(0, 16))
        top.columnconfigure(0, weight=1)
        query_var = tk.StringVar()
        search = tk.Entry(
            top,
            textvariable=query_var,
            bg=CARD,
            fg=TEXT,
            insertbackground=TEXT,
            relief=tk.FLAT,
            highlightbackground=LINE,
            highlightcolor=INDIGO,
            highlightthickness=1,
            font=("Segoe UI", 10),
        )
        search.insert(0, "")
        search.grid(row=0, column=0, sticky="ew", ipady=10)
        search.configure()
        self._button(top, "Buscar / recargar", primary=True, command=lambda: self._refresh_inventory(parent, query_var.get(), allow_empty=True)).grid(row=0, column=1, padx=(12, 0), sticky="e")
        self._button(top, "Crear nuevo artículo", primary=True, command=self._open_create_inventory_item_modal).grid(row=0, column=2, padx=(8, 0), sticky="e")
        self._button(top, "Exportacion de inventario", command=self._export_inventory_visible).grid(row=0, column=3, padx=(8, 0), sticky="e")
        self._button(top, "Diagnosticar estados", command=self._open_inventory_status_diagnostics_modal).grid(row=0, column=4, padx=(8, 0), sticky="e")
        search.bind("<Return>", lambda _event: self._refresh_inventory(parent, query_var.get(), allow_empty=True))
        tk.Label(
            parent,
            text=self._inventory_error
            or ("Cargando inventario real..." if self._inventory_loading else "Inventario real Supabase. Busca por ID, nombre, SKU, familia o referencia. WooCommerce no se toca desde esta vista."),
            bg=ROSE_SOFT if self._inventory_error else INDIGO_SOFT,
            fg=ROSE if self._inventory_error else "#4338CA",
            anchor=tk.W,
            justify=tk.LEFT,
            padx=12,
            pady=9,
        ).pack(fill=tk.X, pady=(0, 14))

        body = tk.Frame(parent, bg=BG)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=3)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        table_card = self._card(body)
        table_card.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        table_card.rowconfigure(1, weight=1)
        table_card.columnconfigure(0, weight=1)
        head = tk.Frame(table_card, bg=CARD)
        head.grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 10))
        head.columnconfigure(0, weight=1)
        tk.Label(head, text="Tabla de inventario", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w")
        columns = ["ID", "Tipo", "Nombre", "Contenido pack", "Precio Woo", "Stock", "Estado"]
        tree = ttk.Treeview(table_card, columns=columns, show="headings", height=14)
        widths = {"ID": 130, "Tipo": 92, "Nombre": 250, "Contenido pack": 330, "Precio Woo": 100, "Stock": 80, "Estado": 95}
        for column in columns:
            tree.heading(column, text=column, anchor=tk.CENTER)
            tree.column(column, width=widths[column], anchor=tk.CENTER)
        item_by_iid: dict[str, InventoryItem] = {}
        items = list(self._inventory_items)
        if items and self._selected_inventory_item not in items:
            self._selected_inventory_item = items[0]
        if not items:
            tree.insert("", tk.END, values=("-", "Sin inventario real cargado", "-", "-", "Info"))
        for index, item in enumerate(items):
            content_preview = self._inventory_pack_contents_text(item, multiline=False)
            if len(content_preview) > 72:
                content_preview = content_preview[:71].rstrip() + "…"
            iid = tree.insert(
                "",
                tk.END,
                values=(
                    item.code,
                    self._inventory_item_type_text(item),
                    item.name,
                    content_preview or "-",
                    item.price,
                    item.stock,
                    item.status,
                ),
            )
            item_by_iid[iid] = item
            if item == self._selected_inventory_item:
                tree.selection_set(iid)
                tree.focus(iid)
        tree.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))

        detail_host = tk.Frame(body, bg=BG)
        detail_host.grid(row=0, column=1, sticky="nsew")

        def render_detail(item: InventoryItem) -> None:
            self._selected_inventory_item = item
            self._render_inventory_detail(detail_host, item)

        def on_select(_event: object | None = None) -> None:
            selection = tree.selection()
            if not selection or selection[0] not in item_by_iid:
                return
            render_detail(item_by_iid[selection[0]])

        tree.bind("<<TreeviewSelect>>", on_select)
        self._render_inventory_detail(detail_host, self._selected_inventory_item)

    def _refresh_inventory(self, parent: tk.Frame, query: str, *, allow_empty: bool = False) -> None:
        query = query.strip()
        if not query and not allow_empty:
            self._inventory_error = "Introduce un texto o ID para buscar inventario real en Supabase."
            self._inventory_loading = False
            if self._current_key == "inventario" and parent.winfo_exists():
                self._show_view("inventario")
            return
        if self._cloud_session is None:
            self._inventory_error = "No hay sesion Supabase activa."
            self._inventory_loading = False
            if self._current_key == "inventario" and parent.winfo_exists():
                self._show_view("inventario")
            return
        self._inventory_query = query
        self._inventory_error = ""
        self._inventory_loading = True
        if self._current_key == "inventario" and parent.winfo_exists():
            self._show_view("inventario")

        def worker() -> None:
            try:
                if query:
                    server_rows = search_cloud_inventory_items(self._cloud_session, query, limit=100)
                    if self._inventory_query_is_code_like(query):
                        rows = server_rows
                    else:
                        all_rows = list_cloud_inventory_items(self._cloud_session, limit=500)
                        rows = self._merge_inventory_rows([server_rows, self._accent_insensitive_inventory_search(all_rows, query)])
                else:
                    rows = list_cloud_inventory_items(self._cloud_session, limit=150)
                items = [self._inventory_item_from_cloud_row(row) for row in rows]
                empty_msg = "Sin resultados reales visibles." if query and not items else ("No hay inventario real visible en Supabase." if not items else "")
                self.after(0, lambda: self._finish_inventory_refresh(items, empty_msg))
            except Exception as exc:
                self.after(0, lambda exc=exc: self._finish_inventory_refresh([], f"No se pudo leer inventario real Supabase: {exc}"))

        threading.Thread(target=worker, daemon=True).start()

    def _finish_inventory_refresh(self, items: list[InventoryItem], error: str) -> None:
        self._inventory_items = list(items)
        self._inventory_error = error
        self._inventory_loading = False
        self._inventory_loaded_once = True
        self._selected_inventory_item = self._inventory_items[0] if self._inventory_items else None
        if self._current_key == "inventario":
            self._show_view("inventario")

    def _inventory_item_from_cloud_row(self, row: dict[str, Any]) -> InventoryItem:
        store = self._number_or_zero(row.get("store_stock"))
        warehouse = self._number_or_zero(row.get("warehouse_stock"))
        total_stock = store + warehouse
        link_status = str(row.get("woo_link_status") or "").strip()
        item_id = str(row.get("item_id") or row.get("woo_id") or "-")
        name = str(row.get("name") or row.get("woo_name") or f"Item {item_id}")
        dimensions = self._clean_inventory_value(row.get("size"), "Sin definir")
        m3 = self._format_optional_m3(row.get("cubic_meters"))
        materials = self._clean_inventory_value(row.get("materials"), "Sin definir")
        subgroup = self._clean_inventory_value(row.get("subgroup"), "Sin definir")
        woo_price = self._format_optional_price(row.get("woo_price"))
        status, status_reasons = self._inventory_status_analysis_from_row(row)
        woo_id = self._clean_inventory_value(row.get("woo_id"), "-")
        woo_parent_id = self._clean_inventory_value(row.get("woo_parent_id"), "-")
        woo_item_kind = self._clean_inventory_value(row.get("woo_item_kind"), "-")
        woo_link_status = self._clean_inventory_value(link_status, "-")
        sync_woo = f"{woo_item_kind} / {woo_id} / {woo_link_status}"
        store_text = self._format_number(store)
        warehouse_text = self._format_number(warehouse)
        total_text = self._format_number(total_stock)
        supplier_provider = self._clean_inventory_value(row.get("supplier_order_provider"), "-")
        return InventoryItem(
            code=item_id,
            name=name,
            price=woo_price,
            stock=total_text,
            status=status,
            family=self._clean_inventory_value(row.get("family"), "Sin definir"),
            provider=supplier_provider,
            m3=m3,
            sku_woo=self._clean_inventory_value(row.get("woo_sku"), "-"),
            measures=dimensions,
            material=materials,
            sync_woo=sync_woo,
            notes=f"Stock tienda: {store_text} | Stock almacen: {warehouse_text} | Actualizado: {row.get('updated_at') or '-'}",
            subgroup=subgroup,
            store_stock=store_text,
            warehouse_stock=warehouse_text,
            stock_total=total_text,
            woo_id=woo_id,
            woo_parent_id=woo_parent_id,
            woo_name=self._clean_inventory_value(row.get("woo_name"), "-"),
            woo_price=woo_price,
            woo_categories=self._clean_inventory_value(row.get("woo_categories"), "Sin definir"),
            woo_item_kind=woo_item_kind,
            woo_link_status=woo_link_status,
            order_calculated_price=self._format_optional_price(row.get("order_calculated_price")),
            weighted_average_cost=self._format_optional_price(row.get("weighted_average_cost")),
            supplier_order_qty=self._clean_inventory_value(row.get("supplier_order_qty"), "-"),
            supplier_order_provider=supplier_provider,
            status_reasons=status_reasons,
            raw=row,
        )

    def _clean_inventory_value(self, value: Any, fallback: str = "-") -> str:
        text = str(value or "").strip()
        return text if text else fallback

    def _format_optional_price(self, value: Any) -> str:
        if value in (None, ""):
            return "Pendiente"
        try:
            number = float(str(value).replace(",", "."))
        except Exception:
            return str(value)
        return f"{number:.2f} €"

    def _is_missing_inventory_value(self, value: Any) -> bool:
        if value is None:
            return True
        text = str(value).strip()
        return text == "" or text.lower() in {"none", "null", "nan", "sin definir", "pendiente", "-"}

    def _inventory_status_analysis_from_row(self, row: dict[str, Any]) -> tuple[str, tuple[str, ...]]:
        """Calcula el semáforo de inventario usando valores reales recibidos de Supabase.

        Regla base: el precio operativo del item es siempre `woo_price`, porque el precio
        de venta vive en WooCommerce. Los estados Woo desconocidos no deben convertir
        todo el inventario en Error: primero se muestran como Warning/Info para diagnóstico.
        """
        reasons: list[tuple[str, str]] = []

        def add(level: str, message: str) -> None:
            reasons.append((level, message))

        item_id = str(row.get("item_id") or "").strip()
        woo_id = str(row.get("woo_id") or "").strip()
        woo_price = row.get("woo_price")
        link_status_raw = str(row.get("woo_link_status") or "").strip()
        link_status = link_status_raw.lower()
        family = row.get("family")
        subgroup = row.get("subgroup")
        materials = row.get("materials")
        size = row.get("size")
        m3 = row.get("cubic_meters")

        try:
            price_number = float(str(woo_price).replace(",", ".")) if woo_price not in (None, "") else None
        except Exception:
            price_number = None

        accepted_link_statuses = {
            "", "-", "ok", "linked", "matched", "synced", "sync", "connected",
            "active", "ready", "found", "linked_by_sku", "matched_by_sku",
            "linked_variation", "variation", "parent", "simple", "variable",
            "woo_synced", "manual", "manual_link", "auto", "auto_link",
        }
        incomplete_link_statuses = {"unlinked", "local_only", "woo_only", "pending", "pending_link"}
        broken_link_statuses = {"broken", "error", "missing", "not_found", "orphan", "woo_missing", "invalid"}

        if not item_id:
            add("Critical", "Falta item_id en inventory_items.")
        if price_number == 0:
            add("Critical", "Precio Woo igual a 0. Es el precio de venta de la web.")
        elif price_number is None:
            add("Warning", "Precio Woo pendiente o no numérico.")

        if link_status in broken_link_statuses:
            add("Error", f"Estado vínculo Woo roto: {link_status_raw}.")
        elif link_status in incomplete_link_statuses:
            add("Warning", f"Vínculo Woo incompleto o pendiente: {link_status_raw}.")
        elif link_status not in accepted_link_statuses:
            add("Warning", f"Estado vínculo Woo desconocido para el semáforo: {link_status_raw}.")

        if self._is_missing_inventory_value(family):
            add("Warning", "Familia sin definir.")
        if self._is_missing_inventory_value(m3):
            add("Warning", "M3 / cubic_meters pendiente.")
        if self._is_missing_inventory_value(materials):
            add("Info", "Materiales sin definir.")
        if self._is_missing_inventory_value(subgroup):
            add("Info", "Subgrupo sin definir.")
        if self._is_missing_inventory_value(size):
            add("Info", "Medidas sin definir.")
        if not woo_id:
            add("Info", "Woo ID no recibido o no vinculado.")

        levels = [level for level, _message in reasons]
        if "Critical" in levels:
            status = "Critical"
        elif "Error" in levels:
            status = "Error"
        elif "Warning" in levels:
            status = "Warning"
        elif "Info" in levels:
            status = "Info"
        else:
            status = "OK"

        readable_reasons = tuple(message for _level, message in reasons) or ("Sin incidencias detectadas con las reglas actuales.",)
        return status, readable_reasons

    def _inventory_status_from_row(self, row: dict[str, Any]) -> str:
        status, _reasons = self._inventory_status_analysis_from_row(row)
        return status

    def _inventory_status_diagnostics(self) -> dict[str, Any]:
        items = list(self._inventory_items)
        status_counts: dict[str, int] = {}
        link_counts: dict[str, int] = {}
        missing_counts: dict[str, int] = {
            "family": 0,
            "subgroup": 0,
            "materials": 0,
            "size": 0,
            "cubic_meters": 0,
            "woo_price": 0,
            "woo_id": 0,
        }
        reason_counts: dict[str, int] = {}

        for item in items:
            status_counts[item.status] = status_counts.get(item.status, 0) + 1
            raw = item.raw or {}
            link = str(raw.get("woo_link_status") or "-").strip() or "-"
            link_counts[link] = link_counts.get(link, 0) + 1
            for field in missing_counts:
                if self._is_missing_inventory_value(raw.get(field)):
                    missing_counts[field] += 1
            for reason in item.status_reasons:
                reason_counts[reason] = reason_counts.get(reason, 0) + 1

        return {
            "total": len(items),
            "status_counts": status_counts,
            "link_counts": link_counts,
            "missing_counts": missing_counts,
            "reason_counts": reason_counts,
        }

    def _open_inventory_status_diagnostics_modal(self) -> None:
        if not self._inventory_items:
            messagebox.showinfo("Diagnóstico inventario", "No hay items cargados. Pulsa Buscar / recargar para leer Supabase primero.")
            return
        diag = self._inventory_status_diagnostics()
        win = tk.Toplevel(self)
        win.title("Diagnóstico de estados de Inventario")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 980, 680)
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)

        header = tk.Frame(win, bg=BG, highlightbackground=LINE, highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)
        tk.Label(header, text="Diagnóstico de estados", bg=INDIGO_SOFT, fg=INDIGO, font=("Segoe UI", 9, "bold"), padx=10, pady=4).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 4))
        tk.Label(header, text=f"Valores reales recibidos desde Supabase · Items cargados: {diag['total']}", bg=BG, fg=TEXT, font=("Segoe UI", 16, "bold")).grid(row=1, column=0, sticky="w", padx=18, pady=(0, 16))
        self._button(header, "Cerrar", command=win.destroy).grid(row=0, column=1, rowspan=2, padx=18, pady=16, sticky="e")

        body = tk.Frame(win, bg=BG)
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=18)
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        left = self._card(body)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        left.rowconfigure(1, weight=1)
        left.columnconfigure(0, weight=1)
        tk.Label(left, text="Resumen recibido", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w", padx=16, pady=(16, 8))
        left_table = ttk.Treeview(left, columns=("Campo", "Valor", "Cantidad"), show="headings", height=14)
        for column, width in {"Campo": 150, "Valor": 210, "Cantidad": 90}.items():
            left_table.heading(column, text=column, anchor=tk.CENTER)
            left_table.column(column, width=width, anchor=tk.CENTER)
        for status, count in sorted((diag["status_counts"] or {}).items()):
            left_table.insert("", tk.END, values=("Estado UI", status, count))
        for link, count in sorted((diag["link_counts"] or {}).items(), key=lambda kv: (-kv[1], kv[0])):
            left_table.insert("", tk.END, values=("woo_link_status", link, count))
        for field, count in sorted((diag["missing_counts"] or {}).items()):
            left_table.insert("", tk.END, values=(f"{field} vacío", "Sí", count))
        left_table.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 16))

        right = self._card(body)
        right.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        right.rowconfigure(1, weight=1)
        right.columnconfigure(0, weight=1)
        tk.Label(right, text="Motivos que generan estado", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w", padx=16, pady=(16, 8))
        reason_table = ttk.Treeview(right, columns=("Motivo", "Cantidad"), show="headings", height=14)
        reason_table.heading("Motivo", text="Motivo", anchor=tk.CENTER)
        reason_table.heading("Cantidad", text="Cantidad", anchor=tk.CENTER)
        reason_table.column("Motivo", width=350, anchor=tk.CENTER)
        reason_table.column("Cantidad", width=90, anchor=tk.CENTER)
        for reason, count in sorted((diag["reason_counts"] or {}).items(), key=lambda kv: (-kv[1], kv[0])):
            reason_table.insert("", tk.END, values=(reason, count))
        reason_table.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 16))


    def _number_or_zero(self, value: Any) -> float:
        try:
            if value in (None, ""):
                return 0.0
            return float(str(value).replace(",", "."))
        except Exception:
            return 0.0

    def _format_number(self, value: Any) -> str:
        try:
            number = float(value)
        except Exception:
            return str(value)
        if number.is_integer():
            return str(int(number))
        return f"{number:.2f}"

    def _format_optional_m3(self, value: Any) -> str:
        if value in (None, ""):
            return "Pendiente"
        return self._format_number(value)

    def _inventory_detail_rows(self, item: InventoryItem) -> list[tuple[str, str]]:
        stock_total = item.stock_total if item.stock_total != "-" else item.stock
        rows = [
            ("ID", item.code),
            ("Código HUB", self._inventory_pack_parent_code(item) or self._clean_inventory_value((item.raw or {}).get("hub_item_code"), "-")),
            ("Tipo item", self._clean_inventory_value((item.raw or {}).get("item_record_type") or (item.raw or {}).get("hub_search_record_type"), "simple")),
            ("Nombre", item.name),
            ("Precio Woo", item.price),
            ("Stock tienda", item.store_stock),
            ("Stock almacen", item.warehouse_stock),
            ("Stock total", f"{stock_total} unidades"),
            ("Familia", item.family),
            ("Subgrupo", item.subgroup),
            ("Materiales", item.material),
            ("Medidas", item.measures),
            ("M3", item.m3),
            ("Woo ID", item.woo_id),
            ("Woo Parent ID", item.woo_parent_id),
            ("SKU Woo", item.sku_woo),
            ("Nombre Woo", item.woo_name),
            ("Categorias Woo", item.woo_categories),
            ("Tipo Woo", item.woo_item_kind),
            ("Estado vinculo Woo", item.woo_link_status),
            ("Coste calculado pedido", item.order_calculated_price),
            ("Coste medio ponderado", item.weighted_average_cost),
            ("Cantidad en pedido proveedor", item.supplier_order_qty),
            ("Proveedor pedido", item.supplier_order_provider),
            ("Estado", item.status),
            ("Notas internas", item.notes),
        ]
        pack_text = self._inventory_pack_contents_text(item, multiline=False)
        if pack_text:
            rows.insert(3, ("Contenido pack", pack_text))
        return rows

    def _export_inventory_visible(self) -> None:
        items = list(self._inventory_items)
        if not items:
            messagebox.showinfo("Inventario", "No hay productos visibles para exportar.")
            return
        path = filedialog.asksaveasfilename(
            title="Exportar inventario visible",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("Todos los archivos", "*.*")],
            initialfile="inventario_visible.csv",
        )
        if not path:
            return
        headers = [
            "ID",
            "Nombre",
            "Familia",
            "Subgrupo",
            "Materiales",
            "Medidas",
            "M3",
            "Stock tienda",
            "Stock almacen",
            "Stock total",
            "Precio Woo",
            "Woo ID",
            "Woo SKU",
            "Estado vinculo Woo",
            "Coste calculado pedido",
            "Coste medio ponderado",
            "Proveedor pedido",
            "Cantidad pedido",
            "Estado",
        ]
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as fh:
                writer = csv.writer(fh, delimiter=";")
                writer.writerow(headers)
                for item in items:
                    writer.writerow([
                        item.code,
                        item.name,
                        item.family,
                        item.subgroup,
                        item.material,
                        item.measures,
                        item.m3,
                        item.store_stock,
                        item.warehouse_stock,
                        item.stock_total if item.stock_total != "-" else item.stock,
                        item.price,
                        item.woo_id,
                        item.sku_woo,
                        item.woo_link_status,
                        item.order_calculated_price,
                        item.weighted_average_cost,
                        item.supplier_order_provider,
                        item.supplier_order_qty,
                        item.status,
                    ])
        except Exception as exc:
            messagebox.showerror("Inventario", f"No se pudo exportar el inventario visible: {exc}")
            return
        messagebox.showinfo("Inventario", f"Inventario visible exportado:\n{path}")


    def _open_create_inventory_item_modal(self) -> None:
        if self._cloud_session is None:
            messagebox.showwarning("Inventario", "Inicia sesión en Supabase para crear artículos.")
            return

        win = tk.Toplevel(self)
        win.title("Crear nuevo artículo")
        win.configure(bg=BG)
        win.geometry("820x720")
        win.transient(self)
        win.grab_set()

        shell = tk.Frame(win, bg=BG)
        shell.pack(fill=tk.BOTH, expand=True, padx=18, pady=18)
        shell.columnconfigure(0, weight=1)
        shell.rowconfigure(1, weight=1)

        tk.Label(
            shell,
            text="Crear nuevo artículo",
            bg=BG,
            fg=TEXT,
            font=("Segoe UI", 18, "bold"),
        ).grid(row=0, column=0, sticky="w")

        subtitle = tk.Label(
            shell,
            text="Se creará en Supabase inventory_items. No toca WooCommerce ni stock externo.",
            bg=BG,
            fg=MUTED,
            font=("Segoe UI", 10),
        )
        subtitle.grid(row=0, column=0, sticky="w", pady=(34, 0))

        card = self._card(shell)
        card.grid(row=1, column=0, sticky="nsew", pady=(18, 12))
        card.columnconfigure(0, weight=1)
        card.rowconfigure(0, weight=1)

        canvas = tk.Canvas(card, bg=CARD, highlightthickness=0)
        scroll = ttk.Scrollbar(card, orient=tk.VERTICAL, command=canvas.yview)
        form = tk.Frame(canvas, bg=CARD)
        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)
        canvas.create_window((0, 0), window=form, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")

        def _sync_scroll(_event: object | None = None) -> None:
            canvas.configure(scrollregion=canvas.bbox("all"))

        form.bind("<Configure>", _sync_scroll)

        fields: dict[str, tk.Entry | tk.Text] = {}
        defaults = {
            "commercial_status": "Normal",
            "packages": "1",
            "store_stock": "0",
            "warehouse_stock": "0",
        }

        def add_entry(row: int, col: int, key: str, label: str, *, width: int = 22) -> None:
            tk.Label(form, text=label, bg=CARD, fg=TEXT, font=("Segoe UI", 9, "bold")).grid(
                row=row, column=col, sticky="w", padx=(18 if col == 0 else 12, 8), pady=(12, 4)
            )
            entry = tk.Entry(
                form,
                bg="white",
                fg=TEXT,
                relief=tk.FLAT,
                highlightbackground=LINE,
                highlightthickness=1,
                font=("Segoe UI", 10),
                width=width,
            )
            entry.insert(0, defaults.get(key, ""))
            entry.grid(row=row, column=col + 1, sticky="ew", padx=(0, 18 if col == 2 else 8), pady=(12, 4), ipady=7)
            fields[key] = entry

        add_entry(0, 0, "item_id", "ID / Referencia *")
        add_entry(0, 2, "heca_reference", "HECA reference")
        add_entry(1, 0, "name", "Nombre *")
        add_entry(1, 2, "commercial_status", "Estado comercial")
        add_entry(2, 0, "family", "Familia")
        add_entry(2, 2, "subgroup", "Subgrupo")
        add_entry(3, 0, "size", "Medida")
        add_entry(3, 2, "materials", "Materiales")
        add_entry(4, 0, "cubic_meters", "M3 unidad")
        add_entry(4, 2, "rotation_c", "Rotación C")
        add_entry(5, 0, "packages", "Bultos")
        add_entry(5, 2, "primary_supplier_price", "Precio proveedor")
        add_entry(6, 0, "pascal_price", "Precio Pascal")
        add_entry(6, 2, "woo_sku", "Woo SKU")
        add_entry(7, 0, "store_stock", "Stock tienda")
        add_entry(7, 2, "warehouse_stock", "Stock almacén")

        tk.Label(form, text="Notas", bg=CARD, fg=TEXT, font=("Segoe UI", 9, "bold")).grid(
            row=8, column=0, sticky="nw", padx=(18, 8), pady=(12, 8)
        )
        notes = tk.Text(form, height=5, bg="white", fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightthickness=1)
        notes.grid(row=8, column=1, columnspan=3, sticky="ew", padx=(0, 18), pady=(12, 8))
        fields["notes"] = notes

        help_box = tk.Label(
            form,
            text="Campos obligatorios: ID y Nombre. Si HECA reference queda vacío, se usará el ID con ceros a la izquierda. Bultos por defecto = 1.",
            bg=INDIGO_SOFT,
            fg="#3730A3",
            wraplength=720,
            justify=tk.LEFT,
            padx=12,
            pady=10,
        )
        help_box.grid(row=9, column=0, columnspan=4, sticky="ew", padx=18, pady=(8, 16))

        def get_payload() -> dict[str, Any]:
            data: dict[str, Any] = {}
            for key, widget in fields.items():
                if isinstance(widget, tk.Text):
                    data[key] = widget.get("1.0", tk.END).strip()
                else:
                    data[key] = widget.get().strip()
            if not data.get("heca_reference") and data.get("item_id"):
                try:
                    data["heca_reference"] = str(int(str(data["item_id"]).strip())).zfill(7)
                except Exception:
                    pass
            return data

        footer = tk.Frame(shell, bg=BG)
        footer.grid(row=2, column=0, sticky="ew")
        footer.columnconfigure(0, weight=1)

        status_var = tk.StringVar(value="Listo para crear.")
        tk.Label(footer, textvariable=status_var, bg=BG, fg=MUTED, anchor=tk.W).grid(row=0, column=0, sticky="ew")

        def preview() -> None:
            try:
                result = preview_create_cloud_inventory_item(self._cloud_session, get_payload())
            except Exception as exc:
                messagebox.showerror("Validación", str(exc))
                return
            payload = result.get("payload") or {}
            if result.get("exists"):
                existing = result.get("existing") or {}
                messagebox.showwarning(
                    "Artículo existente",
                    f"Ya existe el item {existing.get('item_id')}:\n{existing.get('name') or '-'}\n\nNo se creará duplicado.",
                )
                return
            lines = [
                "PREVIEW NUEVO ARTÍCULO",
                "",
                f"ID: {payload.get('item_id')}",
                f"Nombre: {payload.get('name')}",
                f"Familia: {payload.get('family') or '-'}",
                f"Subgrupo: {payload.get('subgroup') or '-'}",
                f"Medida: {payload.get('size') or '-'}",
                f"M3: {payload.get('cubic_meters') or '-'}",
                f"Rotación C: {payload.get('rotation_c') or '-'}",
                f"Bultos: {payload.get('packages') or '-'}",
                f"Precio proveedor: {payload.get('primary_supplier_price') or '-'}",
                f"Precio Pascal: {payload.get('pascal_price') or '-'}",
                "",
                "No toca WooCommerce. No toca stock externo.",
            ]
            messagebox.showinfo("Preview", "\n".join(lines))

        def save() -> None:
            try:
                result = preview_create_cloud_inventory_item(self._cloud_session, get_payload())
            except Exception as exc:
                messagebox.showerror("Validación", str(exc))
                return
            if result.get("exists"):
                existing = result.get("existing") or {}
                messagebox.showwarning(
                    "Artículo existente",
                    f"Ya existe el item {existing.get('item_id')}:\n{existing.get('name') or '-'}",
                )
                return
            payload = result.get("payload") or {}
            confirm_text = (
                f"Se creará el artículo {payload.get('item_id')}:\n"
                f"{payload.get('name')}\n\n"
                "Se guardará en Supabase con log y snapshot.\n"
                "No se tocará WooCommerce.\n\n¿Continuar?"
            )
            if not messagebox.askyesno("Crear artículo", confirm_text):
                return
            try:
                created = create_cloud_inventory_item(self._cloud_session, get_payload())
            except Exception as exc:
                messagebox.showerror("Crear artículo", f"No se pudo crear el artículo.\n\n{exc}")
                return
            messagebox.showinfo("Artículo creado", f"Artículo creado correctamente.\nOperation ID: {created.get('operation_id')}")
            win.destroy()
            self._inventory_loaded_once = False
            self._refresh_inventory(self._content, str(payload.get("item_id") or ""), allow_empty=True)

        self._button(footer, "Cancelar", command=win.destroy).grid(row=0, column=1, padx=(8, 0), sticky="e")
        self._button(footer, "Preview", command=preview).grid(row=0, column=2, padx=(8, 0), sticky="e")
        self._button(footer, "Crear artículo", primary=True, command=save).grid(row=0, column=3, padx=(8, 0), sticky="e")

    def _inventory_pack_parent_code(self, item: InventoryItem | None) -> str:
        raw = item.raw if item and item.raw else {}
        for key in ("hub_item_code", "hub_search_code", "heca_reference"):
            value = str(raw.get(key) or "").strip()
            if value:
                return value
        if item and item.woo_id not in {"", "-"} and (str(raw.get("item_record_type") or raw.get("hub_search_record_type") or "") in {"woo_pack", "manual_pack"} or item.name.lower().startswith("pack woo")):
            return f"WOO-PACK-{item.woo_id}"
        return ""

    def _inventory_pack_contents_text(self, item: InventoryItem | None, *, multiline: bool = False) -> str:
        if item is None:
            return ""
        raw = item.raw or {}
        key = "hub_pack_components_multiline" if multiline else "hub_pack_components_text"
        text = str(raw.get(key) or "").strip()
        if text:
            return text
        sku = str(raw.get("woo_sku") or item.sku_woo or "").strip()
        if "|" not in sku:
            return ""
        counts: dict[str, int] = {}
        order: list[str] = []
        for part in sku.split("|"):
            token = part.strip()
            if not token:
                continue
            if token not in counts:
                counts[token] = 0
                order.append(token)
            counts[token] += 1
        if multiline:
            return "\n".join(f"- {token} x{counts[token]}" for token in order)
        return "; ".join(f"{token} x{counts[token]}" for token in order)

    def _inventory_item_type_text(self, item: InventoryItem) -> str:
        raw = item.raw or {}
        record_type = str(raw.get("item_record_type") or raw.get("hub_search_record_type") or "simple").strip() or "simple"
        labels = {
            "simple": "Simple",
            "woo_pack": "Pack Woo",
            "manual_pack": "Pack manual",
            "alias": "Alias",
        }
        return labels.get(record_type, record_type)

    def _inventory_query_is_code_like(self, query: str) -> bool:
        text = (query or "").strip()
        if not text:
            return False
        # Para códigos/SKU usamos búsqueda exacta/rankeada. Evita que un código como
        # 0201001 active búsquedas contiene sobre todo el inventario y traiga ruido.
        return bool(re.fullmatch(r"[A-Za-z0-9_\-|]+", text))

    def _inventory_pack_component_row(self, parent: tk.Misc, quantity: str, code: str, name: str) -> tk.Frame:
        frame = tk.Frame(parent, bg=SOFT, highlightbackground=LINE, highlightthickness=1)
        qty_box = tk.Frame(frame, bg=INDIGO_SOFT, highlightbackground=LINE, highlightthickness=1)
        qty_box.pack(side=tk.LEFT, padx=(10, 8), pady=8)
        tk.Label(qty_box, text=quantity, bg=INDIGO_SOFT, fg=INDIGO, font=("Segoe UI", 9, "bold"), padx=8, pady=4).pack()

        body = tk.Frame(frame, bg=SOFT)
        body.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10), pady=8)
        tk.Label(body, text=code or "-", bg=SOFT, fg=TEXT, font=("Segoe UI", 9, "bold"), anchor=tk.W).pack(anchor=tk.W)
        tk.Label(body, text=name or "No encontrado en inventario", bg=SOFT, fg=MUTED, font=("Segoe UI", 9), anchor=tk.W, justify=tk.LEFT, wraplength=240).pack(anchor=tk.W, fill=tk.X)
        return frame

    def _inventory_pack_fallback_components(self, item: InventoryItem) -> list[dict[str, str]]:
        sku = str((item.raw or {}).get("woo_sku") or item.sku_woo or "").strip()
        if "|" not in sku:
            return []
        counts: dict[str, int] = {}
        order: list[str] = []
        for part in sku.split("|"):
            token = part.strip()
            if not token:
                continue
            if token not in counts:
                counts[token] = 0
                order.append(token)
            counts[token] += 1
        return [{"component_item_code": token, "component_name": "", "quantity": str(counts[token])} for token in order]

    def _render_inventory_pack_inline_box(self, parent: tk.Misc, item: InventoryItem, *, compact: bool = False) -> None:
        parent_code = self._inventory_pack_parent_code(item)
        initial = self._inventory_pack_contents_text(item, multiline=True)
        if not parent_code and not initial:
            return

        title = "Contenido del pack" if parent_code else "Relación"
        tk.Label(parent, text=title, bg=CARD, fg=TEXT, font=("Segoe UI", 11, "bold")).pack(anchor=tk.W, pady=(12, 4))

        box = tk.Frame(parent, bg=CARD)
        box.pack(fill=tk.X, pady=(0, 8))

        def clear_box() -> None:
            for child in box.winfo_children():
                child.destroy()

        def render_components(components: list[dict[str, object]], note: str = "") -> None:
            if not box.winfo_exists():
                return
            clear_box()
            if note:
                tk.Label(box, text=note, bg=CARD, fg=MUTED, font=("Segoe UI", 8), wraplength=320, justify=tk.LEFT).pack(anchor=tk.W, pady=(0, 6))
            if not components:
                self._detail_row(box, "Estado", "Sin componentes registrados").pack(fill=tk.X, pady=4)
                return
            max_rows = 4 if compact else len(components)
            visible = components[:max_rows]
            for comp in visible:
                qty_raw = str(comp.get("quantity") or "1").strip()
                try:
                    qty_num = float(qty_raw.replace(",", "."))
                    qty_text = f"{int(qty_num)}x" if qty_num.is_integer() else f"{qty_num:g}x"
                except Exception:
                    qty_text = f"{qty_raw}x"
                row = self._inventory_pack_component_row(
                    box,
                    qty_text,
                    str(comp.get("component_item_code") or "").strip(),
                    str(comp.get("component_name") or "").strip(),
                )
                row.pack(fill=tk.X, pady=4)
            hidden = len(components) - len(visible)
            if hidden > 0:
                tk.Label(box, text=f"+ {hidden} componente(s) más en el detalle completo.", bg=CARD, fg=MUTED, font=("Segoe UI", 8)).pack(anchor=tk.W, pady=(4, 0))

        fallback_components = self._inventory_pack_fallback_components(item)
        if fallback_components:
            render_components(fallback_components, note="Cargando nombres y contenido completo del pack...")
        else:
            self._detail_row(box, "Estado", "Cargando contenido del pack...").pack(fill=tk.X, pady=4)

        if not parent_code:
            return

        def worker() -> None:
            try:
                if self._cloud_session is None:
                    raise RuntimeError("No hay sesión Supabase activa.")
                result = fetch_inventory_pack_components(self._cloud_session, parent_code, item.sku_woo)
                components = list(result.get("components") or [])
                source = str(result.get("source") or "")
                if components:
                    note = "" if source == "inventory_item_components" else f"Fuente: {source}"
                    self.after(0, lambda comps=components, note=note: render_components(comps, note=note))
                    return
                fallback = self._inventory_pack_fallback_components(item)
                if fallback:
                    note = "Contenido cargado desde Woo SKU. Faltan nombres en inventory_item_components."
                    self.after(0, lambda comps=fallback, note=note: render_components(comps, note=note))
                    return
                self.after(0, lambda: render_components([], note="Sin componentes registrados para este pack."))
            except Exception as exc:
                fallback = self._inventory_pack_fallback_components(item)
                note = f"No se pudo consultar inventory_item_components: {exc}"
                self.after(0, lambda comps=fallback, note=note: render_components(comps, note=note))

        threading.Thread(target=worker, daemon=True).start()

    def _open_inventory_pack_contents_popup(self, item: InventoryItem) -> None:
        parent_code = self._inventory_pack_parent_code(item)
        if not parent_code:
            messagebox.showinfo("Contenido pack", "Este item no tiene código de pack asociado.")
            return
        win = tk.Toplevel(self)
        win.title(f"Contenido pack - {parent_code}")
        win.configure(bg=BG)
        win.transient(self)
        center_window(win, 760, 520)
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)
        header = tk.Frame(win, bg=BG, highlightbackground=LINE, highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)
        tk.Label(header, text="Contenido del pack", bg=INDIGO_SOFT, fg=INDIGO, font=("Segoe UI", 9, "bold"), padx=10, pady=4).grid(row=0, column=0, sticky="w", padx=18, pady=(14, 4))
        tk.Label(header, text=f"{parent_code} · {item.name}", bg=BG, fg=TEXT, font=("Segoe UI", 16, "bold"), wraplength=620, justify=tk.LEFT).grid(row=1, column=0, sticky="w", padx=18, pady=(0, 14))
        self._button(header, "Cerrar", command=win.destroy).grid(row=0, column=1, rowspan=2, sticky="e", padx=18, pady=14)
        box = tk.Text(win, bg="#0F172A", fg="#E2E8F0", insertbackground="#E2E8F0", relief=tk.FLAT, wrap=tk.WORD, font=("Consolas", 10))
        box.grid(row=1, column=0, sticky="nsew", padx=18, pady=18)
        box.insert("1.0", "Cargando contenido del pack...")
        box.configure(state=tk.DISABLED)

        def render(text: str) -> None:
            if not box.winfo_exists():
                return
            box.configure(state=tk.NORMAL)
            box.delete("1.0", tk.END)
            box.insert("1.0", text)
            box.configure(state=tk.DISABLED)

        def worker() -> None:
            try:
                if self._cloud_session is None:
                    raise RuntimeError("No hay sesión Supabase activa.")
                result = fetch_inventory_pack_components(self._cloud_session, parent_code, item.sku_woo)
                source = result.get("source") or "-"
                content = result.get("multiline") or self._inventory_pack_contents_text(item, multiline=True)
                if not content:
                    content = "Sin componentes registrados para este pack."
                text = (
                    f"Pack: {parent_code}\n"
                    f"Woo ID: {item.woo_id}\n"
                    f"Woo SKU: {item.sku_woo}\n"
                    f"Fuente: {source}\n\n"
                    f"{content}"
                )
            except Exception as exc:
                fallback = self._inventory_pack_contents_text(item, multiline=True)
                text = f"No se pudo consultar inventory_item_components.\n\nERROR: {exc}"
                if fallback:
                    text += f"\n\nFallback desde Woo SKU:\n{fallback}"
            self.after(0, lambda: render(text))

        threading.Thread(target=worker, daemon=True).start()

    def _render_inventory_detail(self, parent: tk.Frame, item: InventoryItem | None) -> None:
        for child in parent.winfo_children():
            child.destroy()
        detail = self._card(parent)
        detail.pack(fill=tk.BOTH, expand=True)
        detail.rowconfigure(0, weight=1)
        detail.columnconfigure(0, weight=1)
        if item is None:
            empty = tk.Frame(detail, bg=CARD)
            empty.grid(row=0, column=0, sticky="nsew", padx=18, pady=18)
            tk.Label(empty, text="Detalles", bg=INDIGO_SOFT, fg=INDIGO, font=("Segoe UI", 9, "bold"), padx=10, pady=4).pack(anchor=tk.W)
            tk.Label(empty, text="Selecciona un item", bg=CARD, fg=TEXT, font=("Segoe UI", 16, "bold")).pack(anchor=tk.W, pady=(12, 4))
            tk.Label(empty, text="El inventario se carga desde Supabase. Si no aparecen items, pulsa Buscar / recargar.", bg=CARD, fg=MUTED, wraplength=300, justify=tk.LEFT).pack(anchor=tk.W)
            return

        scroll_frame = tk.Frame(detail, bg=CARD)
        scroll_frame.grid(row=0, column=0, sticky="nsew", padx=18, pady=(16, 10))
        scroll_frame.rowconfigure(0, weight=1)
        scroll_frame.columnconfigure(0, weight=1)
        canvas = tk.Canvas(scroll_frame, bg=CARD, highlightthickness=0)
        scrollbar = ttk.Scrollbar(scroll_frame, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        scroll_area = tk.Frame(canvas, bg=CARD)
        window_id = canvas.create_window((0, 0), window=scroll_area, anchor="nw")
        scroll_area.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda event: canvas.itemconfigure(window_id, width=event.width))

        tk.Label(scroll_area, text="Detalles", bg=INDIGO_SOFT, fg=INDIGO, font=("Segoe UI", 9, "bold"), padx=10, pady=4).pack(anchor=tk.W)
        tk.Label(scroll_area, text=item.name, bg=CARD, fg=TEXT, font=("Segoe UI", 16, "bold"), wraplength=300, justify=tk.LEFT).pack(
            anchor=tk.W,
            pady=(10, 2),
        )
        tk.Label(scroll_area, text=f"ID: {item.code}", bg=CARD, fg=MUTED, font=("Segoe UI", 9)).pack(anchor=tk.W, pady=(0, 14))

        for label, value in self._inventory_detail_rows(item)[:18]:
            self._detail_row(scroll_area, label, value).pack(fill=tk.X, pady=4)

        self._render_inventory_pack_inline_box(scroll_area, item, compact=True)

        actions = tk.Frame(detail, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        actions.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        detail_pad = (12, 7)
        self._button(actions, "Abrir detalle completo", primary=True, command=lambda: self._open_inventory_detail_window(item)).pack(
            fill=tk.X,
            padx=12,
            pady=detail_pad,
        )
        self._button(actions, "Agregar a Propuesta de precios", command=lambda: self._open_inventory_proposal_modal(item)).pack(
            fill=tk.X,
            padx=12,
            pady=(0, 12),
        )

    def _open_inventory_stock_preview_modal(self, item: InventoryItem) -> None:
        if self._cloud_session is None:
            messagebox.showerror("Inventario", "No hay sesion Supabase activa.")
            return
        if not item.raw:
            messagebox.showinfo("Inventario", "Este item es visual/mock. Busca un item real antes de generar preview.")
            return
        win = tk.Toplevel(self)
        win.title("Preview cambio stock interno")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 760, 560)
        win.columnconfigure(0, weight=1)
        header = tk.Frame(win, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew", padx=18, pady=(18, 0))
        header.columnconfigure(0, weight=1)
        tk.Label(header, text="Preview cambio stock interno", bg=CARD, fg=TEXT, font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 2))
        tk.Label(header, text=f"{item.code} - {item.name}", bg=CARD, fg=MUTED).grid(row=1, column=0, sticky="w", padx=18, pady=(0, 16))
        self._button(header, "Cerrar", command=win.destroy).grid(row=0, column=1, rowspan=2, padx=18, pady=16)

        body = tk.Frame(win, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        body.grid(row=1, column=0, sticky="ew", padx=18, pady=12)
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)
        store_var = tk.StringVar()
        warehouse_var = tk.StringVar()
        notes_var = tk.StringVar(value="Preview desde UI ERP. No aplicar sin confirmacion.")
        self._field(body, "Nuevo stock tienda", store_var).grid(row=0, column=0, sticky="ew", padx=(16, 8), pady=(16, 10))
        self._field(body, "Nuevo stock almacen", warehouse_var).grid(row=0, column=1, sticky="ew", padx=(8, 16), pady=(16, 10))
        self._field(body, "Notas", notes_var).grid(row=1, column=0, columnspan=2, sticky="ew", padx=16, pady=(0, 12))
        result = tk.Text(body, height=12, bg="#0F172A", fg="#E2E8F0", insertbackground="#E2E8F0", relief=tk.FLAT, wrap=tk.WORD, font=("Consolas", 9))
        result.insert("1.0", "Genera preview para validar el cambio. No se escribe en Supabase desde este popup.")
        result.configure(state=tk.DISABLED)
        result.grid(row=2, column=0, columnspan=2, sticky="ew", padx=16, pady=(0, 16))

        def generate_preview() -> None:
            result.configure(state=tk.NORMAL)
            result.delete("1.0", tk.END)
            result.insert("1.0", "Generando preview real...")
            result.configure(state=tk.DISABLED)

            def worker() -> None:
                try:
                    preview = preview_internal_inventory_update(
                        self._cloud_session,
                        int(item.code),
                        store_var.get() or None,
                        warehouse_var.get() or None,
                        notes_var.get(),
                    )
                    text = json.dumps(preview, ensure_ascii=False, indent=2, default=str)
                except Exception as exc:
                    text = f"ERROR: {exc}"
                self.after(0, lambda: render_preview(text))

            threading.Thread(target=worker, daemon=True).start()

        def render_preview(text: str) -> None:
            if not result.winfo_exists():
                return
            result.configure(state=tk.NORMAL)
            result.delete("1.0", tk.END)
            result.insert("1.0", text)
            result.configure(state=tk.DISABLED)

        footer = tk.Frame(win, bg=BG)
        footer.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))
        self._button(footer, "Generar preview", primary=True, command=generate_preview).pack(side=tk.RIGHT)
        self._button(footer, "Cancelar", command=win.destroy).pack(side=tk.RIGHT, padx=(0, 8))

    def _detail_row(self, parent: tk.Misc, label: str, value: str) -> tk.Frame:
        frame = tk.Frame(parent, bg=SOFT, highlightbackground=LINE, highlightthickness=1)
        tk.Label(frame, text=label, bg=SOFT, fg=MUTED, font=("Segoe UI", 9), anchor=tk.W).pack(side=tk.LEFT, padx=12, pady=9)
        tk.Label(
            frame,
            text=value,
            bg=SOFT,
            fg=TEXT,
            font=("Segoe UI", 9, "bold"),
            anchor=tk.E,
            justify=tk.RIGHT,
            wraplength=170,
        ).pack(side=tk.RIGHT, padx=12, pady=9)
        return frame

    def _open_inventory_detail_window(self, item: InventoryItem) -> None:
        win = tk.Toplevel(self)
        win.title(f"Inventario - {item.code}")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 1120, 720)
        win.minsize(980, 620)
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)

        editable_values = self._inventory_editable_initial_values(item)
        edit_vars: dict[str, tk.StringVar] = {field: tk.StringVar(value=value) for field, value in editable_values.items()}
        apply_in_progress = {"value": False}

        def collect_changes() -> dict[str, tuple[str, str]]:
            return self._collect_inventory_detail_changes(editable_values, edit_vars)

        def close_with_guard() -> None:
            if apply_in_progress["value"]:
                return
            changes = collect_changes()
            if not changes:
                win.destroy()
                return
            self._open_inventory_changes_review(
                win,
                item,
                changes,
                on_discard=win.destroy,
                on_applied=lambda: self._after_inventory_item_updated(win),
            )

        win.protocol("WM_DELETE_WINDOW", close_with_guard)

        header = tk.Frame(win, bg=BG, highlightbackground=LINE, highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)
        title = tk.Frame(header, bg=BG)
        title.grid(row=0, column=0, sticky="ew", padx=22, pady=16)
        tk.Label(title, text="Detalle completo - Inventario", bg=INDIGO_SOFT, fg=INDIGO, font=("Segoe UI", 9, "bold"), padx=10, pady=4).pack(anchor=tk.W)
        tk.Label(title, text=item.name, bg=BG, fg=TEXT, font=("Segoe UI", 22, "bold")).pack(anchor=tk.W, pady=(8, 2))
        tk.Label(title, text=f"ID: {item.code} · Precio Woo: {item.price} · Stock: {item.stock}", bg=BG, fg=MUTED).pack(anchor=tk.W)
        self._button(header, "Cerrar", command=close_with_guard).grid(row=0, column=1, padx=22, pady=16, sticky="e")

        body = tk.Frame(win, bg=BG)
        body.grid(row=1, column=0, sticky="nsew", padx=22, pady=22)
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=2)
        body.rowconfigure(0, weight=1)

        left = self._card(body)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        left.rowconfigure(1, weight=1)
        left.columnconfigure(0, weight=1)
        tk.Label(left, text="Detalles editables del item", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w", padx=16, pady=(16, 4))

        details_host = tk.Frame(left, bg=CARD)
        details_host.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 12))
        details_host.rowconfigure(0, weight=1)
        details_host.columnconfigure(0, weight=1)
        details_canvas = tk.Canvas(details_host, bg=CARD, highlightthickness=0)
        details_scrollbar = ttk.Scrollbar(details_host, orient=tk.VERTICAL, command=details_canvas.yview)
        details_canvas.configure(yscrollcommand=details_scrollbar.set)
        details_canvas.grid(row=0, column=0, sticky="nsew")
        details_scrollbar.grid(row=0, column=1, sticky="ns")
        details_content = tk.Frame(details_canvas, bg=CARD)
        details_window_id = details_canvas.create_window((0, 0), window=details_content, anchor="nw")
        details_content.bind("<Configure>", lambda _event: details_canvas.configure(scrollregion=details_canvas.bbox("all")))
        details_canvas.bind("<Configure>", lambda event: details_canvas.itemconfigure(details_window_id, width=event.width))

        tk.Label(
            details_content,
            text="Los cambios son internos del HUB. WooCommerce no se toca desde esta ventana.",
            bg=INDIGO_SOFT,
            fg="#4338CA",
            wraplength=310,
            justify=tk.LEFT,
            padx=10,
            pady=8,
        ).pack(fill=tk.X, pady=(0, 10))

        for label, field in [
            ("Nombre", "name"),
            ("Estado comercial", "commercial_status"),
            ("Familia", "family"),
            ("Subgrupo", "subgroup"),
            ("Medidas", "size"),
            ("Materiales", "materials"),
            ("M3 unidad", "cubic_meters"),
            ("Rotación C", "rotation_c"),
            ("Bultos", "packages"),
            ("Precio proveedor", "primary_supplier_price"),
            ("Precio Pascal", "pascal_price"),
            ("HECA reference", "heca_reference"),
            ("Woo SKU", "woo_sku"),
            ("Stock tienda", "store_stock"),
            ("Stock almacén", "warehouse_stock"),
            ("Notas internas", "notes"),
        ]:
            self._editable_detail_row(details_content, label, edit_vars[field]).pack(fill=tk.X, pady=4)

        readonly_rows = [
            ("ID", item.code),
            ("Precio Woo", item.price),
            ("Stock total", f"{item.stock} unidades"),
            ("Woo ID", item.woo_id),
            ("Woo Parent ID", item.woo_parent_id),
            ("SKU Woo", item.sku_woo),
            ("Nombre Woo", item.woo_name),
            ("Categorias Woo", item.woo_categories),
            ("Tipo Woo", item.woo_item_kind),
            ("Estado vinculo Woo", item.woo_link_status),
            ("Coste calculado pedido", item.order_calculated_price),
            ("Coste medio ponderado", item.weighted_average_cost),
            ("Cantidad en pedido proveedor", item.supplier_order_qty),
            ("Proveedor pedido", item.supplier_order_provider),
            ("Estado", item.status),
        ]
        tk.Label(details_content, text="Datos de lectura", bg=CARD, fg=TEXT, font=("Segoe UI", 11, "bold")).pack(anchor=tk.W, pady=(12, 4))
        for label, value in readonly_rows:
            self._detail_row(details_content, label, value).pack(fill=tk.X, pady=4)

        self._render_inventory_pack_inline_box(details_content, item)

        tk.Label(details_content, text="Motivos del estado", bg=CARD, fg=TEXT, font=("Segoe UI", 11, "bold")).pack(anchor=tk.W, pady=(12, 4))
        reasons_box = tk.Frame(details_content, bg=SOFT, highlightbackground=LINE, highlightthickness=1)
        reasons_box.pack(fill=tk.X, pady=(0, 8))
        for reason in item.status_reasons:
            tk.Label(
                reasons_box,
                text=f"• {reason}",
                bg=SOFT,
                fg=TEXT if item.status == "OK" else (ROSE if item.status == "Critical" else ORANGE if item.status == "Error" else AMBER if item.status == "Warning" else MUTED),
                font=("Segoe UI", 9, "bold" if item.status in {"Critical", "Error", "Warning"} else "normal"),
                anchor=tk.W,
                justify=tk.LEFT,
                wraplength=330,
                padx=12,
                pady=4,
            ).pack(fill=tk.X, anchor=tk.W)

        left_actions = tk.Frame(left, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        left_actions.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 16))
        self._button(
            left_actions,
            "Guardar cambios",
            primary=True,
            command=lambda: self._open_inventory_changes_review(
                win,
                item,
                collect_changes(),
                on_discard=None,
                on_applied=lambda: self._after_inventory_item_updated(win),
            ),
        ).pack(fill=tk.X, padx=12, pady=(12, 7))
        self._button(left_actions, "Agregar a Propuesta de precios", command=lambda: self._open_inventory_proposal_modal(item)).pack(
            fill=tk.X,
            padx=12,
            pady=(0, 12),
        )

        right = tk.Frame(body, bg=BG)
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        price_host = tk.Frame(right, bg=BG)
        price_host.grid(row=0, column=0, sticky="nsew", pady=(0, 14))
        stock_host = tk.Frame(right, bg=BG)
        stock_host.grid(row=1, column=0, sticky="nsew")
        self._render_inventory_history_card(price_host, "Historial de precios", [], "Cargando historial real...", item.price, INDIGO)
        self._render_inventory_history_card(stock_host, "Historial de stock", [], "Cargando historial real...", item.stock, GREEN)

        def history_worker() -> None:
            try:
                if self._cloud_session is None:
                    raise RuntimeError("No hay sesion Supabase activa.")
                history = fetch_inventory_item_history(self._cloud_session, int(item.code), limit=120)
                self.after(0, lambda: self._render_inventory_history(price_host, stock_host, history, item))
            except Exception as exc:
                self.after(0, lambda exc=exc: self._render_inventory_history_error(price_host, stock_host, str(exc), item))

        threading.Thread(target=history_worker, daemon=True).start()

    def _inventory_editable_initial_values(self, item: InventoryItem) -> dict[str, str]:
        raw = item.raw or {}
        return {
            "name": str(raw.get("name") or item.name or ""),
            "family": str(raw.get("family") or ("" if item.family in {"Sin definir", "-"} else item.family)),
            "subgroup": str(raw.get("subgroup") or ("" if item.subgroup in {"Sin definir", "-"} else item.subgroup)),
            "materials": str(raw.get("materials") or ("" if item.material in {"Sin definir", "-"} else item.material)),
            "size": str(raw.get("size") or ("" if item.measures in {"Sin definir", "-"} else item.measures)),
            "cubic_meters": "" if raw.get("cubic_meters") in (None, "") else str(raw.get("cubic_meters")),
            "rotation_c": "" if raw.get("rotation_c") in (None, "") else str(raw.get("rotation_c")),
            "packages": "" if raw.get("packages") in (None, "") else str(raw.get("packages")),
            "primary_supplier_price": "" if raw.get("primary_supplier_price") in (None, "") else str(raw.get("primary_supplier_price")),
            "pascal_price": "" if raw.get("pascal_price") in (None, "") else str(raw.get("pascal_price")),
            "commercial_status": str(raw.get("commercial_status") or "Normal"),
            "heca_reference": str(raw.get("heca_reference") or ""),
            "woo_sku": str(raw.get("woo_sku") or item.sku_woo or ""),
            "store_stock": "" if raw.get("store_stock") in (None, "") else str(raw.get("store_stock")),
            "warehouse_stock": "" if raw.get("warehouse_stock") in (None, "") else str(raw.get("warehouse_stock")),
            "notes": str(raw.get("notes") or ""),
        }

    def _collect_inventory_detail_changes(self, initial: dict[str, str], vars_by_field: dict[str, tk.StringVar]) -> dict[str, tuple[str, str]]:
        changes: dict[str, tuple[str, str]] = {}
        for field, old_value in initial.items():
            new_value = vars_by_field[field].get().strip()
            old_clean = str(old_value or "").strip()
            if new_value != old_clean:
                changes[field] = (old_clean, new_value)
        return changes

    def _editable_detail_row(self, parent: tk.Misc, label: str, variable: tk.StringVar) -> tk.Frame:
        frame = tk.Frame(parent, bg=SOFT, highlightbackground=LINE, highlightthickness=1)
        frame.columnconfigure(1, weight=1)
        tk.Label(frame, text=label, bg=SOFT, fg=MUTED, font=("Segoe UI", 9), anchor=tk.W).grid(row=0, column=0, sticky="w", padx=12, pady=9)
        entry = tk.Entry(frame, textvariable=variable, bg="white", fg=TEXT, insertbackground=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightcolor=INDIGO, highlightthickness=1, font=("Segoe UI", 9))
        entry.grid(row=0, column=1, sticky="ew", padx=(0, 12), pady=8, ipady=4)
        return frame

    def _open_inventory_changes_review(
        self,
        owner: tk.Toplevel,
        item: InventoryItem,
        changes: dict[str, tuple[str, str]],
        *,
        on_discard: object | None,
        on_applied: object | None,
    ) -> None:
        if not changes:
            messagebox.showinfo("Inventario", "No hay cambios pendientes.")
            return
        review = tk.Toplevel(owner)
        review.title("Revisar cambios del item")
        review.configure(bg=BG)
        review.transient(owner)
        review.grab_set()
        center_window(review, 760, 520)
        review.columnconfigure(0, weight=1)
        review.rowconfigure(1, weight=1)

        header = tk.Frame(review, bg=BG, highlightbackground=LINE, highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)
        tk.Label(header, text="Cambios detectados", bg=INDIGO_SOFT, fg=INDIGO, font=("Segoe UI", 9, "bold"), padx=10, pady=4).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 4))
        tk.Label(header, text=f"{item.name} · {item.code}", bg=BG, fg=TEXT, font=("Segoe UI", 16, "bold")).grid(row=1, column=0, sticky="w", padx=18, pady=(0, 16))

        table_card = self._card(review)
        table_card.grid(row=1, column=0, sticky="nsew", padx=18, pady=18)
        table_card.rowconfigure(0, weight=1)
        table_card.columnconfigure(0, weight=1)
        columns = ("Campo", "Valor anterior", "Valor nuevo")
        tree = ttk.Treeview(table_card, columns=columns, show="headings", height=10)
        for column, width in {"Campo": 170, "Valor anterior": 250, "Valor nuevo": 250}.items():
            tree.heading(column, text=column, anchor=tk.CENTER)
            tree.column(column, width=width, anchor=tk.CENTER)
        for field, (before, after) in changes.items():
            tree.insert("", tk.END, values=(field, before or "Sin definir", after or "Sin definir"))
        tree.grid(row=0, column=0, sticky="nsew", padx=14, pady=14)

        footer = tk.Frame(review, bg=BG)
        footer.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))
        footer.columnconfigure(0, weight=1)
        if on_discard is not None:
            self._button(footer, "Descartar cambios", command=lambda: (review.destroy(), on_discard())).grid(row=0, column=0, sticky="w")
        self._button(footer, "Cancelar", command=review.destroy).grid(row=0, column=1, padx=(8, 0), sticky="e")
        self._button(
            footer,
            "Aceptar y guardar",
            primary=True,
            command=lambda: self._apply_inventory_detail_changes(review, item, changes, on_applied),
        ).grid(row=0, column=2, padx=(8, 0), sticky="e")

    def _apply_inventory_detail_changes(self, review: tk.Toplevel, item: InventoryItem, changes: dict[str, tuple[str, str]], on_applied: object | None) -> None:
        if self._cloud_session is None:
            messagebox.showerror("Inventario", "No hay sesion Supabase activa.")
            return
        payload = {field: after for field, (_before, after) in changes.items()}
        for child in review.winfo_children():
            child.configure(cursor="watch") if hasattr(child, "configure") else None
        review.update_idletasks()

        def worker() -> None:
            try:
                result = update_inventory_item_fields(
                    self._cloud_session,
                    int(item.code),
                    payload,
                    notes="Cambio aceptado desde detalle completo de Inventario UI ERP.",
                )
                self.after(0, lambda: finish_ok(result))
            except Exception as exc:
                self.after(0, lambda exc=exc: finish_error(exc))

        def finish_ok(result: dict[str, Any]) -> None:
            if review.winfo_exists():
                review.destroy()
            messagebox.showinfo("Inventario", f"Item actualizado.\noperation_id: {result.get('operation_id')}")
            if on_applied is not None:
                on_applied()

        def finish_error(exc: Exception) -> None:
            if review.winfo_exists():
                for child in review.winfo_children():
                    child.configure(cursor="") if hasattr(child, "configure") else None
            messagebox.showerror("Inventario", f"No se pudo guardar el item: {exc}")

        threading.Thread(target=worker, daemon=True).start()

    def _after_inventory_item_updated(self, detail_window: tk.Toplevel) -> None:
        if detail_window.winfo_exists():
            detail_window.destroy()
        if self._content is not None and self._current_key == "inventario":
            self._inventory_loaded_once = False
            self._refresh_inventory(self._content, self._inventory_query, allow_empty=True)

    def _render_inventory_history(self, price_host: tk.Frame, stock_host: tk.Frame, history: list[dict[str, Any]], item: InventoryItem) -> None:
        price_history = [row for row in history if str(row.get("field") or "").lower() in {"price", "woo_price", "precio", "precio local", "price_local"}]
        stock_history = [row for row in history if str(row.get("field") or "").lower() in {"store_stock", "warehouse_stock", "stock", "stock_total"}]
        self._render_inventory_history_card(price_host, "Historial de precios", price_history, "Sin historial de precios registrado", item.price, INDIGO)
        self._render_inventory_history_card(stock_host, "Historial de stock", stock_history, "Sin historial de stock registrado", item.stock, GREEN)

    def _render_inventory_history_error(self, price_host: tk.Frame, stock_host: tk.Frame, error: str, item: InventoryItem) -> None:
        self._render_inventory_history_card(price_host, "Historial de precios", [], f"No se pudo cargar historial: {error}", item.price, INDIGO)
        self._render_inventory_history_card(stock_host, "Historial de stock", [], f"No se pudo cargar historial: {error}", item.stock, GREEN)

    def _render_inventory_history_card(self, parent: tk.Frame, title: str, history: list[dict[str, Any]], empty_text: str, current_value: str, color: str) -> None:
        for child in parent.winfo_children():
            child.destroy()
        frame = self._card(parent)
        frame.pack(fill=tk.BOTH, expand=True)
        tk.Label(frame, text=title, bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=16, pady=(16, 4))
        tk.Label(frame, text="Historial real guardado en Supabase / caja negra.", bg=CARD, fg=MUTED).pack(anchor=tk.W, padx=16, pady=(0, 10))
        canvas = tk.Canvas(frame, height=145, bg="#FFFFFF", highlightbackground=SOFT, highlightthickness=1)
        canvas.pack(fill=tk.X, padx=16, pady=(0, 8))
        for x in range(24, 560, 86):
            canvas.create_line(x, 18, x, 124, fill=SOFT)
        for y in range(24, 130, 34):
            canvas.create_line(18, y, 560, y, fill=SOFT)
        points: list[tuple[float, float]] = []
        values: list[float] = []
        ordered = list(reversed(history[-8:]))
        for row in ordered:
            raw_value = row.get("after")
            try:
                values.append(float(str(raw_value).replace("€", "").replace(",", ".").strip()))
            except Exception:
                continue
        if values:
            minimum, maximum = min(values), max(values)
            spread = maximum - minimum or 1.0
            width = 520
            step = width / max(len(values) - 1, 1)
            for idx, value in enumerate(values):
                x = 24 + idx * step
                y = 124 - ((value - minimum) / spread) * 96
                points.append((x, y))
            for start, end in zip(points, points[1:]):
                canvas.create_line(*start, *end, fill=color, width=3)
            for x, y in points:
                canvas.create_oval(x - 4, y - 4, x + 4, y + 4, fill=color, outline=color)
        else:
            canvas.create_text(290, 72, text=empty_text, fill=MUTED, font=("Segoe UI", 10, "bold"))
        footer = tk.Frame(frame, bg=CARD)
        footer.pack(fill=tk.X, padx=16, pady=(0, 16))
        tk.Label(footer, text=f"Eventos: {len(history)}", bg=CARD, fg=MUTED, font=("Segoe UI", 9)).pack(side=tk.LEFT)
        tk.Label(footer, text=f"Valor actual: {current_value}", bg=CARD, fg=TEXT, font=("Segoe UI", 9, "bold")).pack(side=tk.RIGHT)

    def _open_inventory_proposal_modal(self, item: InventoryItem) -> None:
        win = tk.Toplevel(self)
        win.title("Agregar a Propuesta de precios")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        win.resizable(False, False)

        card = tk.Frame(win, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=18, pady=18)
        header = tk.Frame(card, bg=CARD)
        header.pack(fill=tk.X, padx=22, pady=(20, 12))
        header.columnconfigure(0, weight=1)
        tk.Label(header, text="Agregar a Propuesta de precios", bg=INDIGO_SOFT, fg=INDIGO, font=("Segoe UI", 9, "bold"), padx=10, pady=4).grid(
            row=0,
            column=0,
            sticky="w",
        )
        tk.Label(header, text="Donde quieres anadir este item?", bg=CARD, fg=TEXT, font=("Segoe UI", 16, "bold")).grid(
            row=1,
            column=0,
            sticky="w",
            pady=(10, 2),
        )
        tk.Label(header, text=f"Producto seleccionado: {item.name} - {item.code}", bg=CARD, fg=MUTED).grid(row=2, column=0, sticky="w")
        self._button(header, "Cerrar", command=win.destroy).grid(row=0, column=1, rowspan=2, sticky="ne")

        options = tk.Frame(card, bg=CARD)
        options.pack(fill=tk.X, padx=22, pady=(0, 12))
        options.columnconfigure(0, weight=1)
        options.columnconfigure(1, weight=1)
        self._proposal_option(
            options,
            "+",
            "Anadir a Nueva Propuesta",
            "Crea una nueva propuesta de precios y anade este producto como primera linea.",
            lambda: self._carry_inventory_item_to_prices(item, win),
        ).grid(row=0, column=0, sticky="ew", padx=(0, 7))
        self._proposal_option(
            options,
            "=",
            "Anadir a Propuesta Existente",
            "Abre el selector de propuestas abiertas para anadir este producto.",
            lambda: self._carry_inventory_item_to_prices(item, win),
        ).grid(row=0, column=1, sticky="ew", padx=(7, 0))

        tk.Label(
            card,
            text="Logica futura: seleccionar opcion -> validar producto -> anadir linea -> confirmar visualmente -> registrar log.",
            bg=INDIGO_SOFT,
            fg="#4338CA",
            wraplength=560,
            justify=tk.LEFT,
            padx=12,
            pady=10,
        ).pack(fill=tk.X, padx=22, pady=(0, 14))

        footer = tk.Frame(card, bg=SOFT, highlightbackground=SOFT, highlightthickness=1)
        footer.pack(fill=tk.X)
        self._button(footer, "Cancelar", command=win.destroy).pack(side=tk.RIGHT, padx=22, pady=14)
        win.update_idletasks()
        center_window(win, max(660, win.winfo_reqwidth() + 36), max(360, win.winfo_reqheight() + 36))

    def _proposal_option(
        self,
        parent: tk.Misc,
        icon: str,
        title: str,
        subtitle: str,
        command: object,
    ) -> tk.Frame:
        frame = tk.Frame(parent, bg=SOFT, highlightbackground=LINE, highlightthickness=1, cursor="hand2")
        tk.Label(frame, text=icon, bg=CARD, fg=INDIGO, font=("Segoe UI", 18, "bold"), width=3).pack(anchor=tk.W, padx=14, pady=(14, 8))
        tk.Label(frame, text=title, bg=SOFT, fg=TEXT, font=("Segoe UI", 11, "bold"), wraplength=230, justify=tk.LEFT).pack(anchor=tk.W, padx=14)
        tk.Label(frame, text=subtitle, bg=SOFT, fg=MUTED, font=("Segoe UI", 9), wraplength=230, justify=tk.LEFT).pack(
            anchor=tk.W,
            padx=14,
            pady=(6, 14),
        )
        for widget in (frame, *frame.winfo_children()):
            widget.bind("<Button-1>", lambda _event: command(), add="+")
            widget.configure(cursor="hand2")
        return frame

    def _carry_inventory_item_to_prices(self, item: InventoryItem, modal: tk.Toplevel) -> None:
        self._proposal_source_item = item
        self._price_mode = "edit"
        if modal.winfo_exists():
            modal.grab_release()
            modal.destroy()
        self._show_view("precios")

    def _build_prices(self, parent: tk.Frame) -> None:
        self._page_header(parent, "Operaciones", "Cambio de Precios", "Propuestas, validacion, aprobacion y publicacion protegida.", ["Nueva propuesta"])
        if self._proposal_source_item is not None:
            item = self._proposal_source_item
            notice = tk.Frame(parent, bg=INDIGO_SOFT, highlightbackground="#C7D2FE", highlightthickness=1)
            notice.pack(fill=tk.X, pady=(0, 14))
            tk.Label(
                notice,
                text=f"Item recibido desde Inventario: {item.name} - {item.code}",
                bg=INDIGO_SOFT,
                fg=INDIGO,
                font=("Segoe UI", 10, "bold"),
                anchor=tk.W,
            ).pack(fill=tk.X, padx=14, pady=(10, 2))
            tk.Label(
                notice,
                text="Flujo previsto: validar producto, crear linea de propuesta, confirmar visualmente y registrar log.",
                bg=INDIGO_SOFT,
                fg="#4338CA",
                font=("Segoe UI", 9),
                anchor=tk.W,
            ).pack(fill=tk.X, padx=14, pady=(0, 10))
        if self._price_mode == "edit":
            self._build_price_edit_workspace(parent)
        else:
            self._build_saved_proposals_workspace(parent)
            if self._cloud_session is not None and not self._price_loaded_once and not self._price_loading:
                self._refresh_price_proposals(parent)

    def _set_price_mode(self, mode: str) -> None:
        if mode == "edit":
            self._prepare_price_edit_state()
        else:
            self._price_edit_notice = ""
        self._price_mode = mode
        self._show_view("precios")

    def _prepare_price_edit_state(self) -> None:
        """Prepare the reusable proposal editor.

        Nueva propuesta: list starts empty.
        Modificar propuesta: list starts with the selected proposal lines.
        """
        if self._price_edit_initialized:
            return
        self._price_edit_initialized = True
        proposal = self._selected_price_proposal
        self._price_edit_lines = list(proposal.lines) if proposal else []
        if self._proposal_source_item is not None and all(line.code != self._proposal_source_item.code for line in self._price_edit_lines):
            item = self._proposal_source_item
            self._price_edit_lines.insert(0, ProposalLine(item.code, item.name, item.price, item.price, "Pendiente", "flat"))
            source = self._price_source_from_inventory_item(item)
            if source:
                self._price_line_sources[item.code] = source

    def _price_reset_edit_state(self) -> None:
        self._price_edit_lines = []
        self._price_edit_initialized = False
        self._price_edit_selected_code = ""
        self._price_edit_notice = ""
        self._price_search_query = ""
        if hasattr(self, "_price_edit_name_var"):
            self._price_edit_name_var.set("")

    def _normalize_search_text(self, value: object) -> str:
        text = str(value or "").lower().strip()
        text = "".join(
            char for char in unicodedata.normalize("NFD", text)
            if unicodedata.category(char) != "Mn"
        )
        return text

    def _inventory_row_matches_query(self, row: dict[str, Any], query: str) -> bool:
        needle = self._normalize_search_text(query)
        if not needle:
            return True
        parts = [
            row.get("item_id"),
            row.get("name"),
            row.get("family"),
            row.get("subgroup"),
            row.get("materials"),
            row.get("size"),
            row.get("woo_sku"),
            row.get("woo_name"),
            row.get("woo_id"),
            row.get("woo_categories"),
            row.get("supplier_order_provider"),
            row.get("heca_reference"),
            row.get("hub_item_code"),
            row.get("base_item_code"),
            row.get("hub_pack_components_text"),
            row.get("hub_search_related_code"),
            row.get("hub_search_related_name"),
        ]
        haystack = self._normalize_search_text(" ".join(str(part or "") for part in parts))
        return needle in haystack

    def _merge_inventory_rows(self, groups: list[list[dict[str, Any]]]) -> list[dict[str, Any]]:
        merged: list[dict[str, Any]] = []
        seen: set[str] = set()
        for rows in groups:
            for row in rows or []:
                key = str(row.get("item_id") or row.get("woo_id") or row)
                if key in seen:
                    continue
                seen.add(key)
                merged.append(row)
        return merged

    def _accent_insensitive_inventory_search(self, rows: list[dict[str, Any]], query: str) -> list[dict[str, Any]]:
        return [row for row in rows if self._inventory_row_matches_query(row, query)]

    def _set_proposal_search(self, query: str) -> None:
        self._proposal_search_query = (query or "").strip()
        self._price_mode = "saved"
        self._show_view("precios")

    def _proposal_matches_search(self, proposal: PriceProposal, query: str) -> bool:
        needle = self._normalize_search_text(query)
        if not needle:
            return True
        haystack_parts = [
            proposal.name,
            proposal.date,
            proposal.status,
            str(proposal.items),
            str(proposal.up),
            str(proposal.down),
        ]
        for line in proposal.lines:
            haystack_parts.extend([line.code, line.name, line.old_price, line.new_price, line.change])
        haystack = self._normalize_search_text(" ".join(str(part or "") for part in haystack_parts))
        return needle in haystack

    def _build_saved_proposals_workspace(self, parent: tk.Frame) -> None:
        search_card = self._card(parent)
        search_card.pack(fill=tk.X, pady=(0, 14))
        search_row = tk.Frame(search_card, bg=CARD)
        search_row.pack(fill=tk.X, padx=14, pady=14)
        search_row.columnconfigure(0, weight=1)
        search = tk.Entry(
            search_row,
            bg=CARD,
            fg=TEXT,
            insertbackground=TEXT,
            relief=tk.FLAT,
            highlightbackground=LINE,
            highlightcolor=INDIGO,
            highlightthickness=1,
            font=("Segoe UI", 10),
        )
        search.insert(0, self._proposal_search_query or "")
        search.grid(row=0, column=0, sticky="ew", ipady=9)
        search.bind("<Return>", lambda _event: self._set_proposal_search(search.get()))
        self._button(search_row, "Buscar", command=lambda: self._set_proposal_search(search.get())).grid(row=0, column=1, padx=(10, 0))
        self._button(search_row, "Limpiar", command=lambda: self._set_proposal_search("")).grid(row=0, column=2, padx=(8, 0))

        body = tk.Frame(parent, bg=BG)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=3, minsize=620)
        body.columnconfigure(1, weight=2, minsize=420)
        body.rowconfigure(0, weight=1)

        list_card = self._card(body)
        list_card.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        list_card.rowconfigure(1, weight=1)
        list_card.columnconfigure(0, weight=1)
        head = tk.Frame(list_card, bg=CARD)
        head.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 10))
        head.columnconfigure(0, weight=1)
        tk.Label(head, text="Propuestas", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w")
        source_proposals = self._price_proposals if (self._cloud_session is not None and self._price_loaded_once) else (self._price_proposals or list(SAVED_PROPOSALS))
        proposals = [proposal for proposal in source_proposals if self._proposal_matches_search(proposal, self._proposal_search_query)]
        state_text = "Cargando reales..." if self._price_loading else f"{len(proposals)} visibles"
        self._status_chip(head, state_text, "Info").grid(row=0, column=1, sticky="e", padx=(0, 8))
        self._button(head, "Actualizar", primary=True, command=lambda: self._refresh_price_proposals(parent)).grid(row=0, column=2, sticky="e")
        if self._price_error:
            tk.Label(list_card, text=self._price_error, bg=INDIGO_SOFT if self._price_loading else ROSE_SOFT, fg=INDIGO if self._price_loading else ROSE, anchor=tk.W).grid(
                row=2,
                column=0,
                sticky="ew",
                padx=16,
                pady=(0, 12),
            )

        table_host = tk.Frame(list_card, bg=CARD)
        table_host.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 16))
        table_host.rowconfigure(0, weight=1)
        table_host.columnconfigure(0, weight=1)
        canvas = tk.Canvas(table_host, bg=CARD, highlightthickness=0)
        scrollbar = ttk.Scrollbar(table_host, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        table = tk.Frame(canvas, bg=CARD)
        table_window = canvas.create_window((0, 0), window=table, anchor="nw")
        table.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda event: canvas.itemconfigure(table_window, width=event.width))
        for i, width in enumerate([4, 1, 1, 1, 1]):
            table.columnconfigure(i, weight=width)
        headers = ["Propuesta", "Items", "Suben", "Bajan", "Estado"]
        for i, label in enumerate(headers):
            tk.Label(table, text=label.upper(), bg=CARD, fg=MUTED, font=("Segoe UI", 8, "bold")).grid(
                row=0,
                column=i,
                sticky="ew",
                padx=4,
                pady=(0, 8),
            )
        detail_host = tk.Frame(body, bg=BG)
        detail_host.grid(row=0, column=1, sticky="nsew")

        def select_proposal(proposal: PriceProposal) -> None:
            self._selected_price_proposal = proposal
            self._show_view("precios")

        if not proposals:
            self._selected_price_proposal = None
            tk.Label(
                table,
                text="No hay propuestas que coincidan con la busqueda.",
                bg=CARD,
                fg=MUTED,
                font=("Segoe UI", 10),
                anchor=tk.W,
                padx=12,
                pady=12,
            ).grid(row=1, column=0, columnspan=len(headers), sticky="ew", pady=8)
            self._render_empty_saved_proposal_detail(detail_host)
            return
        if self._selected_price_proposal not in proposals:
            self._selected_price_proposal = proposals[0]

        for row_index, proposal in enumerate(proposals, start=1):
            selected = proposal == self._selected_price_proposal
            table.rowconfigure(row_index, minsize=64)
            values = [
                f"{proposal.name}\n{proposal.date}",
                str(proposal.items),
                str(proposal.up),
                str(proposal.down),
                proposal.status,
            ]
            for column_index, value in enumerate(values):
                fg, bg = self._proposal_list_cell_style(proposal, column_index, selected)
                label = tk.Label(
                    table,
                    text=value,
                    bg=bg,
                    fg=fg,
                    font=("Segoe UI", 9, "bold" if column_index in (0, 4) else "normal"),
                    padx=10,
                    pady=10,
                    justify=tk.LEFT if column_index == 0 else tk.CENTER,
                    anchor=tk.W if column_index == 0 else tk.CENTER,
                    highlightbackground=LINE,
                    highlightthickness=1,
                    cursor="hand2",
                    wraplength=280 if column_index == 0 else 90,
                )
                label.grid(row=row_index, column=column_index, sticky="nsew", padx=3, pady=3)
                label.bind("<Button-1>", lambda _event, proposal=proposal: select_proposal(proposal), add="+")

        self._render_saved_proposal_detail(detail_host, self._selected_price_proposal)

    def _refresh_price_proposals(self, parent: tk.Frame) -> None:
        if self._cloud_session is None:
            self._price_error = "Inicia sesion para cargar propuestas reales."
            self._price_loading = False
            self._price_loaded_once = True
            if self._current_key == "precios":
                self._show_view("precios")
            return
        self._price_loading = True
        self._price_error = "Cargando propuestas reales..."
        self._show_view("precios")

        def worker() -> None:
            try:
                rows = list_real_price_proposals(self._cloud_session, status="all", limit=100)
                proposals = [self._price_proposal_from_cloud_row(row) for row in rows]
                self.after(0, lambda: self._finish_price_proposals_refresh(proposals, "" if proposals else "No hay propuestas reales visibles."))
            except Exception as exc:
                self.after(0, lambda exc=exc: self._finish_price_proposals_refresh([], f"No se pudieron cargar propuestas reales: {exc}"))

        threading.Thread(target=worker, daemon=True).start()

    def _finish_price_proposals_refresh(self, proposals: list[PriceProposal], error: str) -> None:
        # Si hay sesión cloud, una lista vacía es un estado real: no volver a rellenar con mocks.
        # Los mocks solo sirven cuando no hay sesión/entorno real disponible.
        self._price_proposals = proposals if self._cloud_session is not None else (proposals or list(SAVED_PROPOSALS))
        self._price_error = error
        self._price_loading = False
        self._price_loaded_once = True
        self._selected_price_proposal = self._price_proposals[0] if self._price_proposals else None
        if self._current_key == "precios":
            self._show_view("precios")

    def _price_proposal_from_cloud_row(self, row: dict[str, Any]) -> PriceProposal:
        old_price = self._money_or_none(row.get("old_price"))
        new_price = self._money_or_none(row.get("new_price"))
        delta = None if old_price is None or new_price is None else new_price - old_price
        change = "Pendiente"
        direction = "flat"
        if delta is not None:
            direction = "up" if delta > 0 else "down" if delta < 0 else "flat"
            if old_price not in (None, 0):
                change = f"{(delta / old_price) * 100:+.2f}%"
            else:
                change = f"{delta:+.2f}"
        status = self._price_status_label_from_cloud(row.get("status"))
        source_row = row.get("source_row") if isinstance(row.get("source_row"), dict) else {}
        display_name = str(source_row.get("ui_proposal_name") or row.get("proposal_name") or row.get("title") or row.get("name") or f"Propuesta {row.get('id') or '-'}")
        line = ProposalLine(
            code=f"{row.get('item_kind') or '-'}:{row.get('item_woo_id') or row.get('local_id') or '-'}",
            name=str(row.get("name") or "Producto sin nombre"),
            old_price=self._format_price_value(old_price),
            new_price=self._format_price_value(new_price),
            change=change,
            direction=direction,
        )
        return PriceProposal(
            name=display_name,
            date=self._format_cloud_date(row.get("created_at")),
            items=1,
            up=1 if direction == "up" else 0,
            down=1 if direction == "down" else 0,
            flat=1 if direction == "flat" else 0,
            change=change,
            status=status,
            lines=(line,),
            raw=row,
        )

    def _price_status_label_from_cloud(self, raw: Any) -> str:
        value = str(raw or "").strip().lower()
        return {
            "pending": "Pendiente",
            "approved": "Aprobada",
            "rejected": "Rechazada",
            "publishing": "Publicando",
            "published": "Publicada",
            "rolled_back": "Restaurada",
            "failed": "Fallida",
        }.get(value, str(raw or "-"))

    def _price_status_raw_from_label(self, label: str) -> str:
        value = str(label or "").strip().lower()
        return {
            "pendiente": "pending",
            "aprobada": "approved",
            "rechazada": "rejected",
            "publicando": "publishing",
            "publicada": "published",
            "fallida": "failed",
        }.get(value, value)

    def _format_cloud_date(self, value: Any) -> str:
        text = str(value or "").strip()
        if not text:
            return "-"
        return text.replace("T", " ")[:19]

    def _money_or_none(self, value: Any) -> float | None:
        if value in (None, ""):
            return None
        try:
            return float(str(value).replace(",", "."))
        except Exception:
            return None

    def _format_price_value(self, value: float | None) -> str:
        if value is None:
            return "-"
        return f"{value:.2f}"

    def _proposal_list_cell_style(self, proposal: PriceProposal, column_index: int, selected: bool) -> tuple[str, str]:
        if selected:
            return TEXT, INDIGO_SOFT
        if column_index == 2 and proposal.up:
            return GREEN, GREEN_SOFT
        if column_index == 3 and proposal.down:
            return ROSE, ROSE_SOFT
        if column_index == 4:
            return STATUS_STYLES.get(proposal.status, (MUTED, SOFT))
        return TEXT if column_index == 0 else MUTED, CARD

    def _render_saved_proposal_detail(self, parent: tk.Frame, proposal: PriceProposal) -> None:
        for child in parent.winfo_children():
            child.destroy()
        detail = self._card(parent)
        detail.pack(fill=tk.BOTH, expand=True)
        detail.rowconfigure(1, weight=1)
        detail.columnconfigure(0, weight=1)
        head = tk.Frame(detail, bg=CARD)
        head.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 10))
        head.columnconfigure(0, weight=1)
        tk.Label(head, text="Detalles de propuesta", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w")
        self._status_chip(head, proposal.status, proposal.status).grid(row=0, column=1, sticky="e")

        scroll = tk.Frame(detail, bg=CARD)
        scroll.grid(row=1, column=0, sticky="nsew", padx=16)
        tk.Label(scroll, text=proposal.name, bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold"), wraplength=380, justify=tk.LEFT).pack(anchor=tk.W, fill=tk.X)
        tk.Label(scroll, text=proposal.date, bg=CARD, fg=MUTED).pack(anchor=tk.W, pady=(2, 12))

        summary = tk.Frame(scroll, bg=CARD)
        summary.pack(fill=tk.X, pady=(0, 12))
        for i, (label, value, status) in enumerate(
            [
                ("Items", proposal.items, "Info"),
                ("Suben", proposal.up, "OK"),
                ("Bajan", proposal.down, "Critical" if proposal.down else "Info"),
                ("Igual", proposal.flat, "Info"),
            ]
        ):
            summary.columnconfigure(i, weight=1)
            self._metric(summary, label, str(value), status).grid(row=0, column=i, sticky="ew", padx=(0 if i == 0 else 6, 0))

        for line in proposal.lines:
            self._proposal_line_preview(scroll, line).pack(fill=tk.X, pady=4)

        footer = tk.Frame(detail, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        footer.grid(row=2, column=0, sticky="ew", padx=16, pady=16)
        top_actions = tk.Frame(footer, bg=CARD)
        top_actions.pack(fill=tk.X, padx=12, pady=(12, 7))
        self._button(top_actions, "Modificar", primary=True, command=lambda: self._set_price_mode("edit")).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        self._button(top_actions, "Borrar propuesta", command=lambda: self._open_delete_price_proposal_confirmation(proposal)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0))
        row = tk.Frame(footer, bg=CARD)
        row.pack(fill=tk.X, padx=12, pady=(0, 12))
        can_review = self._proposal_raw_status(proposal) == "pending"
        accept = self._button(row, "Aceptar propuesta", command=lambda: self._open_price_review_modal(proposal, "approved"))
        reject = self._button(row, "Rechazar propuesta", command=lambda: self._open_price_review_modal(proposal, "rejected"))
        if not can_review:
            accept.configure(state=tk.DISABLED)
            reject.configure(state=tk.DISABLED)
        accept.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        reject.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0))

    def _proposal_raw_status(self, proposal: PriceProposal) -> str:
        raw_status = (proposal.raw or {}).get("status")
        if raw_status:
            return str(raw_status).strip().lower()
        return self._price_status_raw_from_label(proposal.status)

    def _open_delete_price_proposal_confirmation(self, proposal: PriceProposal) -> None:
        row = proposal.raw or {}
        proposal_id = row.get("id")
        source_row = row.get("source_row") if isinstance(row.get("source_row"), dict) else {}
        proposal_name = str(source_row.get("ui_proposal_name") or proposal.name or "").strip()
        if not proposal_id and proposal in self._price_proposals:
            if not messagebox.askyesno("Borrar propuesta", f"Borrar la propuesta visual '{proposal.name}'?"):
                return
            self._price_proposals = [item for item in self._price_proposals if item != proposal]
            self._selected_price_proposal = self._price_proposals[0] if self._price_proposals else None
            self._show_view("precios")
            return
        if not proposal_id or self._cloud_session is None:
            messagebox.showinfo("Cambio de Precios", "Esta propuesta no tiene ID real o no hay sesión activa.")
            return
        label = proposal_name or f"ID {proposal_id}"
        if not messagebox.askyesno(
            "Borrar propuesta",
            "Se eliminará la propuesta completa de Cambio de Precios.\n\n"
            f"Propuesta: {label}\n\n"
            "Si la propuesta contiene varios items con el mismo nombre de propuesta, se quitarán todos.\n"
            "Esta acción quedará registrada en Seguridad / Logs.\n\n"
            "¿Continuar?",
        ):
            return
        self._price_error = "Borrando propuesta..."
        self._price_delete_target_id = str(proposal_id)
        self._price_delete_target_name = proposal_name
        self._show_view("precios")

        def worker() -> None:
            try:
                result = delete_real_price_proposal_group(self._cloud_session, str(proposal_id), proposal_name=proposal_name)
                deleted = int(result.get("deleted_count") or 0)
                self.after(0, lambda: self._finish_delete_price_proposal(deleted))
            except Exception as exc:
                self.after(0, lambda exc=exc: self._finish_delete_price_proposal_error(str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def _finish_delete_price_proposal(self, deleted_count: int) -> None:
        messagebox.showinfo("Cambio de Precios", f"Propuesta borrada correctamente. Registros afectados: {deleted_count}")
        deleted_id = str(getattr(self, "_price_delete_target_id", "") or "")
        deleted_name = str(getattr(self, "_price_delete_target_name", "") or "").strip()

        def keep(proposal: PriceProposal) -> bool:
            row = proposal.raw or {}
            source = row.get("source_row") if isinstance(row.get("source_row"), dict) else {}
            row_id = str(row.get("id") or "")
            row_name = str(source.get("ui_proposal_name") or proposal.name or "").strip()
            if deleted_id and row_id == deleted_id:
                return False
            if deleted_name and row_name == deleted_name:
                return False
            return True

        self._price_proposals = [proposal for proposal in self._price_proposals if keep(proposal)]
        self._price_loaded_once = False
        self._selected_price_proposal = self._price_proposals[0] if self._price_proposals else None
        self._price_error = ""
        self._price_mode = "saved"
        self._show_view("precios")
        if self._cloud_session is not None:
            self._refresh_price_proposals(self._content)

    def _finish_delete_price_proposal_error(self, error: str) -> None:
        self._price_error = f"No se pudo borrar la propuesta: {error}"
        if self._current_key == "precios":
            self._show_view("precios")

    def _open_price_proposal_preview(self, proposal: PriceProposal) -> None:
        row = proposal.raw or {}
        proposal_id = row.get("id")
        if not proposal_id or self._cloud_session is None:
            messagebox.showinfo("Cambio de Precios", "Esta propuesta es visual/mock o no hay sesion activa. Carga propuestas reales para generar preview.")
            return
        win = tk.Toplevel(self)
        win.title("Preview propuesta de precio")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 760, 520)
        win.minsize(680, 420)
        body = tk.Frame(win, bg=BG)
        body.pack(fill=tk.BOTH, expand=True, padx=18, pady=18)
        tk.Label(body, text="Preview de seguridad", bg=BG, fg=TEXT, font=("Segoe UI", 18, "bold")).pack(anchor=tk.W)
        tk.Label(body, text="Recalcula validaciones sobre la propuesta real. No toca WooCommerce ni cambia estado.", bg=BG, fg=MUTED).pack(anchor=tk.W, pady=(2, 12))
        result = tk.Text(body, bg="#0F172A", fg="#E2E8F0", insertbackground="#E2E8F0", relief=tk.FLAT, wrap=tk.WORD, font=("Consolas", 9))
        result.pack(fill=tk.BOTH, expand=True)
        result.insert("1.0", "Generando preview real...")
        result.configure(state=tk.DISABLED)
        footer = tk.Frame(body, bg=BG)
        footer.pack(fill=tk.X, pady=(12, 0))
        self._button(footer, "Cerrar", command=win.destroy).pack(side=tk.RIGHT)

        def worker() -> None:
            try:
                preview = preview_existing_price_proposal(self._cloud_session, str(proposal_id), load_settings())
                text = format_existing_price_proposal_preview(preview)
            except Exception as exc:
                text = f"No se pudo generar preview real:\n{exc}"
            self.after(0, lambda: render(text))

        def render(text: str) -> None:
            if not win.winfo_exists():
                return
            result.configure(state=tk.NORMAL)
            result.delete("1.0", tk.END)
            result.insert("1.0", text)
            result.configure(state=tk.DISABLED)

        threading.Thread(target=worker, daemon=True).start()

    def _open_price_review_modal(self, proposal: PriceProposal, decision: str) -> None:
        row = proposal.raw or {}
        proposal_id = row.get("id")
        current_status = self._proposal_raw_status(proposal)
        if not proposal_id or self._cloud_session is None:
            messagebox.showinfo("Cambio de Precios", "Carga una propuesta real con sesion activa antes de revisar.")
            return
        if current_status != "pending":
            messagebox.showinfo("Cambio de Precios", f"Solo se pueden revisar propuestas pendientes. Estado actual: {proposal.status}")
            return

        token = "APROBAR" if decision == "approved" else "RECHAZAR"
        title = "Aprobar propuesta" if decision == "approved" else "Rechazar propuesta"
        win = tk.Toplevel(self)
        win.title(title)
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 820, 620)
        win.minsize(720, 500)

        body = tk.Frame(win, bg=BG)
        body.pack(fill=tk.BOTH, expand=True, padx=18, pady=18)
        tk.Label(body, text=title, bg=BG, fg=TEXT, font=("Segoe UI", 18, "bold")).pack(anchor=tk.W)
        tk.Label(
            body,
            text="Aceptar propuesta publica el precio en WooCommerce con validaciones, snapshot y audit log. Rechazar solo cambia el estado interno.",
            bg=BG,
            fg=MUTED,
        ).pack(anchor=tk.W, pady=(2, 12))

        preview_box = tk.Text(body, bg="#0F172A", fg="#E2E8F0", insertbackground="#E2E8F0", relief=tk.FLAT, wrap=tk.WORD, font=("Consolas", 9), height=16)
        preview_box.pack(fill=tk.BOTH, expand=True)
        preview_box.insert("1.0", "Generando validacion real antes de permitir la decision...")
        preview_box.configure(state=tk.DISABLED)

        confirm_row = tk.Frame(body, bg=BG)
        confirm_row.pack(fill=tk.X, pady=(12, 0))
        tk.Label(confirm_row, text=f"Escribe {token} para confirmar", bg=BG, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        confirm_var = tk.StringVar()
        confirm_entry = tk.Entry(confirm_row, textvariable=confirm_var, bg=CARD, fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightcolor=INDIGO, highlightthickness=1)
        confirm_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=12, ipady=8)

        status_label = tk.Label(body, text="", bg=BG, fg=MUTED, anchor=tk.W)
        status_label.pack(fill=tk.X, pady=(8, 0))

        footer = tk.Frame(body, bg=BG)
        footer.pack(fill=tk.X, pady=(12, 0))

        def render_preview(text: str) -> None:
            if not win.winfo_exists():
                return
            preview_box.configure(state=tk.NORMAL)
            preview_box.delete("1.0", tk.END)
            preview_box.insert("1.0", text)
            preview_box.configure(state=tk.DISABLED)

        def load_preview() -> None:
            try:
                preview = preview_existing_price_proposal(self._cloud_session, str(proposal_id), load_settings())
                text = format_existing_price_proposal_preview(preview)
            except Exception as exc:
                text = f"No se pudo generar preview real:\n{exc}"
            self.after(0, lambda: render_preview(text))

        def execute_review() -> None:
            if confirm_var.get().strip().upper() != token:
                status_label.configure(text=f"Confirmacion invalida. Escribe exactamente {token}.", fg=ROSE)
                return
            review_button.configure(state=tk.DISABLED)
            cancel_button.configure(state=tk.DISABLED)
            status_label.configure(text="Ejecutando flujo protegido...", fg=INDIGO)

            def worker() -> None:
                try:
                    settings = load_settings()
                    review_result = review_latest_real_price_proposal(self._cloud_session, decision, str(proposal_id), settings)
                    review_operation_id = review_result.get("operation_id") or "-"
                    publish_operation_id = None
                    if decision == "approved":
                        publish_result = publish_woocommerce_price(
                            self._cloud_session,
                            proposal_id=str(proposal_id),
                            confirm="PUBLICAR",
                            acknowledge_warnings=True,
                            settings=settings,
                        )
                        publish_operation_id = publish_result.get("operation_id") or "-"
                    self.after(0, lambda: finish_ok(review_operation_id, publish_operation_id))
                except Exception as exc:
                    self.after(0, lambda exc=exc: finish_error(str(exc)))

            threading.Thread(target=worker, daemon=True).start()

        def finish_ok(review_operation_id: str, publish_operation_id: str | None = None) -> None:
            if win.winfo_exists():
                win.grab_release()
                win.destroy()
            if publish_operation_id:
                message = (
                    "Propuesta aceptada y publicada en WooCommerce.\n"
                    f"Revision: {review_operation_id}\n"
                    f"Publicacion Woo: {publish_operation_id}"
                )
            else:
                message = f"Propuesta rechazada correctamente.\nOperacion: {review_operation_id}"
            messagebox.showinfo("Cambio de Precios", message)
            self._price_loaded_once = False
            if self._content is not None:
                self._refresh_price_proposals(self._content)

        def finish_error(error: str) -> None:
            status_label.configure(text=f"No se pudo ejecutar la revision: {error}", fg=ROSE)
            review_button.configure(state=tk.NORMAL)
            cancel_button.configure(state=tk.NORMAL)

        cancel_button = self._button(footer, "Cancelar", command=win.destroy)
        cancel_button.pack(side=tk.RIGHT)
        review_button = self._button(footer, token, primary=True, command=execute_review)
        review_button.pack(side=tk.RIGHT, padx=(0, 8))

        threading.Thread(target=load_preview, daemon=True).start()

    def _proposal_line_preview(self, parent: tk.Misc, line: ProposalLine) -> tk.Frame:
        status = "OK" if line.direction == "up" else "Critical" if line.direction in ("down", "critical") else "Info"
        frame = tk.Frame(parent, bg=SOFT, highlightbackground=LINE, highlightthickness=1)
        frame.columnconfigure(0, weight=1)
        text = tk.Frame(frame, bg=SOFT)
        text.grid(row=0, column=0, sticky="ew", padx=12, pady=9)
        tk.Label(text, text=f"{line.code} - {line.name}", bg=SOFT, fg=TEXT, font=("Segoe UI", 9, "bold"), anchor=tk.W, justify=tk.LEFT, wraplength=300).pack(anchor=tk.W, fill=tk.X)
        tk.Label(text, text=f"{line.old_price} -> {line.new_price}", bg=SOFT, fg=MUTED, font=("Segoe UI", 9)).pack(anchor=tk.W)
        self._status_chip(frame, line.change, status).grid(row=0, column=1, sticky="e", padx=10)
        return frame

    def _build_price_edit_workspace(self, parent: tk.Frame) -> None:
        self._prepare_price_edit_state()
        if (
            self._cloud_session is not None
            and not self._price_available_items
            and not self._price_items_loading
            and not self._price_items_error
        ):
            self.after(50, lambda parent=parent: self._refresh_price_edit_items(parent, self._price_search_query, allow_empty=True))

        body = tk.Frame(parent, bg=BG)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        left = tk.Frame(body, bg=BG)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        right = tk.Frame(body, bg=BG)
        right.grid(row=0, column=1, sticky="nsew")

        search_card = self._card(left)
        search_card.pack(fill=tk.X, pady=(0, 12))
        search_row = tk.Frame(search_card, bg=CARD)
        search_row.pack(fill=tk.X, padx=14, pady=14)
        search_row.columnconfigure(0, weight=1)
        search = tk.Entry(search_row, bg=CARD, fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightcolor=INDIGO, highlightthickness=1)
        search.insert(0, self._price_search_query or "")
        search.grid(row=0, column=0, sticky="ew", ipady=9)
        search.bind("<Return>", lambda _event: self._refresh_price_edit_items(parent, search.get(), allow_empty=True))
        self._button(search_row, "Buscar", command=lambda: self._refresh_price_edit_items(parent, search.get(), allow_empty=True)).grid(row=0, column=1, padx=(10, 0))
        self._button(search_row, "Recargar", command=lambda: self._refresh_price_edit_items(parent, "", allow_empty=True)).grid(row=0, column=2, padx=(8, 0))

        if self._price_items_loading:
            tk.Label(left, text="Cargando items reales desde Supabase...", bg=INDIGO_SOFT, fg=INDIGO, anchor=tk.W, padx=12, pady=8).pack(fill=tk.X, pady=(0, 12))
        if self._price_items_error:
            tk.Label(left, text=self._price_items_error, bg=ROSE_SOFT, fg=ROSE, anchor=tk.W, padx=12, pady=8).pack(fill=tk.X, pady=(0, 12))
        if self._price_edit_notice:
            notice = tk.Label(left, text=self._price_edit_notice, bg=AMBER_SOFT, fg=AMBER, anchor=tk.W, padx=12, pady=8)
            notice.pack(fill=tk.X, pady=(0, 12))

        available_items = list(self._price_available_items or self._inventory_items)
        item_rows = [(item.code, item.name, item.price) for item in available_items]
        if not item_rows and not self._price_items_loading:
            self._price_items_error = self._price_items_error or "No hay items reales cargados para anadir a propuestas. Usa Buscar o Recargar."
        variations_host = tk.Frame(left, bg=BG)
        self._price_pick_table(left, "Items", item_rows, include_all=False, update_variations=True, variation_parent=variations_host)
        variations_host.pack(fill=tk.BOTH, expand=True)
        self._render_price_variations_picker(variations_host)

        panel = self._card(right)
        panel.pack(fill=tk.BOTH, expand=True)
        tk.Label(panel, text="Propuesta", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=16, pady=(16, 8))
        self._price_edit_name_var = getattr(self, "_price_edit_name_var", tk.StringVar(value=""))
        if self._selected_price_proposal and not self._price_edit_name_var.get().strip():
            self._price_edit_name_var.set(self._selected_price_proposal.name)
        name = tk.Entry(panel, textvariable=self._price_edit_name_var, bg=CARD, fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightcolor=INDIGO, highlightthickness=1)
        name.pack(fill=tk.X, padx=16, ipady=9, pady=(0, 12))

        list_outer = tk.Frame(panel, bg=CARD)
        list_outer.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 12))
        list_outer.rowconfigure(0, weight=1)
        list_outer.columnconfigure(0, weight=1)
        list_canvas = tk.Canvas(list_outer, bg=CARD, highlightthickness=0)
        list_scroll = ttk.Scrollbar(list_outer, orient=tk.VERTICAL, command=list_canvas.yview)
        list_canvas.configure(yscrollcommand=list_scroll.set)
        list_canvas.grid(row=0, column=0, sticky="nsew")
        list_scroll.grid(row=0, column=1, sticky="ns")
        list_host = tk.Frame(list_canvas, bg=CARD)
        list_window = list_canvas.create_window((0, 0), window=list_host, anchor="nw")
        list_host.bind("<Configure>", lambda _event: list_canvas.configure(scrollregion=list_canvas.bbox("all")))
        list_canvas.bind("<Configure>", lambda event: list_canvas.itemconfigure(list_window, width=event.width))
        if not self._price_edit_lines:
            tk.Label(list_host, text="Esta propuesta no tiene items.", bg=CARD, fg=MUTED).pack(anchor=tk.W, pady=10)
        else:
            for line in self._price_edit_lines:
                self._proposal_edit_line(list_host, line).pack(fill=tk.X, pady=5)

        footer = tk.Frame(panel, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        footer.pack(fill=tk.X, side=tk.BOTTOM, padx=16, pady=16)
        self._status_chip(footer, f"{len(self._price_edit_lines)} items en propuesta", "Info").pack(side=tk.LEFT, padx=12, pady=12)
        self._button(footer, "Guardar cambios", primary=True, command=self._save_price_edit).pack(side=tk.RIGHT, padx=(6, 12), pady=12)
        self._button(footer, "Cancelar", command=lambda: (self._price_reset_edit_state(), self._set_price_mode("saved"))).pack(side=tk.RIGHT, padx=6, pady=12)

    def _refresh_price_edit_items(self, parent: tk.Frame, query: str = "", allow_empty: bool = True) -> None:
        query = (query or "").strip()
        if not query and not allow_empty:
            self._price_items_error = "Introduce texto para buscar items reales."
            if self._current_key == "precios" and parent.winfo_exists():
                self._show_view("precios")
            return
        if self._cloud_session is None:
            self._price_items_error = "Inicia sesion Supabase para cargar items reales."
            self._price_items_loading = False
            if self._current_key == "precios" and parent.winfo_exists():
                self._show_view("precios")
            return
        self._price_search_query = query
        self._price_items_error = ""
        self._price_items_loading = True
        if self._current_key == "precios" and parent.winfo_exists():
            self._show_view("precios")

        def worker() -> None:
            try:
                if query:
                    server_rows = search_cloud_inventory_items(self._cloud_session, query, limit=100)
                    if self._inventory_query_is_code_like(query):
                        rows = server_rows
                    else:
                        all_rows = list_cloud_inventory_items(self._cloud_session, limit=500)
                        rows = self._merge_inventory_rows([server_rows, self._accent_insensitive_inventory_search(all_rows, query)])
                else:
                    rows = list_cloud_inventory_items(self._cloud_session, limit=150)
                items = [self._inventory_item_from_cloud_row(row) for row in rows]
                error = "Sin resultados reales para esa busqueda." if query and not items else ("No hay items reales visibles en Supabase." if not items else "")
                self.after(0, lambda: self._finish_price_edit_items(items, error))
            except Exception as exc:
                self.after(0, lambda exc=exc: self._finish_price_edit_items([], f"No se pudieron cargar items reales: {exc}"))

        threading.Thread(target=worker, daemon=True).start()

    def _finish_price_edit_items(self, items: list[InventoryItem], error: str) -> None:
        self._price_available_items = list(items)
        self._price_items_error = error
        self._price_items_loading = False
        self._price_line_sources = {
            item.code: self._price_source_from_inventory_item(item)
            for item in items
            if self._price_source_from_inventory_item(item)
        }
        if not self._price_edit_selected_code and items:
            self._price_edit_selected_code = items[0].code
        if self._current_key == "precios":
            self._show_view("precios")

    def _price_source_from_inventory_item(self, item: InventoryItem) -> dict[str, Any]:
        kind = (item.woo_item_kind or "").strip().lower()
        if kind not in {"product", "variation"}:
            raw_kind = (item.raw or {}).get("woo_item_kind") if item.raw else ""
            kind = str(raw_kind or kind or "").strip().lower()
        try:
            woo_id = int(str(item.woo_id or "").strip())
        except Exception:
            try:
                woo_id = int((item.raw or {}).get("woo_id")) if item.raw else 0
            except Exception:
                woo_id = 0
        if kind not in {"product", "variation"} or woo_id <= 0:
            return {}
        return {"item_kind": kind, "woo_id": woo_id, "item_id": item.code}

    def _price_variation_rows_for_selected(self) -> list[tuple[str, str, str]]:
        selected = self._price_edit_selected_code
        if not selected or self._cloud_session is None:
            return []
        item = None
        for candidate in (self._price_available_items or self._inventory_items):
            if candidate.code == selected:
                item = candidate
                break
        if item is None:
            return []
        source = self._price_source_from_inventory_item(item)
        if not source:
            return []
        parent_woo_id = None
        if source.get("item_kind") == "product":
            parent_woo_id = source.get("woo_id")
        else:
            try:
                parent_woo_id = int(str(item.woo_parent_id or "").strip())
            except Exception:
                parent_woo_id = None
        if not parent_woo_id:
            return []
        try:
            resp = (
                self._cloud_session.client.table("product_variations")
                .select("woo_id,parent_woo_id,parent_name,sku,price,regular_price,sale_price,attributes_label,status")
                .eq("parent_woo_id", int(parent_woo_id))
                .order("woo_id")
                .limit(80)
                .execute()
            )
            rows = getattr(resp, "data", None) or []
        except Exception:
            rows = []
        result: list[tuple[str, str, str]] = []
        for row in rows:
            woo_id = row.get("woo_id")
            if woo_id in (None, ""):
                continue
            label = row.get("attributes_label") or row.get("sku") or "variacion"
            parent_name = row.get("parent_name") or item.name
            name = f"{parent_name} - {label}"
            price = row.get("price") if row.get("price") not in (None, "") else row.get("regular_price") or row.get("sale_price") or "0"
            code = str(woo_id)
            self._price_line_sources[code] = {"item_kind": "variation", "woo_id": int(woo_id), "parent_woo_id": parent_woo_id}
            result.append((code, name, self._format_optional_price(price)))
        return result

    def _render_price_variations_picker(self, parent: tk.Misc) -> None:
        for child in parent.winfo_children():
            child.destroy()
        self._price_pick_table(
            parent,
            "Variaciones",
            self._price_variation_rows_for_selected(),
            include_all=True,
            update_variations=False,
        )

    def _price_pick_table(
        self,
        parent: tk.Misc,
        title: str,
        rows: list[tuple[str, str, str]],
        include_all: bool = False,
        update_variations: bool = False,
        variation_parent: tk.Misc | None = None,
    ) -> None:
        card = self._card(parent)
        card.pack(fill=tk.BOTH, expand=True, pady=(0, 12))
        tk.Label(card, text=title, bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=16, pady=(16, 8))
        tree = ttk.Treeview(card, columns=["ID", "Nombre", "Precio"], show="headings", height=4)
        widths = {"ID": 78, "Nombre": 360, "Precio": 78}
        stretches = {"ID": False, "Nombre": True, "Precio": False}
        for column in ["ID", "Nombre", "Precio"]:
            tree.heading(column, text=column, anchor=tk.CENTER)
            tree.column(column, width=widths[column], minwidth=widths[column], anchor=tk.CENTER, stretch=stretches[column])
        selected_iid = None
        for row in rows:
            iid = tree.insert("", tk.END, values=row)
            if row[0] == self._price_edit_selected_code:
                selected_iid = iid
                tree.selection_set(iid)
                tree.focus(iid)
        if selected_iid:
            tree.see(selected_iid)
        tree.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 10))

        if update_variations:
            def on_select(_event: object) -> None:
                selection = tree.selection()
                if not selection:
                    return
                values = tree.item(selection[0], "values")
                if not values:
                    return
                code = str(values[0])
                if code != self._price_edit_selected_code:
                    self._price_edit_selected_code = code
                    if variation_parent is not None:
                        self._render_price_variations_picker(variation_parent)
            tree.bind("<<TreeviewSelect>>", on_select)

        footer = tk.Frame(card, bg=CARD)
        footer.pack(fill=tk.X, padx=16, pady=(0, 16))

        tk.Label(footer, text="Subida %", bg=CARD, fg=MUTED, font=("Segoe UI", 8, "bold")).pack(side=tk.LEFT, padx=(0, 6))
        percent_entry = tk.Entry(footer, width=8, bg=CARD, fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightcolor=INDIGO, highlightthickness=1)
        percent_entry.pack(side=tk.LEFT, ipady=7, padx=(0, 12))
        tk.Label(footer, text="Valor", bg=CARD, fg=MUTED, font=("Segoe UI", 8, "bold")).pack(side=tk.LEFT, padx=(0, 6))
        exact_entry = tk.Entry(footer, width=10, bg=CARD, fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightcolor=INDIGO, highlightthickness=1)
        exact_entry.pack(side=tk.LEFT, ipady=7, padx=(0, 12))

        def selected_rows(all_rows: bool = False) -> list[tuple[str, str, str]]:
            if all_rows:
                return list(rows)
            selection = tree.selection()
            if not selection:
                return []
            return [tuple(str(v) for v in tree.item(selection[0], "values"))]  # type: ignore[list-item]

        spacer = tk.Frame(footer, bg=CARD)
        spacer.pack(side=tk.LEFT, fill=tk.X, expand=True)
        if include_all:
            self._button(
                footer,
                "Anadir Todas Variaciones",
                primary=True,
                command=lambda: self._price_add_rows_to_proposal(selected_rows(True), percent_entry.get(), exact_entry.get()),
            ).pack(side=tk.RIGHT, padx=(6, 0))
        self._button(
            footer,
            "Anadir",
            command=lambda: self._price_add_rows_to_proposal(selected_rows(False), percent_entry.get(), exact_entry.get()),
        ).pack(side=tk.RIGHT, padx=(6, 0))

    def _price_add_rows_to_proposal(self, rows: list[tuple[str, str, str]], percent_text: str, exact_text: str) -> None:
        if not rows:
            messagebox.showwarning("Cambio de Precios", "Selecciona un item o variacion para anadir.")
            return
        percent_text = (percent_text or "").strip()
        exact_text = (exact_text or "").strip()
        if bool(percent_text) == bool(exact_text):
            messagebox.showwarning("Cambio de Precios", "Usa subida en % o subida por valor, pero no ambas a la vez.")
            return
        new_lines = list(self._price_edit_lines)
        duplicates = [row[0] for row in rows if any(line.code == row[0] for line in new_lines)]
        if duplicates:
            answer = messagebox.askyesno(
                "Item ya incluido",
                "Ya existe en la propuesta: " + ", ".join(duplicates) + "\n\nQuieres sobrescribirlo?",
            )
            if not answer:
                self._price_edit_notice = "Anadido cancelado: item ya existia en la propuesta."
                self._show_view("precios")
                return
            new_lines = [line for line in new_lines if line.code not in set(duplicates)]
        try:
            for code, name, price in rows:
                old_price = self._price_parse_money(price)
                proposed = self._price_calculate_new_price(old_price, percent_text, exact_text)
                status, message = self._price_validate_proposed_price(old_price, proposed)
                if status == "Critical":
                    messagebox.showerror("Cambio de Precios", message)
                    return
                change = self._price_change_label(old_price, proposed)
                direction = "up" if proposed > old_price else "down" if proposed < old_price else "flat"
                if code not in self._price_line_sources:
                    for item in (self._price_available_items or self._inventory_items):
                        if item.code == code:
                            source = self._price_source_from_inventory_item(item)
                            if source:
                                self._price_line_sources[code] = source
                            break
                new_lines.append(ProposalLine(code, name, f"{old_price:.2f}", f"{proposed:.2f}", change, direction))
            self._price_edit_lines = new_lines
            self._price_edit_notice = f"{len(rows)} item(s) anadidos a la propuesta."
            self._show_view("precios")
        except ValueError as exc:
            messagebox.showerror("Cambio de Precios", str(exc))

    def _price_parse_money(self, value: str) -> float:
        try:
            return float(str(value or "0").replace("€", "").replace(",", ".").strip())
        except Exception as exc:
            raise ValueError("Precio actual no numerico.") from exc

    def _price_calculate_new_price(self, old_price: float, percent_text: str, exact_text: str) -> float:
        if percent_text:
            percent = float(percent_text.replace(",", "."))
            return round(old_price * (1 + percent / 100), 2)
        exact = float(exact_text.replace(",", "."))
        return round(old_price + exact, 2)

    def _price_validate_proposed_price(self, old_price: float, proposed: float) -> tuple[str, str]:
        if proposed <= 0:
            return "Critical", "Precio propuesto 0 o negativo. Operacion bloqueada."
        if old_price <= 0:
            return "Warning", "Precio anterior 0: se permite asignar precio porque puede ser item nuevo."
        try:
            settings = load_settings()
            block_percent = float(getattr(settings, "price_drop_block_percent", 30.0))
        except Exception:
            block_percent = 30.0
        if proposed < old_price:
            drop_percent = ((old_price - proposed) / old_price) * 100
            if drop_percent >= block_percent:
                return "Critical", f"Bajada del {drop_percent:.2f}% supera el bloqueo configurado ({block_percent:.2f}%)."
        return "OK", "Precio valido."

    def _price_change_label(self, old_price: float, new_price: float) -> str:
        if old_price <= 0:
            return "Nuevo"
        delta = new_price - old_price
        pct = (delta / old_price) * 100
        return f"{delta:+.2f} ({pct:+.2f}%)"

    def _price_select_line_for_edit(self, line: ProposalLine) -> None:
        self._price_edit_selected_code = line.code
        self._price_search_query = line.code
        self._price_available_items = []
        self._price_items_error = ""
        self._price_items_loading = False
        self._price_edit_notice = f"Editando {line.code}: se buscará y seleccionará en la tabla. Si lo añades de nuevo, podrás sobrescribirlo."
        self._show_view("precios")

    def _price_delete_line(self, line: ProposalLine) -> None:
        if not messagebox.askyesno("Borrar item", f"Quitar {line.code} de la propuesta?"):
            return
        self._price_edit_initialized = True
        self._price_edit_lines = [item for item in self._price_edit_lines if item.code != line.code]
        self._price_line_sources.pop(line.code, None)
        self._price_edit_notice = f"{line.code} eliminado de la propuesta."
        self._show_view("precios")

    def _save_price_edit(self) -> None:
        proposal_name = ""
        if hasattr(self, "_price_edit_name_var"):
            proposal_name = self._price_edit_name_var.get().strip()
        if not proposal_name:
            messagebox.showwarning("Cambio de Precios", "Pon un nombre a la propuesta antes de guardar.")
            return
        if not self._price_edit_lines:
            messagebox.showwarning("Cambio de Precios", "No hay items en la propuesta.")
            return
        if self._cloud_session is None:
            messagebox.showwarning("Cambio de Precios", "Inicia sesion Supabase para guardar propuestas reales.")
            return
        missing = [line.code for line in self._price_edit_lines if not self._price_line_sources.get(line.code)]
        if missing:
            messagebox.showwarning(
                "Cambio de Precios",
                "Hay items sin vínculo Woo válido para guardar propuesta real:\n" + ", ".join(missing),
            )
            return
        if not messagebox.askyesno(
            "Guardar propuesta",
            "Se guardaran/actualizaran propuestas reales pendientes en Supabase.\n\nWooCommerce no se toca hasta aceptar la propuesta. Continuar?",
        ):
            return

        lines = list(self._price_edit_lines)
        self._price_edit_notice = "Guardando propuesta real en Supabase..."
        self._show_view("precios")

        def worker() -> None:
            try:
                saved: list[str] = []
                for line in lines:
                    source = self._price_line_sources.get(line.code) or {}
                    new_price = self._price_parse_money(line.new_price)
                    result = create_real_price_proposal(
                        self._cloud_session,
                        str(source.get("item_kind") or ""),
                        int(source.get("woo_id")),
                        float(new_price),
                        notes="Creada desde UI-ERP Cambio de Precios.",
                        acknowledge_price_warning=True,
                    )
                    proposal = result.get("proposal") or {}
                    proposal_id = proposal.get("id")
                    if proposal_id:
                        try:
                            source_row = proposal.get("source_row") if isinstance(proposal.get("source_row"), dict) else {}
                            source_row.update({
                                "ui_proposal_name": proposal_name,
                                "ui_line_code": line.code,
                                "ui_line_name": line.name,
                            })
                            self._cloud_session.client.table("price_change_proposals").update({
                                "source_row": source_row,
                                "notes": f"Propuesta UI: {proposal_name}",
                            }).eq("id", proposal_id).execute()
                        except Exception:
                            pass
                    saved.append(str(proposal_id or source.get("woo_id") or line.code))
                self.after(0, lambda: self._finish_price_edit_saved(saved))
            except Exception as exc:
                self.after(0, lambda exc=exc: self._finish_price_edit_save_error(str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def _finish_price_edit_saved(self, saved: list[str]) -> None:
        messagebox.showinfo("Cambio de Precios", f"Propuesta guardada correctamente. Items guardados: {len(saved)}")
        self._price_reset_edit_state()
        self._price_loaded_once = False
        self._set_price_mode("saved")

    def _finish_price_edit_save_error(self, error: str) -> None:
        self._price_edit_notice = f"No se pudo guardar la propuesta real: {error}"
        if self._current_key == "precios":
            self._show_view("precios")

    def _proposal_edit_line(self, parent: tk.Misc, line: ProposalLine) -> tk.Frame:
        status = "OK" if line.direction == "up" else "Critical" if line.direction in ("down", "critical") else "Info"
        frame = tk.Frame(parent, bg=SOFT, highlightbackground=LINE, highlightthickness=1)
        frame.columnconfigure(0, weight=1)

        top = tk.Frame(frame, bg=SOFT)
        top.grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 6))
        top.columnconfigure(1, weight=1)
        tk.Label(top, text=line.code, bg=SOFT, fg=MUTED, font=("Segoe UI", 8, "bold"), width=12, anchor=tk.W).grid(row=0, column=0, sticky="w")
        tk.Label(
            top,
            text=line.name,
            bg=SOFT,
            fg=TEXT,
            font=("Segoe UI", 10, "bold"),
            anchor=tk.W,
            justify=tk.LEFT,
            wraplength=260,
        ).grid(row=0, column=1, sticky="ew", padx=8)
        actions = tk.Frame(top, bg=SOFT)
        actions.grid(row=0, column=2, sticky="e")
        self._button(actions, "Modificar", command=lambda: self._price_select_line_for_edit(line)).pack(side=tk.LEFT, padx=(0, 6))
        self._button(actions, "Borrar", command=lambda: self._price_delete_line(line)).pack(side=tk.LEFT)

        bottom = tk.Frame(frame, bg=SOFT)
        bottom.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 10))
        bottom.columnconfigure(2, weight=1)
        tk.Label(bottom, text=f"Precio antiguo: {line.old_price}", bg=SOFT, fg=TEXT, font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w")
        tk.Label(bottom, text=f"Precio nuevo: {line.new_price}", bg=SOFT, fg=TEXT, font=("Segoe UI", 9, "bold")).grid(row=0, column=1, sticky="w", padx=(18, 0))
        self._status_chip(bottom, line.change, status).grid(row=0, column=3, sticky="e")
        return frame

    def _load_supplier_orders_from_cloud(self) -> list[SupplierOrder]:
        """Carga pedidos reales desde Supabase para Pedidos.

        No usa datos mock. Si Supabase no devuelve filas, se muestra estado vacío.
        """
        if self._cloud_session is None:
            return []
        rows = list_cloud_supplier_orders(self._cloud_session, limit=100)
        orders: list[SupplierOrder] = []
        for row in rows:
            try:
                order_id = str(row.get("order_id") or row.get("local_order_id") or "")
                item_rows = list_cloud_supplier_order_items(self._cloud_session, order_id) if order_id else []
                items = tuple(self._supplier_order_item_from_cloud_row(item_row) for item_row in item_rows)
                summary = summarize_order_items(item_rows)
                total_m3 = self._format_optional_decimal(summary.get("total_m3"), default="-")
                total_cost = self._format_optional_price(row.get("total_cost") or summary.get("total_cost"))
                status = str(row.get("status") or "Borrador")
                total_items = int(float(row.get("total_items") or len(items) or 0))
                orders.append(
                    SupplierOrder(
                        order_id=cloud_order_display_name(row),
                        provider=str(row.get("provider") or "-"),
                        date=format_order_date(row),
                        items_count=total_items,
                        total_m3=total_m3,
                        status=status,
                        total_cost=total_cost,
                        notes=str(row.get("notes") or ""),
                        items=items,
                        raw=row,
                    )
                )
            except Exception:
                continue
        return orders

    def _supplier_order_item_from_cloud_row(self, row: dict[str, Any]) -> OrderItem:
        source = row.get("source_row") if isinstance(row.get("source_row"), dict) else {}
        code = str(row.get("item_code") or row.get("item_id") or row.get("local_id") or "")
        name = str(row.get("item_name") or source.get("name") or source.get("product_name") or "")
        quantity = int(float(row.get("quantity_ordered") or 0))
        m3_value = source.get("total_m3") or source.get("m3_total") or source.get("cubic_meters_total") or source.get("cubic_meters")
        final_cost_value = row.get("line_cost") or source.get("final_cost") or source.get("coste_final")
        status = str(source.get("status") or source.get("ui_status") or ("OK" if final_cost_value not in (None, "") else "Warning"))
        return OrderItem(
            code=code,
            name=name,
            quantity=quantity,
            m3=self._format_optional_decimal(m3_value, default="Pendiente"),
            final_cost=self._format_optional_price(final_cost_value) if final_cost_value not in (None, "") else "Pendiente",
            status=status,
            raw=row,
        )

    def _ensure_supplier_orders_loaded(self) -> None:
        if self._orders_loaded_once or self._orders_loading:
            return
        self._orders_loading = True
        self._orders_error = ""
        try:
            self._supplier_orders = self._load_supplier_orders_from_cloud()
        except Exception as exc:
            self._supplier_orders = []
            self._orders_error = str(exc)
        finally:
            self._orders_loaded_once = True
            self._orders_loading = False

    def _orders_for_ui(self) -> list[SupplierOrder]:
        if self._cloud_session is not None:
            self._ensure_supplier_orders_loaded()
            return list(self._supplier_orders)
        return list(SUPPLIER_ORDERS)

    def _refresh_supplier_orders(self) -> None:
        self._orders_loaded_once = False
        self._supplier_orders = []
        self._selected_supplier_order = None
        self._show_view("calcular")

    def _format_optional_decimal(self, value: Any, default: str = "-") -> str:
        if value in (None, ""):
            return default
        try:
            number = float(str(value).replace(",", "."))
            if number.is_integer():
                return str(int(number))
            return f"{number:.2f}".replace(".", ",")
        except Exception:
            return str(value)

    def _build_order_calc(self, parent: tk.Frame) -> None:
        self._page_header(parent, "Operaciones", "Pedidos", "Proveedores, pedidos en marcha, calculo, recepcion y exportacion.")
        providers_card = self._card(parent)
        providers_card.pack(fill=tk.X, pady=(0, 14))
        providers_head = tk.Frame(providers_card, bg=CARD)
        providers_head.pack(fill=tk.X, padx=16, pady=(14, 8))
        providers_head.columnconfigure(0, weight=1)
        tk.Label(providers_head, text="Proveedores", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w")
        providers = tk.Frame(providers_card, bg=CARD)
        providers.pack(fill=tk.X, padx=16, pady=(0, 14))
        for i, provider in enumerate(["Ekomat", "Pascal", "Heimei", "Cipta"]):
            providers.columnconfigure(i, weight=1)
            self._order_provider_button(providers, provider).grid(
                row=0,
                column=i,
                sticky="ew",
                padx=(0 if i == 0 else 8, 0),
            )

        body = tk.Frame(parent, bg=BG)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=4, minsize=700)
        body.columnconfigure(1, weight=2, minsize=390)
        body.rowconfigure(0, weight=1)

        orders_card = self._card(body)
        orders_card.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        orders_card.rowconfigure(1, weight=1)
        orders_card.columnconfigure(0, weight=1)
        orders_head = tk.Frame(orders_card, bg=CARD)
        orders_head.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 10))
        orders_head.columnconfigure(0, weight=1)
        tk.Label(orders_head, text="Pedidos en marcha", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w")
        self._button(orders_head, "Actualizar", command=self._refresh_supplier_orders).grid(row=0, column=1, sticky="e", padx=(0, 8))
        orders = self._orders_for_ui()
        self._status_chip(orders_head, f"{len(orders)} abiertos", "Info").grid(row=0, column=2, sticky="e")

        right = tk.Frame(body, bg=BG)
        right.grid(row=0, column=1, sticky="nsew")
        if orders:
            if self._selected_supplier_order not in orders:
                self._selected_supplier_order = orders[0]
            self._order_table(orders_card, right, orders)
            if self._selected_supplier_order is not None:
                self._render_order_detail(right, self._selected_supplier_order)
        else:
            self._empty_orders_state(orders_card, right)

    def _order_provider_button(self, parent: tk.Misc, provider: str) -> tk.Frame:
        frame = tk.Frame(parent, bg=SOFT, highlightbackground=LINE, highlightthickness=1)
        top = tk.Frame(frame, bg=SOFT)
        top.pack(fill=tk.X, padx=10, pady=(10, 8))
        tk.Label(top, text=provider[:1], bg=INDIGO_SOFT, fg=INDIGO, font=("Segoe UI", 12, "bold"), width=2).pack(side=tk.LEFT)
        tk.Label(top, text=provider, bg=SOFT, fg=TEXT, font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT, padx=10)
        self._button(frame, "Calcular nuevo pedido", primary=True, command=lambda provider=provider: self._open_order_calc_flow(provider)).pack(
            fill=tk.X,
            padx=10,
            pady=(0, 10),
        )
        return frame

    def _order_table(self, parent: tk.Frame, detail_parent: tk.Frame, orders: list[SupplierOrder]) -> None:
        columns = ["Pedido", "Proveedor", "Fecha", "Items", "Total M3", "Total", "Estado"]
        tree = ttk.Treeview(parent, columns=columns, show="headings", height=12)
        widths = {"Pedido": 185, "Proveedor": 100, "Fecha": 105, "Items": 75, "Total M3": 85, "Total": 115, "Estado": 100}
        for column in columns:
            tree.heading(column, text=column, anchor=tk.CENTER)
            tree.column(column, width=widths[column], anchor=tk.CENTER)
        order_by_iid: dict[str, SupplierOrder] = {}
        for order in self._orders_sorted_by_date(orders):
            iid = tree.insert("", tk.END, values=(self._order_display_name(order), order.provider, order.date, order.items_count, order.total_m3, order.total_cost, order.status))
            order_by_iid[iid] = order
            if self._selected_supplier_order is not None and order == self._selected_supplier_order:
                tree.selection_set(iid)
                tree.focus(iid)
        yscroll = tk.Scrollbar(parent, orient=tk.VERTICAL, command=tree.yview)
        xscroll = tk.Scrollbar(parent, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=1, column=0, sticky="nsew", padx=(16, 0), pady=(0, 0))
        yscroll.grid(row=1, column=1, sticky="ns", padx=(0, 16), pady=(0, 0))
        xscroll.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 16))

        def on_select(_event: object | None = None) -> None:
            selection = tree.selection()
            if not selection:
                return
            self._selected_supplier_order = order_by_iid[selection[0]]
            self._render_order_detail(detail_parent, self._selected_supplier_order)

        tree.bind("<<TreeviewSelect>>", on_select)

    def _empty_orders_state(self, table_parent: tk.Frame, detail_parent: tk.Frame) -> None:
        message = self._orders_error or "No hay pedidos en marcha."
        box = tk.Frame(table_parent, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        box.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 16), columnspan=2)
        tk.Label(box, text=message, bg=CARD, fg=MUTED, font=("Segoe UI", 11), pady=28).pack(fill=tk.BOTH, expand=True)
        detail = self._card(detail_parent)
        detail.pack(fill=tk.BOTH, expand=True)
        tk.Label(detail, text="Selecciona o crea un pedido para ver sus detalles.", bg=CARD, fg=MUTED, font=("Segoe UI", 11), wraplength=330, justify=tk.LEFT).pack(padx=20, pady=24)

    def _order_display_name(self, order: SupplierOrder) -> str:
        raw = str(order.order_id or "").strip()
        return raw or f"Pedido {order.provider}"

    def _orders_sorted_by_date(self, orders: list[SupplierOrder]) -> list[SupplierOrder]:
        def key(order: SupplierOrder) -> tuple[int, int, int, str]:
            parts = str(order.date or "").split("/")
            if len(parts) == 3:
                try:
                    day, month, year = (int(part) for part in parts)
                    return (year, month, day, order.order_id)
                except Exception:
                    pass
            return (0, 0, 0, order.order_id)

        return sorted(orders, key=key, reverse=True)

    def _order_missing_reasons(self, order: SupplierOrder) -> list[str]:
        reasons: list[str] = []
        for item in order.items:
            item_reasons = self._order_item_missing_reasons(item)
            for reason in item_reasons:
                reasons.append(f"{item.code}: {reason}")
        return reasons


    def _fill_supplier_prices_for_order_items(self, provider: str, items: tuple[OrderItem, ...]) -> tuple[OrderItem, ...]:
        """Fill supplier price and inventory costing fields from Supabase.

        Pricing rule:
        - Pascal uses inventory_items.pascal_price.
        - Every other provider uses inventory_items.primary_supplier_price.

        This does not write to WooCommerce, stock or inventory. It only enriches
        the in-memory order line before painting/calculating.
        """
        if self._cloud_session is None:
            return items
        enriched: list[OrderItem] = []
        for item in items:
            source = item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {}
            supplier_price = get_supplier_price(self._cloud_session, item.code, provider)
            if not supplier_price:
                enriched.append(item)
                continue
            inventory_row = supplier_price.get("item") if isinstance(supplier_price.get("item"), dict) else {}
            price = self._money_float(supplier_price.get("price"))
            updated_source = dict(source)
            current_price = self._money_float(updated_source.get("precio_proveedor") or updated_source.get("precio_excel") or updated_source.get("precio"))
            manual_price = bool(updated_source.get("ui_manual_supplier_price") or updated_source.get("ui_completed_from_order_editor")) and current_price > 0
            if price > 0 and current_price <= 0:
                updated_source.update({"precio_proveedor": price, "precio_excel": price})
                updated_source.update({
                    "supplier_price_source": supplier_price.get("source"),
                    "supplier_price_column": supplier_price.get("column"),
                    "supplier_price_provider": supplier_price.get("supplier"),
                    "supplier_price_matched_by": supplier_price.get("matched_by"),
                    "supplier_price_item_id": supplier_price.get("item_id"),
                    "ui_supplier_price_filled": True,
                })
            elif manual_price:
                updated_source.update({
                    "supplier_price_source": updated_source.get("supplier_price_source") or "manual_order_editor",
                    "supplier_price_column": updated_source.get("supplier_price_column") or "manual",
                    "supplier_price_provider": updated_source.get("supplier_price_provider") or provider,
                    "supplier_price_matched_by": updated_source.get("supplier_price_matched_by") or "manual",
                    "ui_supplier_price_filled": False,
                    "ui_supplier_price_preserved_manual": True,
                })
            else:
                # Si el Excel ya trae precio proveedor, lo respetamos. Solo guardamos
                # metadatos del match sin pisar el precio.
                updated_source.update({
                    "supplier_price_source": updated_source.get("supplier_price_source") or supplier_price.get("source"),
                    "supplier_price_column": updated_source.get("supplier_price_column") or supplier_price.get("column"),
                    "supplier_price_provider": updated_source.get("supplier_price_provider") or supplier_price.get("supplier"),
                    "supplier_price_matched_by": updated_source.get("supplier_price_matched_by") or supplier_price.get("matched_by"),
                    "supplier_price_item_id": updated_source.get("supplier_price_item_id") or supplier_price.get("item_id"),
                    "ui_supplier_price_filled": False,
                })
            updated_source.update(
                {
                    "inventory_name": inventory_row.get("name"),
                    "inventory_m3": inventory_row.get("cubic_meters"),
                    "inventory_rotation_c": inventory_row.get("rotation_c"),
                    "inventory_packages": inventory_row.get("packages"),
                    "inventory_store_stock": inventory_row.get("store_stock"),
                    "inventory_warehouse_stock": inventory_row.get("warehouse_stock"),
                    "inventory_stock_total": self._money_float(inventory_row.get("store_stock"), 0.0) + self._money_float(inventory_row.get("warehouse_stock"), 0.0),
                    "inventory_weighted_average_cost": inventory_row.get("weighted_average_cost"),
                    "inventory_order_calculated_price": inventory_row.get("order_calculated_price"),
                }
            )
            # Si el Excel/PDF no trae M3, aprovechamos el M3 del inventario.
            if self._money_float(updated_source.get("m3_und")) <= 0 and self._money_float(updated_source.get("m3_total")) <= 0:
                inv_m3 = self._money_float(inventory_row.get("cubic_meters"))
                qty = self._parse_units_value(updated_source.get("unidades") or updated_source.get("quantity_ordered") or item.quantity)
                if inv_m3 > 0:
                    updated_source["m3_und"] = inv_m3
                    updated_source["m3_total"] = round(inv_m3 * qty, 6) if qty else inv_m3
            enriched.append(
                OrderItem(
                    code=item.code,
                    name=item.name,
                    quantity=item.quantity,
                    m3=item.m3,
                    final_cost=item.final_cost,
                    status=item.status,
                    raw={"source_row": updated_source},
                )
            )
        return tuple(enriched)

    def _order_item_missing_reasons(self, item: OrderItem) -> list[str]:
        """Return visible blocking reasons for a supplier order calculation line.

        This is intentionally stricter than the final save validation: the goal is
        target-locking bad rows as soon as an Excel/PDF is loaded, before the user
        double-clicks line by line.
        """
        reasons: list[str] = []
        source = item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {}

        def _text(value: Any) -> str:
            return str(value if value is not None else "").strip()

        def _is_empty_like(value: Any) -> bool:
            text = _text(value).lower()
            return text in {"", "-", "none", "null", "nan", "pendiente", "no esta", "no está", "n/a"}

        def _positive_number(value: Any) -> bool:
            try:
                return self._money_float(value) > 0
            except Exception:
                return False

        code = _text(item.code)
        name = _text(item.name)
        qty_raw = source.get("unidades", source.get("quantity_ordered", item.quantity))
        m3_unit_raw = source.get("m3_und", source.get("m3_unit", None))
        m3_total_raw = source.get("m3_total", item.m3)
        price_raw = source.get("precio_proveedor", source.get("precio_excel", source.get("precio", None)))
        cost = _text(item.final_cost).lower()
        status = _text(item.status).lower()

        if _is_empty_like(code):
            reasons.append("falta ID / referencia")
        if _is_empty_like(name):
            reasons.append("falta nombre")
        if not _positive_number(qty_raw):
            reasons.append("falta cantidad")
        # Para calcular necesitamos M3 por unidad o un M3 total válido.
        if not _positive_number(m3_unit_raw) and not _positive_number(m3_total_raw):
            reasons.append("falta M3")
        # Antes del cálculo necesitamos precio de proveedor/base; después del cálculo final_cost ya queda relleno.
        if not _positive_number(price_raw) and cost in {"", "pendiente", "bloqueado", "none", "no esta", "no está"}:
            reasons.append("falta precio proveedor")
        if cost in {"bloqueado"}:
            reasons.append("coste bloqueado")
        if status in {"error", "critical", "bloqueado", "blocked"}:
            reasons.append(f"estado {item.status}")

        # Si el parser marcó motivos en source_row, también los mostramos.
        raw_reasons = source.get("calculation_reasons") or source.get("ui_reasons") or source.get("warnings")
        if isinstance(raw_reasons, list):
            for reason in raw_reasons:
                reason_text = _text(reason)
                if reason_text and reason_text not in reasons:
                    reasons.append(reason_text)
        return reasons

    def _render_order_detail(self, parent: tk.Frame, order: SupplierOrder) -> None:
        for child in parent.winfo_children():
            child.destroy()
        detail = self._card(parent)
        detail.pack(fill=tk.BOTH, expand=True)
        detail.rowconfigure(1, weight=1)
        detail.columnconfigure(0, weight=1)
        head = tk.Frame(detail, bg=CARD)
        head.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 10))
        head.columnconfigure(0, weight=1)
        title = tk.Frame(head, bg=CARD)
        title.grid(row=0, column=0, sticky="ew")
        tk.Label(title, text=order.order_id, bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold"), width=18, anchor=tk.W).pack(anchor=tk.W)
        tk.Label(title, text=f"{order.provider} - {order.date}", bg=CARD, fg=MUTED, width=28, anchor=tk.W).pack(anchor=tk.W)
        self._button(head, "Detalles", command=lambda: self._open_order_detail_window(order)).grid(row=0, column=1, sticky="e")

        content = tk.Frame(detail, bg=CARD)
        content.grid(row=1, column=0, sticky="nsew", padx=16)
        summary = tk.Frame(content, bg=CARD)
        summary.pack(fill=tk.X, pady=(0, 12))
        for i, (label, value, status) in enumerate(
            [("Items", order.items_count, "Info"), ("Total M3", order.total_m3, "OK"), ("Coste final", order.total_cost, "OK" if order.total_cost != "Bloqueado" else "Error"), ("Estado", order.status, order.status)]
        ):
            summary.columnconfigure(i % 2, weight=1)
            self._metric(summary, label, str(value), status).grid(row=i // 2, column=i % 2, sticky="ew", padx=(0 if i % 2 == 0 else 6, 0), pady=(0, 6))
        for item in order.items:
            self._order_item_row(content, item).pack(fill=tk.X, pady=4)
        tk.Label(content, text=order.notes, bg=INDIGO_SOFT, fg="#4338CA", wraplength=330, justify=tk.LEFT, padx=12, pady=10).pack(fill=tk.X, pady=(12, 0))

        missing = self._order_missing_reasons(order)
        missing_box = tk.Frame(content, bg=CARD)
        missing_box.pack(fill=tk.X, pady=(12, 0))
        tk.Label(missing_box, text="Elementos pendientes para calcular", bg=CARD, fg=TEXT, font=("Segoe UI", 10, "bold"), anchor=tk.W).pack(fill=tk.X)
        missing_text = "\n".join(f"• {reason}" for reason in missing[:5]) if missing else "Sin elementos pendientes detectados."
        tk.Label(
            missing_box,
            text=missing_text,
            bg=ROSE_SOFT if missing else GREEN_SOFT,
            fg=ROSE if missing else GREEN,
            wraplength=330,
            justify=tk.LEFT,
            padx=12,
            pady=10,
        ).pack(fill=tk.X, pady=(6, 0))

        footer = tk.Frame(detail, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        footer.grid(row=2, column=0, sticky="ew", padx=16, pady=16)
        top_actions = tk.Frame(footer, bg=CARD)
        top_actions.pack(fill=tk.X, padx=12, pady=(12, 7))
        self._button(top_actions, "Modificar", command=lambda: self._open_order_calc_flow(order.provider, order=order)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        self._button(top_actions, "Recibido", primary=True, command=lambda: self._open_receive_modal(order)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0))
        row = tk.Frame(footer, bg=CARD)
        row.pack(fill=tk.X, padx=12, pady=(0, 12))
        self._button(row, "Borrar pedido", command=lambda: self._cancel_supplier_order_from_ui(order)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        self._button(row, "Exportar", primary=True, command=lambda: self._export_supplier_order_audit_excel(order)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0))

    def _order_item_row(self, parent: tk.Misc, item: OrderItem) -> tk.Frame:
        frame = tk.Frame(parent, bg=SOFT, highlightbackground=LINE, highlightthickness=1)
        text = tk.Frame(frame, bg=SOFT)
        text.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=12, pady=9)
        tk.Label(
            text,
            text=f"{item.code} - {item.name}",
            bg=SOFT,
            fg=TEXT,
            font=("Segoe UI", 9, "bold"),
            anchor=tk.W,
            wraplength=250,
            justify=tk.LEFT,
        ).pack(anchor=tk.W)
        tk.Label(text, text=f"Cantidad: {item.quantity} ud - M3: {item.m3}", bg=SOFT, fg=MUTED, font=("Segoe UI", 9)).pack(anchor=tk.W)
        tk.Label(frame, text=item.final_cost, bg=SOFT, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(side=tk.RIGHT, padx=(8, 12))
        self._status_chip(frame, item.status, item.status).pack(side=tk.RIGHT)
        return frame

    def _normalize_order_header(self, value: Any) -> str:
        text = self._normalize_search_text(value)
        return re.sub(r"[^a-z0-9]+", "", text)

    def _parse_units_value(self, value: Any) -> int:
        if value in (None, ""):
            return 0
        text = str(value).strip().replace(",", ".")
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            return 0
        try:
            return max(0, int(float(match.group(0))))
        except Exception:
            return 0

    def _safe_order_float(self, value: Any) -> float:
        try:
            if value in (None, ""):
                return 0.0
            return float(str(value).replace(",", "."))
        except Exception:
            return 0.0

    def _m3_from_measure_text(self, measure: str) -> float:
        text = str(measure or "").lower().replace(",", ".")
        numbers = re.findall(r"\d+(?:\.\d+)?", text)
        if len(numbers) >= 3:
            try:
                width, length, height = [float(x) for x in numbers[:3]]
                if width > 0 and length > 0 and height > 0:
                    return round((width * length * height) / 1_000_000, 6)
            except Exception:
                return 0.0
        return 0.0

    def _detect_supplier_order_header(self, worksheet: Any) -> tuple[int | None, dict[str, int]]:
        for row_idx in range(1, min(int(getattr(worksheet, "max_row", 0) or 0), 40) + 1):
            values = [worksheet.cell(row_idx, col).value for col in range(1, min(int(getattr(worksheet, "max_column", 0) or 0), 18) + 1)]
            normalized = [self._normalize_order_header(v) for v in values]
            has_ref = any(k in normalized for k in ("referencia", "ref", "codigoarticuloitemcode", "codigoarticulo", "codigo", "itemcode"))
            has_qty = any(k in normalized for k in ("und", "unds", "unidades", "cantidadquantity", "cantidad", "quantity"))
            if has_ref and has_qty:
                return row_idx, {name: index + 1 for index, name in enumerate(normalized) if name}
        return None, {}

    def _supplier_order_col(self, header_map: dict[str, int], *names: str) -> int | None:
        for name in names:
            key = self._normalize_order_header(name)
            if key in header_map:
                return header_map[key]
        return None

    def _order_download_count_exception_codes(self) -> set[str]:
        """Referencias pequeñas/ambiguas que sí cuentan para descarga."""
        return {
            "0727007",
            "0730009",
            "1242001",
            "1242002",
            "1243001",
            "1244001",
            "1245001",
            "1249001",
        }

    def _order_download_excluded_keywords(self) -> tuple[str, ...]:
        """Palabras que normalmente NO cuentan para el reparto de descarga."""
        return ("funda", "cover", "topper", "pillow", "pillows", "almohada", "almohadas")

    def _order_normalize_yes_no(self, value: Any, default: bool | None = None) -> bool | None:
        text = self._normalize_search_text(value)
        if not text:
            return default
        if text in {"si", "sí", "s", "yes", "y", "true", "1", "cuenta"}:
            return True
        if text in {"no", "n", "false", "0", "nocuenta", "no cuenta", "no_cuenta"}:
            return False
        return default

    def _order_line_counts_for_download_auto(self, code: Any, name: Any = "", source: dict[str, Any] | None = None) -> bool:
        source = source or {}
        ref = str(code or source.get("referencia") or source.get("codigo") or "").strip()
        base_ref = ref.split("-")[0].strip()
        if ref in self._order_download_count_exception_codes() or base_ref in self._order_download_count_exception_codes():
            return True
        haystack = self._normalize_search_text(" ".join(str(x or "") for x in (
            ref,
            name,
            source.get("producto"),
            source.get("descripcion"),
            source.get("composicion"),
            source.get("medida"),
            source.get("inventory_name"),
        )))
        return not any(keyword in haystack for keyword in self._order_download_excluded_keywords())

    def _order_line_counts_for_download(self, item_or_source: Any, source: dict[str, Any] | None = None) -> bool:
        if isinstance(item_or_source, OrderItem):
            item = item_or_source
            src = source if isinstance(source, dict) else (item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {})
            manual = self._order_normalize_yes_no(
                src.get("cuenta_para_descarga", src.get("cuenta_reparto_descarga", src.get("cuenta_pedido"))),
                default=None,
            )
            if manual is not None:
                return manual
            return self._order_line_counts_for_download_auto(item.code, item.name, src)
        src = item_or_source if isinstance(item_or_source, dict) else (source or {})
        manual = self._order_normalize_yes_no(
            src.get("cuenta_para_descarga", src.get("cuenta_reparto_descarga", src.get("cuenta_pedido"))),
            default=None,
        )
        if manual is not None:
            return manual
        return self._order_line_counts_for_download_auto(src.get("codigo") or src.get("referencia"), src.get("producto") or src.get("descripcion") or src.get("composicion"), src)

    def _order_line_counts_for_download_label(self, item_or_source: Any, source: dict[str, Any] | None = None) -> str:
        return "Sí" if self._order_line_counts_for_download(item_or_source, source) else "No"

    def _order_line_download_reason(self, item: OrderItem, source: dict[str, Any] | None = None) -> str:
        src = source if isinstance(source, dict) else (item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {})
        manual = self._order_normalize_yes_no(src.get("cuenta_para_descarga", src.get("cuenta_reparto_descarga", src.get("cuenta_pedido"))), default=None)
        auto = self._order_line_counts_for_download_auto(item.code, item.name, src)
        if manual is not None:
            return f"Selección manual: {'Sí' if manual else 'No'}. Regla automática sugerida: {'Sí' if auto else 'No'}."
        ref = str(item.code or "").strip()
        base_ref = ref.split("-")[0].strip()
        if ref in self._order_download_count_exception_codes() or base_ref in self._order_download_count_exception_codes():
            return "Sí por excepción definida: producto grande que cuenta para descarga."
        haystack = self._normalize_search_text(" ".join(str(x or "") for x in (item.name, src.get("producto"), src.get("descripcion"), src.get("composicion"))))
        if any(keyword in haystack for keyword in self._order_download_excluded_keywords()):
            return "No por regla automática: funda/cover/topper/pillow/almohada no reparte descarga."
        return "Sí por regla automática."

    def _order_item_from_loaded_line(self, line: dict[str, Any]) -> OrderItem:
        qty = self._parse_units_value(line.get("unidades") or line.get("quantity") or line.get("quantity_ordered"))
        m3_total = self._safe_order_float(line.get("m3_total"))
        m3_und = self._safe_order_float(line.get("m3_und"))
        m3_text = self._format_optional_decimal(m3_total or m3_und, default="Pendiente")
        code = str(line.get("codigo") or line.get("referencia") or "-").strip()
        name = str(line.get("producto") or line.get("composicion") or line.get("descripcion") or "-").strip()
        source_row = dict(line or {})
        cuenta_descarga = self._order_line_counts_for_download(source_row)
        source_row.update(
            {
                "cuenta_para_descarga": "Sí" if cuenta_descarga else "No",
                "cuenta_reparto_descarga": cuenta_descarga,
                "cuenta_pedido": "Sí" if cuenta_descarga else "No",
                "ui_download_count_rule": "auto_on_load",
            }
        )
        missing = []
        if not code or code == "-":
            missing.append("ID")
        if qty <= 0:
            missing.append("unidades")
        if m3_total <= 0 and m3_und <= 0:
            missing.append("M3")
        status = "Warning" if missing else "OK"
        return OrderItem(
            code=code,
            name=name,
            quantity=qty,
            m3=m3_text,
            final_cost="Pendiente",
            status=status,
            raw={"source_row": source_row},
        )

    def _order_item_source_row(self, item: OrderItem, line: dict[str, Any] | None = None) -> dict[str, Any]:
        source = {}
        if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict):
            source.update(item.raw.get("source_row") or {})
        source.update(dict(line or {}))
        source.update({
            "ui_status": item.status,
            "status": item.status,
            "name": item.name,
            "quantity_ordered": item.quantity,
            "m3_total": source.get("m3_total") or item.m3,
            "final_cost": source.get("final_cost") or item.final_cost,
            "ui_created_from": "erp_order_file_load",
        })
        return source

    def _load_supplier_order_from_excel(self, path: str, provider: str) -> tuple[tuple[OrderItem, ...], list[dict[str, Any]]]:
        import openpyxl

        # Algunos Excel de proveedor traen areas de impresion con nombres
        # definidos que openpyxl no puede resolver. No afectan a los datos del
        # pedido, asi que silenciamos solo ese warning para no ensuciar la consola.
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message=r"Print area cannot be set to Defined name.*",
                category=UserWarning,
                module=r"openpyxl.*",
            )
            wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        header_row, header_map = self._detect_supplier_order_header(ws)
        if header_row is None:
            raise ValueError("No se encontró una fila de cabecera con referencia/código y unidades/cantidad.")

        provider_key = self._normalize_search_text(provider)
        if provider_key.startswith("heimei") or provider_key.startswith("hemei"):
            c_ref = self._supplier_order_col(header_map, "REF", "Referencia")
            c_name = self._supplier_order_col(header_map, "DESCRIPCIÓN", "DESCRIPCION", "Descripcion", "Descripción")
            c_measure = self._supplier_order_col(header_map, "MEDIDA", "Medida")
            c_qty = self._supplier_order_col(header_map, "UND.", "UND", "Unidades")
            c_color = self._supplier_order_col(header_map, "Color")
            c_price = self._supplier_order_col(header_map, "Precio Compra unidad", "Precio compra unidad", "Precio Compra", "Precio")
            c_m3u = self._supplier_order_col(header_map, "M3/Und.", "M3Und", "m3/unit")
            c_m3t = self._supplier_order_col(header_map, "M3/total.", "M3total", "m3/total")
        elif provider_key.startswith("cipta"):
            c_ref = self._supplier_order_col(header_map, "CódigoArtículo (Item Code)", "CodigoArticulo (Item Code)", "CódigoArtículo", "CodigoArticulo", "Item Code", "Código", "Codigo", "REF")
            c_name = self._supplier_order_col(header_map, "Modelo (Model)", "Modelo", "Model", "Familia (Family)", "Familia")
            c_measure = self._supplier_order_col(header_map, "Medida (Size)", "Medida", "Size")
            c_color = self._supplier_order_col(header_map, "Color (Color)", "Color")
            c_qty = self._supplier_order_col(header_map, "Cantidad (Quantity)", "Cantidad", "Quantity", "UND", "Unidades")
            c_price = self._supplier_order_col(header_map, "Precio Compra unidad", "Precio compra unidad", "Precio Compra", "Precio")
            c_m3u = self._supplier_order_col(header_map, "m3/unit", "M3/Und.", "M3Und")
            c_m3t = self._supplier_order_col(header_map, "m3/total", "M3/total.", "M3total")
        else:
            c_ref = self._supplier_order_col(header_map, "Referencia")
            c_name = self._supplier_order_col(header_map, "Composición", "Composicion", "Descripcion", "Descripción")
            c_measure = self._supplier_order_col(header_map, "Medida")
            c_color = self._supplier_order_col(header_map, "Color")
            c_qty = self._supplier_order_col(header_map, "Und.", "Und", "Unidades", "Cantidad")
            c_price = None
            c_m3u = self._supplier_order_col(header_map, "M3/Und.", "M3Und", "m3/unit")
            c_m3t = self._supplier_order_col(header_map, "M3/total.", "M3total", "m3/total")

        if not c_ref or not c_qty:
            raise ValueError("El pedido debe tener al menos referencia/código y unidades/cantidad.")

        raw_lines: list[dict[str, Any]] = []
        items: list[OrderItem] = []
        for row_idx in range(header_row + 1, int(ws.max_row or 0) + 1):
            ref = ws.cell(row_idx, c_ref).value
            qty_raw = ws.cell(row_idx, c_qty).value
            qty = self._parse_units_value(qty_raw)
            if ref in (None, ""):
                continue
            if "total" in self._normalize_search_text(ref):
                continue
            if qty <= 0:
                continue
            measure = str(ws.cell(row_idx, c_measure).value or "").strip() if c_measure else ""
            name = str(ws.cell(row_idx, c_name).value or "").strip() if c_name else ""
            color = str(ws.cell(row_idx, c_color).value or "").strip() if c_color else ""
            m3_und = self._safe_order_float(ws.cell(row_idx, c_m3u).value) if c_m3u else 0.0
            m3_total = self._safe_order_float(ws.cell(row_idx, c_m3t).value) if c_m3t else 0.0
            if m3_und <= 0:
                m3_und = self._m3_from_measure_text(measure)
            if m3_total <= 0 and m3_und > 0:
                m3_total = round(m3_und * qty, 6)
            product = " · ".join(x for x in (name, measure) if x) or str(ref)
            raw = {
                "referencia": str(ref).strip(),
                "codigo": str(ref).strip(),
                "producto": product,
                "composicion": name,
                "medida": measure,
                "color": color,
                "unidades": qty,
                "unidades_raw": qty_raw,
                "m3_und": m3_und,
                "m3_total": m3_total,
                "precio_excel": self._safe_order_float(ws.cell(row_idx, c_price).value) if c_price else 0.0,
                "fila_excel": row_idx,
                "file_type": "XLSX",
            }
            item = self._order_item_from_loaded_line(raw)
            raw_lines.append(raw)
            items.append(item)
        return tuple(items), raw_lines

    def _load_supplier_order_from_pdf(self, path: str, provider: str) -> tuple[tuple[OrderItem, ...], list[dict[str, Any]]]:
        if not self._normalize_search_text(provider).startswith(("heimei", "hemei")):
            raise ValueError("La lectura de PDF está preparada de momento para pedidos Heimei. Para este proveedor usa Excel.")
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise ValueError("Para leer pedidos PDF instala pypdf: pip install pypdf. Si prefieres evitar dependencias, usa Excel.") from exc
        reader = PdfReader(path)
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        if not text.strip():
            raise ValueError("No se pudo extraer texto del PDF. Usa Excel o revisa que el PDF no sea una imagen escaneada.")
        pattern = re.compile(
            r"(?m)^\s*(\d{6,8})\s+(.+?)\s+"
            r"((?:\d+(?:[,.]\d+)?\s*[xX*×]\s*){2}\d+(?:[,.]\d+)?\s*cm)\s+"
            r"(\d+(?:[,.]\d+)?)\s*(?:Pc|pc|PCS|Pcs)?\s*$"
        )
        raw_lines: list[dict[str, Any]] = []
        items: list[OrderItem] = []
        for match in pattern.finditer(text):
            ref, desc, measure, qty_raw = match.groups()
            qty = self._parse_units_value(qty_raw)
            if qty <= 0:
                continue
            m3_und = self._m3_from_measure_text(measure)
            m3_total = round(m3_und * qty, 6) if m3_und else 0.0
            raw = {
                "referencia": ref.strip(),
                "codigo": ref.strip(),
                "producto": f"{desc.strip()} · {measure.strip()}",
                "composicion": desc.strip(),
                "medida": measure.strip(),
                "color": "",
                "unidades": qty,
                "unidades_raw": qty_raw,
                "m3_und": m3_und,
                "m3_total": m3_total,
                "precio_excel": 0.0,
                "fila_excel": 0,
                "file_type": "PDF",
            }
            item = self._order_item_from_loaded_line(raw)
            raw_lines.append(raw)
            items.append(item)
        if not items:
            raise ValueError("No se encontraron líneas de pedido en el PDF. Usa Excel o revisa el formato del PDF.")
        return tuple(items), raw_lines

    def _open_order_calc_flow(self, provider: str, order: SupplierOrder | None = None) -> None:
        win = tk.Toplevel(self)
        win.title(f"Calcular nuevo pedido - {provider}")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 1220, 760)
        win.minsize(1040, 660)
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)
        order_raw = order.raw if isinstance(getattr(order, "raw", None), dict) else {}
        order_source = order_raw.get("source_row") if isinstance(order_raw.get("source_row"), dict) else {}
        order_inputs = order_source.get("inputs") if isinstance(order_source.get("inputs"), dict) else {}
        existing_items = tuple(order.items) if order is not None else tuple()
        existing_items = self._fill_supplier_prices_for_order_items(provider, existing_items) if existing_items else existing_items
        existing_raw_lines = [item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {} for item in existing_items]
        loaded_order_state: dict[str, Any] = {
            "items": existing_items,
            "raw_lines": existing_raw_lines,
            "file_path": "",
            "file_name": str(order_raw.get("order_file") or "") if order is not None else "",
            "file_type": str(order_source.get("file_type") or "BORRADOR") if order is not None else "",
            "order_id": str(order_raw.get("order_id") or "") if order is not None else "",
            "existing_order": order,
        }

        header = tk.Frame(win, bg=BG)
        header.grid(row=0, column=0, sticky="ew", padx=22, pady=(22, 14))
        header.columnconfigure(0, weight=1)
        tk.Label(header, text="Calcular nuevo pedido", bg=BG, fg=TEXT, font=("Segoe UI", 24, "bold")).grid(row=0, column=0, sticky="w")
        actions_head = tk.Frame(header, bg=BG)
        actions_head.grid(row=0, column=1, sticky="e")
        self._button(actions_head, "Cancelar", command=win.destroy).pack(side=tk.LEFT, padx=(0, 8))
        self._button(actions_head, "Guardar pedido", primary=True, command=lambda: self._save_supplier_order_draft_from_calc(win, provider, order_entries, loaded_order_state, existing_order=order)).pack(side=tk.LEFT, padx=(0, 8))
        self._button(actions_head, "Exportar", primary=True, command=lambda: self._export_supplier_order_audit_excel(order, provider=provider, values=self._order_calc_values_from_entries(order_entries), items=tuple(loaded_order_state.get("items") or tuple()))).pack(side=tk.LEFT)

        body = tk.Frame(win, bg=BG)
        body.grid(row=1, column=0, sticky="nsew", padx=22, pady=(0, 22))
        body.columnconfigure(0, weight=1, minsize=390)
        body.columnconfigure(1, weight=2, minsize=760)
        body.rowconfigure(0, weight=1)

        params = self._card(body)
        params.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        params_head = tk.Frame(params, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        params_head.pack(fill=tk.X)
        tk.Label(params_head, text="Pedido", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(side=tk.LEFT, padx=18, pady=16)
        self._status_chip(params_head, f"Proveedor: {provider}", "Info").pack(side=tk.RIGHT, padx=18, pady=16)

        params_body = tk.Frame(params, bg=CARD)
        params_body.pack(fill=tk.BOTH, expand=True, padx=18, pady=18)
        upload_row = tk.Frame(params_body, bg=CARD)
        upload_row.pack(fill=tk.X, pady=(0, 14))
        file_chip = tk.Label(
            upload_row,
            text=str(loaded_order_state.get("file_name") or "Ningún pedido cargado"),
            bg=SOFT,
            fg="#334155",
            font=("Segoe UI", 9, "bold"),
            anchor=tk.W,
            highlightbackground=LINE,
            highlightthickness=1,
            padx=12,
            pady=10,
        )
        type_chip = tk.Label(upload_row, text=str(loaded_order_state.get("file_type") or "-"), bg=INDIGO_SOFT, fg=INDIGO, font=("Segoe UI", 8, "bold"), padx=10, pady=8)

        fields = tk.Frame(params_body, bg=CARD)
        fields.pack(fill=tk.X)
        for column in range(2):
            fields.columnconfigure(column, weight=1)
        is_heimei = provider.lower().startswith("hei")
        field_data = [
            ("Nombre del pedido", str(order_inputs.get("Nombre del pedido") or order.order_id if order else f"PED-{provider[:3].upper()}-BORRADOR")),
            ("Fecha", str(order_inputs.get("Fecha") or order.date if order else "")),
            ("Rentabilidad %", str(order_inputs.get("Rentabilidad %") or "")),
        ]
        if is_heimei:
            field_data.extend(
                [
                    ("Precio en Dolares", str(order_inputs.get("Precio en Dolares") or "")),
                    ("Precio pagado en Euros", str(order_inputs.get("Precio pagado en Euros") or "")),
                    ("Factura transporte", str(order_inputs.get("Factura transporte") or "")),
                    ("Derechos aranceles", str(order_inputs.get("Derechos aranceles") or "")),
                ]
            )
        else:
            field_data.append(("Coste transporte + IVA", str(order_inputs.get("Coste transporte + IVA") or "")))
        order_entries: dict[str, tk.Entry] = {}
        for index, (label, value) in enumerate(field_data):
            field_box = self._input_box(fields, label, value)
            field_box.grid(row=index // 2, column=index % 2, sticky="ew", padx=(0 if index % 2 == 0 else 8, 0), pady=(0, 10))
            entry = getattr(field_box, "_entry", None)
            if isinstance(entry, tk.Entry):
                order_entries[label] = entry

        tk.Label(
            params_body,
            text="Flujo heredado del proveedor: esta ventana ya se abre con el proveedor seleccionado desde la seccion de Pedidos.",
            bg=SOFT,
            fg="#475569",
            font=("Segoe UI", 9),
            wraplength=350,
            justify=tk.LEFT,
            padx=12,
            pady=10,
            highlightbackground=LINE,
            highlightthickness=1,
        ).pack(fill=tk.X, pady=(2, 14))

        def calculate_loaded_order() -> None:
            values = self._order_calc_values_from_entries(order_entries)
            items = tuple(loaded_order_state.get("items") or tuple())
            raw_lines = list(loaded_order_state.get("raw_lines") or [])
            if not items:
                messagebox.showinfo("Pedidos", "Carga un Excel/PDF o abre un borrador con líneas antes de calcular.")
                return
            overlay = self._show_working_overlay(
                "Calculando pedido",
                "Aplicando constantes, precios de proveedor, costes, ponderado y validaciones.\nNo cierres esta ventana.",
            )
            try:
                calculated_items, calculated_raw_lines, summary = self._calculate_supplier_order_in_memory(provider, values, items, raw_lines)
            except ValueError as exc:
                self._close_working_overlay(overlay)
                messagebox.showwarning("Pedidos", str(exc))
                return
            except Exception as exc:
                self._close_working_overlay(overlay)
                messagebox.showerror("Pedidos", f"No se pudo calcular el pedido.\n\n{exc}")
                return
            self._close_working_overlay(overlay)
            loaded_order_state["items"] = calculated_items
            loaded_order_state["raw_lines"] = calculated_raw_lines
            loaded_order_state["calculated"] = True
            loaded_order_state["calculation_summary"] = summary
            redraw_loaded_table(calculated_items)
            messagebox.showinfo(
                "Pedido calculado",
                f"Calculados: {summary.get('ok', 0)} · Pendientes: {summary.get('pending', 0)} · Coste: {summary.get('total_cost', 0):.2f} €",
            )

        actions_left = tk.Frame(params, bg=CARD)
        actions_left.pack(fill=tk.X, padx=18, pady=(0, 18))
        self._button(
            actions_left,
            "Guardar borrador",
            command=lambda: self._save_supplier_order_draft_from_calc(win, provider, order_entries, loaded_order_state, existing_order=order),
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        self._button(actions_left, "Calcular pedido", primary=True, command=lambda: calculate_loaded_order()).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0))

        table_card = self._card(body)
        table_card.grid(row=0, column=1, sticky="nsew")
        table_card.rowconfigure(1, weight=1)
        table_card.columnconfigure(0, weight=1)
        result_head = tk.Frame(table_card, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        result_head.grid(row=0, column=0, sticky="ew")
        result_head.columnconfigure(0, weight=1)
        tk.Label(result_head, text="Resultado del calculo", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w", padx=18, pady=16)
        result_metrics = tk.Frame(result_head, bg=CARD)
        result_metrics.grid(row=0, column=1, sticky="e", padx=18, pady=12)
        result_items_chip = self._status_chip(result_metrics, "Items: 0 · No descarga: 0", "Info")
        result_items_chip.pack(side=tk.LEFT, padx=(0, 8))
        result_m3_chip = self._status_chip(result_metrics, "M3 total: 0,00", "Info")
        result_m3_chip.pack(side=tk.LEFT)

        def update_result_metrics(items_for_metrics: tuple[OrderItem, ...] | None = None) -> None:
            metrics = self._supplier_order_visible_metrics(items_for_metrics or tuple())
            result_items_chip.configure(text=f"Items: {metrics['total_qty']} · No descarga: {metrics['excluded_qty']}")
            result_m3_chip.configure(text=f"M3 total: {self._format_optional_decimal(metrics['total_m3'], default='0,00')}")

        table_area = tk.Frame(table_card, bg=CARD)
        table_area.grid(row=1, column=0, sticky="nsew")
        table_area.rowconfigure(0, weight=1)
        table_area.columnconfigure(0, weight=1)

        def update_loaded_order_line(index: int, updated_item: OrderItem, updated_source: dict[str, Any] | None = None) -> None:
            current_items = list(loaded_order_state.get("items") or tuple())
            current_raw = list(loaded_order_state.get("raw_lines") or [])
            if index < 0 or index >= len(current_items):
                return
            current_items[index] = updated_item
            while len(current_raw) < len(current_items):
                current_raw.append({})
            if updated_source is not None:
                current_raw[index] = updated_source
            else:
                source = updated_item.raw.get("source_row") if isinstance(getattr(updated_item, "raw", None), dict) and isinstance(updated_item.raw.get("source_row"), dict) else {}
                current_raw[index] = source
            loaded_order_state["items"] = tuple(current_items)
            loaded_order_state["raw_lines"] = current_raw
            loaded_order_state["calculated"] = False
            loaded_order_state.pop("calculation_summary", None)
            redraw_loaded_table(tuple(current_items))

        def redraw_loaded_table(items: tuple[OrderItem, ...] | None = None) -> None:
            for child in table_area.winfo_children():
                child.destroy()
            visible_items = items if items is not None else tuple()
            update_result_metrics(visible_items)
            if visible_items:
                self._calculation_tree(table_area, visible_items, on_item_update=update_loaded_order_line, provider=provider).grid(row=0, column=0, sticky="nsew")
            else:
                empty = tk.Frame(table_area, bg=CARD)
                empty.grid(row=0, column=0, sticky="nsew")
                tk.Label(empty, text="Carga un Excel o PDF de pedido para ver sus líneas.", bg=CARD, fg=MUTED, font=("Segoe UI", 11)).pack(expand=True)

        def load_order_file() -> None:
            path = filedialog.askopenfilename(
                title="Seleccionar pedido",
                filetypes=[("Pedidos", "*.xlsx *.xlsm *.pdf"), ("Excel", "*.xlsx *.xlsm"), ("PDF", "*.pdf"), ("Todos los archivos", "*.*")],
            )
            if not path:
                return
            try:
                ext = os.path.splitext(path)[1].lower()
                if ext == ".pdf":
                    items, raw_lines = self._load_supplier_order_from_pdf(path, provider)
                    file_type = "PDF"
                else:
                    items, raw_lines = self._load_supplier_order_from_excel(path, provider)
                    file_type = "XLSX"
                items = self._fill_supplier_prices_for_order_items(provider, tuple(items))
                raw_lines = [
                    item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {}
                    for item in items
                ]
            except Exception as exc:
                messagebox.showerror("No se pudo cargar el pedido", str(exc))
                return
            loaded_order_state.update({
                "items": items,
                "raw_lines": raw_lines,
                "file_path": path,
                "file_name": os.path.basename(path),
                "file_type": file_type,
            })
            file_chip.configure(text=os.path.basename(path))
            type_chip.configure(text=file_type)
            redraw_loaded_table(items)
            messagebox.showinfo("Pedido cargado", f"Se cargaron {len(items)} líneas del pedido.")

        self._button(upload_row, "Cargar pedido", command=load_order_file).pack(side=tk.LEFT, padx=(0, 10))
        file_chip.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        type_chip.pack(side=tk.LEFT)
        redraw_loaded_table(existing_items)


    def _order_calc_values_from_entries(self, order_entries: dict[str, tk.Entry]) -> dict[str, str]:
        values: dict[str, str] = {}
        for label, entry in (order_entries or {}).items():
            try:
                values[str(label)] = str(entry.get() or "").strip()
            except Exception:
                values[str(label)] = ""
        return values

    def _money_float(self, value: Any, default: float = 0.0) -> float:
        if value in (None, ""):
            return default
        text = str(value).strip().replace("€", "").replace("$", "").replace("%", "")
        text = text.replace(".", "").replace(",", ".") if "," in text else text
        try:
            return float(text)
        except Exception:
            try:
                return float(str(value).replace(",", "."))
            except Exception:
                return default

    def _format_eur(self, value: Any) -> str:
        try:
            number = float(value)
        except Exception:
            return "Pendiente"
        return f"{number:.2f} €"

    def _format_usd(self, value: Any) -> str:
        try:
            number = float(value)
        except Exception:
            return "Pendiente"
        return f"{number:.2f} $"

    def _line_source_float(self, source: dict[str, Any], *keys: str) -> float:
        for key in keys:
            value = source.get(key)
            number = self._money_float(value)
            if number:
                return number
        return 0.0


    def _current_business_constant_values(self) -> dict[str, float]:
        """Return calculation constants from Supabase/cache/defaults as numbers."""
        constants = getattr(self, "_business_constants", None) or {key: dict(value) for key, value in DEFAULT_BUSINESS_CONSTANTS.items()}
        if self._cloud_session is not None:
            try:
                constants = list_business_constants(self._cloud_session)
                self._business_constants = constants
            except Exception:
                pass
        result: dict[str, float] = {}
        for key, meta in DEFAULT_BUSINESS_CONSTANTS.items():
            current = constants.get(key, meta) if isinstance(constants, dict) else meta
            result[key] = self._money_float(current.get("value") if isinstance(current, dict) else None) or self._money_float(meta.get("value"))
        # La fórmula legacy usa IVA_RE como multiplicador, no porcentaje bruto.
        # Si en configuración viene 26.2, equivale a 0.262.
        if result.get("IVA_RECARGO_EQUIVALENCIA", 0) > 1:
            result["IVA_RECARGO_EQUIVALENCIA_FACTOR"] = result["IVA_RECARGO_EQUIVALENCIA"] / 100
        else:
            result["IVA_RECARGO_EQUIVALENCIA_FACTOR"] = result.get("IVA_RECARGO_EQUIVALENCIA", 0)
        return result

    def _line_inventory_float(self, source: dict[str, Any], key: str, default: float = 0.0) -> float:
        return self._money_float(source.get(key), default) if hasattr(self, "_money_float") else default

    def _calculate_supplier_order_in_memory(self, provider: str, values: dict[str, str], items: tuple[OrderItem, ...], raw_lines: list[dict[str, Any]]) -> tuple[tuple[OrderItem, ...], list[dict[str, Any]], dict[str, Any]]:
        """Calcula el pedido en memoria usando fórmulas legacy de coste_pedido.py.

        Fórmula general Ekomat/Pascal/Cipta = calcular_coste_unitario_pedido.
        Fórmula Heimei/Tatamis = calcular_coste_unitario_tatamis_pedido.
        """
        provider_key = str(provider or "").strip().lower()
        is_heimei = provider_key.startswith("hei")
        items = self._fill_supplier_prices_for_order_items(provider, items)
        raw_lines = [
            item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {}
            for item in items
        ]
        constants = self._current_business_constant_values()
        rent_percent = self._money_float(values.get("Rentabilidad %"))

        # Conteos para fórmula general. M3 total del camión y productos que cuentan.
        total_m3 = sum(self._money_float((raw_lines[i] if i < len(raw_lines) else {}).get("m3_total") or item.m3) for i, item in enumerate(items))
        total_qty = sum(max(0, int(item.quantity or 0)) for item in items)
        cantidad_total_productos = sum(
            max(0, int(item.quantity or 0))
            for index, item in enumerate(items)
            if self._order_line_counts_for_download(item, raw_lines[index] if index < len(raw_lines) else None)
        )

        if is_heimei:
            precio_dolares = self._money_float(values.get("Precio en Dolares"))
            precio_euros = self._money_float(values.get("Precio pagado en Euros"))
            factura_transporte = self._money_float(values.get("Factura transporte"))
            derechos_aranceles = self._money_float(values.get("Derechos aranceles"))
            if precio_dolares <= 0 or precio_euros <= 0:
                raise ValueError("Para Heimei debes indicar Precio en Dólares y Precio pagado en Euros antes de calcular.")
            tasa_cambio = round(precio_dolares / precio_euros, 6)
            importe_transporte = factura_transporte + derechos_aranceles
            pc_transporte = round((importe_transporte / precio_euros) * 100, 2) if precio_euros else 0
            pc_descarga = round((constants["IMPORTE_DESCARGA_MT"] * 100) / precio_euros, 2) if precio_euros else 0
            pc_varios = round((constants["IMPORTES_VARIOS"] / precio_euros) * 100, 2) if precio_euros else 0
            pc_suma = round(pc_transporte + pc_descarga + constants["PC_GASTOS_FINANCIACION"] + constants["PC_GASTOS_MANIPULACION"] + pc_varios, 2)
            calc_inputs = {
                "formula": "calcular_coste_unitario_tatamis_pedido",
                "precio_dolares": precio_dolares,
                "precio_euros": precio_euros,
                "factura_transporte": factura_transporte,
                "derechos_aranceles": derechos_aranceles,
                "tasa_cambio": tasa_cambio,
                "importe_transporte": importe_transporte,
                "pc_transporte": pc_transporte,
                "pc_descarga": pc_descarga,
                "pc_varios": pc_varios,
                "pc_manipulacion": constants["PC_GASTOS_MANIPULACION"],
                "pc_financiacion": constants["PC_GASTOS_FINANCIACION"],
                "pc_suma": pc_suma,
                "constants": constants,
            }
        else:
            coste_transporte = self._money_float(values.get("Coste transporte + IVA"))
            if coste_transporte <= 0:
                raise ValueError("Indica Coste transporte + IVA antes de calcular este proveedor.")
            if total_m3 <= 0:
                raise ValueError("El pedido no tiene M3 total válido. Revisa las líneas cargadas.")
            if cantidad_total_productos <= 0:
                raise ValueError("El pedido no tiene cantidad total válida para descarga. Revisa las líneas con 'Cuenta para descarga'.")
            ct_m3 = round(coste_transporte / total_m3, 2) if total_m3 else 0
            cd_prod_iva = round(constants["COSTE_TOTAL_DESCARGA_FUTONES_IVA"] / cantidad_total_productos, 2) if cantidad_total_productos else 0
            calc_inputs = {
                "formula": "calcular_coste_unitario_pedido",
                "coste_transporte_iva": coste_transporte,
                "coste_total_descarga_iva": constants["COSTE_TOTAL_DESCARGA_FUTONES_IVA"],
                "m3_total_camion": total_m3,
                "cantidad_total_productos": cantidad_total_productos,
                "ct_m3": ct_m3,
                "cd_prod_iva": cd_prod_iva,
                "constants": constants,
            }

        calculated: list[OrderItem] = []
        calculated_raw: list[dict[str, Any]] = []
        ok = 0
        pending = 0
        total_cost = 0.0
        for index, item in enumerate(items):
            source = dict(raw_lines[index] if index < len(raw_lines) and isinstance(raw_lines[index], dict) else {})
            qty = max(0, int(item.quantity or 0))
            m3_total_line = self._money_float(source.get("m3_total") or item.m3)
            m3_unit = self._money_float(source.get("m3_und")) or (m3_total_line / qty if qty else 0.0)
            if m3_total_line <= 0 and m3_unit > 0 and qty > 0:
                m3_total_line = round(m3_unit * qty, 6)
            price_provider = self._line_source_float(source, "precio_proveedor", "precio_excel", "precio", "base_price")
            rotacion_c = self._money_float(source.get("inventory_rotation_c") or source.get("rotation_c"), 0)
            n_bultos = int(self._money_float(source.get("inventory_packages") or source.get("packages"), 0))
            reasons: list[str] = []
            if not str(item.code or "").strip() or str(item.code).strip() == "-":
                reasons.append("falta ID")
            if qty <= 0:
                reasons.append("falta cantidad")
            if m3_total_line <= 0 and m3_unit <= 0:
                reasons.append("falta M3")
            if price_provider <= 0:
                reasons.append("falta precio proveedor")
            if rotacion_c <= 0:
                reasons.append("falta rotación C")
            if n_bultos <= 0:
                reasons.append("faltan bultos/packages")

            line_details: dict[str, Any] = {}
            if reasons:
                status = "Error"
                final_unit = 0.0
                line_cost = 0.0
                pending += 1
            else:
                if is_heimei:
                    precio_euros_art = round(price_provider / calc_inputs["tasa_cambio"], 2) if calc_inputs["tasa_cambio"] else price_provider
                    gastos_aplicables = round(precio_euros_art * calc_inputs["pc_suma"] / 100, 2)
                    coste_sin_almacenaje = round(precio_euros_art + gastos_aplicables, 2)
                    coste_almacenaje_iva = round(constants["COSTE_DIARIO_ALMACENAJE_M3"] * m3_unit * rotacion_c * 1.21, 4)
                    coste_picking_iva = round(((n_bultos * 0.3) + 4.12) * 1.21, 3)
                    final_unit = round(coste_sin_almacenaje + coste_almacenaje_iva + coste_picking_iva, 2)
                    line_details = {
                        "precio_euros_art": precio_euros_art,
                        "gastos_aplicables": gastos_aplicables,
                        "coste_sin_almacenaje": coste_sin_almacenaje,
                        "coste_almacenaje_iva": coste_almacenaje_iva,
                        "coste_picking_iva": coste_picking_iva,
                    }
                else:
                    ct_m3_prod = round(calc_inputs["ct_m3"] * m3_unit, 2)
                    ct_total_ref = round(qty * ct_m3_prod, 2)
                    cuenta_descarga = self._order_line_counts_for_download(item, source)
                    cd_prod_linea = calc_inputs["cd_prod_iva"] if cuenta_descarga else 0
                    cd_total_ref = round(qty * cd_prod_linea, 3)
                    iva_re = round(price_provider * constants["IVA_RECARGO_EQUIVALENCIA_FACTOR"], 2)
                    precio_con_iva = round(price_provider + iva_re, 2)
                    coste_descarga = round(ct_m3_prod + cd_prod_linea + precio_con_iva, 2)
                    coste_almacenaje_iva = round(constants["COSTE_DIARIO_ALMACENAJE_M3"] * m3_unit * rotacion_c * 1.21, 4)
                    coste_picking_iva = round(((n_bultos * 0.3) + 4.12) * 1.21, 3)
                    final_unit = round(coste_descarga + coste_almacenaje_iva + coste_picking_iva, 2)
                    line_details = {
                        "ct_m3": calc_inputs["ct_m3"],
                        "ct_m3_prod": ct_m3_prod,
                        "ct_total_ref": ct_total_ref,
                        "cd_prod_iva": cd_prod_linea,
                        "cd_prod_iva_global": calc_inputs["cd_prod_iva"],
                        "cuenta_para_descarga": cuenta_descarga,
                        "cd_total_ref": cd_total_ref,
                        "iva_re": iva_re,
                        "precio_con_iva": precio_con_iva,
                        "coste_descarga": coste_descarga,
                        "coste_almacenaje_iva": coste_almacenaje_iva,
                        "coste_picking_iva": coste_picking_iva,
                    }
                if rent_percent > 0 and rent_percent < 100:
                    final_unit = round(final_unit / (1 - rent_percent / 100), 2)
                    line_details["precio_coste_final_con_rentabilidad"] = final_unit
                line_cost = round(final_unit * qty, 2)
                status = "Calculado"
                ok += 1
                total_cost += line_cost

            stock_total_actual = self._money_float(source.get("inventory_stock_total"), 0.0)
            weighted_current = self._money_float(source.get("inventory_weighted_average_cost") or source.get("weighted_average_cost"), 0.0)
            precio_ponderado_lote = 0.0
            if status == "Calculado" and final_unit > 0:
                if stock_total_actual > 0 and weighted_current > 0:
                    precio_ponderado_lote = round(((stock_total_actual * weighted_current) + (qty * final_unit)) / (stock_total_actual + qty), 2) if (stock_total_actual + qty) else final_unit
                else:
                    precio_ponderado_lote = final_unit

            enriched_source = {
                **source,
                "status": status,
                "ui_status": status,
                "m3_total": m3_total_line,
                "m3_und": m3_unit,
                "rotacion_c": rotacion_c,
                "packages": n_bultos,
                "precio_proveedor": price_provider,
                "unit_cost": final_unit,
                "precio_coste_final": final_unit,
                "final_cost": line_cost,
                "line_cost": line_cost,
                "stock_total_actual": stock_total_actual,
                "weighted_average_cost_actual": weighted_current,
                "precio_ponderado_lote": precio_ponderado_lote,
                "calculation_inputs": calc_inputs,
                "calculation_details": line_details,
                "calculation_reasons": reasons,
                "rentabilidad_percent": rent_percent,
                "cuenta_para_descarga": "Sí" if self._order_line_counts_for_download(item, source) else "No",
                "cuenta_reparto_descarga": self._order_line_counts_for_download(item, source),
                "cuenta_pedido": "Sí" if self._order_line_counts_for_download(item, source) else "No",
            }
            calculated_raw.append(enriched_source)
            calculated.append(
                OrderItem(
                    code=item.code,
                    name=item.name,
                    quantity=qty,
                    m3=self._format_optional_decimal(m3_total_line or m3_unit, default="Pendiente"),
                    final_cost=self._format_eur(line_cost) if line_cost else "Pendiente",
                    status=status,
                    raw={"source_row": enriched_source},
                )
            )
        excluded_qty = max(0, total_qty - cantidad_total_productos)
        return tuple(calculated), calculated_raw, {
            "ok": ok,
            "pending": pending,
            "total_cost": total_cost,
            "total_m3": total_m3,
            "total_qty": total_qty,
            "download_qty": cantidad_total_productos,
            "excluded_download_qty": excluded_qty,
            "formula": calc_inputs.get("formula"),
        }

    def _save_supplier_order_draft_from_calc(self, win: tk.Toplevel, provider: str, order_entries: dict[str, tk.Entry], loaded_order_state: dict[str, Any] | None = None, existing_order: SupplierOrder | None = None) -> None:
        """Guarda un pedido borrador real desde la ventana de cálculo.

        Este método faltaba en la clase, por eso Tkinter terminaba delegando el
        atributo al root (`self.tk`) y lanzaba `AttributeError`.
        """
        if self._cloud_session is None:
            messagebox.showwarning("Pedidos", "Inicia sesión en Supabase para guardar pedidos reales.")
            return

        values: dict[str, str] = {}
        for label, entry in (order_entries or {}).items():
            try:
                values[str(label)] = str(entry.get() or "").strip()
            except Exception:
                values[str(label)] = ""

        order_name = values.get("Nombre del pedido", "").strip()
        if not order_name:
            messagebox.showwarning("Pedidos", "Pon un nombre al pedido antes de guardar el borrador.")
            return

        safe_provider = str(provider or "Otros").strip() or "Otros"
        loaded_order_state = loaded_order_state or {}
        loaded_items = tuple(loaded_order_state.get("items") or tuple())
        raw_lines = list(loaded_order_state.get("raw_lines") or [])
        file_type = str(loaded_order_state.get("file_type") or "BORRADOR").upper()
        order_file = str(loaded_order_state.get("file_name") or f"{order_name}.borrador")
        notes = "Borrador guardado desde UI ERP. Pendiente de cálculo."
        if loaded_items:
            notes = f"Borrador guardado desde UI ERP con {len(loaded_items)} líneas cargadas. Pendiente de cálculo."

        item_payloads = []
        for index, item in enumerate(loaded_items):
            source_row = self._order_item_source_row(item, raw_lines[index] if index < len(raw_lines) else None)
            item_payloads.append({
                "code": item.code,
                "name": item.name,
                "quantity": item.quantity,
                "m3": item.m3,
                "final_cost": self._money_float(source_row.get("line_cost") or source_row.get("final_cost") or item.final_cost),
                "line_cost": self._money_float(source_row.get("line_cost") or source_row.get("final_cost") or item.final_cost),
                "unit_cost": self._money_float(source_row.get("unit_cost")),
                "status": item.status,
                "source_row": source_row,
            })

        try:
            existing_raw = existing_order.raw if isinstance(getattr(existing_order, "raw", None), dict) else {}
            existing_order_id = str(loaded_order_state.get("order_id") or existing_raw.get("order_id") or "").strip()
            is_calculated = bool(loaded_order_state.get("calculated"))
            has_errors = any(str(item.get("status") or "").lower() in {"error", "critical", "bloqueado"} for item in item_payloads)
            final_status = "Validación" if has_errors else "Calculado"
            if existing_order_id:
                if is_calculated:
                    update_supplier_order_calculation(
                        self._cloud_session,
                        order_id=existing_order_id,
                        provider=safe_provider,
                        order_name=order_name,
                        order_file=order_file,
                        file_type=file_type,
                        notes="Pedido calculado desde UI ERP." if final_status == "Calculado" else "Pedido calculado con líneas pendientes de revisión.",
                        inputs=values,
                        items=item_payloads,
                        status=final_status,
                    )
                else:
                    update_supplier_order_draft(
                        self._cloud_session,
                        order_id=existing_order_id,
                        provider=safe_provider,
                        order_name=order_name,
                        order_file=order_file,
                        file_type=file_type,
                        notes=notes,
                        inputs=values,
                        items=item_payloads,
                    )
            else:
                created = create_supplier_order_draft(
                    self._cloud_session,
                    provider=safe_provider,
                    order_name=order_name,
                    order_file=order_file,
                    file_type=file_type,
                    notes=notes,
                    inputs=values,
                    items=item_payloads,
                )
                created_order_id = str(created.get("order_id") or "").strip()
                if is_calculated and created_order_id:
                    update_supplier_order_calculation(
                        self._cloud_session,
                        order_id=created_order_id,
                        provider=safe_provider,
                        order_name=order_name,
                        order_file=order_file,
                        file_type=file_type,
                        notes="Pedido calculado desde UI ERP." if final_status == "Calculado" else "Pedido calculado con líneas pendientes de revisión.",
                        inputs=values,
                        items=item_payloads,
                        status=final_status,
                    )
        except Exception as exc:
            messagebox.showerror("Pedidos", f"No se pudo guardar el borrador.\n\n{exc}")
            return

        self._orders_loaded_once = False
        self._supplier_orders = []
        self._selected_supplier_order = None
        try:
            win.destroy()
        except Exception:
            pass
        self._show_view("calcular")
        messagebox.showinfo("Pedidos", "Borrador guardado correctamente.")

    def _input_box(self, parent: tk.Misc, label: str, value: str) -> tk.Frame:
        frame = tk.Frame(parent, bg=CARD)
        tk.Label(frame, text=label.upper(), bg=CARD, fg=MUTED, font=("Segoe UI", 8, "bold"), anchor=tk.W).pack(fill=tk.X, pady=(0, 5))
        entry = tk.Entry(
            frame,
            bg=CARD,
            fg="#334155",
            relief=tk.FLAT,
            highlightbackground=LINE,
            highlightcolor=INDIGO,
            highlightthickness=1,
            font=("Segoe UI", 10),
        )
        entry.insert(0, value)
        entry.pack(fill=tk.X, ipady=8)
        setattr(frame, "_entry", entry)
        return frame

    def _value_card(self, parent: tk.Misc, label: str, value: str) -> tk.Frame:
        frame = tk.Frame(parent, bg=SOFT, highlightbackground=LINE, highlightthickness=1)
        tk.Label(frame, text=label.upper(), bg=SOFT, fg=MUTED, font=("Segoe UI", 7, "bold")).pack(anchor=tk.W, padx=10, pady=(8, 0))
        tk.Label(frame, text=value, bg=SOFT, fg=TEXT, font=("Segoe UI", 11, "bold")).pack(anchor=tk.W, padx=10, pady=(3, 8))
        return frame

    def _calculation_mode_for_items(self, items: tuple[OrderItem, ...], provider: str | None = None) -> str:
        """Return the supplier calculation mode used to draw the calculation table."""
        if str(provider or "").strip().lower().startswith("hei"):
            return "heimei"
        for item in items or tuple():
            source = item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {}
            calc_inputs = source.get("calculation_inputs") if isinstance(source.get("calculation_inputs"), dict) else {}
            if calc_inputs.get("formula") == "calcular_coste_unitario_tatamis_pedido":
                return "heimei"
        return "general"

    def _supplier_order_visible_metrics(self, items: tuple[OrderItem, ...] | list[OrderItem]) -> dict[str, float | int]:
        """Return header metrics for the visible supplier order table.

        total_qty is the ordered unit count. excluded_qty is the ordered unit count
        that will not receive the fixed descarga reparto. total_m3 comes from the
        loaded/calculated line total when available, or from unit M3 × quantity.
        """
        total_qty = 0
        download_qty = 0
        total_m3 = 0.0
        for item in tuple(items or tuple()):
            source = item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {}
            qty = max(0, int(self._money_float(source.get("quantity_ordered") or source.get("unidades") or item.quantity, 0)))
            total_qty += qty
            if self._order_line_counts_for_download(item, source):
                download_qty += qty
            line_m3 = self._money_float(source.get("m3_total"), 0.0)
            if line_m3 <= 0:
                unit_m3 = self._money_float(source.get("m3_und") or source.get("inventory_m3"), 0.0)
                if unit_m3 > 0 and qty > 0:
                    line_m3 = unit_m3 * qty
                else:
                    line_m3 = self._money_float(item.m3, 0.0)
            total_m3 += max(0.0, line_m3)
        return {
            "total_qty": total_qty,
            "download_qty": download_qty,
            "excluded_qty": max(0, total_qty - download_qty),
            "total_m3": round(total_m3, 6),
        }

    def _calculation_tree(self, parent: tk.Misc, items: tuple[OrderItem, ...], on_item_update: Any | None = None, provider: str | None = None) -> tk.Frame:
        """Pedido calculation table with normal horizontal and vertical scroll.

        The previous frozen-column experiment made row selection and editing feel
        split in three tables. For daily work this returns to a single Treeview:
        every row stays together, double-click editing is predictable, and the
        horizontal scroll moves the full calculation table.
        """
        frame = tk.Frame(parent, bg=CARD)
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        mode = self._calculation_mode_for_items(items, provider)
        if mode == "heimei":
            columns = [
                "ID",
                "Producto / Medida",
                "Color",
                "Und.",
                "Cuenta para descarga",
                "M3/Und.",
                "M3 total",
                "Estado",
                "Precio proveedor",
                "Precio en Dolares",
                "Precio en Euros",
                "Tasa cambio",
                "Precio articulo EUR",
                "% Transporte",
                "% Descarga",
                "% Varios",
                "% Manipulacion",
                "% Financiacion",
                "Gastos aplicables",
                "Coste sin almacenaje",
                "Almacenaje + IVA",
                "Picking + IVA",
                "Rentabilidad %",
                "Coste Final Articulo",
                "Precio ponderado lote",
                "Coste Total Cantidad",
            ]
            widths = {
                "ID": 95,
                "Producto / Medida": 260,
                "Color": 105,
                "Und.": 70,
                "Cuenta para descarga": 145,
                "M3/Und.": 90,
                "M3 total": 95,
                "Estado": 145,
                "Precio proveedor": 145,
                "Precio en Dolares": 150,
                "Precio en Euros": 150,
                "Tasa cambio": 125,
                "Precio articulo EUR": 150,
                "% Transporte": 120,
                "% Descarga": 115,
                "% Varios": 105,
                "% Manipulacion": 130,
                "% Financiacion": 130,
                "Gastos aplicables": 150,
                "Coste sin almacenaje": 170,
                "Almacenaje + IVA": 145,
                "Picking + IVA": 130,
                "Rentabilidad %": 125,
                "Coste Final Articulo": 160,
                "Precio ponderado lote": 170,
                "Coste Total Cantidad": 165,
            }
        else:
            columns = [
                "ID",
                "Producto / Medida",
                "Color",
                "Und.",
                "Cuenta para descarga",
                "M3/Und.",
                "M3 total",
                "Estado",
                "Precio proveedor",
                "IVA + RE",
                "Precio compra IVA+RE",
                "Transporte M3/Und.",
                "Descarga/Und.",
                "Coste final con descarga",
                "Almacenaje + IVA",
                "Picking + IVA",
                "Rentabilidad %",
                "Coste Final Articulo",
                "Precio ponderado lote",
                "Coste Total Cantidad",
            ]
            widths = {
                "ID": 95,
                "Producto / Medida": 260,
                "Color": 105,
                "Und.": 70,
                "Cuenta para descarga": 145,
                "M3/Und.": 90,
                "M3 total": 95,
                "Estado": 145,
                "Precio proveedor": 135,
                "IVA + RE": 115,
                "Precio compra IVA+RE": 170,
                "Transporte M3/Und.": 170,
                "Descarga/Und.": 135,
                "Coste final con descarga": 185,
                "Almacenaje + IVA": 145,
                "Picking + IVA": 130,
                "Rentabilidad %": 125,
                "Coste Final Articulo": 160,
                "Precio ponderado lote": 170,
                "Coste Total Cantidad": 165,
            }
        tree = ttk.Treeview(frame, columns=columns, show="headings", height=12)
        for column in columns:
            tree.heading(column, text=column, anchor=tk.CENTER)
            tree.column(column, width=widths.get(column, 130), anchor=tk.CENTER, stretch=False)

        # Sistema de fijado de objetivo: cualquier linea con error/dato
        # bloqueante queda con relleno rojo en toda la fila para ubicar
        # rapidamente que item impide calcular el pedido. Warning queda amarillo.
        tree.tag_configure("error_row", background="#FEE2E2", foreground="#991B1B")
        tree.tag_configure("warning_row", background=AMBER_SOFT, foreground=AMBER)
        tree.tag_configure("ok_row", background="")

        item_by_iid: dict[str, tuple[int, OrderItem]] = {}
        for index, (item, row) in enumerate(zip(items, self._calculation_rows(items, provider=provider))):
            missing = self._order_item_missing_reasons(item)
            status = str(item.status or "").strip().lower()
            source = item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {}
            source_status = str(source.get("ui_status") or source.get("status") or "").strip().lower()
            has_error = bool(missing) or status in {"error", "critical", "bloqueado", "blocked"} or source_status in {"error", "critical", "bloqueado", "blocked"}
            has_warning = status in {"warning", "validacion", "validación"} or source_status in {"warning", "validacion", "validación"}
            tag = "error_row" if has_error else "warning_row" if has_warning else "ok_row"
            iid = tree.insert("", tk.END, values=row, tags=(tag,))
            item_by_iid[iid] = (index, item)

        yscroll = tk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        xscroll = tk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        def on_double_click(event: tk.Event) -> None:
            row_id = tree.identify_row(event.y)
            if not row_id:
                return
            tree.selection_set(row_id)
            tree.focus(row_id)
            found = item_by_iid.get(row_id)
            if found is not None:
                index, item = found
                self._open_order_item_missing_editor(item, on_save=(lambda updated_item, updated_source, index=index: on_item_update(index, updated_item, updated_source)) if callable(on_item_update) else None)

        tree.bind("<Double-1>", on_double_click)
        return frame

    def _open_order_item_missing_editor(self, item: OrderItem, on_save: Any | None = None) -> None:
        """ERP version of the old double-click editor from coste_pedido.py.

        The old window opened on double-click, marked missing fields in red, and
        let the user complete data so the line can be recalculated and saved
        back to inventory_items when the item code matches a Supabase item_id.
        """
        win = tk.Toplevel(self)
        win.title(f"Completar datos - {item.code}")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 700, 700)
        win.minsize(640, 620)
        win.columnconfigure(0, weight=1)

        card = self._card(win)
        card.pack(fill=tk.BOTH, expand=True, padx=18, pady=18)
        header = tk.Frame(card, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        header.pack(fill=tk.X)
        tk.Label(header, text="Completar datos para calcular", bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold")).pack(anchor=tk.W, padx=18, pady=(16, 4))
        tk.Label(header, text=f"{item.code} - {item.name}", bg=CARD, fg=MUTED, wraplength=560, justify=tk.LEFT).pack(anchor=tk.W, padx=18, pady=(0, 14))

        form = tk.Frame(card, bg=CARD)
        form.pack(fill=tk.BOTH, expand=True, padx=18, pady=18)
        for col in range(2):
            form.columnconfigure(col, weight=1)

        source = item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {}
        fields = [
            ("Referencia", item.code, True),
            ("Descripcion", item.name, False),
            ("Unidades", str(item.quantity), False),
            ("M3/Und.", self._format_optional_decimal(source.get("m3_und") or item.m3, default="") if item.m3 != "Pendiente" else "", False),
            ("Rotación C", str(source.get("rotacion_c") or source.get("inventory_rotation_c") or source.get("rotation_c") or ""), False),
            ("Bultos", str(source.get("packages") or source.get("inventory_packages") or "1"), False),
            ("Precio proveedor", str(source.get("precio_proveedor") or source.get("precio_excel") or source.get("precio") or ""), False),
        ]
        entries: dict[str, tk.Entry] = {}
        for index, (label, value, readonly) in enumerate(fields):
            field = tk.Frame(form, bg=CARD)
            field.grid(row=index // 2, column=index % 2, sticky="ew", padx=(0 if index % 2 == 0 else 8, 0), pady=(0, 12))
            tk.Label(field, text=label.upper(), bg=CARD, fg=MUTED, font=("Segoe UI", 8, "bold"), anchor=tk.W).pack(fill=tk.X, pady=(0, 5))
            entry = tk.Entry(field, bg=CARD, fg="#334155", relief=tk.FLAT, highlightbackground=LINE, highlightcolor=INDIGO, highlightthickness=1, font=("Segoe UI", 10))
            entry.insert(0, value)
            if readonly:
                entry.configure(state="readonly", readonlybackground=SOFT)
            entry.pack(fill=tk.X, ipady=8)
            entries[label] = entry

        cuenta_field = tk.Frame(form, bg=CARD)
        cuenta_field.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(0, 12))
        tk.Label(cuenta_field, text="CUENTA PARA DESCARGA", bg=CARD, fg=MUTED, font=("Segoe UI", 8, "bold"), anchor=tk.W).pack(fill=tk.X, pady=(0, 5))
        cuenta_var = tk.StringVar(value=self._order_line_counts_for_download_label(item, source))
        cuenta_combo = ttk.Combobox(cuenta_field, values=("Sí", "No"), textvariable=cuenta_var, state="readonly", font=("Segoe UI", 10), width=12)
        cuenta_combo.pack(side=tk.LEFT, fill=tk.X, expand=False, ipady=5)
        cuenta_reason = tk.Label(
            cuenta_field,
            text=self._order_line_download_reason(item, source),
            bg=INDIGO_SOFT,
            fg="#4338CA",
            wraplength=600,
            justify=tk.LEFT,
            padx=10,
            pady=8,
        )
        cuenta_reason.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))

        def reset_cuenta_auto() -> None:
            auto = self._order_line_counts_for_download_auto(item.code, item.name, source)
            cuenta_var.set("Sí" if auto else "No")
            cuenta_reason.config(text=f"Regla automática aplicada: {'Sí' if auto else 'No'}.")

        self._button(cuenta_field, "Regla auto", command=reset_cuenta_auto).pack(side=tk.RIGHT, padx=(10, 0))

        def mark_if_missing() -> None:
            for label, entry in entries.items():
                value = entry.get().strip() if entry.cget("state") != "readonly" else "readonly"
                missing = False
                if label in {"Descripcion", "Unidades", "M3/Und.", "Rotación C", "Bultos", "Precio proveedor"} and not value:
                    missing = True
                if label in {"M3/Und.", "Rotación C", "Bultos"} and value.lower() in {"pendiente", "0", "0.0"}:
                    missing = True
                entry.configure(highlightbackground=ROSE if missing else LINE, highlightcolor=ROSE if missing else INDIGO, bg=ROSE_SOFT if missing else CARD)

        mark_if_missing()

        msg = tk.Label(
            card,
            text="Los campos marcados en rojo son necesarios. Al aceptar, se actualiza la línea del pedido y, si la referencia coincide con inventory_items.item_id, se guarda M3/medidas/precio de pedido en Supabase con log y snapshot.",
            bg=INDIGO_SOFT,
            fg="#4338CA",
            wraplength=560,
            justify=tk.LEFT,
            padx=12,
            pady=10,
        )
        msg.pack(fill=tk.X, padx=18, pady=(0, 14))

        def accept_line_changes() -> None:
            mark_if_missing()
            desc = entries["Descripcion"].get().strip()
            qty = self._parse_units_value(entries["Unidades"].get())
            m3_unit = self._money_float(entries["M3/Und."].get())
            rotation_c = self._money_float(entries["Rotación C"].get())
            packages = int(self._money_float(entries["Bultos"].get()))
            price_provider = self._money_float(entries["Precio proveedor"].get())
            cuenta = cuenta_var.get().strip() or "Sí"
            cuenta_bool = cuenta == "Sí"
            missing: list[str] = []
            if not desc:
                missing.append("Descripcion")
            if qty <= 0:
                missing.append("Unidades")
            if m3_unit <= 0:
                missing.append("M3/Und.")
            if rotation_c <= 0:
                missing.append("Rotación C")
            if packages <= 0:
                missing.append("Bultos")
            if price_provider <= 0:
                missing.append("Precio proveedor")
            if missing:
                messagebox.showwarning("Pedidos", "Completa los campos obligatorios: " + ", ".join(missing))
                return

            total_m3 = round(m3_unit * qty, 6)
            updated_source = dict(source)
            updated_source.update({
                "producto": desc,
                "descripcion": desc,
                "composicion": desc,
                "unidades": qty,
                "quantity_ordered": qty,
                "m3_und": m3_unit,
                "m3_total": total_m3,
                "rotacion_c": rotation_c,
                "inventory_rotation_c": rotation_c,
                "packages": packages,
                "inventory_packages": packages,
                "precio_excel": price_provider,
                "precio_proveedor": price_provider,
                "supplier_price_source": "manual_order_editor",
                "supplier_price_column": "manual",
                "supplier_price_matched_by": "manual",
                "ui_manual_supplier_price": True,
                "cuenta_pedido": cuenta,
                "cuenta_para_descarga": cuenta,
                "cuenta_reparto_descarga": cuenta_bool,
                "status": "OK",
                "ui_status": "OK",
                "calculation_reasons": [],
                "ui_completed_from_order_editor": True,
            })
            updated_item = OrderItem(
                code=item.code,
                name=desc,
                quantity=qty,
                m3=self._format_optional_decimal(total_m3, default="Pendiente"),
                final_cost="Pendiente",
                status="OK",
                raw={"source_row": updated_source},
            )

            # Si la referencia coincide con inventory_items.item_id, guardamos los
            # campos que sirven para que este bug no vuelva a aparecer en pedidos
            # futuros. No tocamos WooCommerce.
            if self._cloud_session is not None and str(item.code or "").strip().isdigit():
                try:
                    update_inventory_item_fields(
                        self._cloud_session,
                        int(str(item.code).strip()),
                        {
                            "name": desc,
                            "cubic_meters": m3_unit,
                            "rotation_c": rotation_c,
                            "packages": packages,
                            "primary_supplier_price": price_provider,
                        },
                        notes="Datos completados desde editor de línea de Pedido. WooCommerce no fue tocado.",
                    )
                    updated_source["inventory_update_status"] = "OK"
                except Exception as exc:
                    updated_source["inventory_update_status"] = "ERROR"
                    updated_source["inventory_update_error"] = str(exc)
                    messagebox.showwarning(
                        "Inventario",
                        "La línea se actualizó en el pedido, pero no se pudo actualizar inventory_items.\n\n" + str(exc),
                    )
            if callable(on_save):
                on_save(updated_item, updated_source)
            win.destroy()

        buttons = tk.Frame(card, bg=CARD)
        buttons.pack(fill=tk.X, padx=18, pady=(0, 18))
        self._button(buttons, "Cancelar", command=win.destroy).pack(side=tk.RIGHT, padx=(8, 0))
        self._button(buttons, "Aceptar", primary=True, command=accept_line_changes).pack(side=tk.RIGHT)

    def _calculation_rows(self, items: tuple[OrderItem, ...], provider: str | None = None) -> list[tuple[str, ...]]:
        """Return calculation rows using supplier-specific formulas and labels.

        General suppliers (Ekomat/Pascal/Cipta) no longer show the Heimei USD
        fields. Heimei keeps the order USD/EUR totals and exchange-rate columns
        because that formula actually needs them. This keeps the visible table
        aligned with the calculation path that leads to Coste Final Articulo.
        """
        rows = []
        mode = self._calculation_mode_for_items(items, provider)

        def eur(value: Any, default: str = "Pendiente") -> str:
            return self._format_eur(value) if value not in (None, "", 0, 0.0) else default

        def usd(value: Any, default: str = "Pendiente") -> str:
            return self._format_usd(value) if value not in (None, "", 0, 0.0) else default

        def num(value: Any, default: str = "Pendiente") -> str:
            return self._format_optional_decimal(value, default=default) if value not in (None, "") else default

        for item in items:
            missing = self._order_item_missing_reasons(item)
            source = item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {}
            cuenta = self._order_line_counts_for_download_label(item, source)
            estado = "Faltan: " + ", ".join(missing) if missing else (item.status or "OK")
            calc_inputs = source.get("calculation_inputs") if isinstance(source.get("calculation_inputs"), dict) else {}
            calc_details = source.get("calculation_details") if isinstance(source.get("calculation_details"), dict) else {}
            precio_base_num = self._money_float(source.get("precio_proveedor") or source.get("precio_excel") or source.get("precio"))
            precio_ponderado_lote = source.get("precio_ponderado_lote")
            coste_total_cantidad = source.get("line_cost") or source.get("final_cost")
            coste_unitario_final = source.get("precio_coste_final") or source.get("unit_cost")
            rentabilidad = source.get("rentabilidad_percent")
            common = (
                item.code,
                item.name,
                str(source.get("color") or "-"),
                str(item.quantity),
                cuenta,
                num(source.get("m3_und") or source.get("inventory_m3") or item.m3),
                num(source.get("m3_total") or item.m3),
                estado,
            )
            if mode == "heimei":
                rows.append(
                    common
                    + (
                        usd(precio_base_num),
                        usd(calc_inputs.get("precio_dolares")),
                        eur(calc_inputs.get("precio_euros")),
                        num(calc_inputs.get("tasa_cambio")),
                        eur(calc_details.get("precio_euros_art")),
                        num(calc_inputs.get("pc_transporte")),
                        num(calc_inputs.get("pc_descarga")),
                        num(calc_inputs.get("pc_varios")),
                        num(calc_inputs.get("pc_manipulacion")),
                        num(calc_inputs.get("pc_financiacion")),
                        eur(calc_details.get("gastos_aplicables")),
                        eur(calc_details.get("coste_sin_almacenaje")),
                        eur(calc_details.get("coste_almacenaje_iva")),
                        eur(calc_details.get("coste_picking_iva")),
                        num(rentabilidad),
                        eur(coste_unitario_final),
                        eur(precio_ponderado_lote),
                        eur(coste_total_cantidad, default=("Pendiente" if item.final_cost == "Bloqueado" else item.final_cost)),
                    )
                )
            else:
                rows.append(
                    common
                    + (
                        eur(precio_base_num),
                        eur(calc_details.get("iva_re")),
                        eur(calc_details.get("precio_con_iva")),
                        eur(calc_details.get("ct_m3_prod")),
                        eur(calc_details.get("cd_prod_iva")),
                        eur(calc_details.get("coste_descarga")),
                        eur(calc_details.get("coste_almacenaje_iva")),
                        eur(calc_details.get("coste_picking_iva")),
                        num(rentabilidad),
                        eur(coste_unitario_final),
                        eur(precio_ponderado_lote),
                        eur(coste_total_cantidad, default=("Pendiente" if item.final_cost == "Bloqueado" else item.final_cost)),
                    )
                )
        return rows

    def _order_detail_value_cards(self, order: SupplierOrder) -> list[tuple[str, str, str]]:
        calc_inputs: dict[str, Any] = {}
        for item in order.items:
            source = item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {}
            candidate = source.get("calculation_inputs")
            if isinstance(candidate, dict) and candidate:
                calc_inputs = candidate
                break
        if order.provider.lower().startswith("hei"):
            return [
                ("Precio en Euros", self._format_eur(calc_inputs.get("precio_euros")) if calc_inputs else "Pendiente", "Info"),
                ("Precio en Dolares", self._format_eur(calc_inputs.get("precio_dolares")) if calc_inputs else "Pendiente", "Info"),
                ("Factura transporte", self._format_eur(calc_inputs.get("factura_transporte")) if calc_inputs else "Pendiente", "Info"),
                ("Derechos aranceles", self._format_eur(calc_inputs.get("derechos_aranceles")) if calc_inputs else "Pendiente", "Info"),
                ("Coste total pedido", order.total_cost, "OK" if order.total_cost != "Bloqueado" else "Error"),
            ]
        return [
            ("Coste transporte + IVA", self._format_eur(calc_inputs.get("coste_transporte_iva")) if calc_inputs else "Pendiente", "Info"),
            ("M3 total camión", self._format_optional_decimal(calc_inputs.get("m3_total_camion"), default="Pendiente") if calc_inputs else "Pendiente", "Info"),
            ("Coste transporte M3", self._format_eur(calc_inputs.get("ct_m3")) if calc_inputs else "Pendiente", "Info"),
            ("Descarga por producto", self._format_eur(calc_inputs.get("cd_prod_iva")) if calc_inputs else "Pendiente", "Info"),
            ("Coste total pedido", order.total_cost, "OK" if order.total_cost != "Bloqueado" else "Error"),
        ]


    def _cancel_supplier_order_from_ui(self, order: SupplierOrder) -> None:
        """Cancela lógicamente un pedido desde la UI ERP."""
        if self._cloud_session is None:
            messagebox.showwarning("Pedidos", "Inicia sesión en Supabase para borrar/cancelar pedidos.")
            return
        raw = order.raw if isinstance(getattr(order, "raw", None), dict) else {}
        order_id = str(raw.get("order_id") or raw.get("id") or "").strip()
        if not order_id:
            messagebox.showwarning(
                "Pedidos",
                "No se encontró el ID real del pedido en Supabase. No se usará el nombre visual para borrar.",
            )
            return
        status = str(order.status or "").strip().lower()
        if status in {"recibido completo", "received", "received_full"}:
            messagebox.showwarning("Pedidos", "No se puede borrar/cancelar directamente un pedido recibido completo.")
            return
        if not messagebox.askyesno(
            "Borrar pedido",
            f"Se cancelará el pedido '{self._order_display_name(order)}'.\n\nNo se borrará histórico ni logs.\n\n¿Continuar?",
        ):
            return
        try:
            cancel_supplier_order(self._cloud_session, order_id, reason="Cancelado desde UI ERP")
        except Exception as exc:
            messagebox.showerror("Pedidos", f"No se pudo cancelar el pedido.\n\n{exc}")
            return

        self._supplier_orders = [
            existing
            for existing in self._supplier_orders
            if str((existing.raw or {}).get("order_id") or (existing.raw or {}).get("id") or "") != order_id
        ]
        self._selected_supplier_order = self._supplier_orders[0] if self._supplier_orders else None
        self._orders_loaded_once = False
        messagebox.showinfo("Pedido cancelado", "El pedido se quitó de la bandeja.")
        self._show_view("calcular")

    def _open_order_detail_window(self, order: SupplierOrder) -> None:
        win = tk.Toplevel(self)
        win.title(f"Pedido - {order.order_id}")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 1240, 720)
        win.minsize(980, 620)
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)
        header = tk.Frame(win, bg=BG, highlightbackground=LINE, highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)
        tk.Label(header, text=f"Detalle completo - {order.order_id}", bg=BG, fg=TEXT, font=("Segoe UI", 20, "bold")).grid(row=0, column=0, sticky="w", padx=22, pady=(18, 2))
        tk.Label(header, text=f"{order.provider} - {order.date} - tabla de calculo y resumen economico", bg=BG, fg=MUTED).grid(row=1, column=0, sticky="w", padx=22, pady=(0, 18))
        self._button(header, "Cerrar", command=win.destroy).grid(row=0, column=1, rowspan=2, padx=22, pady=18)
        body = tk.Frame(win, bg=BG)
        body.grid(row=1, column=0, sticky="nsew", padx=22, pady=22)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)
        self._calculation_tree(body, order.items, provider=order.provider).grid(row=0, column=0, sticky="nsew")
        values = tk.Frame(body, bg=BG)
        values.grid(row=1, column=0, sticky="ew", pady=(14, 0))
        detail_values = self._order_detail_value_cards(order)
        for i, (label, value, status) in enumerate(detail_values):
            values.columnconfigure(i % 4, weight=1)
            self._metric(values, label, value, status).grid(row=i // 4, column=i % 4, sticky="ew", padx=(0 if i % 4 == 0 else 8, 0), pady=(0, 8))
        footer = tk.Frame(win, bg=BG)
        footer.grid(row=2, column=0, sticky="ew", padx=22, pady=(0, 18))
        self._button(footer, "Exportar detalle", primary=True, command=lambda: self._export_supplier_order_audit_excel(order)).pack(side=tk.RIGHT)
        self._button(footer, "Cerrar", command=win.destroy).pack(side=tk.RIGHT, padx=(0, 8))


    def _supplier_order_actual_id(self, order: SupplierOrder | None) -> str:
        if order is None:
            return ""
        raw = order.raw if isinstance(getattr(order, "raw", None), dict) else {}
        return str(raw.get("order_id") or raw.get("id") or "").strip()

    def _safe_sheet_name(self, value: str) -> str:
        clean = re.sub(r"[:\\/?*\[\]]", "_", str(value or "Sheet")).strip() or "Sheet"
        return clean[:31]

    def _autosize_worksheet(self, ws, *, min_width: int = 10, max_width: int = 42) -> None:
        for column_cells in ws.columns:
            max_len = 0
            letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), max_width))
            ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))

    def _audit_value(self, value: Any) -> Any:
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False, default=str)
        return value

    def _excel_number(self, value: Any) -> Any:
        try:
            if value in (None, ""):
                return None
            return float(str(value).replace("€", "").replace("%", "").replace(",", ".").strip())
        except Exception:
            return value

    def _apply_report_sheet_style(self, ws, *, freeze: str = "A2", money_cols: tuple[int, ...] = (), percent_cols: tuple[int, ...] = (), dollar_cols: tuple[int, ...] = ()) -> None:
        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        soft_fill = PatternFill("solid", fgColor="F8FAFC")

        if freeze:
            ws.freeze_panes = freeze

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.row > 1 and cell.row % 2 == 0:
                    cell.fill = soft_fill

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column in money_cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=column).number_format = '#,##0.00 €'

        for column in dollar_cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=column).number_format = '#,##0.00 $'

        for column in percent_cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=column).number_format = '0.00'

        try:
            ws.auto_filter.ref = ws.dimensions
        except Exception:
            pass

        self._autosize_worksheet(ws)

    def _make_report_title(self, ws, title: str, subtitle: str = "", *, last_col: int = 8) -> None:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(size=22, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="0F172A")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 34

        if subtitle:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
            sub_cell = ws.cell(row=2, column=1, value=subtitle)
            sub_cell.font = Font(size=11, color="475569")
            sub_cell.fill = PatternFill("solid", fgColor="E2E8F0")
            sub_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[2].height = 24

    def _write_kpi_card(self, ws, row: int, col: int, label: str, value: Any, *, fill: str = "EEF2FF") -> None:
        ws.cell(row=row, column=col, value=label)
        ws.cell(row=row + 1, column=col, value=value)
        ws.cell(row=row, column=col).font = Font(size=9, bold=True, color="475569")
        ws.cell(row=row + 1, column=col).font = Font(size=16, bold=True, color="0F172A")
        for r in (row, row + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin", color="CBD5E1"),
                right=Side(style="thin", color="CBD5E1"),
                top=Side(style="thin", color="CBD5E1"),
                bottom=Side(style="thin", color="CBD5E1"),
            )

    def _export_supplier_order_audit_excel(
        self,
        order: SupplierOrder | None,
        *,
        provider: str = "",
        values: dict[str, str] | None = None,
        items: tuple[OrderItem, ...] | None = None,
        path: str | None = None,
    ) -> None:
        """Exporta un pedido calculado con formato visual y auditoría completa."""
        provider_name = provider or (order.provider if order else "Proveedor")
        order_name = self._order_display_name(order) if order else str((values or {}).get("Nombre del pedido") or "Pedido calculado")
        values = values or {}
        items = items if items is not None else tuple(order.items if order else tuple())
        if not items:
            messagebox.showinfo("Exportar pedido", "No hay líneas calculadas para exportar.")
            return

        if not path:
            default_name = re.sub(r"[^A-Za-z0-9_-]+", "_", f"pedido_auditoria_{order_name}")[:80] + ".xlsx"
            path = filedialog.asksaveasfilename(
                title="Exportar auditoría de pedido",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
                initialfile=default_name,
            )
        if not path:
            return

        constants = self._current_business_constant_values() if hasattr(self, "_current_business_constant_values") else {}
        wb = Workbook()

        # Datos agregados
        total_qty = 0
        total_line_cost = 0.0
        calculated_count = 0
        pending_count = 0
        unit_cost_values: list[float] = []
        formula_name = ""
        input_candidates: dict[str, Any] = {}

        for item in items:
            source = item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {}
            calc_inputs = source.get("calculation_inputs") if isinstance(source.get("calculation_inputs"), dict) else {}
            if calc_inputs and not input_candidates:
                input_candidates = calc_inputs
            if calc_inputs.get("formula"):
                formula_name = str(calc_inputs.get("formula"))
            total_qty += int(item.quantity or 0)
            line_cost = self._money_float(source.get("line_cost") or source.get("final_cost"), 0.0)
            unit_cost = self._money_float(source.get("precio_coste_final") or source.get("unit_cost"), 0.0)
            total_line_cost += line_cost
            if unit_cost:
                unit_cost_values.append(unit_cost)
            if str(item.status or "").lower() == "calculado":
                calculated_count += 1
            else:
                pending_count += 1

        avg_unit_cost = round(sum(unit_cost_values) / len(unit_cost_values), 2) if unit_cost_values else 0.0
        visible_metrics = self._supplier_order_visible_metrics(items)

        # Hoja portada / resumen
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        self._make_report_title(
            ws_summary,
            f"Pedido {order_name}",
            f"Proveedor: {provider_name} | Auditoría de cálculo FutonHUB",
            last_col=8,
        )

        self._write_kpi_card(ws_summary, 4, 1, "Estado", order.status if order else "Calculado", fill="DCFCE7" if pending_count == 0 else "FEF3C7")
        self._write_kpi_card(ws_summary, 4, 2, "Líneas", len(items), fill="E0F2FE")
        self._write_kpi_card(ws_summary, 4, 3, "Unidades", total_qty, fill="EEF2FF")
        self._write_kpi_card(ws_summary, 4, 4, "Coste total", total_line_cost, fill="ECFDF5")
        self._write_kpi_card(ws_summary, 4, 5, "M3 total", visible_metrics.get("total_m3", 0), fill="E0F2FE")
        self._write_kpi_card(ws_summary, 4, 6, "No descarga", visible_metrics.get("excluded_qty", 0), fill="FEF3C7" if visible_metrics.get("excluded_qty", 0) else "DCFCE7")
        self._write_kpi_card(ws_summary, 4, 7, "Pendientes", pending_count, fill="FEE2E2" if pending_count else "DCFCE7")
        self._write_kpi_card(ws_summary, 4, 8, "Fórmula", formula_name or "-", fill="F1F5F9")

        ws_summary.cell(row=8, column=1, value="Resumen general")
        ws_summary.cell(row=8, column=1).font = Font(size=14, bold=True, color="0F172A")
        summary_rows = [
            ("Pedido", order_name),
            ("ID real Supabase", self._supplier_order_actual_id(order)),
            ("Proveedor", provider_name),
            ("Fecha", order.date if order else values.get("Fecha", "")),
            ("Estado", order.status if order else ""),
            ("Total líneas", len(items)),
            ("Total unidades", total_qty),
            ("Unidades que cuentan para descarga", visible_metrics.get("download_qty", 0)),
            ("Unidades que NO cuentan para descarga", visible_metrics.get("excluded_qty", 0)),
            ("M3 total pedido", visible_metrics.get("total_m3", 0)),
            ("Total coste cantidades", total_line_cost),
            ("Coste medio unitario", avg_unit_cost),
            ("Exportado desde", "UI ERP Pedidos"),
        ]
        row_cursor = 9
        for label, value in summary_rows:
            ws_summary.cell(row=row_cursor, column=1, value=label)
            ws_summary.cell(row=row_cursor, column=2, value=value)
            row_cursor += 1

        ws_summary.cell(row=8, column=4, value="Inputs usados")
        ws_summary.cell(row=8, column=4).font = Font(size=14, bold=True, color="0F172A")
        input_row = 9
        merged_inputs = dict(values)
        for k, v in input_candidates.items():
            if k not in merged_inputs and k != "constants":
                merged_inputs[k] = v
        for key, value in merged_inputs.items():
            ws_summary.cell(row=input_row, column=4, value=key)
            ws_summary.cell(row=input_row, column=5, value=self._audit_value(value))
            input_row += 1

        for row in range(9, max(row_cursor, input_row)):
            for col in [1, 4]:
                ws_summary.cell(row=row, column=col).font = Font(bold=True, color="334155")
                ws_summary.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F1F5F9")

        for col in range(1, 9):
            ws_summary.column_dimensions[get_column_letter(col)].width = 22
        ws_summary.column_dimensions["H"].width = 32
        for row in ws_summary.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(
                    left=Side(style="thin", color="CBD5E1"),
                    right=Side(style="thin", color="CBD5E1"),
                    top=Side(style="thin", color="CBD5E1"),
                    bottom=Side(style="thin", color="CBD5E1"),
                )
        for cell in [ws_summary.cell(4, 4), ws_summary.cell(5, 4), ws_summary.cell(9, 2), ws_summary.cell(16, 2)]:
            cell.number_format = '#,##0.00 €'

        # Constantes
        ws_const = wb.create_sheet("Constantes usadas")
        ws_const.append(("Constante", "Valor usado", "Tipo"))
        for key, value in sorted(constants.items()):
            tipo = "Factor" if str(key).endswith("_FACTOR") else "Constante"
            ws_const.append((key, self._excel_number(value), tipo))
        self._apply_report_sheet_style(ws_const, freeze="A2")
        for row in range(2, ws_const.max_row + 1):
            ws_const.cell(row=row, column=2).number_format = '0.000000'

        # Líneas calculadas
        ws_lines = wb.create_sheet("Líneas calculadas")
        headers = [
            "ID",
            "Nombre",
            "Cantidad",
            "M3 unidad",
            "M3 total línea",
            "Precio proveedor",
            "Precio en Dolares",
            "Precio en Euros",
            "Precio artículo EUR",
            "Tasa cambio",
            "% transporte",
            "% descarga",
            "% varios",
            "% manipulación",
            "% financiación",
            "Transporte ref / Factura",
            "Aranceles / Descarga ref",
            "IVA + RE",
            "Coste descarga",
            "Almacenaje IVA",
            "Picking IVA",
            "Rentabilidad %",
            "Coste Final Artículo",
            "Precio ponderado lote",
            "Coste Total Cantidad",
            "Estado",
            "Motivos / errores",
            "Origen precio proveedor",
            "Match precio",
        ]
        ws_lines.append(headers)

        ws_detail = wb.create_sheet("Detalle fórmula")
        ws_detail.append(("ID", "Nombre", "Grupo", "Campo", "Valor"))

        for item in items:
            source = item.raw.get("source_row") if isinstance(getattr(item, "raw", None), dict) and isinstance(item.raw.get("source_row"), dict) else {}
            calc_inputs = source.get("calculation_inputs") if isinstance(source.get("calculation_inputs"), dict) else {}
            calc_details = source.get("calculation_details") if isinstance(source.get("calculation_details"), dict) else {}
            reasons = source.get("calculation_reasons") or []

            ws_lines.append(
                [
                    item.code,
                    item.name,
                    item.quantity,
                    self._excel_number(source.get("m3_und")),
                    self._excel_number(source.get("m3_total")),
                    self._excel_number(source.get("precio_proveedor")),
                    self._excel_number(calc_inputs.get("precio_dolares")),
                    self._excel_number(calc_inputs.get("precio_euros")),
                    self._excel_number(calc_details.get("precio_euros_art") or calc_details.get("precio_con_iva")),
                    self._excel_number(calc_inputs.get("tasa_cambio")),
                    self._excel_number(calc_inputs.get("pc_transporte")),
                    self._excel_number(calc_inputs.get("pc_descarga")),
                    self._excel_number(calc_inputs.get("pc_varios")),
                    self._excel_number(calc_inputs.get("pc_manipulacion")),
                    self._excel_number(calc_inputs.get("pc_financiacion")),
                    self._excel_number(calc_inputs.get("factura_transporte") if calc_inputs.get("factura_transporte") is not None else calc_details.get("ct_total_ref")),
                    self._excel_number(calc_inputs.get("derechos_aranceles") if calc_inputs.get("derechos_aranceles") is not None else calc_details.get("cd_total_ref")),
                    self._excel_number(calc_details.get("iva_re")),
                    self._excel_number(calc_details.get("coste_descarga")),
                    self._excel_number(calc_details.get("coste_almacenaje_iva")),
                    self._excel_number(calc_details.get("coste_picking_iva")),
                    self._excel_number(source.get("rentabilidad_percent")),
                    self._excel_number(source.get("precio_coste_final") or source.get("unit_cost")),
                    self._excel_number(source.get("precio_ponderado_lote")),
                    self._excel_number(source.get("line_cost") or source.get("final_cost")),
                    item.status,
                    ", ".join(str(reason) for reason in reasons) if isinstance(reasons, list) else str(reasons or ""),
                    source.get("supplier_price_source"),
                    f"{source.get('supplier_price_matched_by') or ''} → {source.get('supplier_price_item_id') or ''}",
                ]
            )

            for group_name, data in [
                ("source_row", source),
                ("calculation_inputs", calc_inputs),
                ("calculation_details", calc_details),
                ("constants_snapshot", calc_inputs.get("constants") if isinstance(calc_inputs, dict) else {}),
            ]:
                if isinstance(data, dict):
                    for key, value in data.items():
                        ws_detail.append((item.code, item.name, group_name, key, self._audit_value(value)))

        money_cols = (8, 9, 16, 17, 18, 19, 20, 21, 23, 24, 25)
        dollar_cols = (6, 7)
        percent_cols = (11, 12, 13, 14, 15, 22)
        self._apply_report_sheet_style(ws_lines, freeze="A2", money_cols=money_cols, percent_cols=percent_cols, dollar_cols=dollar_cols)
        self._apply_report_sheet_style(ws_detail, freeze="A2")

        # Columnes especiales
        ws_lines.column_dimensions["B"].width = 38
        ws_lines.column_dimensions["X"].width = 34
        ws_lines.column_dimensions["Y"].width = 28
        ws_lines.column_dimensions["Z"].width = 24

        # Totales al final de líneas
        total_row = ws_lines.max_row + 2
        ws_lines.cell(row=total_row, column=20, value="TOTALES")
        ws_lines.cell(row=total_row, column=21, value=avg_unit_cost)
        ws_lines.cell(row=total_row, column=22, value="Referencia")
        ws_lines.cell(row=total_row, column=23, value=total_line_cost)
        for col in range(20, 24):
            cell = ws_lines.cell(row=total_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F172A")
            cell.border = Border(
                left=Side(style="thin", color="CBD5E1"),
                right=Side(style="thin", color="CBD5E1"),
                top=Side(style="thin", color="CBD5E1"),
                bottom=Side(style="thin", color="CBD5E1"),
            )
        ws_lines.cell(row=total_row, column=21).number_format = '#,##0.00 €'
        ws_lines.cell(row=total_row, column=23).number_format = '#,##0.00 €'

        # Resaltar estados
        red_fill = PatternFill("solid", fgColor="FEE2E2")
        green_fill = PatternFill("solid", fgColor="DCFCE7")
        yellow_fill = PatternFill("solid", fgColor="FEF3C7")
        status_col = headers.index("Estado") + 1
        reasons_col = headers.index("Motivos / errores") + 1
        for row in range(2, ws_lines.max_row):
            status = str(ws_lines.cell(row=row, column=status_col).value or "").lower()
            reasons = str(ws_lines.cell(row=row, column=reasons_col).value or "")
            if status in {"error", "critical", "bloqueado"} or reasons:
                fill = red_fill
            elif status == "calculado":
                fill = green_fill
            else:
                fill = yellow_fill

            # No reasignamos `cell.fill = cell.fill`: OpenPyXL devuelve StyleProxy
            # y puede lanzar TypeError: unhashable type: 'StyleProxy'.
            # Pintamos solo columnas de estado/errores para mantener la tabla limpia.
            ws_lines.cell(row=row, column=status_col).fill = fill
            ws_lines.cell(row=row, column=reasons_col).fill = fill

        # Ajuste visual final
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False

        wb.save(path)
        messagebox.showinfo("Exportar pedido", f"Auditoría exportada correctamente:\n{path}")

    def _open_receive_modal(self, order: SupplierOrder) -> None:
        if self._cloud_session is None:
            messagebox.showwarning("Pedidos", "Inicia sesión en Supabase para recibir pedidos.")
            return
        actual_order_id = self._supplier_order_actual_id(order) or str(order.order_id or "")
        if not actual_order_id:
            messagebox.showerror("Recepción de pedido", "No se pudo determinar el ID real del pedido.")
            return

        win = tk.Toplevel(self)
        win.title("Recepción de pedido")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 980, 680)
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)

        header = tk.Frame(win, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew", padx=18, pady=(18, 0))
        header.columnconfigure(0, weight=1)
        tk.Label(header, text="Recepción de pedido", bg=CARD, fg=TEXT, font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 2))
        tk.Label(header, text=f"{order.order_id} · {order.provider} · stock interno Supabase", bg=CARD, fg=MUTED).grid(row=1, column=0, sticky="w", padx=18, pady=(0, 16))
        self._button(header, "Cerrar", command=win.destroy).grid(row=0, column=1, rowspan=2, padx=18, pady=16)

        body = tk.Frame(win, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=12)
        body.rowconfigure(2, weight=1)
        body.columnconfigure(0, weight=1)

        notice = tk.Label(
            body,
            text="Preview obligatorio. No toca WooCommerce ni Hexa. Solo suma stock en inventory_items y marca cantidades recibidas.",
            bg=INDIGO_SOFT,
            fg="#3730A3",
            justify=tk.LEFT,
            anchor=tk.W,
            padx=12,
            pady=9,
        )
        notice.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 8))

        controls = tk.Frame(body, bg=CARD)
        controls.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 8))
        controls.columnconfigure(5, weight=1)

        tk.Label(controls, text="Destino", bg=CARD, fg=MUTED, font=("Segoe UI", 8, "bold")).grid(row=0, column=0, sticky="w")
        dest_combo = ttk.Combobox(controls, values=["Almacén", "Tienda"], state="readonly", width=14)
        dest_combo.set("Almacén")
        dest_combo.grid(row=0, column=1, sticky="w", padx=(8, 18))

        mode_var = tk.StringVar(value="Pendiente")
        ttk.Radiobutton(controls, text="Recibir pendiente", variable=mode_var, value="Pendiente").grid(row=0, column=2, sticky="w", padx=(0, 12))
        ttk.Radiobutton(controls, text="Recibir cero", variable=mode_var, value="Cero").grid(row=0, column=3, sticky="w", padx=(0, 12))

        note_var = tk.StringVar()
        tk.Label(controls, text="Nota", bg=CARD, fg=MUTED, font=("Segoe UI", 8, "bold")).grid(row=1, column=0, sticky="w", pady=(8, 0))
        note_entry = tk.Entry(controls, textvariable=note_var, bg="white", fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightthickness=1)
        note_entry.grid(row=1, column=1, columnspan=5, sticky="ew", padx=(8, 0), pady=(8, 0), ipady=6)

        table_host = tk.Frame(body, bg=CARD)
        table_host.grid(row=2, column=0, sticky="nsew", padx=16, pady=(0, 12))
        table_host.rowconfigure(0, weight=1)
        table_host.columnconfigure(0, weight=1)

        columns = ["ID", "Nombre", "Pedida", "Recibida previa", "Recibir ahora", "Pendiente"]
        tree = ttk.Treeview(table_host, columns=columns, show="headings", height=12)
        widths = {"ID": 105, "Nombre": 330, "Pedida": 90, "Recibida previa": 125, "Recibir ahora": 120, "Pendiente": 95}
        for column in columns:
            tree.heading(column, text=column, anchor=tk.CENTER)
            tree.column(column, width=widths[column], anchor=tk.CENTER if column != "Nombre" else tk.W)
        yscroll = ttk.Scrollbar(table_host, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")

        line_state: dict[str, dict[str, Any]] = {}
        for item in order.items:
            raw = item.raw if isinstance(item.raw, dict) else {}
            qty_ordered = self._money_float(raw.get("quantity_ordered") or item.quantity, 0.0)
            qty_prev = self._money_float(raw.get("quantity_received"), 0.0)
            pending = max(0.0, qty_ordered - qty_prev)
            iid = str(raw.get("id") or item.code)
            receive_now = pending
            line_state[iid] = {
                "line_id": raw.get("id"),
                "item_code": item.code,
                "name": item.name,
                "quantity_ordered": qty_ordered,
                "quantity_received_before": qty_prev,
                "quantity_received_now": receive_now,
                "pending": pending,
            }
            tree.insert(
                "",
                tk.END,
                iid=iid,
                values=(
                    item.code,
                    item.name,
                    self._format_optional_decimal(qty_ordered, default="0"),
                    self._format_optional_decimal(qty_prev, default="0"),
                    self._format_optional_decimal(receive_now, default="0"),
                    self._format_optional_decimal(pending, default="0"),
                ),
            )

        def repaint() -> None:
            for iid, data in line_state.items():
                if not tree.exists(iid):
                    continue
                tree.item(
                    iid,
                    values=(
                        data["item_code"],
                        data["name"],
                        self._format_optional_decimal(data["quantity_ordered"], default="0"),
                        self._format_optional_decimal(data["quantity_received_before"], default="0"),
                        self._format_optional_decimal(data["quantity_received_now"], default="0"),
                        self._format_optional_decimal(data["pending"], default="0"),
                    ),
                )

        def apply_mode() -> None:
            mode = mode_var.get()
            for data in line_state.values():
                data["quantity_received_now"] = data["pending"] if mode == "Pendiente" else 0.0
            repaint()

        mode_var.trace_add("write", lambda *_args: apply_mode())

        def edit_selected_qty(_event: object | None = None) -> None:
            selection = tree.selection()
            if not selection:
                return
            iid = selection[0]
            data = line_state.get(iid)
            if not data:
                return
            dialog = tk.Toplevel(win)
            dialog.title("Cantidad recibida")
            dialog.configure(bg=BG)
            dialog.transient(win)
            dialog.grab_set()
            center_window(dialog, 420, 220)
            tk.Label(dialog, text=data["name"], bg=BG, fg=TEXT, font=("Segoe UI", 12, "bold"), wraplength=360, justify=tk.LEFT).pack(anchor=tk.W, padx=18, pady=(18, 6))
            tk.Label(dialog, text=f"Pendiente: {data['pending']:g}", bg=BG, fg=MUTED).pack(anchor=tk.W, padx=18)
            qty_var = tk.StringVar(value=str(data["quantity_received_now"]).replace(".", ","))
            entry = tk.Entry(dialog, textvariable=qty_var, bg="white", fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightthickness=1)
            entry.pack(fill=tk.X, padx=18, pady=14, ipady=8)
            entry.focus_set()

            def accept() -> None:
                value = self._money_float(qty_var.get(), -1)
                if value < 0:
                    messagebox.showerror("Cantidad", "La cantidad recibida no puede ser negativa.")
                    return
                if value > data["pending"]:
                    if not messagebox.askyesno("Cantidad superior", "La cantidad supera lo pendiente. ¿Quieres continuar?"):
                        return
                data["quantity_received_now"] = value
                repaint()
                dialog.destroy()

            footer = tk.Frame(dialog, bg=BG)
            footer.pack(fill=tk.X, padx=18, pady=(0, 18))
            self._button(footer, "Cancelar", command=dialog.destroy).pack(side=tk.RIGHT)
            self._button(footer, "Aceptar", primary=True, command=accept).pack(side=tk.RIGHT, padx=(0, 8))

        tree.bind("<Double-1>", edit_selected_qty)

        footer = tk.Frame(win, bg=BG)
        footer.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))
        footer.columnconfigure(0, weight=1)

        status_var = tk.StringVar(value="Doble click sobre una línea para cambiar la cantidad recibida.")
        tk.Label(footer, textvariable=status_var, bg=BG, fg=MUTED, anchor=tk.W).grid(row=0, column=0, sticky="ew")

        def build_payload() -> list[dict[str, Any]]:
            return [
                {
                    "line_id": data.get("line_id"),
                    "item_code": data.get("item_code"),
                    "quantity_received_now": data.get("quantity_received_now"),
                }
                for data in line_state.values()
                if self._money_float(data.get("quantity_received_now"), 0.0) > 0
            ]

        def destination_key() -> str:
            return "store" if dest_combo.get().lower().startswith("ti") else "warehouse"

        def preview_action() -> dict[str, Any] | None:
            try:
                result = preview_receive_supplier_order(
                    self._cloud_session,
                    order_id=actual_order_id,
                    received_lines=build_payload(),
                    destination=destination_key(),
                    notes=note_var.get().strip(),
                )
            except Exception as exc:
                messagebox.showerror("Preview recepción", str(exc))
                return None
            if result.get("errors"):
                messagebox.showerror("Preview recepción", "\n".join(str(e) for e in result.get("errors", [])))
                return None
            lines = result.get("lines") or []
            msg = [
                "PREVIEW RECEPCIÓN",
                "",
                f"Destino: {dest_combo.get()}",
                f"Líneas a recibir: {len(lines)}",
                f"Unidades a recibir: {result.get('total_receive')}",
                f"Nuevo estado: {result.get('new_status')}",
                "",
            ]
            for line in lines[:12]:
                msg.append(
                    f"{line.get('item_code')} · {line.get('item_name')}: "
                    f"+{line.get('quantity_received_now'):g} "
                    f"Stock tienda {line.get('store_stock_before'):g}→{line.get('store_stock_after'):g} · "
                    f"almacén {line.get('warehouse_stock_before'):g}→{line.get('warehouse_stock_after'):g}"
                )
            if len(lines) > 12:
                msg.append(f"... y {len(lines) - 12} líneas más")
            messagebox.showinfo("Preview recepción", "\n".join(msg))
            return result

        def confirm_action() -> None:
            preview_result = preview_action()
            if not preview_result:
                return
            if not messagebox.askyesno(
                "Confirmar recepción",
                "Se actualizará stock interno en Supabase y el estado del pedido.\n\nNo toca WooCommerce ni Hexa.\n¿Continuar?",
            ):
                return
            try:
                result = receive_supplier_order(
                    self._cloud_session,
                    order_id=actual_order_id,
                    received_lines=build_payload(),
                    destination=destination_key(),
                    notes=note_var.get().strip(),
                )
            except Exception as exc:
                messagebox.showerror("Recepción de pedido", f"No se pudo aplicar la recepción.\n\n{exc}")
                return
            messagebox.showinfo("Recepción aplicada", f"Pedido recibido.\nOperation ID: {result.get('operation_id')}\nEstado: {result.get('order', {}).get('status')}")
            win.destroy()
            self._refresh_supplier_orders()

        self._button(footer, "Cancelar", command=win.destroy).grid(row=0, column=1, padx=(8, 0), sticky="e")
        self._button(footer, "Preview", command=preview_action).grid(row=0, column=2, padx=(8, 0), sticky="e")
        self._button(footer, "Confirmar recepción", primary=True, command=confirm_action).grid(row=0, column=3, padx=(8, 0), sticky="e")


    def _open_delete_order_confirmation(self, order: SupplierOrder) -> None:
        messagebox.showwarning(
            "Borrar pedido",
            f"Confirmacion futura para eliminar/cancelar {order.order_id}.\n\nNo se ha modificado ningun dato real.",
        )

    def _filtered_woo_sync_rows(self) -> list[dict[str, Any]]:
        rows = list(self._woo_sync_rows or [])
        text = str(getattr(self, "_woo_sync_filter_text", "") or "").strip().lower()
        review_filter = str(getattr(self, "_woo_sync_filter_review", "Todos") or "Todos")
        status_filter = str(getattr(self, "_woo_sync_filter_status", "Todos") or "Todos")
        link_filter = str(getattr(self, "_woo_sync_filter_link", "Todos") or "Todos")

        def haystack(row: dict[str, Any]) -> str:
            woo = row.get("woo") or {}
            cls = row.get("classification_after") or {}
            supa = row.get("supabase_match") or {}
            review = row.get("review") or {}
            issues = row.get("issues") or []
            return " ".join(
                str(x or "")
                for x in [
                    woo.get("woo_id"), woo.get("sku"), woo.get("name"), woo.get("categories"),
                    cls.get("family"), cls.get("subgroup"), cls.get("size"), cls.get("materials"),
                    supa.get("item_id"), supa.get("name"), row.get("match_method"),
                    " ".join(review.get("reasons") or []),
                    " ".join(str(issue.get("code") or issue.get("message") or issue) for issue in issues),
                ]
            ).lower()

        def keep(row: dict[str, Any]) -> bool:
            if text and text not in haystack(row):
                return False
            review = row.get("review") or {}
            needs_review = bool(review.get("needs_review"))
            if review_filter == "Solo revisar" and not needs_review:
                return False
            if review_filter == "Solo OK" and needs_review:
                return False
            status = str(row.get("status") or "Info")
            if status_filter != "Todos" and status != status_filter:
                return False
            has_link_candidate = bool((row.get("manual_link_candidate") or {}).get("available"))
            has_match = bool(row.get("supabase_match"))
            if link_filter == "Candidato enlace manual" and not has_link_candidate:
                return False
            if link_filter == "Sin enlace Supabase" and has_match:
                return False
            if link_filter == "Safe to apply later" and not row.get("safe_to_apply_later"):
                return False
            if link_filter == "Pack/composición" and not ((row.get("classification_after") or {}).get("is_pack") or row.get("classification_kind") == "pack_or_composition"):
                return False
            if link_filter == "Medida pendiente" and (row.get("classification_after") or {}).get("size"):
                return False
            if link_filter == "Material pendiente" and (row.get("classification_after") or {}).get("materials"):
                return False
            return True

        return [row for row in rows if keep(row)]

    def _build_woocommerce(self, parent: tk.Frame) -> None:
        self._page_header(
            parent,
            "Gestión",
            "WooCommerce",
            "Lectura, autoclasificación y comparativa contra Supabase. Preview sin escrituras.",
        )

        actions = tk.Frame(parent, bg=BG)
        actions.pack(fill=tk.X, pady=(0, 14))
        self._button(actions, "Sincronizar + Autoclasificar", primary=True, command=lambda: self._refresh_woo_sync_preview(parent)).pack(side=tk.LEFT, padx=(0, 8))
        self._button(actions, "Exportar JSON preview", command=self._export_woo_sync_preview_json).pack(side=tk.LEFT, padx=(0, 8))
        self._button(actions, "Preview publicación precio", command=self._open_woo_publish_preview_modal).pack(side=tk.LEFT)

        if self._woo_sync_error:
            tk.Label(
                parent,
                text=self._woo_sync_error,
                bg=ROSE_SOFT,
                fg=ROSE,
                anchor=tk.W,
                justify=tk.LEFT,
                padx=12,
                pady=9,
                wraplength=980,
            ).pack(fill=tk.X, pady=(0, 14))
        elif self._woo_sync_loading:
            tk.Label(
                parent,
                text="Leyendo WooCommerce, autoclasificando y comparando con Supabase...",
                bg=INDIGO_SOFT,
                fg=INDIGO,
                anchor=tk.W,
                justify=tk.LEFT,
                padx=12,
                pady=9,
            ).pack(fill=tk.X, pady=(0, 14))
        else:
            tk.Label(
                parent,
                text="v53: preview seguro Woo ↔ Supabase. Incluye revisión, edición de clasificación, filtros y enlace manual protegido. WooCommerce no se toca.",
                bg=BLUE_SOFT,
                fg=BLUE,
                anchor=tk.W,
                justify=tk.LEFT,
                padx=12,
                pady=9,
                wraplength=980,
            ).pack(fill=tk.X, pady=(0, 14))

        counters = (self._woo_sync_preview or {}).get("counters", {})
        summary = tk.Frame(parent, bg=BG)
        summary.pack(fill=tk.X, pady=(0, 18))
        metrics = [
            ("Items Woo", str(counters.get("total_woo_items", 0)), "Info"),
            ("Enlazados OK", str(counters.get("linked_ok", 0)), "OK"),
            ("Sin enlace", str(counters.get("no_match", 0)), "Warning" if counters.get("no_match", 0) else "OK"),
            ("Warnings", str(counters.get("warnings", 0)), "Warning" if counters.get("warnings", 0) else "OK"),
            ("Errores/Critical", f"{counters.get('errors', 0)} / {counters.get('critical', 0)}", "Error" if counters.get("errors", 0) or counters.get("critical", 0) else "OK"),
        ]
        for index, (label, value, status) in enumerate(metrics):
            summary.columnconfigure(index, weight=1)
            self._metric(summary, label, value, status).grid(row=0, column=index, sticky="ew", padx=(0 if index == 0 else 8, 0))

        filters = self._card(parent)
        filters.pack(fill=tk.X, pady=(0, 14))
        filters.columnconfigure(1, weight=1)
        tk.Label(filters, text="Filtros de revisión", bg=CARD, fg=TEXT, font=("Segoe UI", 11, "bold")).grid(row=0, column=0, sticky="w", padx=14, pady=(12, 6), columnspan=8)

        filter_text_var = tk.StringVar(value=getattr(self, "_woo_sync_filter_text", ""))
        review_var = tk.StringVar(value=getattr(self, "_woo_sync_filter_review", "Todos"))
        status_var = tk.StringVar(value=getattr(self, "_woo_sync_filter_status", "Todos"))
        link_var = tk.StringVar(value=getattr(self, "_woo_sync_filter_link", "Todos"))

        def apply_filters() -> None:
            self._woo_sync_filter_text = filter_text_var.get().strip()
            self._woo_sync_filter_review = review_var.get()
            self._woo_sync_filter_status = status_var.get()
            self._woo_sync_filter_link = link_var.get()
            if self._current_key == "woocommerce" and self._content.winfo_exists():
                for child in self._content.winfo_children():
                    child.destroy()
                self._build_woocommerce(self._content)

        def clear_filters() -> None:
            filter_text_var.set("")
            review_var.set("Todos")
            status_var.set("Todos")
            link_var.set("Todos")
            apply_filters()

        tk.Label(filters, text="Buscar", bg=CARD, fg=MUTED, font=("Segoe UI", 9, "bold")).grid(row=1, column=0, sticky="w", padx=(14, 6), pady=(0, 12))
        search_entry = tk.Entry(filters, textvariable=filter_text_var, bg="white", fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightthickness=1)
        search_entry.grid(row=1, column=1, sticky="ew", padx=(0, 10), pady=(0, 12), ipady=5)
        search_entry.bind("<Return>", lambda _event: apply_filters())

        tk.Label(filters, text="Revisión", bg=CARD, fg=MUTED, font=("Segoe UI", 9, "bold")).grid(row=1, column=2, sticky="w", padx=(0, 6), pady=(0, 12))
        ttk.Combobox(filters, textvariable=review_var, values=["Todos", "Solo revisar", "Solo OK"], state="readonly", width=14).grid(row=1, column=3, sticky="ew", padx=(0, 10), pady=(0, 12), ipady=3)
        tk.Label(filters, text="Estado", bg=CARD, fg=MUTED, font=("Segoe UI", 9, "bold")).grid(row=1, column=4, sticky="w", padx=(0, 6), pady=(0, 12))
        ttk.Combobox(filters, textvariable=status_var, values=["Todos", "OK", "Info", "Warning", "Error", "Critical"], state="readonly", width=12).grid(row=1, column=5, sticky="ew", padx=(0, 10), pady=(0, 12), ipady=3)
        tk.Label(filters, text="Caso", bg=CARD, fg=MUTED, font=("Segoe UI", 9, "bold")).grid(row=1, column=6, sticky="w", padx=(0, 6), pady=(0, 12))
        ttk.Combobox(filters, textvariable=link_var, values=["Todos", "Candidato enlace manual", "Sin enlace Supabase", "Safe to apply later", "Pack/composición", "Medida pendiente", "Material pendiente"], state="readonly", width=23).grid(row=1, column=7, sticky="ew", padx=(0, 10), pady=(0, 12), ipady=3)
        self._button(filters, "Aplicar", primary=True, command=apply_filters).grid(row=1, column=8, sticky="ew", padx=(0, 8), pady=(0, 12))
        self._button(filters, "Limpiar", command=clear_filters).grid(row=1, column=9, sticky="ew", padx=(0, 14), pady=(0, 12))

        visible_count = len(self._filtered_woo_sync_rows())
        total_count = len(self._woo_sync_rows or [])
        tk.Label(filters, text=f"Mostrando {visible_count} de {total_count} líneas", bg=CARD, fg=MUTED, font=("Segoe UI", 9)).grid(row=2, column=0, columnspan=10, sticky="w", padx=14, pady=(0, 12))

        body = tk.Frame(parent, bg=BG)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=4)
        body.columnconfigure(1, weight=2)
        body.rowconfigure(0, weight=1)

        table_card = self._card(body)
        table_card.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        table_card.columnconfigure(0, weight=1)
        table_card.rowconfigure(1, weight=1)

        table_head = tk.Frame(table_card, bg=CARD)
        table_head.grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 10))
        table_head.columnconfigure(0, weight=1)
        tk.Label(table_head, text="Preview Woo ↔ Supabase", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w")
        self._status_chip(table_head, "Preview sin escrituras", "Info").grid(row=0, column=1, sticky="e")

        table_frame = tk.Frame(table_card, bg=CARD)
        table_frame.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        columns = ["Estado", "Revisión", "Woo ID", "Tipo", "SKU", "Nombre Woo", "Item Supabase", "Familia", "Medida", "Materiales", "Enlace manual", "Acción propuesta"]
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=14)
        widths = {
            "Estado": 90,
            "Revisión": 110,
            "Woo ID": 90,
            "Tipo": 90,
            "SKU": 120,
            "Nombre Woo": 250,
            "Item Supabase": 130,
            "Familia": 140,
            "Medida": 110,
            "Materiales": 140,
            "Enlace manual": 125,
            "Acción propuesta": 190,
        }
        for column in columns:
            tree.heading(column, text=column, anchor=tk.CENTER)
            tree.column(column, width=widths[column], anchor=tk.CENTER if column not in {"Nombre Woo", "Acción propuesta"} else tk.W, stretch=False)

        for status, color in [("OK", "#F0FDF4"), ("Info", "#EFF6FF"), ("Warning", "#FFFBEB"), ("Error", "#FFF7ED"), ("Critical", "#FFF1F2")]:
            tree.tag_configure(status, background=color)

        row_by_iid: dict[str, dict[str, Any]] = {}
        rows = self._filtered_woo_sync_rows()
        if not rows:
            empty_msg = "Pulsa Sincronizar + Autoclasificar" if not self._woo_sync_rows else "Sin resultados con estos filtros"
            tree.insert("", tk.END, values=("Info", "", "", "", "", empty_msg, "", "", "", "", "", ""))
        for row in rows:
            woo = row.get("woo") or {}
            cls = row.get("classification_after") or {}
            supa = row.get("supabase_match") or {}
            issues = row.get("issues") or []
            action = "Sin cambios"
            if row.get("proposed_supabase_update"):
                action = "Rellenar campos vacíos en Supabase"
            elif not supa:
                action = "Revisar / crear enlace"
            elif issues:
                action = "Revisar incidencias"
            status = str(row.get("status") or "Info")
            iid = tree.insert(
                "",
                tk.END,
                values=(
                    status,
                    "REVISAR" if (row.get("review") or {}).get("needs_review") else "OK",
                    woo.get("woo_id") or "",
                    woo.get("item_kind") or "",
                    woo.get("sku") or "",
                    woo.get("name") or "",
                    supa.get("item_id") or "-",
                    cls.get("family") or "-",
                    cls.get("size") or "-",
                    cls.get("materials") or "-",
                    "Disponible" if (row.get("manual_link_candidate") or {}).get("available") else "No",
                    action,
                ),
                tags=(status,),
            )
            row_by_iid[iid] = row

        yscroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        xscroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        detail_host = tk.Frame(body, bg=BG)
        detail_host.grid(row=0, column=1, sticky="nsew")

        def render_detail(row: dict[str, Any] | None) -> None:
            self._render_woo_sync_detail(detail_host, row)

        def on_select(_event: object | None = None) -> None:
            selection = tree.selection()
            if not selection:
                return
            render_detail(row_by_iid.get(selection[0]))

        tree.bind("<<TreeviewSelect>>", on_select)
        render_detail(rows[0] if rows else None)

    def _refresh_woo_sync_preview(self, parent: tk.Frame) -> None:
        if self._cloud_session is None:
            messagebox.showwarning("WooCommerce", "Inicia sesión en Supabase para comparar Woo con inventario.")
            return
        self._woo_sync_loading = True
        self._woo_sync_error = ""
        overlay = self._show_working_overlay(
            "Sincronizando WooCommerce",
            "Leyendo productos y variaciones, autoclasificando y comparando contra Supabase.\nNo cierres esta ventana.",
        )
        self._build_woocommerce(parent)

        def worker() -> None:
            try:
                preview = build_woo_sync_preview(self._cloud_session)
                self.after(0, lambda: (self._close_working_overlay(overlay), self._finish_woo_sync_preview(parent, preview, "")))
            except Exception as exc:
                self.after(0, lambda exc=exc: (self._close_working_overlay(overlay), self._finish_woo_sync_preview(parent, None, f"No se pudo sincronizar WooCommerce: {exc}")))

        import threading
        threading.Thread(target=worker, daemon=True).start()

    def _finish_woo_sync_preview(self, parent: tk.Frame, preview: dict[str, Any] | None, error: str) -> None:
        self._woo_sync_loading = False
        self._woo_sync_error = error
        if preview is not None:
            self._woo_sync_preview = preview
            self._woo_sync_rows = list(preview.get("items") or [])
        # Auto-refresh real: parent puede haber sido destruido al reconstruir la vista.
        # Si seguimos en WooCommerce, redibujamos usando self._content.
        if self._current_key == "woocommerce" and self._content.winfo_exists():
            for child in self._content.winfo_children():
                child.destroy()
            self._build_woocommerce(self._content)

    def _open_woo_classification_edit_modal(self, row: dict[str, Any]) -> None:
        if not row:
            return
        woo = row.get("woo") or {}
        cls = row.get("classification_after") or {}

        win = tk.Toplevel(self)
        win.title("Editar clasificación preview")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 620, 590)
        win.columnconfigure(0, weight=1)

        tk.Label(win, text="Editar clasificación preview", bg=BG, fg=TEXT, font=("Segoe UI", 17, "bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(18, 4))
        tk.Label(win, text=str(woo.get("name") or "-"), bg=BG, fg=MUTED, wraplength=560, justify=tk.LEFT).grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 14))

        form = self._card(win)
        form.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 12))
        form.columnconfigure(1, weight=1)

        family_values = ["Futones", "Tatamis", "Complementos", "Sofás Cama", "Camas Japonesas", "Ofertas / Packs", "Otros / Sin clasificar"]
        subgroup_values = ["Futón", "Tatami", "Funda futón", "Cojines", "Topper", "Mesita", "Sofá cama", "Cama japonesa", "Base tatami", "Pack futón + funda", "Pack futón + cojines", "Pack futón + funda + cojines", "Pack tatami + futón", "Pack", "Complemento", ""]
        status_values = ["Normal", "Outlet", "Oferta"]
        confidence_values = ["Alta", "Media", "Baja"]
        kind_values = ["single_item", "pack_or_composition"]

        vars_by_field: dict[str, tk.StringVar] = {}

        def add_combo(row_index: int, label: str, field: str, values: list[str]) -> None:
            tk.Label(form, text=label, bg=CARD, fg=MUTED, font=("Segoe UI", 9, "bold")).grid(row=row_index, column=0, sticky="w", padx=14, pady=(10, 2))
            var = tk.StringVar(value=str(cls.get(field) or ""))
            vars_by_field[field] = var
            combo = ttk.Combobox(form, textvariable=var, values=values, state="readonly")
            combo.grid(row=row_index, column=1, sticky="ew", padx=(8, 14), pady=(10, 2), ipady=4)

        def add_entry(row_index: int, label: str, field: str) -> None:
            tk.Label(form, text=label, bg=CARD, fg=MUTED, font=("Segoe UI", 9, "bold")).grid(row=row_index, column=0, sticky="w", padx=14, pady=(10, 2))
            var = tk.StringVar(value=str(cls.get(field) or ""))
            vars_by_field[field] = var
            entry = tk.Entry(form, textvariable=var, bg="white", fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightthickness=1)
            entry.grid(row=row_index, column=1, sticky="ew", padx=(8, 14), pady=(10, 2), ipady=6)

        add_combo(0, "Familia", "family", family_values)
        add_combo(1, "Subgrupo", "subgroup", subgroup_values)
        add_entry(2, "Medida", "size")
        add_entry(3, "Materiales", "materials")
        add_combo(4, "Estado comercial", "commercial_status", status_values)
        add_combo(5, "Es pack", "is_pack", ["0", "1"])
        add_combo(6, "Confianza", "confidence", confidence_values)
        add_combo(7, "Tipo clasificación", "classification_kind", kind_values)

        tk.Label(
            win,
            text="Esto solo modifica el preview y el JSON exportado. No escribe en Supabase.",
            bg=INDIGO_SOFT,
            fg=INDIGO,
            anchor=tk.W,
            justify=tk.LEFT,
            padx=12,
            pady=9,
            wraplength=560,
        ).grid(row=3, column=0, sticky="ew", padx=18, pady=(0, 12))

        footer = tk.Frame(win, bg=BG)
        footer.grid(row=4, column=0, sticky="ew", padx=18, pady=(0, 18))

        def apply_edit() -> None:
            edited = {field: var.get() for field, var in vars_by_field.items()}
            try:
                updated = apply_manual_classification_edit(row, edited)
            except Exception as exc:
                messagebox.showerror("Editar clasificación", f"No se pudo aplicar el ajuste.\n\n{exc}")
                return
            target_woo_id = (row.get("woo") or {}).get("woo_id")
            target_parent_id = (row.get("woo") or {}).get("parent_woo_id")
            target_sku = (row.get("woo") or {}).get("sku")
            replaced = False
            for idx, existing in enumerate(self._woo_sync_rows):
                existing_woo = existing.get("woo") or {}
                if existing_woo.get("woo_id") == target_woo_id and existing_woo.get("parent_woo_id") == target_parent_id and existing_woo.get("sku") == target_sku:
                    self._woo_sync_rows[idx] = updated
                    replaced = True
                    break
            if self._woo_sync_preview is not None:
                items = list(self._woo_sync_preview.get("items") or [])
                for idx, existing in enumerate(items):
                    existing_woo = existing.get("woo") or {}
                    if existing_woo.get("woo_id") == target_woo_id and existing_woo.get("parent_woo_id") == target_parent_id and existing_woo.get("sku") == target_sku:
                        items[idx] = updated
                        break
                self._woo_sync_preview["items"] = items
                self._woo_sync_preview["manual_edits_applied"] = True
            win.destroy()
            if self._current_key == "woocommerce" and self._content.winfo_exists():
                for child in self._content.winfo_children():
                    child.destroy()
                self._build_woocommerce(self._content)

        self._button(footer, "Cancelar", command=win.destroy).pack(side=tk.RIGHT)
        self._button(footer, "Aplicar al preview", primary=True, command=apply_edit).pack(side=tk.RIGHT, padx=(0, 8))

    def _open_woo_manual_link_modal(self, row: dict[str, Any]) -> None:
        if self._cloud_session is None:
            messagebox.showwarning("WooCommerce", "Inicia sesión en Supabase para enlazar manualmente.")
            return
        manual_link = row.get("manual_link_candidate") or {}
        if not manual_link.get("available"):
            messagebox.showwarning("Enlace manual", manual_link.get("reason") or "Este Woo no es candidato para enlace manual.")
            return
        woo = row.get("woo") or {}
        win = tk.Toplevel(self)
        win.title("Enlace manual Woo ↔ Supabase")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 980, 640)
        win.columnconfigure(0, weight=1)
        win.rowconfigure(3, weight=1)

        tk.Label(win, text="Enlace manual Woo ↔ Supabase", bg=BG, fg=TEXT, font=("Segoe UI", 17, "bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(18, 4))
        tk.Label(win, text=f"Woo {woo.get('woo_id')} · SKU {woo.get('sku') or '-'} · {woo.get('name') or '-'}", bg=BG, fg=MUTED, wraplength=900, justify=tk.LEFT).grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 12))

        search_card = self._card(win)
        search_card.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 12))
        search_card.columnconfigure(1, weight=1)
        query_var = tk.StringVar(value=str(woo.get("sku") or woo.get("name") or "")[:80])
        tk.Label(search_card, text="Buscar item Supabase sin Woo", bg=CARD, fg=MUTED, font=("Segoe UI", 9, "bold")).grid(row=0, column=0, sticky="w", padx=14, pady=12)
        query_entry = tk.Entry(search_card, textvariable=query_var, bg="white", fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightthickness=1)
        query_entry.grid(row=0, column=1, sticky="ew", padx=(0, 10), pady=12, ipady=5)

        table_card = self._card(win)
        table_card.grid(row=3, column=0, sticky="nsew", padx=18, pady=(0, 12))
        table_card.columnconfigure(0, weight=1)
        table_card.rowconfigure(0, weight=1)
        cols = ["item_id", "Nombre", "Familia", "Medida", "Materiales", "Stock", "Woo"]
        tree = ttk.Treeview(table_card, columns=cols, show="headings", height=10)
        widths = {"item_id": 90, "Nombre": 330, "Familia": 130, "Medida": 110, "Materiales": 130, "Stock": 90, "Woo": 130}
        for col in cols:
            tree.heading(col, text=col, anchor=tk.CENTER)
            tree.column(col, width=widths[col], anchor=tk.CENTER if col != "Nombre" else tk.W, stretch=False)
        yscroll = ttk.Scrollbar(table_card, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.grid(row=0, column=0, sticky="nsew", padx=(14, 0), pady=14)
        yscroll.grid(row=0, column=1, sticky="ns", pady=14, padx=(0, 14))
        rows_by_iid: dict[str, dict[str, Any]] = {}

        preview_box = tk.Text(win, bg="#0F172A", fg="#E2E8F0", insertbackground="#E2E8F0", relief=tk.FLAT, wrap=tk.WORD, font=("Consolas", 9), height=8)
        preview_box.grid(row=4, column=0, sticky="ew", padx=18, pady=(0, 12))
        preview_box.insert("1.0", "Selecciona un item Supabase y pulsa Preview enlace.")
        preview_box.configure(state=tk.DISABLED)

        footer = tk.Frame(win, bg=BG)
        footer.grid(row=5, column=0, sticky="ew", padx=18, pady=(0, 18))
        footer.columnconfigure(0, weight=1)

        selected_preview: dict[str, Any] | None = None

        def set_preview_text(text: str) -> None:
            preview_box.configure(state=tk.NORMAL)
            preview_box.delete("1.0", tk.END)
            preview_box.insert("1.0", text)
            preview_box.configure(state=tk.DISABLED)

        def search_candidates() -> None:
            for child in tree.get_children():
                tree.delete(child)
            rows_by_iid.clear()
            try:
                candidates = search_manual_link_inventory_candidates(self._cloud_session, query_var.get(), limit=50)
            except Exception as exc:
                set_preview_text(f"No se pudo buscar candidatos:\n{exc}")
                return
            if not candidates:
                tree.insert("", tk.END, values=("", "Sin candidatos sin Woo", "", "", "", "", ""))
                return
            for cand in candidates:
                iid = tree.insert("", tk.END, values=(
                    cand.get("item_id") or "",
                    cand.get("name") or "-",
                    cand.get("family") or "-",
                    cand.get("size") or "-",
                    cand.get("materials") or "-",
                    f"{cand.get('store_stock') or 0}/{cand.get('warehouse_stock') or 0}",
                    cand.get("woo_link_status") or "Sin Woo",
                ))
                rows_by_iid[iid] = cand

        def selected_item_id() -> int | None:
            selection = tree.selection()
            if not selection:
                return None
            cand = rows_by_iid.get(selection[0])
            if not cand:
                return None
            try:
                return int(cand.get("item_id"))
            except Exception:
                return None

        def preview_link() -> None:
            nonlocal selected_preview
            item_id = selected_item_id()
            if item_id is None:
                messagebox.showwarning("Enlace manual", "Selecciona primero un item Supabase candidato.", parent=win)
                return
            try:
                selected_preview = preview_manual_woo_link(self._cloud_session, row, item_id)
                set_preview_text(format_manual_woo_link_preview(selected_preview))
            except Exception as exc:
                selected_preview = None
                set_preview_text(f"Enlace bloqueado:\n{exc}")

        def apply_link() -> None:
            item_id = selected_item_id()
            if item_id is None:
                messagebox.showwarning("Enlace manual", "Selecciona primero un item Supabase candidato.", parent=win)
                return
            try:
                preview = preview_manual_woo_link(self._cloud_session, row, item_id)
            except Exception as exc:
                messagebox.showerror("Enlace manual", f"No se puede aplicar.\n\n{exc}", parent=win)
                return
            token = simpledialog.askstring("Confirmar enlace", format_manual_woo_link_preview(preview) + "\n\nEscribe ENLAZAR para aplicar:", parent=win)
            if str(token or "").strip().upper() != "ENLAZAR":
                return
            overlay = self._show_working_overlay("Enlazando Woo ↔ Supabase", "Aplicando enlace manual con snapshot y audit_log. WooCommerce no se toca.")
            try:
                result = apply_manual_woo_link(self._cloud_session, row, item_id, load_settings())
            except Exception as exc:
                self._close_working_overlay(overlay)
                messagebox.showerror("Enlace manual", f"No se pudo aplicar el enlace.\n\n{exc}", parent=win)
                return
            self._close_working_overlay(overlay)
            messagebox.showinfo("Enlace manual", f"Enlace aplicado.\nOperation ID: {result.get('operation_id')}", parent=win)
            win.destroy()
            self._refresh_woo_sync_preview(self._content)

        self._button(search_card, "Buscar", primary=True, command=search_candidates).grid(row=0, column=2, sticky="ew", padx=(0, 14), pady=12)
        query_entry.bind("<Return>", lambda _event: search_candidates())
        self._button(footer, "Cerrar", command=win.destroy).grid(row=0, column=1, padx=(8, 0), sticky="e")
        self._button(footer, "Preview enlace", command=preview_link).grid(row=0, column=2, padx=(8, 0), sticky="e")
        self._button(footer, "Aplicar enlace", primary=True, command=apply_link).grid(row=0, column=3, padx=(8, 0), sticky="e")
        search_candidates()

    def _export_woo_sync_preview_json(self) -> None:
        if not self._woo_sync_preview:
            messagebox.showinfo("WooCommerce", "Primero ejecuta Sincronizar + Autoclasificar.")
            return
        path = filedialog.asksaveasfilename(
            title="Exportar preview Woo JSON",
            defaultextension=".json",
            filetypes=[("JSON", "*.json")],
            initialfile="futonhub_woo_sync_preview.json",
        )
        if not path:
            return
        try:
            export_woo_preview_json(self._woo_sync_preview, path)
        except Exception as exc:
            messagebox.showerror("Exportar JSON", f"No se pudo exportar.\n\n{exc}")
            return
        messagebox.showinfo("Exportar JSON", f"JSON generado:\n{path}")

    def _render_woo_sync_detail(self, parent: tk.Frame, row: dict[str, Any] | None) -> None:
        for child in parent.winfo_children():
            child.destroy()
        outer = self._card(parent)
        outer.pack(fill=tk.BOTH, expand=True)
        outer.rowconfigure(0, weight=1)
        outer.columnconfigure(0, weight=1)

        canvas = tk.Canvas(outer, bg=CARD, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
        detail = tk.Frame(canvas, bg=CARD)
        detail.columnconfigure(0, weight=1)
        detail_window = canvas.create_window((0, 0), window=detail, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        def _sync_scroll_region(_event: object | None = None) -> None:
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _sync_width(event: object) -> None:
            try:
                canvas.itemconfigure(detail_window, width=event.width)
            except Exception:
                pass

        detail.bind("<Configure>", _sync_scroll_region)
        canvas.bind("<Configure>", _sync_width)
        canvas.bind("<MouseWheel>", lambda event: canvas.yview_scroll(int(-1 * (event.delta / 120)), "units"))

        if not row:
            tk.Label(detail, text="Detalle", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=18, pady=(16, 8))
            self._status_row(detail, "Sin selección", "Ejecuta la sincronización y selecciona una línea.", "Info").pack(fill=tk.X, padx=18, pady=8)
            return

        woo = row.get("woo") or {}
        cls = row.get("classification_after") or {}
        supa = row.get("supabase_match") or {}
        proposed = row.get("proposed_supabase_update") or {}
        issues = row.get("issues") or []
        status = str(row.get("status") or "Info")

        top = tk.Frame(detail, bg=CARD)
        top.pack(fill=tk.X, padx=18, pady=(16, 10))
        top.columnconfigure(0, weight=1)
        tk.Label(top, text="Detalle Woo Sync", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w")
        self._status_chip(top, status, status).grid(row=0, column=1, sticky="e")

        tk.Label(detail, text=str(woo.get("name") or "-"), bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold"), wraplength=330, justify=tk.LEFT).pack(anchor=tk.W, padx=18, pady=(0, 2))
        tk.Label(detail, text=f"Woo {woo.get('woo_id')} · SKU {woo.get('sku') or '-'} · {woo.get('item_kind')}", bg=CARD, fg=MUTED).pack(anchor=tk.W, padx=18, pady=(0, 12))

        for label, value in [
            ("Match", row.get("match_method") or "-"),
            ("Item Supabase", supa.get("item_id") or "-"),
            ("Familia detectada", cls.get("family") or "-"),
            ("Subgrupo", cls.get("subgroup") or "-"),
            ("Medida", cls.get("size") or "-"),
            ("Materiales", cls.get("materials") or "-"),
            ("Confianza", cls.get("confidence") or "-"),
            ("Precio Woo", woo.get("price") or "-"),
            ("Categorías", woo.get("categories") or "-"),
        ]:
            self._detail_row(detail, label, value).pack(fill=tk.X, padx=18, pady=3)

        review = row.get("review") or {}
        if review.get("needs_review"):
            self._status_row(
                detail,
                "REVISAR",
                " · ".join(review.get("reasons") or ["Revisión pendiente"]),
                review.get("severity") or "Warning",
            ).pack(fill=tk.X, padx=18, pady=(10, 4))
        else:
            self._status_row(detail, "OK", "Sin indicadores de revisión.", "OK").pack(fill=tk.X, padx=18, pady=(10, 4))

        manual_link = row.get("manual_link_candidate") or {}
        tk.Label(detail, text="Enlace manual Supabase ↔ Woo", bg=CARD, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, padx=18, pady=(12, 4))
        if manual_link.get("available"):
            self._status_row(
                detail,
                "Disponible",
                "Este Woo no parece enlazado. En v53 puedes enlazarlo con un item Supabase sin Woo, con preview y bloqueo de duplicados.",
                "Warning",
            ).pack(fill=tk.X, padx=18, pady=4)
        else:
            self._status_row(
                detail,
                "No disponible",
                str(manual_link.get("reason") or "No es candidato para enlace manual."),
                "Info",
            ).pack(fill=tk.X, padx=18, pady=4)

        tk.Label(detail, text="Propuesta segura para Supabase", bg=CARD, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, padx=18, pady=(12, 4))
        if proposed:
            for key, value in proposed.items():
                self._detail_row(detail, key, value).pack(fill=tk.X, padx=18, pady=2)
        else:
            self._status_row(detail, "Sin escritura propuesta", "No hay cambios seguros o requiere revisión manual.", "Info").pack(fill=tk.X, padx=18, pady=4)

        tk.Label(detail, text="Incidencias", bg=CARD, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, padx=18, pady=(12, 4))
        if not issues:
            self._status_row(detail, "OK", "Sin incidencias detectadas.", "OK").pack(fill=tk.X, padx=18, pady=4)
        for issue in issues[:8]:
            self._status_row(detail, str(issue.get("code") or issue.get("severity") or "Incidencia"), str(issue.get("message") or issue), str(issue.get("severity") or "Warning")).pack(fill=tk.X, padx=18, pady=4)

        actions = tk.Frame(detail, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        actions.pack(fill=tk.X, padx=18, pady=(14, 18))
        self._button(actions, "Editar clasificación preview", primary=True, command=lambda r=row: self._open_woo_classification_edit_modal(r)).pack(fill=tk.X, padx=12, pady=(12, 6))
        if manual_link.get("available"):
            self._button(actions, "Enlazar con item Supabase", command=lambda r=row: self._open_woo_manual_link_modal(r)).pack(fill=tk.X, padx=12, pady=(0, 6))
        self._button(actions, "Ver JSON línea", command=lambda r=row: self._open_json_value_modal("Woo Sync JSON", r)).pack(fill=tk.X, padx=12, pady=(0, 12))


    def _render_woo_detail(self, parent: tk.Frame, difference: WooDifference) -> None:
        for child in parent.winfo_children():
            child.destroy()
        detail = self._card(parent)
        detail.pack(fill=tk.BOTH, expand=True)
        detail.rowconfigure(0, weight=1)
        detail.columnconfigure(0, weight=1)

        scroll_area = tk.Frame(detail, bg=CARD)
        scroll_area.grid(row=0, column=0, sticky="nsew", padx=18, pady=(16, 10))
        top = tk.Frame(scroll_area, bg=CARD)
        top.pack(fill=tk.X, pady=(0, 12))
        top.columnconfigure(0, weight=1)
        tk.Label(top, text="Detalle", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w")
        self._status_chip(top, difference.action, difference.status).grid(row=0, column=1, sticky="e")
        tk.Label(scroll_area, text=difference.name, bg=CARD, fg=TEXT, font=("Segoe UI", 16, "bold"), wraplength=330, justify=tk.LEFT).pack(
            anchor=tk.W,
            pady=(0, 2),
        )
        tk.Label(scroll_area, text=f"{difference.local_id} - Woo ID {difference.woo_id}", bg=CARD, fg=MUTED, font=("Segoe UI", 9)).pack(
            anchor=tk.W,
            pady=(0, 14),
        )

        for label, value in [
            ("Campo", difference.field),
            ("Base local", difference.local_value),
            ("WooCommerce", difference.woo_value),
            ("Diferencia", difference.difference),
            ("Clasificacion", difference.classification),
            ("Accion", difference.action),
            ("Estado", difference.status),
        ]:
            self._detail_row(scroll_area, label, value).pack(fill=tk.X, pady=4)

        tk.Label(
            scroll_area,
            text=difference.detail,
            bg=AMBER_SOFT if difference.status in {"Warning", "Error", "Critical"} else BLUE_SOFT,
            fg="#92400E" if difference.status in {"Warning", "Error", "Critical"} else BLUE,
            wraplength=330,
            justify=tk.LEFT,
            padx=12,
            pady=10,
        ).pack(fill=tk.X, pady=(14, 0))

        for title, subtitle, status in [
            ("Actualizar base de datos", "Trae cambios detectados desde WooCommerce hacia la base local del HUB.", "Info"),
            ("Auto-clasificar", "Propone familia, subgrupo, materiales y medidas usando nombre, atributos y variaciones.", "Info"),
            ("Revision manual", "Los casos dudosos quedan marcados para revision antes de guardarse.", "Warning"),
        ]:
            self._status_row(scroll_area, title, subtitle, status).pack(fill=tk.X, pady=(10, 0))

        actions = tk.Frame(detail, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        actions.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        self._button(actions, "Actualizar base de datos", primary=True, command=lambda: self._open_woo_action_modal("Actualizar base de datos")).pack(
            fill=tk.X,
            padx=12,
            pady=(12, 7),
        )
        row = tk.Frame(actions, bg=CARD)
        row.pack(fill=tk.X, padx=12, pady=(0, 12))
        row.columnconfigure(0, weight=1)
        row.columnconfigure(1, weight=1)
        self._button(row, "Auto-clasificar", command=lambda: self._open_woo_action_modal("Auto-clasificar")).grid(row=0, column=0, sticky="ew", padx=(0, 6))
        self._button(row, "Revisar manual", command=lambda: self._open_woo_action_modal("Revisar manual")).grid(row=0, column=1, sticky="ew", padx=(6, 0))

    def _open_woo_publish_preview_modal(self) -> None:
        if self._cloud_session is None:
            messagebox.showinfo("WooCommerce", "Inicia sesion para generar preview real de publicacion.")
            return
        win = tk.Toplevel(self)
        win.title("Preview publicacion WooCommerce")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 820, 600)
        win.minsize(720, 500)

        body = tk.Frame(win, bg=BG)
        body.pack(fill=tk.BOTH, expand=True, padx=18, pady=18)
        tk.Label(body, text="Preview publicacion WooCommerce", bg=BG, fg=TEXT, font=("Segoe UI", 18, "bold")).pack(anchor=tk.W)
        tk.Label(
            body,
            text="Evalua propuestas aprobadas contra WooCommerce. Solo lectura: no publica, no sincroniza y no ejecuta PUT.",
            bg=BG,
            fg=MUTED,
        ).pack(anchor=tk.W, pady=(2, 12))

        result = tk.Text(body, bg="#0F172A", fg="#E2E8F0", insertbackground="#E2E8F0", relief=tk.FLAT, wrap=tk.WORD, font=("Consolas", 9))
        result.pack(fill=tk.BOTH, expand=True)
        result.insert("1.0", "Generando preview real...")
        result.configure(state=tk.DISABLED)

        footer = tk.Frame(body, bg=BG)
        footer.pack(fill=tk.X, pady=(12, 0))
        self._button(footer, "Cerrar", command=win.destroy).pack(side=tk.RIGHT)

        def worker() -> None:
            try:
                preview = preview_woocommerce_publish(self._cloud_session, limit=50, settings=load_settings())
                text = format_woocommerce_publish_preview(preview)
            except Exception as exc:
                text = f"No se pudo generar preview WooCommerce:\n{exc}"
            self.after(0, lambda: render(text))

        def render(text: str) -> None:
            if not win.winfo_exists():
                return
            result.configure(state=tk.NORMAL)
            result.delete("1.0", tk.END)
            result.insert("1.0", text)
            result.configure(state=tk.DISABLED)

        threading.Thread(target=worker, daemon=True).start()

    def _open_woo_action_modal(self, action: str) -> None:
        difference = self._selected_woo_difference
        win = tk.Toplevel(self)
        win.title(action)
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 720, 420)
        win.columnconfigure(0, weight=1)

        header = tk.Frame(win, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew", padx=18, pady=(18, 0))
        header.columnconfigure(0, weight=1)
        tk.Label(header, text=action, bg=CARD, fg=TEXT, font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 2))
        tk.Label(
            header,
            text=f"{difference.local_id} - {difference.name}",
            bg=CARD,
            fg=MUTED,
            font=("Segoe UI", 10),
        ).grid(row=1, column=0, sticky="w", padx=18, pady=(0, 16))
        self._button(header, "Cerrar", command=win.destroy).grid(row=0, column=1, rowspan=2, sticky="e", padx=18, pady=16)

        body = tk.Frame(win, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        body.grid(row=1, column=0, sticky="ew", padx=18, pady=12)
        if difference.status == "Critical":
            title = "Operacion bloqueada por Critical"
            text = "El prototipo no permite avanzar con cambios mientras exista un Critical seleccionado."
            status = "Critical"
        else:
            title = "Preview visual"
            text = "Esta accion todavia no toca datos reales. La conexion real llegara despues con validacion, confirmacion y log."
            status = "Info"
        self._status_row(body, title, text, status).pack(fill=tk.X, padx=16, pady=(16, 8))
        for label, value in [
            ("Campo", difference.field),
            ("Base local", difference.local_value),
            ("WooCommerce", difference.woo_value),
            ("Accion propuesta", difference.action),
        ]:
            self._detail_row(body, label, value).pack(fill=tk.X, padx=16, pady=4)

        footer = tk.Frame(win, bg=BG)
        footer.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))
        self._button(footer, "Cerrar", command=win.destroy).pack(side=tk.RIGHT)
        self._button(footer, "Aplicar bloqueado" if difference.status == "Critical" else "Guardar preview", primary=True).pack(side=tk.RIGHT, padx=(0, 8))

    def _build_suppliers(self, parent: tk.Frame) -> None:
        self._page_header(parent, "Gestion", "Proveedores", "Informacion operativa, notas y acceso a calculo de pedidos.", ["Exportar", "Nuevo proveedor"])
        grid = tk.Frame(parent, bg=BG)
        grid.pack(fill=tk.X)
        for i, (provider, status) in enumerate([("Ekomat", "OK"), ("Pascal", "Info"), ("Heimei", "Warning"), ("Cipta", "OK")]):
            grid.columnconfigure(i, weight=1)
            self._provider_card(grid, provider, status).grid(row=0, column=i, sticky="ew", padx=(0 if i == 0 else 8, 0))


    def _format_supplier_price_value(self, value: Any) -> str:
        if value in (None, "", "NO ESTA"):
            return ""
        try:
            return f"{float(str(value).replace(',', '.')):.2f}"
        except Exception:
            return str(value)

    def _supplier_price_status(self, row: dict[str, Any]) -> str:
        primary = self._money_float(row.get("primary_supplier_price"), 0.0)
        pascal = self._money_float(row.get("pascal_price"), 0.0)
        if primary > 0 and pascal > 0:
            return "Doble"
        if primary > 0:
            return "Principal"
        if pascal > 0:
            return "Pascal"
        return "Sin precio"

    def _build_supplier_prices(self, parent: tk.Frame) -> None:
        self._page_header(
            parent,
            "Gestion",
            "Precio Proveedores",
            "Control visual de precios proveedor leídos desde Supabase inventory_items.",
            ["Actualizar"],
        )

        if self._cloud_session is None:
            empty = self._card(parent)
            empty.pack(fill=tk.BOTH, expand=True)
            tk.Label(
                empty,
                text="Inicia sesión en Supabase para consultar y editar precios de proveedor.",
                bg=CARD,
                fg=MUTED,
                font=("Segoe UI", 12),
            ).pack(expand=True)
            return

        state: dict[str, Any] = {
            "rows": [],
            "selected": None,
        }

        layout = tk.Frame(parent, bg=BG)
        layout.pack(fill=tk.BOTH, expand=True)
        layout.columnconfigure(0, weight=3)
        layout.columnconfigure(1, weight=2)
        layout.rowconfigure(1, weight=1)

        top = self._card(layout)
        top.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 14))
        top.columnconfigure(1, weight=1)
        tk.Label(top, text="Buscar", bg=CARD, fg=TEXT, font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w", padx=16, pady=14)
        search_var = tk.StringVar()
        search_entry = tk.Entry(top, textvariable=search_var, bg="white", fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightthickness=1)
        search_entry.grid(row=0, column=1, sticky="ew", padx=(0, 10), pady=14, ipady=7)

        table_card = self._card(layout)
        table_card.grid(row=1, column=0, sticky="nsew", padx=(0, 14))
        table_card.rowconfigure(1, weight=1)
        table_card.columnconfigure(0, weight=1)

        head = tk.Frame(table_card, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        head.grid(row=0, column=0, sticky="ew")
        head.columnconfigure(0, weight=1)
        tk.Label(head, text="Items y precios", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w", padx=16, pady=14)
        count_label = tk.Label(head, text="0 items", bg=CARD, fg=MUTED, font=("Segoe UI", 9, "bold"))
        count_label.grid(row=0, column=1, sticky="e", padx=16, pady=14)

        table_host = tk.Frame(table_card, bg=CARD)
        table_host.grid(row=1, column=0, sticky="nsew")
        table_host.rowconfigure(0, weight=1)
        table_host.columnconfigure(0, weight=1)

        columns = ("ID", "Nombre", "Principal", "Pascal", "Familia", "Subgrupo", "Estado")
        tree = ttk.Treeview(table_host, columns=columns, show="headings", selectmode="browse")
        widths = {
            "ID": 90,
            "Nombre": 360,
            "Principal": 110,
            "Pascal": 110,
            "Familia": 140,
            "Subgrupo": 140,
            "Estado": 110,
        }
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=widths.get(col, 120), minwidth=70, anchor=tk.CENTER if col != "Nombre" else tk.W)
        tree.tag_configure("ok", background="#ECFDF5")
        tree.tag_configure("double", background="#EEF2FF")
        tree.tag_configure("missing", background="#FFF1F2")
        yscroll = ttk.Scrollbar(table_host, orient=tk.VERTICAL, command=tree.yview)
        xscroll = ttk.Scrollbar(table_host, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew", padx=(16, 0), pady=(0, 0))
        yscroll.grid(row=0, column=1, sticky="ns", padx=(0, 16))
        xscroll.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 16))

        detail = self._card(layout)
        detail.grid(row=1, column=1, sticky="nsew")
        detail.columnconfigure(0, weight=1)

        def set_detail(row: dict[str, Any] | None) -> None:
            for child in detail.winfo_children():
                child.destroy()
            if not row:
                tk.Label(detail, text="Selecciona un item", bg=CARD, fg=TEXT, font=("Segoe UI", 16, "bold")).pack(anchor=tk.W, padx=18, pady=(18, 8))
                tk.Label(detail, text="Aquí podrás revisar y editar precio proveedor principal y precio Pascal.", bg=CARD, fg=MUTED, wraplength=360, justify=tk.LEFT).pack(anchor=tk.W, padx=18)
                return

            tk.Label(detail, text=str(row.get("name") or "-"), bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold"), wraplength=390, justify=tk.LEFT).pack(anchor=tk.W, padx=18, pady=(18, 6))
            meta = f"ID {row.get('item_id')} · {row.get('family') or '-'} · {row.get('subgroup') or '-'}"
            tk.Label(detail, text=meta, bg=CARD, fg=MUTED, font=("Segoe UI", 9)).pack(anchor=tk.W, padx=18, pady=(0, 14))

            form = tk.Frame(detail, bg=CARD)
            form.pack(fill=tk.X, padx=18)
            form.columnconfigure(1, weight=1)

            tk.Label(form, text="Precio proveedor principal", bg=CARD, fg=TEXT, font=("Segoe UI", 9, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 6))
            primary_var = tk.StringVar(value=self._format_supplier_price_value(row.get("primary_supplier_price")))
            primary_entry = tk.Entry(form, textvariable=primary_var, bg="white", fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightthickness=1)
            primary_entry.grid(row=0, column=1, sticky="ew", padx=(12, 0), pady=(0, 6), ipady=6)

            tk.Label(form, text="Precio Pascal", bg=CARD, fg=TEXT, font=("Segoe UI", 9, "bold")).grid(row=1, column=0, sticky="w", pady=(0, 6))
            pascal_var = tk.StringVar(value=self._format_supplier_price_value(row.get("pascal_price")))
            pascal_entry = tk.Entry(form, textvariable=pascal_var, bg="white", fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightthickness=1)
            pascal_entry.grid(row=1, column=1, sticky="ew", padx=(12, 0), pady=(0, 6), ipady=6)

            tk.Label(form, text="Motivo / nota", bg=CARD, fg=TEXT, font=("Segoe UI", 9, "bold")).grid(row=2, column=0, sticky="nw", pady=(0, 6))
            reason_text = tk.Text(form, height=4, bg="white", fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightthickness=1)
            reason_text.grid(row=2, column=1, sticky="ew", padx=(12, 0), pady=(0, 6))

            info = tk.Frame(detail, bg=SOFT, highlightbackground=LINE, highlightthickness=1)
            info.pack(fill=tk.X, padx=18, pady=(12, 8))
            for idx_info, (label, value) in enumerate([
                ("HECA", row.get("heca_reference") or "-"),
                ("Woo SKU", row.get("woo_sku") or "-"),
                ("Medidas", row.get("size") or "-"),
                ("Materiales", row.get("materials") or "-"),
                ("Actualizado", row.get("updated_at") or "-"),
            ]):
                tk.Label(info, text=f"{label}: {value}", bg=SOFT, fg="#334155", anchor=tk.W, justify=tk.LEFT, wraplength=360).pack(fill=tk.X, padx=10, pady=(6 if idx_info == 0 else 2, 2))

            def save_prices() -> None:
                before_primary = self._format_supplier_price_value(row.get("primary_supplier_price"))
                before_pascal = self._format_supplier_price_value(row.get("pascal_price"))
                new_primary = primary_var.get().strip()
                new_pascal = pascal_var.get().strip()
                if before_primary == new_primary and before_pascal == new_pascal:
                    messagebox.showinfo("Precio Proveedores", "No hay cambios para guardar.")
                    return
                reason = reason_text.get("1.0", tk.END).strip()
                msg = (
                    f"Item {row.get('item_id')}\n\n"
                    f"Principal: {before_primary or 'vacío'} → {new_primary or 'vacío'}\n"
                    f"Pascal: {before_pascal or 'vacío'} → {new_pascal or 'vacío'}\n\n"
                    "Se guardará en Supabase con log y snapshot.\n¿Continuar?"
                )
                if not messagebox.askyesno("Guardar precios proveedor", msg):
                    return
                try:
                    result = update_supplier_price_inventory_item(
                        self._cloud_session,
                        row.get("item_id"),
                        primary_supplier_price=new_primary,
                        pascal_price=new_pascal,
                        reason=reason,
                    )
                except Exception as exc:
                    messagebox.showerror("Precio Proveedores", f"No se pudieron guardar los precios.\n\n{exc}")
                    return
                messagebox.showinfo("Precio Proveedores", f"Precios guardados.\nOperation ID: {result.get('operation_id')}")
                load_rows(select_item_id=row.get("item_id"))

            actions_bar = tk.Frame(detail, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
            actions_bar.pack(fill=tk.X, padx=18, pady=(8, 10))
            self._button(actions_bar, "Guardar cambios", primary=True, command=save_prices).pack(side=tk.RIGHT, padx=10, pady=10)
            self._button(actions_bar, "Cancelar", command=lambda: set_detail(state.get("selected"))).pack(side=tk.RIGHT, pady=10)
            tk.Label(
                actions_bar,
                text="Guarda en Supabase con log + snapshot",
                bg=CARD,
                fg=MUTED,
                font=("Segoe UI", 8),
            ).pack(side=tk.LEFT, padx=10, pady=10)


        def render_rows(rows: list[dict[str, Any]], select_item_id: Any = None) -> None:
            tree.delete(*tree.get_children())
            state["rows"] = rows
            by_iid: dict[str, dict[str, Any]] = {}
            for row in rows:
                status = self._supplier_price_status(row)
                tag = "double" if status == "Doble" else "ok" if status in {"Principal", "Pascal"} else "missing"
                iid = str(row.get("item_id"))
                tree.insert(
                    "",
                    tk.END,
                    iid=iid,
                    values=(
                        row.get("item_id"),
                        row.get("name"),
                        self._format_supplier_price_value(row.get("primary_supplier_price")) or "-",
                        self._format_supplier_price_value(row.get("pascal_price")) or "-",
                        row.get("family") or "-",
                        row.get("subgroup") or "-",
                        status,
                    ),
                    tags=(tag,),
                )
                by_iid[iid] = row
            tree._row_by_iid = by_iid  # type: ignore[attr-defined]
            count_label.configure(text=f"{len(rows)} items")
            target = str(select_item_id or "")
            if target and target in by_iid:
                tree.selection_set(target)
                tree.see(target)
                state["selected"] = by_iid[target]
                set_detail(by_iid[target])
            elif rows:
                first = str(rows[0].get("item_id"))
                tree.selection_set(first)
                state["selected"] = rows[0]
                set_detail(rows[0])
            else:
                state["selected"] = None
                set_detail(None)

        def load_rows(select_item_id: Any = None) -> None:
            try:
                rows = list_supplier_price_inventory_items(self._cloud_session, query=search_var.get(), limit=700)
            except Exception as exc:
                messagebox.showerror("Precio Proveedores", f"No se pudieron cargar precios.\n\n{exc}")
                rows = []
            render_rows(rows, select_item_id=select_item_id)

        def on_select(_event: object | None = None) -> None:
            selection = tree.selection()
            if not selection:
                return
            row = getattr(tree, "_row_by_iid", {}).get(selection[0])
            state["selected"] = row
            set_detail(row)

        tree.bind("<<TreeviewSelect>>", on_select)
        self._button(top, "Buscar", primary=True, command=load_rows).grid(row=0, column=2, sticky="e", padx=(0, 10), pady=14)
        self._button(top, "Actualizar", command=load_rows).grid(row=0, column=3, sticky="e", padx=(0, 16), pady=14)
        search_entry.bind("<Return>", lambda _event: load_rows())
        load_rows()

    def _build_reports(self, parent: tk.Frame) -> None:
        head = tk.Frame(parent, bg=BG)
        head.pack(fill=tk.X, pady=(0, 16))
        head.columnconfigure(0, weight=1)
        tk.Label(head, text="Informes / Exportaciones", bg=BG, fg=TEXT, font=("Segoe UI", 24, "bold")).grid(row=0, column=0, sticky="w")
        self._button(head, "Nueva exportacion", primary=True, command=self._open_new_export_modal).grid(row=0, column=1, sticky="e")

        top = tk.Frame(parent, bg=BG)
        top.pack(fill=tk.X, pady=(0, 16))
        top.columnconfigure(0, weight=1)
        search = tk.Entry(
            top,
            bg=CARD,
            fg=TEXT,
            insertbackground=TEXT,
            relief=tk.FLAT,
            highlightbackground=LINE,
            highlightcolor=INDIGO,
            highlightthickness=1,
            font=("Segoe UI", 10),
        )
        search.insert(0, "Buscar por archivo, modulo, usuario, formato o estado...")
        search.grid(row=0, column=0, sticky="ew", ipady=11)
        self._button(top, "Actualizar registro").grid(row=0, column=1, sticky="e", padx=(12, 0))

        summary = tk.Frame(parent, bg=BG)
        summary.pack(fill=tk.X, pady=(0, 18))
        for index, (label, value, status) in enumerate(
            [
                ("Exportaciones", "128", "Info"),
                ("Este mes", "21", "OK"),
                ("Errores", "2", "Error"),
                ("Ultima", "Hoy", "Info"),
            ]
        ):
            summary.columnconfigure(index, weight=1)
            self._metric(summary, label, value, status).grid(row=0, column=index, sticky="ew", padx=(0 if index == 0 else 8, 0))

        body = tk.Frame(parent, bg=BG)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=4)
        body.columnconfigure(1, weight=2)
        body.rowconfigure(0, weight=1)

        table_card = self._card(body)
        table_card.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        table_card.columnconfigure(0, weight=1)
        table_card.rowconfigure(1, weight=1)
        table_head = tk.Frame(table_card, bg=CARD)
        table_head.grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 10))
        table_head.columnconfigure(0, weight=1)
        tk.Label(table_head, text="Registro de exportaciones", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w")
        self._status_chip(table_head, "Ultimas 100", "Info").grid(row=0, column=1, sticky="e")

        table_frame = tk.Frame(table_card, bg=CARD)
        table_frame.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)
        columns = ["Fecha / Hora", "Tipo", "Modulo", "Formato", "Usuario / Rol", "Estado", "Archivo", ""]
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=14)
        widths = {
            "Fecha / Hora": 145,
            "Tipo": 175,
            "Modulo": 145,
            "Formato": 80,
            "Usuario / Rol": 110,
            "Estado": 95,
            "Archivo": 210,
            "": 70,
        }
        for column in columns:
            tree.heading(column, text=column, anchor=tk.CENTER)
            tree.column(column, width=widths[column], anchor=tk.CENTER, stretch=False)
        tree.tag_configure("OK", background="#F0FDF4")
        tree.tag_configure("Info", background="#EFF6FF")
        tree.tag_configure("Warning", background="#FFFBEB")
        tree.tag_configure("Error", background="#FFF7ED")
        export_by_iid: dict[str, ExportRecord] = {}
        for record in EXPORT_RECORDS:
            visible_status = "Generado" if record.status in {"OK", "Warning"} else record.status
            iid = tree.insert(
                "",
                tk.END,
                values=(record.date, record.report_type, record.module, record.format, record.user_role, visible_status, record.file_name, "Ver"),
                tags=(record.status,),
            )
            export_by_iid[iid] = record
            if record == self._selected_export_record:
                tree.selection_set(iid)
                tree.focus(iid)
        yscroll = tk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        xscroll = tk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        detail_host = tk.Frame(body, bg=BG)
        detail_host.grid(row=0, column=1, sticky="nsew")

        def render_detail(record: ExportRecord) -> None:
            self._selected_export_record = record
            self._render_export_detail(detail_host, record)

        def on_select(_event: object | None = None) -> None:
            selection = tree.selection()
            if not selection:
                return
            render_detail(export_by_iid[selection[0]])

        tree.bind("<<TreeviewSelect>>", on_select)
        render_detail(self._selected_export_record)

    def _render_export_detail(self, parent: tk.Frame, record: ExportRecord) -> None:
        for child in parent.winfo_children():
            child.destroy()
        detail = self._card(parent)
        detail.pack(fill=tk.BOTH, expand=True)
        detail.rowconfigure(0, weight=1)
        detail.columnconfigure(0, weight=1)

        scroll_area = tk.Frame(detail, bg=CARD)
        scroll_area.grid(row=0, column=0, sticky="nsew", padx=18, pady=(16, 10))
        top = tk.Frame(scroll_area, bg=CARD)
        top.pack(fill=tk.X, pady=(0, 12))
        top.columnconfigure(0, weight=1)
        tk.Label(top, text="Detalle", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w")
        self._status_chip(top, "Generado" if record.status == "OK" else record.status, record.status).grid(row=0, column=1, sticky="e")
        tk.Label(scroll_area, text=record.file_name, bg=CARD, fg=TEXT, font=("Segoe UI", 16, "bold"), wraplength=330, justify=tk.LEFT).pack(
            anchor=tk.W,
            pady=(0, 2),
        )
        tk.Label(scroll_area, text=f"{record.report_type} - {record.format}", bg=CARD, fg=MUTED, font=("Segoe UI", 9)).pack(anchor=tk.W, pady=(0, 14))
        for label, value in [
            ("Modulo", record.module),
            ("Tipo", record.report_type),
            ("Formato", record.format),
            ("Usuario / Rol", record.user_role),
            ("Fecha", record.date),
            ("Estado", "Generado" if record.status == "OK" else record.status),
            ("Filas", record.rows),
            ("Referencia", record.reference),
        ]:
            self._detail_row(scroll_area, label, value).pack(fill=tk.X, pady=4)
        for title, text in [("Filtros usados", record.filters), ("Columnas incluidas", record.columns), ("Ruta interna", record.path)]:
            self._status_row(scroll_area, title, text, "Info").pack(fill=tk.X, pady=(10, 0))

        actions = tk.Frame(detail, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        actions.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        self._button(actions, "Descargar", primary=True, command=lambda: self._open_export_action_modal("Descargar", record)).pack(fill=tk.X, padx=12, pady=(12, 7))
        row = tk.Frame(actions, bg=CARD)
        row.pack(fill=tk.X, padx=12, pady=(0, 12))
        row.columnconfigure(0, weight=1)
        row.columnconfigure(1, weight=1)
        self._button(row, "Regenerar", command=lambda: self._open_export_action_modal("Regenerar", record)).grid(row=0, column=0, sticky="ew", padx=(0, 6))
        self._button(row, "Eliminar registro", command=lambda: self._open_export_action_modal("Eliminar registro", record)).grid(row=0, column=1, sticky="ew", padx=(6, 0))

    def _open_new_export_modal(self) -> None:
        win = tk.Toplevel(self)
        win.title("Nueva exportacion")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 860, 520)
        win.columnconfigure(0, weight=1)
        header = tk.Frame(win, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew", padx=18, pady=(18, 0))
        header.columnconfigure(0, weight=1)
        tk.Label(header, text="Nueva exportacion", bg=CARD, fg=TEXT, font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 2))
        tk.Label(header, text="Generar informe o archivo desde un modulo del ERP.", bg=CARD, fg=MUTED).grid(row=1, column=0, sticky="w", padx=18, pady=(0, 16))
        self._button(header, "Cerrar", command=win.destroy).grid(row=0, column=1, rowspan=2, padx=18, pady=16)

        body = tk.Frame(win, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        body.grid(row=1, column=0, sticky="ew", padx=18, pady=12)
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)
        self._combo_field(body, "Modulo", ["Inventario", "Cambio de Precios", "Pedidos", "WooCommerce", "Seguridad / Logs"], "Inventario").grid(row=0, column=0, sticky="ew", padx=(16, 8), pady=(16, 10))
        self._combo_field(
            body,
            "Tipo de informe",
            ["Inventario completo", "Propuesta de precios", "Pedido calculado", "Detalle de pedido", "Auditoria de logs", "Incidencias WooCommerce"],
            "Inventario completo",
        ).grid(row=0, column=1, sticky="ew", padx=(8, 16), pady=(16, 10))
        self._combo_field(body, "Formato", ["XLSX", "PDF", "CSV"], "XLSX").grid(row=1, column=0, sticky="ew", padx=(16, 8), pady=(0, 10))
        name_var = tk.StringVar(value="exportacion_futonhub")
        self._field(body, "Nombre archivo", name_var).grid(row=1, column=1, sticky="ew", padx=(8, 16), pady=(0, 10))

        checks = tk.Frame(body, bg=CARD)
        checks.grid(row=2, column=0, columnspan=2, sticky="ew", padx=16, pady=(4, 12))
        for index, (label, checked) in enumerate(
            [
                ("Incluir filtros aplicados", True),
                ("Registrar en logs", True),
                ("Incluir incidencias", False),
                ("Abrir al generar", False),
            ]
        ):
            checks.columnconfigure(index % 2, weight=1)
            var = tk.BooleanVar(value=checked)
            tk.Checkbutton(checks, text=label, variable=var, bg=CARD, fg=TEXT, activebackground=CARD, anchor=tk.W).grid(
                row=index // 2,
                column=index % 2,
                sticky="ew",
                pady=3,
            )
        tk.Label(
            body,
            text="La exportacion debe quedar registrada con modulo, usuario, filtros, formato, estado y archivo generado.",
            bg=INDIGO_SOFT,
            fg="#4338CA",
            wraplength=760,
            justify=tk.LEFT,
            padx=12,
            pady=10,
        ).grid(row=3, column=0, columnspan=2, sticky="ew", padx=16, pady=(0, 16))

        footer = tk.Frame(win, bg=BG)
        footer.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))
        self._button(footer, "Generar exportacion", primary=True).pack(side=tk.RIGHT)
        self._button(footer, "Cancelar", command=win.destroy).pack(side=tk.RIGHT, padx=(0, 8))

    def _open_export_action_modal(self, action: str, record: ExportRecord) -> None:
        messagebox.showinfo(action, f"{action}: {record.file_name}\n\nAccion visual. La logica real se conectara con validacion y log.")

    def _build_settings(self, parent: tk.Frame) -> None:
        tk.Label(parent, text="Configuracion", bg=BG, fg=TEXT, font=("Segoe UI", 24, "bold")).pack(anchor=tk.W, pady=(0, 14))
        tab_host = tk.Frame(parent, bg=BG)
        tab_host.pack(fill=tk.X, pady=(0, 16))
        content = tk.Frame(parent, bg=BG)
        content.pack(fill=tk.BOTH, expand=True)

        def show_tab(tab: str) -> None:
            self._settings_tab = tab
            for child in tab_host.winfo_children():
                child.destroy()
            for label in ["Generales", "Calculos", "Seguridad"]:
                self._button(tab_host, label, primary=label == tab, command=lambda value=label: show_tab(value)).pack(side=tk.LEFT, padx=(0, 8))
            for child in content.winfo_children():
                child.destroy()
            if tab == "Generales":
                self._render_settings_general(content)
            elif tab == "Calculos":
                self._render_settings_calculations(content)
            else:
                self._render_settings_security(content)

        show_tab(self._settings_tab)

    def _render_settings_general(self, parent: tk.Frame) -> None:
        body = tk.Frame(parent, bg=BG)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=2)
        body.columnconfigure(1, weight=1)
        card = self._card(body)
        card.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        card.columnconfigure(0, weight=1)
        tk.Label(card, text="Generales", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 10))
        form = tk.Frame(card, bg=CARD)
        form.grid(row=1, column=0, sticky="ew", padx=18)
        form.columnconfigure(0, weight=1)
        form.columnconfigure(1, weight=1)
        self._field(form, "Entorno", tk.StringVar(value="FutonHUB Produccion")).grid(row=0, column=0, sticky="ew", padx=(0, 8), pady=8)
        self._combo_field(form, "Modo", ["Online + Local", "Solo local", "Solo lectura"], "Online + Local").grid(row=0, column=1, sticky="ew", padx=(8, 0), pady=8)
        self._field(form, "Rol actual", tk.StringVar(value="Admin")).grid(row=1, column=0, sticky="ew", padx=(0, 8), pady=8)
        self._combo_field(form, "Tema", ["Claro ERP", "Oscuro ERP", "Sistema"], "Claro ERP").grid(row=1, column=1, sticky="ew", padx=(8, 0), pady=8)
        self._field(form, "Ruta base local", tk.StringVar(value="GestorWoo/data/gestorwoo.sqlite3")).grid(row=2, column=0, columnspan=2, sticky="ew", pady=8)
        footer = tk.Frame(card, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        footer.grid(row=2, column=0, sticky="ew", padx=18, pady=(12, 18))
        self._button(footer, "Guardar generales", primary=True).pack(side=tk.RIGHT, padx=12, pady=12)
        self._button(footer, "Cancelar").pack(side=tk.RIGHT, pady=12)

        side = self._card(body)
        side.grid(row=0, column=1, sticky="nsew")
        tk.Label(side, text="Conexiones", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=16, pady=(16, 8))
        for label, value, status in [
            ("SQLite local", "OK", "OK"),
            ("Supabase", "Online", "OK"),
            ("WooCommerce", "Conectado", "OK"),
            ("Backups", "Activo", "Info"),
        ]:
            self._status_row(side, label, value, status).pack(fill=tk.X, padx=16, pady=5)

    def _render_settings_calculations(self, parent: tk.Frame) -> None:
        body = tk.Frame(parent, bg=BG)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=2)
        body.columnconfigure(1, weight=1)

        if self._cloud_session is not None:
            try:
                self._business_constants = list_business_constants(self._cloud_session)
            except Exception:
                if not getattr(self, "_business_constants", None):
                    self._business_constants = {key: dict(value) for key, value in DEFAULT_BUSINESS_CONSTANTS.items()}
        elif not getattr(self, "_business_constants", None):
            self._business_constants = {key: dict(value) for key, value in DEFAULT_BUSINESS_CONSTANTS.items()}

        card = self._card(body)
        card.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        tk.Label(card, text="Constantes del negocio", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=18, pady=(16, 8))

        entries: dict[str, tk.Entry] = {}
        for key, meta in DEFAULT_BUSINESS_CONSTANTS.items():
            current = self._business_constants.get(key, meta)
            row = self._constant_row(
                card,
                key,
                str(current.get("description") or meta.get("description") or key),
                str(current.get("value", meta.get("value", ""))),
                str(current.get("unit") or meta.get("unit") or ""),
            )
            row.pack(fill=tk.X, padx=18, pady=4)
            entry = getattr(row, "_entry", None)
            if isinstance(entry, tk.Entry):
                entries[key] = entry

        def refresh_constants() -> None:
            if self._cloud_session is None:
                messagebox.showwarning("Configuracion", "Inicia sesión Supabase para leer constantes reales.")
                return
            try:
                self._business_constants = list_business_constants(self._cloud_session)
            except Exception as exc:
                messagebox.showerror("Configuracion", f"No se pudieron leer las constantes.\n\n{exc}")
                return
            self._show_view("configuracion")

        def save_constants_from_ui() -> None:
            if self._cloud_session is None:
                messagebox.showwarning("Configuracion", "Inicia sesión Supabase para guardar constantes.")
                return
            values: dict[str, float] = {}
            for key, entry in entries.items():
                raw = str(entry.get() or "").strip()
                try:
                    values[key] = float(raw.replace(",", "."))
                except Exception:
                    messagebox.showwarning("Configuracion", f"Valor inválido para {key}: {raw}")
                    return
            if not messagebox.askyesno(
                "Guardar constantes",
                "Se guardarán las constantes de cálculo en Supabase y se registrará log/snapshot.\n\n¿Continuar?",
            ):
                return
            try:
                result = save_business_constants(self._cloud_session, values)
                self._business_constants = list_business_constants(self._cloud_session)
            except Exception as exc:
                messagebox.showerror("Configuracion", f"No se pudieron guardar las constantes.\n\n{exc}")
                return
            messagebox.showinfo("Configuracion", f"Constantes guardadas.\nOperation ID: {result.get('operation_id')}")
            self._show_view("configuracion")

        footer = tk.Frame(card, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        footer.pack(fill=tk.X, padx=18, pady=(12, 18))
        self._button(footer, "Guardar calculos", primary=True, command=save_constants_from_ui).pack(side=tk.RIGHT, padx=12, pady=12)
        self._button(footer, "Recargar", command=refresh_constants).pack(side=tk.RIGHT, pady=12)
        self._button(footer, "Cancelar", command=lambda: self._show_view("configuracion")).pack(side=tk.RIGHT, padx=(0, 8), pady=12)

        side = self._card(body)
        side.grid(row=0, column=1, sticky="nsew")
        tk.Label(side, text="Impacto", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=16, pady=(16, 8))
        tk.Label(
            side,
            text="Los cambios en constantes afectan cálculo de pedidos, costes finales, rentabilidad y exportaciones. Deben quedar registrados en logs.",
            bg=AMBER_SOFT,
            fg="#92400E",
            wraplength=300,
            justify=tk.LEFT,
            padx=12,
            pady=10,
        ).pack(fill=tk.X, padx=16, pady=8)
        for key in DEFAULT_BUSINESS_CONSTANTS:
            meta = self._business_constants.get(key, DEFAULT_BUSINESS_CONSTANTS[key])
            self._status_row(side, key, f"{meta.get('value')} {meta.get('unit', '')}", "Info").pack(fill=tk.X, padx=16, pady=4)

    def _render_settings_security(self, parent: tk.Frame) -> None:
        body = tk.Frame(parent, bg=BG)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=2)
        body.columnconfigure(1, weight=1)
        card = self._card(body)
        card.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        tk.Label(card, text="Seguridad", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=18, pady=(16, 8))
        for title, subtitle in [
            ("Preview interno obligatorio", "Validacion antes de operaciones sensibles"),
            ("Bloquear precios en 0", "Critical automatico si un precio queda en cero"),
            ("Confirmacion por palabra", "Requiere escribir CONFIRMAR"),
            ("Cancelar operacion ante Critical", "Bloqueo completo del flujo"),
            ("Backups automaticos", "Antes de cambios sensibles"),
            ("Registro de operaciones", "Guardar accion, usuario, fecha y resultado"),
        ]:
            self._setting_switch_row(card, title, subtitle).pack(fill=tk.X, padx=18, pady=5)
        footer = tk.Frame(card, bg=CARD, highlightbackground=SOFT, highlightthickness=1)
        footer.pack(fill=tk.X, padx=18, pady=(12, 18))
        self._button(footer, "Guardar seguridad", primary=True).pack(side=tk.RIGHT, padx=12, pady=12)
        self._button(footer, "Cancelar").pack(side=tk.RIGHT, pady=12)

        side = self._card(body)
        side.grid(row=0, column=1, sticky="nsew")
        tk.Label(side, text="Reglas activas", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=16, pady=(16, 8))
        for label, value, status in [("Warning", "No bloquea", "Warning"), ("Error", "Bloquea paso", "Error"), ("Critical", "Bloquea todo", "Critical"), ("Logs", "Obligatorios", "Info")]:
            self._status_row(side, label, value, status).pack(fill=tk.X, padx=16, pady=5)

    def _build_security(self, parent: tk.Frame) -> None:
        self._render_security_workspace(parent)
        self._refresh_security_data(parent)

    def _render_security_workspace(self, parent: tk.Frame) -> None:
        for child in parent.winfo_children():
            child.destroy()

        if self._cloud_session is None or str(getattr(self._cloud_session, "role", "") or "").lower() != "admin":
            self._page_header(parent, "Sistema", "Seguridad / Logs", "Acceso restringido a administradores.")
            card = self._card(parent)
            card.pack(fill=tk.BOTH, expand=True)
            tk.Label(
                card,
                text="Acceso denegado",
                bg=CARD,
                fg=ROSE,
                font=("Segoe UI", 18, "bold"),
            ).pack(anchor=tk.W, padx=20, pady=(20, 8))
            tk.Label(
                card,
                text="Los workers no pueden ver logs ni snapshots de seguridad.",
                bg=CARD,
                fg=MUTED,
                font=("Segoe UI", 11),
            ).pack(anchor=tk.W, padx=20)
            return

        self._page_header(
            parent,
            "Sistema",
            "Seguridad / Logs",
            "Auditoría de operaciones, cambios y actividad multiusuario.",
        )

        actions = tk.Frame(parent, bg=BG)
        actions.pack(fill=tk.X, pady=(0, 12))
        self._button(actions, "Actualizar", primary=True, command=lambda: self._refresh_security_data(parent)).pack(side=tk.LEFT, padx=(0, 8))
        self._button(actions, "Exportar visible", command=self._export_visible_security_logs).pack(side=tk.LEFT, padx=(0, 8))
        self._button(actions, "Limpiar filtros", command=lambda: self._clear_security_filters(parent)).pack(side=tk.LEFT)

        filters = self._card(parent)
        filters.pack(fill=tk.X, pady=(0, 14))
        for i in range(8):
            filters.columnconfigure(i, weight=1 if i in {1, 3, 5, 7} else 0)

        def entry(row: int, col: int, label: str, attr: str, width: int = 18) -> tk.Entry:
            tk.Label(filters, text=label, bg=CARD, fg=MUTED, font=("Segoe UI", 8, "bold")).grid(row=row, column=col, sticky="w", padx=(14 if col == 0 else 8, 4), pady=(12 if row == 0 else 6, 2))
            var = tk.StringVar(value=str(getattr(self, attr, "") or ""))
            ent = tk.Entry(filters, textvariable=var, bg="white", fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightthickness=1, width=width)
            ent.grid(row=row, column=col + 1, sticky="ew", padx=(0, 8), pady=(12 if row == 0 else 6, 2), ipady=6)
            ent.bind("<Return>", lambda _event, a=attr, v=var: self._apply_security_filter(parent, a, v.get()))
            ent.bind("<FocusOut>", lambda _event, a=attr, v=var: setattr(self, a, v.get().strip()))
            return ent

        entry(0, 0, "Texto", "_security_filter_text", 26)
        entry(0, 2, "Usuario", "_security_filter_user", 20)

        tk.Label(filters, text="Módulo", bg=CARD, fg=MUTED, font=("Segoe UI", 8, "bold")).grid(
            row=0, column=4, sticky="w", padx=(8, 4), pady=(12, 2)
        )
        module_values = [
            "Todos",
            "Inventario",
            "Pedidos",
            "Precio Proveedores",
            "Cambio de Precios",
            "WooCommerce",
            "Configuración",
            "Seguridad",
            "Sistema",
        ]
        module_var = tk.StringVar(value=self._security_filter_module or "Todos")
        module_combo = ttk.Combobox(filters, textvariable=module_var, values=module_values, state="readonly", width=20)
        module_combo.grid(row=0, column=5, sticky="ew", padx=(0, 8), pady=(12, 2), ipady=4)

        def _module_changed(_event: object | None = None) -> None:
            value = module_var.get().strip()
            self._security_filter_module = "" if value == "Todos" else value
            self._refresh_security_data(parent)

        module_combo.bind("<<ComboboxSelected>>", _module_changed)

        entry(0, 6, "Estado", "_security_filter_status", 14)
        entry(1, 0, "Severidad", "_security_filter_severity", 14)
        entry(1, 2, "Desde YYYY-MM-DD", "_security_filter_date_from", 14)
        entry(1, 4, "Hasta YYYY-MM-DD", "_security_filter_date_to", 14)
        self._button(filters, "Aplicar filtros", primary=True, command=lambda: self._refresh_security_data(parent)).grid(row=1, column=6, columnspan=2, sticky="e", padx=14, pady=(6, 10))

        kpis = security_log_kpis(self._security_visible_rows)
        summary = tk.Frame(parent, bg=BG)
        summary.pack(fill=tk.X, pady=(0, 14))
        kpi_items = [
            ("Eventos hoy", str(kpis.get("events_today", 0)), "Info"),
            ("Errores hoy", str(kpis.get("errors_today", 0)), "Error" if kpis.get("errors_today", 0) else "OK"),
            ("Críticos", str(kpis.get("critical", 0)), "Critical" if kpis.get("critical", 0) else "OK"),
            ("Última operación", str(kpis.get("last_operation") or "-")[:38], "Info"),
            ("Último usuario", str(kpis.get("last_user") or "-")[:38], "Info"),
        ]
        for index, (label, value, status) in enumerate(kpi_items):
            summary.columnconfigure(index, weight=1)
            self._metric(summary, label, value, status).grid(row=0, column=index, sticky="ew", padx=(0 if index == 0 else 8, 0))

        status_text = self._security_error or f"Logs visibles: {len(self._security_visible_rows)} · Snapshots cargados: {len(self._security_snapshots)}"
        if status_text:
            is_error = self._security_error and not self._security_error.startswith("Cargando")
            tk.Label(
                parent,
                text=status_text,
                bg=ROSE_SOFT if is_error else INDIGO_SOFT,
                fg=ROSE if is_error else "#4338CA",
                anchor=tk.W,
                justify=tk.LEFT,
                padx=12,
                pady=9,
                wraplength=980,
            ).pack(fill=tk.X, pady=(0, 14))

        card = self._card(parent)
        card.pack(fill=tk.BOTH, expand=True)
        card.rowconfigure(1, weight=1)
        card.columnconfigure(0, weight=1)

        card_head = tk.Frame(card, bg=CARD)
        card_head.grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 10))
        card_head.columnconfigure(0, weight=1)
        tk.Label(card_head, text="Eventos auditados", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w")
        self._status_chip(card_head, "Últimos 200", "Info").grid(row=0, column=1, sticky="e")

        table_frame = tk.Frame(card, bg=CARD)
        table_frame.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        columns = ["Fecha / Hora", "Usuario", "Rol", "Módulo", "Acción", "Estado", "Severidad", "Entidad", "ID Entidad", "Operation ID", "Mensaje"]
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        widths = {
            "Fecha / Hora": 165,
            "Usuario": 210,
            "Rol": 80,
            "Módulo": 145,
            "Acción": 170,
            "Estado": 90,
            "Severidad": 90,
            "Entidad": 130,
            "ID Entidad": 110,
            "Operation ID": 180,
            "Mensaje": 360,
        }
        for column in columns:
            tree.heading(column, text=column, anchor=tk.CENTER)
            tree.column(column, width=widths[column], anchor=tk.CENTER if column != "Mensaje" else tk.W, stretch=False)

        tag_colors = {
            "OK": "#ECFDF5",
            "INFO": "#EFF6FF",
            "WARNING": "#FFFBEB",
            "ERROR": "#FFF1F2",
            "CRITICAL": "#FFE4E6",
            "BLOCKED": "#F1F5F9",
        }
        for tag, color in tag_colors.items():
            tree.tag_configure(tag, background=color)

        row_by_iid: dict[str, dict[str, Any]] = {}
        rows = self._security_visible_rows
        if not rows:
            tree.insert("", tk.END, values=("", "", "", "Sin logs visibles", "", "", "", "", "", "", "Pulsa Actualizar o revisa filtros."), tags=("INFO",))
        for row in rows:
            tag = str(row.get("severity") or row.get("status") or "INFO").upper()
            if tag not in tag_colors:
                tag = str(row.get("status") or "INFO").upper()
            if tag not in tag_colors:
                tag = "INFO"
            iid = tree.insert(
                "",
                tk.END,
                values=(
                    self._format_datetime_short(row.get("created_at")),
                    row.get("user_email") or "-",
                    row.get("role") or "-",
                    row.get("visual_module") or row.get("module") or "-",
                    row.get("visual_action") or row.get("action") or "-",
                    row.get("status") or "-",
                    row.get("severity") or "-",
                    row.get("entity_type") or "-",
                    row.get("entity_id") or "-",
                    row.get("operation_id") or "-",
                    row.get("message") or row.get("error_detail") or "-",
                ),
                tags=(tag,),
            )
            row_by_iid[iid] = row

        yscroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        xscroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        def on_open(_event: object | None = None) -> None:
            selection = tree.selection()
            if not selection:
                return
            row = row_by_iid.get(selection[0])
            if row:
                self._open_security_log_detail(row)

        tree.bind("<Double-1>", on_open)
        tree.bind("<Return>", on_open)

    def _apply_security_filter(self, parent: tk.Frame, attr: str, value: str) -> None:
        setattr(self, attr, str(value or "").strip())
        self._refresh_security_data(parent)

    def _clear_security_filters(self, parent: tk.Frame) -> None:
        self._security_filter_text = ""
        self._security_filter_user = ""
        self._security_filter_module = ""
        self._security_filter_status = ""
        self._security_filter_severity = ""
        self._security_filter_date_from = ""
        self._security_filter_date_to = ""
        self._refresh_security_data(parent)

    def _refresh_security_data(self, parent: tk.Frame) -> None:
        if self._cloud_session is None or str(getattr(self._cloud_session, "role", "") or "").lower() != "admin":
            self._security_error = "Acceso denegado. Seguridad/Logs solo está disponible para admin."
            self._render_security_workspace(parent)
            return
        self._security_error = "Cargando logs reales..."
        self._render_security_workspace(parent)

        def worker() -> None:
            try:
                rows, snapshots, active_count, stale_count = self._fetch_security_data()
                self.after(
                    0,
                    lambda: self._finish_security_refresh(parent, rows, snapshots, active_count, stale_count, ""),
                )
            except Exception as exc:
                self.after(
                    0,
                    lambda exc=exc: self._finish_security_refresh(parent, [], [], 0, 0, f"No se pudieron cargar logs reales: {exc}"),
                )

        threading.Thread(target=worker, daemon=True).start()

    def _fetch_security_data(self) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int, int]:
        if self._cloud_session is None:
            return [], [], 0, 0
        filters = {
            "text": self._security_filter_text,
            "user": self._security_filter_user,
            "module": self._security_filter_module,
            "status": self._security_filter_status,
            "severity": self._security_filter_severity,
            "date_from": self._security_filter_date_from,
            "date_to": self._security_filter_date_to,
        }
        rows = list_security_audit_logs(self._cloud_session, filters=filters, limit=200)
        operation_ids = [str(row.get("operation_id") or "") for row in rows if row.get("operation_id")]
        snapshots: list[dict[str, Any]] = []
        for operation_id in operation_ids[:50]:
            try:
                snapshot = get_snapshot_by_operation(self._cloud_session, operation_id)
                if snapshot:
                    snapshots.append(snapshot)
            except Exception:
                pass
        settings = load_settings()
        try:
            active_count = len(active_locks(settings.db_path))
            stale_count = len(stale_locks(settings.db_path))
        except Exception:
            active_count = 0
            stale_count = 0
        return rows, snapshots, active_count, stale_count

    def _finish_security_refresh(
        self,
        parent: tk.Frame,
        rows: list[dict[str, Any]],
        snapshots: list[dict[str, Any]],
        active_count: int,
        stale_count: int,
        error: str,
    ) -> None:
        self._security_log_rows = rows
        self._security_visible_rows = rows
        self._security_events = [self._security_event_from_row(row) for row in rows]
        self._security_snapshots = snapshots
        self._security_local_locks = active_count
        self._security_stale_locks = stale_count
        self._security_error = error
        if self._current_key == "seguridad" and parent.winfo_exists():
            self._render_security_workspace(parent)

    def _security_event_from_row(self, row: dict[str, Any]) -> SecurityEvent:
        severity = str(row.get("severity") or row.get("status") or "Info")
        level = self._normalize_security_level(severity)
        user = row.get("user_email") or "-"
        role = row.get("role") or "-"
        entity_type = row.get("entity_type") or ""
        entity_id = row.get("entity_id") or ""
        reference = f"{entity_type}:{entity_id}" if entity_type or entity_id else str(row.get("operation_id") or "-")
        message = str(row.get("message") or row.get("error_detail") or "")
        return SecurityEvent(
            date=str(row.get("created_at") or ""),
            level=level,
            module=str(row.get("visual_module") or row.get("module") or "-"),
            action=str(row.get("visual_action") or row.get("action") or "-"),
            user_role=f"{user} / {role}",
            result=str(row.get("status") or level),
            reference=reference,
            message=message,
            payload=json.dumps(row, ensure_ascii=False, indent=2, default=str),
        )

    def _normalize_security_level(self, raw: str) -> str:
        value = str(raw or "").strip().upper()
        if value in {"CRITICAL"}:
            return "Critical"
        if value in {"ERROR", "BLOCKED", "REJECTED"}:
            return "Error"
        if value in {"WARNING", "WARN"}:
            return "Warning"
        if value in {"OK", "SUCCESS"}:
            return "OK"
        return "Info"

    def _security_counts(self) -> dict[str, int]:
        counts = {"OK": 0, "Info": 0, "Warning": 0, "Error": 0, "Critical": 0}
        for event in self._security_events:
            counts[event.level] = counts.get(event.level, 0) + 1
        return counts

    def _format_datetime_short(self, value: Any) -> str:
        text = str(value or "")
        if not text:
            return "-"
        try:
            from datetime import datetime
            dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
            return dt.strftime("%d/%m/%Y %H:%M")
        except Exception:
            return text[:16]

    def _open_security_log_detail(self, row: dict[str, Any]) -> None:
        win = tk.Toplevel(self)
        win.title("Detalle de log")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 1180, 760)
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)

        header = tk.Frame(win, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew", padx=18, pady=(18, 0))
        header.columnconfigure(0, weight=1)
        title = f"{row.get('visual_module') or row.get('module')} · {row.get('visual_action') or row.get('action')}"
        tk.Label(header, text="Detalle del evento", bg=CARD, fg=TEXT, font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 2))
        tk.Label(header, text=title, bg=CARD, fg=MUTED).grid(row=1, column=0, sticky="w", padx=18, pady=(0, 16))
        self._button(header, "Cerrar", command=win.destroy).grid(row=0, column=1, rowspan=2, padx=18, pady=16)

        body = tk.Frame(win, bg=BG)
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=12)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)

        summary = tk.Frame(body, bg=BG)
        summary.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        summary_items = [
            ("Fecha", self._format_datetime_short(row.get("created_at")), "Info"),
            ("Usuario", str(row.get("user_email") or "-")[:36], "Info"),
            ("Estado", str(row.get("status") or "-"), self._normalize_security_level(row.get("status"))),
            ("Severidad", str(row.get("severity") or "-"), self._normalize_security_level(row.get("severity"))),
        ]
        for index, (label, value, status) in enumerate(summary_items):
            summary.columnconfigure(index, weight=1)
            self._metric(summary, label, value, status).grid(row=0, column=index, sticky="ew", padx=(0 if index == 0 else 8, 0))

        pane = tk.PanedWindow(body, orient=tk.HORIZONTAL, bg=BG, sashwidth=6, bd=0)
        pane.grid(row=1, column=0, sticky="nsew")

        left = self._card(pane)
        right = self._card(pane)
        pane.add(left, minsize=520)
        pane.add(right, minsize=520)

        left.columnconfigure(0, weight=1)
        left.rowconfigure(1, weight=1)
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        tk.Label(left, text="Resumen técnico", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w", padx=16, pady=(16, 8))
        summary_box = tk.Frame(left, bg=CARD)
        summary_box.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 16))

        for label, value in [
            ("Operation ID", row.get("operation_id")),
            ("Rol", row.get("role")),
            ("Máquina", row.get("machine_name")),
            ("Módulo", row.get("visual_module") or row.get("module")),
            ("Acción", row.get("visual_action") or row.get("action")),
            ("Entidad", row.get("entity_type")),
            ("ID entidad", row.get("entity_id")),
            ("Mensaje", row.get("message")),
            ("Error detail", row.get("error_detail")),
        ]:
            self._detail_row(summary_box, label, value or "-").pack(fill=tk.X, pady=3)

        tk.Label(right, text="Before / After", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).grid(row=0, column=0, sticky="w", padx=16, pady=(16, 8))
        diff_host = tk.Frame(right, bg=CARD)
        diff_host.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 16))
        self._render_before_after_diff(diff_host, row.get("before_data"), row.get("after_data"))

        snapshot = None
        if self._cloud_session is not None and row.get("operation_id"):
            try:
                snapshot = get_snapshot_by_operation(self._cloud_session, str(row.get("operation_id")))
            except Exception:
                snapshot = None

        snapshot_card = self._card(win)
        snapshot_card.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))
        snapshot_card.columnconfigure(0, weight=1)
        tk.Label(snapshot_card, text="Snapshot asociado", bg=CARD, fg=TEXT, font=("Segoe UI", 12, "bold")).grid(row=0, column=0, sticky="w", padx=14, pady=(12, 4))
        if snapshot:
            snap_text = f"{snapshot.get('module')} · {snapshot.get('action')} · {snapshot.get('entity_type')}:{snapshot.get('entity_id')} · {snapshot.get('reason') or '-'}"
            tk.Label(snapshot_card, text=snap_text, bg=CARD, fg=MUTED, anchor=tk.W, wraplength=880, justify=tk.LEFT).grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 12))
            self._button(snapshot_card, "Ver before snapshot", command=lambda snap=snapshot: self._open_snapshot_detail_modal(snap)).grid(row=0, column=1, rowspan=2, padx=14, pady=12)
            self._button(snapshot_card, "Restaurar estado anterior", command=lambda snap=snapshot, parent=win: self._restore_snapshot_from_security_detail(snap, parent)).grid(row=0, column=2, rowspan=2, padx=(0, 14), pady=12)
        else:
            tk.Label(snapshot_card, text="No hay snapshot asociado visible para este operation_id.", bg=CARD, fg=MUTED).grid(row=1, column=0, sticky="w", padx=14, pady=(0, 12))

    def _render_before_after_diff(self, parent: tk.Frame, before_data: Any, after_data: Any) -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)
        columns = ["Campo", "Antes", "Después"]
        tree = ttk.Treeview(parent, columns=columns, show="headings", height=13)
        widths = {"Campo": 180, "Antes": 250, "Después": 250}
        for column in columns:
            tree.heading(column, text=column, anchor=tk.CENTER)
            tree.column(column, width=widths[column], anchor=tk.W if column != "Campo" else tk.CENTER)
        diff = build_before_after_diff(before_data, after_data)
        if not diff:
            tree.insert("", tk.END, values=("Sin diferencias simples", "-", "-"))
        for row in diff:
            before = row.get("before") or ""
            after = row.get("after") or ""
            if row.get("is_complex"):
                before = "Ver JSON / valor largo"
                after = "Ver JSON / valor largo"
            tree.insert("", tk.END, values=(row.get("field"), before, after))
        yscroll = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=tree.yview)
        xscroll = ttk.Scrollbar(parent, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        def show_json(_event: object | None = None) -> None:
            selection = tree.selection()
            if not selection:
                return
            values = tree.item(selection[0], "values")
            if not values:
                return
            field = values[0]
            selected = next((item for item in diff if item.get("field") == field), None)
            if selected:
                self._open_json_value_modal(f"Before / After · {field}", {"before": selected.get("before"), "after": selected.get("after")})

        tree.bind("<Double-1>", show_json)

    def _open_json_value_modal(self, title: str, value: Any) -> None:
        win = tk.Toplevel(self)
        win.title(title)
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 820, 560)
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)
        tk.Label(win, text=title, bg=BG, fg=TEXT, font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(18, 8))
        text = tk.Text(win, bg="#0F172A", fg="#E2E8F0", insertbackground="#E2E8F0", relief=tk.FLAT, wrap=tk.NONE, font=("Consolas", 9))
        try:
            text.insert("1.0", json.dumps(value, ensure_ascii=False, indent=2, default=str))
        except Exception:
            text.insert("1.0", str(value))
        text.configure(state=tk.DISABLED)
        yscroll = ttk.Scrollbar(win, orient=tk.VERTICAL, command=text.yview)
        xscroll = ttk.Scrollbar(win, orient=tk.HORIZONTAL, command=text.xview)
        text.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        text.grid(row=1, column=0, sticky="nsew", padx=(18, 0), pady=(0, 18))
        yscroll.grid(row=1, column=1, sticky="ns", padx=(0, 18), pady=(0, 18))
        xscroll.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))

    def _open_snapshot_detail_modal(self, snapshot: dict[str, Any]) -> None:
        win = tk.Toplevel(self)
        win.title("Snapshot asociado")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 900, 620)
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)
        tk.Label(win, text="Snapshot asociado", bg=BG, fg=TEXT, font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(18, 4))
        tk.Label(win, text=str(snapshot.get("operation_id") or ""), bg=BG, fg=MUTED).grid(row=0, column=0, sticky="w", padx=18, pady=(48, 12))
        text = tk.Text(win, bg="#0F172A", fg="#E2E8F0", insertbackground="#E2E8F0", relief=tk.FLAT, wrap=tk.NONE, font=("Consolas", 9))
        text.insert("1.0", json.dumps(snapshot, ensure_ascii=False, indent=2, default=str))
        text.configure(state=tk.DISABLED)
        yscroll = ttk.Scrollbar(win, orient=tk.VERTICAL, command=text.yview)
        xscroll = ttk.Scrollbar(win, orient=tk.HORIZONTAL, command=text.xview)
        text.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        text.grid(row=1, column=0, sticky="nsew", padx=(18, 0), pady=(0, 18))
        yscroll.grid(row=1, column=1, sticky="ns", padx=(0, 18), pady=(0, 18))
        xscroll.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))

    def _export_visible_security_logs(self) -> None:
        if not self._security_visible_rows:
            messagebox.showinfo("Seguridad / Logs", "No hay logs visibles para exportar.")
            return
        path = filedialog.asksaveasfilename(
            title="Exportar logs visibles",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="futonhub_seguridad_logs.xlsx",
        )
        if not path:
            return
        try:
            export_security_logs_excel(self._security_visible_rows, self._security_snapshots, path)
        except Exception as exc:
            messagebox.showerror("Exportar logs", f"No se pudo exportar.\n\n{exc}")
            return
        messagebox.showinfo("Exportar logs", f"Exportación creada:\n{path}")

    def _restore_snapshot_from_security_detail(self, snapshot: dict[str, Any], parent_window: tk.Toplevel | None = None) -> None:
        if self._cloud_session is None or str(getattr(self._cloud_session, "role", "") or "").lower() != "admin":
            messagebox.showerror("Restaurar snapshot", "Solo admin puede restaurar snapshots.")
            return
        try:
            preview = preview_restore_snapshot(self._cloud_session, snapshot)
        except Exception as exc:
            messagebox.showerror("Restaurar snapshot", f"No se pudo preparar el preview.\n\n{exc}")
            return

        if not preview.get("supported"):
            messagebox.showwarning("Restaurar snapshot", preview.get("reason") or "Este snapshot no es restaurable todavía.")
            return

        changes = preview.get("changes") or []
        lines = [
            "PREVIEW DE RESTAURACIÓN",
            "",
            f"Snapshot operation_id: {snapshot.get('operation_id')}",
            f"Cambios a restaurar: {len(changes)}",
            "",
        ]
        for change in changes[:12]:
            lines.append(f"• {change.get('description') or change.get('table')}")
        if len(changes) > 12:
            lines.append(f"... y {len(changes) - 12} cambios más")
        lines.append("")
        if preview.get("special_restore") == "woocommerce_publish":
            lines.extend([
                "Esto restaurará el precio anterior directamente en WooCommerce,",
                "verificará el resultado mediante una lectura posterior y actualizará Supabase.",
            ])
        else:
            lines.extend([
                "Esto restaurará datos internos en Supabase al estado anterior del snapshot.",
                "No toca WooCommerce ni Hexa.",
            ])
        if not messagebox.askyesno("Preview restauración", "\n".join(lines)):
            return

        confirm_win = tk.Toplevel(parent_window or self)
        confirm_win.title("Confirmar restauración")
        confirm_win.configure(bg=BG)
        confirm_win.transient(parent_window or self)
        confirm_win.grab_set()
        center_window(confirm_win, 560, 300)
        confirm_win.columnconfigure(0, weight=1)

        tk.Label(
            confirm_win,
            text="Confirmación crítica",
            bg=BG,
            fg=ROSE,
            font=("Segoe UI", 16, "bold"),
        ).grid(row=0, column=0, sticky="w", padx=18, pady=(18, 6))
        tk.Label(
            confirm_win,
            text=(
                "Vas a restaurar datos desde un snapshot.\n"
                "Esta operación también quedará registrada en Seguridad / Logs.\n\n"
                "Escribe RESTAURAR para continuar."
            ),
            bg=BG,
            fg=TEXT,
            justify=tk.LEFT,
            wraplength=500,
        ).grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 12))

        confirm_var = tk.StringVar()
        entry = tk.Entry(confirm_win, textvariable=confirm_var, bg="white", fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightthickness=1)
        entry.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 14), ipady=8)
        entry.focus_set()

        footer = tk.Frame(confirm_win, bg=BG)
        footer.grid(row=3, column=0, sticky="ew", padx=18, pady=(0, 18))

        def execute_restore() -> None:
            if confirm_var.get().strip() != "RESTAURAR":
                messagebox.showerror("Confirmación", "Debes escribir RESTAURAR exactamente.")
                return
            try:
                result = restore_snapshot_to_previous_state(self._cloud_session, snapshot)
            except Exception as exc:
                messagebox.showerror("Restaurar snapshot", f"No se pudo restaurar.\n\n{exc}")
                return
            messagebox.showinfo(
                "Restauración aplicada",
                f"Estado anterior restaurado.\nOperation ID: {result.get('operation_id')}\nCambios: {len(result.get('restored') or [])}",
            )
            confirm_win.destroy()
            if parent_window is not None and parent_window.winfo_exists():
                parent_window.destroy()
            self._inventory_loaded_once = False
            self._orders_loaded_once = False
            self._refresh_security_data(self._content)

        self._button(footer, "Cancelar", command=confirm_win.destroy).pack(side=tk.RIGHT)
        self._button(footer, "Restaurar", primary=True, command=execute_restore).pack(side=tk.RIGHT, padx=(0, 8))

    def _open_security_event_details(self, event: SecurityEvent) -> None:
        try:
            payload = json.loads(event.payload) if event.payload else {}
        except Exception:
            payload = {}
        if isinstance(payload, dict):
            self._open_security_log_detail(payload)
        else:
            self._open_json_value_modal("Detalle del evento", event.payload)

    def _open_snapshots_modal(self) -> None:
        win = tk.Toplevel(self)
        win.title("Snapshots")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        center_window(win, 980, 600)
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)
        header = tk.Frame(win, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew", padx=18, pady=(18, 0))
        header.columnconfigure(0, weight=1)
        tk.Label(header, text="Operation snapshots", bg=CARD, fg=TEXT, font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 2))
        tk.Label(header, text="Snapshots relacionados con los logs visibles.", bg=CARD, fg=MUTED).grid(row=1, column=0, sticky="w", padx=18, pady=(0, 16))
        self._button(header, "Cerrar", command=win.destroy).grid(row=0, column=1, rowspan=2, padx=18, pady=16)
        card = self._card(win)
        card.grid(row=1, column=0, sticky="nsew", padx=18, pady=12)
        card.columnconfigure(0, weight=1)
        card.rowconfigure(0, weight=1)
        columns = ["Fecha", "Operación", "Módulo", "Acción", "Entidad", "Razón"]
        tree = ttk.Treeview(card, columns=columns, show="headings", height=12)
        widths = {"Fecha": 170, "Operación": 190, "Módulo": 135, "Acción": 165, "Entidad": 180, "Razón": 280}
        for column in columns:
            tree.heading(column, text=column, anchor=tk.CENTER)
            tree.column(column, width=widths[column], anchor=tk.CENTER if column != "Razón" else tk.W, stretch=False)
        snap_by_iid: dict[str, dict[str, Any]] = {}
        for row in self._security_snapshots:
            iid = tree.insert(
                "",
                tk.END,
                values=(
                    self._format_datetime_short(row.get("created_at")),
                    row.get("operation_id", ""),
                    row.get("module", ""),
                    row.get("action", ""),
                    f"{row.get('entity_type','')}:{row.get('entity_id','')}",
                    row.get("reason", ""),
                ),
            )
            snap_by_iid[iid] = row
        if not self._security_snapshots:
            tree.insert("", tk.END, values=("", "", "Sin snapshots visibles", "", "", ""))
        yscroll = ttk.Scrollbar(card, orient=tk.VERTICAL, command=tree.yview)
        xscroll = ttk.Scrollbar(card, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew", padx=(16, 0), pady=16)
        yscroll.grid(row=0, column=1, sticky="ns", padx=(0, 16), pady=16)
        xscroll.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 16))

        def open_selected(_event: object | None = None) -> None:
            selection = tree.selection()
            if not selection:
                return
            snap = snap_by_iid.get(selection[0])
            if snap:
                self._open_snapshot_detail_modal(snap)

        tree.bind("<Double-1>", open_selected)

    def _table_page(
        self,
        parent: tk.Frame,
        title: str,
        columns: list[str],
        rows: list[tuple[str, ...]],
        *,
        detail_title: str,
        detail_lines: list[str],
    ) -> None:
        body = tk.Frame(parent, bg=BG)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=3)
        body.columnconfigure(1, weight=1)
        table_card = self._card(body)
        table_card.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        tk.Label(table_card, text=title, bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=16, pady=(16, 8))
        tree = ttk.Treeview(table_card, columns=columns, show="headings", height=10)
        for column in columns:
            tree.heading(column, text=column, anchor=tk.CENTER)
            tree.column(column, width=130, anchor=tk.CENTER)
        for row in rows:
            tree.insert("", tk.END, values=row)
        tree.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 16))

        detail = self._card(body)
        detail.grid(row=0, column=1, sticky="nsew")
        tk.Label(detail, text=detail_title, bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=16, pady=(16, 8))
        for line in detail_lines:
            tk.Label(detail, text=line, bg=CARD, fg="#334155", anchor=tk.W, justify=tk.LEFT).pack(fill=tk.X, padx=16, pady=5)
        self._button(detail, "Abrir flujo", primary=True).pack(fill=tk.X, padx=16, pady=(16, 6))
        self._button(detail, "Generar preview").pack(fill=tk.X, padx=16, pady=(0, 16))

    def _system_status(self, parent: tk.Frame) -> tk.Frame:
        frame = self._card(parent)
        tk.Label(frame, text="Estado del sistema", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=16, pady=(16, 8))
        for label, detail, status in [
            ("Supabase", "Conectado", "OK"),
            ("WooCommerce", "API disponible", "OK"),
            ("Modo protegido", "Preview obligatorio", "OK"),
            ("Locks", "Sin bloqueos activos", "Info"),
        ]:
            self._status_row(frame, label, detail, status).pack(fill=tk.X, padx=16, pady=5)
        return frame

    def _update_dashboard_detail(self, parent: tk.Frame, title: str, items: list[tuple[str, str, str]]) -> None:
        for child in parent.winfo_children():
            child.destroy()
        frame = self._card(parent)
        frame.pack(fill=tk.BOTH, expand=True)
        tk.Label(frame, text=title, bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=16, pady=(16, 8))
        tk.Label(
            frame,
            text="Vista detallada del estado seleccionado.",
            bg=CARD,
            fg=MUTED,
            font=("Segoe UI", 9),
        ).pack(anchor=tk.W, padx=16, pady=(0, 10))
        for label, detail, status in items:
            self._status_row(frame, label, detail, status).pack(fill=tk.X, padx=16, pady=5)

    def _quick_actions(self, parent: tk.Frame) -> tk.Frame:
        frame = self._card(parent)
        tk.Label(frame, text="Acciones rapidas", bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(anchor=tk.W, padx=16, pady=(16, 8))
        for label, key in [
            ("Pedidos", "calcular"),
            ("Crear propuesta de precio", "precios"),
            ("WooCommerce", "woocommerce"),
        ]:
            self._button(frame, label, primary=label == "Crear propuesta de precio", command=lambda key=key: self._show_view(key)).pack(fill=tk.X, padx=16, pady=5)
        return frame

    def _show_working_overlay(self, title: str, message: str = "Trabajando...") -> tk.Toplevel:
        """Blocking work overlay anchored to the ERP window, not to a random monitor."""
        self.update_idletasks()
        overlay = tk.Toplevel(self)
        overlay.title(title)
        overlay.configure(bg=BG)
        overlay.transient(self)
        overlay.grab_set()
        overlay.resizable(False, False)
        try:
            overlay.attributes("-topmost", True)
        except Exception:
            pass

        width, height = 500, 210
        try:
            root_x = self.winfo_rootx()
            root_y = self.winfo_rooty()
            root_w = max(self.winfo_width(), 800)
            root_h = max(self.winfo_height(), 520)
            x = root_x + max((root_w - width) // 2, 0)
            y = root_y + max((root_h - height) // 2, 0)
            overlay.geometry(f"{width}x{height}+{x}+{y}")
        except Exception:
            center_window(overlay, width, height)

        blocker = tk.Frame(overlay, bg=BG)
        blocker.pack(fill=tk.BOTH, expand=True)
        card = tk.Frame(blocker, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=16, pady=16)
        tk.Label(card, text=title, bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold")).pack(anchor=tk.W, padx=18, pady=(18, 6))
        tk.Label(card, text=message, bg=CARD, fg=MUTED, font=("Segoe UI", 10), justify=tk.LEFT, wraplength=430).pack(anchor=tk.W, padx=18, pady=(0, 12))
        tk.Label(card, text="Ventana bloqueada mientras termina la operación.", bg=CARD, fg=INDIGO, font=("Segoe UI", 9, "bold")).pack(anchor=tk.W, padx=18, pady=(0, 8))
        progress = ttk.Progressbar(card, mode="indeterminate", length=420)
        progress.pack(fill=tk.X, padx=18, pady=(0, 18))
        progress.start(12)

        overlay.protocol("WM_DELETE_WINDOW", lambda: None)
        try:
            overlay.lift(self)
            overlay.focus_force()
            overlay.update_idletasks()
            overlay.update()
        except Exception:
            pass
        return overlay

    def _close_working_overlay(self, overlay: tk.Toplevel | None) -> None:
        if overlay is None:
            return
        try:
            overlay.grab_release()
        except Exception:
            pass
        try:
            overlay.attributes("-topmost", False)
        except Exception:
            pass
        try:
            overlay.destroy()
        except Exception:
            pass

    def _metric(self, parent: tk.Frame, label: str, value: str, status: str, *, command: object | None = None) -> tk.Frame:
        fg, bg = STATUS_STYLES.get(status, STATUS_STYLES.get(str(status).strip().title(), (BLUE, BLUE_SOFT)))
        frame = tk.Frame(parent, bg=bg, highlightbackground=LINE, highlightthickness=1)
        tk.Label(frame, text=label.upper(), bg=bg, fg=MUTED, font=("Segoe UI", 8, "bold")).pack(anchor=tk.W, padx=14, pady=(12, 4))
        tk.Label(frame, text=value, bg=bg, fg=fg, font=("Segoe UI", 20, "bold")).pack(anchor=tk.W, padx=14, pady=(0, 12))
        if command is not None:
            for widget in (frame, *frame.winfo_children()):
                widget.bind("<Button-1>", lambda _event: command(), add="+")
                widget.configure(cursor="hand2")
        return frame

    def _provider_card(self, parent: tk.Frame, provider: str, status: str) -> tk.Frame:
        return self._simple_card(parent, provider, f"Proveedor activo - acceso a Calcular Pedido {provider}.", status)

    def _simple_card(self, parent: tk.Misc, title: str, subtitle: str, status: str) -> tk.Frame:
        frame = self._card(parent)
        top = tk.Frame(frame, bg=CARD)
        top.pack(fill=tk.X, padx=16, pady=(16, 8))
        tk.Label(top, text=title, bg=CARD, fg=TEXT, font=("Segoe UI", 14, "bold")).pack(side=tk.LEFT)
        self._status_chip(top, status, status).pack(side=tk.RIGHT)
        tk.Label(frame, text=subtitle, bg=CARD, fg=MUTED, wraplength=260, justify=tk.LEFT).pack(fill=tk.X, padx=16, pady=(0, 16))
        return frame

    def _status_row(self, parent: tk.Misc, title: str, subtitle: str, status: str) -> tk.Frame:
        frame = tk.Frame(parent, bg=SOFT, highlightbackground=LINE, highlightthickness=1)
        text = tk.Frame(frame, bg=SOFT)
        text.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=12, pady=9)
        tk.Label(text, text=title, bg=SOFT, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(anchor=tk.W)
        tk.Label(text, text=subtitle, bg=SOFT, fg=MUTED, font=("Segoe UI", 9)).pack(anchor=tk.W)
        self._status_chip(frame, status, status).pack(side=tk.RIGHT, padx=10)
        return frame

    def _status_chip(self, parent: tk.Misc, text: str, status: str) -> tk.Label:
        fg, bg = STATUS_STYLES.get(status, (BLUE, BLUE_SOFT))
        return tk.Label(parent, text=text, bg=bg, fg=fg, font=("Segoe UI", 8, "bold"), padx=9, pady=4)

    def _button(self, parent: tk.Misc, text: str, *, primary: bool = False, command: object | None = None) -> tk.Button:
        return tk.Button(
            parent,
            text=text,
            command=command,
            bg=INDIGO if primary else CARD,
            fg="white" if primary else "#334155",
            activebackground="#4338CA" if primary else SOFT,
            activeforeground="white" if primary else TEXT,
            bd=0,
            relief=tk.FLAT,
            padx=13,
            pady=9,
            font=("Segoe UI", 9, "bold"),
            highlightbackground=LINE,
            highlightthickness=1,
        )

    def _field(self, parent: tk.Misc, label: str, variable: tk.StringVar, *, show: str | None = None) -> tk.Frame:
        frame = tk.Frame(parent, bg=CARD)
        tk.Label(
            frame,
            text=label.upper(),
            bg=CARD,
            fg=MUTED,
            font=("Segoe UI", 8, "bold"),
            anchor=tk.W,
        ).pack(fill=tk.X, pady=(0, 5))
        entry = tk.Entry(
            frame,
            textvariable=variable,
            show=show or "",
            bg="white",
            fg=TEXT,
            insertbackground=TEXT,
            relief=tk.FLAT,
            highlightbackground=LINE,
            highlightcolor=INDIGO,
            highlightthickness=1,
            font=("Segoe UI", 10),
        )
        entry.pack(fill=tk.X, ipady=8)
        return frame

    def _combo_field(self, parent: tk.Misc, label: str, values: list[str], selected: str) -> tk.Frame:
        frame = tk.Frame(parent, bg=CARD)
        tk.Label(
            frame,
            text=label.upper(),
            bg=CARD,
            fg=MUTED,
            font=("Segoe UI", 8, "bold"),
            anchor=tk.W,
        ).pack(fill=tk.X, pady=(0, 5))
        combo = ttk.Combobox(frame, values=values, state="readonly", font=("Segoe UI", 10))
        combo.set(selected)
        combo.pack(fill=tk.X, ipady=5)
        return frame

    def _constant_row(self, parent: tk.Misc, name: str, description: str, value: str, unit: str) -> tk.Frame:
        frame = tk.Frame(parent, bg=SOFT, highlightbackground=LINE, highlightthickness=1)
        frame.columnconfigure(0, weight=1)
        text = tk.Frame(frame, bg=SOFT)
        text.grid(row=0, column=0, sticky="ew", padx=12, pady=9)
        tk.Label(text, text=name, bg=SOFT, fg=TEXT, font=("Segoe UI", 9, "bold")).pack(anchor=tk.W)
        tk.Label(text, text=description, bg=SOFT, fg=MUTED, font=("Segoe UI", 8)).pack(anchor=tk.W)
        entry = tk.Entry(frame, bg="white", fg=TEXT, relief=tk.FLAT, highlightbackground=LINE, highlightthickness=1, width=10)
        entry.insert(0, value)
        entry.grid(row=0, column=1, padx=(0, 8), pady=9, ipady=5)
        frame._entry = entry  # type: ignore[attr-defined]
        tk.Label(frame, text=unit, bg=SOFT, fg=MUTED, font=("Segoe UI", 9, "bold"), width=8, anchor=tk.W).grid(row=0, column=2, padx=(0, 12), pady=9)
        return frame

    def _setting_switch_row(self, parent: tk.Misc, title: str, subtitle: str) -> tk.Frame:
        frame = tk.Frame(parent, bg=SOFT, highlightbackground=LINE, highlightthickness=1)
        frame.columnconfigure(0, weight=1)
        text = tk.Frame(frame, bg=SOFT)
        text.grid(row=0, column=0, sticky="ew", padx=12, pady=9)
        tk.Label(text, text=title, bg=SOFT, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(anchor=tk.W)
        tk.Label(text, text=subtitle, bg=SOFT, fg=MUTED, font=("Segoe UI", 9)).pack(anchor=tk.W)
        switch = tk.Frame(frame, bg=GREEN_SOFT, highlightbackground="#BBF7D0", highlightthickness=1, width=54, height=26)
        switch.grid(row=0, column=1, padx=12, pady=9)
        switch.grid_propagate(False)
        tk.Frame(switch, bg=GREEN, width=20, height=20).place(x=29, y=2)
        return frame

    def _card(self, parent: tk.Misc) -> tk.Frame:
        return tk.Frame(parent, bg=CARD, highlightbackground=LINE, highlightthickness=1)


def run_erp_prototype() -> None:
    app = FutonHubErpPrototype()
    app.mainloop()
