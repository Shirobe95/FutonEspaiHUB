from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any
import json


AUDIT_LOG_COLUMNS = (
    "id,operation_id,user_id,user_email,role,device_id,machine_name,module,action,"
    "status,severity,entity_type,entity_id,before_data,after_data,message,error_detail,created_at"
)

SNAPSHOT_COLUMNS = (
    "id,operation_id,user_id,module,action,entity_type,entity_id,before_data,reason,created_at"
)


def _safe_text(value: Any) -> str:
    return "" if value is None else str(value)


def _clean_filter(value: Any) -> str:
    return str(value or "").strip()


def _json_safe(value: Any) -> Any:
    try:
        return json.loads(json.dumps(value, ensure_ascii=False, default=str))
    except Exception:
        return {"_raw": str(value)}


def _money_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "."))
    except Exception:
        return None


def _effective_woo_price(data: dict[str, Any] | None) -> float | None:
    data = data or {}
    sale = _money_or_none(data.get("sale_price"))
    if sale is not None and sale > 0:
        return sale
    regular = _money_or_none(data.get("regular_price"))
    if regular is not None and regular > 0:
        return regular
    return _money_or_none(data.get("price"))


def _format_price_value(value: Any) -> str | None:
    amount = _money_or_none(value)
    if amount is None:
        return None
    return f"{amount:.2f}"


def _contains_text(row: dict[str, Any], needle: str) -> bool:
    if not needle:
        return True
    haystack = " ".join(
        [
            _safe_text(row.get("operation_id")),
            _safe_text(row.get("user_email")),
            _safe_text(row.get("role")),
            _safe_text(row.get("module")),
            _safe_text(row.get("action")),
            _safe_text(row.get("status")),
            _safe_text(row.get("severity")),
            _safe_text(row.get("entity_type")),
            _safe_text(row.get("entity_id")),
            _safe_text(row.get("message")),
            _safe_text(row.get("error_detail")),
        ]
    ).lower()
    return needle.lower() in haystack


def _date_iso_start(value: str) -> str:
    value = _clean_filter(value)
    if not value:
        return ""
    if "T" in value:
        return value
    return f"{value}T00:00:00+00:00"


def _date_iso_end(value: str) -> str:
    value = _clean_filter(value)
    if not value:
        return ""
    if "T" in value:
        return value
    return f"{value}T23:59:59+00:00"


def normalize_visual_module(module: Any) -> str:
    text = _safe_text(module).strip()
    low = text.lower()
    mapping = {
        "inventory_items": "Inventario",
        "inventario": "Inventario",
        "supplier_orders": "Pedidos",
        "pedidos": "Pedidos",
        "precio proveedores": "Precio Proveedores",
        "business_constants": "Configuración",
        "configuracion": "Configuración",
        "configuración": "Configuración",
        "price_proposals": "Cambio de Precios",
        "woocommerce": "WooCommerce",
    }
    return mapping.get(low, text or "-")


def normalize_visual_action(module: Any, action: Any) -> str:
    action_text = _safe_text(action).strip()
    key = f"{_safe_text(module).lower()}::{action_text.lower()}"
    mapping = {
        "inventory_items::create_inventory_item": "Crear artículo",
        "inventory_items::inventory_item_field_update": "Editar artículo",
        "supplier_orders::create_draft": "Crear borrador",
        "supplier_orders::update_draft": "Actualizar borrador",
        "supplier_orders::calculate_order": "Calcular pedido",
        "supplier_orders::receive_order": "Recibir pedido",
        "pedidos::cancel_supplier_order": "Cancelar pedido",
        "supplier_orders::cancel_supplier_order": "Cancelar pedido",
        "precio proveedores::update_supplier_prices": "Actualizar precio",
        "business_constants::update": "Cambiar constante",
        "price_proposals::accept": "Aceptar propuesta",
    }
    if key in mapping:
        return mapping[key]
    fallback = action_text.replace("_", " ").strip().capitalize()
    return fallback or "-"


def list_audit_logs(session, filters: dict[str, Any] | None = None, limit: int = 200) -> list[dict[str, Any]]:
    filters = filters or {}
    limit = max(1, min(int(limit or 200), 1000))

    query = session.client.table("audit_logs").select(AUDIT_LOG_COLUMNS)

    user = _clean_filter(filters.get("user"))
    module = _clean_filter(filters.get("module"))
    status = _clean_filter(filters.get("status"))
    severity = _clean_filter(filters.get("severity"))
    operation_id = _clean_filter(filters.get("operation_id"))
    entity_id = _clean_filter(filters.get("entity_id"))
    date_from = _date_iso_start(filters.get("date_from"))
    date_to = _date_iso_end(filters.get("date_to"))

    if user:
        query = query.ilike("user_email", f"%{user}%")
    if module:
        module_lookup = {
            "inventario": ["inventory_items", "inventario", "Inventario"],
            "pedidos": ["supplier_orders", "pedidos", "Pedidos"],
            "precio proveedores": ["Precio Proveedores", "precio proveedores", "supplier_prices"],
            "cambio de precios": ["price_proposals", "Cambio de Precios", "precios"],
            "woocommerce": ["woocommerce", "WooCommerce"],
            "configuración": ["business_constants", "configuración", "configuracion", "Configuración"],
            "configuracion": ["business_constants", "configuración", "configuracion", "Configuración"],
            "seguridad": ["security", "Seguridad"],
            "sistema": ["system", "Sistema"],
        }
        alternatives = module_lookup.get(module.lower(), [module])
        if len(alternatives) == 1:
            query = query.ilike("module", f"%{alternatives[0]}%")
        else:
            # Supabase/postgrest OR sobre el mismo campo.
            or_filter = ",".join(f"module.ilike.%{value}%" for value in alternatives)
            query = query.or_(or_filter)
    if status:
        query = query.ilike("status", f"%{status}%")
    if severity:
        query = query.ilike("severity", f"%{severity}%")
    if operation_id:
        query = query.ilike("operation_id", f"%{operation_id}%")
    if entity_id:
        query = query.ilike("entity_id", f"%{entity_id}%")
    if date_from:
        query = query.gte("created_at", date_from)
    if date_to:
        query = query.lte("created_at", date_to)

    response = query.order("created_at", desc=True).limit(limit).execute()
    rows = list(getattr(response, "data", None) or [])
    text = _clean_filter(filters.get("text"))
    if text:
        rows = [row for row in rows if _contains_text(row, text)]
    return [enrich_audit_log_row(row) for row in rows]


def enrich_audit_log_row(row: dict[str, Any]) -> dict[str, Any]:
    module = row.get("module")
    action = row.get("action")
    enriched = dict(row)
    enriched["visual_module"] = normalize_visual_module(module)
    enriched["visual_action"] = normalize_visual_action(module, action)
    enriched["visual_operation"] = f"{enriched['visual_module']} · {enriched['visual_action']}"
    return enriched


def list_operation_snapshots(session, *, operation_id: str = "", limit: int = 50) -> list[dict[str, Any]]:
    limit = max(1, min(int(limit or 50), 200))
    query = session.client.table("operation_snapshots").select(SNAPSHOT_COLUMNS)
    operation_id = _clean_filter(operation_id)
    if operation_id:
        query = query.eq("operation_id", operation_id)
    response = query.order("created_at", desc=True).limit(limit).execute()
    return list(getattr(response, "data", None) or [])


def get_snapshot_by_operation(session, operation_id: str) -> dict[str, Any] | None:
    rows = list_operation_snapshots(session, operation_id=operation_id, limit=1)
    return rows[0] if rows else None


def _flatten_for_diff(value: Any, prefix: str = "") -> dict[str, Any]:
    if value is None:
        return {}
    if isinstance(value, dict):
        out: dict[str, Any] = {}
        for key, child in value.items():
            child_key = f"{prefix}.{key}" if prefix else str(key)
            if isinstance(child, dict):
                nested = _flatten_for_diff(child, child_key)
                if nested:
                    out.update(nested)
                else:
                    out[child_key] = child
            elif isinstance(child, list):
                out[child_key] = child
            else:
                out[child_key] = child
        return out
    return {prefix or "value": value}


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False, indent=2)
        except Exception:
            return str(value)
    return str(value)


def build_before_after_diff(before_data: Any, after_data: Any) -> list[dict[str, Any]]:
    before = _flatten_for_diff(before_data)
    after = _flatten_for_diff(after_data)
    keys = sorted(set(before.keys()) | set(after.keys()))
    rows: list[dict[str, Any]] = []
    for key in keys:
        b = before.get(key)
        a = after.get(key)
        if _display_value(b) == _display_value(a):
            continue
        rows.append(
            {
                "field": key,
                "before": _display_value(b),
                "after": _display_value(a),
                "is_complex": isinstance(b, (dict, list)) or isinstance(a, (dict, list)) or len(_display_value(b)) > 160 or len(_display_value(a)) > 160,
            }
        )
    return rows


def security_log_kpis(rows: list[dict[str, Any]]) -> dict[str, Any]:
    today = datetime.now(timezone.utc).date()
    events_today = 0
    errors_today = 0
    critical = 0
    last_operation = "-"
    last_user = "-"
    for idx, row in enumerate(rows):
        if idx == 0:
            last_operation = row.get("visual_operation") or row.get("action") or "-"
            last_user = row.get("user_email") or "-"
        created = _safe_text(row.get("created_at"))
        try:
            dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
            if dt.date() == today:
                events_today += 1
                if _safe_text(row.get("status")).upper() in {"ERROR", "CRITICAL", "BLOCKED"} or _safe_text(row.get("severity")).upper() in {"ERROR", "CRITICAL"}:
                    errors_today += 1
        except Exception:
            pass
        if _safe_text(row.get("severity")).upper() == "CRITICAL" or _safe_text(row.get("status")).upper() == "CRITICAL":
            critical += 1
    return {
        "events_today": events_today,
        "errors_today": errors_today,
        "critical": critical,
        "last_operation": last_operation,
        "last_user": last_user,
    }


def export_security_logs_excel(rows: list[dict[str, Any]], snapshots: list[dict[str, Any]], output_path: str | Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    output_path = Path(output_path)
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    ws_logs = wb.create_sheet("Logs visibles")
    ws_snap = wb.create_sheet("Snapshots")

    title_fill = PatternFill("solid", fgColor="0F172A")
    header_fill = PatternFill("solid", fgColor="1E293B")
    soft_fill = PatternFill("solid", fgColor="E0E7FF")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    ws_summary["A1"] = "FutonHUB · Seguridad / Logs"
    ws_summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws_summary["A1"].fill = title_fill
    ws_summary.merge_cells("A1:E1")
    kpis = security_log_kpis(rows)
    for idx, (label, value) in enumerate(
        [
            ("Eventos hoy", kpis.get("events_today")),
            ("Errores hoy", kpis.get("errors_today")),
            ("Críticos", kpis.get("critical")),
            ("Última operación", kpis.get("last_operation")),
            ("Último usuario", kpis.get("last_user")),
        ],
        start=3,
    ):
        ws_summary.cell(row=idx, column=1, value=label)
        ws_summary.cell(row=idx, column=2, value=value)
        ws_summary.cell(row=idx, column=1).font = Font(bold=True)
        ws_summary.cell(row=idx, column=1).fill = soft_fill
        ws_summary.cell(row=idx, column=1).border = border
        ws_summary.cell(row=idx, column=2).border = border

    headers = [
        "Fecha",
        "Usuario",
        "Rol",
        "Módulo",
        "Acción",
        "Estado",
        "Severidad",
        "Entidad",
        "ID Entidad",
        "Operation ID",
        "Mensaje",
        "Error",
    ]
    ws_logs.append(headers)
    for row in rows:
        ws_logs.append(
            [
                row.get("created_at"),
                row.get("user_email"),
                row.get("role"),
                row.get("visual_module") or row.get("module"),
                row.get("visual_action") or row.get("action"),
                row.get("status"),
                row.get("severity"),
                row.get("entity_type"),
                row.get("entity_id"),
                row.get("operation_id"),
                row.get("message"),
                row.get("error_detail"),
            ]
        )

    ws_snap.append(["Fecha", "Operation ID", "Módulo", "Acción", "Entidad", "ID Entidad", "Razón", "Before data"])
    for snap in snapshots:
        ws_snap.append(
            [
                snap.get("created_at"),
                snap.get("operation_id"),
                snap.get("module"),
                snap.get("action"),
                snap.get("entity_type"),
                snap.get("entity_id"),
                snap.get("reason"),
                json.dumps(_json_safe(snap.get("before_data")), ensure_ascii=False, indent=2),
            ]
        )

    for ws in (ws_logs, ws_snap):
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            width = min(48, max(10, max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 80) + 1)) + 2))
            ws.column_dimensions[letter].width = width

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# =====================================================
# Snapshot rollback / restore
# =====================================================

def _as_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _strip_non_writable(row: dict[str, Any]) -> dict[str, Any]:
    """Remove common generated/read-only fields before update."""
    blocked = {
        "id",
        "created_at",
        "created_by",
        "user_id",
        "user_email",
    }
    return {k: v for k, v in dict(row or {}).items() if k not in blocked}


def _snapshot_entity(snapshot: dict[str, Any]) -> tuple[str, str, str, str]:
    module = str(snapshot.get("module") or "").strip()
    action = str(snapshot.get("action") or "").strip()
    entity_type = str(snapshot.get("entity_type") or "").strip()
    entity_id = str(snapshot.get("entity_id") or "").strip()
    return module, action, entity_type, entity_id


def preview_restore_snapshot(session, snapshot: dict[str, Any]) -> dict[str, Any]:
    """Preview a safe restore based on operation_snapshots.before_data.

    Supported:
    - inventory_items / inventory_item
    - supplier_orders single order snapshots
    - supplier_orders receive_order compound snapshot with order + lines + inventory

    Not supported:
    - snapshots with no before_data
    - generic unknown entities
    - creation snapshots that only contain created_payload
    """
    module, action, entity_type, entity_id = _snapshot_entity(snapshot)
    before_data = _as_dict(snapshot.get("before_data"))
    operation_id = str(snapshot.get("operation_id") or "")

    if not before_data:
        return {
            "supported": False,
            "reason": "El snapshot no contiene before_data restaurable.",
            "operation_id": operation_id,
            "changes": [],
        }

    if "created_payload" in before_data:
        return {
            "supported": False,
            "reason": "Este snapshot es de creación. Restaurar implicaría borrar el registro creado y se deja para rollback v2 con reglas específicas.",
            "operation_id": operation_id,
            "changes": [],
        }

    low_module = module.lower()
    low_entity = entity_type.lower()
    changes: list[dict[str, Any]] = []

    # Publicación Woo: snapshot compuesto con propuesta, espejo cloud y Woo anterior.
    if low_module == "woocommerce_publish" or action == "admin_publish_woocommerce_price":
        bundle = before_data
        proposal = _as_dict(bundle.get("proposal"))
        cloud_item = _as_dict(bundle.get("cloud_item"))
        woo_before = _as_dict(bundle.get("woo_before"))
        if not proposal or not woo_before:
            return {
                "supported": False,
                "reason": "El snapshot de publicación Woo no contiene proposal/woo_before restaurables.",
                "operation_id": operation_id,
                "changes": [],
            }
        return {
            "supported": True,
            "reason": "",
            "operation_id": operation_id,
            "module": module,
            "action": action,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "special_restore": "woocommerce_publish",
            "proposal": proposal,
            "cloud_item": cloud_item,
            "woo_before": woo_before,
            "changes": [
                {
                    "table": "WooCommerce",
                    "match_field": "woo_id",
                    "match_value": proposal.get("item_woo_id") or proposal.get("local_id"),
                    "before": {
                        "regular_price": woo_before.get("regular_price"),
                        "sale_price": woo_before.get("sale_price"),
                        "price": woo_before.get("price"),
                    },
                    "description": "Restaurar precio efectivo anterior en WooCommerce y espejo Supabase",
                }
            ],
        }

    # Recepción de pedido: snapshot compuesto con order, lines e inventory.
    if (low_module in {"supplier_orders", "pedidos"} and action == "receive_order") or (
        "order" in before_data and "inventory" in before_data
    ):
        order = _as_dict(before_data.get("order"))
        lines = before_data.get("lines") if isinstance(before_data.get("lines"), list) else []
        inventory = before_data.get("inventory") if isinstance(before_data.get("inventory"), list) else []
        if order:
            changes.append({
                "table": "supplier_orders",
                "match_field": "order_id",
                "match_value": str(order.get("order_id") or entity_id),
                "before": order,
                "description": f"Restaurar pedido {order.get('order_id') or entity_id}",
            })
        for line in lines:
            if isinstance(line, dict):
                changes.append({
                    "table": "supplier_order_items",
                    "match_field": "id" if line.get("id") not in (None, "") else "item_code",
                    "match_value": line.get("id") if line.get("id") not in (None, "") else line.get("item_code"),
                    "extra_match": {"order_id": order.get("order_id") or entity_id} if line.get("id") in (None, "") else {},
                    "before": line,
                    "description": f"Restaurar línea {line.get('item_code') or line.get('id')}",
                })
        for inv in inventory:
            if isinstance(inv, dict):
                changes.append({
                    "table": "inventory_items",
                    "match_field": "item_id",
                    "match_value": inv.get("item_id"),
                    "before": inv,
                    "description": f"Restaurar stock/item {inv.get('item_id')}",
                })
        return {
            "supported": bool(changes),
            "reason": "" if changes else "No se encontraron datos restaurables en recepción.",
            "operation_id": operation_id,
            "module": module,
            "action": action,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "changes": changes,
        }

    # Inventory item direct restore.
    if low_module in {"inventory_items", "inventario"} or low_entity in {"inventory_item", "inventory_items"}:
        before = before_data
        item_id = before.get("item_id") or entity_id
        if item_id in (None, ""):
            return {"supported": False, "reason": "No se pudo determinar item_id para restaurar inventory_items.", "operation_id": operation_id, "changes": []}
        changes.append({
            "table": "inventory_items",
            "match_field": "item_id",
            "match_value": item_id,
            "before": before,
            "description": f"Restaurar artículo {item_id}",
        })
        return {
            "supported": True,
            "reason": "",
            "operation_id": operation_id,
            "module": module,
            "action": action,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "changes": changes,
        }

    # Supplier order direct restore.
    if low_module in {"supplier_orders", "pedidos"} or low_entity in {"supplier_order", "supplier_orders"}:
        before = before_data
        order_id = before.get("order_id") or entity_id
        if order_id in (None, ""):
            return {"supported": False, "reason": "No se pudo determinar order_id para restaurar supplier_orders.", "operation_id": operation_id, "changes": []}
        changes.append({
            "table": "supplier_orders",
            "match_field": "order_id",
            "match_value": str(order_id),
            "before": before,
            "description": f"Restaurar pedido {order_id}",
        })
        return {
            "supported": True,
            "reason": "",
            "operation_id": operation_id,
            "module": module,
            "action": action,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "changes": changes,
        }

    return {
        "supported": False,
        "reason": f"Rollback no soportado todavía para módulo={module}, entidad={entity_type}.",
        "operation_id": operation_id,
        "changes": [],
    }


def restore_snapshot_to_previous_state(session, snapshot: dict[str, Any]) -> dict[str, Any]:
    from futonhub.cloud.audit import AuditEvent, new_operation_id, write_audit_event
    from futonhub.core.config import load_settings

    preview = preview_restore_snapshot(session, snapshot)
    if not preview.get("supported"):
        raise ValueError(preview.get("reason") or "Snapshot no soportado para restauración.")
    settings = load_settings()
    restore_operation_id = new_operation_id("RESTORE")
    restored: list[dict[str, Any]] = []
    errors: list[str] = []

    if preview.get("special_restore") == "woocommerce_publish":
        from futonhub.cloud.services.inventory import sync_woocommerce_price_inventory_state
        from gestorwoo.woocommerce import WooCommerceClient
        proposal = _as_dict(preview.get("proposal"))
        cloud_item = _as_dict(preview.get("cloud_item"))
        woo_before = _as_dict(preview.get("woo_before"))
        kind = str(proposal.get("item_kind") or "").strip().lower()
        woo_id = int(proposal.get("item_woo_id") or proposal.get("local_id") or 0)
        parent_id = cloud_item.get("parent_woo_id") or (_as_dict(proposal.get("source_row")).get("item_snapshot") or {}).get("parent_woo_id")
        payload = {
            "regular_price": str(woo_before.get("regular_price") or ""),
            "sale_price": str(woo_before.get("sale_price") or ""),
        }
        client = WooCommerceClient(settings.woocommerce_url, settings.consumer_key, settings.consumer_secret)
        if kind == "product":
            client.update_product_pricing(woo_id, payload)
            verified = client.get(f"products/{woo_id}").json()
            mirror_table = "products"
        elif kind == "variation":
            if not parent_id:
                raise ValueError(f"No se pudo determinar parent_woo_id para rollback de variación {woo_id}.")
            client.update_variation_pricing(int(parent_id), woo_id, payload)
            verified = client.get(f"products/{int(parent_id)}/variations/{woo_id}").json()
            mirror_table = "product_variations"
        else:
            raise ValueError("item_kind inválido en snapshot Woo.")

        expected_regular = str(woo_before.get("regular_price") or "")
        expected_sale = str(woo_before.get("sale_price") or "")
        got_regular = str(verified.get("regular_price") or "")
        got_sale = str(verified.get("sale_price") or "")
        if got_regular != expected_regular or got_sale != expected_sale:
            raise ValueError(
                f"Rollback Woo no verificado. Esperado regular={expected_regular!r}, sale={expected_sale!r}; "
                f"obtenido regular={got_regular!r}, sale={got_sale!r}."
            )

        snapshot_before = _as_dict(snapshot.get("before_data"))
        preview_row = _as_dict(snapshot_before.get("preview_row"))
        rollback_from_price = (
            proposal.get("new_price")
            if proposal.get("new_price") not in (None, "")
            else preview_row.get("new_price")
        )
        try:
            inventory_sync_result = sync_woocommerce_price_inventory_state(
                session,
                operation_id=restore_operation_id,
                proposal=proposal,
                cloud_item=cloud_item,
                woo_id=woo_id,
                before_price=_format_price_value(rollback_from_price),
                verified_price=_format_price_value(_effective_woo_price(verified)),
                action="restore_woocommerce_price_snapshot",
                message="Precio Woo restaurado desde snapshot y verificado.",
                metadata={
                    "source_operation_id": snapshot.get("operation_id"),
                    "proposal_id": proposal.get("id"),
                    "item_kind": kind,
                    "woo_id": woo_id,
                    "parent_woo_id": parent_id,
                    "payload": _json_safe(payload),
                },
            )
        except Exception as exc:
            write_audit_event(
                session,
                AuditEvent(
                    operation_id=restore_operation_id,
                    module="woocommerce_publish",
                    action="restore_woocommerce_price_snapshot_partial_internal_sync_failed",
                    status="ERROR",
                    severity="CRITICAL",
                    entity_type="price_change_proposal",
                    entity_id=str(proposal.get("id") or snapshot.get("entity_id") or ""),
                    before_data={"snapshot": _json_safe(snapshot)},
                    after_data={"woo_restored": True, "payload": payload, "internal_sync_error": str(exc)},
                    message="WooCommerce fue restaurado y verificado, pero fallo la sincronizacion interna de Inventario/historial.",
                    error_detail=str(exc),
                ),
                settings,
            )
            raise

        session.client.table(mirror_table).update({
            "price": str(verified.get("price") or ""),
            "regular_price": got_regular,
            "sale_price": got_sale,
        }).eq("woo_id", woo_id).execute()
        proposal_id = proposal.get("id")
        if proposal_id:
            rollback_source = {
                **_as_dict(proposal.get("source_row")),
                "rolled_back": True,
                "rolled_back_from_operation_id": snapshot.get("operation_id"),
                "rollback_operation_id": restore_operation_id,
                "woo_after_rollback_verified": _json_safe(verified),
                "inventory_sync": _json_safe(inventory_sync_result),
                "inventory_history": _json_safe(inventory_sync_result.get("history")),
                "inventory_history_resolution": _json_safe(inventory_sync_result.get("resolution")),
            }
            rollback_update = {
                "status": "rolled_back",
                "error_message": None,
                "source_row": rollback_source,
            }
            try:
                session.client.table("price_change_proposals").update(rollback_update).eq("id", proposal_id).execute()
            except Exception as exc:
                # Compatibilidad con instalaciones cuyo CHECK de status todavía no incluye rolled_back.
                # Woo ya fue restaurado y verificado, así que no convertimos un rollback exitoso
                # en un falso fallo de la operación. Conservamos status=published y marcamos
                # explícitamente el rollback dentro de source_row.
                msg = str(exc)
                if "price_change_proposals_status_check" not in msg and "23514" not in msg:
                    raise
                session.client.table("price_change_proposals").update({
                    "status": "published",
                    "error_message": None,
                    "source_row": {
                        **rollback_source,
                        "rolled_back_status_fallback": True,
                    },
                }).eq("id", proposal_id).execute()
        restored.append({
            "table": "WooCommerce",
            "match_field": "woo_id",
            "match_value": woo_id,
            "rows": [_json_safe(verified)],
            "description": "Precio Woo restaurado y verificado",
        })
        try:
            write_audit_event(
                session,
                AuditEvent(
                    operation_id=restore_operation_id,
                    module="woocommerce_publish",
                    action="restore_woocommerce_price_snapshot",
                    status="OK",
                    severity="WARNING",
                    entity_type="price_change_proposal",
                    entity_id=str(proposal_id or snapshot.get("entity_id") or ""),
                    before_data={"snapshot": _json_safe(snapshot)},
                    after_data={"woo_verified": _json_safe(verified), "payload": payload},
                    message="Precio Woo restaurado desde snapshot y verificado mediante lectura posterior.",
                ),
                settings,
            )
        except Exception:
            pass
        return {
            "operation_id": restore_operation_id,
            "restored": restored,
            "preview": preview,
            "inventory_sync": inventory_sync_result,
            "inventory_history": inventory_sync_result.get("history"),
            "inventory_history_resolution": inventory_sync_result.get("resolution"),
        }

    for change in preview.get("changes", []):
        table = str(change.get("table") or "")
        match_field = str(change.get("match_field") or "")
        match_value = change.get("match_value")
        before = _strip_non_writable(_as_dict(change.get("before")))
        if not table or not match_field or match_value in (None, "") or not before:
            errors.append(f"Cambio inválido: {change}")
            continue
        try:
            payload = dict(before)
            # Trazabilidad de restauración sin pisar source_row si existe.
            source_row = payload.get("source_row")
            if not isinstance(source_row, dict):
                source_row = {}
            source_row = {
                **source_row,
                "restored_from_operation_id": snapshot.get("operation_id"),
                "restore_operation_id": restore_operation_id,
                "restored_by_email": getattr(session, "email", None),
            }
            payload["source_row"] = source_row

            query = session.client.table(table).update(payload).eq(match_field, match_value)
            extra = change.get("extra_match")
            if isinstance(extra, dict):
                for key, value in extra.items():
                    if value not in (None, ""):
                        query = query.eq(key, value)
            response = query.execute()
            restored.append({
                "table": table,
                "match_field": match_field,
                "match_value": match_value,
                "rows": getattr(response, "data", None) or [],
                "description": change.get("description"),
            })
        except Exception as exc:
            errors.append(f"{table}.{match_field}={match_value}: {exc}")

    status = "OK" if not errors else "ERROR"
    try:
        write_audit_event(
            session,
            AuditEvent(
                operation_id=restore_operation_id,
                module="Seguridad",
                action="restore_snapshot",
                status=status,
                severity="INFO" if not errors else "ERROR",
                entity_type=str(snapshot.get("entity_type") or "snapshot"),
                entity_id=str(snapshot.get("entity_id") or snapshot.get("operation_id") or ""),
                before_data={"snapshot": _json_safe(snapshot), "preview": _json_safe(preview)},
                after_data={"restored": _json_safe(restored), "errors": errors},
                message=f"Restauración desde snapshot {snapshot.get('operation_id')}",
                error_detail="; ".join(errors),
            ),
            settings,
        )
    except Exception:
        pass

    if errors:
        raise ValueError("Restauración con errores: " + "; ".join(errors))
    return {
        "operation_id": restore_operation_id,
        "restored": restored,
        "preview": preview,
    }
