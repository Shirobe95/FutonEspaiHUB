from __future__ import annotations

import inspect
import json
import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from futonhub.cloud.services import orders as orders_service  # noqa: E402
from futonhub.cloud.services.supplier_prices import (  # noqa: E402
    SupplierOrderCodeAmbiguityError,
    resolve_supplier_order_effective_price,
    resolve_supplier_order_inventory_items,
)
from futonhub.core.codes import (  # noqa: E402
    is_inventory_pack_row,
    is_supplier_order_eligible_inventory_row,
    normalize_inventory_numeric_code,
)
from futonhub.ui.erp import prototype as prototype_module  # noqa: E402
from futonhub.ui.erp.prototype import FutonHubErpPrototype  # noqa: E402
from futonhub.ui.erp.shared_ui import OrderItem  # noqa: E402


class Entry:
    def __init__(self, value: str) -> None:
        self.value = value

    def get(self) -> str:
        return self.value


class Window:
    def __init__(self) -> None:
        self.destroyed = False

    def destroy(self) -> None:
        self.destroyed = True


class Session:
    user_id = "USER-1"
    email = "admin@example.com"
    role = "admin"


class MemoryQuery:
    def __init__(self, client: "MemoryClient", table: str) -> None:
        self.client = client
        self.table = table
        self.action = "select"
        self.payload = None
        self.filters: list[tuple[str, str, object]] = []
        self.limit_value: int | None = None
        self.range_value: tuple[int, int] | None = None

    def select(self, _columns: str) -> "MemoryQuery":
        self.action = "select"
        return self

    def insert(self, payload: object) -> "MemoryQuery":
        self.action = "insert"
        self.payload = payload
        return self

    def update(self, payload: dict[str, object]) -> "MemoryQuery":
        self.action = "update"
        self.payload = payload
        return self

    def delete(self) -> "MemoryQuery":
        self.action = "delete"
        return self

    def eq(self, field: str, value: object) -> "MemoryQuery":
        self.filters.append(("eq", field, value))
        return self

    def neq(self, field: str, value: object) -> "MemoryQuery":
        self.filters.append(("neq", field, value))
        return self

    def order(self, _field: str, desc: bool = False) -> "MemoryQuery":
        return self

    def limit(self, value: int) -> "MemoryQuery":
        self.limit_value = value
        return self

    def range(self, start: int, end: int) -> "MemoryQuery":
        self.range_value = (start, end)
        return self

    def _matches(self, row: dict[str, object]) -> bool:
        for operator, field, value in self.filters:
            if operator == "eq" and str(row.get(field)) != str(value):
                return False
            if operator == "neq" and str(row.get(field)) == str(value):
                return False
        return True

    def execute(self) -> SimpleNamespace:
        self.client.execute_counts[self.table] = self.client.execute_counts.get(self.table, 0) + 1
        rows = self.client.tables.setdefault(self.table, [])
        if self.action == "select":
            selected = [json.loads(json.dumps(row)) for row in rows if self._matches(row)]
            if self.range_value is not None:
                start, end = self.range_value
                selected = selected[start : end + 1]
            if self.limit_value is not None:
                selected = selected[: self.limit_value]
            return SimpleNamespace(data=selected)
        if self.action == "delete":
            deleted = [row for row in rows if self._matches(row)]
            self.client.tables[self.table] = [row for row in rows if not self._matches(row)]
            return SimpleNamespace(data=json.loads(json.dumps(deleted)))
        if self.action == "update":
            updated = []
            for row in rows:
                if self._matches(row):
                    row.update(json.loads(json.dumps(self.payload)))
                    updated.append(json.loads(json.dumps(row)))
            return SimpleNamespace(data=updated)
        payloads = self.payload if isinstance(self.payload, list) else [self.payload]
        inserted = []
        for payload in payloads:
            row = json.loads(json.dumps(payload))
            if self.table == "supplier_orders":
                row.setdefault("order_id", f"ORDER-{len(rows) + 1}")
            else:
                row.setdefault("id", len(rows) + 1)
            rows.append(row)
            inserted.append(json.loads(json.dumps(row)))
        return SimpleNamespace(data=inserted)


class MemoryClient:
    def __init__(self) -> None:
        self.execute_counts: dict[str, int] = {}
        self.tables: dict[str, list[dict[str, object]]] = {
            "inventory_items": [],
            "supplier_orders": [],
            "supplier_order_items": [],
        }

    def table(self, name: str) -> MemoryQuery:
        return MemoryQuery(self, name)


class MemorySession(Session):
    def __init__(self) -> None:
        self.client = MemoryClient()


def app() -> FutonHubErpPrototype:
    instance = FutonHubErpPrototype.__new__(FutonHubErpPrototype)
    instance._cloud_session = None
    instance._orders_loaded_once = True
    instance._supplier_orders = []
    instance._selected_supplier_order = None
    constants = {
        "IMPORTE_DESCARGA_MT": 0.0,
        "PC_GASTOS_MANIPULACION": 0.0,
        "PC_GASTOS_FINANCIACION": 0.0,
        "IMPORTES_VARIOS": 0.0,
        "COSTE_TOTAL_DESCARGA_FUTONES_IVA": 0.0,
        "COSTE_DESCARGA_FUTONES_UNIDAD": 0.0,
        "IVA_RECARGO_EQUIVALENCIA": 0.0,
        "IVA_RECARGO_EQUIVALENCIA_FACTOR": 0.0,
        "COSTE_DIARIO_ALMACENAJE_M3": 0.0,
    }
    instance._current_business_constant_values = lambda: constants
    instance._fill_supplier_prices_for_order_items = lambda _provider, items: items
    instance._show_view = lambda _key: None
    return instance


def general_item() -> OrderItem:
    source = {
        "m3_total": 1,
        "m3_und": 1,
        "precio_proveedor": 93.65,
        "rotation_c": 1,
        "packages": 1,
        "cuenta_reparto_descarga": True,
        "price_input_source": "global_margin",
        "use_global_rentability": True,
    }
    return OrderItem("201014", "Futon prueba", 1, "1", "Pendiente", "OK", raw={"source_row": source})


def heimei_item() -> OrderItem:
    source = {
        "m3_total": 1,
        "m3_und": 1,
        "precio_proveedor": 94.652,
        "rotation_c": 1,
        "packages": 1,
        "cuenta_reparto_descarga": True,
        "price_input_source": "global_margin",
        "use_global_rentability": True,
    }
    return OrderItem("301014", "Tatami prueba", 1, "1", "Pendiente", "OK", raw={"source_row": source})


def inventory_row(
    item_id: int | str,
    *,
    name: str = "Futon inventario",
    heca_reference: str = "",
    hub_item_code: str = "",
    woo_sku: str = "",
    item_record_type: str = "simple",
    is_pack: bool = False,
    base_item_code: str = "",
    woo_id: int = 0,
    woo_item_kind: str = "",
    woo_parent_id: int = 0,
    woo_link_status: str = "",
    source_row: dict[str, object] | None = None,
    primary_supplier_price: object = 93.65,
    pascal_price: object = 0,
    rotation_c: float = 2,
    store_stock: object = 0,
    warehouse_stock: object = 0,
    weighted_average_cost: object = 0,
) -> dict[str, object]:
    return {
        "item_id": item_id,
        "name": name,
        "heca_reference": heca_reference,
        "hub_item_code": hub_item_code,
        "woo_sku": woo_sku,
        "item_record_type": item_record_type,
        "is_pack": is_pack,
        "base_item_code": base_item_code,
        "woo_id": woo_id,
        "woo_item_kind": woo_item_kind,
        "woo_parent_id": woo_parent_id,
        "woo_link_status": woo_link_status,
        "primary_supplier_price": primary_supplier_price,
        "pascal_price": pascal_price,
        "cubic_meters": 1,
        "rotation_c": rotation_c,
        "packages": 1,
        "store_stock": store_stock,
        "warehouse_stock": warehouse_stock,
        "weighted_average_cost": weighted_average_cost,
        "order_calculated_price": 0,
        "updated_at": "2026-06-19T08:00:00+00:00",
        "source_row": source_row or {},
    }


class SupplierOrderCostTests(unittest.TestCase):
    def woo_item(
        self,
        *,
        code: str = "201014",
        woo_id: int = 123,
        woo_item_kind: str = "product",
        woo_parent_id: int | None = None,
        price_input_source: str | None = None,
        extra_source: dict[str, object] | None = None,
    ) -> OrderItem:
        source: dict[str, object] = {
            "m3_total": 1,
            "m3_und": 1,
            "precio_proveedor": 93.65,
            "rotation_c": 1,
            "packages": 1,
            "cuenta_reparto_descarga": True,
            "inventory_stock_total": 1,
            "inventory_weighted_average_cost": 80,
            "woo_id": woo_id,
            "woo_item_kind": woo_item_kind,
        }
        if woo_parent_id is not None:
            source["woo_parent_id"] = woo_parent_id
        if price_input_source is not None:
            source["price_input_source"] = price_input_source
        if extra_source:
            source.update(extra_source)
        return OrderItem(code, "Futon Woo", 1, "1", "Pendiente", "OK", raw={"source_row": source})

    def test_general_supplier_separates_cost_margin_and_pvp(self) -> None:
        ui = app()

        items, _raw, summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
            (general_item(),),
            [],
        )

        source = items[0].raw["source_row"]
        self.assertEqual(source["unit_cost"], 100.0)
        self.assertEqual(source["line_cost"], 100.0)
        self.assertEqual(summary["total_cost"], 100.0)
        self.assertEqual(source["rentabilidad_percent"], 30.0)
        self.assertEqual(source["pvp_unit"], 142.86)
        self.assertEqual(source["pvp_line"], 142.86)

    def test_sales_margin_example_5268_and_40_produces_8780(self) -> None:
        ui = app()

        self.assertEqual(ui._supplier_order_pvp("52.68", "40"), 87.80)

    def test_pvp_example_8780_and_5268_produces_40_margin(self) -> None:
        ui = app()

        self.assertEqual(ui._supplier_order_margin_from_pvp("52.68", "87.80"), 40.00)

    def test_margin_to_pvp_to_margin_roundtrip(self) -> None:
        ui = app()

        pvp = ui._supplier_order_pvp("52.68", "40")
        margin = ui._supplier_order_margin_from_pvp("52.68", pvp)

        self.assertAlmostEqual(margin, 40.0, places=2)

    def test_pvp_to_margin_to_pvp_roundtrip(self) -> None:
        ui = app()

        margin = ui._supplier_order_margin_from_pvp("52.68", "87.80")
        pvp = ui._supplier_order_pvp("52.68", margin)

        self.assertAlmostEqual(pvp, 87.80, places=2)

    def test_zero_margin_keeps_pvp_equal_to_real_cost(self) -> None:
        ui = app()

        items, _raw, summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "0", "Coste transporte + IVA": "1"},
            (general_item(),),
            [],
        )

        source = items[0].raw["source_row"]
        self.assertEqual(source["unit_cost"], 100.0)
        self.assertEqual(source["pvp_unit"], 100.0)
        self.assertEqual(source["pvp_line"], 100.0)
        self.assertEqual(summary["total_cost"], 100.0)

    def test_heimei_supplier_separates_cost_margin_and_pvp(self) -> None:
        ui = app()

        items, _raw, summary = ui._calculate_supplier_order_in_memory(
            "heimei",
            {
                "Rentabilidad %": "30",
                "Precio en Dolares": "100",
                "Precio pagado en Euros": "100",
                "Factura transporte": "0",
                "Derechos aranceles": "0",
            },
            (heimei_item(),),
            [],
        )

        source = items[0].raw["source_row"]
        self.assertEqual(source["unit_cost"], 100.0)
        self.assertEqual(source["line_cost"], 100.0)
        self.assertEqual(summary["total_cost"], 100.0)
        self.assertEqual(source["pvp_unit"], 142.86)
        self.assertEqual(source["pvp_line"], 142.86)

    def test_sales_margin_formula_is_used_in_order_calculation(self) -> None:
        source = inspect.getsource(FutonHubErpPrototype._calculate_supplier_order_in_memory)

        self.assertIn('price_input_source == "pvp"', source)
        self.assertIn('price_input_source == "woo_price"', source)
        self.assertIn("_supplier_order_margin_from_pvp(precio_ponderado_lote, pvp_unit)", source)
        self.assertIn("_supplier_order_pvp(precio_ponderado_lote, effective_rent_percent", source)
        self.assertNotIn("* (1 + rent_percent / 100)", source)

    def test_woo_price_is_default_source_and_margin_uses_weighted_cost(self) -> None:
        ui = app()
        with patch.object(ui, "_supplier_order_fetch_woo_price", return_value={"price": 120, "checked_at": "now", "woo_price_item_id": 123, "woo_price_item_kind": "product", "woo_price_parent_id": ""}):
            items, _raw, summary = ui._calculate_supplier_order_in_memory(
                "ekomat",
                {"Margen de Venta %": "40", "Coste transporte + IVA": "1"},
                (self.woo_item(),),
                [],
            )

        source = items[0].raw["source_row"]
        self.assertEqual(source["precio_ponderado_lote"], 90.0)
        self.assertEqual(source["pvp_unit"], 120.0)
        self.assertEqual(source["rentabilidad_percent"], 25.0)
        self.assertEqual(source["price_input_source"], "woo_price")
        self.assertEqual(source["pvp_source_label"], "WooCommerce")
        self.assertEqual(summary["total_cost"], 100.0)

    def test_woo_price_does_not_fallback_to_global_margin(self) -> None:
        ui = app()
        with patch.object(ui, "_supplier_order_fetch_woo_price", return_value={"price": 120, "checked_at": "now", "woo_price_item_id": 123, "woo_price_item_kind": "product", "woo_price_parent_id": ""}):
            items, _raw, _summary = ui._calculate_supplier_order_in_memory(
                "ekomat",
                {"Margen de Venta %": "40", "Coste transporte + IVA": "1"},
                (self.woo_item(),),
                [],
            )

        source = items[0].raw["source_row"]
        self.assertEqual(source["pvp_unit"], 120.0)
        self.assertNotEqual(source["pvp_unit"], 150.0)
        self.assertEqual(source["rentabilidad_percent"], 25.0)

    def test_missing_woo_price_marks_pvp_and_margin_pending(self) -> None:
        ui = app()
        with patch.object(ui, "_supplier_order_fetch_woo_price", return_value={"price": None, "error": "sin price", "checked_at": "now", "woo_price_item_id": 123, "woo_price_item_kind": "product", "woo_price_parent_id": ""}):
            items, _raw, summary = ui._calculate_supplier_order_in_memory(
                "ekomat",
                {"Margen de Venta %": "40", "Coste transporte + IVA": "1"},
                (self.woo_item(),),
                [],
            )

        source = items[0].raw["source_row"]
        self.assertEqual(items[0].status, "Warning")
        self.assertIsNone(source["pvp_unit"])
        self.assertIsNone(source["rentabilidad_percent"])
        self.assertEqual(source["pvp_source"], "pending")
        self.assertEqual(source["woo_price_error"], "sin price")
        self.assertEqual(summary["total_cost"], 100.0)

    def test_zero_woo_price_is_controlled_pending(self) -> None:
        ui = app()
        with patch.object(ui, "_supplier_order_fetch_woo_price", return_value={"price": 0, "error": "precio cero", "checked_at": "now", "woo_price_item_id": 123, "woo_price_item_kind": "product", "woo_price_parent_id": ""}):
            items, _raw, _summary = ui._calculate_supplier_order_in_memory(
                "ekomat",
                {"Margen de Venta %": "40", "Coste transporte + IVA": "1"},
                (self.woo_item(),),
                [],
            )

        source = items[0].raw["source_row"]
        self.assertEqual(items[0].status, "Warning")
        self.assertIsNone(source["pvp_unit"])
        self.assertEqual(source["pvp_source_label"], "Pendiente")

    def test_manual_pvp_wins_over_woo(self) -> None:
        ui = app()
        with patch.object(ui, "_supplier_order_fetch_woo_price", side_effect=AssertionError("Woo no debe consultarse")):
            items, _raw, _summary = ui._calculate_supplier_order_in_memory(
                "ekomat",
                {"Margen de Venta %": "40", "Coste transporte + IVA": "1"},
                (self.woo_item(price_input_source="pvp", extra_source={"pvp_unit": 80}),),
                [],
            )

        source = items[0].raw["source_row"]
        self.assertEqual(source["pvp_unit"], 80.0)
        self.assertEqual(source["rentabilidad_percent"], -12.5)
        self.assertEqual(source["pvp_source_label"], "Manual P.V.P.")

    def test_manual_margin_wins_over_woo(self) -> None:
        ui = app()
        with patch.object(ui, "_supplier_order_fetch_woo_price", side_effect=AssertionError("Woo no debe consultarse")):
            items, _raw, _summary = ui._calculate_supplier_order_in_memory(
                "ekomat",
                {"Margen de Venta %": "10", "Coste transporte + IVA": "1"},
                (self.woo_item(price_input_source="individual_margin", extra_source={"rentabilidad_individual_percent": 40, "use_global_rentability": False}),),
                [],
            )

        source = items[0].raw["source_row"]
        self.assertEqual(source["pvp_unit"], 150.0)
        self.assertEqual(source["pvp_source_label"], "Manual Margen")

    def test_explicit_global_margin_wins_over_woo(self) -> None:
        ui = app()
        with patch.object(ui, "_supplier_order_fetch_woo_price", side_effect=AssertionError("Woo no debe consultarse")):
            items, _raw, _summary = ui._calculate_supplier_order_in_memory(
                "ekomat",
                {"Margen de Venta %": "40", "Coste transporte + IVA": "1"},
                (self.woo_item(price_input_source="global_margin", extra_source={"use_global_rentability": True}),),
                [],
            )

        source = items[0].raw["source_row"]
        self.assertEqual(source["pvp_unit"], 150.0)
        self.assertEqual(source["pvp_source_label"], "Margen global")

    def test_variation_woo_price_uses_variation_endpoint(self) -> None:
        ui = app()

        class Response:
            def json(self) -> dict[str, str]:
                return {"price": "120"}

        class Client:
            endpoints: list[str] = []

            def __init__(self, *_args: object) -> None:
                pass

            def get(self, endpoint: str) -> Response:
                self.endpoints.append(endpoint)
                return Response()

            def put(self, *_args: object) -> None:
                raise AssertionError("Woo no debe escribirse")

        with patch.object(prototype_module, "load_settings", return_value=SimpleNamespace(woocommerce_url="https://woo.test", consumer_key="ck", consumer_secret="cs")):
            with patch.object(prototype_module, "WooCommerceClient", Client):
                result = ui._supplier_order_fetch_woo_price({"woo_id": 456, "woo_item_kind": "variation", "woo_parent_id": 123})

        self.assertEqual(result["price"], 120.0)
        self.assertEqual(Client.endpoints, ["products/123/variations/456"])

    def test_product_woo_price_uses_product_endpoint_and_no_writes(self) -> None:
        ui = app()

        class Response:
            def json(self) -> dict[str, str]:
                return {"price": "120"}

        class Client:
            endpoints: list[str] = []

            def __init__(self, *_args: object) -> None:
                pass

            def get(self, endpoint: str) -> Response:
                self.endpoints.append(endpoint)
                return Response()

            def put(self, *_args: object) -> None:
                raise AssertionError("Woo no debe escribirse")

        with patch.object(prototype_module, "load_settings", return_value=SimpleNamespace(woocommerce_url="https://woo.test", consumer_key="ck", consumer_secret="cs")):
            with patch.object(prototype_module, "WooCommerceClient", Client):
                result = ui._supplier_order_fetch_woo_price({"woo_id": 123, "woo_item_kind": "product"})

        self.assertEqual(result["price"], 120.0)
        self.assertEqual(Client.endpoints, ["products/123"])

    def test_woo_product_404_returns_clear_pending_error(self) -> None:
        ui = app()

        class Client:
            def __init__(self, *_args: object) -> None:
                pass

            def get(self, _endpoint: str) -> object:
                raise prototype_module.WooCommerceError("Error WooCommerce 404: not found")

            def put(self, *_args: object) -> None:
                raise AssertionError("Woo no debe escribirse")

            def post(self, *_args: object) -> None:
                raise AssertionError("Woo no debe escribirse")

        with patch.object(prototype_module, "load_settings", return_value=SimpleNamespace(woocommerce_url="https://woo.test", consumer_key="ck", consumer_secret="cs")):
            with patch.object(prototype_module, "WooCommerceClient", Client):
                result = ui._supplier_order_fetch_woo_price({"woo_id": 123, "woo_item_kind": "product"})

        self.assertIsNone(result["price"])
        self.assertIn("producto 123", result["error"])

    def test_woo_variation_404_returns_parent_and_variation_ids(self) -> None:
        ui = app()

        class Client:
            def __init__(self, *_args: object) -> None:
                pass

            def get(self, _endpoint: str) -> object:
                raise prototype_module.WooCommerceError("Error WooCommerce 404: not found")

            def put(self, *_args: object) -> None:
                raise AssertionError("Woo no debe escribirse")

        with patch.object(prototype_module, "load_settings", return_value=SimpleNamespace(woocommerce_url="https://woo.test", consumer_key="ck", consumer_secret="cs")):
            with patch.object(prototype_module, "WooCommerceClient", Client):
                result = ui._supplier_order_fetch_woo_price({"woo_id": 456, "woo_item_kind": "variation", "woo_parent_id": 123})

        self.assertIsNone(result["price"])
        self.assertIn("variacion 456", result["error"])
        self.assertIn("producto padre 123", result["error"])

    def test_variation_without_parent_is_pending(self) -> None:
        ui = app()

        result = ui._supplier_order_fetch_woo_price({"woo_id": 456, "woo_item_kind": "variation"})

        self.assertIsNone(result["price"])
        self.assertIn("parent Woo", result["error"])

    def test_partial_woo_failure_does_not_break_order(self) -> None:
        ui = app()

        def fetch(source: dict[str, object]) -> dict[str, object]:
            if source.get("woo_id") == 1:
                return {"price": 120, "checked_at": "now", "woo_price_item_id": 1, "woo_price_item_kind": "product", "woo_price_parent_id": ""}
            return {"price": None, "error": "fallo conexion", "checked_at": "now", "woo_price_item_id": 2, "woo_price_item_kind": "product", "woo_price_parent_id": ""}

        with patch.object(ui, "_supplier_order_fetch_woo_price", side_effect=fetch):
            items, _raw, summary = ui._calculate_supplier_order_in_memory(
                "ekomat",
                {"Margen de Venta %": "40", "Coste transporte + IVA": "1"},
                (self.woo_item(woo_id=1), self.woo_item(woo_id=2)),
                [],
            )

        self.assertEqual(items[0].status, "Calculado")
        self.assertEqual(items[1].status, "Warning")
        self.assertEqual(summary["total_cost"], 199.0)

    def test_calculate_loaded_order_uses_background_thread(self) -> None:
        source = inspect.getsource(FutonHubErpPrototype._open_order_calc_flow)

        self.assertIn("threading.Thread(target=worker, daemon=True).start()", source)
        self.assertIn("self.after(0, lambda result=result: finish_calculation(result))", source)
        self.assertIn("set_calculation_buttons_enabled(False)", source)
        self.assertIn("set_calculation_buttons_enabled(True)", source)

    def test_historical_line_without_marker_uses_woo_but_manual_signal_is_preserved(self) -> None:
        ui = app()
        self.assertEqual(ui._supplier_order_price_input_source({"use_global_rentability": True}, "global"), "global_margin")
        self.assertEqual(ui._supplier_order_price_input_source({"use_global_rentability": False, "rentabilidad_individual_percent": 15}, "individual"), "individual_margin")
        self.assertEqual(ui._supplier_order_price_input_source({"pvp_source": "manual_pvp", "pvp_unit": 80}, "individual"), "pvp")
        self.assertEqual(ui._supplier_order_price_input_source({}, "global"), "woo_price")

    def test_editor_initial_global_checkbox_depends_on_price_input_source(self) -> None:
        ui = app()

        self.assertFalse(ui._supplier_order_initial_use_global("woo_price"))
        self.assertTrue(ui._supplier_order_initial_use_global("global_margin"))
        self.assertFalse(ui._supplier_order_initial_use_global("pvp"))
        self.assertFalse(ui._supplier_order_initial_use_global("individual_margin"))

    def test_calculation_rows_do_not_invent_pvp_before_calculation(self) -> None:
        ui = app()

        def fail_pvp(*_args: object, **_kwargs: object) -> float:
            raise AssertionError("_calculation_rows no debe calcular P.V.P.")

        ui._supplier_order_pvp = fail_pvp
        item = general_item()
        source = dict(item.raw["source_row"])
        source.pop("pvp_unit", None)
        source.pop("status", None)
        item = OrderItem(item.code, item.name, item.quantity, item.m3, item.final_cost, item.status, raw={"source_row": source})

        rows = ui._calculation_rows((item,), provider="ekomat")

        self.assertEqual(rows[0][4], "Pendiente")

    def test_calculation_rows_show_persisted_pvp_after_calculation(self) -> None:
        ui = app()
        item = general_item()
        source = dict(item.raw["source_row"])
        source.update({"status": "Calculado", "pvp_unit": 120, "rentabilidad_percent": 25})
        item = OrderItem(item.code, item.name, item.quantity, item.m3, item.final_cost, "Calculado", raw={"source_row": source})

        rows = ui._calculation_rows((item,), provider="ekomat")

        self.assertTrue(str(rows[0][4]).startswith("120.00"))

    def test_editor_initial_margin_uses_woo_margin_without_global(self) -> None:
        ui = app()

        source = {"price_input_source": "woo_price", "rentabilidad_percent": 63.92}

        self.assertEqual(ui._supplier_order_editor_initial_margin(source, 0, "woo_price"), "63.92")
        self.assertFalse(ui._supplier_order_initial_use_global("woo_price"))

    def test_editor_initial_margin_for_woo_without_margin_is_blank(self) -> None:
        ui = app()

        self.assertEqual(ui._supplier_order_editor_initial_margin({"price_input_source": "woo_price"}, 30, "woo_price"), "")

    def test_editor_initial_margin_respects_global_and_individual_sources(self) -> None:
        ui = app()

        self.assertEqual(ui._supplier_order_editor_initial_margin({}, 30, "global_margin"), "30")
        self.assertTrue(ui._supplier_order_initial_use_global("global_margin"))
        self.assertEqual(
            ui._supplier_order_editor_initial_margin({"rentabilidad_individual_percent": 22.5}, 30, "individual_margin"),
            "22.5",
        )

    def test_editing_margin_on_woo_line_becomes_individual_margin(self) -> None:
        ui = app()
        updated = ui._supplier_order_update_line_rentabilidad(
            {
                "precio_ponderado_lote": 90,
                "precio_coste_final": 100,
                "pvp_unit": 120,
                "price_input_source": "woo_price",
                "pvp_source": "woocommerce",
                "pvp_source_label": "WooCommerce",
            },
            use_global=False,
            global_percent=40,
            individual_percent=25,
            input_source="margin",
            quantity=1,
        )

        self.assertEqual(updated["price_input_source"], "individual_margin")
        self.assertEqual(updated["pvp_source_label"], "Manual Margen")

    def test_editor_accepting_woo_line_keeps_dedicated_woo_route(self) -> None:
        source = inspect.getsource(FutonHubErpPrototype._open_order_item_missing_editor)

        self.assertIn('if input_source == "woo":', source)
        self.assertIn('updated_source["price_input_source"] = "woo_price"', source)
        self.assertLess(source.index('if input_source == "woo":'), source.index('else:\n                updated_source = self._supplier_order_update_line_rentabilidad'))

    def test_raw_order_line_inherits_woo_reference_from_resolved_inventory(self) -> None:
        ui = app()
        del ui.__dict__["_fill_supplier_prices_for_order_items"]
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(
                201002,
                name="Futon Woo vinculado",
                heca_reference="0201002",
                woo_id=123,
                woo_item_kind="product",
                woo_parent_id=0,
                woo_sku="SKU-201002",
                woo_link_status="linked",
            )
        ]
        ui._cloud_session = session

        item = OrderItem("0201002", "Pendiente", 1, "1", "Pendiente", "OK", raw={"source_row": {"m3_total": 1, "m3_und": 1, "cuenta_reparto_descarga": True}})
        enriched = ui._fill_supplier_prices_for_order_items("ekomat", (item,))
        source = enriched[0].raw["source_row"]

        self.assertEqual(source["inventory_matched_item_id"], 201002)
        self.assertEqual(source["woo_id"], 123)
        self.assertEqual(source["inventory_woo_id"], 123)
        self.assertEqual(source["woo_item_kind"], "product")
        self.assertEqual(ui._supplier_order_woo_reference(source), {"woo_id": 123, "woo_item_kind": "product", "woo_parent_id": 0})

    def test_raw_order_line_calculates_pvp_from_live_woo_after_inventory_resolution(self) -> None:
        ui = app()
        del ui.__dict__["_fill_supplier_prices_for_order_items"]
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(
                201002,
                name="Futon Woo vinculado",
                heca_reference="0201002",
                woo_id=123,
                woo_item_kind="product",
                primary_supplier_price=93.65,
                rotation_c=1,
                store_stock=1,
                weighted_average_cost=80,
            )
        ]
        ui._cloud_session = session
        item = OrderItem("0201002", "Pendiente", 1, "1", "Pendiente", "OK", raw={"source_row": {"m3_total": 1, "m3_und": 1, "cuenta_reparto_descarga": True}})
        enriched = ui._fill_supplier_prices_for_order_items("ekomat", (item,))

        with patch.object(ui, "_supplier_order_fetch_woo_price", return_value={"price": 120, "error": "", "checked_at": "now", "woo_price_item_id": 123, "woo_price_item_kind": "product", "woo_price_parent_id": ""}):
            calculated, _raw, _summary = ui._calculate_supplier_order_in_memory("ekomat", {"Margen de Venta %": "40", "Coste transporte + IVA": "1"}, enriched, [])

        source = calculated[0].raw["source_row"]
        self.assertEqual(source["status"], "Calculado")
        self.assertEqual(source["pvp_unit"], 120.0)
        self.assertEqual(source["rentabilidad_percent"], 25.0)
        self.assertEqual(source["price_input_source"], "woo_price")
        self.assertEqual(source["pvp_source_label"], "WooCommerce")
        self.assertEqual(source["woo_price_error"], "")

    def test_inventory_without_woo_reference_remains_pending_for_woo_price(self) -> None:
        ui = app()
        item = self.woo_item(woo_id=0, extra_source={"woo_item_kind": "", "inventory_woo_id": "", "supplier_price_woo_id": ""})

        calculated, _raw, _summary = ui._calculate_supplier_order_in_memory("ekomat", {"Margen de Venta %": "40", "Coste transporte + IVA": "1"}, (item,), [])
        source = calculated[0].raw["source_row"]

        self.assertEqual(source["status"], "Warning")
        self.assertIsNone(source["pvp_unit"])
        self.assertIsNone(source["rentabilidad_percent"])
        self.assertIn("No se pudo resolver Woo ID", source["woo_price_error"])

    def test_inventory_variation_reference_uses_parent_endpoint(self) -> None:
        ui = app()

        class Response:
            def json(self) -> dict[str, str]:
                return {"price": "120"}

        class Client:
            endpoints: list[str] = []

            def __init__(self, *_args: object) -> None:
                pass

            def get(self, endpoint: str) -> Response:
                self.endpoints.append(endpoint)
                return Response()

            def put(self, *_args: object) -> None:
                raise AssertionError("Woo no debe escribirse")

            def post(self, *_args: object) -> None:
                raise AssertionError("Woo no debe escribirse")

        with patch.object(prototype_module, "load_settings", return_value=SimpleNamespace(woocommerce_url="https://woo.test", consumer_key="ck", consumer_secret="cs")):
            with patch.object(prototype_module, "WooCommerceClient", Client):
                result = ui._supplier_order_fetch_woo_price({"inventory_woo_id": 456, "inventory_woo_item_kind": "variation", "inventory_woo_parent_id": 123})

        self.assertEqual(result["price"], 120.0)
        self.assertEqual(Client.endpoints, ["products/123/variations/456"])

    def test_inventory_variation_without_parent_is_pending_without_traceback(self) -> None:
        ui = app()

        result = ui._supplier_order_fetch_woo_price({"inventory_woo_id": 456, "inventory_woo_item_kind": "variation"})

        self.assertIsNone(result["price"])
        self.assertEqual(result["error"], "Variacion sin parent Woo ID.")

    def test_inventory_supabase_price_is_not_used_as_automatic_pvp(self) -> None:
        ui = app()
        item = self.woo_item(extra_source={"inventory_woo_price": 999})
        with patch.object(ui, "_supplier_order_fetch_woo_price", return_value={"price": 120, "error": "", "checked_at": "now", "woo_price_item_id": 123, "woo_price_item_kind": "product", "woo_price_parent_id": ""}):
            calculated, _raw, _summary = ui._calculate_supplier_order_in_memory("ekomat", {"Margen de Venta %": "40", "Coste transporte + IVA": "1"}, (item,), [])

        self.assertEqual(calculated[0].raw["source_row"]["pvp_unit"], 120.0)
        self.assertNotEqual(calculated[0].raw["source_row"]["pvp_unit"], 999)

    def test_partial_raw_inventory_resolution_failure_does_not_break_order(self) -> None:
        ui = app()
        item_ok = self.woo_item(woo_id=123)
        item_missing = self.woo_item(code="201015", woo_id=0, extra_source={"woo_item_kind": ""})

        def fetch(source: dict[str, object]) -> dict[str, object]:
            reference = ui._supplier_order_woo_reference(source)
            if reference["woo_id"] == 123:
                return {"price": 120, "error": "", "checked_at": "now", "woo_price_item_id": 123, "woo_price_item_kind": "product", "woo_price_parent_id": ""}
            return {"price": None, "error": "No se pudo resolver Woo ID para la linea.", "checked_at": "now", "woo_price_item_id": "", "woo_price_item_kind": "", "woo_price_parent_id": ""}

        with patch.object(ui, "_supplier_order_fetch_woo_price", side_effect=fetch):
            calculated, _raw, _summary = ui._calculate_supplier_order_in_memory("ekomat", {"Margen de Venta %": "40", "Coste transporte + IVA": "1"}, (item_ok, item_missing), [])

        self.assertEqual(calculated[0].status, "Calculado")
        self.assertEqual(calculated[1].status, "Warning")

    def test_global_margin_validation_accepts_zero_and_blocks_100_or_negative(self) -> None:
        ui = app()

        self.assertEqual(ui._supplier_order_pvp(100, 0), 100.0)
        with self.assertRaises(ValueError):
            ui._supplier_order_pvp(100, 100)
        with self.assertRaises(ValueError):
            ui._supplier_order_pvp(100, 101)
        with self.assertRaises(ValueError):
            ui._supplier_order_pvp(100, -1)

    def test_individual_negative_margin_reconstructs_pvp(self) -> None:
        ui = app()

        self.assertEqual(ui._supplier_order_pvp(100, -25, allow_negative=True), 80.0)

    def test_pvp_zero_and_negative_are_blocked(self) -> None:
        ui = app()

        with self.assertRaises(ValueError):
            ui._supplier_order_margin_from_pvp(100, 0)
        with self.assertRaises(ValueError):
            ui._supplier_order_margin_from_pvp(100, -1)

    def test_pvp_below_cost_calculates_negative_margin_without_traceback(self) -> None:
        ui = app()

        self.assertEqual(ui._supplier_order_margin_from_pvp(100, 80), -25.0)

    def test_empty_pvp_field_is_validated_without_traceback(self) -> None:
        ui = app()

        with self.assertRaisesRegex(ValueError, "vac"):
            ui._parse_supplier_order_pvp_unit("")

    def test_inventory_numeric_code_normalization_is_shared_and_safe(self) -> None:
        self.assertEqual(normalize_inventory_numeric_code("0201001"), "201001")
        self.assertEqual(normalize_inventory_numeric_code("000201001"), "201001")
        self.assertEqual(normalize_inventory_numeric_code("0000"), "0")
        self.assertEqual(normalize_inventory_numeric_code("AB0201001"), "AB0201001")

    def test_shared_pack_detector_uses_persisted_pack_signals(self) -> None:
        self.assertTrue(is_inventory_pack_row({"hub_item_code": "WOO-PACK-1"}))
        self.assertTrue(is_inventory_pack_row({"item_id": "WOO-PACK-2"}))
        self.assertTrue(is_inventory_pack_row({"item_record_type": "woo_pack"}))
        self.assertTrue(is_inventory_pack_row({"is_pack": True}))
        self.assertTrue(is_inventory_pack_row({"woo_sku": "0724001|0201001"}))
        self.assertFalse(is_inventory_pack_row({"item_id": 724001, "woo_sku": "0724001", "item_record_type": "simple"}))

    def test_supplier_order_eligibility_accepts_base_and_rejects_derived_rows(self) -> None:
        self.assertTrue(
            is_supplier_order_eligible_inventory_row(
                inventory_row(724004, heca_reference="0724004", item_record_type="simple")
            )
        )
        self.assertFalse(
            is_supplier_order_eligible_inventory_row(
                inventory_row(
                    930000012860,
                    hub_item_code="WOO-ITEM-12860",
                    woo_sku="0724003-724004",
                    item_record_type="woo_item",
                    woo_item_kind="variation",
                    woo_parent_id=12856,
                )
            )
        )
        self.assertFalse(is_supplier_order_eligible_inventory_row(inventory_row(990001, base_item_code="724004")))
        self.assertFalse(
            is_supplier_order_eligible_inventory_row(
                inventory_row(990002, source_row={"relation_type": "component", "component_item_code": "0724004"})
            )
        )
        self.assertFalse(
            is_supplier_order_eligible_inventory_row(
                inventory_row(990003, source_row={"hub_search_record_type": "search_alias"})
            )
        )

    def test_normal_item_wins_when_equivalent_pack_is_present(self) -> None:
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(724001, name="Articulo normal", heca_reference="724001", rotation_c=4),
            inventory_row(
                990001,
                name="Pack Woo",
                heca_reference="0724001",
                hub_item_code="WOO-PACK-990001",
                item_record_type="woo_pack",
                is_pack=True,
            ),
        ]

        resolved = resolve_supplier_order_inventory_items(session, ["0724001"], "ekomat")

        self.assertEqual(resolved["0724001"]["item_id"], 724001)
        self.assertEqual(resolved["0724001"]["item"]["rotation_c"], 4)
        self.assertEqual(resolved["0724001"]["price"], 93.65)
        self.assertEqual(resolved["0724001"]["matched_by"], "canonical:item_id")

    def test_multiple_equivalent_packs_do_not_create_order_ambiguity(self) -> None:
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(724001, name="Articulo normal", heca_reference="724001"),
            inventory_row(
                990001,
                name="Pack prefijo",
                heca_reference="0724001",
                hub_item_code="WOO-PACK-990001",
            ),
            inventory_row(
                990002,
                name="Pack SKU compuesto",
                heca_reference="000724001",
                woo_sku="0724001|0201001",
            ),
        ]

        resolved = resolve_supplier_order_inventory_items(session, ["0724001"], "ekomat")

        self.assertEqual(resolved["0724001"]["item_id"], 724001)
        self.assertEqual(session.client.execute_counts["inventory_items"], 1)

    def test_pack_prefix_and_compound_sku_are_never_order_candidates(self) -> None:
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row("WOO-PACK-1", hub_item_code="WOO-PACK-1"),
            inventory_row(990002, heca_reference="0724001", woo_sku="0724001|0201001"),
        ]

        resolved = resolve_supplier_order_inventory_items(
            session,
            ["WOO-PACK-1", "0724001"],
            "ekomat",
        )

        self.assertEqual(resolved, {})

    def test_alphanumeric_normal_order_code_still_resolves_exactly(self) -> None:
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(810001, heca_reference="AB-0724001", hub_item_code="AB-0724001"),
        ]

        resolved = resolve_supplier_order_inventory_items(session, ["AB-0724001"], "ekomat")

        self.assertEqual(resolved["AB-0724001"]["item_id"], 810001)
        self.assertEqual(resolved["AB-0724001"]["matched_by"], "exact:hub_item_code")

    def test_order_code_resolver_matches_both_leading_zero_directions_in_one_batch(self) -> None:
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(201001, heca_reference="201001"),
            inventory_row(999001, heca_reference="0201002", hub_item_code="000201002"),
        ]

        resolved = resolve_supplier_order_inventory_items(
            session,
            ["0201001", "201002", "AB0201001"],
            "ekomat",
        )

        self.assertEqual(resolved["0201001"]["item_id"], 201001)
        self.assertEqual(resolved["0201001"]["matched_by"], "canonical:item_id")
        self.assertEqual(resolved["201002"]["item_id"], 999001)
        self.assertIn(resolved["201002"]["matched_by"], {"canonical:heca_reference", "canonical:hub_item_code"})
        self.assertNotIn("AB0201001", resolved)
        self.assertEqual(session.client.execute_counts["inventory_items"], 1)

    def test_exact_order_code_match_has_priority_over_canonical_match(self) -> None:
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(201001, name="Canonico", heca_reference="201001"),
            inventory_row(880001, name="Exacto", heca_reference="0201001"),
        ]

        resolved = resolve_supplier_order_inventory_items(session, ["0201001"], "ekomat")

        self.assertEqual(resolved["0201001"]["item_id"], 201001)
        self.assertEqual(resolved["0201001"]["matched_by"], "canonical:item_id")

    def test_canonical_order_code_ambiguity_is_blocked(self) -> None:
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(880001, name="Uno", hub_item_code="0777777"),
            inventory_row(880002, name="Dos", hub_item_code="777777"),
        ]

        with self.assertRaisesRegex(SupplierOrderCodeAmbiguityError, "canonical:hub_item_code"):
            resolve_supplier_order_inventory_items(session, ["000777777"], "ekomat")

    def test_item_id_exact_beats_lower_priority_hub_item_code(self) -> None:
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(724004, name="Base por item_id", hub_item_code="BASE-724004"),
            inventory_row(880004, name="Alias menor prioridad", hub_item_code="724004"),
        ]

        resolved = resolve_supplier_order_inventory_items(session, ["724004"], "ekomat")

        self.assertEqual(resolved["724004"]["item_id"], 724004)
        self.assertEqual(resolved["724004"]["matched_by"], "exact:item_id")

    def test_item_id_canonical_beats_lower_priority_alias_canonical(self) -> None:
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(724004, name="Base por item_id"),
            inventory_row(880004, name="Alias menor prioridad", hub_item_code="0724004"),
        ]

        resolved = resolve_supplier_order_inventory_items(session, ["000724004"], "ekomat")

        self.assertEqual(resolved["000724004"]["item_id"], 724004)
        self.assertEqual(resolved["000724004"]["matched_by"], "canonical:item_id")

    def test_real_0724004_base_wins_over_woo_synthetic_variation(self) -> None:
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(
                724004,
                name="Futón de Algodón 150 x 200 x 14 cm.",
                heca_reference="0724004",
                primary_supplier_price=80.64,
                rotation_c=0.051780822,
            ),
            inventory_row(
                930000012860,
                name="14,5 cm, 140x200x14,5 cm, Crudo",
                hub_item_code="WOO-ITEM-12860",
                woo_sku="0724003-724004",
                item_record_type="woo_item",
                woo_item_kind="variation",
                woo_parent_id=12856,
                primary_supplier_price=324,
            ),
        ]

        resolved = resolve_supplier_order_inventory_items(session, ["0724004"], "ekomat")

        self.assertEqual(resolved["0724004"]["item_id"], 724004)
        self.assertEqual(resolved["0724004"]["matched_by"], "canonical:item_id")
        self.assertEqual(resolved["0724004"]["price"], 80.64)
        self.assertEqual(resolved["0724004"]["item"]["rotation_c"], 0.051780822)

    def test_pascal_effective_price_prefers_valid_pascal_price(self) -> None:
        result = resolve_supplier_order_effective_price(
            {"pascal_price": 120, "primary_supplier_price": 114.26},
            "Pascal",
        )

        self.assertEqual(result["price"], 120)
        self.assertEqual(result["source"], "pascal")
        self.assertFalse(result["fallback_used"])
        self.assertEqual(result["warning"], "")

    def test_pascal_missing_or_invalid_uses_primary_fallback(self) -> None:
        for missing in (None, "", "NO ESTA", "no numerico"):
            with self.subTest(pascal_price=missing):
                result = resolve_supplier_order_effective_price(
                    {"pascal_price": missing, "primary_supplier_price": 114.26},
                    "Pascal",
                )
                self.assertEqual(result["price"], 114.26)
                self.assertEqual(result["source"], "primary_fallback_for_pascal")
                self.assertEqual(result["requested_provider"], "pascal")
                self.assertTrue(result["fallback_used"])
                self.assertIn("114,26 €", result["warning"])

    def test_pascal_without_any_valid_price_remains_manual_required(self) -> None:
        for primary in (None, "", "NO ESTA", "incorrecto"):
            with self.subTest(primary_supplier_price=primary):
                result = resolve_supplier_order_effective_price(
                    {"pascal_price": None, "primary_supplier_price": primary},
                    "Pascal",
                )
                self.assertEqual(result["price"], 0)
                self.assertEqual(result["source"], "pascal")
                self.assertFalse(result["fallback_used"])

    def test_pascal_line_without_any_price_keeps_manual_price_blocker(self) -> None:
        ui = app()
        del ui.__dict__["_fill_supplier_prices_for_order_items"]
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(
                780004,
                name="Futon Pascal",
                heca_reference="0780004",
                primary_supplier_price="NO ESTA",
                pascal_price=None,
                rotation_c=0.051780822,
            )
        ]
        ui._cloud_session = session
        item = OrderItem(
            "0780004",
            "Pendiente",
            1,
            "Pendiente",
            "Pendiente",
            "OK",
            raw={"source_row": {"cuenta_reparto_descarga": True}},
        )

        enriched = ui._fill_supplier_prices_for_order_items("Pascal", (item,))

        self.assertIn("falta precio proveedor", ui._order_item_missing_reasons(enriched[0]))
        self.assertNotIn("precio_proveedor", enriched[0].raw["source_row"])

    def test_non_pascal_keeps_primary_price_rule(self) -> None:
        result = resolve_supplier_order_effective_price(
            {"pascal_price": 120, "primary_supplier_price": 114.26},
            "Ekomat",
        )

        self.assertEqual(result["price"], 114.26)
        self.assertEqual(result["source"], "primary")
        self.assertFalse(result["fallback_used"])

    def test_pascal_fallback_enriches_line_without_inventory_write_and_recalculates(self) -> None:
        ui = app()
        del ui.__dict__["_fill_supplier_prices_for_order_items"]
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(
                780004,
                name="Futon Pascal",
                heca_reference="0780004",
                primary_supplier_price=114.26,
                pascal_price=None,
                rotation_c=0.051780822,
            )
        ]
        ui._cloud_session = session
        item = OrderItem(
            "0780004",
            "Pendiente",
            1,
            "Pendiente",
            "Pendiente",
            "OK",
            raw={"source_row": {"cuenta_reparto_descarga": True}},
        )
        before_inventory = json.loads(json.dumps(session.client.tables["inventory_items"]))

        enriched = ui._fill_supplier_prices_for_order_items("Pascal", (item,))
        source = enriched[0].raw["source_row"]
        self.assertEqual(enriched[0].code, "0780004")
        self.assertEqual(enriched[0].name, "Futon Pascal")
        self.assertEqual(source["precio_proveedor"], 114.26)
        self.assertEqual(source["inventory_rotation_c"], 0.051780822)
        self.assertEqual(source["inventory_packages"], 1)
        self.assertEqual(source["inventory_m3"], 1)
        self.assertEqual(source["supplier_price_source"], "primary_fallback_for_pascal")
        self.assertEqual(source["supplier_price_requested_provider"], "pascal")
        self.assertTrue(source["supplier_price_fallback_used"])
        self.assertIn("Precio Pascal no disponible", source["supplier_price_warning"])
        self.assertNotIn("falta precio proveedor", ui._order_item_missing_reasons(enriched[0]))
        self.assertEqual(session.client.tables["inventory_items"], before_inventory)
        enriched[0].raw["source_row"]["price_input_source"] = "global_margin"
        enriched[0].raw["source_row"]["use_global_rentability"] = True

        first, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "Pascal",
            {"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
            enriched,
            [],
        )
        second, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "Pascal",
            {"Rentabilidad %": "30", "Coste transporte + IVA": "11"},
            first,
            [],
        )
        first_source = first[0].raw["source_row"]
        second_source = second[0].raw["source_row"]
        self.assertGreater(second_source["unit_cost"], first_source["unit_cost"])
        self.assertEqual(second_source["precio_proveedor"], 114.26)
        self.assertEqual(second_source["supplier_price_source"], "primary_fallback_for_pascal")
        self.assertEqual(
            second_source["pvp_unit"],
            ui._supplier_order_pvp(second_source["unit_cost"], 30),
        )
        self.assertEqual(session.client.tables["inventory_items"], before_inventory)

    def test_pascal_fallback_trace_survives_reload_and_exports_origin(self) -> None:
        ui = app()
        source = {
            **general_item().raw["source_row"],
            "precio_proveedor": 114.26,
            "precio_excel": 114.26,
            "unit_cost": 120,
            "precio_coste_final": 120,
            "line_cost": 120,
            "final_cost": 120,
            "rentabilidad_percent": 30,
            "pvp_unit": 171.43,
            "supplier_price_source": "primary_fallback_for_pascal",
            "supplier_price_source_label": "Principal fallback Pascal",
            "supplier_price_requested_provider": "pascal",
            "supplier_price_fallback_used": True,
            "supplier_price_item_id": 780004,
            "supplier_price_warning": "Precio Pascal no disponible. Se usa precio principal: 114.26 EUR",
        }
        cloud_row = json.loads(
            json.dumps(
                {
                    "id": 1,
                    "order_id": "ORDER-PASCAL",
                    "item_code": "0780004",
                    "item_name": "Futon Pascal",
                    "quantity_ordered": 1,
                    "quantity_received": 0,
                    "unit_cost": 120,
                    "line_cost": 120,
                    "updated_at": "2026-06-19T00:00:00+00:00",
                    "source_row": source,
                }
            )
        )

        reloaded = ui._supplier_order_item_from_cloud_row(cloud_row)
        reloaded_source = reloaded.raw["source_row"]
        self.assertEqual(reloaded.code, "0780004")
        self.assertEqual(reloaded_source["precio_proveedor"], 114.26)
        self.assertTrue(reloaded_source["supplier_price_fallback_used"])
        self.assertEqual(ui._supplier_order_price_origin_label(reloaded_source), "Principal fallback Pascal")

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "pascal-fallback.xlsx"
            with patch.object(prototype_module.messagebox, "showinfo"):
                ui._export_supplier_order_audit_excel(
                    None,
                    provider="Pascal",
                    values={"Nombre del pedido": "PED-PASCAL"},
                    items=(reloaded,),
                    path=str(path),
                )
            ws = load_workbook(path)["Líneas calculadas"]
            headers = [cell.value for cell in ws[1]]
            origin_column = headers.index("Origen precio proveedor") + 1
            self.assertEqual(ws.cell(row=2, column=origin_column).value, "Principal fallback Pascal")

    def test_equivalent_order_code_fills_name_rotation_price_and_calculates_without_changing_code(self) -> None:
        ui = app()
        del ui.__dict__["_fill_supplier_prices_for_order_items"]
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(
                201001,
                name="Futon Canonico",
                heca_reference="201001",
                primary_supplier_price=93.65,
                rotation_c=3,
            ),
            inventory_row(
                990001,
                name="Pack equivalente excluido",
                heca_reference="0201001",
                hub_item_code="WOO-PACK-990001",
                woo_sku="0201001|0724001",
                item_record_type="woo_pack",
                is_pack=True,
            ),
            inventory_row(
                990002,
                name="Componente auxiliar excluido",
                hub_item_code="0201001",
                item_record_type="component",
                source_row={"relation_type": "component", "component_item_code": "0201001"},
            ),
            inventory_row(
                990003,
                name="Alias de búsqueda excluido",
                heca_reference="0201001",
                base_item_code="201001",
                source_row={"hub_search_record_type": "search_alias"},
            ),
            inventory_row(
                930000012999,
                name="Variación Woo sintética excluida",
                hub_item_code="WOO-ITEM-12999",
                woo_sku="0201001",
                item_record_type="woo_item",
                woo_item_kind="variation",
                woo_parent_id=12000,
            ),
        ]
        ui._cloud_session = session
        source = {
            "m3_total": 1,
            "m3_und": 1,
            "packages": 1,
            "cuenta_reparto_descarga": True,
        }
        original = OrderItem("0201001", "Pendiente", 1, "1", "Pendiente", "OK", raw={"source_row": source})

        enriched = ui._fill_supplier_prices_for_order_items("ekomat", (original,))
        enriched_source = enriched[0].raw["source_row"]
        enriched_source["price_input_source"] = "global_margin"
        enriched_source["use_global_rentability"] = True
        self.assertEqual(enriched[0].code, "0201001")
        self.assertEqual(enriched[0].name, "Futon Canonico")
        self.assertEqual(enriched_source["inventory_rotation_c"], 3)
        self.assertEqual(enriched_source["precio_proveedor"], 93.65)
        self.assertEqual(enriched_source["supplier_price_item_id"], 201001)
        self.assertEqual(enriched_source["inventory_matched_item_id"], 201001)
        self.assertEqual(ui._order_item_missing_reasons(enriched[0]), [])

        calculated, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
            enriched,
            [],
        )
        calculated_source = calculated[0].raw["source_row"]
        self.assertEqual(calculated[0].code, "0201001")
        self.assertGreater(calculated_source["unit_cost"], 0)
        self.assertGreater(calculated_source["precio_ponderado_lote"], 0)
        self.assertEqual(calculated_source["rentabilidad_percent"], 30)
        self.assertEqual(
            calculated_source["pvp_unit"],
            ui._supplier_order_pvp(calculated_source["unit_cost"], 30),
        )

    def test_equivalent_codes_do_not_duplicate_inventory_or_change_cost_results(self) -> None:
        ui = app()
        del ui.__dict__["_fill_supplier_prices_for_order_items"]
        session = MemorySession()
        session.client.tables["inventory_items"] = [inventory_row(201001, heca_reference="0201001")]
        ui._cloud_session = session

        def calculate(code: str) -> dict[str, object]:
            item = general_item()
            item = OrderItem(code, item.name, item.quantity, item.m3, item.final_cost, item.status, raw=item.raw)
            calculated, _raw, _summary = ui._calculate_supplier_order_in_memory(
                "ekomat",
                {"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
                (item,),
                [],
            )
            return calculated[0].raw["source_row"]

        with_zero = calculate("0201001")
        without_zero = calculate("201001")

        for field in ("unit_cost", "line_cost", "precio_ponderado_lote", "rentabilidad_percent", "pvp_unit"):
            self.assertEqual(with_zero[field], without_zero[field])
        self.assertEqual(len(session.client.tables["inventory_items"]), 1)

    def test_global_rentabilidad_is_used_when_line_has_no_individual_value(self) -> None:
        ui = app()
        items, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
            (general_item(),),
            [],
        )

        source = items[0].raw["source_row"]
        self.assertEqual(source["rentabilidad_percent"], 30.0)
        self.assertEqual(source["rentabilidad_source"], "global")
        self.assertTrue(source["use_global_rentability"])
        self.assertEqual(source["pvp_unit"], 142.86)

    def test_individual_rentabilidad_overrides_global_without_being_overwritten(self) -> None:
        ui = app()
        item = general_item()
        source = dict(item.raw["source_row"])
        source.update({"rentabilidad_individual_percent": 20, "use_global_rentability": False})
        item = OrderItem(item.code, item.name, item.quantity, item.m3, item.final_cost, item.status, raw={"source_row": source})

        items, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
            (item,),
            [],
        )

        result = items[0].raw["source_row"]
        self.assertEqual(result["rentabilidad_individual_percent"], 20)
        self.assertEqual(result["rentabilidad_percent"], 20.0)
        self.assertEqual(result["rentabilidad_source"], "individual")
        self.assertFalse(result["use_global_rentability"])
        self.assertEqual(result["pvp_unit"], 125.0)

    def test_individual_rentabilidad_recalculates_from_new_real_cost(self) -> None:
        ui = app()
        item = general_item()
        source = dict(item.raw["source_row"])
        source.update({"rentabilidad_individual_percent": 30, "use_global_rentability": False})
        item = OrderItem(item.code, item.name, item.quantity, item.m3, item.final_cost, item.status, raw={"source_row": source})

        first, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "10", "Coste transporte + IVA": "1"},
            (item,),
            [],
        )
        second, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "60", "Coste transporte + IVA": "11"},
            first,
            [],
        )

        first_source = first[0].raw["source_row"]
        result = second[0].raw["source_row"]
        self.assertGreater(result["unit_cost"], first_source["unit_cost"])
        self.assertEqual(result["rentabilidad_individual_percent"], 30)
        self.assertEqual(result["rentabilidad_percent"], 30.0)
        self.assertEqual(result["pvp_unit"], ui._supplier_order_pvp(result["unit_cost"], 30))

    def test_editing_only_rentabilidad_keeps_real_cost_fields_unchanged(self) -> None:
        ui = app()
        original = {
            "unit_cost": 100.0,
            "precio_coste_final": 100.0,
            "line_cost": 200.0,
            "final_cost": 200.0,
            "precio_ponderado_lote": 95.0,
            "rentabilidad_percent": 30.0,
            "pvp_unit": 142.86,
            "pvp_line": 285.72,
        }

        updated = ui._supplier_order_update_line_rentabilidad(
            original,
            use_global=False,
            global_percent=40,
            individual_percent=20,
            quantity=2,
        )

        for key in ("unit_cost", "precio_coste_final", "line_cost", "final_cost", "precio_ponderado_lote"):
            self.assertEqual(updated[key], original[key])
        self.assertEqual(updated["rentabilidad_individual_percent"], 20.0)
        self.assertEqual(updated["pvp_unit"], 118.75)
        self.assertEqual(updated["pvp_line"], 237.5)

    def test_editing_pvp_calculates_margin_and_disables_global_without_cost_changes(self) -> None:
        ui = app()
        original = {
            "unit_cost": 52.68,
            "precio_coste_final": 52.68,
            "line_cost": 105.36,
            "final_cost": 105.36,
            "precio_ponderado_lote": 50.0,
            "rentabilidad_percent": 30.0,
            "use_global_rentability": True,
        }

        updated = ui._supplier_order_update_line_rentabilidad(
            original,
            use_global=True,
            global_percent=30,
            pvp_unit=87.80,
            input_source="pvp",
            quantity=2,
        )

        for key in ("unit_cost", "precio_coste_final", "line_cost", "final_cost", "precio_ponderado_lote"):
            self.assertEqual(updated[key], original[key])
        self.assertFalse(updated["use_global_rentability"])
        self.assertEqual(updated["rentabilidad_individual_percent"], 43.05)
        self.assertEqual(updated["rentabilidad_percent"], 43.05)
        self.assertEqual(updated["pvp_unit"], 87.8)
        self.assertEqual(updated["pvp_line"], 175.6)
        self.assertEqual(updated["price_input_source"], "pvp")

    def test_sales_margin_uses_weighted_cost_instead_of_final_cost(self) -> None:
        ui = app()
        source = {
            "unit_cost": 100.0,
            "precio_coste_final": 100.0,
            "line_cost": 100.0,
            "precio_ponderado_lote": 90.0,
        }

        updated = ui._supplier_order_update_line_rentabilidad(
            source,
            use_global=False,
            global_percent=30,
            individual_percent=40,
            input_source="margin",
            quantity=1,
        )

        self.assertEqual(updated["pvp_unit"], 150.0)
        self.assertNotEqual(updated["pvp_unit"], 166.67)

    def test_pvp_inverse_margin_uses_weighted_cost_instead_of_final_cost(self) -> None:
        ui = app()
        source = {
            "unit_cost": 100.0,
            "precio_coste_final": 100.0,
            "line_cost": 100.0,
            "precio_ponderado_lote": 90.0,
        }

        updated = ui._supplier_order_update_line_rentabilidad(
            source,
            use_global=True,
            global_percent=30,
            pvp_unit=120,
            input_source="pvp",
            quantity=1,
        )

        self.assertEqual(updated["rentabilidad_individual_percent"], 25.0)
        self.assertNotEqual(updated["rentabilidad_individual_percent"], 16.67)
        self.assertEqual(updated["price_input_source"], "pvp")

    def test_real_weighted_cost_feeds_global_margin_pvp_without_changing_cost_total(self) -> None:
        ui = app()
        item = general_item()
        source = dict(item.raw["source_row"])
        source.update({"inventory_stock_total": 10, "inventory_weighted_average_cost": 80})
        item = OrderItem(item.code, item.name, item.quantity, item.m3, item.final_cost, item.status, raw={"source_row": source})

        calculated, _raw, summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Margen de Venta %": "30", "Coste transporte + IVA": "1"},
            (item,),
            [],
        )

        result = calculated[0].raw["source_row"]
        self.assertEqual(result["unit_cost"], 100.0)
        self.assertEqual(result["line_cost"], 100.0)
        self.assertEqual(summary["total_cost"], 100.0)
        self.assertEqual(result["precio_ponderado_lote"], 81.82)
        self.assertEqual(result["pvp_unit"], 116.89)
        self.assertEqual(result["price_input_source"], "global_margin")

    def test_global_margin_source_uses_weighted_cost(self) -> None:
        ui = app()
        source = {
            "unit_cost": 100.0,
            "precio_coste_final": 100.0,
            "line_cost": 100.0,
            "precio_ponderado_lote": 90.0,
            "rentabilidad_individual_percent": 20,
            "use_global_rentability": False,
        }

        updated = ui._supplier_order_update_line_rentabilidad(
            source,
            use_global=True,
            global_percent=40,
            quantity=1,
        )

        self.assertEqual(updated["price_input_source"], "global_margin")
        self.assertEqual(updated["rentabilidad_percent"], 40.0)
        self.assertEqual(updated["pvp_unit"], 150.0)

    def test_without_previous_stock_weighted_cost_equals_final_cost(self) -> None:
        ui = app()

        weighted = ui._supplier_order_weighted_unit_cost(
            final_unit=100,
            quantity=1,
            current_stock=0,
            current_weighted_cost=80,
        )

        self.assertEqual(weighted, 100.0)
        self.assertEqual(ui._supplier_order_pvp(weighted, 40), 166.67)

    def test_negative_individual_margin_from_pvp_survives_order_recalculation(self) -> None:
        ui = app()
        item = general_item()
        source = dict(item.raw["source_row"])
        source.update(
            {
                "unit_cost": 100.0,
                "precio_coste_final": 100.0,
                "line_cost": 100.0,
                "precio_ponderado_lote": 100.0,
            }
        )
        source = ui._supplier_order_update_line_rentabilidad(
            source,
            use_global=True,
            global_percent=30,
            pvp_unit=80,
            input_source="pvp",
            quantity=1,
        )
        item = OrderItem(item.code, item.name, item.quantity, item.m3, item.final_cost, item.status, raw={"source_row": source})

        recalculated, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Margen de Venta %": "30", "Coste transporte + IVA": "1"},
            (item,),
            [],
        )

        result = recalculated[0].raw["source_row"]
        self.assertFalse(result["use_global_rentability"])
        self.assertEqual(result["rentabilidad_source"], "individual")
        self.assertEqual(result["price_input_source"], "pvp")
        self.assertEqual(result["rentabilidad_individual_percent"], -25.0)
        self.assertEqual(result["rentabilidad_percent"], -25.0)
        self.assertEqual(result["unit_cost"], 100.0)
        self.assertEqual(result["pvp_unit"], 80.0)
        self.assertEqual(result["precio_ponderado_lote"], 100.0)

    def test_pvp_price_input_source_keeps_pvp_and_recalculates_margin_from_new_cost(self) -> None:
        ui = app()
        item = general_item()
        source = dict(item.raw["source_row"])
        source.update(
            {
                "unit_cost": 100.0,
                "precio_coste_final": 100.0,
                "line_cost": 100.0,
                "precio_ponderado_lote": 100.0,
            }
        )
        source = ui._supplier_order_update_line_rentabilidad(
            source,
            use_global=True,
            global_percent=30,
            pvp_unit=80,
            input_source="pvp",
            quantity=1,
        )
        item = OrderItem(item.code, item.name, item.quantity, item.m3, item.final_cost, item.status, raw={"source_row": source})

        recalculated, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Margen de Venta %": "30", "Coste transporte + IVA": "11"},
            (item,),
            [],
        )

        result = recalculated[0].raw["source_row"]
        self.assertEqual(result["unit_cost"], 110.0)
        self.assertEqual(result["precio_ponderado_lote"], 110.0)
        self.assertEqual(result["price_input_source"], "pvp")
        self.assertEqual(result["pvp_unit"], 80.0)
        self.assertEqual(result["rentabilidad_source"], "individual")
        self.assertEqual(result["rentabilidad_individual_percent"], -37.5)
        self.assertEqual(result["rentabilidad_percent"], -37.5)

    def test_negative_individual_margin_can_be_edited_directly_and_rebuilds_pvp(self) -> None:
        ui = app()
        source = {
            "unit_cost": 100.0,
            "precio_coste_final": 100.0,
            "line_cost": 100.0,
            "precio_ponderado_lote": 100.0,
        }

        updated = ui._supplier_order_update_line_rentabilidad(
            source,
            use_global=False,
            global_percent=30,
            individual_percent=-25,
            input_source="margin",
            quantity=1,
        )

        self.assertFalse(updated["use_global_rentability"])
        self.assertEqual(updated["rentabilidad_source"], "individual")
        self.assertEqual(updated["price_input_source"], "individual_margin")
        self.assertEqual(updated["rentabilidad_individual_percent"], -25.0)
        self.assertEqual(updated["pvp_unit"], 80.0)

    def test_individual_margin_source_keeps_margin_and_rebuilds_pvp_from_new_cost(self) -> None:
        ui = app()
        item = general_item()
        source = dict(item.raw["source_row"])
        source.update({"rentabilidad_individual_percent": -25, "use_global_rentability": False, "price_input_source": "individual_margin"})
        item = OrderItem(item.code, item.name, item.quantity, item.m3, item.final_cost, item.status, raw={"source_row": source})

        recalculated, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Margen de Venta %": "30", "Coste transporte + IVA": "11"},
            (item,),
            [],
        )

        result = recalculated[0].raw["source_row"]
        self.assertEqual(result["unit_cost"], 110.0)
        self.assertEqual(result["price_input_source"], "individual_margin")
        self.assertEqual(result["rentabilidad_individual_percent"], -25)
        self.assertEqual(result["rentabilidad_percent"], -25.0)
        self.assertEqual(result["pvp_unit"], 88.0)

    def test_use_global_option_falls_back_to_global(self) -> None:
        ui = app()

        percent, source = ui._supplier_order_effective_rentabilidad(
            {"rentabilidad_individual_percent": 20, "use_global_rentability": True},
            30,
        )

        self.assertEqual((percent, source), (30, "global"))
        updated = ui._supplier_order_update_line_rentabilidad(
            {
                "unit_cost": 100,
                "line_cost": 100,
                "rentabilidad_individual_percent": 20,
                "use_global_rentability": False,
            },
            use_global=True,
            global_percent=30,
            quantity=1,
        )
        self.assertNotIn("rentabilidad_individual_percent", updated)
        self.assertEqual(updated["rentabilidad_percent"], 30.0)
        self.assertEqual(updated["price_input_source"], "global_margin")
        self.assertTrue(updated["use_global_rentability"])

    def test_calculation_rows_show_cost_margin_and_pvp_in_order(self) -> None:
        ui = app()
        items, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
            (general_item(),),
            [],
        )

        row = ui._calculation_rows(items, provider="ekomat")[0]

        self.assertEqual(row[:6], ("201014", "Futon prueba", "100.00 €", "100.00 €", "142.86 €", "30"))

    def test_legacy_rows_without_pvp_unit_do_not_get_visual_pvp_fallback(self) -> None:
        ui = app()
        item = OrderItem(
            "201014",
            "Futon antiguo",
            1,
            "1",
            "100.00 €",
            "Calculado",
            raw={
                "source_row": {
                    "m3_total": 1,
                    "m3_und": 1,
                    "unit_cost": 100,
                    "precio_coste_final": 100,
                    "line_cost": 100,
                    "rentabilidad_percent": 30,
                    "cuenta_reparto_descarga": True,
                }
            },
        )

        row = ui._calculation_rows((item,), provider="ekomat")[0]

        self.assertEqual(row[:6], ("201014", "Futon antiguo", "100.00 €", "Pendiente", "Pendiente", "30"))
        self.assertNotIn("pvp_unit", item.raw["source_row"])

    def test_save_payload_keeps_real_cost_and_pvp_separate(self) -> None:
        ui = app()
        ui._cloud_session = Session()
        items, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
            (general_item(),),
            [],
        )
        entries = {"Nombre del pedido": Entry("PED-TEST"), "Rentabilidad %": Entry("30"), "Coste transporte + IVA": Entry("1")}

        with (
            patch.object(prototype_module, "create_supplier_order_draft", return_value={"order_id": "ORDER-1"}),
            patch.object(prototype_module, "update_supplier_order_calculation") as update_calc,
            patch.object(prototype_module.messagebox, "showinfo"),
        ):
            ui._save_supplier_order_draft_from_calc(Window(), "ekomat", entries, {"items": items, "raw_lines": [items[0].raw["source_row"]], "calculated": True})

        payload_item = update_calc.call_args.kwargs["items"][0]
        self.assertEqual(payload_item["unit_cost"], 100.0)
        self.assertEqual(payload_item["line_cost"], 100.0)
        self.assertEqual(payload_item["source_row"]["pvp_unit"], 142.86)
        self.assertEqual(payload_item["source_row"]["pvp_line"], 142.86)
        self.assertEqual(payload_item["source_row"]["rentabilidad_source"], "global")

    def test_service_serialization_and_reload_preserve_individual_rentabilidad(self) -> None:
        ui = app()
        session = MemorySession()
        session.client.tables["supplier_orders"].append(
            {
                "order_id": "ORDER-1",
                "local_order_id": "PED-TEST",
                "provider": "ekomat",
                "order_file": "pedido.xlsx",
                "status": "Borrador",
                "total_items": 2,
                "total_cost": 0,
                "notes": "",
                "created_at": "2026-06-18T08:00:00+00:00",
                "updated_at": "2026-06-18T08:00:00+00:00",
                "source_row": {"ui_order_name": "PED-TEST", "inputs": {"Rentabilidad %": "30"}},
            }
        )
        individual = general_item()
        individual_source = dict(individual.raw["source_row"])
        individual_source.update({"rentabilidad_individual_percent": 20, "use_global_rentability": False})
        individual = OrderItem(
            individual.code,
            individual.name,
            individual.quantity,
            individual.m3,
            individual.final_cost,
            individual.status,
            raw={"source_row": individual_source},
        )
        global_item = general_item()
        global_item = OrderItem(
            "201015",
            "Futon global",
            global_item.quantity,
            global_item.m3,
            global_item.final_cost,
            global_item.status,
            raw=global_item.raw,
        )
        calculated, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
            (individual, global_item),
            [],
        )
        payloads = []
        for item in calculated:
            source = item.raw["source_row"]
            payloads.append(
                {
                    "code": item.code,
                    "name": item.name,
                    "quantity": item.quantity,
                    "unit_cost": source["unit_cost"],
                    "line_cost": source["line_cost"],
                    "final_cost": source["line_cost"],
                    "status": item.status,
                    "source_row": source,
                }
            )

        settings = SimpleNamespace(sync_role="admin", machine_name="TEST")
        with (
            patch.object(orders_service, "load_settings", return_value=settings),
            patch.object(orders_service, "new_operation_id", return_value="ORDERCALC-TEST"),
            patch.object(orders_service, "write_snapshot"),
            patch.object(orders_service, "write_audit_event"),
        ):
            orders_service.update_supplier_order_calculation(
                session,
                order_id="ORDER-1",
                provider="ekomat",
                order_name="PED-TEST",
                order_file="pedido.xlsx",
                file_type="XLSX",
                inputs={"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
                items=payloads,
            )

        ui._cloud_session = session
        reloaded_orders = ui._load_supplier_orders_from_cloud()
        self.assertEqual(len(reloaded_orders), 1)
        self.assertEqual(len(reloaded_orders[0].items), 2)
        reloaded_individual = reloaded_orders[0].items[0].raw["source_row"]
        reloaded_global = reloaded_orders[0].items[1].raw["source_row"]
        self.assertEqual(reloaded_individual["rentabilidad_individual_percent"], 20)
        self.assertFalse(reloaded_individual["use_global_rentability"])
        self.assertTrue(reloaded_global["use_global_rentability"])

        recalculated, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "40", "Coste transporte + IVA": "11"},
            reloaded_orders[0].items,
            [],
        )
        individual_after = recalculated[0].raw["source_row"]
        global_after = recalculated[1].raw["source_row"]
        self.assertEqual(individual_after["rentabilidad_percent"], 20.0)
        self.assertEqual(individual_after["rentabilidad_source"], "individual")
        self.assertEqual(global_after["rentabilidad_percent"], 40.0)
        self.assertEqual(global_after["rentabilidad_source"], "global")
        self.assertEqual(
            individual_after["pvp_unit"],
            ui._supplier_order_pvp(individual_after["unit_cost"], 20),
        )
        self.assertEqual(
            global_after["pvp_unit"],
            ui._supplier_order_pvp(global_after["unit_cost"], 40),
        )

    def test_service_reload_preserves_negative_individual_margin_and_pvp(self) -> None:
        ui = app()
        session = MemorySession()
        session.client.tables["supplier_orders"].append(
            {
                "order_id": "ORDER-NEG",
                "local_order_id": "PED-NEG",
                "provider": "ekomat",
                "order_file": "pedido-neg.xlsx",
                "status": "Borrador",
                "total_items": 1,
                "total_cost": 0,
                "notes": "",
                "created_at": "2026-07-01T08:00:00+00:00",
                "updated_at": "2026-07-01T08:00:00+00:00",
                "source_row": {"ui_order_name": "PED-NEG", "inputs": {"Margen de Venta %": "30"}},
            }
        )
        item = general_item()
        source_row = dict(item.raw["source_row"])
        source_row.update(
            {
                "unit_cost": 100.0,
                "precio_coste_final": 100.0,
                "line_cost": 100.0,
                "precio_ponderado_lote": 100.0,
            }
        )
        source = ui._supplier_order_update_line_rentabilidad(
            source_row,
            use_global=True,
            global_percent=30,
            pvp_unit=80,
            input_source="pvp",
            quantity=1,
        )
        item = OrderItem(item.code, item.name, item.quantity, item.m3, item.final_cost, item.status, raw={"source_row": source})
        calculated, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Margen de Venta %": "30", "Coste transporte + IVA": "1"},
            (item,),
            [],
        )
        payload_source = calculated[0].raw["source_row"]
        payload = {
            "code": calculated[0].code,
            "name": calculated[0].name,
            "quantity": calculated[0].quantity,
            "unit_cost": payload_source["unit_cost"],
            "line_cost": payload_source["line_cost"],
            "final_cost": payload_source["line_cost"],
            "status": calculated[0].status,
            "source_row": payload_source,
        }
        settings = SimpleNamespace(sync_role="admin", machine_name="TEST")
        with (
            patch.object(orders_service, "load_settings", return_value=settings),
            patch.object(orders_service, "new_operation_id", return_value="ORDERCALC-NEG"),
            patch.object(orders_service, "write_snapshot"),
            patch.object(orders_service, "write_audit_event"),
        ):
            orders_service.update_supplier_order_calculation(
                session,
                order_id="ORDER-NEG",
                provider="ekomat",
                order_name="PED-NEG",
                order_file="pedido-neg.xlsx",
                file_type="XLSX",
                inputs={"Margen de Venta %": "30", "Coste transporte + IVA": "1"},
                items=[payload],
            )

        ui._cloud_session = session
        reloaded = ui._load_supplier_orders_from_cloud()[0].items[0]
        reloaded_source = reloaded.raw["source_row"]
        self.assertEqual(reloaded_source["rentabilidad_individual_percent"], -25.0)
        self.assertFalse(reloaded_source["use_global_rentability"])
        self.assertEqual(reloaded_source["price_input_source"], "pvp")
        self.assertEqual(reloaded_source["pvp_unit"], 80.0)

        recalculated, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Margen de Venta %": "45", "Coste transporte + IVA": "11"},
            (reloaded,),
            [],
        )
        result = recalculated[0].raw["source_row"]
        self.assertEqual(result["rentabilidad_source"], "individual")
        self.assertEqual(result["price_input_source"], "pvp")
        self.assertEqual(result["unit_cost"], 110.0)
        self.assertEqual(result["precio_ponderado_lote"], 110.0)
        self.assertEqual(result["rentabilidad_individual_percent"], -37.5)
        self.assertEqual(result["rentabilidad_percent"], -37.5)
        self.assertEqual(result["pvp_unit"], 80.0)
        self.assertEqual(recalculated[0].status, "Calculado")

    def test_historical_serialized_line_without_new_fields_uses_global_rentabilidad(self) -> None:
        ui = app()
        historical_row = {
            "id": 1,
            "order_id": "ORDER-OLD",
            "item_code": "201014",
            "item_name": "Futon historico",
            "quantity_ordered": 1,
            "quantity_received": 0,
            "unit_cost": 100,
            "line_cost": 100,
            "updated_at": "2025-01-01T00:00:00+00:00",
            "source_row": {
                **general_item().raw["source_row"],
                "unit_cost": 100,
                "line_cost": 100,
                "precio_coste_final": 100,
                "rentabilidad_percent": 15,
            },
        }

        item = ui._supplier_order_item_from_cloud_row(json.loads(json.dumps(historical_row)))
        recalculated, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "35", "Coste transporte + IVA": "1"},
            (item,),
            [],
        )

        source = recalculated[0].raw["source_row"]
        self.assertEqual(source["rentabilidad_percent"], 35.0)
        self.assertEqual(source["rentabilidad_source"], "global")
        self.assertTrue(source["use_global_rentability"])
        self.assertEqual(source["pvp_unit"], ui._supplier_order_pvp(source["unit_cost"], 35))

    def test_leading_zero_code_survives_service_reload_historical_recalculation_and_export(self) -> None:
        ui = app()
        del ui.__dict__["_fill_supplier_prices_for_order_items"]
        session = MemorySession()
        session.client.tables["inventory_items"] = [
            inventory_row(201001, name="Futon Canonico", heca_reference="201001")
        ]
        session.client.tables["supplier_orders"].append(
            {
                "order_id": "ORDER-ZERO",
                "local_order_id": "PED-ZERO",
                "provider": "ekomat",
                "order_file": "pedido-zero.xlsx",
                "status": "Borrador",
                "total_items": 1,
                "total_cost": 0,
                "notes": "",
                "created_at": "2025-01-01T00:00:00+00:00",
                "updated_at": "2025-01-01T00:00:00+00:00",
                "source_row": {"ui_order_name": "PED-ZERO", "inputs": {"Rentabilidad %": "30"}},
            }
        )
        ui._cloud_session = session
        historical = OrderItem(
            "0201001",
            "Pendiente",
            1,
            "1",
            "Pendiente",
            "OK",
            raw={
                "source_row": {
                    "m3_total": 1,
                    "m3_und": 1,
                    "packages": 1,
                    "cuenta_reparto_descarga": True,
                }
            },
        )
        calculated, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
            (historical,),
            [],
        )
        source = calculated[0].raw["source_row"]
        payload = {
            "code": calculated[0].code,
            "name": calculated[0].name,
            "quantity": calculated[0].quantity,
            "unit_cost": source["unit_cost"],
            "line_cost": source["line_cost"],
            "final_cost": source["line_cost"],
            "status": calculated[0].status,
            "source_row": source,
        }
        settings = SimpleNamespace(sync_role="admin", machine_name="TEST")
        with (
            patch.object(orders_service, "load_settings", return_value=settings),
            patch.object(orders_service, "new_operation_id", return_value="ORDERCALC-ZERO"),
            patch.object(orders_service, "write_snapshot"),
            patch.object(orders_service, "write_audit_event"),
        ):
            orders_service.update_supplier_order_calculation(
                session,
                order_id="ORDER-ZERO",
                provider="ekomat",
                order_name="PED-ZERO",
                order_file="pedido-zero.xlsx",
                file_type="XLSX",
                inputs={"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
                items=[payload],
            )

        reloaded = ui._load_supplier_orders_from_cloud()[0].items[0]
        self.assertEqual(reloaded.code, "0201001")
        recalculated, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
            (reloaded,),
            [],
        )
        self.assertEqual(recalculated[0].code, "0201001")
        self.assertEqual(recalculated[0].raw["source_row"]["supplier_price_item_id"], 201001)

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "pedido-zero.xlsx"
            with patch.object(prototype_module.messagebox, "showinfo"):
                ui._export_supplier_order_audit_excel(
                    None,
                    provider="ekomat",
                    values={"Nombre del pedido": "PED-ZERO"},
                    items=recalculated,
                    path=str(path),
                )
            ws = load_workbook(path)["Líneas calculadas"]
            self.assertEqual(ws.cell(row=2, column=1).value, "0201001")

    def test_export_contains_cost_margin_and_pvp_columns(self) -> None:
        ui = app()
        items, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "30", "Coste transporte + IVA": "1"},
            (general_item(),),
            [],
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "pedido.xlsx"
            with patch.object(prototype_module.messagebox, "showinfo"):
                ui._export_supplier_order_audit_excel(None, provider="ekomat", values={"Nombre del pedido": "PED-TEST"}, items=items, path=str(path))
            wb = load_workbook(path)
            ws = wb["Líneas calculadas"]
            headers = [cell.value for cell in ws[1]]
            cost_index = next(index for index, header in enumerate(headers) if str(header).startswith("Coste Final Art"))

            self.assertEqual(headers[cost_index : cost_index + 5], [headers[cost_index], "Precio ponderado lote", "P.V.P.", "Margen de Venta %", "Origen P.V.P."])
            self.assertEqual(ws.cell(row=2, column=cost_index + 1).value, 100)
            self.assertEqual(ws.cell(row=2, column=cost_index + 3).value, 142.86)
            self.assertEqual(ws.cell(row=2, column=cost_index + 4).value, 30)
            self.assertEqual(ws.cell(row=2, column=cost_index + 5).value, "Margen global")

    def test_weighted_cost_does_not_depend_on_rentabilidad(self) -> None:
        ui = app()
        low, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "10", "Coste transporte + IVA": "1"},
            (general_item(),),
            [],
        )
        high, _raw, _summary = ui._calculate_supplier_order_in_memory(
            "ekomat",
            {"Rentabilidad %": "60", "Coste transporte + IVA": "1"},
            (general_item(),),
            [],
        )

        self.assertEqual(
            low[0].raw["source_row"]["precio_ponderado_lote"],
            high[0].raw["source_row"]["precio_ponderado_lote"],
        )
        self.assertNotEqual(low[0].raw["source_row"]["pvp_unit"], high[0].raw["source_row"]["pvp_unit"])

    def test_editor_exposes_individual_and_global_rentability_controls(self) -> None:
        source = inspect.getsource(FutonHubErpPrototype._open_order_item_missing_editor)

        self.assertIn("rentabilidad_individual_percent", source)
        self.assertIn("use_global_rentability_var", source)
        self.assertIn("Usar margen global", source)
        self.assertIn("MARGEN DE VENTA", source)
        self.assertIn("pvp_unit_var", source)
        self.assertIn("refresh_from_pvp", source)
        self.assertIn("live_update", source)
        self.assertIn("allow_negative=True", source)
        self.assertIn("initial_price_input_source", source)
        self.assertIn('last_price_input_source = {"value": "pvp" if initial_price_input_source == "pvp" else "woo" if initial_price_input_source == "woo_price" else "margin"}', source)
        self.assertIn('if initial_price_input_source == "pvp"', source)
        self.assertNotIn("Los campos marcados en rojo", source)
        self.assertIn("_supplier_order_editor_window_metrics", source)
        self.assertIn("tk.Canvas(body", source)
        self.assertIn("body_scrollbar", source)
        self.assertIn("buttons.grid(row=2", source)
        self.assertNotIn("center_window(win, 700, 640)", source)
        self.assertIn('"Aceptar"', source)
        self.assertIn('"Cancelar"', source)
        self.assertIn("Ponderado calculado", source)
        self.assertNotIn("Coste Final calculado", source)
        self.assertIn("_supplier_order_update_line_rentabilidad", source)
        self.assertIn("supplier_price_warning", source)
        self.assertIn('"Origen precio"', source)
        self.assertNotIn('entries["Coste Final"]', source)

    def test_order_item_editor_height_is_capped_for_720px_screen(self) -> None:
        ui = app()

        width, height, min_width, min_height = ui._supplier_order_editor_window_metrics(1280, 720)

        self.assertLessEqual(height, 600)
        self.assertLessEqual(min_height, height)
        self.assertLessEqual(width, 700)
        self.assertLessEqual(min_width, width)

    def test_inventory_secondary_update_failures_do_not_show_popup(self) -> None:
        source = inspect.getsource(FutonHubErpPrototype._open_order_item_missing_editor)

        self.assertIn('inventory_update_status"] = "NO_CHANGES"', source)
        self.assertIn('inventory_update_status"] = "ERROR"', source)
        self.assertIn("[SUPPLIER_ORDER_INVENTORY_UPDATE]", source)
        self.assertNotIn('messagebox.showwarning(\n                        "Inventario"', source)
        self.assertIn("if callable(on_save):", source)
        self.assertIn("win.destroy()", source)

    def test_new_column_order_starts_with_cost_margin_pvp_and_weighted(self) -> None:
        source = inspect.getsource(FutonHubErpPrototype._calculation_tree)

        self.assertIn('"ID",\n                "Nombre",\n                "Coste Final",\n                "Ponderado",\n                "P.V.P.",\n                "Margen de Venta"', source)

    def test_receive_flow_does_not_use_pvp_fields_for_costs(self) -> None:
        receive_source = inspect.getsource(orders_service.receive_supplier_order)

        self.assertIn("unit_cost", orders_service.SUPPLIER_ORDER_ITEM_COLUMNS)
        self.assertIn("line_cost", orders_service.SUPPLIER_ORDER_ITEM_COLUMNS)
        self.assertNotIn("pvp", orders_service.SUPPLIER_ORDER_ITEM_COLUMNS.lower())
        self.assertNotIn("pvp_", receive_source)


if __name__ == "__main__":
    unittest.main()
